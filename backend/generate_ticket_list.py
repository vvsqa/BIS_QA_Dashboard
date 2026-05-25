"""Generate formatted ticket list Excel with PM links and bug counts."""
import json, os, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import Counter

DATA_DIR = os.path.join(os.path.dirname(__file__), 'data')
REPORTS_DIR = os.path.join(os.path.dirname(__file__), 'reports')
PM_URL = 'https://www.bissafety.app/pm/tickets#!/'

with open(os.path.join(DATA_DIR, 'pm_tickets_cache.json')) as f:
    tickets = json.load(f)
try:
    with open(os.path.join(DATA_DIR, 'redmine_cache.json')) as f:
        bugs_cache = json.load(f)
except:
    bugs_cache = {}

ticket_map = {t['ticket_id']: t for t in tickets}

ids = list(dict.fromkeys([
    18752, 18739, 18740, 20240, 17730, 20457, 20316, 18910, 16526, 19666,
    17175, 20515, 19729, 20318, 19691, 19414, 19190, 15413, 17541, 20410,
    19983, 19024, 20542, 20304, 19434, 20094, 20452, 18975, 18899,
    19860, 18773, 18025, 20161, 20395, 18261, 20238, 20257, 19068
]))

# Manual notes for specific tickets
TICKET_NOTES = {
    19860: 'Plan to test next week (Still in development)',
    18773: 'Plan to test next week (Still in development)',
    18025: 'Already in PRE',
    20161: 'Need confirmation (Lallu)',
    20395: 'In staging',
    18261: 'In staging',
    20238: 'In staging',
    20257: 'In staging',
    19068: 'In staging',
}

wb = Workbook()
ws = wb.active
ws.title = 'Ticket Details'
ws.sheet_properties.tabColor = '0D47A1'

# Styles
hdr_font = Font(bold=True, color='FFFFFF', size=10, name='Calibri')
data_font = Font(size=10, name='Calibri')
link_font = Font(size=10, name='Calibri', color='2196F3', underline='single', bold=True)
bold_font = Font(bold=True, size=10, name='Calibri')
tb = Border(
    left=Side(style='thin', color='B0BEC5'), right=Side(style='thin', color='B0BEC5'),
    top=Side(style='thin', color='B0BEC5'), bottom=Side(style='thin', color='B0BEC5')
)

hdr_fill = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
alt_row = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
red_bg = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
green_bg = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
amber_bg = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
blue_bg = PatternFill(start_color='BBDEFB', end_color='BBDEFB', fill_type='solid')
purple_bg = PatternFill(start_color='E1BEE7', end_color='E1BEE7', fill_type='solid')

ca = Alignment(horizontal='center', vertical='center', wrap_text=True)
la = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Title
ws.merge_cells('A1:Q1')
ws.cell(row=1, column=1, value='QA Ticket Status Report').font = Font(bold=True, size=14, name='Calibri', color='1A237E')
ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 30

ws.merge_cells('A2:Q2')
ws.cell(row=2, column=1, value=f'{len(ids)} tickets  |  {datetime.date.today().strftime("%d %b %Y")}').font = Font(size=10, name='Calibri', color='757575')
ws.cell(row=2, column=1).alignment = Alignment(horizontal='center')

# Headers
headers = ['#', 'Ticket', 'Title', 'Status', 'Activity', 'Module', 'Priority',
           'QC Tester', 'Backend Dev', 'Frontend Dev', 'Platform',
           'Bugs', 'Open', 'Rel. to QA', 'Closed', 'Type', 'Notes']
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=c, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = ca
    cell.border = tb
ws.row_dimensions[4].height = 28

widths = {'A': 5, 'B': 12, 'C': 55, 'D': 22, 'E': 22, 'F': 24, 'G': 16,
          'H': 20, 'I': 24, 'J': 24, 'K': 10, 'L': 8, 'M': 8, 'N': 10, 'O': 8, 'P': 14, 'Q': 40}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

for i, tid in enumerate(ids):
    row = i + 5
    t = ticket_map.get(tid)
    if not t:
        ws.cell(row=row, column=1, value=i + 1).border = tb
        ws.cell(row=row, column=2, value=tid).border = tb
        c3 = ws.cell(row=row, column=3, value='NOT FOUND IN PM')
        c3.border = tb
        c3.font = Font(color='C62828', bold=True, size=10)
        continue

    b = bugs_cache.get(str(tid)) or bugs_cache.get(tid) or {'total': 0, 'open': 0, 'closed': 0, 'released_to_qa': 0}
    status = t['status']

    # Activity
    if status == 'QC Testing Hold':
        activity = 'On Hold'
    elif status == 'QC Testing in Progress':
        activity = 'QC In Progress'
    elif status == 'QC Testing' and t.get('qc_tester'):
        activity = 'Assigned, Not Started'
    elif status == 'QC Testing' and not t.get('qc_tester'):
        activity = 'QA Unassigned'
    elif status == 'QC Review Fail':
        activity = 'QC Review Failed'
    elif status == 'BIS Testing':
        activity = 'BIS Testing'
    elif status == 'Approved for Live':
        activity = 'Prod Verification'
    elif status == 'In Progress' and t.get('qc_tester'):
        activity = 'Dev Refix'
    elif status == 'In Progress':
        activity = 'Dev In Progress'
    elif status == 'Code Review Passed':
        activity = 'Ready for QC'
    elif status in ('Start Code Review', 'Code Review Failed'):
        activity = 'Code Review'
    elif status in ('Closed', 'Moved to Live'):
        activity = 'Closed'
    else:
        activity = status

    # Status color
    status_colors = {
        'QC Review Fail': red_bg,
        'QC Testing Hold': amber_bg,
        'QC Testing in Progress': green_bg,
        'QC Testing': blue_bg,
        'BIS Testing': purple_bg,
        'Approved for Live': purple_bg,
        'Closed': green_bg,
        'Moved to Live': green_bg,
    }
    status_fill = status_colors.get(status)
    row_fill = alt_row if i % 2 == 0 else None

    vals = [
        i + 1, tid, t['title'], status, activity, t.get('module', '-'), t['priority'],
        t.get('qc_tester') or '-', t.get('backend_developer') or '-',
        t.get('frontend_developer') or '-', t.get('platform', 'Web'),
        b['total'] or '', b['open'] or '', b.get('released_to_qa', 0) or '',
        b['closed'] or '', t.get('ticket_type', '-'), TICKET_NOTES.get(tid, '')
    ]

    center_cols = {1, 2, 4, 5, 7, 11, 12, 13, 14, 15, 16}
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = data_font
        cell.border = tb
        cell.alignment = ca if c in center_cols else la
        if row_fill and c != 4:
            cell.fill = row_fill

    # Ticket ID hyperlink
    id_cell = ws.cell(row=row, column=2)
    id_cell.hyperlink = f'{PM_URL}{tid}'
    id_cell.font = link_font
    id_cell.value = tid

    # Status color
    if status_fill:
        ws.cell(row=row, column=4).fill = status_fill
        ws.cell(row=row, column=4).font = bold_font

    # Activity color
    act_colors = {
        'QC Review Failed': red_bg, 'Dev Refix': red_bg, 'On Hold': amber_bg,
        'QC In Progress': green_bg, 'BIS Testing': purple_bg, 'Ready for QC': blue_bg,
    }
    act_fill = act_colors.get(activity)
    if act_fill:
        ws.cell(row=row, column=5).fill = act_fill

    # Bug highlights
    if b['open'] > 0:
        ws.cell(row=row, column=13).fill = red_bg
        ws.cell(row=row, column=13).font = Font(bold=True, size=10, color='C62828', name='Calibri')
    if b.get('released_to_qa', 0) > 0:
        ws.cell(row=row, column=14).fill = green_bg

    # Notes highlight
    note = TICKET_NOTES.get(tid, '')
    if note:
        notes_cell = ws.cell(row=row, column=17)
        notes_cell.fill = amber_bg
        notes_cell.font = Font(bold=True, size=10, name='Calibri', color='E65100')

    ws.row_dimensions[row].height = 30

# Freeze header
ws.freeze_panes = 'A5'
ws.auto_filter.ref = f'A4:Q{4 + len(ids)}'

# Summary
sr = 5 + len(ids) + 2
ws.merge_cells(f'A{sr}:C{sr}')
ws.cell(row=sr, column=1, value='Summary').font = Font(bold=True, size=12, name='Calibri', color='1A237E')
sr += 1

status_counts = Counter()
for tid in ids:
    t = ticket_map.get(tid)
    if t:
        status_counts[t['status']] += 1

ws.cell(row=sr, column=2, value='Status').font = bold_font
ws.cell(row=sr, column=3, value='Count').font = bold_font
sr += 1
for s, c in status_counts.most_common():
    ws.cell(row=sr, column=2, value=s).font = data_font
    ws.cell(row=sr, column=3, value=c).font = bold_font
    ws.cell(row=sr, column=2).border = tb
    ws.cell(row=sr, column=3).border = tb
    sr += 1

sr += 1
total_bugs = sum((bugs_cache.get(str(tid)) or {}).get('total', 0) for tid in ids)
open_bugs = sum((bugs_cache.get(str(tid)) or {}).get('open', 0) for tid in ids)
ws.cell(row=sr, column=2, value='Total Bugs').font = data_font
ws.cell(row=sr, column=3, value=total_bugs).font = bold_font
sr += 1
ws.cell(row=sr, column=2, value='Open Bugs').font = Font(bold=True, size=10, color='C62828', name='Calibri')
ws.cell(row=sr, column=3, value=open_bugs).font = Font(bold=True, size=10, color='C62828', name='Calibri')

os.makedirs(REPORTS_DIR, exist_ok=True)
path = os.path.join(REPORTS_DIR, 'Custom_Ticket_List.xlsx')
wb.save(path)
print(f'Saved: {path} ({len(ids)} tickets)')
