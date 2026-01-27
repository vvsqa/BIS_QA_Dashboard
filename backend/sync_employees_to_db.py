"""
Employee Excel Import Script

This script imports employee data from Excel files into the database.

Usage:
    python sync_employees_to_db.py --file "path/to/employee_details.xlsx"
    python sync_employees_to_db.py --file "C:\\Users\\Vishnu\\OneDrive\\Desktop\\employee details.xlsx"
"""

import os
import sys
import argparse
from datetime import datetime

# Fix Unicode output on Windows
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook

from database import SessionLocal
from models import Employee

# Column mapping from Excel headers to database fields
COLUMN_MAPPING = {
    "Name": "name",
    "Email": "email",
    "Employee ID": "employee_id",
    "Role": "role",
    "Location": "location",
    "Introduced/Joined": "date_of_joining",
    "Team": "team",
    "Category": "category",
    "Lead": "lead",
}


def parse_string(value):
    """Parse string value, returning None for empty values"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return str(value)
    value = str(value).strip()
    if value.lower() in ('', 'n/a', 'na', '-', 'none'):
        return None
    return value


def parse_datetime_value(value):
    """Parse datetime from Excel cell"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        # Try common date formats
        formats = [
            "%Y-%m-%d",
            "%d-%m-%Y",
            "%m/%d/%Y",
            "%d/%m/%Y",
            "%Y-%m-%d %H:%M:%S",
        ]
        for fmt in formats:
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                continue
    return None


def find_header_row(ws):
    """Find the header row in the worksheet"""
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        row_values = [str(cell).strip() if cell else "" for cell in row]
        if any("Name" in val for val in row_values) or any("Employee ID" in val for val in row_values):
            return row_idx
    return 1


def map_headers(header_row):
    """Map Excel headers to database field names"""
    column_map = {}
    for idx, header in enumerate(header_row):
        if header is None:
            continue
        header_str = str(header).strip()
        for excel_header, db_field in COLUMN_MAPPING.items():
            if excel_header.lower() == header_str.lower():
                column_map[idx] = db_field
                break
    return column_map


def import_employees(filepath):
    """Import employee data from Excel file"""
    print(f"\n{'='*60}")
    print(f"Importing Employees: {filepath}")
    print(f"Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*60}")
    
    if not os.path.exists(filepath):
        print(f"ERROR: File not found: {filepath}")
        return False, 0, 0
    
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        
        # Find header row
        header_row_idx = find_header_row(ws)
        print(f"Header row found at row: {header_row_idx}")
        
        # Get header row
        header_row = []
        for cell in ws[header_row_idx]:
            header_row.append(cell.value)
        
        # Map headers to database fields
        column_map = map_headers(header_row)
        print(f"Mapped columns: {column_map}")
        
        if 'name' not in column_map.values():
            print("ERROR: Could not find Name column in the Excel file")
            return False, 0, 0
        
        # Find key column indices
        name_col = None
        employee_id_col = None
        for idx, field in column_map.items():
            if field == 'name':
                name_col = idx
            if field == 'employee_id':
                employee_id_col = idx
        
        db = SessionLocal()
        imported = 0
        updated = 0
        skipped = 0
        
        try:
            for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
                # Get name first
                if name_col is None or name_col >= len(row):
                    continue
                
                name = parse_string(row[name_col])
                if not name:
                    skipped += 1
                    continue
                
                # Get employee_id for lookup
                employee_id = None
                if employee_id_col is not None and employee_id_col < len(row):
                    employee_id = parse_string(row[employee_id_col])
                
                # Check if record exists (by employee_id or name)
                existing = None
                if employee_id:
                    existing = db.query(Employee).filter(Employee.employee_id == employee_id).first()
                if not existing:
                    existing = db.query(Employee).filter(Employee.name == name).first()
                
                if existing:
                    record = existing
                    updated += 1
                else:
                    record = Employee()
                    record.created_on = datetime.utcnow()
                    imported += 1
                
                # Update fields based on column mapping
                for col_idx, db_field in column_map.items():
                    if col_idx >= len(row):
                        continue
                    
                    value = row[col_idx]
                    
                    if db_field == 'date_of_joining':
                        setattr(record, db_field, parse_datetime_value(value))
                    elif db_field == 'team':
                        # Normalize team names
                        team_value = parse_string(value)
                        if team_value:
                            team_value = team_value.upper()
                            if 'QA' in team_value:
                                team_value = 'QA'
                            elif 'DEV' in team_value or 'DEVELOPMENT' in team_value:
                                team_value = 'DEVELOPMENT'
                        setattr(record, db_field, team_value)
                    else:
                        setattr(record, db_field, parse_string(value))
                
                record.updated_on = datetime.utcnow()
                record.is_active = True
                
                if not existing:
                    db.add(record)
                
                # Commit every 50 records
                if (imported + updated) % 50 == 0:
                    db.commit()
                    print(f"  Progress: {imported} imported, {updated} updated...")
            
            db.commit()
            print(f"\nImport completed successfully!")
            print(f"  New employees: {imported}")
            print(f"  Updated employees: {updated}")
            print(f"  Skipped rows: {skipped}")
            
            wb.close()
            return True, imported, updated
            
        except Exception as e:
            db.rollback()
            print(f"ERROR during import: {e}")
            import traceback
            traceback.print_exc()
            return False, 0, 0
        finally:
            db.close()
            
    except Exception as e:
        print(f"ERROR opening Excel file: {e}")
        import traceback
        traceback.print_exc()
        return False, 0, 0


def main():
    parser = argparse.ArgumentParser(
        description="Import employee data from Excel files",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python sync_employees_to_db.py --file "C:/path/to/employee_details.xlsx"
        """
    )
    parser.add_argument('--file', '-f', type=str, required=True, help="Path to Excel file to import")
    
    args = parser.parse_args()
    
    success, imported, updated = import_employees(args.file)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
