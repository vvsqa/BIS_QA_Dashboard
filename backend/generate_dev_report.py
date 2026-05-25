"""Generate Dev Team Excel reports — weekly and monthly with time period selection."""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import Counter, defaultdict
from datetime import date, timedelta

try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(__file__), '.env'))
except ImportError:
    pass

DATA_FILE = os.path.join(os.path.dirname(__file__), 'data', 'pm_tickets_cache.json')
REDMINE_FILE = os.path.join(os.path.dirname(__file__), 'data', 'redmine_cache.json')
REPORTS_DIR = os.path.join(os.path.dirname(__file__), 'reports')

with open(DATA_FILE) as f:
    tickets = json.load(f)
try:
    with open(REDMINE_FILE) as f:
        bugs_cache = json.load(f)
except (FileNotFoundError, json.JSONDecodeError):
    bugs_cache = {}

today = date.today()
d7 = today - timedelta(days=7)
d14 = today - timedelta(days=14)
d30 = today - timedelta(days=30)
d60 = today - timedelta(days=60)

def parse_d(v):
    if not v: return None
    try: return date.fromisoformat(str(v)[:10])
    except: return None

def sum_h(tl, f): return round(sum(t.get(f) or 0 for t in tl), 1)
def get_bugs(tid):
    return bugs_cache.get(str(tid)) or bugs_cache.get(tid) or {'total':0,'open':0,'closed':0,'released_to_qa':0}

# Styles
hf = Font(bold=True, color='FFFFFF', size=11)
sf = Font(bold=True, size=12)
tb = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
green_fill = PatternFill(start_color='1B5E20', end_color='1B5E20', fill_type='solid')
blue_fill = PatternFill(start_color='0D47A1', end_color='0D47A1', fill_type='solid')
purple_fill = PatternFill(start_color='6A1B9A', end_color='6A1B9A', fill_type='solid')
orange_fill = PatternFill(start_color='E65100', end_color='E65100', fill_type='solid')
red_fill = PatternFill(start_color='B71C1C', end_color='B71C1C', fill_type='solid')
teal_fill = PatternFill(start_color='00695C', end_color='00695C', fill_type='solid')
light_red = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
light_green = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
light_blue = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
ca = Alignment(horizontal='center', vertical='center')

def write_header(ws, row, headers, fill):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = hf; cell.fill = fill; cell.alignment = ca; cell.border = tb

def write_row(ws, row, vals):
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.border = tb; cell.alignment = Alignment(horizontal='center' if isinstance(v, (int, float)) else 'left')

def write_section(ws, row, title):
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = sf; cell.fill = light_blue
    return row + 1


def generate_dev_report(period_start, period_end, prev_start, prev_end, label):
    """Generate dev team report for given period."""
    wb = Workbook()

    # Filter tickets by period (created_on or closed_on in range)
    def in_period(t, start, end):
        created = parse_d(t.get('created_on'))
        closed = parse_d(t.get('closed_on'))
        return (created and start <= created <= end) or (closed and start <= closed <= end)

    period_tickets = [t for t in tickets if in_period(t, period_start, period_end)]
    prev_tickets = [t for t in tickets if in_period(t, prev_start, prev_end)]

    # Only relevant statuses
    DEV_STATUSES = {'Ready For Development', 'In Progress', 'Hold/Pending', 'Start Code Review',
                    'Code Review Failed', 'Code Review Passed', 'Express Lane Review', 'Testing In Progress'}
    QA_STATUSES = {'QC Testing', 'QC Testing in Progress', 'QC Testing Hold', 'QC Review Fail',
                   'BIS Testing', 'Tested - Awaiting Fixes', 'Approved for Live', 'Moved to Live'}
    active_statuses = DEV_STATUSES | QA_STATUSES
    all_active = [t for t in tickets if t['status'] in active_statuses]

    # ===== Sheet 1: Executive Summary =====
    ws = wb.active; ws.title = 'Executive Summary'
    ws.column_dimensions['A'].width = 45; ws.column_dimensions['B'].width = 18; ws.column_dimensions['C'].width = 18; ws.column_dimensions['D'].width = 18

    ws.cell(row=1, column=1, value=f'Dev Team Report — {period_start.strftime("%b %d")} to {period_end.strftime("%b %d, %Y")}').font = Font(bold=True, size=14)

    # Separate dev vs QA statuses
    # Dev workflow as confirmed by user
    DEV_STATUSES_ORDER = ['Ready For Development', 'In Progress', 'Hold/Pending', 'Start Code Review', 'Code Review Failed', 'Code Review Passed', 'Express Lane Review', 'Testing In Progress']
    QA_STATUSES_ORDER = ['QC Testing', 'QC Testing in Progress', 'QC Testing Hold', 'QC Review Fail',
                         'BIS Testing', 'Tested - Awaiting Fixes', 'Approved for Live', 'Moved to Live']

    status_counts = Counter(t['status'] for t in all_active)
    dev_active = [t for t in all_active if t['status'] in set(DEV_STATUSES_ORDER)]
    qa_active = [t for t in all_active if t['status'] in set(QA_STATUSES_ORDER)]

    r = write_section(ws, 3, 'Development Pipeline')
    write_header(ws, r, ['Status', 'Count', 'Dev Est Hours', 'Dev Actual Hours', 'Overrun Tickets'], green_fill); r += 1
    for s in DEV_STATUSES_ORDER:
        c = status_counts.get(s, 0)
        if c == 0: continue
        tl = [t for t in dev_active if t['status'] == s]
        ov = sum(1 for t in tl if (t.get('actual_dev_hours') or 0) > (t.get('dev_estimate_hours') or 999))
        write_row(ws, r, [s, c, sum_h(tl, 'dev_estimate_hours'), sum_h(tl, 'actual_dev_hours'), ov]); r += 1
    write_row(ws, r, ['Dev Total', len(dev_active), sum_h(dev_active, 'dev_estimate_hours'), sum_h(dev_active, 'actual_dev_hours'), ''])
    ws.cell(row=r, column=1).font = Font(bold=True); r += 1

    r = write_section(ws, r + 1, 'QA Pipeline (for reference)')
    write_header(ws, r, ['Status', 'Count', 'QA Est Hours', 'QA Actual Hours', 'Bugs Open'], orange_fill); r += 1
    for s in QA_STATUSES_ORDER:
        c = status_counts.get(s, 0)
        if c == 0: continue
        tl = [t for t in qa_active if t['status'] == s]
        bugs_open = sum(get_bugs(t['ticket_id'])['open'] for t in tl)
        write_row(ws, r, [s, c, sum_h(tl, 'qa_estimate_hours'), sum_h(tl, 'qa_actual_hours'), bugs_open]); r += 1
    write_row(ws, r, ['QA Total', len(qa_active), sum_h(qa_active, 'qa_estimate_hours'), sum_h(qa_active, 'qa_actual_hours'), ''])
    ws.cell(row=r, column=1).font = Font(bold=True); r += 1

    r = write_section(ws, r + 1, 'Bug Summary')
    write_header(ws, r, ['Metric', 'Count'], blue_fill); r += 1
    total_bugs = sum(get_bugs(t['ticket_id'])['total'] for t in all_active)
    open_bugs = sum(get_bugs(t['ticket_id'])['open'] for t in all_active)
    closed_bugs = sum(get_bugs(t['ticket_id'])['closed'] for t in all_active)
    write_row(ws, r, ['Total Bugs (active tickets)', total_bugs]); r += 1
    write_row(ws, r, ['Open Bugs', open_bugs]); r += 1
    write_row(ws, r, ['Closed Bugs', closed_bugs]); r += 1

    r = write_section(ws, r + 1, 'Dev Hours Summary')
    write_header(ws, r, ['Metric', 'Hours'], blue_fill); r += 1
    write_row(ws, r, ['Total Dev Estimated', sum_h(dev_active, 'dev_estimate_hours')]); r += 1
    write_row(ws, r, ['Total Dev Actual', sum_h(dev_active, 'actual_dev_hours')]); r += 1
    overrun = [t for t in dev_active if (t.get('actual_dev_hours') or 0) > (t.get('dev_estimate_hours') or 999)]
    write_row(ws, r, ['Tickets with Dev Overrun', len(overrun)]); r += 1
    total_overrun = sum((t.get('actual_dev_hours') or 0) - (t.get('dev_estimate_hours') or 0) for t in overrun)
    write_row(ws, r, ['Total Overrun Hours', round(total_overrun, 1)]); r += 1
    refix_count = sum(1 for t in dev_active if t.get('qc_tester'))
    write_row(ws, r, ['Dev Refix Tickets (from QA fail)', refix_count]); r += 1

    # Top modules by bugs
    r = write_section(ws, r + 1, 'Top 10 Modules by Bug Count')
    write_header(ws, r, ['Module', 'Total Bugs', 'Open', 'Closed'], orange_fill); r += 1
    mod_bugs = defaultdict(lambda: {'total': 0, 'open': 0, 'closed': 0})
    for t in all_active:
        b = get_bugs(t['ticket_id'])
        if b['total'] > 0:
            mod_bugs[t.get('module', 'Other')]['total'] += b['total']
            mod_bugs[t.get('module', 'Other')]['open'] += b['open']
            mod_bugs[t.get('module', 'Other')]['closed'] += b['closed']
    for mod, b in sorted(mod_bugs.items(), key=lambda x: -x[1]['total'])[:10]:
        write_row(ws, r, [mod, b['total'], b['open'], b['closed']]); r += 1

    # ===== Sheet 2: Status Breakdown =====
    ws2 = wb.create_sheet('Status Breakdown')
    ws2.column_dimensions['A'].width = 25
    for c in 'BCDEFGH': ws2.column_dimensions[c].width = 16

    r2 = 1
    ws2.cell(row=r2, column=1, value='Development Statuses').font = sf; ws2.cell(row=r2, column=1).fill = light_blue; r2 += 1
    write_header(ws2, r2, ['Status', 'Current', f'Period ({period_start.strftime("%b %d")}-{period_end.strftime("%b %d")})',
                           f'Previous ({prev_start.strftime("%b %d")}-{prev_end.strftime("%b %d")})', 'Change',
                           'Dev Est Hrs', 'Dev Actual Hrs', 'Overrun Count'], green_fill); r2 += 1
    for s in DEV_STATUSES_ORDER:
        c = status_counts.get(s, 0)
        if c == 0: continue
        p_count = sum(1 for t in period_tickets if t['status'] == s)
        pp_count = sum(1 for t in prev_tickets if t['status'] == s)
        tl = [t for t in dev_active if t['status'] == s]
        ov = sum(1 for t in tl if (t.get('actual_dev_hours') or 0) > (t.get('dev_estimate_hours') or 999))
        write_row(ws2, r2, [s, c, p_count, pp_count, p_count - pp_count, sum_h(tl, 'dev_estimate_hours'), sum_h(tl, 'actual_dev_hours'), ov])
        if p_count - pp_count > 0: ws2.cell(row=r2, column=5).fill = light_green
        elif p_count - pp_count < 0: ws2.cell(row=r2, column=5).fill = light_red
        r2 += 1

    r2 += 1
    ws2.cell(row=r2, column=1, value='QA Statuses (for reference)').font = sf; ws2.cell(row=r2, column=1).fill = light_blue; r2 += 1
    write_header(ws2, r2, ['Status', 'Current', f'Period', f'Previous', 'Change',
                           'QA Est Hrs', 'QA Actual Hrs', 'Bugs Open'], orange_fill); r2 += 1
    for s in QA_STATUSES_ORDER:
        c = status_counts.get(s, 0)
        if c == 0: continue
        p_count = sum(1 for t in period_tickets if t['status'] == s)
        pp_count = sum(1 for t in prev_tickets if t['status'] == s)
        tl = [t for t in qa_active if t['status'] == s]
        bugs_open = sum(get_bugs(t['ticket_id'])['open'] for t in tl)
        write_row(ws2, r2, [s, c, p_count, pp_count, p_count - pp_count, sum_h(tl, 'qa_estimate_hours'), sum_h(tl, 'qa_actual_hours'), bugs_open])
        r2 += 1

    # ===== Sheet 3: Developer Performance =====
    ws3 = wb.create_sheet('Developer Performance')
    write_header(ws3, 1, ['Developer', 'Active Tickets', 'In Progress', 'Code Review', 'CR Passed',
                           'Bugs Total', 'Bugs Open', 'Bugs Closed', 'Refix Count',
                           'Dev Est Hrs', 'Dev Actual Hrs', 'Overrun Hrs', 'Overrun Tickets',
                           'Efficiency %', 'Top Module'], purple_fill)
    dev_data = defaultdict(lambda: {'tickets': [], 'bugs': {'total': 0, 'open': 0, 'closed': 0},
                                     'refix': 0, 'statuses': Counter(), 'modules': Counter()})
    for t in all_active:
        for d in t.get('developers', []):
            if not d: continue
            dev_data[d]['tickets'].append(t)
            b = get_bugs(t['ticket_id'])
            dev_data[d]['bugs']['total'] += b['total']
            dev_data[d]['bugs']['open'] += b['open']
            dev_data[d]['bugs']['closed'] += b['closed']
            if t.get('qc_tester') and t['status'] in DEV_STATUSES:
                dev_data[d]['refix'] += 1
            dev_data[d]['statuses'][t['status']] += 1
            dev_data[d]['modules'][t.get('module', 'Other')] += 1

    # Sort by: refix desc, then bugs desc, then overrun desc, then ticket count desc
    for i, (name, d) in enumerate(sorted(dev_data.items(), key=lambda x: (-x[1]['refix'], -x[1]['bugs']['total'], -len(x[1]['tickets']))), 2):
        tl = d['tickets']
        est = sum_h(tl, 'dev_estimate_hours'); act = sum_h(tl, 'actual_dev_hours')
        ov_tickets = sum(1 for t in tl if (t.get('actual_dev_hours') or 0) > (t.get('dev_estimate_hours') or 999))
        ov_hrs = sum((t.get('actual_dev_hours') or 0) - (t.get('dev_estimate_hours') or 0) for t in tl if (t.get('actual_dev_hours') or 0) > (t.get('dev_estimate_hours') or 999))
        eff = round((act / est) * 100, 1) if est else 0
        top_mod = d['modules'].most_common(1)[0][0] if d['modules'] else '-'
        vals = [name, len(tl), d['statuses'].get('In Progress', 0), d['statuses'].get('Start Code Review', 0) + d['statuses'].get('Code Review Failed', 0),
                d['statuses'].get('Code Review Passed', 0), d['bugs']['total'], d['bugs']['open'], d['bugs']['closed'],
                d['refix'], est, act, round(ov_hrs, 1), ov_tickets, f'{eff}%', top_mod]
        write_row(ws3, i, vals)
        if ov_tickets > 0: ws3.cell(row=i, column=12).fill = light_red
    ws3.column_dimensions['A'].width = 30; ws3.column_dimensions['O'].width = 22
    for c in 'BCDEFGHIJKLMN': ws3.column_dimensions[c].width = 14

    # ===== Sheet 4: Rankings (sorted by context) =====
    ws4 = wb.create_sheet('Rankings')
    ws4.column_dimensions['A'].width = 8; ws4.column_dimensions['B'].width = 30
    for c in 'CDEFGH': ws4.column_dimensions[c].width = 16

    r4 = 1
    # Ranking 1: Most Refix (top concern)
    ws4.cell(row=r4, column=1, value='Developers with Most Refix (QC Fail Returns)').font = sf; ws4.cell(row=r4, column=1).fill = light_red; r4 += 1
    write_header(ws4, r4, ['Rank', 'Developer', 'Refix Count', 'Total Bugs', 'Open Bugs', 'Overrun Tickets', 'Overrun Hrs', 'Active Tickets'], red_fill); r4 += 1
    refix_ranked = sorted([(n, d) for n, d in dev_data.items() if d['refix'] > 0], key=lambda x: -x[1]['refix'])
    for rank, (name, d) in enumerate(refix_ranked, 1):
        tl = d['tickets']
        ov_t = sum(1 for t in tl if (t.get('actual_dev_hours') or 0) > (t.get('dev_estimate_hours') or 999))
        ov_h = sum((t.get('actual_dev_hours') or 0) - (t.get('dev_estimate_hours') or 0) for t in tl if (t.get('actual_dev_hours') or 0) > (t.get('dev_estimate_hours') or 999))
        write_row(ws4, r4, [rank, name, d['refix'], d['bugs']['total'], d['bugs']['open'], ov_t, round(ov_h, 1), len(tl)])
        ws4.cell(row=r4, column=3).fill = light_red
        r4 += 1

    r4 += 1
    # Ranking 2: Most Bugs
    ws4.cell(row=r4, column=1, value='Developers with Most Bugs Reported').font = sf; ws4.cell(row=r4, column=1).fill = light_red; r4 += 1
    write_header(ws4, r4, ['Rank', 'Developer', 'Total Bugs', 'Open Bugs', 'Closed Bugs', 'Refix Count', 'Active Tickets', ''], red_fill); r4 += 1
    bug_ranked = sorted([(n, d) for n, d in dev_data.items() if d['bugs']['total'] > 0], key=lambda x: -x[1]['bugs']['total'])
    for rank, (name, d) in enumerate(bug_ranked[:20], 1):
        write_row(ws4, r4, [rank, name, d['bugs']['total'], d['bugs']['open'], d['bugs']['closed'], d['refix'], len(d['tickets']), ''])
        if d['bugs']['open'] > 5: ws4.cell(row=r4, column=4).fill = light_red
        r4 += 1

    r4 += 1
    # Ranking 3: Most Time Overrun
    ws4.cell(row=r4, column=1, value='Developers with Most Time Overrun').font = sf; ws4.cell(row=r4, column=1).fill = light_red; r4 += 1
    write_header(ws4, r4, ['Rank', 'Developer', 'Overrun Tickets', 'Overrun Hrs', 'Dev Est Hrs', 'Dev Actual Hrs', 'Active Tickets', ''], orange_fill); r4 += 1
    overrun_data = []
    for name, d in dev_data.items():
        tl = d['tickets']
        ov_t = sum(1 for t in tl if (t.get('actual_dev_hours') or 0) > (t.get('dev_estimate_hours') or 999))
        if ov_t > 0:
            ov_h = sum((t.get('actual_dev_hours') or 0) - (t.get('dev_estimate_hours') or 0) for t in tl if (t.get('actual_dev_hours') or 0) > (t.get('dev_estimate_hours') or 999))
            overrun_data.append((name, ov_t, ov_h, d))
    for rank, (name, ov_t, ov_h, d) in enumerate(sorted(overrun_data, key=lambda x: -x[2]), 1):
        tl = d['tickets']
        write_row(ws4, r4, [rank, name, ov_t, round(ov_h, 1), sum_h(tl, 'dev_estimate_hours'), sum_h(tl, 'actual_dev_hours'), len(tl), ''])
        r4 += 1

    # ===== Sheet 5: Module Quality =====
    ws5 = wb.create_sheet('Module Quality')
    write_header(ws5, 1, ['Module', 'Active Tickets', 'Total Bugs', 'Open Bugs', 'Closed Bugs',
                           'Refix Tickets', 'Dev Est Hrs', 'Dev Actual Hrs', 'Overrun Hrs', 'Developers'], teal_fill)
    mod_data = defaultdict(lambda: {'tickets': [], 'bugs': {'total': 0, 'open': 0, 'closed': 0}, 'refix': 0, 'devs': set()})
    for t in all_active:
        mod = t.get('module', 'Other')
        mod_data[mod]['tickets'].append(t)
        b = get_bugs(t['ticket_id'])
        mod_data[mod]['bugs']['total'] += b['total']; mod_data[mod]['bugs']['open'] += b['open']; mod_data[mod]['bugs']['closed'] += b['closed']
        if t.get('qc_tester'): mod_data[mod]['refix'] += 1
        for d in t.get('developers', []): mod_data[mod]['devs'].add(d)

    for i, (mod, m) in enumerate(sorted(mod_data.items(), key=lambda x: -x[1]['bugs']['total']), 2):
        tl = m['tickets']
        ov_hrs = sum((t.get('actual_dev_hours') or 0) - (t.get('dev_estimate_hours') or 0) for t in tl if (t.get('actual_dev_hours') or 0) > (t.get('dev_estimate_hours') or 999))
        write_row(ws5, i, [mod, len(tl), m['bugs']['total'], m['bugs']['open'], m['bugs']['closed'], m['refix'],
                           sum_h(tl, 'dev_estimate_hours'), sum_h(tl, 'actual_dev_hours'), round(ov_hrs, 1),
                           ', '.join(sorted(m['devs'] - {''})[:5])])
    ws5.column_dimensions['A'].width = 25; ws5.column_dimensions['J'].width = 50
    for c in 'BCDEFGHI': ws5.column_dimensions[c].width = 14

    # ===== Sheet 6: Ticket Details (Dev statuses only) =====
    ws6 = wb.create_sheet('Ticket Details')
    write_header(ws6, 1, ['Ticket', 'Title', 'Status', 'Priority', 'Module', 'Platform', 'Backend Dev', 'Frontend Dev',
                           'Dev Est', 'Dev Actual', 'Overrun', 'Refix', 'Bugs', 'Open Bugs', 'ETA', 'Created'], blue_fill)
    dev_only = [t for t in all_active if t['status'] in set(DEV_STATUSES_ORDER)]
    for i, t in enumerate(sorted(dev_only, key=lambda x: -x['ticket_id']), 2):
        est = t.get('dev_estimate_hours') or 0; act = t.get('actual_dev_hours') or 0
        ov = round(act - est, 1) if est > 0 and act > est else 0
        b = get_bugs(t['ticket_id'])
        is_refix = 'Yes' if t.get('qc_tester') and t['status'] in DEV_STATUSES else ''
        write_row(ws6, i, [t['ticket_id'], t['title'], t['status'], t['priority'], t.get('module', '-'), t.get('platform', '-'),
                           t.get('backend_developer', '-'), t.get('frontend_developer', '-'),
                           est, act, ov if ov > 0 else '', is_refix, b['total'] if b['total'] > 0 else '', b['open'] if b['open'] > 0 else '',
                           t.get('eta', '-'), t.get('created_on', '-')])
        if ov > 0: ws6.cell(row=i, column=11).fill = light_red
        if is_refix: ws6.cell(row=i, column=12).fill = light_red
        if b['open'] > 0: ws6.cell(row=i, column=14).fill = light_red
    ws6.column_dimensions['A'].width = 10; ws6.column_dimensions['B'].width = 55; ws6.column_dimensions['E'].width = 22
    for c in 'CDFGHIJKLMNOP': ws6.column_dimensions[c].width = 14

    # ===== Sheet 7: Developer Ticket Mapping =====
    ws7 = wb.create_sheet('Developer Ticket Mapping')
    ws7.column_dimensions['A'].width = 30; ws7.column_dimensions['B'].width = 12; ws7.column_dimensions['C'].width = 55
    ws7.column_dimensions['D'].width = 22; ws7.column_dimensions['E'].width = 18; ws7.column_dimensions['F'].width = 22
    ws7.column_dimensions['G'].width = 14; ws7.column_dimensions['H'].width = 14

    ws7.column_dimensions['A'].width = 12; ws7.column_dimensions['B'].width = 50; ws7.column_dimensions['C'].width = 22
    ws7.column_dimensions['D'].width = 16; ws7.column_dimensions['E'].width = 22; ws7.column_dimensions['F'].width = 14
    ws7.column_dimensions['G'].width = 14; ws7.column_dimensions['H'].width = 14; ws7.column_dimensions['I'].width = 10
    ws7.column_dimensions['J'].width = 10; ws7.column_dimensions['K'].width = 12; ws7.column_dimensions['L'].width = 10
    ws7.column_dimensions['M'].width = 16; ws7.column_dimensions['N'].width = 12

    # Load cycle tracker for fail loop counts
    CYCLE_FILE = os.path.join(os.path.dirname(__file__), 'data', 'qc_cycle_tracker.json')
    try:
        with open(CYCLE_FILE) as f:
            cycle_tracker = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        cycle_tracker = {}

    ws7.column_dimensions['O'].width = 12; ws7.column_dimensions['P'].width = 12
    COL_COUNT = 16

    r7 = 1
    status_order = {'In Progress': 0, 'Hold/Pending': 1, 'Start Code Review': 2, 'Code Review Failed': 3,
                    'Code Review Passed': 4, 'QC Testing': 5, 'QC Testing in Progress': 6, 'QC Testing Hold': 7,
                    'QC Review Fail': 8, 'BIS Testing': 9, 'Approved for Live': 10}

    # Sort developers: most refix first, then most bugs, then most tickets
    for name, d in sorted(dev_data.items(), key=lambda x: (-x[1]['refix'], -x[1]['bugs']['total'], -len(x[1]['tickets']))):
        tl = d['tickets']
        if not tl:
            continue

        # Developer header with summary
        refix_label = f' | {d["refix"]} refix' if d['refix'] > 0 else ''
        bugs_label = f' | {d["bugs"]["total"]} bugs ({d["bugs"]["open"]} open)' if d['bugs']['total'] > 0 else ''
        ov_count = sum(1 for t in tl if (t.get('actual_dev_hours') or 0) > (t.get('dev_estimate_hours') or 999))
        ov_label = f' | {ov_count} overrun' if ov_count > 0 else ''
        header_text = f'{name} ({len(tl)} tickets{refix_label}{bugs_label}{ov_label})'
        cell = ws7.cell(row=r7, column=1, value=header_text)
        cell.font = Font(bold=True, size=11, color='FFFFFF')
        fill = red_fill if d['refix'] > 0 else (orange_fill if d['bugs']['total'] > 0 else purple_fill)
        cell.fill = fill
        for c in range(2, COL_COUNT + 1):
            ws7.cell(row=r7, column=c).fill = fill
        r7 += 1

        # Column headers
        write_header(ws7, r7, ['Ticket', 'Title', 'Status', 'Priority', 'Module', 'Dev Est', 'Dev Actual',
                                'Overrun', 'Refix', 'Fail Cycles', 'Bugs', 'Open', 'Rel. to QA', 'Closed',
                                'QC Tester', 'ETA'], blue_fill)
        r7 += 1

        # Sort: refix tickets first, then by status workflow
        sorted_tickets = sorted(tl, key=lambda t: (0 if t.get('qc_tester') else 1, status_order.get(t['status'], 99), t['ticket_id']))

        for t in sorted_tickets:
            est = t.get('dev_estimate_hours') or 0; act = t.get('actual_dev_hours') or 0
            ov = round(act - est, 1) if est > 0 and act > est else 0
            b = get_bugs(t['ticket_id'])
            is_refix = 'Yes' if t.get('qc_tester') else ''
            ct = cycle_tracker.get(str(t['ticket_id']), {})
            cycles = ct.get('cycle_count', 0)
            write_row(ws7, r7, [t['ticket_id'], t['title'], t['status'], t['priority'],
                                t.get('module', '-'), est, act, ov if ov > 0 else '',
                                is_refix, cycles if cycles > 0 else '',
                                b['total'] if b['total'] > 0 else '', b['open'] if b['open'] > 0 else '',
                                b.get('released_to_qa', 0) if b.get('released_to_qa', 0) > 0 else '',
                                b['closed'] if b['closed'] > 0 else '',
                                t.get('qc_tester', '-'), t.get('eta', '-')])
            if ov > 0: ws7.cell(row=r7, column=8).fill = light_red
            if is_refix: ws7.cell(row=r7, column=9).fill = light_red
            if cycles > 0: ws7.cell(row=r7, column=10).fill = light_red
            if b['open'] > 0: ws7.cell(row=r7, column=12).fill = light_red
            r7 += 1

        r7 += 1  # Blank row between developers

    return wb


if __name__ == '__main__':
    import sys
    os.makedirs(REPORTS_DIR, exist_ok=True)

    # Check for custom dates from command line args
    custom_start = None
    custom_end = None
    for arg in sys.argv[1:]:
        if arg.startswith('--start='):
            custom_start = date.fromisoformat(arg.split('=')[1])
        elif arg.startswith('--end='):
            custom_end = date.fromisoformat(arg.split('=')[1])

    if custom_start and custom_end:
        days_diff = (custom_end - custom_start).days
        prev_start = custom_start - timedelta(days=days_diff)
        prev_end = custom_start
        wb = generate_dev_report(custom_start, custom_end, prev_start, prev_end, 'Custom')
        path = os.path.join(REPORTS_DIR, f'Dev_Report_{today.strftime("%Y%m%d")}.xlsx')
        wb.save(path)
        print(f'Custom: {path}')
    else:
        # All time: use a very old start date to capture everything
        all_start = date(2020, 1, 1)
        prev_half = today - timedelta(days=((today - all_start).days // 2))
        wb = generate_dev_report(all_start, today, all_start, prev_half, 'All Time')
        path = os.path.join(REPORTS_DIR, f'Dev_Report_{today.strftime("%Y%m%d")}.xlsx')
        wb.save(path)
        print(f'All Time: {path}')
