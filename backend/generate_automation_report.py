"""Generate Automation Utilization Report — weekly Excel showing how automated test cases are utilized."""
import json, os, sys, re, time, base64, logging
from datetime import date, timedelta, datetime
from collections import defaultdict

try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(__file__), '.env'))
except ImportError:
    pass

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

REPORTS_DIR = os.path.join(os.path.dirname(__file__), 'reports')
PM_CACHE = os.path.join(os.path.dirname(__file__), 'data', 'pm_tickets_cache.json')
MODULE_CACHE = os.path.join(os.path.dirname(__file__), 'data', 'testrail_module_cache.json')
TESTRAIL_URL = os.environ.get('TESTRAIL_URL', 'https://bistrainer.testrail.io')
TESTRAIL_EMAIL = os.environ.get('TESTRAIL_EMAIL', '')
TESTRAIL_API_KEY = os.environ.get('TESTRAIL_API_KEY', '')
PROJECT_ID = int(os.environ.get('TESTRAIL_AUTOMATION_PROJECT_ID', '18'))

API_BASE = f'{TESTRAIL_URL}/index.php?/api/v2'
CRED = base64.b64encode(f'{TESTRAIL_EMAIL}:{TESTRAIL_API_KEY}'.encode()).decode()
HEADERS = {'Authorization': f'Basic {CRED}', 'Content-Type': 'application/json'}

# Parse args
start_date = None
end_date = None
for arg in sys.argv[1:]:
    if arg.startswith('--start='):
        start_date = date.fromisoformat(arg.split('=', 1)[1])
    elif arg.startswith('--end='):
        end_date = date.fromisoformat(arg.split('=', 1)[1])

today = date.today()
if not start_date or not end_date:
    # Default: past 7 days
    end_date = today
    start_date = today - timedelta(days=7)

prev_start = start_date - timedelta(days=(end_date - start_date).days)
prev_end = start_date

logger.info(f'Automation Report: {start_date} to {end_date}')


def tr_get(endpoint, params=None):
    """TestRail API GET with rate limiting."""
    try:
        resp = requests.get(f'{API_BASE}{endpoint}', headers=HEADERS, params=params or {}, timeout=30)
        if resp.status_code == 200:
            return resp.json()
        logger.warning(f'TestRail {endpoint}: status {resp.status_code}')
    except Exception as e:
        logger.error(f'TestRail {endpoint}: {e}')
    return None


def fetch_all_paginated(endpoint, key, params=None):
    """Fetch all items with pagination."""
    items = []
    offset = 0
    base_params = params or {}
    while True:
        p = {**base_params, 'limit': 250, 'offset': offset}
        data = tr_get(endpoint, p)
        if not data:
            break
        batch = data.get(key, data) if isinstance(data, dict) else data
        if not batch:
            break
        items.extend(batch)
        offset += len(batch)
        if len(batch) < 250:
            break
        time.sleep(0.3)
    return items


# ===== FETCH DATA =====

# 1. Fetch all plans from Project 18 (plans = ticket-based test executions)
start_ts = int(datetime.combine(start_date, datetime.min.time()).timestamp())
end_ts = int(datetime.combine(end_date, datetime.max.time()).timestamp())
prev_start_ts = int(datetime.combine(prev_start, datetime.min.time()).timestamp())
prev_end_ts = int(datetime.combine(prev_end, datetime.max.time()).timestamp())

logger.info('Fetching test plans from Project 18...')
all_plans = fetch_all_paginated(f'/get_plans/{PROJECT_ID}', 'plans')
logger.info(f'Total plans: {len(all_plans)}')

# Also fetch standalone runs
all_standalone_runs = fetch_all_paginated(f'/get_runs/{PROJECT_ID}', 'runs')
logger.info(f'Total standalone runs: {len(all_standalone_runs)}')

# Filter plans by date
period_plans = [p for p in all_plans if p.get('created_on') and start_ts <= p['created_on'] <= end_ts]
prev_plans = [p for p in all_plans if p.get('created_on') and prev_start_ts <= p['created_on'] <= prev_end_ts]
period_runs = [r for r in all_standalone_runs if r.get('created_on') and start_ts <= r['created_on'] <= end_ts]

logger.info(f'Plans in period: {len(period_plans)}, previous: {len(prev_plans)}, standalone runs: {len(period_runs)}')


def build_run_details_from_plans(plans_list, runs_list=None):
    """Build run details from plans (which have summary counts) + optional standalone runs."""
    run_details = []
    all_tests = []

    for p in plans_list:
        name = str(p.get('name', '')).strip()
        match = re.match(r'^(\d+)', name)
        ticket_id = int(match.group(1)) if match and int(match.group(1)) > 100 else None

        passed = p.get('passed_count', 0) or 0
        failed = p.get('failed_count', 0) or 0
        blocked = p.get('blocked_count', 0) or 0
        retest_count = p.get('retest_count', 0) or 0
        untested = p.get('untested_count', 0) or 0
        total = passed + failed + blocked + retest_count + untested

        run_details.append({
            'run_id': p['id'],
            'name': name,
            'ticket_id': ticket_id,
            'type': 'plan',
            'created_on': datetime.fromtimestamp(p['created_on']).strftime('%Y-%m-%d') if p.get('created_on') else '',
            'total': total,
            'passed': passed,
            'failed': failed,
            'blocked': blocked,
            'retest': retest_count,
            'untested': untested,
            'pass_rate': round(passed / total * 100, 1) if total else 0,
        })

        # Build test-level entries from plan counts (no individual test fetch needed for summary)
        for _ in range(passed):
            all_tests.append({'status_id': 1, '_run_name': name, '_ticket_id': ticket_id, '_plan_id': p['id']})
        for _ in range(failed):
            all_tests.append({'status_id': 5, '_run_name': name, '_ticket_id': ticket_id, '_plan_id': p['id']})
        for _ in range(retest_count):
            all_tests.append({'status_id': 4, '_run_name': name, '_ticket_id': ticket_id, '_plan_id': p['id']})
        for _ in range(blocked):
            all_tests.append({'status_id': 2, '_run_name': name, '_ticket_id': ticket_id, '_plan_id': p['id']})

    # Add standalone runs
    for r in (runs_list or []):
        name = str(r.get('name', '')).strip()
        match = re.match(r'^(\d+)', name)
        ticket_id = int(match.group(1)) if match and int(match.group(1)) > 100 else None
        passed = r.get('passed_count', 0) or 0
        failed = r.get('failed_count', 0) or 0
        untested = r.get('untested_count', 0) or 0
        blocked = r.get('blocked_count', 0) or 0
        retest_count = r.get('retest_count', 0) or 0
        total = passed + failed + blocked + retest_count + untested
        if total == 0:
            continue
        run_details.append({
            'run_id': r['id'], 'name': name, 'ticket_id': ticket_id, 'type': 'run',
            'created_on': datetime.fromtimestamp(r['created_on']).strftime('%Y-%m-%d') if r.get('created_on') else '',
            'total': total, 'passed': passed, 'failed': failed, 'blocked': blocked,
            'retest': retest_count, 'untested': untested,
            'pass_rate': round(passed / total * 100, 1) if total else 0,
        })
        for _ in range(passed):
            all_tests.append({'status_id': 1, '_run_name': name, '_ticket_id': ticket_id})
        for _ in range(failed):
            all_tests.append({'status_id': 5, '_run_name': name, '_ticket_id': ticket_id})

    return run_details, all_tests


logger.info('Building current period data...')
run_details, all_tests = build_run_details_from_plans(period_plans, period_runs)

logger.info('Building previous period data...')
prev_run_details, prev_tests = build_run_details_from_plans(prev_plans)

# 3. Load PM ticket data for QA hours
pm_tickets = {}
if os.path.exists(PM_CACHE):
    with open(PM_CACHE) as f:
        for t in json.load(f):
            pm_tickets[t['ticket_id']] = t

# 4. Load module stats cache for automated case counts
module_stats = {}
if os.path.exists(MODULE_CACHE):
    with open(MODULE_CACHE) as f:
        module_stats = json.load(f)

# ===== COMPUTE METRICS =====

total_automated_cases = sum(m.get('automated', 0) for m in module_stats.values())
total_cases_all = sum(m.get('total_cases', 0) for m in module_stats.values())

# Total cases executed (from plan summary counts)
total_cases_executed = sum(r['total'] for r in run_details)
prev_cases_executed = sum(r['total'] for r in prev_run_details)

total_executed = len(all_tests)
total_passed = sum(1 for t in all_tests if t.get('status_id') == 1)
total_failed = sum(1 for t in all_tests if t.get('status_id') == 5)
total_retest = sum(1 for t in all_tests if t.get('status_id') == 4)
total_blocked = sum(1 for t in all_tests if t.get('status_id') == 2)
total_untested = total_executed - total_passed - total_failed - total_retest - total_blocked

prev_total_executed = len(prev_tests)
prev_total_passed = sum(1 for t in prev_tests if t.get('status_id') == 1)

utilization_pct = round(total_cases_executed / total_automated_cases * 100, 1) if total_automated_cases else 0
pass_rate = round(total_passed / total_executed * 100, 1) if total_executed else 0
fail_rate = round(total_failed / total_executed * 100, 1) if total_executed else 0

# Time saved: sum QA estimate hours from PM tickets linked to runs
time_saved = 0
ticket_ids_in_runs = set(r['ticket_id'] for r in run_details if r['ticket_id'])
for tid in ticket_ids_in_runs:
    pm = pm_tickets.get(tid)
    if pm:
        time_saved += pm.get('qa_estimate_hours', 0) or 0

prev_time_saved = 0
prev_ticket_ids = set(r['ticket_id'] for r in prev_run_details if r['ticket_id'])
for tid in prev_ticket_ids:
    pm = pm_tickets.get(tid)
    if pm:
        prev_time_saved += pm.get('qa_estimate_hours', 0) or 0

# Module-wise breakdown
module_execution = defaultdict(lambda: {'executed': 0, 'passed': 0, 'failed': 0, 'retest': 0, 'untested': 0, 'ticket_ids': set()})
for t in all_tests:
    # Map case to module via section (simplified: use test title or run name)
    tid = t.get('_ticket_id')
    if tid:
        pm = pm_tickets.get(tid)
        mod = pm.get('module', 'Unassigned') if pm else 'Unassigned'
    else:
        mod = 'Unassigned'
    module_execution[mod]['executed'] += 1
    if t.get('status_id') == 1: module_execution[mod]['passed'] += 1
    elif t.get('status_id') == 5: module_execution[mod]['failed'] += 1
    elif t.get('status_id') == 4: module_execution[mod]['retest'] += 1
    else: module_execution[mod]['untested'] += 1
    if tid: module_execution[mod]['ticket_ids'].add(tid)

# Failed/retest cases
failed_cases = [t for t in all_tests if t.get('status_id') in (4, 5)]

# ===== GENERATE EXCEL =====

# Styles
hf = Font(bold=True, color='FFFFFF', size=10)
sf = Font(bold=True, size=13, color='1a1a2e')
df = Font(size=9)
lf = Font(color='0563C1', underline='single', size=9)
tb = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
ca = Alignment(horizontal='center', vertical='center')
la = Alignment(horizontal='left', vertical='center', wrap_text=True)
header_fill = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
green_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
red_fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
amber_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
blue_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
alt_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
teal_fill = PatternFill(start_color='0D47A1', end_color='0D47A1', fill_type='solid')
purple_fill = PatternFill(start_color='6A1B9A', end_color='6A1B9A', fill_type='solid')


def write_header(ws, row, headers, fill=header_fill):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = hf; cell.fill = fill; cell.alignment = ca; cell.border = tb


def write_row(ws, row, vals, is_alt=False):
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = df; cell.border = tb
        cell.alignment = ca if isinstance(v, (int, float)) else la
        if is_alt: cell.fill = alt_fill


wb = Workbook()

# ===== SHEET 1: EXECUTIVE SUMMARY =====
ws = wb.active
ws.title = 'Executive Summary'
ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 18

ws.merge_cells('A1:D1')
ws['A1'].value = f'Automation Utilization Report — {start_date.strftime("%b %d")} to {end_date.strftime("%b %d, %Y")}'
ws['A1'].font = sf

r = 3
write_header(ws, r, ['Metric', 'Current Period', 'Previous Period', 'Change'])
r += 1
metrics = [
    ('Total Automated Cases (TestRail)', total_automated_cases, total_automated_cases, 0),
    ('Total Cases Executed', total_cases_executed, prev_cases_executed, total_cases_executed - prev_cases_executed),
    ('Total Test Executions', total_executed, prev_total_executed, total_executed - prev_total_executed),
    ('Utilization %', f'{utilization_pct}%', f'{round(prev_cases_executed/total_automated_cases*100,1) if total_automated_cases else 0}%', ''),
    ('Passed', total_passed, prev_total_passed, total_passed - prev_total_passed),
    ('Failed', total_failed, sum(1 for t in prev_tests if t.get('status_id') == 5), ''),
    ('Pass Rate', f'{pass_rate}%', f'{round(prev_total_passed/prev_total_executed*100,1) if prev_total_executed else 0}%', ''),
    ('Test Runs Created', len(period_runs), len(prev_runs), len(period_runs) - len(prev_runs)),
    ('', '', '', ''),
    ('TIME SAVED BY AUTOMATION', '', '', ''),
    ('Manual QA Hours Saved (from PM estimates)', round(time_saved, 1), round(prev_time_saved, 1), round(time_saved - prev_time_saved, 1)),
    ('Tickets Covered by Automation', len(ticket_ids_in_runs), len(prev_ticket_ids), len(ticket_ids_in_runs) - len(prev_ticket_ids)),
]
for label, curr, prev, change in metrics:
    write_row(ws, r, [label, curr, prev, change], r % 2 == 0)
    if 'TIME SAVED' in str(label):
        ws.cell(row=r, column=1).font = Font(bold=True, size=10, color='1a1a2e')
    if label == 'Manual QA Hours Saved (from PM estimates)':
        ws.cell(row=r, column=2).fill = green_fill
        ws.cell(row=r, column=2).font = Font(bold=True, size=11, color='1B5E20')
    r += 1

# ===== SHEET 2: MODULE-WISE UTILIZATION =====
ws2 = wb.create_sheet('Module Utilization')
ws2.column_dimensions['A'].width = 25
for c in 'BCDEFGHIJ':
    ws2.column_dimensions[c].width = 14

ws2.merge_cells('A1:J1')
ws2['A1'].value = f'Module-wise Automation Utilization — {start_date.strftime("%b %d")} to {end_date.strftime("%b %d, %Y")}'
ws2['A1'].font = sf

write_header(ws2, 3, ['Module', 'Total Cases', 'Automated', 'Auto %', 'Executed', 'Passed', 'Failed',
                       'Pass Rate', 'Utilization %', 'QA Hours Saved'])
r = 4
for mod_name in sorted(module_stats.keys()):
    ms = module_stats[mod_name]
    me = module_execution.get(mod_name, {})
    auto = ms.get('automated', 0)
    exec_count = me.get('executed', 0)
    passed = me.get('passed', 0)
    failed = me.get('failed', 0)
    util = round(exec_count / auto * 100, 1) if auto else 0
    pr = round(passed / exec_count * 100, 1) if exec_count else 0
    # QA hours saved for this module
    mod_hrs = sum(pm_tickets.get(tid, {}).get('qa_estimate_hours', 0) or 0 for tid in me.get('ticket_ids', set()))
    write_row(ws2, r, [mod_name, ms.get('total_cases', 0), auto,
                       f'{round(auto/ms.get("total_cases",1)*100)}%' if ms.get('total_cases') else '0%',
                       exec_count, passed, failed, f'{pr}%', f'{util}%', round(mod_hrs, 1)], r % 2 == 0)
    if util == 0 and auto > 0:
        ws2.cell(row=r, column=9).fill = red_fill
    elif util < 50:
        ws2.cell(row=r, column=9).fill = amber_fill
    else:
        ws2.cell(row=r, column=9).fill = green_fill
    r += 1

# ===== SHEET 3: TEST RUN DETAILS =====
ws3 = wb.create_sheet('Test Run Details')
ws3.column_dimensions['A'].width = 12
ws3.column_dimensions['B'].width = 45
ws3.column_dimensions['C'].width = 12
ws3.column_dimensions['D'].width = 14
for c in 'EFGHIJK':
    ws3.column_dimensions[c].width = 12

ws3.merge_cells('A1:K1')
ws3['A1'].value = f'Test Runs — {start_date.strftime("%b %d")} to {end_date.strftime("%b %d, %Y")}'
ws3['A1'].font = sf

write_header(ws3, 3, ['Run ID', 'Run Name', 'Ticket ID', 'Created', 'Total', 'Passed', 'Failed',
                       'Blocked', 'Retest', 'Pass Rate', 'QA Hrs Saved'])
r = 4
for rd in sorted(run_details, key=lambda x: -x['total']):
    pm_hrs = pm_tickets.get(rd['ticket_id'], {}).get('qa_estimate_hours', 0) if rd['ticket_id'] else 0
    write_row(ws3, r, [rd['run_id'], rd['name'], rd['ticket_id'] or '-', rd['created_on'],
                       rd['total'], rd['passed'], rd['failed'], rd['blocked'], rd['retest'],
                       f'{rd["pass_rate"]}%', round(pm_hrs or 0, 1)], r % 2 == 0)
    # Link ticket ID
    if rd['ticket_id']:
        cell = ws3.cell(row=r, column=3)
        cell.hyperlink = f'https://www.bissafety.app/pm/tickets#!/{rd["ticket_id"]}'
        cell.font = lf
    # Color pass rate
    pr_cell = ws3.cell(row=r, column=10)
    if rd['pass_rate'] >= 90: pr_cell.fill = green_fill
    elif rd['pass_rate'] >= 70: pr_cell.fill = amber_fill
    else: pr_cell.fill = red_fill
    r += 1

# ===== SHEET 4: FAILED/RETEST CASES =====
ws4 = wb.create_sheet('Failed & Retest Cases')
ws4.column_dimensions['A'].width = 12
ws4.column_dimensions['B'].width = 50
ws4.column_dimensions['C'].width = 30
ws4.column_dimensions['D'].width = 12
ws4.column_dimensions['E'].width = 40

ws4.merge_cells('A1:E1')
ws4['A1'].value = f'Failed & Retest Cases — {start_date.strftime("%b %d")} to {end_date.strftime("%b %d, %Y")}'
ws4['A1'].font = sf

write_header(ws4, 3, ['Case ID', 'Title', 'Section', 'Status', 'Run Name'], purple_fill)
r = 4
status_names = {1: 'Passed', 2: 'Blocked', 3: 'Untested', 4: 'Retest', 5: 'Failed'}
for t in sorted(failed_cases, key=lambda x: x.get('status_id', 0)):
    status = status_names.get(t.get('status_id'), 'Unknown')
    write_row(ws4, r, [t.get('case_id', ''), t.get('title', ''), t.get('section', {}).get('name', '') if isinstance(t.get('section'), dict) else '',
                       status, t.get('_run_name', '')], r % 2 == 0)
    # Link case to TestRail
    if t.get('case_id'):
        cell = ws4.cell(row=r, column=1)
        cell.hyperlink = f'{TESTRAIL_URL}/index.php?/cases/view/{t["case_id"]}'
        cell.font = lf
    # Color status
    if status == 'Failed': ws4.cell(row=r, column=4).fill = red_fill
    elif status == 'Retest': ws4.cell(row=r, column=4).fill = amber_fill
    r += 1

if not failed_cases:
    ws4.cell(row=4, column=1, value='No failed or retest cases this period').font = Font(italic=True, color='666666')

# Freeze headers
for ws_item in [ws, ws2, ws3, ws4]:
    ws_item.freeze_panes = ws_item.cell(row=4, column=1)

# Save
os.makedirs(REPORTS_DIR, exist_ok=True)
out_path = os.path.join(REPORTS_DIR, f'Automation_Utilization_Report_{end_date.strftime("%Y%m%d")}.xlsx')
wb.save(out_path)
logger.info(f'Report saved: {out_path}')
print(f'Report: {out_path}')
