"""
QA Cycle Time Dashboard Excel Generator

Generates a comprehensive Excel file with:
- Executive Dashboard with KPI widgets
- Historical Impact Analysis
- Improvement Tracker
- Ticket Data with embedded formulas
- Platform Analysis
- Methodology reference

All calculated columns use Excel formulas (highlighted in yellow).
"""

from datetime import datetime, timedelta
from pathlib import Path
from collections import defaultdict, Counter
from typing import List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule

# Status definitions - tracking all existing PM statuses
# QA-related statuses (tickets actively in QA)
QA_STATUSES = {'QC Testing', 'QC Testing in Progress', 'QC Review Fail', 'QC Testing On-hold', 'QC Testing Hold', 'Tested - Awaiting Fixes'}
QA_START_STATUSES = {'QC Testing', 'QC Testing in Progress'}  # When ticket enters QA
QA_END_STATUSES = {'BIS Testing', 'Closed', 'Approved for Live', 'Moved to Live'}  # When ticket exits QA successfully
QA_HOLD_STATUSES = {'QC Testing On-hold', 'QC Testing Hold', 'Hold/Pending'}  # Hold statuses (time not counted)
QA_FAIL_STATUSES = {'QC Review Fail', 'Tested - Awaiting Fixes'}  # Failed QA review

# Colors
DARK_BLUE = '1F4E78'
LIGHT_BLUE = 'D6EAF8'
WHITE = 'FFFFFF'
GRAY = '95A5A6'
GREEN = '27AE60'
LIGHT_GREEN = 'D5F5E3'
YELLOW = 'F7DC6F'
LIGHT_YELLOW = 'FCF3CF'
ORANGE = 'E67E22'
RED = 'E74C3C'
LIGHT_RED = 'FADBD8'
PURPLE = '9B59B6'
LIGHT_PURPLE = 'E8DAEF'

BASELINE_DATE = datetime(2026, 2, 24)


def generate_qa_dashboard_excel(
    ticket_history: Dict[int, List[Dict]],
    ticket_lookup: Dict[int, Any],
    output_path: Path = None
) -> Path:
    """
    Generate comprehensive QA Dashboard Excel with formulas.
    
    Args:
        ticket_history: Dict mapping ticket_id to list of status changes
        ticket_lookup: Dict mapping ticket_id to ticket metadata
        output_path: Optional output path, defaults to reports directory
    
    Returns:
        Path to generated Excel file
    """
    
    # Calculate metrics
    results = []
    status_counts = Counter()
    monthly_qa_times = defaultdict(list)
    weekly_qa_times = defaultdict(list)
    platform_qa_times = defaultdict(list)
    priority_qa_times = defaultdict(list)
    cycle_distribution = defaultdict(list)
    
    for ticket_id, history in ticket_history.items():
        if not history:
            continue
            
        first_created = history[0]['date']
        current_status = history[-1]['new_status'] if history else None
        status_counts[current_status] += 1
        
        qa_start = None
        qa_end = None
        qa_cycles = 0
        fail_count = 0
        total_hold_hours = 0.0
        hold_start = None
        
        for h in history:
            if h['new_status'] in QA_START_STATUSES:
                if qa_start is None:
                    qa_start = h['date']
                qa_cycles += 1
            
            if h['new_status'] in QA_END_STATUSES and qa_start is not None:
                qa_end = h['date']
            
            if h['new_status'] in QA_HOLD_STATUSES:
                hold_start = h['date']
            elif hold_start is not None and h['new_status'] not in QA_HOLD_STATUSES:
                hold_duration = (h['date'] - hold_start).total_seconds() / 3600
                total_hold_hours += hold_duration
                hold_start = None
            
            if h['new_status'] in QA_FAIL_STATUSES:
                fail_count += 1
        
        qa_business_days = None
        if qa_start and qa_end:
            gross_hours = (qa_end - qa_start).total_seconds() / 3600
            net_hours = max(gross_hours - total_hold_hours, 0)
            qa_business_days = net_hours / 8
            
            month_key = qa_end.strftime('%Y-%m')
            monthly_qa_times[month_key].append(qa_business_days)
            
            week_key = qa_end.strftime('%Y-W%W')
            weekly_qa_times[week_key].append(qa_business_days)
            
            ticket = ticket_lookup.get(ticket_id)
            if ticket:
                platform = getattr(ticket, 'subdepartment', None) or 'Unknown'
                priority = getattr(ticket, 'priority', None) or 'Unknown'
                platform_qa_times[platform].append(qa_business_days)
                priority_qa_times[priority].append(qa_business_days)
            
            cycle_key = '1 cycle' if qa_cycles == 1 else ('2 cycles' if qa_cycles == 2 else '3+ cycles')
            cycle_distribution[cycle_key].append(qa_business_days)
        
        ticket = ticket_lookup.get(ticket_id)
        results.append({
            'ticket_id': ticket_id,
            'title': getattr(ticket, 'title', '') if ticket else '',
            'current_status': current_status,
            'priority': getattr(ticket, 'priority', '') if ticket else '',
            'subdepartment': getattr(ticket, 'subdepartment', '') if ticket else '',
            'qc_tester': getattr(ticket, 'qc_tester', '') if ticket else '',
            'backend_dev': getattr(ticket, 'backend_developer', '') if ticket else '',
            'frontend_dev': getattr(ticket, 'frontend_developer', '') if ticket else '',
            'first_activity': first_created,
            'last_activity': history[-1]['date'] if history else None,
            'total_status_changes': len(history),
            'qa_start': qa_start,
            'qa_end': qa_end,
            'qa_hold_hours': round(total_hold_hours, 2),
            'qa_business_days': round(qa_business_days, 2) if qa_business_days else None,
            'qa_cycles': qa_cycles,
            'qa_fail_count': fail_count,
        })
    
    results.sort(key=lambda x: x['ticket_id'], reverse=True)
    
    # Stats
    qa_completed = [r for r in results if r['qa_business_days'] is not None]
    avg_qa_days = sum(r['qa_business_days'] for r in qa_completed) / len(qa_completed) if qa_completed else 0
    sorted_days = sorted([r['qa_business_days'] for r in qa_completed])
    median_qa_days = sorted_days[len(sorted_days)//2] if sorted_days else 0
    total_fails = sum(r['qa_fail_count'] for r in results)
    avg_cycles = sum(r['qa_cycles'] for r in results) / len(results) if results else 0
    first_pass = len([r for r in qa_completed if r['qa_cycles'] == 1])
    first_pass_rate = (first_pass / len(qa_completed) * 100) if qa_completed else 0
    
    # Styles
    header_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')
    header_font = Font(color=WHITE, bold=True, size=11)
    formula_fill = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    wb = Workbook()
    
    # ===== SHEET 1: EXECUTIVE DASHBOARD =====
    ws_dash = wb.active
    ws_dash.title = 'Executive Dashboard'
    
    ws_dash.merge_cells('B2:P2')
    ws_dash['B2'] = 'QA PERFORMANCE DASHBOARD'
    ws_dash['B2'].font = Font(bold=True, size=24, color=DARK_BLUE)
    ws_dash['B2'].alignment = Alignment(horizontal='center')
    
    ws_dash.merge_cells('B3:P3')
    ws_dash['B3'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")} | Baseline: {BASELINE_DATE.strftime("%Y-%m-%d")}'
    ws_dash['B3'].font = Font(italic=True, size=10, color=GRAY)
    ws_dash['B3'].alignment = Alignment(horizontal='center')
    
    # Widgets Row 1
    widget_row = 5
    widgets = [
        ('B', len(results), 'Total Tickets', LIGHT_BLUE, DARK_BLUE),
        ('D', len(qa_completed), 'QA Completed', LIGHT_GREEN, GREEN),
        ('F', round(avg_qa_days, 1), 'Avg QA Days', LIGHT_YELLOW, ORANGE),
        ('H', round(median_qa_days, 1), 'Median Days', LIGHT_PURPLE, PURPLE),
        ('J', f"{round(first_pass_rate, 1)}%", 'First Pass Rate', LIGHT_GREEN, GREEN),
        ('L', round(avg_cycles, 2), 'Avg Cycles', LIGHT_BLUE, DARK_BLUE),
    ]
    
    for col, value, label, bg_color, text_color in widgets:
        ws_dash.merge_cells(f'{col}{widget_row}:{chr(ord(col)+1)}{widget_row+2}')
        cell = ws_dash[f'{col}{widget_row}']
        cell.value = value
        cell.font = Font(bold=True, size=24, color=text_color)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.border = thin_border
        ws_dash[f'{col}{widget_row+3}'] = label
        ws_dash[f'{col}{widget_row+3}'].font = Font(size=10, color=GRAY)
        ws_dash[f'{col}{widget_row+3}'].alignment = Alignment(horizontal='center')
    
    # Widgets Row 2
    widget_row2 = 10
    widgets2 = [
        ('B', total_fails, 'Total QA Fails', LIGHT_RED, RED),
        ('D', status_counts.get('QC Testing', 0) + status_counts.get('QC Testing in Progress', 0), 'In QA Now', LIGHT_YELLOW, ORANGE),
        ('F', round(sum(r['qa_hold_hours'] for r in results if r['qa_hold_hours']), 1), 'Total Hold Hours', LIGHT_BLUE, DARK_BLUE),
    ]
    
    for col, value, label, bg_color, text_color in widgets2:
        ws_dash.merge_cells(f'{col}{widget_row2}:{chr(ord(col)+1)}{widget_row2+2}')
        cell = ws_dash[f'{col}{widget_row2}']
        cell.value = value
        cell.font = Font(bold=True, size=24, color=text_color)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.border = thin_border
        ws_dash[f'{col}{widget_row2+3}'] = label
        ws_dash[f'{col}{widget_row2+3}'].font = Font(size=10, color=GRAY)
        ws_dash[f'{col}{widget_row2+3}'].alignment = Alignment(horizontal='center')
    
    # Platform Table
    plat_row = 15
    ws_dash[f'B{plat_row}'] = 'QA TIME BY PLATFORM'
    ws_dash[f'B{plat_row}'].font = Font(bold=True, size=14, color=DARK_BLUE)
    
    for col, h in enumerate(['Platform', 'Tickets', 'Avg Days', 'Total Days'], 2):
        ws_dash.cell(row=plat_row+1, column=col, value=h).font = header_font
        ws_dash.cell(row=plat_row+1, column=col).fill = header_fill
    
    for idx, (plat, times) in enumerate(sorted(platform_qa_times.items(), key=lambda x: -len(x[1]))[:8]):
        row = plat_row + 2 + idx
        ws_dash.cell(row=row, column=2, value=plat or 'Unknown')
        ws_dash.cell(row=row, column=3, value=len(times))
        ws_dash.cell(row=row, column=4, value=round(sum(times)/len(times), 1) if times else 0)
        ws_dash.cell(row=row, column=5, value=round(sum(times), 1))
    
    # Priority Table
    prio_row = 15
    ws_dash[f'G{prio_row}'] = 'QA TIME BY PRIORITY'
    ws_dash[f'G{prio_row}'].font = Font(bold=True, size=14, color=DARK_BLUE)
    
    for col, h in enumerate(['Priority', 'Tickets', 'Avg Days', 'Total Days'], 7):
        ws_dash.cell(row=prio_row+1, column=col, value=h).font = header_font
        ws_dash.cell(row=prio_row+1, column=col).fill = header_fill
    
    priority_order = ['URGENT', 'High (Bugs)', 'High', 'Medium', 'Low']
    sorted_priorities = sorted(priority_qa_times.items(), 
                               key=lambda x: priority_order.index(x[0]) if x[0] in priority_order else 99)
    
    for idx, (prio, times) in enumerate(sorted_priorities[:8]):
        row = prio_row + 2 + idx
        ws_dash.cell(row=row, column=7, value=prio or 'Unknown')
        ws_dash.cell(row=row, column=8, value=len(times))
        ws_dash.cell(row=row, column=9, value=round(sum(times)/len(times), 1) if times else 0)
        ws_dash.cell(row=row, column=10, value=round(sum(times), 1))
    
    # Reduction Targets with formulas
    target_row = 26
    ws_dash[f'B{target_row}'] = 'CYCLE TIME REDUCTION TARGETS'
    ws_dash[f'B{target_row}'].font = Font(bold=True, size=14, color=DARK_BLUE)
    
    for col, h in enumerate(['Target', 'Days', 'Reduction', 'Status'], 2):
        ws_dash.cell(row=target_row+1, column=col, value=h).font = header_font
        ws_dash.cell(row=target_row+1, column=col).fill = header_fill
    
    baseline_cell = f'C{target_row+2}'
    ws_dash.cell(row=target_row+2, column=2, value='Current Baseline')
    ws_dash.cell(row=target_row+2, column=3, value=round(avg_qa_days, 2))
    ws_dash.cell(row=target_row+2, column=4, value='-')
    ws_dash.cell(row=target_row+2, column=5, value='CURRENT')
    ws_dash[baseline_cell].fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type='solid')
    
    for idx, (label, mult, status) in enumerate([
        ('10% Reduction', 0.9, 'TARGET'),
        ('20% Reduction', 0.8, 'TARGET'),
        ('30% Reduction (Stretch)', 0.7, 'STRETCH'),
        ('50% Reduction (Aspirational)', 0.5, 'ASPIRATIONAL'),
    ]):
        row = target_row + 3 + idx
        ws_dash.cell(row=row, column=2, value=label)
        ws_dash.cell(row=row, column=3, value=f'={baseline_cell}*{mult}').fill = formula_fill
        ws_dash.cell(row=row, column=4, value=f'={baseline_cell}-C{row}').fill = formula_fill
        ws_dash.cell(row=row, column=5, value=status)
    
    # Status Distribution
    stat_row = 26
    ws_dash[f'G{stat_row}'] = 'STATUS DISTRIBUTION'
    ws_dash[f'G{stat_row}'].font = Font(bold=True, size=14, color=DARK_BLUE)
    
    ws_dash.cell(row=stat_row+1, column=7, value='Status').font = header_font
    ws_dash.cell(row=stat_row+1, column=7).fill = header_fill
    ws_dash.cell(row=stat_row+1, column=8, value='Count').font = header_font
    ws_dash.cell(row=stat_row+1, column=8).fill = header_fill
    
    top_statuses = status_counts.most_common(8)
    for idx, (status, count) in enumerate(top_statuses):
        ws_dash.cell(row=stat_row+2+idx, column=7, value=status or 'Unknown')
        ws_dash.cell(row=stat_row+2+idx, column=8, value=count)
    
    # Pie Chart
    if top_statuses:
        pie = PieChart()
        pie.title = "Status Distribution"
        labels = Reference(ws_dash, min_col=7, min_row=stat_row+2, max_row=stat_row+1+len(top_statuses))
        data = Reference(ws_dash, min_col=8, min_row=stat_row+1, max_row=stat_row+1+len(top_statuses))
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.width = 10
        pie.height = 7
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        ws_dash.add_chart(pie, "L15")
    
    # Monthly Trend
    trend_row = 38
    ws_dash[f'B{trend_row}'] = 'MONTHLY QA CYCLE TIME TREND'
    ws_dash[f'B{trend_row}'].font = Font(bold=True, size=14, color=DARK_BLUE)
    
    for col, h in enumerate(['Month', 'Avg Days', 'Tickets'], 2):
        ws_dash.cell(row=trend_row+1, column=col, value=h).font = header_font
        ws_dash.cell(row=trend_row+1, column=col).fill = header_fill
    
    monthly_sorted = sorted(monthly_qa_times.items())[-12:]
    for idx, (month, times) in enumerate(monthly_sorted):
        ws_dash.cell(row=trend_row+2+idx, column=2, value=month)
        ws_dash.cell(row=trend_row+2+idx, column=3, value=round(sum(times)/len(times), 1) if times else 0)
        ws_dash.cell(row=trend_row+2+idx, column=4, value=len(times))
    
    # Line Chart
    if monthly_sorted:
        line = LineChart()
        line.title = "QA Cycle Time Trend"
        line.y_axis.title = "Days"
        data = Reference(ws_dash, min_col=3, min_row=trend_row+1, max_row=trend_row+1+len(monthly_sorted))
        cats = Reference(ws_dash, min_col=2, min_row=trend_row+2, max_row=trend_row+1+len(monthly_sorted))
        line.add_data(data, titles_from_data=True)
        line.set_categories(cats)
        line.width = 14
        line.height = 8
        ws_dash.add_chart(line, "F38")
    
    for col in range(1, 18):
        ws_dash.column_dimensions[get_column_letter(col)].width = 12
    
    # ===== SHEET 2: HISTORICAL IMPACT =====
    ws_impact = wb.create_sheet('Historical Impact')
    
    ws_impact['B2'] = 'HISTORICAL IMPACT: HOW REWORK AFFECTED QA TIMELINES'
    ws_impact['B2'].font = Font(bold=True, size=18, color=DARK_BLUE)
    ws_impact.merge_cells('B2:I2')
    
    ws_impact['B4'] = 'QA CYCLE TIME BY REWORK COUNT'
    ws_impact['B4'].font = Font(bold=True, size=14, color=DARK_BLUE)
    
    for col, h in enumerate(['QA Cycles', 'Tickets', 'Avg Days', 'Min', 'Max', 'Total Days', '% Tickets', '% Time'], 2):
        ws_impact.cell(row=5, column=col, value=h).font = header_font
        ws_impact.cell(row=5, column=col).fill = header_fill
    
    total_completed = len(qa_completed)
    total_qa_time = sum(r['qa_business_days'] for r in qa_completed)
    
    cycle_data = []
    for cycle_key in ['1 cycle', '2 cycles', '3+ cycles']:
        times = cycle_distribution.get(cycle_key, [])
        if times:
            cycle_data.append({
                'label': cycle_key,
                'count': len(times),
                'avg': sum(times) / len(times),
                'min': min(times),
                'max': max(times),
                'total': sum(times),
                'pct_tickets': len(times) / total_completed * 100 if total_completed else 0,
                'pct_time': sum(times) / total_qa_time * 100 if total_qa_time else 0,
            })
    
    for idx, cd in enumerate(cycle_data):
        row = 6 + idx
        ws_impact.cell(row=row, column=2, value=cd['label'])
        ws_impact.cell(row=row, column=3, value=cd['count'])
        ws_impact.cell(row=row, column=4, value=round(cd['avg'], 1))
        ws_impact.cell(row=row, column=5, value=round(cd['min'], 1))
        ws_impact.cell(row=row, column=6, value=round(cd['max'], 1))
        ws_impact.cell(row=row, column=7, value=round(cd['total'], 1))
        ws_impact.cell(row=row, column=8, value=f"{round(cd['pct_tickets'], 1)}%")
        ws_impact.cell(row=row, column=9, value=f"{round(cd['pct_time'], 1)}%")
    
    # Insights
    insight_row = 11
    ws_impact[f'B{insight_row}'] = 'KEY INSIGHTS'
    ws_impact[f'B{insight_row}'].font = Font(bold=True, size=14, color=DARK_BLUE)
    
    if len(cycle_data) >= 2:
        ratio = cycle_data[1]['avg'] / cycle_data[0]['avg'] if cycle_data[0]['avg'] > 0 else 0
        ws_impact[f'B{insight_row+1}'] = f"• Tickets with 2 cycles take {round(ratio, 1)}x longer than first-pass tickets"
        ws_impact[f'B{insight_row+1}'].font = Font(size=12)
    
    if len(cycle_data) >= 3:
        ws_impact[f'B{insight_row+2}'] = f"• {round(cycle_data[2]['pct_tickets'], 1)}% of tickets (3+ cycles) consume {round(cycle_data[2]['pct_time'], 1)}% of QA time"
        ws_impact[f'B{insight_row+2}'].font = Font(size=12)
    
    if cycle_data:
        baseline_avg = cycle_data[0]['avg']
        extra_days = sum((cd['avg'] - baseline_avg) * cd['count'] for cd in cycle_data[1:])
        ws_impact[f'B{insight_row+3}'] = f"• Total extra days lost to rework: {round(extra_days, 0)} days"
        ws_impact[f'B{insight_row+3}'].font = Font(size=12, color=RED)
    
    # Bar Chart
    if cycle_data:
        bar = BarChart()
        bar.title = "Avg QA Days by Cycle Count"
        bar.y_axis.title = "Days"
        data = Reference(ws_impact, min_col=4, min_row=5, max_row=5+len(cycle_data))
        cats = Reference(ws_impact, min_col=2, min_row=6, max_row=5+len(cycle_data))
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        bar.width = 10
        bar.height = 7
        ws_impact.add_chart(bar, "K4")
    
    for col in range(1, 12):
        ws_impact.column_dimensions[get_column_letter(col)].width = 14
    
    # ===== SHEET 3: TICKET DATA WITH FORMULAS =====
    ws_data = wb.create_sheet('Ticket Data')
    
    data_headers = ['Ticket ID', 'Title', 'Status', 'Priority', 'Platform', 'QC Tester',
                    'Backend Dev', 'Frontend Dev', 'First Activity', 'Last Activity',
                    'Changes', 'QA Start', 'QA End', 'Hold Hrs', 'Gross Hrs',
                    'Net Hrs', 'Biz Days', 'Cycles', 'Fails']
    
    for col, h in enumerate(data_headers, 1):
        ws_data.cell(row=1, column=col, value=h).font = header_font
        ws_data.cell(row=1, column=col).fill = header_fill
    
    for row_idx, r in enumerate(results, 2):
        ws_data.cell(row=row_idx, column=1, value=r['ticket_id'])
        ws_data.cell(row=row_idx, column=2, value=(r['title'][:40] + '...') if r['title'] and len(r['title']) > 40 else r['title'])
        ws_data.cell(row=row_idx, column=3, value=r['current_status'])
        ws_data.cell(row=row_idx, column=4, value=r['priority'])
        ws_data.cell(row=row_idx, column=5, value=r['subdepartment'])
        ws_data.cell(row=row_idx, column=6, value=r['qc_tester'])
        ws_data.cell(row=row_idx, column=7, value=r['backend_dev'])
        ws_data.cell(row=row_idx, column=8, value=r['frontend_dev'])
        ws_data.cell(row=row_idx, column=9, value=r['first_activity'])
        ws_data.cell(row=row_idx, column=10, value=r['last_activity'])
        ws_data.cell(row=row_idx, column=11, value=r['total_status_changes'])
        ws_data.cell(row=row_idx, column=12, value=r['qa_start'])
        ws_data.cell(row=row_idx, column=13, value=r['qa_end'])
        ws_data.cell(row=row_idx, column=14, value=r['qa_hold_hours'])
        
        # Formula columns (highlighted yellow)
        ws_data.cell(row=row_idx, column=15, value=f'=IF(AND(L{row_idx}<>"",M{row_idx}<>""),(M{row_idx}-L{row_idx})*24,"")').fill = formula_fill
        ws_data.cell(row=row_idx, column=16, value=f'=IF(O{row_idx}<>"",MAX(O{row_idx}-N{row_idx},0),"")').fill = formula_fill
        ws_data.cell(row=row_idx, column=17, value=f'=IF(P{row_idx}<>"",P{row_idx}/8,"")').fill = formula_fill
        
        ws_data.cell(row=row_idx, column=18, value=r['qa_cycles'])
        ws_data.cell(row=row_idx, column=19, value=r['qa_fail_count'])
    
    # Format date columns
    for col in [9, 10, 12, 13]:
        for row_idx in range(2, len(results) + 2):
            cell = ws_data.cell(row=row_idx, column=col)
            if cell.value and not str(cell.value).startswith('='):
                cell.number_format = 'YYYY-MM-DD HH:MM'
    
    ws_data.auto_filter.ref = f"A1:S{len(results)+1}"
    
    # Conditional formatting for Business Days
    if len(results) > 0:
        ws_data.conditional_formatting.add(
            f'Q2:Q{len(results)+1}',
            ColorScaleRule(start_type='min', start_color='27AE60',
                           mid_type='percentile', mid_value=50, mid_color='F7DC6F',
                           end_type='max', end_color='E74C3C')
        )
    
    col_widths = [10, 30, 18, 12, 15, 15, 15, 15, 16, 16, 8, 16, 16, 8, 8, 8, 10, 7, 6]
    for i, width in enumerate(col_widths, 1):
        ws_data.column_dimensions[get_column_letter(i)].width = width
    
    # ===== SHEET 4: FORMULAS & METHODOLOGY =====
    ws_ref = wb.create_sheet('Formulas & Methodology')
    
    ref_content = [
        ['QA CYCLE TIME CALCULATION METHODOLOGY', ''],
        ['', ''],
        ['METRIC', 'FORMULA / DEFINITION'],
        ['QA Gross Hours', '= (QA End - QA Start) * 24'],
        ['QA Net Hours', '= MAX(Gross Hours - Hold Hours, 0)'],
        ['QA Business Days', '= Net Hours / 8'],
        ['', ''],
        ['FORMULA COLUMNS (Yellow in Ticket Data)', ''],
        ['Column O (Gross Hrs)', '=IF(AND(L{row}<>"",M{row}<>""),(M{row}-L{row})*24,"")'],
        ['Column P (Net Hrs)', '=IF(O{row}<>"",MAX(O{row}-N{row},0),"")'],
        ['Column Q (Biz Days)', '=IF(P{row}<>"",P{row}/8,"")'],
        ['', ''],
        ['STATUS DEFINITIONS', ''],
        ['QA-Related Statuses', 'QC Testing, QC Testing in Progress, QC Review Fail, QC Testing On-hold, QC Testing Hold, Tested - Awaiting Fixes'],
        ['QA Start Statuses', 'QC Testing, QC Testing in Progress'],
        ['QA End Statuses', 'BIS Testing, Closed, Approved for Live, Moved to Live'],
        ['QA Hold Statuses', 'QC Testing On-hold, QC Testing Hold, Hold/Pending'],
        ['QA Fail Statuses', 'QC Review Fail, Tested - Awaiting Fixes'],
        ['', ''],
        ['KPI DEFINITIONS', ''],
        ['First Pass Rate', '% of tickets completing QA in exactly 1 cycle'],
        ['QA Cycles', 'Number of times ticket entered QA status'],
        ['Rework Cost', 'Extra days spent on tickets with 2+ cycles'],
        ['', ''],
        ['CLIENT GOALS', ''],
        ['1. Reduce Cycle Time', 'Target: 20-30% reduction from baseline'],
        ['2. Reduce Duplicate Testing', 'Track: QA Cycles, Fail Count, Rework Cost'],
        ['3. Increase Automated Cases', 'Track via TestRail (separate dashboard)'],
        ['4. Increase Automation Utilization', 'Track via TestRail (separate dashboard)'],
    ]
    
    for row_idx, (col1, col2) in enumerate(ref_content, 1):
        ws_ref.cell(row=row_idx, column=1, value=col1)
        ws_ref.cell(row=row_idx, column=2, value=col2)
        if row_idx in [1, 8, 13, 19, 24]:
            ws_ref.cell(row=row_idx, column=1).font = Font(bold=True, size=14, color=DARK_BLUE)
        elif row_idx == 3:
            ws_ref.cell(row=row_idx, column=1).font = header_font
            ws_ref.cell(row=row_idx, column=1).fill = header_fill
            ws_ref.cell(row=row_idx, column=2).font = header_font
            ws_ref.cell(row=row_idx, column=2).fill = header_fill
    
    ws_ref.column_dimensions['A'].width = 35
    ws_ref.column_dimensions['B'].width = 55
    
    # Save
    if output_path is None:
        output_path = Path("reports") / f"QA_Dashboard_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    output_path.parent.mkdir(exist_ok=True)
    wb.save(output_path)
    
    return output_path
