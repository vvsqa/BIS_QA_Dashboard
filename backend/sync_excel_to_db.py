"""
Excel to Database Sync Script for Ticket Tracking Data

This script imports ticket tracking data from Excel files into the database.
It supports:
- One-time import: python sync_excel_to_db.py --file path/to/file.xlsx
- Folder watching: python sync_excel_to_db.py --watch
- Downloads folder watching: python sync_excel_to_db.py --watch-downloads
- Manual trigger via API

Usage:
    python sync_excel_to_db.py --file "path/to/ticket_data.xlsx"
    python sync_excel_to_db.py --watch
    python sync_excel_to_db.py --watch-downloads  # Watch Downloads folder for TicketReport_* files
"""

import os
import sys
import argparse
import time
import re
import shutil
from datetime import datetime
from pathlib import Path

# Fix Unicode output on Windows
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

from database import SessionLocal
from models import TicketTracking, TicketStatusHistory

# Configuration
IMPORTS_FOLDER = os.getenv("IMPORTS_FOLDER", os.path.join(os.path.dirname(__file__), "imports"))

# Get Downloads folder path (works on Windows, Mac, Linux)
def get_downloads_folder():
    """Get the user's Downloads folder path"""
    if sys.platform == 'win32':
        import winreg
        try:
            with winreg.OpenKey(winreg.HKEY_CURRENT_USER, 
                r'SOFTWARE\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders') as key:
                downloads = winreg.QueryValueEx(key, '{374DE290-123F-4565-9164-39C4925E467B}')[0]
                return downloads
        except:
            pass
    # Fallback to standard path
    return os.path.join(Path.home(), 'Downloads')

DOWNLOADS_FOLDER = os.getenv("DOWNLOADS_FOLDER", get_downloads_folder())

# File pattern for TicketReport files
TICKET_REPORT_PATTERN = re.compile(r'^TicketReport_\d{8}_\d{6}\.xlsx$', re.IGNORECASE)

# Column mapping from Excel headers to database fields
# These match the exact column headers from TicketReport Excel exports
COLUMN_MAPPING = {
    # Ticket ID
    "Ticket Number": "ticket_id",
    "Ticket Numb": "ticket_id",
    
    # Status
    "Status": "status",
    
    # Team Members
    "Backend Developer": "backend_developer",
    "Frontend Developer": "frontend_developer",
    "QC Tester": "qc_tester",
    "Current Assignee": "current_assignee",
    "Developer": "developer_assigned",
    
    # ETA
    "ETA": "eta",
    
    # Development Time
    "Development Estimated Time": "dev_estimate_hours",
    "Development Estim": "dev_estimate_hours",
    "Development Estimate": "dev_estimate_hours",
    
    # Actual Development Time
    "Actual Development Spend": "actual_dev_hours",
    "Actual Development": "actual_dev_hours",
    "Actual Development Time": "actual_dev_hours",
    "Actual Dev": "actual_dev_hours",
    
    # QA Estimated Time (labeled as "Other Estimated Time" in Excel)
    "Other Estimated Time": "qa_estimate_hours",
    "Other Estimated": "qa_estimate_hours",
    "QA Estimate": "qa_estimate_hours",
    
    # Actual QA Time
    "Actual QA/QC Spend": "actual_qa_hours",
    "Actual QA": "actual_qa_hours",
}


def parse_float(value):
    """Parse a value to float, handling various formats"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value) if value != 0 else None
    if isinstance(value, str):
        value = value.strip()
        if not value or value.lower() in ('', 'n/a', 'na', '-', 'not assigned'):
            return None
        try:
            # Remove any commas or spaces
            value = value.replace(',', '').replace(' ', '')
            return float(value)
        except ValueError:
            return None
    return None


def parse_datetime_value(value):
    """Parse datetime from Excel cell"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        # Try common date formats
        formats = [
            "%Y-%m-%d",
            "%d-%m-%Y",
            "%m/%d/%Y",
            "%d/%m/%Y",
            "%Y-%m-%d %H:%M:%S",
        ]
        for fmt in formats:
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                continue
    return None


def parse_string(value):
    """Parse string value, returning None for empty/placeholder values"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return str(int(value)) if isinstance(value, float) and value.is_integer() else str(value)
    value = str(value).strip()
    if value.lower() in ('', 'n/a', 'na', '-', 'not assigned', 'none'):
        return None
    return value


def parse_ticket_id(value):
    """Parse ticket ID to integer"""
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    if isinstance(value, str):
        value = value.strip()
        try:
            return int(float(value))
        except ValueError:
            return None
    return None


def find_header_row(ws):
    """Find the header row in the worksheet"""
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        # Check if this row looks like a header row
        row_values = [str(cell).strip() if cell else "" for cell in row]
        if any("Ticket" in val for val in row_values) or any("Status" in val for val in row_values):
            return row_idx
    return 1  # Default to first row


def map_headers(header_row):
    """Map Excel headers to database field names"""
    column_map = {}
    for idx, header in enumerate(header_row):
        if header is None:
            continue
        header_str = str(header).strip()
        # Try to find matching field
        for excel_header, db_field in COLUMN_MAPPING.items():
            if excel_header.lower() in header_str.lower() or header_str.lower() in excel_header.lower():
                column_map[idx] = db_field
                break
    return column_map


def import_excel_file(filepath):
    """Import data from an Excel file into the database"""
    print(f"\n{'='*60}")
    print(f"Importing: {filepath}")
    print(f"Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*60}")
    
    if not os.path.exists(filepath):
        print(f"ERROR: File not found: {filepath}")
        return False, 0, 0
    
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        
        # Find header row
        header_row_idx = find_header_row(ws)
        print(f"Header row found at row: {header_row_idx}")
        
        # Get header row
        header_row = []
        for cell in ws[header_row_idx]:
            header_row.append(cell.value)
        
        # Map headers to database fields
        column_map = map_headers(header_row)
        print(f"Mapped columns: {column_map}")
        
        if 'ticket_id' not in column_map.values():
            print("ERROR: Could not find Ticket ID column in the Excel file")
            return False, 0, 0
        
        # Find ticket_id column index
        ticket_id_col = None
        for idx, field in column_map.items():
            if field == 'ticket_id':
                ticket_id_col = idx
                break
        
        db = SessionLocal()
        imported = 0
        updated = 0
        skipped = 0
        
        try:
            # Iterate through data rows
            for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
                # Get ticket_id first
                if ticket_id_col is None or ticket_id_col >= len(row):
                    continue
                    
                ticket_id = parse_ticket_id(row[ticket_id_col])
                if ticket_id is None:
                    skipped += 1
                    continue
                
                # Check if record exists
                existing = db.query(TicketTracking).filter(TicketTracking.ticket_id == ticket_id).first()
                
                if existing:
                    record = existing
                    updated += 1
                else:
                    record = TicketTracking()
                    record.ticket_id = ticket_id
                    imported += 1
                
                # Capture previous status for history tracking
                previous_status = existing.status if existing else None
                previous_updated_on = existing.updated_on if existing else None
                
                # Build a dict of new values first
                new_values = {}
                for col_idx, db_field in column_map.items():
                    if col_idx >= len(row):
                        continue
                    
                    value = row[col_idx]
                    
                    if db_field == 'ticket_id':
                        continue  # Already handled
                    elif db_field == 'eta':
                        new_values[db_field] = parse_datetime_value(value)
                    elif db_field in ('dev_estimate_hours', 'actual_dev_hours', 'qa_estimate_hours', 'actual_qa_hours'):
                        new_values[db_field] = parse_float(value)
                    else:
                        new_values[db_field] = parse_string(value)
                
                # Apply new values to record
                for db_field, value in new_values.items():
                    setattr(record, db_field, value)
                
                record.updated_on = datetime.now()
                
                # Track status change if status has changed
                new_status = new_values.get('status')
                if new_status and new_status != previous_status:
                    # Calculate duration in previous status (if we have previous data)
                    duration_hours = None
                    if previous_status and previous_updated_on:
                        duration_seconds = (datetime.now() - previous_updated_on).total_seconds()
                        duration_hours = round(duration_seconds / 3600, 2)
                    
                    # Create status history record
                    history = TicketStatusHistory(
                        ticket_id=ticket_id,
                        previous_status=previous_status,
                        new_status=new_status,
                        changed_on=datetime.now(),
                        current_assignee=new_values.get('current_assignee'),
                        qc_tester=new_values.get('qc_tester'),
                        duration_in_previous_status=duration_hours,
                        source='sync'
                    )
                    db.add(history)
                
                if not existing:
                    db.add(record)
                
                # Commit every 100 records
                if (imported + updated) % 100 == 0:
                    db.commit()
                    print(f"  Progress: {imported} imported, {updated} updated...")
            
            db.commit()
            print(f"\nImport completed successfully!")
            print(f"  New records: {imported}")
            print(f"  Updated records: {updated}")
            print(f"  Skipped rows: {skipped}")
            
            wb.close()
            return True, imported, updated
            
        except Exception as e:
            db.rollback()
            print(f"ERROR during import: {e}")
            import traceback
            traceback.print_exc()
            return False, 0, 0
        finally:
            db.close()
            
    except Exception as e:
        print(f"ERROR opening Excel file: {e}")
        import traceback
        traceback.print_exc()
        return False, 0, 0


class ExcelFileHandler(FileSystemEventHandler):
    """Handler for file system events on Excel files"""
    
    def __init__(self, file_pattern=None, copy_to_imports=False):
        self.last_modified = {}
        self.debounce_seconds = 3  # Wait 3 seconds before processing
        self.file_pattern = file_pattern  # Optional regex pattern to match
        self.copy_to_imports = copy_to_imports  # Copy files to imports folder
    
    def _matches_pattern(self, filepath):
        """Check if file matches the expected pattern"""
        if self.file_pattern is None:
            return True
        filename = os.path.basename(filepath)
        return bool(self.file_pattern.match(filename))
    
    def on_created(self, event):
        if event.is_directory:
            return
        if event.src_path.endswith('.xlsx') or event.src_path.endswith('.xls'):
            if not self._matches_pattern(event.src_path):
                return
            # Wait longer for file to be fully written (downloads can be slow)
            time.sleep(2)
            self._process_file(event.src_path)
    
    def on_modified(self, event):
        if event.is_directory:
            return
        if event.src_path.endswith('.xlsx') or event.src_path.endswith('.xls'):
            if not self._matches_pattern(event.src_path):
                return
            # Debounce - only process if not recently processed
            now = time.time()
            last = self.last_modified.get(event.src_path, 0)
            if now - last < self.debounce_seconds:
                return
            self.last_modified[event.src_path] = now
            # Wait longer for file to be fully written
            time.sleep(2)
            self._process_file(event.src_path)
    
    def _process_file(self, filepath):
        """Process an Excel file"""
        filename = os.path.basename(filepath)
        print(f"\n[{datetime.now().strftime('%H:%M:%S')}] Detected file: {filename}")
        
        # If copy_to_imports is enabled, copy the file to imports folder
        if self.copy_to_imports:
            try:
                Path(IMPORTS_FOLDER).mkdir(parents=True, exist_ok=True)
                dest_path = os.path.join(IMPORTS_FOLDER, filename)
                shutil.copy2(filepath, dest_path)
                print(f"  Copied to: {dest_path}")
            except Exception as e:
                print(f"  Warning: Could not copy file: {e}")
        
        import_excel_file(filepath)


def start_watcher(folder_path, file_pattern=None, copy_to_imports=False):
    """Start watching a folder for Excel file changes"""
    # Create folder if it doesn't exist (only for imports folder)
    if not copy_to_imports:
        Path(folder_path).mkdir(parents=True, exist_ok=True)
    
    print(f"\n{'='*60}")
    print("Excel File Watcher Started")
    print(f"{'='*60}")
    print(f"Watching folder: {folder_path}")
    if file_pattern:
        print(f"File pattern: TicketReport_YYYYMMDD_HHMMSS.xlsx")
    else:
        print(f"Drop .xlsx files in this folder for automatic import")
    if copy_to_imports:
        print(f"Files will be copied to: {IMPORTS_FOLDER}")
    print(f"Press Ctrl+C to stop")
    print(f"{'='*60}\n")
    
    event_handler = ExcelFileHandler(file_pattern=file_pattern, copy_to_imports=copy_to_imports)
    observer = Observer()
    observer.schedule(event_handler, folder_path, recursive=False)
    observer.start()
    
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        print("\nStopping watcher...")
        observer.stop()
    
    observer.join()
    print("Watcher stopped.")


def start_downloads_watcher():
    """Start watching the Downloads folder for TicketReport files"""
    print(f"\n{'='*60}")
    print("Downloads Folder Watcher Started")
    print(f"{'='*60}")
    print(f"Watching: {DOWNLOADS_FOLDER}")
    print(f"Looking for: TicketReport_YYYYMMDD_HHMMSS.xlsx")
    print(f"Files will be copied to: {IMPORTS_FOLDER}")
    print(f"\nJust download your TicketReport file and it will be")
    print(f"automatically imported into the database!")
    print(f"\nPress Ctrl+C to stop")
    print(f"{'='*60}\n")
    
    # Create imports folder
    Path(IMPORTS_FOLDER).mkdir(parents=True, exist_ok=True)
    
    event_handler = ExcelFileHandler(
        file_pattern=TICKET_REPORT_PATTERN,
        copy_to_imports=True
    )
    observer = Observer()
    observer.schedule(event_handler, DOWNLOADS_FOLDER, recursive=False)
    observer.start()
    
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        print("\nStopping watcher...")
        observer.stop()
    
    observer.join()
    print("Watcher stopped.")


def import_latest_from_downloads():
    """Find and import the most recent TicketReport file from Downloads"""
    if not os.path.exists(DOWNLOADS_FOLDER):
        print(f"Downloads folder not found: {DOWNLOADS_FOLDER}")
        return False, 0, 0
    
    # Find all matching files
    matching_files = []
    for f in os.listdir(DOWNLOADS_FOLDER):
        if TICKET_REPORT_PATTERN.match(f):
            filepath = os.path.join(DOWNLOADS_FOLDER, f)
            matching_files.append((filepath, os.path.getmtime(filepath)))
    
    if not matching_files:
        print(f"No TicketReport files found in {DOWNLOADS_FOLDER}")
        print("Looking for files matching: TicketReport_YYYYMMDD_HHMMSS.xlsx")
        return False, 0, 0
    
    # Sort by modification time (newest first)
    matching_files.sort(key=lambda x: x[1], reverse=True)
    latest_file = matching_files[0][0]
    
    print(f"Found {len(matching_files)} TicketReport file(s)")
    print(f"Importing most recent: {os.path.basename(latest_file)}")
    
    # Copy to imports folder
    Path(IMPORTS_FOLDER).mkdir(parents=True, exist_ok=True)
    dest_path = os.path.join(IMPORTS_FOLDER, os.path.basename(latest_file))
    shutil.copy2(latest_file, dest_path)
    print(f"Copied to: {dest_path}")
    
    return import_excel_file(latest_file)


def main():
    parser = argparse.ArgumentParser(
        description="Import ticket tracking data from Excel files",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python sync_excel_to_db.py --file "C:/path/to/file.xlsx"   Import a specific file
  python sync_excel_to_db.py --watch                          Watch imports folder
  python sync_excel_to_db.py --watch-downloads                Watch Downloads folder for TicketReport_* files
  python sync_excel_to_db.py --import-latest                  Import most recent TicketReport from Downloads
        """
    )
    parser.add_argument('--file', '-f', type=str, help="Path to Excel file to import")
    parser.add_argument('--watch', '-w', action='store_true', help="Watch imports folder for new files")
    parser.add_argument('--watch-downloads', '-d', action='store_true', 
                        help="Watch Downloads folder for TicketReport_* files (auto-imports on download)")
    parser.add_argument('--import-latest', '-l', action='store_true',
                        help="Import the most recent TicketReport file from Downloads")
    parser.add_argument('--folder', type=str, default=IMPORTS_FOLDER, 
                        help="Folder to watch (default: backend/imports)")
    
    args = parser.parse_args()
    
    if args.file:
        # Single file import
        success, imported, updated = import_excel_file(args.file)
        sys.exit(0 if success else 1)
    elif args.watch_downloads:
        # Watch Downloads folder for TicketReport files
        start_downloads_watcher()
    elif args.import_latest:
        # Import the most recent TicketReport from Downloads
        success, imported, updated = import_latest_from_downloads()
        sys.exit(0 if success else 1)
    elif args.watch:
        # Start folder watcher
        start_watcher(args.folder)
    else:
        # Default: Check if there are any files in the imports folder
        folder = args.folder
        if os.path.exists(folder):
            files = [f for f in os.listdir(folder) if f.endswith('.xlsx') or f.endswith('.xls')]
            if files:
                print(f"Found {len(files)} Excel file(s) in {folder}")
                for f in files:
                    import_excel_file(os.path.join(folder, f))
            else:
                print(f"No Excel files found in {folder}")
                print("\nUsage options:")
                print("  --file <path>       Import a specific Excel file")
                print("  --watch             Watch imports folder for new files")
                print("  --watch-downloads   Watch Downloads folder for TicketReport_* files")
                print("  --import-latest     Import most recent TicketReport from Downloads")
        else:
            print(f"Imports folder does not exist: {folder}")
            print("\nUsage options:")
            print("  --file <path>       Import a specific Excel file")
            print("  --watch             Watch imports folder for new files")
            print("  --watch-downloads   Watch Downloads folder for TicketReport_* files")
            print("  --import-latest     Import most recent TicketReport from Downloads")


if __name__ == "__main__":
    main()
