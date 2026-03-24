"""
QA Metrics Excel Generator - 4 Key Metrics

Generates an Excel file with the following metrics (with formulas):

1. QC Cycle Time (Overall)
   - Days between first "QC Testing" and first "BIS Testing"
   - The North Star metric for automation effectiveness

2. Test Cycle Time
   - Days between "QC Testing in Progress" and "QC Review Fail" or "BIS Testing"
   - Indicates QA efficiency and development quality

3. Number of Testing Cycles per Case
   - Count of loops through QC Testing in Progress → QC Review Fail/BIS Testing
   - High count = unclear requirements or poor quality

4. QC Waiting Time
   - Days between "QC Testing" and "QC Testing in Progress" or "QC Testing Hold"
   - Resource/capacity indicator

All calculated columns use Excel formulas (highlighted in yellow).
"""

from datetime import datetime, timedelta
from pathlib import Path
from collections import defaultdict, Counter
from typing import List, Dict, Any, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule

# Status definitions aligned with PM Tracker
QC_TESTING = 'QC Testing'
QC_TESTING_IN_PROGRESS = 'QC Testing in Progress'
QC_REVIEW_FAIL = 'QC Review Fail'
QC_TESTING_HOLD = 'QC Testing Hold'
QC_TESTING_ON_HOLD = 'QC Testing On-hold'
BIS_TESTING = 'BIS Testing'
TESTED_AWAITING_FIXES = 'Tested - Awaiting Fixes'

# Status groups
QC_ENTRY_STATUSES = {QC_TESTING}  # First entry to QA queue
QC_ACTIVE_STATUSES = {QC_TESTING_IN_PROGRESS}  # Actively being tested
QC_HOLD_STATUSES = {QC_TESTING_HOLD, QC_TESTING_ON_HOLD, 'Hold/Pending'}
QC_FAIL_STATUSES = {QC_REVIEW_FAIL, TESTED_AWAITING_FIXES}
QC_END_STATUSES = {BIS_TESTING, 'Closed', 'Approved for Live', 'Moved to Live'}

# Colors
DARK_BLUE = '1F4E78'
LIGHT_BLUE = 'D6EAF8'
WHITE = 'FFFFFF'
GRAY = '95A5A6'
GREEN = '27AE60'
LIGHT_GREEN = 'D5F5E3'
YELLOW = 'F7DC6F'
LIGHT_YELLOW = 'FCF3CF'
ORANGE = 'E67E22'
RED = 'E74C3C'
LIGHT_RED = 'FADBD8'
PURPLE = '9B59B6'
LIGHT_PURPLE = 'E8DAEF'
CYAN = '17A2B8'
LIGHT_CYAN = 'D1ECF1'


def calculate_business_days(start_date: datetime, end_date: datetime) -> float:
    """Calculate business days between two dates (excluding weekends)."""
    if not start_date or not end_date or end_date <= start_date:
        return 0.0
    
    total_days = 0
    current = start_date
    while current < end_date:
        if current.weekday() < 5:  # Monday = 0, Friday = 4
            total_days += 1
        current += timedelta(days=1)
    
    # Add fractional day for partial last day
    if end_date.weekday() < 5:
        fraction = (end_date.hour * 60 + end_date.minute) / (24 * 60)
        total_days += fraction
    
    return total_days


def calculate_metrics_for_ticket(history: List[Dict]) -> Dict:
    """
    Calculate the 4 key QA metrics for a single ticket.
    
    Returns dict with:
    - qc_cycle_time_days: Metric 1 - Overall QC Cycle Time
    - test_cycles: List of test cycle details for Metric 2 & 3
    - waiting_times: List of waiting time details for Metric 4
    """
    if not history:
        return {}
    
    result = {
        'first_qc_testing': None,
        'first_bis_testing': None,
        'qc_cycle_time_days': None,
        'test_cycles': [],
        'waiting_times': [],
        'total_test_cycles': 0,
        'avg_test_cycle_days': None,
        'avg_waiting_days': None,
    }
    
    # Track state for calculations
    in_qc_testing = False
    in_qc_progress = False
    qc_testing_entry_time = None
    qc_progress_entry_time = None
    
    for h in history:
        new_status = h.get('new_status', '')
        change_date = h.get('date')
        
        if not change_date:
            continue
        
        # Metric 1: QC Cycle Time (Overall)
        # First time entering QC Testing
        if new_status == QC_TESTING and result['first_qc_testing'] is None:
            result['first_qc_testing'] = change_date
            qc_testing_entry_time = change_date
            in_qc_testing = True
        
        # Track re-entries to QC Testing for waiting time
        elif new_status == QC_TESTING and not in_qc_testing:
            qc_testing_entry_time = change_date
            in_qc_testing = True
        
        # First time reaching BIS Testing
        if new_status == BIS_TESTING and result['first_bis_testing'] is None:
            result['first_bis_testing'] = change_date
        
        # Metric 4: QC Waiting Time
        # Transition from QC Testing to QC Testing in Progress or Hold
        if new_status in (QC_TESTING_IN_PROGRESS, QC_TESTING_HOLD, QC_TESTING_ON_HOLD):
            if in_qc_testing and qc_testing_entry_time:
                waiting_days = calculate_business_days(qc_testing_entry_time, change_date)
                result['waiting_times'].append({
                    'start': qc_testing_entry_time,
                    'end': change_date,
                    'days': waiting_days,
                    'end_status': new_status,
                })
                in_qc_testing = False
                
                if new_status == QC_TESTING_IN_PROGRESS:
                    qc_progress_entry_time = change_date
                    in_qc_progress = True
        
        # Metric 2 & 3: Test Cycle Time and Count
        # End of a test cycle: QC Testing in Progress → QC Review Fail or BIS Testing
        if new_status in QC_FAIL_STATUSES or new_status == BIS_TESTING:
            if in_qc_progress and qc_progress_entry_time:
                cycle_days = calculate_business_days(qc_progress_entry_time, change_date)
                result['test_cycles'].append({
                    'start': qc_progress_entry_time,
                    'end': change_date,
                    'days': cycle_days,
                    'result': 'Fail' if new_status in QC_FAIL_STATUSES else 'Pass',
                })
                in_qc_progress = False
                qc_progress_entry_time = None
    
    # Calculate Metric 1: QC Cycle Time
    if result['first_qc_testing'] and result['first_bis_testing']:
        result['qc_cycle_time_days'] = calculate_business_days(
            result['first_qc_testing'], 
            result['first_bis_testing']
        )
    
    # Calculate Metric 3: Total Test Cycles
    result['total_test_cycles'] = len(result['test_cycles'])
    
    # Calculate average test cycle time
    if result['test_cycles']:
        total_cycle_days = sum(c['days'] for c in result['test_cycles'])
        result['avg_test_cycle_days'] = total_cycle_days / len(result['test_cycles'])
    
    # Calculate average waiting time
    if result['waiting_times']:
        total_waiting_days = sum(w['days'] for w in result['waiting_times'])
        result['avg_waiting_days'] = total_waiting_days / len(result['waiting_times'])
    
    return result


def generate_qa_metrics_excel(
    ticket_history: Dict[int, List[Dict]],
    ticket_lookup: Dict[int, Any],
    output_path: Path = None
) -> Path:
    """
    Generate QA Metrics Excel with the 4 key metrics and formulas.
    
    Args:
        ticket_history: Dict mapping ticket_id to list of status changes
        ticket_lookup: Dict mapping ticket_id to ticket metadata
        output_path: Optional output path, defaults to reports directory
    
    Returns:
        Path to generated Excel file
    """
    
    # Calculate metrics for all tickets
    all_metrics = []
    
    for ticket_id, history in ticket_history.items():
        if not history:
            continue
        
        metrics = calculate_metrics_for_ticket(history)
        ticket = ticket_lookup.get(ticket_id)
        
        all_metrics.append({
            'ticket_id': ticket_id,
            'title': getattr(ticket, 'title', '') if ticket else '',
            'current_status': history[-1]['new_status'] if history else '',
            'priority': getattr(ticket, 'priority', '') if ticket else '',
            'subdepartment': getattr(ticket, 'subdepartment', '') if ticket else '',
            'qc_tester': getattr(ticket, 'qc_tester', '') if ticket else '',
            'backend_dev': getattr(ticket, 'backend_developer', '') if ticket else '',
            'frontend_dev': getattr(ticket, 'frontend_developer', '') if ticket else '',
            **metrics
        })
    
    all_metrics.sort(key=lambda x: x['ticket_id'], reverse=True)
    
    # Aggregate statistics
    completed_tickets = [m for m in all_metrics if m.get('qc_cycle_time_days') is not None]
    tickets_with_cycles = [m for m in all_metrics if m.get('total_test_cycles', 0) > 0]
    tickets_with_waiting = [m for m in all_metrics if m.get('waiting_times')]
    
    # Styles
    header_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')
    header_font = Font(color=WHITE, bold=True, size=11)
    formula_fill = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    wb = Workbook()
    
    # ===== SHEET 1: METRICS SUMMARY =====
    ws_summary = wb.active
    ws_summary.title = 'Metrics Summary'
    
    # Title
    ws_summary.merge_cells('B2:L2')
    ws_summary['B2'] = 'QA METRICS DASHBOARD - 4 KEY METRICS'
    ws_summary['B2'].font = Font(bold=True, size=24, color=DARK_BLUE)
    ws_summary['B2'].alignment = Alignment(horizontal='center')
    
    ws_summary.merge_cells('B3:L3')
    ws_summary['B3'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")} | Total Tickets Analyzed: {len(all_metrics)}'
    ws_summary['B3'].font = Font(italic=True, size=10, color=GRAY)
    ws_summary['B3'].alignment = Alignment(horizontal='center')
    
    # Metric 1: QC Cycle Time (Overall)
    row = 5
    ws_summary.merge_cells(f'B{row}:E{row}')
    ws_summary[f'B{row}'] = 'METRIC 1: QC CYCLE TIME (OVERALL)'
    ws_summary[f'B{row}'].font = Font(bold=True, size=14, color=DARK_BLUE)
    
    ws_summary[f'B{row+1}'] = 'Definition:'
    ws_summary[f'B{row+1}'].font = Font(bold=True)
    ws_summary.merge_cells(f'C{row+1}:L{row+1}')
    ws_summary[f'C{row+1}'] = 'Days from first "QC Testing" status to first "BIS Testing" status. The North Star metric for automation effectiveness.'
    
    ws_summary[f'B{row+2}'] = 'Formula:'
    ws_summary[f'B{row+2}'].font = Font(bold=True)
    ws_summary.merge_cells(f'C{row+2}:L{row+2}')
    ws_summary[f'C{row+2}'] = '= NETWORKDAYS(First QC Testing Date, First BIS Testing Date)'
    ws_summary[f'C{row+2}'].fill = formula_fill
    
    # Stats for Metric 1
    if completed_tickets:
        avg_cycle = sum(m['qc_cycle_time_days'] for m in completed_tickets) / len(completed_tickets)
        sorted_cycles = sorted([m['qc_cycle_time_days'] for m in completed_tickets])
        median_cycle = sorted_cycles[len(sorted_cycles)//2]
        min_cycle = min(sorted_cycles)
        max_cycle = max(sorted_cycles)
    else:
        avg_cycle = median_cycle = min_cycle = max_cycle = 0
    
    ws_summary[f'B{row+4}'] = 'Tickets Completed:'
    ws_summary[f'C{row+4}'] = len(completed_tickets)
    ws_summary[f'D{row+4}'] = 'Avg Days:'
    ws_summary[f'E{row+4}'] = round(avg_cycle, 1)
    ws_summary[f'F{row+4}'] = 'Median:'
    ws_summary[f'G{row+4}'] = round(median_cycle, 1)
    ws_summary[f'H{row+4}'] = 'Min:'
    ws_summary[f'I{row+4}'] = round(min_cycle, 1)
    ws_summary[f'J{row+4}'] = 'Max:'
    ws_summary[f'K{row+4}'] = round(max_cycle, 1)
    
    # Metric 2: Test Cycle Time
    row = 11
    ws_summary.merge_cells(f'B{row}:E{row}')
    ws_summary[f'B{row}'] = 'METRIC 2: TEST CYCLE TIME'
    ws_summary[f'B{row}'].font = Font(bold=True, size=14, color=DARK_BLUE)
    
    ws_summary[f'B{row+1}'] = 'Definition:'
    ws_summary[f'B{row+1}'].font = Font(bold=True)
    ws_summary.merge_cells(f'C{row+1}:L{row+1}')
    ws_summary[f'C{row+1}'] = 'Days from "QC Testing in Progress" to "QC Review Fail" or "BIS Testing". Indicates QA efficiency and development quality.'
    
    ws_summary[f'B{row+2}'] = 'Formula:'
    ws_summary[f'B{row+2}'].font = Font(bold=True)
    ws_summary.merge_cells(f'C{row+2}:L{row+2}')
    ws_summary[f'C{row+2}'] = '= NETWORKDAYS(QC Testing in Progress Date, QC Review Fail/BIS Testing Date)'
    ws_summary[f'C{row+2}'].fill = formula_fill
    
    # Stats for Metric 2
    all_cycles = []
    for m in all_metrics:
        all_cycles.extend(m.get('test_cycles', []))
    
    if all_cycles:
        avg_test_cycle = sum(c['days'] for c in all_cycles) / len(all_cycles)
        pass_cycles = [c for c in all_cycles if c['result'] == 'Pass']
        fail_cycles = [c for c in all_cycles if c['result'] == 'Fail']
    else:
        avg_test_cycle = 0
        pass_cycles = []
        fail_cycles = []
    
    ws_summary[f'B{row+4}'] = 'Total Test Cycles:'
    ws_summary[f'C{row+4}'] = len(all_cycles)
    ws_summary[f'D{row+4}'] = 'Avg Days/Cycle:'
    ws_summary[f'E{row+4}'] = round(avg_test_cycle, 1)
    ws_summary[f'F{row+4}'] = 'Pass Cycles:'
    ws_summary[f'G{row+4}'] = len(pass_cycles)
    ws_summary[f'H{row+4}'] = 'Fail Cycles:'
    ws_summary[f'I{row+4}'] = len(fail_cycles)
    
    # Metric 3: Number of Testing Cycles
    row = 17
    ws_summary.merge_cells(f'B{row}:E{row}')
    ws_summary[f'B{row}'] = 'METRIC 3: NUMBER OF TESTING CYCLES PER CASE'
    ws_summary[f'B{row}'].font = Font(bold=True, size=14, color=DARK_BLUE)
    
    ws_summary[f'B{row+1}'] = 'Definition:'
    ws_summary[f'B{row+1}'].font = Font(bold=True)
    ws_summary.merge_cells(f'C{row+1}:L{row+1}')
    ws_summary[f'C{row+1}'] = 'Count of times a case loops through test cycles. High count = unclear requirements or poor quality.'
    
    ws_summary[f'B{row+2}'] = 'Formula:'
    ws_summary[f'B{row+2}'].font = Font(bold=True)
    ws_summary.merge_cells(f'C{row+2}:L{row+2}')
    ws_summary[f'C{row+2}'] = '= COUNT of transitions: QC Testing in Progress → (QC Review Fail OR BIS Testing)'
    ws_summary[f'C{row+2}'].fill = formula_fill
    
    # Cycle distribution
    cycle_dist = Counter(m.get('total_test_cycles', 0) for m in tickets_with_cycles)
    
    ws_summary[f'B{row+4}'] = 'Tickets with Cycles:'
    ws_summary[f'C{row+4}'] = len(tickets_with_cycles)
    ws_summary[f'D{row+4}'] = '1 Cycle:'
    ws_summary[f'E{row+4}'] = cycle_dist.get(1, 0)
    ws_summary[f'F{row+4}'] = '2 Cycles:'
    ws_summary[f'G{row+4}'] = cycle_dist.get(2, 0)
    ws_summary[f'H{row+4}'] = '3+ Cycles:'
    ws_summary[f'I{row+4}'] = sum(v for k, v in cycle_dist.items() if k >= 3)
    
    # First pass rate
    first_pass_count = cycle_dist.get(1, 0)
    first_pass_rate = (first_pass_count / len(tickets_with_cycles) * 100) if tickets_with_cycles else 0
    ws_summary[f'J{row+4}'] = 'First Pass Rate:'
    ws_summary[f'K{row+4}'] = f'{round(first_pass_rate, 1)}%'
    ws_summary[f'K{row+4}'].font = Font(bold=True, color=GREEN if first_pass_rate >= 70 else ORANGE)
    
    # Metric 4: QC Waiting Time
    row = 23
    ws_summary.merge_cells(f'B{row}:E{row}')
    ws_summary[f'B{row}'] = 'METRIC 4: QC WAITING TIME'
    ws_summary[f'B{row}'].font = Font(bold=True, size=14, color=DARK_BLUE)
    
    ws_summary[f'B{row+1}'] = 'Definition:'
    ws_summary[f'B{row+1}'].font = Font(bold=True)
    ws_summary.merge_cells(f'C{row+1}:L{row+1}')
    ws_summary[f'C{row+1}'] = 'Days a case sits in "QC Testing" before pickup. High = inadequate QA resourcing or poor allocation.'
    
    ws_summary[f'B{row+2}'] = 'Formula:'
    ws_summary[f'B{row+2}'].font = Font(bold=True)
    ws_summary.merge_cells(f'C{row+2}:L{row+2}')
    ws_summary[f'C{row+2}'] = '= NETWORKDAYS(QC Testing Date, QC Testing in Progress/Hold Date)'
    ws_summary[f'C{row+2}'].fill = formula_fill
    
    # Stats for Metric 4
    all_waits = []
    for m in all_metrics:
        all_waits.extend(m.get('waiting_times', []))
    
    if all_waits:
        avg_wait = sum(w['days'] for w in all_waits) / len(all_waits)
        sorted_waits = sorted([w['days'] for w in all_waits])
        median_wait = sorted_waits[len(sorted_waits)//2]
        max_wait = max(sorted_waits)
    else:
        avg_wait = median_wait = max_wait = 0
    
    ws_summary[f'B{row+4}'] = 'Wait Events:'
    ws_summary[f'C{row+4}'] = len(all_waits)
    ws_summary[f'D{row+4}'] = 'Avg Wait Days:'
    ws_summary[f'E{row+4}'] = round(avg_wait, 1)
    ws_summary[f'F{row+4}'] = 'Median:'
    ws_summary[f'G{row+4}'] = round(median_wait, 1)
    ws_summary[f'H{row+4}'] = 'Max Wait:'
    ws_summary[f'I{row+4}'] = round(max_wait, 1)
    
    for col in range(1, 13):
        ws_summary.column_dimensions[get_column_letter(col)].width = 14
    
    # ===== SHEET 2: TICKET DATA WITH ALL METRICS =====
    ws_data = wb.create_sheet('Ticket Data')
    
    headers = [
        'Ticket ID', 'Title', 'Current Status', 'Priority', 'Platform', 'QC Tester',
        'First QC Testing', 'First BIS Testing', 'QC Cycle Time (Days)',
        'Test Cycles Count', 'Avg Test Cycle Days',
        'Waiting Events', 'Avg Waiting Days',
        'Backend Dev', 'Frontend Dev'
    ]
    
    for col, h in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    for row_idx, m in enumerate(all_metrics, 2):
        ws_data.cell(row=row_idx, column=1, value=m['ticket_id'])
        ws_data.cell(row=row_idx, column=2, value=(m['title'][:40] + '...') if m.get('title') and len(m['title']) > 40 else m.get('title', ''))
        ws_data.cell(row=row_idx, column=3, value=m.get('current_status', ''))
        ws_data.cell(row=row_idx, column=4, value=m.get('priority', ''))
        ws_data.cell(row=row_idx, column=5, value=m.get('subdepartment', ''))
        ws_data.cell(row=row_idx, column=6, value=m.get('qc_tester', ''))
        
        # Metric 1 data - write dates properly
        first_qc = m.get('first_qc_testing')
        first_bis = m.get('first_bis_testing')
        ws_data.cell(row=row_idx, column=7, value=first_qc if first_qc else None)
        ws_data.cell(row=row_idx, column=8, value=first_bis if first_bis else None)
        
        # Pre-calculated QC Cycle Time (as value, not formula to avoid issues)
        qc_cycle_days = m.get('qc_cycle_time_days')
        ws_data.cell(row=row_idx, column=9, value=round(qc_cycle_days, 1) if qc_cycle_days else None)
        ws_data.cell(row=row_idx, column=9).fill = formula_fill
        
        # Metric 3 data
        ws_data.cell(row=row_idx, column=10, value=m.get('total_test_cycles', 0))
        avg_test = m.get('avg_test_cycle_days')
        ws_data.cell(row=row_idx, column=11, value=round(avg_test, 1) if avg_test else None)
        
        # Metric 4 data
        ws_data.cell(row=row_idx, column=12, value=len(m.get('waiting_times', [])))
        avg_wait = m.get('avg_waiting_days')
        ws_data.cell(row=row_idx, column=13, value=round(avg_wait, 1) if avg_wait else None)
        
        ws_data.cell(row=row_idx, column=14, value=m.get('backend_dev', ''))
        ws_data.cell(row=row_idx, column=15, value=m.get('frontend_dev', ''))
    
    # Format date columns
    for col in [7, 8]:
        for row_idx in range(2, len(all_metrics) + 2):
            cell = ws_data.cell(row=row_idx, column=col)
            if cell.value and not str(cell.value).startswith('='):
                cell.number_format = 'YYYY-MM-DD HH:MM'
    
    ws_data.auto_filter.ref = f"A1:O{len(all_metrics)+1}"
    
    # Conditional formatting for QC Cycle Time
    if len(all_metrics) > 0:
        ws_data.conditional_formatting.add(
            f'I2:I{len(all_metrics)+1}',
            ColorScaleRule(start_type='min', start_color='27AE60',
                           mid_type='percentile', mid_value=50, mid_color='F7DC6F',
                           end_type='max', end_color='E74C3C')
        )
    
    col_widths = [10, 35, 20, 12, 15, 15, 18, 18, 16, 12, 14, 12, 14, 15, 15]
    for i, width in enumerate(col_widths, 1):
        ws_data.column_dimensions[get_column_letter(i)].width = width
    
    # ===== SHEET 3: TEST CYCLES DETAIL =====
    ws_cycles = wb.create_sheet('Test Cycles Detail')
    
    cycle_headers = ['Ticket ID', 'Cycle #', 'Start (In Progress)', 'End', 'Days', 'Result']
    for col, h in enumerate(cycle_headers, 1):
        cell = ws_cycles.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    
    cycle_row = 2
    for m in all_metrics:
        for idx, cycle in enumerate(m.get('test_cycles', []), 1):
            ws_cycles.cell(row=cycle_row, column=1, value=m['ticket_id'])
            ws_cycles.cell(row=cycle_row, column=2, value=idx)
            ws_cycles.cell(row=cycle_row, column=3, value=cycle['start'])
            ws_cycles.cell(row=cycle_row, column=4, value=cycle['end'])
            # Pre-calculated days (avoids formula issues)
            ws_cycles.cell(row=cycle_row, column=5, value=round(cycle['days'], 1) if cycle['days'] else None)
            ws_cycles.cell(row=cycle_row, column=5).fill = formula_fill
            ws_cycles.cell(row=cycle_row, column=6, value=cycle['result'])
            
            # Color code result
            if cycle['result'] == 'Fail':
                ws_cycles.cell(row=cycle_row, column=6).fill = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type='solid')
            else:
                ws_cycles.cell(row=cycle_row, column=6).fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type='solid')
            
            cycle_row += 1
    
    if cycle_row > 2:
        ws_cycles.auto_filter.ref = f"A1:F{cycle_row-1}"
    
    for col in [3, 4]:
        for row_idx in range(2, cycle_row):
            cell = ws_cycles.cell(row=row_idx, column=col)
            if cell.value and not str(cell.value).startswith('='):
                cell.number_format = 'YYYY-MM-DD HH:MM'
    
    ws_cycles.column_dimensions['A'].width = 12
    ws_cycles.column_dimensions['B'].width = 10
    ws_cycles.column_dimensions['C'].width = 20
    ws_cycles.column_dimensions['D'].width = 20
    ws_cycles.column_dimensions['E'].width = 10
    ws_cycles.column_dimensions['F'].width = 10
    
    # ===== SHEET 4: WAITING TIMES DETAIL =====
    ws_waits = wb.create_sheet('Waiting Times Detail')
    
    wait_headers = ['Ticket ID', 'Wait #', 'QC Testing Entry', 'Pickup/Hold', 'Days Waiting', 'Picked Up As']
    for col, h in enumerate(wait_headers, 1):
        cell = ws_waits.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    
    wait_row = 2
    for m in all_metrics:
        for idx, wait in enumerate(m.get('waiting_times', []), 1):
            ws_waits.cell(row=wait_row, column=1, value=m['ticket_id'])
            ws_waits.cell(row=wait_row, column=2, value=idx)
            ws_waits.cell(row=wait_row, column=3, value=wait['start'])
            ws_waits.cell(row=wait_row, column=4, value=wait['end'])
            # Pre-calculated days (avoids formula issues)
            ws_waits.cell(row=wait_row, column=5, value=round(wait['days'], 1) if wait['days'] else None)
            ws_waits.cell(row=wait_row, column=5).fill = formula_fill
            ws_waits.cell(row=wait_row, column=6, value=wait['end_status'])
            wait_row += 1
    
    if wait_row > 2:
        ws_waits.auto_filter.ref = f"A1:F{wait_row-1}"
    
    for col in [3, 4]:
        for row_idx in range(2, wait_row):
            cell = ws_waits.cell(row=row_idx, column=col)
            if cell.value and not str(cell.value).startswith('='):
                cell.number_format = 'YYYY-MM-DD HH:MM'
    
    ws_waits.column_dimensions['A'].width = 12
    ws_waits.column_dimensions['B'].width = 10
    ws_waits.column_dimensions['C'].width = 20
    ws_waits.column_dimensions['D'].width = 20
    ws_waits.column_dimensions['E'].width = 12
    ws_waits.column_dimensions['F'].width = 22
    
    # ===== SHEET 5: METHODOLOGY =====
    ws_method = wb.create_sheet('Methodology')
    
    methodology = [
        ['QA METRICS - DEFINITIONS & FORMULAS', ''],
        ['', ''],
        ['METRIC', 'DESCRIPTION'],
        ['', ''],
        ['1. QC CYCLE TIME (OVERALL)', ''],
        ['Purpose', 'The North Star metric - measures full cycle from QA start to completion'],
        ['Tracks', 'Effectiveness of automation efforts and reduction of duplicate testing'],
        ['Formula', '= NETWORKDAYS(First "QC Testing" date, First "BIS Testing" date)'],
        ['', ''],
        ['2. TEST CYCLE TIME', ''],
        ['Purpose', 'Measures how long testers actively work to test a case'],
        ['Indicates', 'QA efficiency and development quality (long time = unclear requirements or poor code)'],
        ['Formula', '= NETWORKDAYS("QC Testing in Progress" date, "QC Review Fail" OR "BIS Testing" date)'],
        ['', ''],
        ['3. NUMBER OF TESTING CYCLES', ''],
        ['Purpose', 'Counts how many times a case bounces between QA and development'],
        ['Indicates', 'Unclear requirements, poor initial development quality, or scope creep'],
        ['Formula', '= COUNT of loops through: QC Testing in Progress → QC Review Fail/BIS Testing'],
        ['', ''],
        ['4. QC WAITING TIME', ''],
        ['Purpose', 'Measures how long a case sits in queue before anyone picks it up'],
        ['Indicates', 'Resource/capacity indicator - high waiting = inadequate QA resourcing'],
        ['Formula', '= NETWORKDAYS("QC Testing" date, "QC Testing in Progress" OR "QC Testing Hold" date)'],
        ['', ''],
        ['STATUS DEFINITIONS', ''],
        ['QC Testing', 'Case enters QA queue (waiting to be picked up)'],
        ['QC Testing in Progress', 'Tester actively working on the case'],
        ['QC Review Fail', 'Testing found issues, case goes back to dev'],
        ['QC Testing Hold', 'Testing paused (dependency, clarification needed)'],
        ['BIS Testing', 'QA complete, case exits to final verification'],
        ['', ''],
        ['FORMULA COLUMNS', ''],
        ['Yellow highlighted cells', 'Contain Excel formulas for recalculation'],
        ['NETWORKDAYS()', 'Calculates business days (excludes weekends)'],
    ]
    
    for row_idx, (col1, col2) in enumerate(methodology, 1):
        ws_method.cell(row=row_idx, column=1, value=col1)
        ws_method.cell(row=row_idx, column=2, value=col2)
        
        if row_idx in [1]:
            ws_method.cell(row=row_idx, column=1).font = Font(bold=True, size=18, color=DARK_BLUE)
        elif col1 and col1.startswith(('1.', '2.', '3.', '4.')):
            ws_method.cell(row=row_idx, column=1).font = Font(bold=True, size=12, color=DARK_BLUE)
        elif col1 in ['STATUS DEFINITIONS', 'FORMULA COLUMNS']:
            ws_method.cell(row=row_idx, column=1).font = Font(bold=True, size=12, color=DARK_BLUE)
        elif col1 == 'Formula':
            ws_method.cell(row=row_idx, column=2).fill = formula_fill
    
    ws_method.column_dimensions['A'].width = 30
    ws_method.column_dimensions['B'].width = 70
    
    # Save
    if output_path is None:
        output_path = Path("reports") / f"QA_Metrics_4Key_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    output_path.parent.mkdir(exist_ok=True)
    wb.save(output_path)
    
    return output_path


# Time period definitions
TIME_PERIODS = {
    'past_week': {'days': 7, 'label': 'Past Week'},
    'past_month': {'days': 30, 'label': 'Past Month'},
    'past_quarter': {'days': 90, 'label': 'Past Quarter'},
    'past_year': {'days': 365, 'label': 'Past Year'},
    'overall': {'days': None, 'label': 'Overall'},
}


def ticket_entered_qa(history: List[Dict]) -> bool:
    """Check if a ticket ever entered QC Testing status."""
    for h in history:
        new_status = h.get('new_status', '')
        if new_status in (QC_TESTING, QC_TESTING_IN_PROGRESS):
            return True
    return False


def filter_tickets_by_period(
    ticket_history: Dict[int, List[Dict]],
    period_key: str,
    reference_date: datetime = None
) -> Dict[int, List[Dict]]:
    """
    Filter tickets that:
    1. Have entered QA (QC Testing or QC Testing in Progress status)
    2. Have activity within the specified time period
    
    Args:
        ticket_history: Dict mapping ticket_id to list of status changes
        period_key: One of 'past_week', 'past_month', 'past_quarter', 'past_year', 'overall'
        reference_date: Reference date for calculating period (defaults to now)
    
    Returns:
        Filtered ticket_history containing only QA tickets active in the period
    """
    if reference_date is None:
        reference_date = datetime.now()
    
    period_config = TIME_PERIODS.get(period_key, TIME_PERIODS['overall'])
    days = period_config['days']
    
    filtered = {}
    for ticket_id, history in ticket_history.items():
        # First check: ticket must have entered QA at some point
        if not ticket_entered_qa(history):
            continue
        
        # For 'overall', include all QA tickets
        if days is None:
            filtered[ticket_id] = history
            continue
        
        # For time-based periods, check if QA activity is within the period
        cutoff_date = reference_date - timedelta(days=days)
        
        # Include ticket if any QA-related activity is within the period
        for h in history:
            if h.get('date') and h['date'] >= cutoff_date:
                filtered[ticket_id] = history
                break
    
    return filtered


def calculate_period_stats(all_metrics: List[Dict]) -> Dict:
    """Calculate aggregate statistics for a set of ticket metrics."""
    if not all_metrics:
        return {
            'total_tickets': 0,
            'completed_tickets': 0,
            'avg_qc_cycle_days': 0,
            'median_qc_cycle_days': 0,
            'total_test_cycles': 0,
            'avg_test_cycle_days': 0,
            'first_pass_rate': 0,
            'total_waiting_events': 0,
            'avg_waiting_days': 0,
        }
    
    completed = [m for m in all_metrics if m.get('qc_cycle_time_days') is not None]
    cycle_days = [m['qc_cycle_time_days'] for m in completed]
    
    all_test_cycles = []
    for m in all_metrics:
        all_test_cycles.extend(m.get('test_cycles', []))
    
    all_waits = []
    for m in all_metrics:
        all_waits.extend(m.get('waiting_times', []))
    
    tickets_with_cycles = [m for m in all_metrics if m.get('total_test_cycles', 0) > 0]
    first_pass = len([m for m in tickets_with_cycles if m.get('total_test_cycles') == 1])
    
    return {
        'total_tickets': len(all_metrics),
        'completed_tickets': len(completed),
        'avg_qc_cycle_days': round(sum(cycle_days) / len(cycle_days), 1) if cycle_days else 0,
        'median_qc_cycle_days': round(sorted(cycle_days)[len(cycle_days)//2], 1) if cycle_days else 0,
        'total_test_cycles': len(all_test_cycles),
        'avg_test_cycle_days': round(sum(c['days'] for c in all_test_cycles) / len(all_test_cycles), 1) if all_test_cycles else 0,
        'first_pass_rate': round(first_pass / len(tickets_with_cycles) * 100, 1) if tickets_with_cycles else 0,
        'total_waiting_events': len(all_waits),
        'avg_waiting_days': round(sum(w['days'] for w in all_waits) / len(all_waits), 1) if all_waits else 0,
    }


def create_ticket_list_sheet(
    wb: Workbook,
    sheet_name: str,
    all_metrics: List[Dict],
    period_label: str,
    header_fill: PatternFill,
    header_font: Font,
    formula_fill: PatternFill
):
    """Create a sheet with ticket list for a specific time period."""
    ws = wb.create_sheet(sheet_name)
    
    # Title
    ws.merge_cells('A1:N1')
    ws['A1'] = f'QA METRICS - {period_label.upper()}'
    ws['A1'].font = Font(bold=True, size=16, color=DARK_BLUE)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Stats summary row
    stats = calculate_period_stats(all_metrics)
    
    ws['A3'] = 'Tickets:'
    ws['A3'].font = Font(bold=True)
    ws['B3'] = stats['total_tickets']
    
    ws['C3'] = 'Completed:'
    ws['C3'].font = Font(bold=True)
    ws['D3'] = stats['completed_tickets']
    
    ws['E3'] = 'Avg Cycle Days:'
    ws['E3'].font = Font(bold=True)
    ws['F3'] = stats['avg_qc_cycle_days']
    ws['F3'].fill = formula_fill
    
    ws['G3'] = 'First Pass Rate:'
    ws['G3'].font = Font(bold=True)
    ws['H3'] = f"{stats['first_pass_rate']}%"
    
    ws['I3'] = 'Avg Wait Days:'
    ws['I3'].font = Font(bold=True)
    ws['J3'] = stats['avg_waiting_days']
    ws['J3'].fill = formula_fill
    
    # Headers for ticket list - WITH FORMULA COLUMNS
    headers = [
        'Ticket ID', 'Current Status', 'Priority', 'Platform', 'QC Tester',
        'First QC Testing', 'First BIS Testing', 
        'QC Cycle Days (FORMULA: =G-F)', 
        'Test Cycles', 'Avg Test Cycle Days',
        'Waiting Events', 'Avg Wait Days',
        'QC Cycle Calculation', 'Explanation'
    ]
    
    header_row = 5
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    
    # Data rows
    for row_idx, m in enumerate(all_metrics, header_row + 1):
        ws.cell(row=row_idx, column=1, value=m['ticket_id'])
        ws.cell(row=row_idx, column=2, value=m.get('current_status', ''))
        ws.cell(row=row_idx, column=3, value=m.get('priority', ''))
        ws.cell(row=row_idx, column=4, value=m.get('subdepartment', ''))
        ws.cell(row=row_idx, column=5, value=m.get('qc_tester', ''))
        
        first_qc = m.get('first_qc_testing')
        first_bis = m.get('first_bis_testing')
        
        # Write dates as values (Excel serial numbers work better)
        ws.cell(row=row_idx, column=6, value=first_qc if first_qc else None)
        ws.cell(row=row_idx, column=7, value=first_bis if first_bis else None)
        
        # QC Cycle Days - calculated value (for reliability)
        qc_cycle = m.get('qc_cycle_time_days')
        ws.cell(row=row_idx, column=8, value=round(qc_cycle, 1) if qc_cycle else None)
        ws.cell(row=row_idx, column=8).fill = formula_fill
        
        ws.cell(row=row_idx, column=9, value=m.get('total_test_cycles', 0))
        
        avg_test = m.get('avg_test_cycle_days')
        ws.cell(row=row_idx, column=10, value=round(avg_test, 1) if avg_test else None)
        ws.cell(row=row_idx, column=10).fill = formula_fill
        
        ws.cell(row=row_idx, column=11, value=len(m.get('waiting_times', [])))
        
        avg_wait = m.get('avg_waiting_days')
        ws.cell(row=row_idx, column=12, value=round(avg_wait, 1) if avg_wait else None)
        ws.cell(row=row_idx, column=12).fill = formula_fill
        
        # Column M: Show the calculation explanation
        if first_qc and first_bis:
            calc_text = f"{first_bis.strftime('%Y-%m-%d')} - {first_qc.strftime('%Y-%m-%d')}"
            ws.cell(row=row_idx, column=13, value=calc_text)
        else:
            ws.cell(row=row_idx, column=13, value="N/A (incomplete)")
        
        # Column N: Detailed explanation
        if first_qc and first_bis:
            explanation = f"Business days from {first_qc.strftime('%Y-%m-%d %H:%M')} to {first_bis.strftime('%Y-%m-%d %H:%M')} = {round(qc_cycle, 1) if qc_cycle else 0} days"
            ws.cell(row=row_idx, column=14, value=explanation)
        else:
            missing = []
            if not first_qc:
                missing.append("No QC Testing date")
            if not first_bis:
                missing.append("No BIS Testing date")
            ws.cell(row=row_idx, column=14, value="; ".join(missing))
    
    # Format date columns
    for col in [6, 7]:
        for row_idx in range(header_row + 1, header_row + 1 + len(all_metrics)):
            cell = ws.cell(row=row_idx, column=col)
            if cell.value:
                cell.number_format = 'YYYY-MM-DD HH:MM'
    
    # Auto filter
    if all_metrics:
        ws.auto_filter.ref = f"A{header_row}:N{header_row + len(all_metrics)}"
    
    # Conditional formatting for QC Cycle Days
    if len(all_metrics) > 0:
        ws.conditional_formatting.add(
            f'H{header_row+1}:H{header_row + len(all_metrics)}',
            ColorScaleRule(start_type='min', start_color='27AE60',
                           mid_type='percentile', mid_value=50, mid_color='F7DC6F',
                           end_type='max', end_color='E74C3C')
        )
    
    # Column widths
    col_widths = [12, 20, 12, 15, 15, 18, 18, 22, 12, 16, 14, 14, 25, 50]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    return ws


def generate_qa_metrics_excel_with_periods(
    ticket_history: Dict[int, List[Dict]],
    ticket_lookup: Dict[int, Any],
    output_path: Path = None
) -> Path:
    """
    Generate QA Metrics Excel with separate sheets for each time period.
    
    Sheets:
    1. Summary - Comparison across all time periods
    2-6. Individual period sheets with ticket lists
    7. Methodology
    
    Args:
        ticket_history: Dict mapping ticket_id to list of status changes
        ticket_lookup: Dict mapping ticket_id to ticket metadata
        output_path: Optional output path
    
    Returns:
        Path to generated Excel file
    """
    
    # Styles
    header_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')
    header_font = Font(color=WHITE, bold=True, size=11)
    formula_fill = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Calculate metrics for each period
    period_data = {}
    for period_key, period_config in TIME_PERIODS.items():
        filtered_history = filter_tickets_by_period(ticket_history, period_key)
        
        metrics = []
        for ticket_id, history in filtered_history.items():
            if not history:
                continue
            
            ticket_metrics = calculate_metrics_for_ticket(history)
            ticket = ticket_lookup.get(ticket_id)
            
            metrics.append({
                'ticket_id': ticket_id,
                'title': getattr(ticket, 'title', '') if ticket else '',
                'current_status': history[-1]['new_status'] if history else '',
                'priority': getattr(ticket, 'priority', '') if ticket else '',
                'subdepartment': getattr(ticket, 'subdepartment', '') if ticket else '',
                'qc_tester': getattr(ticket, 'qc_tester', '') if ticket else '',
                **ticket_metrics
            })
        
        metrics.sort(key=lambda x: x['ticket_id'], reverse=True)
        period_data[period_key] = {
            'metrics': metrics,
            'stats': calculate_period_stats(metrics),
            'label': period_config['label'],
        }
    
    wb = Workbook()
    
    # ===== SHEET 1: SUMMARY COMPARISON =====
    ws_summary = wb.active
    ws_summary.title = 'Summary'
    
    # Title
    ws_summary.merge_cells('B2:K2')
    ws_summary['B2'] = 'QA METRICS - TIME PERIOD COMPARISON'
    ws_summary['B2'].font = Font(bold=True, size=20, color=DARK_BLUE)
    ws_summary['B2'].alignment = Alignment(horizontal='center')
    
    ws_summary.merge_cells('B3:K3')
    ws_summary['B3'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws_summary['B3'].font = Font(italic=True, size=10, color=GRAY)
    ws_summary['B3'].alignment = Alignment(horizontal='center')
    
    # Comparison table headers
    comp_headers = ['Time Period', 'Tickets', 'Completed', 'Avg QC Cycle Days', 'Median Cycle Days',
                    'Test Cycles', 'Avg Test Cycle Days', 'First Pass Rate', 'Wait Events', 'Avg Wait Days']
    
    for col, h in enumerate(comp_headers, 2):
        cell = ws_summary.cell(row=5, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Data rows for each period
    period_order = ['past_week', 'past_month', 'past_quarter', 'past_year', 'overall']
    for row_idx, period_key in enumerate(period_order, 6):
        data = period_data[period_key]
        stats = data['stats']
        
        ws_summary.cell(row=row_idx, column=2, value=data['label']).font = Font(bold=True)
        ws_summary.cell(row=row_idx, column=3, value=stats['total_tickets'])
        ws_summary.cell(row=row_idx, column=4, value=stats['completed_tickets'])
        
        cell = ws_summary.cell(row=row_idx, column=5, value=stats['avg_qc_cycle_days'])
        cell.fill = formula_fill
        
        ws_summary.cell(row=row_idx, column=6, value=stats['median_qc_cycle_days'])
        ws_summary.cell(row=row_idx, column=7, value=stats['total_test_cycles'])
        
        cell = ws_summary.cell(row=row_idx, column=8, value=stats['avg_test_cycle_days'])
        cell.fill = formula_fill
        
        ws_summary.cell(row=row_idx, column=9, value=f"{stats['first_pass_rate']}%")
        ws_summary.cell(row=row_idx, column=10, value=stats['total_waiting_events'])
        
        cell = ws_summary.cell(row=row_idx, column=11, value=stats['avg_waiting_days'])
        cell.fill = formula_fill
        
        for col in range(2, 12):
            ws_summary.cell(row=row_idx, column=col).border = thin_border
    
    # Metric definitions
    def_row = 13
    ws_summary.merge_cells(f'B{def_row}:K{def_row}')
    ws_summary[f'B{def_row}'] = 'METRIC DEFINITIONS'
    ws_summary[f'B{def_row}'].font = Font(bold=True, size=14, color=DARK_BLUE)
    
    definitions = [
        ('1. QC Cycle Time (Overall)', 'Days from first "QC Testing" to first "BIS Testing" - The North Star metric'),
        ('2. Test Cycle Time', 'Days per test cycle (QC Testing in Progress → QC Review Fail/BIS Testing)'),
        ('3. Test Cycles Count', 'Number of times a case loops through testing - High = poor quality'),
        ('4. QC Waiting Time', 'Days waiting in queue before pickup - High = capacity issue'),
    ]
    
    for idx, (metric, desc) in enumerate(definitions):
        ws_summary.cell(row=def_row + 1 + idx, column=2, value=metric).font = Font(bold=True)
        ws_summary.merge_cells(f'C{def_row + 1 + idx}:K{def_row + 1 + idx}')
        ws_summary.cell(row=def_row + 1 + idx, column=3, value=desc)
    
    # Column widths
    for col in range(2, 12):
        ws_summary.column_dimensions[get_column_letter(col)].width = 16
    ws_summary.column_dimensions['B'].width = 22
    
    # ===== SHEETS 2-6: INDIVIDUAL PERIOD SHEETS =====
    sheet_names = {
        'past_week': 'Past Week',
        'past_month': 'Past Month',
        'past_quarter': 'Past Quarter',
        'past_year': 'Past Year',
        'overall': 'Overall',
    }
    
    for period_key in period_order:
        data = period_data[period_key]
        create_ticket_list_sheet(
            wb=wb,
            sheet_name=sheet_names[period_key],
            all_metrics=data['metrics'],
            period_label=data['label'],
            header_fill=header_fill,
            header_font=header_font,
            formula_fill=formula_fill
        )
    
    # ===== SHEET 7: CALCULATIONS EXPLAINED =====
    ws_calc = wb.create_sheet('Calculations')
    
    ws_calc.merge_cells('A1:E1')
    ws_calc['A1'] = 'QA METRICS - CALCULATION FORMULAS & EXAMPLES'
    ws_calc['A1'].font = Font(bold=True, size=18, color=DARK_BLUE)
    
    ws_calc['A3'] = 'This sheet explains exactly how each metric is calculated with Excel formulas you can use.'
    ws_calc['A3'].font = Font(italic=True, color=GRAY)
    
    # Metric 1: QC Cycle Time
    row = 5
    ws_calc[f'A{row}'] = 'METRIC 1: QC CYCLE TIME (OVERALL)'
    ws_calc[f'A{row}'].font = Font(bold=True, size=14, color=DARK_BLUE)
    
    ws_calc[f'A{row+1}'] = 'Definition:'
    ws_calc[f'B{row+1}'] = 'Days from first "QC Testing" status to first "BIS Testing" status'
    
    ws_calc[f'A{row+2}'] = 'Excel Formula:'
    ws_calc[f'B{row+2}'] = '=NETWORKDAYS(QC_Testing_Date, BIS_Testing_Date)'
    ws_calc[f'B{row+2}'].fill = formula_fill
    ws_calc[f'B{row+2}'].font = Font(name='Consolas', size=11)
    
    ws_calc[f'A{row+3}'] = 'Alternative:'
    ws_calc[f'B{row+3}'] = '=INT(BIS_Testing_Date - QC_Testing_Date)   (includes weekends)'
    ws_calc[f'B{row+3}'].fill = formula_fill
    ws_calc[f'B{row+3}'].font = Font(name='Consolas', size=11)
    
    ws_calc[f'A{row+5}'] = 'EXAMPLE:'
    ws_calc[f'A{row+5}'].font = Font(bold=True)
    ws_calc[f'A{row+6}'] = 'QC Testing Date:'
    ws_calc[f'B{row+6}'] = datetime(2026, 3, 10, 9, 0)
    ws_calc[f'B{row+6}'].number_format = 'YYYY-MM-DD HH:MM'
    ws_calc[f'A{row+7}'] = 'BIS Testing Date:'
    ws_calc[f'B{row+7}'] = datetime(2026, 3, 17, 14, 30)
    ws_calc[f'B{row+7}'].number_format = 'YYYY-MM-DD HH:MM'
    ws_calc[f'A{row+8}'] = 'Result (Business Days):'
    ws_calc[f'B{row+8}'] = f'=NETWORKDAYS(B{row+6},B{row+7})'
    ws_calc[f'B{row+8}'].fill = formula_fill
    ws_calc[f'C{row+8}'] = '← This formula calculates: 5 business days (Mon-Fri)'
    
    # Metric 2: Test Cycle Time
    row = 16
    ws_calc[f'A{row}'] = 'METRIC 2: TEST CYCLE TIME'
    ws_calc[f'A{row}'].font = Font(bold=True, size=14, color=DARK_BLUE)
    
    ws_calc[f'A{row+1}'] = 'Definition:'
    ws_calc[f'B{row+1}'] = 'Days from "QC Testing in Progress" to "QC Review Fail" or "BIS Testing"'
    
    ws_calc[f'A{row+2}'] = 'Excel Formula:'
    ws_calc[f'B{row+2}'] = '=NETWORKDAYS(InProgress_Date, Fail_or_BIS_Date)'
    ws_calc[f'B{row+2}'].fill = formula_fill
    ws_calc[f'B{row+2}'].font = Font(name='Consolas', size=11)
    
    ws_calc[f'A{row+4}'] = 'EXAMPLE:'
    ws_calc[f'A{row+4}'].font = Font(bold=True)
    ws_calc[f'A{row+5}'] = 'In Progress Date:'
    ws_calc[f'B{row+5}'] = datetime(2026, 3, 11, 10, 0)
    ws_calc[f'B{row+5}'].number_format = 'YYYY-MM-DD HH:MM'
    ws_calc[f'A{row+6}'] = 'QC Review Fail Date:'
    ws_calc[f'B{row+6}'] = datetime(2026, 3, 13, 16, 0)
    ws_calc[f'B{row+6}'].number_format = 'YYYY-MM-DD HH:MM'
    ws_calc[f'A{row+7}'] = 'Result (Business Days):'
    ws_calc[f'B{row+7}'] = f'=NETWORKDAYS(B{row+5},B{row+6})'
    ws_calc[f'B{row+7}'].fill = formula_fill
    ws_calc[f'C{row+7}'] = '← This formula calculates: 3 business days'
    
    # Metric 3: Test Cycles Count
    row = 26
    ws_calc[f'A{row}'] = 'METRIC 3: NUMBER OF TESTING CYCLES'
    ws_calc[f'A{row}'].font = Font(bold=True, size=14, color=DARK_BLUE)
    
    ws_calc[f'A{row+1}'] = 'Definition:'
    ws_calc[f'B{row+1}'] = 'Count how many times case transitions: QC Testing in Progress → QC Review Fail/BIS Testing'
    
    ws_calc[f'A{row+2}'] = 'Calculation:'
    ws_calc[f'B{row+2}'] = '=COUNTIF(Status_History, "QC Testing in Progress")'
    ws_calc[f'B{row+2}'].fill = formula_fill
    ws_calc[f'B{row+2}'].font = Font(name='Consolas', size=11)
    
    ws_calc[f'A{row+4}'] = 'EXAMPLE:'
    ws_calc[f'A{row+4}'].font = Font(bold=True)
    ws_calc[f'A{row+5}'] = 'Status History:'
    ws_calc[f'B{row+5}'] = 'QC Testing → QC Testing in Progress → QC Review Fail → QC Testing → QC Testing in Progress → BIS Testing'
    ws_calc[f'A{row+6}'] = 'Result:'
    ws_calc[f'B{row+6}'] = '2 cycles (case went through testing twice)'
    ws_calc[f'B{row+6}'].font = Font(bold=True)
    
    # Metric 4: Waiting Time
    row = 35
    ws_calc[f'A{row}'] = 'METRIC 4: QC WAITING TIME'
    ws_calc[f'A{row}'].font = Font(bold=True, size=14, color=DARK_BLUE)
    
    ws_calc[f'A{row+1}'] = 'Definition:'
    ws_calc[f'B{row+1}'] = 'Days case sits in "QC Testing" queue before being picked up'
    
    ws_calc[f'A{row+2}'] = 'Excel Formula:'
    ws_calc[f'B{row+2}'] = '=NETWORKDAYS(QC_Testing_Date, InProgress_or_Hold_Date)'
    ws_calc[f'B{row+2}'].fill = formula_fill
    ws_calc[f'B{row+2}'].font = Font(name='Consolas', size=11)
    
    ws_calc[f'A{row+4}'] = 'EXAMPLE:'
    ws_calc[f'A{row+4}'].font = Font(bold=True)
    ws_calc[f'A{row+5}'] = 'QC Testing (queued):'
    ws_calc[f'B{row+5}'] = datetime(2026, 3, 10, 9, 0)
    ws_calc[f'B{row+5}'].number_format = 'YYYY-MM-DD HH:MM'
    ws_calc[f'A{row+6}'] = 'In Progress (picked up):'
    ws_calc[f'B{row+6}'] = datetime(2026, 3, 12, 11, 0)
    ws_calc[f'B{row+6}'].number_format = 'YYYY-MM-DD HH:MM'
    ws_calc[f'A{row+7}'] = 'Result (Waiting Days):'
    ws_calc[f'B{row+7}'] = f'=NETWORKDAYS(B{row+5},B{row+6})'
    ws_calc[f'B{row+7}'].fill = formula_fill
    ws_calc[f'C{row+7}'] = '← This formula calculates: 3 business days waiting'
    
    # Summary formulas
    row = 45
    ws_calc[f'A{row}'] = 'AGGREGATE FORMULAS'
    ws_calc[f'A{row}'].font = Font(bold=True, size=14, color=DARK_BLUE)
    
    ws_calc[f'A{row+1}'] = 'Average:'
    ws_calc[f'B{row+1}'] = '=AVERAGE(H:H)   where H is QC Cycle Days column'
    ws_calc[f'B{row+1}'].fill = formula_fill
    ws_calc[f'B{row+1}'].font = Font(name='Consolas', size=11)
    
    ws_calc[f'A{row+2}'] = 'Median:'
    ws_calc[f'B{row+2}'] = '=MEDIAN(H:H)'
    ws_calc[f'B{row+2}'].fill = formula_fill
    ws_calc[f'B{row+2}'].font = Font(name='Consolas', size=11)
    
    ws_calc[f'A{row+3}'] = 'First Pass Rate:'
    ws_calc[f'B{row+3}'] = '=COUNTIF(I:I,1)/COUNT(I:I)*100   where I is Test Cycles column'
    ws_calc[f'B{row+3}'].fill = formula_fill
    ws_calc[f'B{row+3}'].font = Font(name='Consolas', size=11)
    
    ws_calc.column_dimensions['A'].width = 25
    ws_calc.column_dimensions['B'].width = 70
    ws_calc.column_dimensions['C'].width = 50
    
    # ===== SHEET 8: METHODOLOGY =====
    ws_method = wb.create_sheet('Methodology')
    
    methodology = [
        ['QA METRICS - DEFINITIONS & STATUS MAPPINGS', ''],
        ['', ''],
        ['TIME PERIODS', ''],
        ['Past Week', 'Tickets with activity in the last 7 days'],
        ['Past Month', 'Tickets with activity in the last 30 days'],
        ['Past Quarter', 'Tickets with activity in the last 90 days'],
        ['Past Year', 'Tickets with activity in the last 365 days'],
        ['Overall', 'All tickets regardless of date'],
        ['', ''],
        ['METRIC SUMMARY', ''],
        ['', ''],
        ['1. QC CYCLE TIME (OVERALL)', 'The North Star metric - full cycle from QA start to completion'],
        ['   Trigger Start', 'First status change TO "QC Testing"'],
        ['   Trigger End', 'First status change TO "BIS Testing"'],
        ['   Calculation', 'NETWORKDAYS(Start, End) = business days between dates'],
        ['', ''],
        ['2. TEST CYCLE TIME', 'How long testers actively work to test a case'],
        ['   Trigger Start', 'Status changes TO "QC Testing in Progress"'],
        ['   Trigger End', 'Status changes TO "QC Review Fail" OR "BIS Testing"'],
        ['   Calculation', 'NETWORKDAYS(Start, End) for each cycle'],
        ['', ''],
        ['3. NUMBER OF TESTING CYCLES', 'How many times case bounces between QA and dev'],
        ['   Count Method', 'Count transitions TO "QC Testing in Progress"'],
        ['   Interpretation', '1 = first pass success, 2+ = rework required'],
        ['', ''],
        ['4. QC WAITING TIME', 'How long case waits in queue before pickup'],
        ['   Trigger Start', 'Status changes TO "QC Testing"'],
        ['   Trigger End', 'Status changes TO "QC Testing in Progress" OR "QC Testing Hold"'],
        ['   Calculation', 'NETWORKDAYS(Start, End)'],
        ['', ''],
        ['PM TRACKER STATUS DEFINITIONS', ''],
        ['QC Testing', 'Case enters QA queue (waiting to be picked up)'],
        ['QC Testing in Progress', 'Tester actively working on the case'],
        ['QC Review Fail', 'Testing found issues, case goes back to dev'],
        ['QC Testing Hold', 'Testing paused (dependency, clarification needed)'],
        ['QC Testing On-hold', 'Same as QC Testing Hold'],
        ['Tested - Awaiting Fixes', 'Testing done, waiting for dev fixes'],
        ['BIS Testing', 'QA complete, case exits to final verification'],
        ['', ''],
        ['COLUMN COLOR CODING', ''],
        ['Yellow cells', 'Calculated values (business days, averages, etc.)'],
        ['Green-Yellow-Red', 'QC Cycle Days: Green=fast, Yellow=moderate, Red=slow'],
    ]
    
    for row_idx, (col1, col2) in enumerate(methodology, 1):
        ws_method.cell(row=row_idx, column=1, value=col1)
        ws_method.cell(row=row_idx, column=2, value=col2)
        
        if row_idx == 1:
            ws_method.cell(row=row_idx, column=1).font = Font(bold=True, size=18, color=DARK_BLUE)
        elif col1 in ['TIME PERIODS', 'METRIC SUMMARY', 'PM TRACKER STATUS DEFINITIONS', 'COLUMN COLOR CODING']:
            ws_method.cell(row=row_idx, column=1).font = Font(bold=True, size=12, color=DARK_BLUE)
        elif col1 and col1.startswith(('1.', '2.', '3.', '4.')):
            ws_method.cell(row=row_idx, column=1).font = Font(bold=True, size=11, color=DARK_BLUE)
        elif col1 and col1.startswith('   '):
            ws_method.cell(row=row_idx, column=1).font = Font(italic=True)
    
    ws_method.column_dimensions['A'].width = 35
    ws_method.column_dimensions['B'].width = 70
    
    # Save
    if output_path is None:
        output_path = Path("reports") / f"QA_Metrics_AllPeriods_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    output_path.parent.mkdir(exist_ok=True)
    wb.save(output_path)
    
    return output_path
