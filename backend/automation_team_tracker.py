"""Automation Team Contribution Tracker — tracks individual scripting and execution activity."""
import json, os, logging, subprocess, re
from datetime import date, datetime
from collections import defaultdict
from typing import Dict, List, Optional

logger = logging.getLogger(__name__)

TRACKER_FILE = os.path.join(os.path.dirname(__file__), 'data', 'automation_team_tracker.json')
GIT_REPO_PATH = os.environ.get('AUTOMATION_GIT_REPO', r'D:\Vishnu VS\bis-automation')

TEAM = {
    'Vishnu VS': {'git_authors': ['Vishnu VS'], 'email': 'vishnu.vs@techversantinfotech.com'},
    'Vivek V Nair': {'git_authors': ['vivekv-techversant'], 'email': 'vivek.v@techversantinfotech.com'},
    'Varsha Dcruz P': {'git_authors': ['varshadcruz'], 'email': ''},
}


def _load_tracker() -> Dict:
    if os.path.exists(TRACKER_FILE):
        try:
            with open(TRACKER_FILE, 'r') as f:
                return json.load(f)
        except Exception:
            pass
    return {'members': {name: {'entries': []} for name in TEAM}, 'last_sync': None}


def _save_tracker(data: Dict):
    os.makedirs(os.path.dirname(TRACKER_FILE), exist_ok=True)
    with open(TRACKER_FILE, 'w') as f:
        json.dump(data, f, indent=2, default=str)


def import_from_excel(excel_path: str, person: str = 'Vishnu VS') -> Dict:
    """Import historical scripting data from automation_dashboard.xlsx."""
    from openpyxl import load_workbook

    wb = load_workbook(excel_path, data_only=True)
    tracker = _load_tracker()
    if person not in tracker['members']:
        tracker['members'][person] = {'entries': []}

    # Clear existing excel-sourced entries for this person
    tracker['members'][person]['entries'] = [
        e for e in tracker['members'][person]['entries'] if e.get('source') != 'excel'
    ]

    imported = 0
    # Read each "Cases" sheet
    module_map = {
        'Classroom Calendar': 'Classroom Calendar',
        'Online Course': 'Online Course',
        'User Administration': 'User Management',
    }

    for sheet_name in wb.sheetnames:
        if 'Cases' not in sheet_name:
            continue

        ws = wb[sheet_name]
        # Detect module from sheet name
        module = None
        for key, mod in module_map.items():
            if key in sheet_name:
                module = mod
                break
        if not module:
            continue

        # Find header row (look for "Test Case ID" or "S.No")
        header_row = None
        headers = {}
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=False):
            for cell in row:
                if cell.value and str(cell.value).strip().lower() in ('s.no', 'test case id', 'title'):
                    header_row = cell.row
                    break
            if header_row:
                break

        if not header_row:
            continue

        # Map header columns
        for cell in ws[header_row]:
            if cell.value:
                headers[str(cell.value).strip().lower()] = cell.column - 1

        case_id_col = headers.get('test case id', headers.get('case id'))
        date_col = headers.get('scripted date')
        status_col = headers.get('automation status')

        if case_id_col is None:
            continue

        # Group by scripted date
        date_counts = defaultdict(int)
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or len(row) <= (case_id_col or 0):
                continue
            case_id = row[case_id_col] if case_id_col is not None else None
            if not case_id:
                continue

            scripted_date = None
            if date_col is not None and len(row) > date_col:
                raw = row[date_col]
                if isinstance(raw, datetime):
                    scripted_date = raw.strftime('%Y-%m-%d')
                elif isinstance(raw, date):
                    scripted_date = raw.isoformat()
                elif raw:
                    try:
                        scripted_date = datetime.strptime(str(raw).strip(), '%d-%b-%Y').strftime('%Y-%m-%d')
                    except Exception:
                        try:
                            scripted_date = datetime.strptime(str(raw).strip()[:10], '%Y-%m-%d').strftime('%Y-%m-%d')
                        except Exception:
                            pass

            if scripted_date:
                date_counts[scripted_date] += 1

        # Create entries per date
        for dt, count in sorted(date_counts.items()):
            tracker['members'][person]['entries'].append({
                'date': dt,
                'module': module,
                'cases_scripted': count,
                'cases_executed': 0,
                'activity': 'scripting',
                'source': 'excel',
                'notes': f'Imported from {sheet_name}',
            })
            imported += count

    # Also import execution tracker data
    if 'Execution Tracker' in wb.sheetnames:
        ws = wb['Execution Tracker']
        # Find header
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=False):
            headers = {}
            for cell in row:
                if cell.value:
                    headers[str(cell.value).strip().lower()] = cell.column - 1
            if 'case id' in headers or 'module' in headers:
                break

        mod_col = headers.get('module')
        exec_col = headers.get('exec count')
        date_col = headers.get('last run date')

        if mod_col is not None and exec_col is not None:
            exec_by_module_date = defaultdict(lambda: defaultdict(int))
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) <= max(mod_col, exec_col):
                    continue
                mod = row[mod_col]
                exec_count = row[exec_col]
                run_date = None
                if date_col is not None and len(row) > date_col and row[date_col]:
                    raw = row[date_col]
                    if isinstance(raw, (datetime, date)):
                        run_date = raw.strftime('%Y-%m-%d') if isinstance(raw, datetime) else raw.isoformat()
                    else:
                        try:
                            run_date = datetime.strptime(str(raw).strip()[:10], '%Y-%m-%d').strftime('%Y-%m-%d')
                        except Exception:
                            pass

                if mod and exec_count and int(exec_count) > 0 and run_date:
                    exec_by_module_date[mod][run_date] += int(exec_count)

            for mod, dates in exec_by_module_date.items():
                for dt, count in dates.items():
                    tracker['members'][person]['entries'].append({
                        'date': dt,
                        'module': mod,
                        'cases_scripted': 0,
                        'cases_executed': count,
                        'activity': 'execution',
                        'source': 'excel',
                        'notes': f'Execution from tracker',
                    })

    _save_tracker(tracker)
    logger.info(f'Imported {imported} scripted cases for {person}')
    return {'imported_cases': imported, 'person': person}


def sync_git_activity() -> Dict:
    """Sync git commit activity for all team members."""
    tracker = _load_tracker()
    if not os.path.exists(GIT_REPO_PATH):
        return {'error': f'Git repo not found at {GIT_REPO_PATH}'}

    # Build author -> person mapping
    author_to_person = {}
    for person, info in TEAM.items():
        for author in info['git_authors']:
            author_to_person[author.lower()] = person

    # Get git log since last sync or last 30 days
    since = tracker.get('last_sync') or '2026-03-01'
    try:
        result = subprocess.run(
            ['git', 'log', '--all', f'--since={since}', '--format=%an|%ad|%s', '--date=short', '--numstat'],
            cwd=GIT_REPO_PATH, capture_output=True, text=True, timeout=30
        )
        if result.returncode != 0:
            return {'error': result.stderr[:200]}
    except Exception as e:
        return {'error': str(e)}

    # Parse git output: count new spec files per author per day
    commits = defaultdict(lambda: defaultdict(lambda: {'spec_files': 0, 'commits': 0, 'messages': []}))
    current_author = None
    current_date = None

    for line in result.stdout.split('\n'):
        line = line.strip()
        if not line:
            continue
        if '|' in line and len(line.split('|')) >= 3:
            parts = line.split('|', 2)
            author = parts[0].strip().lower()
            current_date = parts[1].strip()
            msg = parts[2].strip()
            person = author_to_person.get(author)
            if person:
                current_author = person
                commits[person][current_date]['commits'] += 1
                commits[person][current_date]['messages'].append(msg)
            else:
                current_author = None
        elif current_author and current_date:
            # numstat line: "123\t0\tpath/to/file.spec.ts"
            parts = line.split('\t')
            if len(parts) >= 3:
                filepath = parts[2]
                added = parts[0]
                if (filepath.endswith('.spec.ts') or filepath.endswith('.test.ts')) and added != '0' and added != '-':
                    commits[current_author][current_date]['spec_files'] += 1

    # Add entries
    added = 0
    for person, dates in commits.items():
        if person not in tracker['members']:
            tracker['members'][person] = {'entries': []}
        existing_dates = {(e['date'], e.get('source')) for e in tracker['members'][person]['entries']}
        for dt, data in dates.items():
            if (dt, 'git') in existing_dates:
                continue
            tracker['members'][person]['entries'].append({
                'date': dt,
                'module': '',
                'cases_scripted': data['spec_files'],
                'cases_executed': 0,
                'activity': 'scripting',
                'source': 'git',
                'notes': f'{data["commits"]} commits, {data["spec_files"]} spec files',
            })
            added += 1

    tracker['last_sync'] = date.today().isoformat()
    _save_tracker(tracker)
    return {'synced': added, 'members': list(commits.keys())}


def add_manual_entry(person: str, entry_date: str, module: str,
                     cases_scripted: int = 0, cases_executed: int = 0,
                     activity: str = 'scripting', notes: str = '') -> Dict:
    """Add a manual entry."""
    tracker = _load_tracker()
    if person not in tracker['members']:
        tracker['members'][person] = {'entries': []}

    tracker['members'][person]['entries'].append({
        'date': entry_date,
        'module': module,
        'cases_scripted': cases_scripted,
        'cases_executed': cases_executed,
        'activity': activity,
        'source': 'manual',
        'notes': notes,
    })
    _save_tracker(tracker)
    return {'success': True}


def get_member_weekly_activity(person: str) -> Dict:
    """Get weekly Mon-Fri breakdown for a specific team member (like the Excel weekly report)."""
    tracker = _load_tracker()
    data = tracker.get('members', {}).get(person, {})
    entries = data.get('entries', [])

    # Group entries by week (Mon start) then by day
    from datetime import timedelta as td
    weekly = defaultdict(lambda: {
        'days': {d: {'scripted': 0, 'executed': 0, 'notes': []} for d in ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']},
        'scripted': 0, 'executed': 0, 'modules': defaultdict(int),
    })

    DAY_NAMES = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']

    for e in entries:
        try:
            d = date.fromisoformat(e['date'])
        except Exception:
            continue
        week_start = d - td(days=d.weekday())
        wk_key = f'{week_start.isoformat()} - {(week_start + td(days=4)).isoformat()}'
        day_name = DAY_NAMES[d.weekday()]
        if day_name in ('Sat', 'Sun'):
            continue

        scripted = e.get('cases_scripted', 0)
        executed = e.get('cases_executed', 0)
        weekly[wk_key]['days'][day_name]['scripted'] += scripted
        weekly[wk_key]['days'][day_name]['executed'] += executed
        if e.get('notes'):
            weekly[wk_key]['days'][day_name]['notes'].append(e['notes'])
        weekly[wk_key]['scripted'] += scripted
        weekly[wk_key]['executed'] += executed
        mod = e.get('module') or 'General'
        if scripted > 0:
            weekly[wk_key]['modules'][mod] += scripted

    # Build cumulative
    weeks_list = []
    cumulative = 0
    for wk_key in sorted(weekly.keys()):
        w = weekly[wk_key]
        cumulative += w['scripted']
        # Format daily breakdown with module info
        day_details = {}
        for day_name in ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']:
            dd = w['days'][day_name]
            day_details[day_name] = {
                'scripted': dd['scripted'],
                'executed': dd['executed'],
                'notes': ' | '.join(dd['notes']) if dd['notes'] else '',
            }

        # Module breakdown for this week
        mod_breakdown = ', '.join(f'{m}: {c}' for m, c in sorted(w['modules'].items(), key=lambda x: -x[1])) if w['modules'] else ''

        weeks_list.append({
            'week': wk_key,
            'days': day_details,
            'scripted': w['scripted'],
            'executed': w['executed'],
            'cumulative': cumulative,
            'module_breakdown': mod_breakdown,
        })

    return {
        'person': person,
        'weeks': weeks_list,
        'total_scripted': cumulative,
        'total_executed': sum(w['executed'] for w in weeks_list),
    }


def get_team_stats() -> Dict:
    """Get aggregated stats per team member."""
    tracker = _load_tracker()
    members = []

    for person, data in tracker.get('members', {}).items():
        entries = data.get('entries', [])
        total_scripted = sum(e.get('cases_scripted', 0) for e in entries)
        total_executed = sum(e.get('cases_executed', 0) for e in entries)

        # By module
        module_stats = defaultdict(lambda: {'scripted': 0, 'executed': 0})
        for e in entries:
            mod = e.get('module') or 'General'
            module_stats[mod]['scripted'] += e.get('cases_scripted', 0)
            module_stats[mod]['executed'] += e.get('cases_executed', 0)

        # By week (last 8 weeks)
        weekly = defaultdict(lambda: {'scripted': 0, 'executed': 0})
        for e in entries:
            try:
                d = date.fromisoformat(e['date'])
                week_start = d - __import__('datetime').timedelta(days=d.weekday())
                week_key = week_start.isoformat()
                weekly[week_key]['scripted'] += e.get('cases_scripted', 0)
                weekly[week_key]['executed'] += e.get('cases_executed', 0)
            except Exception:
                pass

        weekly_list = [{'week': k, **v} for k, v in sorted(weekly.items())[-8:]]

        members.append({
            'name': person,
            'total_scripted': total_scripted,
            'total_executed': total_executed,
            'modules': [{'module': k, **v} for k, v in sorted(module_stats.items(), key=lambda x: -x[1]['scripted'])],
            'weekly_trend': weekly_list,
            'entry_count': len(entries),
            'last_entry': max((e['date'] for e in entries), default=None) if entries else None,
        })

    members.sort(key=lambda m: -m['total_scripted'])
    return {'members': members, 'last_sync': tracker.get('last_sync')}
