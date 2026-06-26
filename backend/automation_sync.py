"""Automation module sync + metrics (rebuilt 2026-06).

Source of truth: TestRail Project 18 (suite 137 Web, 847 Mobile).
- Catalog (AutomationCase): one row per case with automation status, automated-by, module.
- Executions (AutomationExecution): one row per case-in-a-run for utilization/reuse.
- Daily snapshot (AutomationSnapshot): one row/day of computed metrics for growth charts.

Reuses the auth/constants from sync_automation_testrail; adds 429-aware request handling
(TestRail rate-limits aggressive scans) and the section->module resolver from pm_live_data logic.
"""
import os
import time
import requests
from collections import defaultdict
from datetime import datetime, date, timedelta

from database import SessionLocal
from models import (
    AutomationCase, AutomationExecution, AutomationSnapshot, AppSetting, SyncLog,
)
from sync_automation_testrail import API_BASE, headers, AUTOMATION_STATUS_MAP, STATUS_IDS, parse_datetime
from pm_live_data import load_module_ownership

PROJECT_ID = 18
SUITES = {137: "Web", 847: "Mobile"}
STATUS_KEY = "custom_case_automated"
AUTOMATED_BY_KEY = "custom_case_automated_by"

# custom_case_automated_by is a DROPDOWN. SOURCE OF TRUTH = TestRail's own field config:
#   GET /api/v2/get_case_fields -> custom_case_automated_by options = "1, Vishnu V S | 2, Varsha D | 3, Vivek V"
# i.e. id 1 = Vishnu, id 2 = Varsha, id 3 = Vivek. Do NOT swap these — an earlier "the labels are
# misleading" override had Vishnu<->Vivek backwards (1<->3) and mislabelled ~1,100 of Vishnu's cases as
# Vivek's. If the counts ever look swapped again, re-check get_case_fields rather than flipping this map.
TEAM = ["Vishnu VS", "Varsha Dcruz P", "Vivek V Nair"]
# Corrected defaults (used only as a fallback if the live TestRail field can't be read).
BYID_TO_PERSON = {1: "Vishnu VS", 2: "Varsha Dcruz P", 3: "Vivek V Nair"}
# Accept the various name forms used across the app/TestRail -> dropdown id (for write-back).
PERSON_TO_BYID = {
    "vishnu vs": 1, "vishnu v s": 1, "vishnuvs": 1, "vishnu": 1,
    "varsha": 2, "varsha d": 2, "varsha dcruz": 2, "varsha dcruz p": 2,
    "vivek": 3, "vivek v": 3, "vivek v nair": 3,
}


def refresh_automated_by_map():
    """PERMANENT FIX: pull the automated_by id->name mapping straight from TestRail's field config so it
    always matches the source of truth and can never be swapped by a wrong hardcoded value again. Maps
    the TestRail labels (e.g. 'Vishnu V S') to our canonical TEAM names by first name. Updates the module
    globals in place; on any failure the corrected hardcoded defaults above stand."""
    global BYID_TO_PERSON, PERSON_TO_BYID
    try:
        fields = _request("GET", "get_case_fields")
        for f in (fields or []):
            if f.get("system_name") != "custom_case_automated_by":
                continue
            for cfg in f.get("configs", []):
                items = ((cfg.get("options") or {}).get("items") or "")
                m = {}
                for line in items.splitlines():
                    if "," not in line:
                        continue
                    sid, label = line.split(",", 1)
                    try:
                        sid = int(sid.strip())
                    except ValueError:
                        continue
                    first = (label.strip().split() or [""])[0].lower()
                    canon = next((t for t in TEAM if t.lower().startswith(first)), label.strip())
                    m[sid] = canon
                if m:
                    BYID_TO_PERSON = m
                    PERSON_TO_BYID = dict(PERSON_TO_BYID)
                    for i, t in m.items():
                        PERSON_TO_BYID[t.lower()] = i
                    return m
    except Exception:
        pass
    return BYID_TO_PERSON

DEFAULT_MANUAL_MINUTES = 15


# ---------------------------------------------------------------------------
# TestRail client (429-aware)
# ---------------------------------------------------------------------------
def _request(method, path, params=None, json_body=None, max_retries=6):
    url = f"{API_BASE}/{path}"
    delay = 2.0
    last = None
    for _ in range(max_retries):
        resp = requests.request(method, url, headers=headers, params=params, json=json_body, timeout=90)
        last = resp
        if resp.status_code == 429:
            try:
                wait = float(resp.headers.get("Retry-After", delay))
            except (TypeError, ValueError):
                wait = delay
            time.sleep(max(wait, delay))
            delay = min(delay * 2, 30)
            continue
        resp.raise_for_status()
        return resp.json()
    last.raise_for_status()
    return last.json()


def _paged(path, params=None, key=None):
    out, offset, limit = [], 0, 250
    params = dict(params or {})
    while True:
        params.update(offset=offset, limit=limit)
        data = _request("GET", path, params=params)
        if isinstance(data, list):
            items = data
        elif isinstance(data, dict):
            items = data.get(key, []) if key else []
        else:
            items = []
        if not items:
            break
        out.extend(items)
        if len(items) < limit:
            break
        offset += limit
        time.sleep(0.25)
    return out


def fetch_cases(suite_id):
    return _paged(f"get_cases/{PROJECT_ID}", {"suite_id": suite_id}, "cases")


def fetch_sections(suite_id):
    return _paged(f"get_sections/{PROJECT_ID}", {"suite_id": suite_id}, "sections")


def fetch_plans():
    return _paged(f"get_plans/{PROJECT_ID}", {"is_completed": 0}, "plans")


def fetch_runs_for_project():
    return _paged(f"get_runs/{PROJECT_ID}", {"is_completed": 0}, "runs")


def fetch_plan(plan_id):
    return _request("GET", f"get_plan/{plan_id}")


def fetch_tests(run_id):
    return _paged(f"get_tests/{run_id}", {}, "tests")


def fetch_case_fields():
    data = _request("GET", "get_case_fields")
    return data.get("case_fields", data) if isinstance(data, dict) else data


def update_case(case_id, fields):
    """Write-back: POST update_case/{case_id} with {custom_...: value}. Raises on non-2xx."""
    return _request("POST", f"update_case/{case_id}", json_body=fields)


# ---------------------------------------------------------------------------
# Settings (key/value)
# ---------------------------------------------------------------------------
def get_setting(db, key, default=None):
    row = db.query(AppSetting).filter(AppSetting.key == key).first()
    return row.value if row else default


def set_setting(db, key, value):
    row = db.query(AppSetting).filter(AppSetting.key == key).first()
    if row:
        row.value = str(value)
    else:
        db.add(AppSetting(key=key, value=str(value)))
    db.commit()


def get_manual_minutes(db):
    try:
        return int(float(get_setting(db, "manual_minutes_per_case", DEFAULT_MANUAL_MINUTES)))
    except (TypeError, ValueError):
        return DEFAULT_MANUAL_MINUTES


def person_to_byid(person):
    return PERSON_TO_BYID.get((person or "").strip().lower())


# ---------------------------------------------------------------------------
# Section -> module resolver (lifted from pm_live_data._fetch_testrail_module_stats)
# ---------------------------------------------------------------------------
def build_section_resolver():
    ownership = load_module_ownership()
    mapping = ownership.get("testrail_mapping", {})
    sec_map = {}
    for suite_id in SUITES:
        for s in fetch_sections(suite_id):
            if s.get("id"):
                sec_map[s["id"]] = s

    def ancestry(sec_id):
        names, seen = [], set()
        while sec_id and sec_id not in seen:
            seen.add(sec_id)
            s = sec_map.get(sec_id)
            if not s:
                break
            names.append((s.get("name") or "").lower())
            sec_id = s.get("parent_id")
        return names

    def resolve(sec_id):
        if not sec_id:
            return None
        names = ancestry(sec_id)
        for mod, cfg in mapping.items():
            mapped = set(cfg.get("section_ids", []))
            cid, seen = sec_id, set()
            while cid and cid not in seen:
                seen.add(cid)
                if cid in mapped:
                    return mod
                p = sec_map.get(cid)
                cid = p.get("parent_id") if p else None
            for kw in cfg.get("keywords", []):
                kw = kw.lower()
                if any(kw in n for n in names):
                    return mod
        return None

    return resolve, sec_map


def extract_ticket_id(*texts):
    import re
    for t in texts:
        if not t:
            continue
        m = re.match(r"^\s*(\d{3,})\s*[_\-\s]", t)  # "18400_..", "14176 - Staging"
        if m:
            return int(m.group(1))
        m = re.search(r"#(\d{3,})", t)
        if m:
            return int(m.group(1))
    return None


# ---------------------------------------------------------------------------
# Sync: catalog
# ---------------------------------------------------------------------------
def upsert_case_catalog(db, resolve, sec_map):
    """Upsert AutomationCase rows from TestRail cases. Returns {case_id: (module, is_automated)}."""
    now = datetime.utcnow()
    catalog = {}
    seen_status = set()
    for suite_id, suite_name in SUITES.items():
        try:
            cases = fetch_cases(suite_id)
        except Exception as e:
            print(f"[automation_sync] cases suite {suite_id} failed: {e}")
            continue
        for c in cases:
            cid = c.get("id")
            if not cid:
                continue
            status_id = c.get(STATUS_KEY)
            status = AUTOMATION_STATUS_MAP.get(status_id)
            by_id = c.get(AUTOMATED_BY_KEY)
            person = BYID_TO_PERSON.get(by_id)
            sec_id = c.get("section_id")
            module = resolve(sec_id)
            sec_name = (sec_map.get(sec_id) or {}).get("name")

            row = db.query(AutomationCase).filter(AutomationCase.case_id == cid).first()
            is_new = row is None
            old_status = None if is_new else row.automation_status
            if is_new:
                row = AutomationCase(case_id=cid, created_on=now)
                db.add(row)
            # Stamp transition dates ONLY for genuine transitions observed after go-live — NOT for
            # pre-existing baseline on first sight (else week-1 counts the whole backlog). Going
            # forward, a case flipping into Planned/Automated gets a real app-observed date.
            if not is_new:
                if status == "Planned" and old_status != "Planned":
                    row.planned_on = now
                    row.planned_by = person
                if status == "Automated" and old_status != "Automated":
                    row.automated_on = now
            row.suite_id = suite_id
            row.suite_name = suite_name
            row.section_id = sec_id
            row.section_name = sec_name
            row.module = module
            row.title = (c.get("title") or "")[:500]
            row.priority = str(c.get("priority_id") or "")
            row.automation_status = status
            row.automation_status_id = status_id
            row.automatable = (status_id != 4)
            # automated_by: TestRail (custom_case_automated_by) is authoritative ONLY when set.
            # When blank, preserve any app/seed attribution (e.g. Vishnu's cases that aren't
            # tagged in TestRail) rather than wiping it to null.
            if by_id is not None:
                row.automated_by = person
                row.automated_by_id = by_id
            row.last_synced = now
            catalog[cid] = (module, status == "Automated")
        db.commit()
    return catalog


# ---------------------------------------------------------------------------
# Sync: executions
# ---------------------------------------------------------------------------
def _upsert_run(db, catalog, run_tuple):
    """Fetch + upsert all tests for one run. Returns rows touched."""
    run_id, run_name, plan_id, ticket_id, run_created = run_tuple
    if not run_id:
        return 0
    today = date.today()
    now = datetime.utcnow()
    try:
        tests = fetch_tests(run_id)
    except Exception as e:
        print(f"[automation_sync] tests run {run_id} failed: {e}")
        return 0
    existing = {e.test_id: e for e in db.query(AutomationExecution).filter(
        AutomationExecution.run_id == run_id).all()}
    n = 0
    for t in tests:
        test_id = t.get("id")
        if not test_id:
            continue
        case_id = t.get("case_id")
        module, is_auto = catalog.get(case_id, (None, False))
        status_id = t.get("status_id")
        row = existing.get(test_id)
        if not row:
            row = AutomationExecution(test_id=test_id, executed_on=today, created_on=now)
            db.add(row)
        row.case_id = case_id
        row.run_id = run_id
        row.plan_id = plan_id
        row.ticket_id = ticket_id
        row.module = module
        row.status_id = status_id
        row.status_name = STATUS_IDS.get(status_id)
        row.is_automated_case = bool(is_auto)
        row.run_created_on = run_created
        row.last_synced = now
        n += 1
    db.commit()
    time.sleep(0.2)
    return n


def upsert_executions(db, catalog):
    """Standalone runs first (big regression runs land early), then plan runs — each run is
    written immediately so partial progress is visible and a slow plan traversal can't lose data."""
    total = 0
    try:
        for r in fetch_runs_for_project():
            total += _upsert_run(db, catalog, (r.get("id"), r.get("name"), r.get("plan_id"),
                                               extract_ticket_id(r.get("name")),
                                               parse_datetime(r.get("created_on"))))
    except Exception as e:
        print(f"[automation_sync] standalone runs failed: {e}")
    try:
        for p in fetch_plans():
            det = fetch_plan(p.get("id"))
            ptid = extract_ticket_id(p.get("name"))
            for entry in (det.get("entries") or []) if det else []:
                for run in entry.get("runs", []):
                    total += _upsert_run(db, catalog, (
                        run.get("id"), run.get("name"), p.get("id"),
                        extract_ticket_id(run.get("name"), p.get("name")) or ptid,
                        parse_datetime(run.get("created_on"))))
    except Exception as e:
        print(f"[automation_sync] plan runs failed: {e}")
    return total


# ---------------------------------------------------------------------------
# Metrics
# ---------------------------------------------------------------------------
def _week_start(d=None):
    d = d or date.today()
    return d - timedelta(days=d.weekday())  # Monday


def compute_metrics(db, week_start=None):
    """Full payload: overview + per-module + per-ticket + status mix + team + planning."""
    M = get_manual_minutes(db)
    cases = db.query(AutomationCase).all()
    execs = db.query(AutomationExecution).all()

    # Per-module aggregation
    mod = {}

    def mslot(m):
        m = m or "Unmapped"
        return mod.setdefault(m, {
            "module": m, "total_cases": 0, "automated_cases": 0, "automatable_cases": 0,
            "planned_cases": 0, "total_executions": 0, "automated_executions": 0,
        })

    status_mix = {}
    for c in cases:
        s = mslot(c.module)
        s["total_cases"] += 1
        if c.automatable:
            s["automatable_cases"] += 1
        if c.automation_status == "Automated":
            s["automated_cases"] += 1
        if c.automation_status == "Planned":
            s["planned_cases"] += 1
        status_mix[c.automation_status or "Unset"] = status_mix.get(c.automation_status or "Unset", 0) + 1

    ticket = {}
    for e in execs:
        s = mslot(e.module)
        s["total_executions"] += 1
        if e.is_automated_case:
            s["automated_executions"] += 1
        if e.ticket_id:
            ts = ticket.setdefault(e.ticket_id, {"ticket_id": e.ticket_id, "module": e.module,
                                                 "total_executions": 0, "automated_executions": 0})
            ts["total_executions"] += 1
            if e.is_automated_case:
                ts["automated_executions"] += 1

    modules = []
    for s in mod.values():
        auto_exec = s["automated_executions"]
        s["coverage_pct"] = round(s["automated_cases"] / s["automatable_cases"] * 100, 1) if s["automatable_cases"] else 0.0
        s["utilization_pct"] = round(auto_exec / s["total_executions"] * 100, 1) if s["total_executions"] else 0.0
        s["reuse_ratio"] = round(auto_exec / s["automated_cases"], 1) if s["automated_cases"] else 0.0
        s["time_saved_minutes"] = auto_exec * M
        s["time_saved_hours"] = round(auto_exec * M / 60, 1)
        modules.append(s)
    modules.sort(key=lambda x: x["time_saved_hours"], reverse=True)

    tickets = []
    for t in ticket.values():
        t["utilization_pct"] = round(t["automated_executions"] / t["total_executions"] * 100, 1) if t["total_executions"] else 0.0
        t["time_saved_hours"] = round(t["automated_executions"] * M / 60, 1)
        tickets.append(t)
    tickets.sort(key=lambda x: x["time_saved_hours"], reverse=True)

    total_cases = sum(s["total_cases"] for s in modules)
    total_automatable = sum(s["automatable_cases"] for s in modules)
    total_automated = sum(s["automated_cases"] for s in modules)
    total_exec = sum(s["total_executions"] for s in modules)
    auto_exec = sum(s["automated_executions"] for s in modules)
    top = max((m for m in modules if m["total_executions"] > 0), key=lambda x: x["utilization_pct"], default=None)

    overview = {
        "total_cases": total_cases,
        "automated_cases": total_automated,
        "automatable_cases": total_automatable,
        "coverage_pct": round(total_automated / total_automatable * 100, 1) if total_automatable else 0.0,
        "total_executions": total_exec,
        "automated_executions": auto_exec,
        "manual_executions": total_exec - auto_exec,
        "utilization_pct": round(auto_exec / total_exec * 100, 1) if total_exec else 0.0,
        "time_saved_hours": round(auto_exec * M / 60, 1),
        "manual_minutes_per_case": M,
        "top_module": top["module"] if top else None,
        "top_module_utilization": top["utilization_pct"] if top else 0.0,
        "status_mix": status_mix,
    }
    return {
        "overview": overview,
        "modules": modules,
        "tickets": tickets[:25],
        "team": compute_team(db, week_start),
        "planning": compute_planning(db, week_start),
    }


def compute_team(db, week_start=None):
    """Per-member cumulative + weekly (8 weeks ending at the selected week) Automated counts."""
    cases = db.query(AutomationCase).filter(AutomationCase.automation_status == "Automated").all()
    weeks = [(_week_start(week_start) - timedelta(weeks=k)) for k in range(7, -1, -1)]
    members = []
    for person in TEAM:
        pcases = [c for c in cases if c.automated_by == person]
        weekly = []
        for ws in weeks:
            we = ws + timedelta(days=6)
            n = sum(1 for c in pcases if c.automated_on and ws <= c.automated_on.date() <= we)
            weekly.append({"week": ws.isoformat(), "scripted": n})
        this_week = weekly[-1]["scripted"] if weekly else 0
        members.append({
            "name": person,
            "total_scripted": len(pcases),
            "this_week": this_week,
            "weekly": weekly,
        })
    return {"members": members, "week_start": _week_start(week_start).isoformat()}


def compute_planning(db, week_start=None):
    """Planned (targeted) automation work + backlog, surfaced by PERSON (incl. Unassigned) and by
    MODULE so cases show even when automated_by isn't tagged yet. 'planned' = all cases currently
    in Planned status; they roll forward until marked Automated."""
    ws = _week_start(week_start)
    we = ws + timedelta(days=6)
    planned = db.query(AutomationCase).filter(AutomationCase.automation_status == "Planned").all()
    automated = db.query(AutomationCase).filter(AutomationCase.automation_status == "Automated").all()

    labels = TEAM + ["Unassigned"]

    def pkey(c):
        return c.automated_by if c.automated_by in TEAM else "Unassigned"

    planned_by_person = {p: 0 for p in labels}
    for c in planned:
        planned_by_person[pkey(c)] += 1

    automated_week = [c for c in automated if c.automated_on and ws <= c.automated_on.date() <= we]
    aut_week_by_person = {p: 0 for p in labels}
    for c in automated_week:
        aut_week_by_person[pkey(c)] += 1

    by_module = defaultdict(int)
    for c in planned:
        by_module[c.module or "Unmapped"] += 1
    planned_by_module = sorted(({"module": k, "planned": v} for k, v in by_module.items()),
                               key=lambda x: -x["planned"])

    def brief(c):
        return {"case_id": c.case_id, "title": c.title, "module": c.module,
                "automated_by": c.automated_by or "Unassigned",
                "planned_on": c.planned_on.date().isoformat() if c.planned_on else None}

    return {
        "week_start": ws.isoformat(),
        "planned_total": len(planned),
        "automated_this_week_total": len(automated_week),
        "by_person": [{"person": p, "planned": planned_by_person[p],
                       "automated_this_week": aut_week_by_person[p]} for p in labels],
        "by_module": planned_by_module,
        "backlog": [brief(c) for c in planned[:500]],
    }


import json as _json
_WEEK_PLAN_FILE = os.path.join(os.path.dirname(__file__), "data", "automation_week_plan.json")


def _load_week_plan():
    try:
        with open(_WEEK_PLAN_FILE) as f:
            return _json.load(f)
    except Exception:
        return {}


def _save_week_plan(d):
    try:
        os.makedirs(os.path.dirname(_WEEK_PLAN_FILE), exist_ok=True)
        with open(_WEEK_PLAN_FILE, "w") as f:
            _json.dump(d, f)
    except Exception:
        pass


def compute_weekly_automation(db, week_start=None):
    """Weekly scripting view for the combined report + live dashboard. The team's model (no dates):
    a case **Planned + automated_by(name)** is next week's committed plan for that person; as each is
    scripted it flips to **Automated**. Planned cases with **no name** are the general backlog.

      next_week  — Planned + named, grouped by module (count + per-person + case list).
      backlog    — Planned + unnamed (Unassigned), grouped by module.
      this_week  — per person {planned, scripted}: scripted = Automated with automated_on in [Mon,Sun];
                   planned = the named-plan that was committed for THIS week, read from a weekly snapshot
                   (captured when it was tagged as "next week"), so it survives the plan rolling forward.
      daily      — per-day scripted counts per person for the last 14 days (drives the dashboard).
    """
    ws = _week_start(week_start)
    we = ws + timedelta(days=6)
    nws, nwe = ws + timedelta(days=7), ws + timedelta(days=13)

    def named(c):
        return c.automated_by in TEAM

    planned = db.query(AutomationCase).filter(AutomationCase.automation_status == "Planned").all()
    automated = db.query(AutomationCase).filter(AutomationCase.automation_status == "Automated").all()
    not_automated = db.query(AutomationCase).filter(AutomationCase.automation_status == "Not Automated").all()

    next_week_cases = [c for c in planned if named(c)]       # Planned + name → next week's plan
    # Backlog = the "Not Automated" pool (the cases still to be automated) + any un-named Planned cases
    # (so nothing is orphaned). Driven by the Not Automated status per the team's convention.
    backlog_cases = not_automated + [c for c in planned if not named(c)]

    def by_module(cases):
        mods = defaultdict(list)
        for c in cases:
            mods[c.module or "Unmapped"].append({
                "case_id": c.case_id, "title": c.title, "module": c.module or "Unmapped",
                "automated_by": c.automated_by or "Unassigned",
            })
        rows = [{"module": m, "count": len(v), "cases": v} for m, v in mods.items()]
        rows.sort(key=lambda x: -x["count"])
        by_person = defaultdict(int)
        for c in cases:
            by_person[c.automated_by if named(c) else "Unassigned"] += 1
        return {"by_module": rows, "total": sum(r["count"] for r in rows),
                "by_person": [{"person": p, "count": n} for p, n in sorted(by_person.items(), key=lambda x: -x[1])]}

    next_week = by_module(next_week_cases)
    backlog = by_module(backlog_cases)

    # scripted this week, per person
    scripted_tw = {p: 0 for p in TEAM}
    for c in automated:
        if c.automated_on and ws <= c.automated_on.date() <= we and c.automated_by in TEAM:
            scripted_tw[c.automated_by] += 1

    # Weekly plan snapshot: today's named-plan IS next week's plan — record it under next-week's key
    # (latest tag through the week wins). "This week's plan" = the snapshot captured for THIS week.
    snap = _load_week_plan()
    nw_by_person = {p: 0 for p in TEAM}
    for c in next_week_cases:
        nw_by_person[c.automated_by] += 1
    snap[nws.isoformat()] = nw_by_person
    _save_week_plan(snap)
    planned_tw = snap.get(ws.isoformat())  # may be None if this week was never snapshotted

    this_week_by_person = []
    for p in TEAM:
        this_week_by_person.append({
            "person": p,
            "planned": (planned_tw or {}).get(p) if planned_tw else None,  # None = no snapshot yet
            "scripted": scripted_tw[p],
        })

    today = date.today()
    days = [today - timedelta(days=k) for k in range(13, -1, -1)]
    daily = []
    for d in days:
        per = {p: 0 for p in TEAM}
        for c in automated:
            if c.automated_on and c.automated_on.date() == d and c.automated_by in TEAM:
                per[c.automated_by] += 1
        daily.append({"date": d.isoformat(), "total": sum(per.values()), **per})

    return {
        "week_start": ws.isoformat(), "week_end": we.isoformat(),
        "next_week_start": nws.isoformat(), "next_week_end": nwe.isoformat(),
        "this_week": {
            "by_person": this_week_by_person,
            "scripted_total": sum(scripted_tw.values()),
            "planned_total": (sum(planned_tw.values()) if planned_tw else None),
            "planned_known": planned_tw is not None,
        },
        "next_week": next_week,
        "backlog": backlog,
        "daily": daily,
        "plan_progress": plan_progress(db),
    }


def plan_progress(db):
    """Tracker for the Excel-supplied plan (override.planned): of each person's planned case IDs, how many are
    now Automated in TestRail (synced) + a per-day breakdown. This is the 'scripted off the plan' count that
    fills in daily from next Monday as the team flips cases Planned->Automated in TestRail. None if no plan."""
    try:
        import automation_override as OV
        ov = OV.load_override()
        if not ov.get("enabled"):
            return None
        by_person_ids = {p: set(int(x) for x in (v.get("case_ids") or []))
                         for p, v in (ov.get("planned") or {}).items()}
    except Exception:
        return None
    all_ids = set().union(*by_person_ids.values()) if by_person_ids else set()
    if not all_ids:
        return None
    cases = {c.case_id: c for c in db.query(AutomationCase).filter(AutomationCase.case_id.in_(all_ids)).all()}
    by_person = []
    for person, ids in by_person_ids.items():
        done = sum(1 for i in ids if i in cases and cases[i].automation_status == "Automated")
        by_person.append({"person": person, "planned": len(ids), "scripted": done})
    today = date.today()
    days = [today - timedelta(days=k) for k in range(13, -1, -1)]
    daily = []
    for dday in days:
        n = sum(1 for i in all_ids if i in cases and cases[i].automation_status == "Automated"
                and cases[i].automated_on and cases[i].automated_on.date() == dday)
        daily.append({"date": dday.isoformat(), "scripted": n})
    return {"by_person": by_person, "total_planned": len(all_ids),
            "total_scripted": sum(p["scripted"] for p in by_person), "daily": daily}


def write_daily_snapshot(db):
    payload = compute_metrics(db)
    today = date.today()
    row = db.query(AutomationSnapshot).filter(AutomationSnapshot.snapshot_date == today).first()
    if not row:
        row = AutomationSnapshot(snapshot_date=today)
        db.add(row)
    row.payload = payload
    db.commit()
    return payload


def growth_series(db, days=60):
    rows = db.query(AutomationSnapshot).order_by(AutomationSnapshot.snapshot_date).all()
    rows = rows[-days:]
    out = []
    for r in rows:
        ov = (r.payload or {}).get("overview", {})
        out.append({
            "date": r.snapshot_date.isoformat(),
            "automated_cases": ov.get("automated_cases", 0),
            "automated_executions": ov.get("automated_executions", 0),
            "time_saved_hours": ov.get("time_saved_hours", 0),
            "coverage_pct": ov.get("coverage_pct", 0),
        })
    return out


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------
def seed_from_excel(path, person="Vishnu VS"):
    """One-time seed of already-automated cases from the desktop 'QA Automation Scripting Report'
    xlsx. The per-module '… — Cases' sheets list Test Case ID / Automation Status / Scripted Date
    (no owner column) — this file is the user's (Vishnu's) machine export, so Automated rows are
    attributed to `person` with their real scripted date. Only updates cases already in the catalog.
    """
    import openpyxl
    bid = person_to_byid(person)
    pname = BYID_TO_PERSON.get(bid, person)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    db = SessionLocal()
    seeded = skipped = notfound = 0
    try:
        for sn in wb.sheetnames:
            if "case" not in sn.lower():
                continue
            ws = wb[sn]
            header = None
            col = {}
            for row in ws.iter_rows(values_only=True):
                if header is None:
                    if row and any(str(c).strip().lower() == "test case id" for c in row if c):
                        header = [str(c).strip().lower() if c else "" for c in row]
                        col = {name: i for i, name in enumerate(header)}
                    continue
                if not row:
                    continue
                try:
                    raw_id = row[col.get("test case id")]
                    status = row[col.get("automation status")] if "automation status" in col else None
                except (IndexError, TypeError):
                    continue
                if not raw_id or (status and str(status).strip().lower() != "automated"):
                    skipped += 1
                    continue
                digits = "".join(ch for ch in str(raw_id) if ch.isdigit())
                if not digits:
                    continue
                cid = int(digits)
                c = db.query(AutomationCase).filter(AutomationCase.case_id == cid).first()
                if not c:
                    notfound += 1
                    continue
                scripted = None
                if "scripted date" in col and row[col["scripted date"]]:
                    sd = row[col["scripted date"]]
                    if isinstance(sd, datetime):
                        scripted = sd
                    else:
                        for fmt in ("%d-%b-%Y", "%d-%b-%y", "%Y-%m-%d", "%d/%m/%Y"):
                            try:
                                scripted = datetime.strptime(str(sd).strip(), fmt)
                                break
                            except ValueError:
                                continue
                c.automation_status = "Automated"
                c.automation_status_id = 3
                c.automatable = True
                # Fill-only: never overwrite an existing (TestRail/app) attribution.
                if not c.automated_by:
                    c.automated_by = pname
                    c.automated_by_id = bid
                if scripted and not c.automated_on:
                    c.automated_on = scripted
                seeded += 1
            db.commit()
        write_daily_snapshot(db)
        return {"seeded": seeded, "skipped": skipped, "not_in_catalog": notfound, "person": pname}
    finally:
        db.close()
        wb.close()


def run_automation_sync(include_executions=True):
    db = SessionLocal()
    started = datetime.utcnow()
    try:
        refresh_automated_by_map()  # always align id->name to TestRail's field config (permanent anti-swap)
        resolve, sec_map = build_section_resolver()
        catalog = upsert_case_catalog(db, resolve, sec_map)
        exec_count = 0
        if include_executions:
            exec_count = upsert_executions(db, catalog)
        write_daily_snapshot(db)
        msg = f"cases={len(catalog)} executions={exec_count}"
        elapsed = (datetime.utcnow() - started).total_seconds()
        try:
            db.add(SyncLog(sync_source="automation", success=True, message=msg,
                           total_records=len(catalog), started_at=started,
                           completed_at=datetime.utcnow(), duration_seconds=elapsed))
            db.commit()
        except Exception:
            db.rollback()
        print(f"[automation_sync] done in {elapsed:.0f}s: {msg}")
        return {"status": "success", "cases": len(catalog), "executions": exec_count,
                "elapsed_seconds": round(elapsed, 1)}
    except Exception as e:
        try:
            db.add(SyncLog(sync_source="automation", success=False, message=str(e)[:500],
                           started_at=started, completed_at=datetime.utcnow()))
            db.commit()
        except Exception:
            db.rollback()
        print(f"[automation_sync] ERROR: {e}")
        return {"status": "error", "error": str(e)}
    finally:
        db.close()


if __name__ == "__main__":
    import sys
    inc = "--cases-only" not in sys.argv
    print(run_automation_sync(include_executions=inc))
