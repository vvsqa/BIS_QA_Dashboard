"""
Timesheet Excel Import Script

This script imports timesheet data from PmTimeTracker Excel files into the database.
Supports watching the Downloads folder for new files.

Usage:
    python sync_timesheet_to_db.py --file "path/to/PmTimeTracker_*.xlsx"
    python sync_timesheet_to_db.py --watch-downloads
    python sync_timesheet_to_db.py --import-latest
"""

import os
import sys
import argparse
import time
import re
import shutil
from datetime import datetime, date
from pathlib import Path

# Fix Unicode output on Windows
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook
from sqlalchemy.dialects.postgresql import insert

from database import SessionLocal
from models import Timesheet

# Configuration
IMPORTS_FOLDER = os.getenv("IMPORTS_FOLDER", os.path.join(os.path.dirname(__file__), "imports"))

# Get Downloads folder path
def get_downloads_folder():
    """Get the user's Downloads folder path"""
    if sys.platform == 'win32':
        import winreg
        try:
            with winreg.OpenKey(winreg.HKEY_CURRENT_USER, 
                r'SOFTWARE\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders') as key:
                downloads = winreg.QueryValueEx(key, '{374DE290-123F-4565-9164-39C4925E467B}')[0]
                return downloads
        except:
            pass
    return os.path.join(Path.home(), 'Downloads')

DOWNLOADS_FOLDER = os.getenv("DOWNLOADS_FOLDER", get_downloads_folder())

# File pattern for PmTimeTracker files
TIMETRACKER_PATTERN = re.compile(r'^PmTimeTracker_.*\.xlsx$', re.IGNORECASE)

# Column mapping
COLUMN_MAPPING = {
    "EMPLOYEE_NAME": "employee_name",
    "TICKET_ID": "ticket_id",
    "DATE": "date",
    "TIME_LOGGED": "time_logged",
    "TEAM": "team",
}


def parse_string(value):
    """Parse string value"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return str(value)
    return str(value).strip() or None


def parse_date(value):
    """Parse date from Excel cell"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        formats = ["%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y"]
        for fmt in formats:
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


def parse_time_to_minutes(value):
    """Parse time string (HH:MM:SS) to total minutes"""
    if value is None:
        return 0
    
    time_str = str(value).strip()
    if not time_str:
        return 0
    
    try:
        # Handle HH:MM:SS format
        parts = time_str.split(':')
        if len(parts) == 3:
            hours, minutes, seconds = int(parts[0]), int(parts[1]), int(parts[2])
            return hours * 60 + minutes + (1 if seconds >= 30 else 0)
        elif len(parts) == 2:
            hours, minutes = int(parts[0]), int(parts[1])
            return hours * 60 + minutes
        else:
            return int(float(time_str) * 60)  # Assume hours as decimal
    except (ValueError, IndexError):
        return 0


def parse_ticket_id(value):
    """Parse ticket ID to integer"""
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    if isinstance(value, str):
        try:
            return int(float(value.strip()))
        except ValueError:
            return None
    return None


def find_header_row(ws):
    """Find the header row in the worksheet"""
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        row_values = [str(cell).strip().upper() if cell else "" for cell in row]
        if "EMPLOYEE_NAME" in row_values or "TICKET_ID" in row_values:
            return row_idx
    return 1


def map_headers(header_row):
    """Map Excel headers to database field names"""
    column_map = {}
    for idx, header in enumerate(header_row):
        if header is None:
            continue
        header_str = str(header).strip().upper()
        for excel_header, db_field in COLUMN_MAPPING.items():
            if excel_header == header_str:
                column_map[idx] = db_field
                break
    return column_map


def import_timesheet(filepath):
    """Import timesheet data from Excel file"""
    print(f"\n{'='*60}")
    print(f"Importing Timesheet: {filepath}")
    print(f"Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*60}")
    
    if not os.path.exists(filepath):
        print(f"ERROR: File not found: {filepath}")
        return False, 0, 0
    
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        
        # Find header row
        header_row_idx = find_header_row(ws)
        print(f"Header row found at row: {header_row_idx}")
        
        # Get header row
        header_row = []
        for cell in ws[header_row_idx]:
            header_row.append(cell.value)
        
        # Map headers to database fields
        column_map = map_headers(header_row)
        print(f"Mapped columns: {column_map}")
        
        required_fields = {'employee_name', 'ticket_id', 'date'}
        if not required_fields.issubset(set(column_map.values())):
            print(f"ERROR: Missing required columns. Found: {set(column_map.values())}")
            return False, 0, 0
        
        db = SessionLocal()
        imported = 0
        updated = 0
        skipped = 0
        
        try:
            batch = []
            batch_size = 500
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
                # Extract values
                record_data = {}
                for col_idx, db_field in column_map.items():
                    if col_idx >= len(row):
                        continue
                    value = row[col_idx]
                    
                    if db_field == 'employee_name':
                        record_data[db_field] = parse_string(value)
                    elif db_field == 'ticket_id':
                        record_data[db_field] = parse_ticket_id(value)
                    elif db_field == 'date':
                        record_data[db_field] = parse_date(value)
                    elif db_field == 'time_logged':
                        record_data[db_field] = parse_string(value)
                        record_data['time_logged_minutes'] = parse_time_to_minutes(value)
                    elif db_field == 'team':
                        record_data[db_field] = parse_string(value)
                
                # Validate required fields
                if not record_data.get('employee_name') or not record_data.get('ticket_id') or not record_data.get('date'):
                    skipped += 1
                    continue
                
                record_data['created_on'] = datetime.utcnow()
                batch.append(record_data)
                
                # Process batch
                if len(batch) >= batch_size:
                    result = process_batch(db, batch)
                    imported += result['inserted']
                    updated += result['updated']
                    batch = []
                    print(f"  Progress: {imported} imported, {updated} updated...")
            
            # Process remaining batch
            if batch:
                result = process_batch(db, batch)
                imported += result['inserted']
                updated += result['updated']
            
            db.commit()
            print(f"\nImport completed successfully!")
            print(f"  New entries: {imported}")
            print(f"  Updated entries: {updated}")
            print(f"  Skipped rows: {skipped}")
            
            wb.close()
            return True, imported, updated
            
        except Exception as e:
            db.rollback()
            print(f"ERROR during import: {e}")
            import traceback
            traceback.print_exc()
            return False, 0, 0
        finally:
            db.close()
            
    except Exception as e:
        print(f"ERROR opening Excel file: {e}")
        import traceback
        traceback.print_exc()
        return False, 0, 0


def process_batch(db, batch):
    """Process a batch of timesheet entries using upsert"""
    inserted = 0
    updated = 0
    
    for record_data in batch:
        # Check if exists
        existing = db.query(Timesheet).filter(
            Timesheet.employee_name == record_data['employee_name'],
            Timesheet.ticket_id == record_data['ticket_id'],
            Timesheet.date == record_data['date']
        ).first()
        
        if existing:
            # Update existing
            existing.time_logged = record_data.get('time_logged')
            existing.time_logged_minutes = record_data.get('time_logged_minutes', 0)
            existing.team = record_data.get('team')
            updated += 1
        else:
            # Insert new
            record = Timesheet(
                employee_name=record_data['employee_name'],
                ticket_id=record_data['ticket_id'],
                date=record_data['date'],
                time_logged=record_data.get('time_logged'),
                time_logged_minutes=record_data.get('time_logged_minutes', 0),
                team=record_data.get('team'),
                created_on=datetime.utcnow()
            )
            db.add(record)
            inserted += 1
    
    db.flush()
    return {'inserted': inserted, 'updated': updated}


def import_latest_from_downloads():
    """Find and import the most recent PmTimeTracker file from Downloads"""
    if not os.path.exists(DOWNLOADS_FOLDER):
        print(f"Downloads folder not found: {DOWNLOADS_FOLDER}")
        return False, 0, 0
    
    # Find all matching files
    matching_files = []
    for f in os.listdir(DOWNLOADS_FOLDER):
        if TIMETRACKER_PATTERN.match(f):
            filepath = os.path.join(DOWNLOADS_FOLDER, f)
            matching_files.append((filepath, os.path.getmtime(filepath)))
    
    if not matching_files:
        print(f"No PmTimeTracker files found in {DOWNLOADS_FOLDER}")
        return False, 0, 0
    
    # Sort by modification time (newest first)
    matching_files.sort(key=lambda x: x[1], reverse=True)
    latest_file = matching_files[0][0]
    
    print(f"Found {len(matching_files)} PmTimeTracker file(s)")
    print(f"Importing most recent: {os.path.basename(latest_file)}")
    
    # Copy to imports folder
    Path(IMPORTS_FOLDER).mkdir(parents=True, exist_ok=True)
    dest_path = os.path.join(IMPORTS_FOLDER, os.path.basename(latest_file))
    shutil.copy2(latest_file, dest_path)
    print(f"Copied to: {dest_path}")
    
    return import_timesheet(latest_file)


def start_downloads_watcher():
    """Start watching the Downloads folder for PmTimeTracker files"""
    try:
        from watchdog.observers import Observer
        from watchdog.events import FileSystemEventHandler
    except ImportError:
        print("ERROR: watchdog package required. Install with: pip install watchdog")
        return
    
    class TimesheetFileHandler(FileSystemEventHandler):
        def __init__(self):
            self.last_modified = {}
            self.debounce_seconds = 3
        
        def on_created(self, event):
            if event.is_directory:
                return
            if TIMETRACKER_PATTERN.match(os.path.basename(event.src_path)):
                time.sleep(2)  # Wait for file to be fully written
                self._process_file(event.src_path)
        
        def on_modified(self, event):
            if event.is_directory:
                return
            if TIMETRACKER_PATTERN.match(os.path.basename(event.src_path)):
                now = time.time()
                last = self.last_modified.get(event.src_path, 0)
                if now - last < self.debounce_seconds:
                    return
                self.last_modified[event.src_path] = now
                time.sleep(2)
                self._process_file(event.src_path)
        
        def _process_file(self, filepath):
            filename = os.path.basename(filepath)
            print(f"\n[{datetime.now().strftime('%H:%M:%S')}] Detected: {filename}")
            
            # Copy to imports folder
            try:
                Path(IMPORTS_FOLDER).mkdir(parents=True, exist_ok=True)
                dest_path = os.path.join(IMPORTS_FOLDER, filename)
                shutil.copy2(filepath, dest_path)
                print(f"  Copied to: {dest_path}")
            except Exception as e:
                print(f"  Warning: Could not copy file: {e}")
            
            import_timesheet(filepath)
    
    print(f"\n{'='*60}")
    print("Timesheet Watcher Started")
    print(f"{'='*60}")
    print(f"Watching: {DOWNLOADS_FOLDER}")
    print(f"Looking for: PmTimeTracker_*.xlsx")
    print(f"Files will be copied to: {IMPORTS_FOLDER}")
    print(f"\nJust download your timesheet file and it will be")
    print(f"automatically imported into the database!")
    print(f"\nPress Ctrl+C to stop")
    print(f"{'='*60}\n")
    
    Path(IMPORTS_FOLDER).mkdir(parents=True, exist_ok=True)
    
    event_handler = TimesheetFileHandler()
    observer = Observer()
    observer.schedule(event_handler, DOWNLOADS_FOLDER, recursive=False)
    observer.start()
    
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        print("\nStopping watcher...")
        observer.stop()
    
    observer.join()
    print("Watcher stopped.")


def main():
    parser = argparse.ArgumentParser(
        description="Import timesheet data from PmTimeTracker Excel files",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python sync_timesheet_to_db.py --file "path/to/PmTimeTracker_*.xlsx"
  python sync_timesheet_to_db.py --watch-downloads
  python sync_timesheet_to_db.py --import-latest
        """
    )
    parser.add_argument('--file', '-f', type=str, help="Path to Excel file to import")
    parser.add_argument('--watch-downloads', '-w', action='store_true',
                        help="Watch Downloads folder for PmTimeTracker_* files")
    parser.add_argument('--import-latest', '-l', action='store_true',
                        help="Import the most recent PmTimeTracker file from Downloads")
    
    args = parser.parse_args()
    
    if args.file:
        success, imported, updated = import_timesheet(args.file)
        sys.exit(0 if success else 1)
    elif args.watch_downloads:
        start_downloads_watcher()
    elif args.import_latest:
        success, imported, updated = import_latest_from_downloads()
        sys.exit(0 if success else 1)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
