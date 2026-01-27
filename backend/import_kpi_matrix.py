"""
Direct import script for KPI Matrix Excel file
Run this script to import KPIs directly into the database
"""
import openpyxl
import re
import sys
import io

# Fix Unicode encoding for Windows console
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

from database import SessionLocal
from models import KPI

def import_kpi_matrix(file_path):
    """Import KPI matrix from Excel file"""
    db = SessionLocal()
    try:
        workbook = openpyxl.load_workbook(file_path)
        
        # Map sheet names to role names (normalize to match database)
        role_mapping = {
            'Software Engineer': 'SOFTWARE ENGINEER',
            'Lead': 'LEAD',
            'Project Manager': 'PROJECT MANAGER',
            'Department Heads': 'DEPARTMENT HEAD',
            'QA Engineer': 'QA ENGINEER',
            'QA Manager': 'QA MANAGER'
        }
        
        # Determine team from role
        def get_team_from_role(role_name):
            if 'QA' in role_name.upper():
                return 'QA'
            elif 'SOFTWARE ENGINEER' in role_name.upper() or 'LEAD' in role_name.upper():
                return 'DEVELOPMENT'
            else:
                return None  # For PM, Department Heads, etc.
        
        total_imported = 0
        total_updated = 0
        sheet_summary = []
        
        # Process each sheet (each sheet represents a role)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            role_name = role_mapping.get(sheet_name, sheet_name.upper())
            team = get_team_from_role(role_name)
            
            print(f"\nProcessing sheet: {sheet_name} -> Role: {role_name}, Team: {team}")
            
            imported_count = 0
            updated_count = 0
            current_kra_group = None
            
            # Process rows starting from row 2
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                # Skip if KPI name (column B) is empty
                if not row[1] or not str(row[1]).strip():
                    # If KRA Group (column A) has value, update current KRA
                    if row[0] and str(row[0]).strip():
                        current_kra_group = str(row[0]).strip()
                    continue
                
                # Extract data from columns
                # Column A: KRA Group (category)
                # Column B: KPI Name
                # Column C: Weight %
                # Column F: Evaluation Guideline (description)
                
                kpi_name = str(row[1]).strip()
                if not kpi_name or kpi_name.lower() in ['kpi', 'none']:
                    continue
                
                # Use KRA Group from column A, or current_kra_group, or None
                category = None
                if row[0] and str(row[0]).strip():
                    category = str(row[0]).strip()
                    current_kra_group = category
                elif current_kra_group:
                    category = current_kra_group
                
                # Weight from column C (convert percentage to decimal if needed)
                weight_value = row[2]
                if weight_value is not None:
                    try:
                        weight = float(weight_value)
                        # If weight is > 1, assume it's a percentage, convert to decimal
                        if weight > 1:
                            weight = weight / 100.0
                    except (ValueError, TypeError):
                        weight = 1.0
                else:
                    weight = 1.0
                
                # Description from column F (Evaluation Guideline)
                description = None
                if len(row) > 5 and row[5]:
                    description = str(row[5]).strip()
                
                # Generate KPI code from name (sanitize and make unique)
                # Role name can be up to 30 chars, so leave 70 for KPI name
                role_prefix = role_name.replace(' ', '_')[:30]
                kpi_code_base = re.sub(r'[^a-zA-Z0-9]', '_', kpi_name.upper())[:65]
                kpi_code = f"{role_prefix}_{kpi_code_base}"[:100]  # Ensure total length <= 100
                
                # Check if KPI already exists
                existing = db.query(KPI).filter(
                    KPI.kpi_code == kpi_code
                ).first()
                
                if existing:
                    # Update existing
                    existing.kpi_name = kpi_name
                    existing.description = description
                    existing.role = role_name
                    existing.team = team
                    existing.category = category
                    existing.weight = weight
                    updated_count += 1
                    # Only print first few to avoid too much output
                    if updated_count <= 5:
                        print(f"  Updated: {kpi_name[:50]}")
                else:
                    # Create new - commit immediately to avoid bulk insert conflicts
                    new_kpi = KPI(
                        kpi_code=kpi_code,
                        kpi_name=kpi_name,
                        description=description,
                        role=role_name,
                        team=team,
                        category=category,
                        weight=weight
                    )
                    db.add(new_kpi)
                    db.flush()  # Flush to get ID but don't commit yet
                    imported_count += 1
                    # Only print first few to avoid too much output
                    if imported_count <= 5:
                        print(f"  Imported: {kpi_name[:50]} (Weight: {weight})")
            
            # Commit after each sheet to avoid bulk insert issues
            db.commit()
            
            total_imported += imported_count
            total_updated += updated_count
            sheet_summary.append({
                "sheet": sheet_name,
                "role": role_name,
                "imported": imported_count,
                "updated": updated_count
            })
            
            print(f"  Summary: {imported_count} imported, {updated_count} updated")
        
        print("\n" + "="*80)
        print("IMPORT SUMMARY")
        print("="*80)
        print(f"Total imported: {total_imported}")
        print(f"Total updated: {total_updated}")
        print(f"Total processed: {total_imported + total_updated}")
        print(f"Sheets processed: {len(workbook.sheetnames)}")
        print("\nPer-sheet breakdown:")
        for summary in sheet_summary:
            print(f"  {summary['sheet']} ({summary['role']}): {summary['imported']} imported, {summary['updated']} updated")
        
        return {
            "imported": total_imported,
            "updated": total_updated,
            "total": total_imported + total_updated,
            "sheets_processed": len(workbook.sheetnames),
            "sheet_details": sheet_summary
        }
            
    except Exception as e:
        db.rollback()
        import traceback
        print(f"ERROR: {str(e)}")
        print(traceback.format_exc())
        raise
    finally:
        db.close()


if __name__ == "__main__":
    file_path = r'C:\Users\Vishnu\Downloads\KPI Matrix.xlsx'
    print("="*80)
    print("KPI MATRIX IMPORT")
    print("="*80)
    print(f"Importing from: {file_path}")
    print()
    
    try:
        result = import_kpi_matrix(file_path)
        print("\n[SUCCESS] Import completed successfully!")
    except Exception as e:
        print(f"\n[ERROR] Import failed: {e}")
