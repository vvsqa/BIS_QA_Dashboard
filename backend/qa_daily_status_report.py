"""
QA Daily Status Report - Excel Generator

Fetches live data from PM Tracker API and generates a polished Excel dashboard:
  Sheet 1: Executive Summary  - KPI cards, status pie charts (Web vs Mobile)
  Sheet 2: Team Workload      - Each QA member's ongoing tickets
  Sheet 3: QC Status Dashboard - Tickets by QC status, split Web / Mobile
  Sheet 4: Unassigned Tickets  - QC Testing tickets with blank QC Tester
  Sheet 5: QC Review Fail     - Tickets back with Dev after QA rejection

Usage:
    python qa_daily_status_report.py              # generate today's report
    python qa_daily_status_report.py --out DIR     # custom output directory
"""

import sys, os, json, argparse
from datetime import datetime
from pathlib import Path
from collections import defaultdict, Counter

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties as ChartGP
from openpyxl.chart.title import title_maker
from openpyxl.chart.text import RichText as ChartRichText
from openpyxl.chart.axis import ChartLines
from openpyxl.drawing.text import (
    Paragraph as DrawParagraph, ParagraphProperties as DrawParagraphProps,
    CharacterProperties as DrawCharProps,
)
from openpyxl.drawing.line import LineProperties

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
load_dotenv()

PM_API_URL = os.getenv(
    "PM_API_URL", "https://www.bissafety.app/rest/v.01/pm/ticket-export"
)
PM_API_KEY = os.getenv("PM_API_KEY", "")
REPORTS_DIR = Path(__file__).parent / "reports"

CLOSED_STATUSES = {"Closed", "Moved to Live"}
QC_ACTIVE_STATUSES = {
    "QC Testing", "QC Testing in Progress", "QC Testing Hold", "QC Review Fail",
}
MOBILE_KEYWORDS = {"Mobile", "SafeTapp"}

# ---------------------------------------------------------------------------
# Dark Theme Palette  (matches reference dashboard images)
# ---------------------------------------------------------------------------
# Background layers
BG_BASE      = "0B1120"   # deepest background
BG_PANEL     = "111B2E"   # card / panel surface
BG_ELEVATED  = "162036"   # elevated card / hover
BG_HEADER    = "0D1526"   # section header bar
BG_ROW_A     = "0F1729"   # table row alternating A
BG_ROW_B     = "131D30"   # table row alternating B
BG_TABLE_HDR = "1A2744"   # table column header

# Accents
ACC_CYAN     = "00D2FF"   # primary accent (like reference)
ACC_TEAL     = "00C9A7"   # secondary accent
ACC_BLUE     = "3B82F6"   # info / web
ACC_GREEN    = "22C55E"   # success / in-progress
ACC_YELLOW   = "FACC15"   # warning / hold
ACC_ORANGE   = "F97316"   # caution
ACC_RED      = "EF4444"   # fail / critical
ACC_PURPLE   = "A855F7"   # unassigned / special
ACC_PINK     = "EC4899"   # alert

# Text
TXT_WHITE    = "F1F5F9"   # primary text
TXT_LIGHT    = "94A3B8"   # secondary text
TXT_DIM      = "64748B"   # muted text
TXT_DARK     = "1E293B"   # text on light bg

# Status badge presets  (badge_bg, badge_text)
STATUS_BADGE = {
    "QC Testing":              ("1E3A5F", ACC_CYAN),
    "QC Testing in Progress":  ("14532D", ACC_GREEN),
    "QC Testing Hold":         ("422006", ACC_YELLOW),
    "QC Review Fail":          ("450A0A", ACC_RED),
    "Tested - Awaiting Fixes": ("3B0764", ACC_PURPLE),
    "BIS Testing":             ("134E4A", ACC_TEAL),
    "In Progress":             ("431407", ACC_ORANGE),
    "Approved for Live":       ("14532D", ACC_GREEN),
    "Testing In Progress":     ("14532D", ACC_GREEN),
    "Start Code Review":       ("1E3A5F", ACC_BLUE),
    "Code Review Failed":      ("450A0A", ACC_RED),
    "Code Review Passed":      ("14532D", ACC_GREEN),
    "Technical Review":        ("3B0764", ACC_PURPLE),
    "Planning":                ("1E3A5F", ACC_BLUE),
    "Backlog":                 ("1C1917", TXT_DIM),
    "NEW":                     ("1E3A5F", ACC_CYAN),
}

# KPI card accent colors (matches reference colored-top-border cards)
KPI_COLORS = [ACC_CYAN, ACC_BLUE, ACC_GREEN, ACC_ORANGE, ACC_RED,
              ACC_TEAL, ACC_GREEN, ACC_PURPLE, ACC_CYAN, ACC_YELLOW]


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------
def _f(h):
    return PatternFill(start_color=h, end_color=h, fill_type="solid")

def _border(color="1E293B", style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

SUBTLE_BORDER = _border("1E293B")
NO_BORDER = Border()
CENTER = Alignment(horizontal="center", vertical="center")
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_CENTER = Alignment(vertical="center", wrap_text=True)
LEFT_PAD = Alignment(vertical="center", wrap_text=True, indent=1)


SCREEN_COLS = 26   # A-Z  covers a typical 1920px screen
SCREEN_ROWS = 60   # covers visible scroll area

def _paint_area(ws, r1, c1, r2, c2, bg=BG_BASE):
    """Fill a rectangular area with background colour."""
    fill = _f(bg)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).fill = fill

def _paint_screen(ws, content_rows=60, bg=BG_BASE):
    """Paint the full visible screen area dark so no white is visible."""
    fill = _f(bg)
    for r in range(1, max(content_rows, SCREEN_ROWS) + 1):
        for c in range(1, SCREEN_COLS + 1):
            cell = ws.cell(row=r, column=c)
            # Only paint if cell has no fill yet (don't overwrite styled cells)
            if cell.fill.patternType is None:
                cell.fill = fill


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _header_row(ws, row, headers, start=1):
    """Dark table column header."""
    for ci, h in enumerate(headers, start):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = _f(BG_TABLE_HDR)
        c.font = Font(name="Calibri", size=9, bold=True, color=ACC_CYAN)
        c.alignment = CENTER_WRAP
        c.border = _border("1E293B")
    ws.row_dimensions[row].height = 26


def _data_row(ws, row, vals, start=1):
    """Alternating dark data row."""
    bg = BG_ROW_A if row % 2 == 0 else BG_ROW_B
    for ci, v in enumerate(vals, start):
        c = ws.cell(row=row, column=ci, value=v)
        c.fill = _f(bg)
        c.font = Font(name="Calibri", size=10, color=TXT_LIGHT)
        c.alignment = LEFT_CENTER
        c.border = _border("1E293B")
    ws.row_dimensions[row].height = 22


def _status_badge(ws, row, col, status):
    """Apply status pill styling."""
    cell = ws.cell(row=row, column=col)
    badge = STATUS_BADGE.get(status)
    if badge:
        cell.fill = _f(badge[0])
        cell.font = Font(name="Calibri", size=9, bold=True, color=badge[1])
    cell.alignment = CENTER


def _section_bar(ws, row, text, c1, c2, accent=ACC_CYAN):
    """Full-width section header with accent underline."""
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cell = ws.cell(row=row, column=c1, value=f"    {text}")
    cell.font = Font(name="Calibri", size=13, bold=True, color=accent)
    cell.fill = _f(BG_HEADER)
    cell.alignment = Alignment(vertical="center")
    cell.border = Border(
        bottom=Side(style="medium", color=accent),
        left=Side(style="thick", color=accent),
    )
    for c in range(c1 + 1, c2 + 1):
        ws.cell(row=row, column=c).fill = _f(BG_HEADER)
        ws.cell(row=row, column=c).border = Border(bottom=Side(style="medium", color=accent))
    ws.row_dimensions[row].height = 32


def _kpi_card(ws, row, col, label, value, accent, w=2):
    """
    3-row KPI card:
      row   : accent-coloured top bar
      row+1 : large value
      row+2 : label
    """
    # Top accent bar
    for c in range(col, col + w):
        ws.cell(row=row, column=c).fill = _f(accent)
        ws.cell(row=row, column=c).border = Border(
            left=Side(style="thin", color=accent),
            right=Side(style="thin", color=accent),
        )
    ws.row_dimensions[row].height = 5  # thin accent line

    # Value
    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + w - 1)
    vc = ws.cell(row=row + 1, column=col, value=value)
    vc.font = Font(name="Calibri", size=26, bold=True, color=TXT_WHITE)
    vc.alignment = CENTER
    for c in range(col, col + w):
        ws.cell(row=row + 1, column=c).fill = _f(BG_PANEL)
        ws.cell(row=row + 1, column=c).border = Border(
            left=Side(style="thin", color=accent),
            right=Side(style="thin", color=accent),
        )
    ws.row_dimensions[row + 1].height = 40

    # Label
    ws.merge_cells(start_row=row + 2, start_column=col, end_row=row + 2, end_column=col + w - 1)
    lc = ws.cell(row=row + 2, column=col, value=label)
    lc.font = Font(name="Calibri", size=8, bold=True, color=TXT_DIM)
    lc.alignment = CENTER
    for c in range(col, col + w):
        ws.cell(row=row + 2, column=c).fill = _f(BG_PANEL)
        ws.cell(row=row + 2, column=c).border = Border(
            left=Side(style="thin", color=accent),
            right=Side(style="thin", color=accent),
            bottom=Side(style="thin", color=accent),
        )
    ws.row_dimensions[row + 2].height = 18


# ---------------------------------------------------------------------------
# Data fetching
# ---------------------------------------------------------------------------
def fetch_pm_tickets():
    headers = {"authID": PM_API_KEY, "Content-Type": "application/json"}
    print("  Fetching tickets from PM Tracker API...")
    resp = requests.get(PM_API_URL, headers=headers, timeout=60)
    resp.raise_for_status()
    data = resp.json()
    tickets = data.get("data", data) if isinstance(data, dict) else data
    if isinstance(tickets, list):
        tickets = [t for t in tickets if isinstance(t, dict)]
    print(f"  Received {len(tickets)} tickets")
    return tickets


def classify_platform(subdept):
    if not subdept:
        return "Web"
    return "Mobile" if subdept.strip() in MOBILE_KEYWORDS else "Web"


def _make_text_props(color, size=900, bold=True):
    """Create ChartRichText with the given font colour and size."""
    rpr = DrawCharProps(solidFill=color, sz=size, b=bold)
    ppr = DrawParagraphProps(defRPr=rpr)
    return ChartRichText(p=[DrawParagraph(pPr=ppr, endParaRPr=rpr)])


def _dark_chart(chart, title_text=None, hide_legend=False):
    """Apply dark theme to any chart: dark backgrounds, bright readable text, data labels."""
    # Chart area background (outer frame)
    chart_gp = ChartGP()
    chart_gp.solidFill = BG_PANEL
    chart_gp.line = LineProperties(solidFill="1E293B")
    chart.graphical_properties = chart_gp

    # Plot area background (inner area)
    plot_gp = ChartGP()
    plot_gp.solidFill = BG_BASE
    plot_gp.line = LineProperties(solidFill="1E293B")
    chart.plot_area.spPr = plot_gp

    # Axis text - bright white, large enough to read
    if hasattr(chart, "x_axis") and chart.x_axis is not None:
        chart.x_axis.txPr = _make_text_props(TXT_WHITE, 850)
        chart.x_axis.spPr = ChartGP(ln=LineProperties(solidFill="334155"))
        chart.x_axis.delete = False  # ensure axis is shown

    if hasattr(chart, "y_axis") and chart.y_axis is not None:
        chart.y_axis.txPr = _make_text_props(TXT_WHITE, 850)
        chart.y_axis.spPr = ChartGP(ln=LineProperties(solidFill="334155"))
        chart.y_axis.delete = False
        # Subtle gridlines
        gl = ChartLines()
        gl.spPr = ChartGP(ln=LineProperties(solidFill="1E293B"))
        chart.y_axis.majorGridlines = gl

    # Data labels on all series - white text, show value
    # Preserve existing dLbls if already configured (e.g. showCatName)
    for s in chart.series:
        if s.dLbls is None:
            s.dLbls = DataLabelList()
            s.dLbls.showVal = True
            s.dLbls.showCatName = False
            s.dLbls.showSerName = False
            s.dLbls.txPr = _make_text_props(TXT_WHITE, 900, bold=True)

    # Title - cyan accent, large
    if title_text:
        chart.title = title_maker(title_text)
        if chart.title and hasattr(chart.title, 'txPr'):
            chart.title.txPr = _make_text_props(ACC_CYAN, 1200)

    # Legend
    if hide_legend:
        chart.legend = None
    elif chart.legend:
        chart.legend.txPr = _make_text_props(TXT_WHITE, 850)


# ---------------------------------------------------------------------------
# Sheet 1: Executive Summary
# ---------------------------------------------------------------------------
def _build_summary(wb, ongoing, unassigned, all_tickets):
    ws = wb.active
    ws.title = "Executive Summary"
    ws.sheet_properties.tabColor = ACC_CYAN

    # Layout: cols A(spacer) B-K(content) L(spacer)
    _set_col_widths(ws, [2, 18, 18, 3, 18, 18, 3, 18, 18, 3, 18, 2])
    # Shrink columns beyond content to avoid wide white gaps
    for c in range(13, SCREEN_COLS + 1):
        ws.column_dimensions[get_column_letter(c)].width = 8

    # Paint entire visible screen dark first
    _paint_area(ws, 1, 1, 95, SCREEN_COLS, BG_BASE)

    # ---- Title ----
    ws.merge_cells("B2:K2")
    t = ws.cell(row=2, column=2, value="QA DAILY STATUS REPORT")
    t.font = Font(name="Calibri", size=26, bold=True, color=ACC_CYAN)
    t.alignment = CENTER
    ws.row_dimensions[2].height = 50

    # Subtle separator line
    for c in range(2, 12):
        ws.cell(row=3, column=c).border = Border(top=Side(style="medium", color=ACC_CYAN))

    ws.merge_cells("B3:K3")
    ws.cell(row=3, column=2, value=datetime.now().strftime(
        "Generated:  %A, %B %d, %Y   |   %I:%M %p"
    )).font = Font(name="Calibri", size=10, color=TXT_DIM)
    ws.cell(row=3, column=2).alignment = CENTER

    # ---- KPI Cards  (row 5-7, 9-11) ----
    _section_bar(ws, 5, "KEY METRICS", 2, 11, ACC_CYAN)

    # Compute counts
    counts = Counter()
    for t_ in ongoing:
        counts[t_.get("Status", "")] += 1
    n_testing    = counts.get("QC Testing", 0)
    n_inprog     = counts.get("QC Testing in Progress", 0)
    n_hold       = counts.get("QC Testing Hold", 0) + counts.get("QC Testing On-hold", 0)
    n_fail       = counts.get("QC Review Fail", 0)
    n_bis        = counts.get("BIS Testing", 0)
    n_approved   = counts.get("Approved for Live", 0)
    n_awaiting   = counts.get("Tested - Awaiting Fixes", 0)
    n_team = len({(t_.get("QCTester") or "").strip() for t_ in ongoing if (t_.get("QCTester") or "").strip()})

    # Row 1:  5 cards at row 7
    _kpi_card(ws, 7, 2,  "TOTAL QA ONGOING",  len(ongoing), ACC_CYAN)
    _kpi_card(ws, 7, 5,  "QC TESTING",        n_testing,    ACC_BLUE)
    _kpi_card(ws, 7, 8,  "QC IN PROGRESS",    n_inprog,     ACC_GREEN)
    _kpi_card(ws, 7, 11, "QC ON HOLD",        n_hold,       ACC_ORANGE, w=1)
    # Adjust single-width card
    ws.cell(row=7, column=11).fill = _f(ACC_ORANGE)
    ws.merge_cells(start_row=7+1, start_column=11, end_row=7+1, end_column=11)
    ws.merge_cells(start_row=7+2, start_column=11, end_row=7+2, end_column=11)

    # Row 2:  5 cards at row 11
    _kpi_card(ws, 11, 2,  "QC REVIEW FAIL",   n_fail,       ACC_RED)
    _kpi_card(ws, 11, 5,  "BIS TESTING",      n_bis,        ACC_TEAL)
    _kpi_card(ws, 11, 8,  "UNASSIGNED",       len(unassigned), ACC_PURPLE)
    _kpi_card(ws, 11, 11, "TEAM MEMBERS",     n_team,       ACC_CYAN, w=1)
    ws.cell(row=11, column=11).fill = _f(ACC_CYAN)

    # ---- Web vs Mobile Breakdown ----
    _section_bar(ws, 15, "QC STATUS  |  WEB vs MOBILE", 2, 6, ACC_CYAN)

    qc_ordered = ["QC Testing", "QC Testing in Progress", "QC Testing Hold", "QC Review Fail"]
    row = 17
    for ci, h in enumerate(["STATUS", "WEB", "MOBILE", "TOTAL"], 2):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = _f(BG_TABLE_HDR)
        c.font = Font(name="Calibri", size=9, bold=True, color=ACC_CYAN)
        c.alignment = CENTER
        c.border = _border("1E293B")
    ws.column_dimensions["F"].width = 12
    ws.row_dimensions[row].height = 24

    web_tot = mob_tot = 0
    for si, status in enumerate(qc_ordered):
        r = row + 1 + si
        matched = [t_ for t_ in all_tickets if t_.get("Status") == status]
        wc = sum(1 for t_ in matched if classify_platform(t_.get("Subdepartment")) == "Web")
        mc = sum(1 for t_ in matched if classify_platform(t_.get("Subdepartment")) == "Mobile")
        web_tot += wc; mob_tot += mc
        badge = STATUS_BADGE.get(status, (BG_ROW_A, TXT_LIGHT))
        bg = BG_ROW_A if si % 2 == 0 else BG_ROW_B

        sc = ws.cell(row=r, column=2, value=status)
        sc.fill = _f(badge[0])
        sc.font = Font(name="Calibri", size=10, bold=True, color=badge[1])
        sc.border = _border("1E293B")
        sc.alignment = LEFT_PAD

        for ci, val in enumerate([wc, mc, wc + mc], 3):
            c = ws.cell(row=r, column=ci, value=val)
            c.fill = _f(bg)
            c.font = Font(name="Calibri", size=12, bold=True, color=TXT_WHITE)
            c.alignment = CENTER
            c.border = _border("1E293B")
        ws.row_dimensions[r].height = 24

    # Total row
    tr = row + 1 + len(qc_ordered)
    for ci, val in enumerate(["TOTAL", web_tot, mob_tot, web_tot + mob_tot], 2):
        c = ws.cell(row=tr, column=ci, value=val)
        c.fill = _f(ACC_CYAN)
        c.font = Font(name="Calibri", size=11, bold=True, color=BG_BASE)
        c.alignment = CENTER
        c.border = _border(ACC_CYAN)

    # ---- Charts ----
    # Hidden data region (far below visible area)
    DR = 100

    # -- Pie chart data --
    ws.cell(row=DR, column=2, value="Web").font = Font(size=1, color=BG_BASE)
    ws.cell(row=DR, column=3, value=web_tot).font = Font(size=1, color=BG_BASE)
    ws.cell(row=DR+1, column=2, value="Mobile").font = Font(size=1, color=BG_BASE)
    ws.cell(row=DR+1, column=3, value=mob_tot).font = Font(size=1, color=BG_BASE)

    pie = PieChart()
    pie.style = 10
    pie.width = 13
    pie.height = 9
    pie.set_categories(Reference(ws, min_col=2, min_row=DR, max_row=DR+1))
    pie.add_data(Reference(ws, min_col=3, min_row=DR, max_row=DR+1), titles_from_data=False)
    s = pie.series[0]
    for idx, clr in enumerate([ACC_BLUE, ACC_TEAL]):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = clr
        s.data_points.append(pt)
    _dark_chart(pie, "Web vs Mobile")
    # Override data labels for pie - show name + value + percent
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = True
    pie.dataLabels.showCatName = True
    pie.dataLabels.showSerName = False
    pie.dataLabels.txPr = _make_text_props(TXT_WHITE, 1100)
    ws.add_chart(pie, "G15")   # right of breakdown table

    # -- Bar chart: QC Status Distribution --
    # Use short labels so they fit on the axis
    short_labels = ["QC Testing", "In Progress", "On Hold", "Review Fail"]
    bar_colors   = [ACC_BLUE,     ACC_GREEN,     ACC_ORANGE, ACC_RED]
    for si, (status, label) in enumerate(zip(qc_ordered, short_labels)):
        ws.cell(row=DR+3+si, column=2, value=label).font = Font(size=1, color=BG_BASE)
        cnt = sum(1 for t_ in all_tickets if t_.get("Status") == status)
        ws.cell(row=DR+3+si, column=3, value=cnt).font = Font(size=1, color=BG_BASE)

    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.width = 13
    bar.height = 10
    bar.add_data(Reference(ws, min_col=3, min_row=DR+3, max_row=DR+6), titles_from_data=False)
    bar.set_categories(Reference(ws, min_col=2, min_row=DR+3, max_row=DR+6))
    bar.shape = 4
    series = bar.series[0]
    for i, clr in enumerate(bar_colors):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = clr
        series.data_points.append(pt)
    # Show status name on each bar so the chart is self-explanatory
    series.dLbls = DataLabelList()
    series.dLbls.showVal = True
    series.dLbls.showCatName = True
    series.dLbls.showSerName = False
    series.dLbls.txPr = _make_text_props(TXT_WHITE, 900, bold=True)
    _dark_chart(bar, "QC Status Distribution", hide_legend=True)
    ws.add_chart(bar, "B26")

    # -- Horizontal bar chart: Ticket Load per QA Member --
    by_tester = defaultdict(int)
    for t_ in ongoing:
        qc = (t_.get("QCTester") or "").strip()
        if qc:
            by_tester[qc] += 1
    sorted_t = sorted(by_tester.items(), key=lambda x: -x[1])

    for i, (name, cnt) in enumerate(sorted_t):
        # Use first name + last initial for readability
        parts = name.split()
        short = parts[0] + (" " + parts[-1][0] + "." if len(parts) > 1 else "")
        ws.cell(row=DR+10+i, column=2, value=short).font = Font(size=1, color=BG_BASE)
        ws.cell(row=DR+10+i, column=3, value=cnt).font = Font(size=1, color=BG_BASE)

    if sorted_t:
        tbar = BarChart()
        tbar.type = "bar"
        tbar.style = 10
        tbar.width = 13
        tbar.height = max(10, 2 + len(sorted_t))
        end = DR+10+len(sorted_t)-1
        tbar.add_data(Reference(ws, min_col=3, min_row=DR+10, max_row=end), titles_from_data=False)
        tbar.set_categories(Reference(ws, min_col=2, min_row=DR+10, max_row=end))
        tbar.series[0].graphicalProperties.solidFill = ACC_CYAN
        # Show member name + count on each bar
        s = tbar.series[0]
        s.dLbls = DataLabelList()
        s.dLbls.showVal = True
        s.dLbls.showCatName = True
        s.dLbls.showSerName = False
        s.dLbls.txPr = _make_text_props(TXT_WHITE, 850, bold=True)
        _dark_chart(tbar, "Ticket Load per QA Member", hide_legend=True)
        ws.add_chart(tbar, "G38")

    _paint_screen(ws, 110)
    ws.sheet_view.showGridLines = False


# ---------------------------------------------------------------------------
# Sheet 2: Team Workload
# ---------------------------------------------------------------------------
def _build_team_sheet(wb, ongoing):
    ws = wb.create_sheet("Team Workload")
    ws.sheet_properties.tabColor = ACC_BLUE

    _set_col_widths(ws, [4, 22, 11, 72, 24, 18, 18, 13, 20])
    COLS = 9
    for c in range(COLS + 1, SCREEN_COLS + 1):
        ws.column_dimensions[get_column_letter(c)].width = 8

    by_tester = defaultdict(list)
    for t_ in ongoing:
        qc = (t_.get("QCTester") or "").strip()
        if qc:
            by_tester[qc].append(t_)

    # Title
    row = 1
    _paint_area(ws, 1, 1, 3, COLS, BG_BASE)
    ws.merge_cells(f"A1:I1")
    c = ws.cell(row=1, column=1, value="    QA TEAM WORKLOAD")
    c.font = Font(name="Calibri", size=22, bold=True, color=ACC_CYAN)
    c.alignment = Alignment(vertical="center")
    for col in range(1, COLS+1):
        ws.cell(row=1, column=col).fill = _f(BG_BASE)
        ws.cell(row=1, column=col).border = Border(bottom=Side(style="medium", color=ACC_CYAN))
    ws.row_dimensions[1].height = 48

    ws.merge_cells("A2:I2")
    ws.cell(row=2, column=1, value=f"    {datetime.now().strftime('%A, %B %d, %Y  |  %I:%M %p')}   |   {len(by_tester)} members   |   {len(ongoing)} tickets").font = Font(name="Calibri", size=10, color=TXT_DIM)

    row = 4
    headers = ["#", "QC Tester", "Ticket", "Title", "Status", "Priority", "Module", "ETA", "Assignee"]
    status_order = {
        "QC Testing in Progress": 0, "Testing In Progress": 1,
        "QC Testing": 2, "QC Review Fail": 3, "Tested - Awaiting Fixes": 4,
        "QC Testing Hold": 5, "BIS Testing": 6, "In Progress": 7, "Approved for Live": 8,
    }

    for tester in sorted(by_tester.keys()):
        tasks = by_tester[tester]
        tasks.sort(key=lambda x: status_order.get(x.get("Status", ""), 99))

        # Tester name bar (merge A-H, keep I separate for count)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COLS-1)
        cell = ws.cell(row=row, column=1, value=f"    {tester}")
        cell.font = Font(name="Calibri", size=12, bold=True, color=TXT_WHITE)
        cell.fill = _f(BG_HEADER)
        cell.alignment = Alignment(vertical="center")
        cell.border = Border(left=Side(style="thick", color=ACC_CYAN), bottom=Side(style="thin", color="1E293B"))
        for c in range(2, COLS):
            ws.cell(row=row, column=c).fill = _f(BG_HEADER)
            ws.cell(row=row, column=c).border = Border(bottom=Side(style="thin", color="1E293B"))
        # Ticket count badge on right column
        cnt_cell = ws.cell(row=row, column=COLS)
        cnt_cell.value = f"{len(tasks)} tickets"
        cnt_cell.font = Font(name="Calibri", size=10, bold=True, color=ACC_CYAN)
        cnt_cell.fill = _f(BG_HEADER)
        cnt_cell.alignment = Alignment(horizontal="right", vertical="center")
        cnt_cell.border = Border(bottom=Side(style="thin", color="1E293B"))
        ws.row_dimensions[row].height = 30
        row += 1

        _header_row(ws, row, headers)
        row += 1

        for idx, t_ in enumerate(tasks, 1):
            status = t_.get("Status", "")
            _data_row(ws, row, [
                idx, tester, t_.get("TicketNumber", ""),
                (t_.get("TicketTitle") or "")[:80],
                status, t_.get("Priority", ""),
                t_.get("Subdepartment", ""),
                t_.get("ETA") or "-",
                t_.get("CurrentAssignee", ""),
            ])
            _status_badge(ws, row, 5, status)
            # Cyan left accent
            ws.cell(row=row, column=1).border = Border(
                left=Side(style="thick", color=ACC_CYAN),
                top=Side(style="thin", color="1E293B"),
                bottom=Side(style="thin", color="1E293B"),
                right=Side(style="thin", color="1E293B"),
            )
            row += 1

        # Gap row
        _paint_area(ws, row, 1, row, SCREEN_COLS, BG_BASE)
        ws.row_dimensions[row].height = 10
        row += 1

    _paint_area(ws, row, 1, row + 3, SCREEN_COLS, BG_BASE)
    _paint_screen(ws, row + 3)
    ws.sheet_view.showGridLines = False


# ---------------------------------------------------------------------------
# Sheet 3: QC Status Dashboard
# ---------------------------------------------------------------------------
def _build_status_dashboard(wb, all_tickets):
    ws = wb.create_sheet("QC Status Dashboard")
    ws.sheet_properties.tabColor = ACC_GREEN
    COLS = 9
    for c in range(COLS + 1, SCREEN_COLS + 1):
        ws.column_dimensions[get_column_letter(c)].width = 8

    _set_col_widths(ws, [4, 11, 72, 24, 18, 18, 22, 13, 20])

    statuses = ["QC Testing", "QC Testing in Progress", "QC Testing Hold", "QC Review Fail"]
    platforms = ["Web", "Mobile"]
    plat_colors = {"Web": ACC_BLUE, "Mobile": ACC_TEAL}

    # Title
    _paint_area(ws, 1, 1, 3, COLS, BG_BASE)
    ws.merge_cells("A1:I1")
    ws.cell(row=1, column=1, value="    QC STATUS DASHBOARD").font = Font(name="Calibri", size=22, bold=True, color=ACC_GREEN)
    ws.cell(row=1, column=1).alignment = Alignment(vertical="center")
    for c in range(1, COLS+1):
        ws.cell(row=1, column=c).fill = _f(BG_BASE)
        ws.cell(row=1, column=c).border = Border(bottom=Side(style="medium", color=ACC_GREEN))
    ws.row_dimensions[1].height = 48

    ws.merge_cells("A2:I2")
    ws.cell(row=2, column=1, value=f"    {datetime.now().strftime('%B %d, %Y')}   |   Breakdown by QC Status   |   Web & Mobile").font = Font(name="Calibri", size=10, color=TXT_DIM)

    row = 4
    headers = ["#", "Ticket", "Title", "Status", "Priority", "Module", "QC Tester", "ETA", "Assignee"]

    for platform in platforms:
        p_color = plat_colors[platform]

        # Platform mega header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COLS)
        cell = ws.cell(row=row, column=1, value=f"    {platform.upper()} PLATFORM")
        cell.font = Font(name="Calibri", size=16, bold=True, color=p_color)
        cell.fill = _f(BG_HEADER)
        cell.alignment = Alignment(vertical="center")
        cell.border = Border(left=Side(style="thick", color=p_color), bottom=Side(style="medium", color=p_color))
        for c in range(2, COLS+1):
            ws.cell(row=row, column=c).fill = _f(BG_HEADER)
            ws.cell(row=row, column=c).border = Border(bottom=Side(style="medium", color=p_color))
        ws.row_dimensions[row].height = 38
        row += 1

        for status in statuses:
            matched = [
                t_ for t_ in all_tickets
                if t_.get("Status") == status
                and classify_platform(t_.get("Subdepartment")) == platform
            ]
            if not matched:
                continue

            badge = STATUS_BADGE.get(status, (BG_ROW_A, TXT_LIGHT))

            # Status sub-header
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COLS)
            sc = ws.cell(row=row, column=1, value=f"      {status}    ({len(matched)})")
            sc.font = Font(name="Calibri", size=11, bold=True, color=badge[1])
            sc.fill = _f(badge[0])
            sc.alignment = Alignment(vertical="center")
            sc.border = Border(left=Side(style="thick", color=p_color))
            for c in range(2, COLS+1):
                ws.cell(row=row, column=c).fill = _f(badge[0])
            ws.row_dimensions[row].height = 26
            row += 1

            _header_row(ws, row, headers)
            row += 1

            matched.sort(key=lambda x: (x.get("Priority", ""), x.get("TicketNumber", 0)))
            for idx, t_ in enumerate(matched, 1):
                _data_row(ws, row, [
                    idx, t_.get("TicketNumber", ""),
                    (t_.get("TicketTitle") or "")[:80], status,
                    t_.get("Priority", ""), t_.get("Subdepartment", ""),
                    (t_.get("QCTester") or "").strip() or "UNASSIGNED",
                    t_.get("ETA") or "-", t_.get("CurrentAssignee", ""),
                ])
                _status_badge(ws, row, 4, status)
                ws.cell(row=row, column=1).border = Border(
                    left=Side(style="thick", color=p_color),
                    top=Side(style="thin", color="1E293B"),
                    bottom=Side(style="thin", color="1E293B"),
                    right=Side(style="thin", color="1E293B"),
                )
                row += 1

            _paint_area(ws, row, 1, row, COLS, BG_BASE)
            ws.row_dimensions[row].height = 8
            row += 1

        _paint_area(ws, row, 1, row, COLS, BG_BASE)
        ws.row_dimensions[row].height = 14
        row += 1

    _paint_area(ws, row, 1, row + 3, COLS, BG_BASE)
    _paint_screen(ws, row + 3)
    ws.sheet_view.showGridLines = False


# ---------------------------------------------------------------------------
# Sheet 4: Unassigned Tickets
# ---------------------------------------------------------------------------
def _build_unassigned(wb, unassigned):
    ws = wb.create_sheet("Unassigned Tickets")
    ws.sheet_properties.tabColor = ACC_PURPLE
    COLS = 10
    for c in range(COLS + 1, SCREEN_COLS + 1):
        ws.column_dimensions[get_column_letter(c)].width = 8

    _set_col_widths(ws, [4, 11, 72, 24, 18, 18, 14, 22, 13, 20])

    # Title
    ws.merge_cells("A1:J1")
    ws.cell(row=1, column=1, value="    UNASSIGNED QC TICKETS").font = Font(name="Calibri", size=22, bold=True, color=ACC_PURPLE)
    ws.cell(row=1, column=1).alignment = Alignment(vertical="center")
    for c in range(1, COLS+1):
        ws.cell(row=1, column=c).fill = _f(BG_BASE)
        ws.cell(row=1, column=c).border = Border(bottom=Side(style="medium", color=ACC_PURPLE))
    ws.row_dimensions[1].height = 48

    ws.merge_cells("A2:J2")
    ws.cell(row=2, column=1, value="    Tickets in QC Testing status with no QC Tester assigned").font = Font(name="Calibri", size=10, color=TXT_DIM, italic=True)

    # Alert banner
    ws.merge_cells("A3:J3")
    alert = ws.cell(row=3, column=1, value=f"    {len(unassigned)} TICKETS NEED IMMEDIATE ASSIGNMENT")
    alert.font = Font(name="Calibri", size=14, bold=True, color=BG_BASE)
    alert.alignment = Alignment(vertical="center")
    for c in range(1, COLS+1):
        ws.cell(row=3, column=c).fill = _f(ACC_YELLOW)
    ws.row_dimensions[3].height = 36

    row = 5
    headers = ["#", "Ticket", "Title", "Status", "Priority", "Module", "Type", "Reported By", "ETA", "Assignee"]
    _header_row(ws, row, headers)
    row += 1

    unassigned.sort(key=lambda x: (x.get("Priority", ""), x.get("TicketNumber", 0)))
    for idx, t_ in enumerate(unassigned, 1):
        _data_row(ws, row, [
            idx, t_.get("TicketNumber", ""),
            (t_.get("TicketTitle") or "")[:80], t_.get("Status", ""),
            t_.get("Priority", ""), t_.get("Subdepartment", ""),
            t_.get("Type", ""), t_.get("ReportedBy", ""),
            t_.get("ETA") or "-", t_.get("CurrentAssignee", ""),
        ])
        _status_badge(ws, row, 4, t_.get("Status", ""))
        ws.cell(row=row, column=1).border = Border(
            left=Side(style="thick", color=ACC_PURPLE),
            top=Side(style="thin", color="1E293B"),
            bottom=Side(style="thin", color="1E293B"),
            right=Side(style="thin", color="1E293B"),
        )
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COLS)
    ws.cell(row=row, column=1, value=f"    Total unassigned: {len(unassigned)}").font = Font(name="Calibri", size=12, bold=True, color=ACC_PURPLE)
    for c in range(1, COLS+1):
        ws.cell(row=row, column=c).fill = _f(BG_HEADER)
    _paint_area(ws, row+1, 1, row+3, COLS, BG_BASE)
    _paint_screen(ws, row + 3)
    ws.sheet_view.showGridLines = False


# ---------------------------------------------------------------------------
# Sheet 5: QC Review Fail
# ---------------------------------------------------------------------------
def _build_review_fail(wb, all_tickets):
    ws = wb.create_sheet("QC Review Fail (Dev)")
    ws.sheet_properties.tabColor = ACC_RED
    COLS = 9
    for c in range(COLS + 1, SCREEN_COLS + 1):
        ws.column_dimensions[get_column_letter(c)].width = 8

    _set_col_widths(ws, [4, 11, 72, 18, 18, 22, 22, 22, 13])

    qc_fail = [t_ for t_ in all_tickets if t_.get("Status") == "QC Review Fail"]
    web_fail = [t_ for t_ in qc_fail if classify_platform(t_.get("Subdepartment")) == "Web"]
    mob_fail = [t_ for t_ in qc_fail if classify_platform(t_.get("Subdepartment")) == "Mobile"]

    _paint_area(ws, 1, 1, 4, COLS, BG_BASE)

    # Title
    ws.merge_cells("A1:I1")
    ws.cell(row=1, column=1, value="    QC REVIEW FAIL  |  BACK WITH DEVELOPERS").font = Font(name="Calibri", size=22, bold=True, color=ACC_RED)
    ws.cell(row=1, column=1).alignment = Alignment(vertical="center")
    for c in range(1, COLS+1):
        ws.cell(row=1, column=c).fill = _f(BG_BASE)
        ws.cell(row=1, column=c).border = Border(bottom=Side(style="medium", color=ACC_RED))
    ws.row_dimensions[1].height = 48

    ws.merge_cells("A2:I2")
    ws.cell(row=2, column=1, value="    Tickets that failed QA and need developer fixes").font = Font(name="Calibri", size=10, color=TXT_DIM, italic=True)

    # Alert
    ws.merge_cells("A3:I3")
    ws.cell(row=3, column=1, value=f"    {len(qc_fail)} TICKETS NEED DEVELOPER ATTENTION").font = Font(name="Calibri", size=14, bold=True, color=TXT_WHITE)
    for c in range(1, COLS+1):
        ws.cell(row=3, column=c).fill = _f(ACC_RED)
    ws.row_dimensions[3].height = 36

    row = 5
    headers = ["#", "Ticket", "Title", "Priority", "Module", "QC Tester", "Backend Dev", "Frontend Dev", "ETA"]

    for platform, tlist, p_color in [("WEB", web_fail, ACC_BLUE), ("MOBILE", mob_fail, ACC_TEAL)]:
        if not tlist:
            continue

        _section_bar(ws, row, f"{platform}    ({len(tlist)} tickets)", 1, COLS, p_color)
        row += 1
        _header_row(ws, row, headers)
        row += 1

        for idx, t_ in enumerate(tlist, 1):
            _data_row(ws, row, [
                idx, t_.get("TicketNumber", ""),
                (t_.get("TicketTitle") or "")[:80],
                t_.get("Priority", ""), t_.get("Subdepartment", ""),
                (t_.get("QCTester") or "").strip(),
                t_.get("BackendDeveloper", ""),
                t_.get("FrontendDeveloper", ""),
                t_.get("ETA") or "-",
            ])
            ws.cell(row=row, column=1).border = Border(
                left=Side(style="thick", color=ACC_RED),
                top=Side(style="thin", color="1E293B"),
                bottom=Side(style="thin", color="1E293B"),
                right=Side(style="thin", color="1E293B"),
            )
            row += 1

        _paint_area(ws, row, 1, row, COLS, BG_BASE)
        ws.row_dimensions[row].height = 10
        row += 1

    # Summary footer
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COLS)
    ws.cell(row=row, column=1, value=f"    Total: {len(qc_fail)}    |    Web: {len(web_fail)}    |    Mobile: {len(mob_fail)}").font = Font(name="Calibri", size=12, bold=True, color=ACC_RED)
    for c in range(1, COLS+1):
        ws.cell(row=row, column=c).fill = _f(BG_HEADER)
    _paint_area(ws, row+1, 1, row+3, COLS, BG_BASE)
    _paint_screen(ws, row + 3)
    ws.sheet_view.showGridLines = False


# ---------------------------------------------------------------------------
# Main generator
# ---------------------------------------------------------------------------
def generate_report(output_dir=None):
    print("\n" + "=" * 60)
    print("  QA DAILY STATUS REPORT GENERATOR")
    print("=" * 60)

    all_tickets = fetch_pm_tickets()

    ongoing = [
        t for t in all_tickets
        if (t.get("QCTester") or "").strip()
        and (t.get("Status") or "").strip() not in CLOSED_STATUSES
    ]
    unassigned = [
        t for t in all_tickets
        if not (t.get("QCTester") or "").strip()
        and (t.get("Status") or "").strip() in QC_ACTIVE_STATUSES
    ]

    print(f"  Ongoing QA tickets: {len(ongoing)}")
    print(f"  Unassigned QC tickets: {len(unassigned)}")
    print("  Building Excel report...")

    wb = Workbook()
    _build_summary(wb, ongoing, unassigned, all_tickets)
    _build_team_sheet(wb, ongoing)
    _build_status_dashboard(wb, all_tickets)
    _build_unassigned(wb, unassigned)
    _build_review_fail(wb, all_tickets)

    out_dir = Path(output_dir) if output_dir else REPORTS_DIR
    out_dir.mkdir(parents=True, exist_ok=True)

    # Single file - always overwrite the same name
    filepath = out_dir / "QA_Daily_Status_Report.xlsx"

    # Close the file in Excel if it's open (Windows only)
    if sys.platform == "win32" and filepath.exists():
        try:
            import subprocess
            subprocess.run(
                ["taskkill", "/F", "/FI", f"WINDOWTITLE eq *QA_Daily_Status*"],
                capture_output=True, timeout=5,
            )
        except Exception:
            pass
        # Brief pause so Excel releases the file lock
        import time; time.sleep(1)

    # Delete old file if it exists
    if filepath.exists():
        try:
            filepath.unlink()
        except PermissionError:
            # File still locked - use timestamped name as fallback
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = out_dir / f"QA_Daily_Status_{ts}.xlsx"
            print(f"  Previous report is open, saving as: {filepath.name}")

    wb.save(str(filepath))

    print(f"\n  Report saved: {filepath}")
    print(f"  File size: {filepath.stat().st_size / 1024:.1f} KB")
    print("=" * 60 + "\n")
    return filepath


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="QA Daily Status Report Generator")
    parser.add_argument("--out", type=str, default=None, help="Output directory")
    args = parser.parse_args()
    try:
        path = generate_report(output_dir=args.out)
        if sys.platform == "win32":
            os.startfile(str(path))
    except Exception as e:
        print(f"\n  ERROR: {e}", file=sys.stderr)
        import traceback; traceback.print_exc()
        sys.exit(1)
