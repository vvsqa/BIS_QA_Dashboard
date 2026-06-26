"""Manual (Excel-supplied) override for the automation Team + Planning tabs.

TestRail attribution is currently unreliable, so per-person weekly/scripted counts and planned
cases are taken from uploaded Excel files and shown (clearly badged) until TestRail is fixed.
Sources:
- automation_dashboard.xlsx  -> Vishnu VS team numbers (Weekly Report tab: this-week, all-time, weekly trend).
- UA_Planned_Cases_By_Team.xlsx -> planned-next-week case lists per person (Vishnu / Varsha / Vivek).
The store lives at data/automation_override.json; toggle with `enabled`.
"""
import os
import re
import json
from datetime import datetime

from automation_sync import BYID_TO_PERSON, person_to_byid

OVERRIDE_FILE = os.path.join(os.path.dirname(__file__), "data", "automation_override.json")
TEAM_ORDER = ["Vishnu VS", "Varsha Dcruz P", "Vivek V Nair"]


def _canon(name):
    """Map an Excel name ('Vishnu V S', 'Varsha D', 'Vivek V') to the app's canonical name."""
    bid = person_to_byid(name or "")
    return BYID_TO_PERSON.get(bid)


def load_override():
    if os.path.exists(OVERRIDE_FILE):
        try:
            with open(OVERRIDE_FILE, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def save_override(d):
    os.makedirs(os.path.dirname(OVERRIDE_FILE), exist_ok=True)
    with open(OVERRIDE_FILE, "w", encoding="utf-8") as f:
        json.dump(d, f, indent=2, ensure_ascii=False)


def is_enabled():
    return bool(load_override().get("enabled"))


def set_enabled(flag):
    d = load_override()
    d["enabled"] = bool(flag)
    save_override(d)
    return d


# ---------------------------------------------------------------------------
# Parsers
# ---------------------------------------------------------------------------
def parse_vishnu_dashboard(path):
    """Weekly Report tab -> {this_week, total_scripted, weekly:[{week(iso), scripted}]}."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Weekly Report"]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()

    this_week = total = executed = None
    pass_rate = None
    for i, r in enumerate(rows):
        joined = " ".join(str(c) for c in r if c not in (None, ""))
        if "PERFORMANCE OVERVIEW" in joined:
            for j in range(i + 1, min(i + 4, len(rows))):
                nums = [c for c in rows[j] if isinstance(c, (int, float))]
                if len(nums) >= 2:
                    this_week, total = int(nums[0]), int(nums[1])
                    if len(nums) >= 3:
                        executed = int(nums[2])
                    for c in rows[j]:
                        if isinstance(c, str) and c.strip().endswith("%"):
                            try:
                                pass_rate = float(c.strip().rstrip("%"))
                            except ValueError:
                                pass
                    break
            break

    weekly = []
    for i, r in enumerate(rows):
        low = [str(c).strip().lower() if c else "" for c in r]
        if "week" in low and "scripted" in low and "cumulative" in low:
            wkc = low.index("week")
            sc = low.index("scripted")
            for rr in rows[i + 1:]:
                joined = " ".join(str(c) for c in rr if c not in (None, ""))
                if "grand total" in joined.lower():
                    break
                val = rr[sc] if sc < len(rr) else None
                label = rr[wkc] if wkc < len(rr) else None
                if isinstance(val, (int, float)):
                    m = re.search(r"(\d{2}-[A-Za-z]{3}-\d{4})", str(label or ""))
                    wk = (str(label)[:14] if label else f"W{len(weekly) + 1}")
                    if m:
                        try:
                            wk = datetime.strptime(m.group(1), "%d-%b-%Y").date().isoformat()
                        except ValueError:
                            pass
                    weekly.append({"week": wk, "scripted": int(val)})
            break

    return {
        "this_week": this_week or 0,
        "total_scripted": total or (weekly[-1]["scripted"] if weekly else 0),
        "this_week_executed": executed or 0,
        "this_week_pass_rate": pass_rate or 0,
        "this_week_passed": round((executed or 0) * (pass_rate or 0) / 100),
        "weekly": weekly,
    }


MODULE_SHEET_MAP = {
    "classroom calendar": "Classroom Calendar",
    "online course": "Online Course",
    "user administration": "User Management",
}


def parse_module_counts(path):
    """Per-module 'X — Cases' sheets -> {app_module: {automated, not_automatable, total, automatable}}.
    These are the authoritative automated-case counts (TestRail attribution is wrong)."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = {}
    for sn in wb.sheetnames:
        if "cases" not in sn.lower():
            continue
        key = re.sub(r"[^a-z ]", "", sn.lower().replace("cases", "")).strip()
        mod = MODULE_SHEET_MAP.get(key)
        if not mod:
            continue
        ws = wb[sn]
        sidx = None
        stt = {}
        for r in ws.iter_rows(values_only=True):
            if sidx is None:
                low = [str(c).strip().lower() if c else "" for c in r]
                if "automation status" in low:
                    sidx = low.index("automation status")
                continue
            if sidx < len(r) and r[sidx]:
                st = str(r[sidx]).strip()
                stt[st] = stt.get(st, 0) + 1
        automated = stt.get("Automated", 0)
        not_auto = stt.get("Not Automatable", 0)
        total = sum(stt.values())
        out[mod] = {"automated": automated, "not_automatable": not_auto,
                    "total": total, "automatable": total - not_auto}
    wb.close()
    return out


def parse_planned(path):
    """Each sheet 'Name (N)' -> {canonical_person: {count, case_ids}}."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = {}
    for sn in wb.sheetnames:
        person = _canon(re.sub(r"\(.*?\)", "", sn).strip())
        if not person:
            continue
        ws = wb[sn]
        case_ids = []
        cidx = None
        for r in ws.iter_rows(values_only=True):
            if cidx is None:
                low = [str(c).strip().lower() if c else "" for c in r]
                if "case id" in low:
                    cidx = low.index("case id")
                continue
            if cidx < len(r) and r[cidx]:
                digits = "".join(ch for ch in str(r[cidx]) if ch.isdigit())
                if digits:
                    case_ids.append(int(digits))
        out[person] = {"count": len(case_ids), "case_ids": case_ids}
    wb.close()
    return out


def build_override(dashboard_path, planned_path, enabled=True):
    existing = load_override()
    team = dict(existing.get("team", {}))  # preserve manual per-person entries (e.g. Varsha/Vivek pasted scripted)
    modules = dict(existing.get("modules", {}))
    if dashboard_path and os.path.exists(dashboard_path):
        team["Vishnu VS"] = parse_vishnu_dashboard(dashboard_path)
        modules.update(parse_module_counts(dashboard_path))
    planned = parse_planned(planned_path) if (planned_path and os.path.exists(planned_path)) else existing.get("planned", {})
    d = {
        "enabled": enabled,
        "imported_at": datetime.now().date().isoformat(),
        "team": team,
        "planned": planned,
        "modules": modules,
    }
    save_override(d)
    return d


def set_manual_scripted(person, case_ids, total_scripted=None, week=None):
    """Record a person's THIS-WEEK scripted cases (pasted list) into the override team store.
    Cumulative (total_scripted) stays pending (None) until supplied. Preserved across rebuilds."""
    from datetime import date, timedelta
    canon = _canon(person) or person
    case_ids = [int(c) for c in case_ids]
    d = load_override()
    team = d.setdefault("team", {})
    wk = week or (date.today() - timedelta(days=date.today().weekday())).isoformat()
    entry = team.get(canon, {})
    entry["this_week"] = len(case_ids)
    entry["this_week_case_ids"] = case_ids
    if total_scripted is not None:
        entry["total_scripted"] = int(total_scripted)
    else:
        entry.setdefault("total_scripted", None)  # cumulative pending
    weekly = [w for w in entry.get("weekly", []) if w.get("week") != wk]
    weekly.append({"week": wk, "scripted": len(case_ids)})
    entry["weekly"] = sorted(weekly, key=lambda w: w["week"])
    team[canon] = entry
    d["enabled"] = True
    save_override(d)
    return {"person": canon, "this_week": len(case_ids), "total_scripted": entry["total_scripted"]}


def set_weekly(person, scripted, backlog=None, cumulative=None, week=None):
    """Count-based manual weekly entry (no case IDs): this-week scripted + backlog + cumulative for a person.
    Used when the team supplies counts (not pasted case lists). Preserved across rebuilds."""
    from datetime import date, timedelta
    canon = _canon(person) or person
    d = load_override()
    team = d.setdefault("team", {})
    wk = week or (date.today() - timedelta(days=date.today().weekday())).isoformat()
    entry = team.get(canon, {})
    entry["this_week"] = int(scripted)
    if backlog is not None:
        entry["backlog"] = int(backlog)
    if cumulative is not None:
        entry["total_scripted"] = int(cumulative)
    else:
        entry.setdefault("total_scripted", None)
    weekly = [w for w in entry.get("weekly", []) if w.get("week") != wk]
    weekly.append({"week": wk, "scripted": int(scripted)})
    entry["weekly"] = sorted(weekly, key=lambda w: w["week"])
    team[canon] = entry
    save_override(d)
    return {"person": canon, "this_week": int(scripted), "backlog": entry.get("backlog"),
            "total_scripted": entry.get("total_scripted")}


def set_cumulative(person, total):
    """Override ONLY a person's all-time cumulative (total_scripted) — e.g. to account for historical
    cases automated by them but never tagged with custom_case_automated_by in TestRail. Leaves this_week
    (and the weekly series) absent so apply_to_team keeps those on LIVE synced data."""
    canon = _canon(person) or person
    d = load_override()
    team = d.setdefault("team", {})
    entry = team.get(canon, {})
    entry["total_scripted"] = int(total)
    team[canon] = entry
    save_override(d)
    return {"person": canon, "total_scripted": int(total), "this_week": entry.get("this_week")}


# ---------------------------------------------------------------------------
# Apply to computed payloads
# ---------------------------------------------------------------------------
def apply_to_team(payload):
    ov = load_override()
    if not ov.get("enabled"):
        return payload
    team_ov = ov.get("team", {})
    out = []
    # Override ONLY the persons that have an entry; everyone else keeps their LIVE computed values
    # (e.g. Vishnu stays on synced TestRail data).
    for m in payload.get("members", []):
        m = dict(m)
        person = m.get("name")
        if person in team_ov:
            t = team_ov[person]
            # Field-level override: only replace a field when the override sets it (not None), so a
            # person can be overridden on SOME fields and stay LIVE on others (e.g. Vishnu's all-time
            # cumulative is fixed at 1180 while his this_week / weekly remain live synced).
            if t.get("this_week") is not None:
                m["this_week"] = t.get("this_week")
            if t.get("total_scripted") is not None:
                m["total_scripted"] = t.get("total_scripted")
            if t.get("weekly"):
                m["weekly"] = t.get("weekly")
            if t.get("backlog") is not None:
                m["backlog"] = t.get("backlog")
            m["pending"] = False
        out.append(m)
    payload["members"] = out
    payload["override"] = {"active": True, "imported_at": ov.get("imported_at")}
    return payload


def apply_to_modules(modules_list, overview):
    """Correct per-module automated-case counts (Classroom Calendar / Online Course / User
    Management) from the uploaded sheet, and recompute the overview coverage rollups. Does NOT
    touch executions/utilization/time-saved (those come from run data)."""
    ov = load_override()
    if not ov.get("enabled"):
        return
    mod_ov = ov.get("modules", {})
    if not mod_ov:
        return
    for m in modules_list:
        o = mod_ov.get(m.get("module"))
        if not o:
            continue
        autobl = o.get("automatable", o.get("automated", 0))
        m["automated_cases"] = o.get("automated", 0)
        m["total_cases"] = o.get("total", m.get("total_cases", 0))
        m["automatable_cases"] = autobl
        m["coverage_pct"] = round(o["automated"] / autobl * 100, 1) if autobl else 0.0
        m["source"] = "uploaded"
    if overview is not None:
        overview["automated_cases"] = sum(m.get("automated_cases", 0) for m in modules_list)
        overview["total_cases"] = sum(m.get("total_cases", 0) for m in modules_list)
        overview["automatable_cases"] = sum(m.get("automatable_cases", 0) for m in modules_list)
        autobl = overview["automatable_cases"]
        overview["coverage_pct"] = round(overview["automated_cases"] / autobl * 100, 1) if autobl else 0.0


def apply_to_planning(payload, module_lookup=None):
    ov = load_override()
    if not ov.get("enabled"):
        return payload
    planned_ov = ov.get("planned", {})
    team_ov = ov.get("team", {})

    by_person, total = [], 0
    for person in TEAM_ORDER:
        cnt = planned_ov.get(person, {}).get("count", 0)
        total += cnt if isinstance(cnt, (int, float)) else 0
        aw = team_ov.get(person, {}).get("this_week")  # None = pending upload (no uploaded scripted data yet)
        by_person.append({"person": person, "planned": cnt, "automated_this_week": aw})

    backlog, by_mod = [], {}
    for person in TEAM_ORDER:
        for cid in planned_ov.get(person, {}).get("case_ids", []):
            mod = (module_lookup or {}).get(cid) or "User Administration"
            by_mod[mod] = by_mod.get(mod, 0) + 1
            if len(backlog) < 500:
                backlog.append({"case_id": cid, "title": "", "module": mod, "automated_by": person})

    payload["by_person"] = by_person
    payload["planned_total"] = total
    payload["automated_this_week_total"] = sum((p["automated_this_week"] or 0) for p in by_person)
    # Manual planned-by-module entries (counts supplied without case lists, e.g. "Forms 200 next week").
    for mod, c in (ov.get("planned_by_module") or {}).items():
        if mod in by_mod:
            by_mod[mod] += c
        else:
            by_mod[mod] = c
    payload["by_module"] = sorted(({"module": k, "planned": v} for k, v in by_mod.items()), key=lambda x: -x["planned"])
    payload["backlog"] = backlog
    payload["override"] = {"active": True, "imported_at": ov.get("imported_at")}
    return payload
