from fastapi import FastAPI, Query, HTTPException, UploadFile, File, Body, Request, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from sqlalchemy.orm import Session
from sqlalchemy import String, func, or_
from datetime import datetime, timedelta, date
from typing import Optional, List, Dict, Any, Tuple
from collections import defaultdict
from pydantic import BaseModel, Field
import re
import calendar
import tempfile
import os
import shutil
import time
import logging

logger = logging.getLogger(__name__)

# Load .env from backend directory so PM_API_KEY etc. are available
from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))
CLIENT_DEFAULT_PASSWORD = os.getenv("CLIENT_DEFAULT_PASSWORD", "BIS@123")

# Module ids clients can be granted access to. Null/empty in DB = use this default.
DEFAULT_CLIENT_ALLOWED_MODULES = ["home", "ticket_dashboard", "tickets", "all_bugs"]

from database import SessionLocal
from models import (
    Bug, TestPlan, TestRun, TestCase, TestResult, TicketTracking, QATicketFlag,
    Employee, EmployeeSkill, Timesheet, EmployeeGoal, EmployeeReview, KPI, KPIRating,
    TicketStatusHistory, TicketPriorityHistory, BugStatusHistory,
    EnhancedTimesheet, LeaveEntry, TimeSheetSubmission, TimeSheetEntry, TimeSheetEntryReview, TimeSheetApprovalLog, PlannedTask, WeeklyPlan,
    EmployeeNameMapping, Holiday, SyncLog,
    DevPlanningWeek, DevPlannedTask, DevPlannedAllocation, DevPlanningAuditLog,
    QAPlanningWeek, QAPlannedTask, QAPlannedAllocation, QATaskHoldHistory,
    User, AdminConfig, ClientProfile,
    AutomationTestRun, AutomationTestCase,
    PerformanceSnapshot, QAFlowSnapshot, TicketMovementSnapshot,
)
from auth import (
    authenticate_user, hash_password, create_access_token,
    get_current_user, get_current_user_optional, require_role, require_reports_access,
    get_visible_employee_ids, is_planning_lead, can_access_employee, get_db,
    can_access_employee_profile, can_edit_employee_profile, can_manage_tasks_for,
)
from dev_planning import (
    get_planning_week_dates,
    is_working_day,
    get_working_days_list,
    get_leave_hours_for_employees,
    get_allocated_hours_for_week,
    get_development_employees,
    get_or_create_planning_week,
    get_planning_week,
    simulate_allocation_distribution,
    get_available_hours_on_date,
    get_next_available_date,
    get_availability_summary,
    check_duplicate_task,
    create_allocations_for_task,
    log_audit,
    HOURS_PER_DAY,
    HOURS_PER_WEEK,
    ALLOCATION_PCT_VALID,
    GENERIC_CATEGORIES,
    TASK_CATEGORIES,
    PLANNING_STATES,
)
from google_sheets_sync import GoogleSheetsSync, get_sheets_sync_status
from sheets_scheduler import get_scheduler, start_auto_sync, stop_auto_sync
from pm_tracker_scheduler import start_pm_auto_sync, stop_pm_auto_sync, get_pm_scheduler_status, unpause_pm_sync
from sync_health import sync_health
from redmine_scheduler import start_redmine_auto_sync, stop_redmine_auto_sync, get_redmine_scheduler_status
from monthly_report_scheduler import (
    start_monthly_report_scheduler,
    stop_monthly_report_scheduler,
    get_monthly_report_status,
    get_monthly_report_scheduler,
)
from testrail_scheduler import (
    start_testrail_auto_sync,
    stop_testrail_auto_sync,
    get_testrail_scheduler_status,
    trigger_testrail_sync_now,
)
from google_sheets_export import (
    start_sheets_export_scheduler,
    stop_sheets_export_scheduler,
    trigger_manual_export,
    get_sheets_exporter,
    GOOGLE_EXPORT_AVAILABLE,
)
from pm_sync_runner import run_pm_api_sync
from pm_api_sync import PMApiClient
from sync_utils import upsert_tickets, log_sync_operation, get_last_sync_info, cleanup_sync_history
from config.pm_tracker_config import ENABLE_SYNC_LOGGING
from sync_redmine_to_db import sync_redmine_bugs
from qa_planning import (
    get_qa_overview_data, get_qa_employees, get_qa_employees_for_planner,
    get_qa_qc_review_fail_data, get_qc_fail_count, QC_FAIL_STATUSES, PRIORITY_ORDER as QA_PRIORITY_ORDER,
    get_qa_allocated_hours_for_week, get_or_create_qa_planning_week, get_qa_planning_week,
    _allocation_not_released,
    get_qa_available_hours_on_date,
    get_qa_next_available_date,
    get_qa_availability_summary,
    simulate_qa_allocation_distribution,
    create_qa_allocations_for_task,
    QA_QC_STATUSES,
    CLOSED_STATUSES,
    get_qa_ticket_suggestions,
    calculate_qc_priority_score,
    get_status_durations,
    get_qc_cycle_details,
    get_qc_cycles_summary,
    get_ageing_overview,
    get_ageing_bottlenecks,
    get_ticket_flow_rate,
    get_bis_to_closed_tracking,
    get_qa_activity_summary,
)


# ===== PYDANTIC MODELS =====

class EmployeeCreate(BaseModel):
    employee_id: str
    name: str
    email: str
    role: Optional[str] = None
    location: Optional[str] = None
    date_of_joining: Optional[datetime] = None
    team: str
    category: Optional[str] = None
    employment_status: Optional[str] = "Ongoing Employee"
    lead: Optional[str] = None

class EmployeeUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    role: Optional[str] = None  # Job title (SOFTWARE ENGINEER, etc.)
    designation: Optional[str] = None  # Designation (e.g., "Software Engineer", "QA Lead")
    location: Optional[str] = None
    mode_of_work: Optional[str] = None  # Onsite, Remote, Hybrid
    date_of_joining: Optional[datetime] = None
    team: Optional[str] = None
    category: Optional[str] = None
    employment_status: Optional[str] = None  # Ongoing Employee, Serving Notice Period, Resigned
    lead: Optional[str] = None
    manager: Optional[str] = None
    previous_experience: Optional[float] = None
    bis_introduced_date: Optional[datetime] = None
    platform: Optional[str] = None
    photo_url: Optional[str] = None
    is_active: Optional[bool] = None
    mapping_data: Optional[dict] = None
    access_role: Optional[str] = None  # Access role (ADMIN, MANAGER_DEV, MANAGER_QA, LEAD_DEV, LEAD_QA, EMPLOYEE, CLIENT)
    # Notice period tracking
    resignation_date: Optional[datetime] = None  # Date resignation was submitted
    expected_lwd: Optional[datetime] = None  # Expected Last Working Day (manual override)

class GoalCreate(BaseModel):
    goal_type: str  # 'goal', 'strength', 'improvement'
    title: str
    description: Optional[str] = None
    target_date: Optional[date] = None
    created_by: str

class GoalUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    target_date: Optional[date] = None
    status: Optional[str] = None
    progress: Optional[int] = None

class ReviewCreate(BaseModel):
    review_period: str
    review_date: date
    technical_rating: int
    productivity_rating: int
    quality_rating: int
    communication_rating: int
    strengths_summary: Optional[str] = None
    improvements_summary: Optional[str] = None
    manager_comments: Optional[str] = None
    recommendation: str
    salary_hike_percent: Optional[float] = None
    reviewed_by: str

class KPICreate(BaseModel):
    kpi_code: str
    kpi_name: str
    description: Optional[str] = None
    role: str
    team: Optional[str] = None
    category: Optional[str] = None
    weight: Optional[float] = 1.0

class KPIRatingCreate(BaseModel):
    kpi_id: int
    quarter: str  # "2025-Q1"
    rating: Optional[float] = None  # Deprecated, use manager_rating
    self_rating: Optional[float] = None
    lead_rating: Optional[float] = None
    manager_rating: Optional[float] = None
    self_comments: Optional[str] = None
    lead_comments: Optional[str] = None
    manager_comments: Optional[str] = None
    rated_by: str  # "self", "lead", or "manager"
    salary_hike_percent: Optional[float] = None
    reviewed_by: str


class LoginRequest(BaseModel):
    email: str
    password: str


class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str


app = FastAPI()

# Configure logging
if ENABLE_SYNC_LOGGING:
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )

# Static uploads (employee photos)
UPLOADS_ROOT = os.path.join(os.path.dirname(__file__), "uploads")
PROFILE_PHOTO_DIR = os.path.join(UPLOADS_ROOT, "profile_photos")
os.makedirs(PROFILE_PHOTO_DIR, exist_ok=True)
app.mount("/uploads", StaticFiles(directory=UPLOADS_ROOT), name="uploads")

# GZIP compression for faster network transfer
from starlette.middleware.gzip import GZipMiddleware
app.add_middleware(GZipMiddleware, minimum_size=1000)

# Enable CORS for frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition"],
)

# Startup and shutdown events for auto-sync schedulers
@app.on_event("startup")
async def startup_event():
    """Start auto-sync schedulers on application startup."""
    try:
        if start_auto_sync():
            print("[OK] Google Sheets auto-sync started")
        else:
            print("[INFO] Google Sheets auto-sync is disabled (set SHEETS_AUTO_SYNC=true to enable)")
    except Exception as e:
        print(f"[WARNING] Failed to start Google Sheets auto-sync: {e}")
    try:
        if start_pm_auto_sync():
            print("[OK] PM Tracker API auto-sync started (data kept up to date)")
        else:
            print("[INFO] PM Tracker auto-sync is disabled (set PM_AUTO_SYNC=true to enable)")
    except Exception as e:
        print(f"[WARNING] Failed to start PM Tracker auto-sync: {e}")
    try:
        if start_redmine_auto_sync():
            print("[OK] Redmine auto-sync started (bug data kept up to date)")
        else:
            print("[INFO] Redmine auto-sync is disabled (set REDMINE_AUTO_SYNC=true to enable)")
    except Exception as e:
        print(f"[WARNING] Failed to start Redmine auto-sync: {e}")
    try:
        if start_testrail_auto_sync():
            print("[OK] TestRail auto-sync started (test data kept up to date)")
        else:
            print("[INFO] TestRail auto-sync is disabled (set TESTRAIL_AUTO_SYNC=true to enable)")
    except Exception as e:
        print(f"[WARNING] Failed to start TestRail auto-sync: {e}")
    try:
        if start_sheets_export_scheduler():
            print("[OK] Google Sheets Export auto-sync started (exports every hour)")
        else:
            print("[INFO] Google Sheets Export is disabled (set SHEETS_EXPORT_AUTO_SYNC=true to enable)")
    except Exception as e:
        print(f"[WARNING] Failed to start Google Sheets Export scheduler: {e}")
    try:
        if start_monthly_report_scheduler():
            print("[OK] Monthly Team Report scheduler started (runs on last day of every month)")
        else:
            print("[INFO] Monthly Team Report scheduler is disabled (set MONTHLY_REPORT_AUTO=true to enable)")
    except Exception as e:
        print(f"[WARNING] Failed to start Monthly Team Report scheduler: {e}")
    # One-time Google Sheets sync on startup (ensures fresh data regardless of auto-sync setting)
    import threading
    def _startup_sheets_sync():
        try:
            sync = GoogleSheetsSync()
            result = sync.sync_all()
            print(f"[OK] Startup Google Sheets sync completed: {result}")
        except Exception as e:
            print(f"[WARNING] Startup Google Sheets sync failed: {e}")
    threading.Thread(target=_startup_sheets_sync, daemon=True).start()
    print("[INFO] Google Sheets startup sync triggered in background")

@app.on_event("shutdown")
async def shutdown_event():
    """Stop schedulers on application shutdown."""
    try:
        stop_auto_sync()
        print("[OK] Google Sheets auto-sync stopped")
    except Exception as e:
        print(f"[WARNING] Error stopping Google Sheets auto-sync: {e}")
    try:
        stop_pm_auto_sync()
        print("[OK] PM Tracker auto-sync stopped")
    except Exception as e:
        print(f"[WARNING] Error stopping PM Tracker auto-sync: {e}")
    try:
        stop_redmine_auto_sync()
        print("[OK] Redmine auto-sync stopped")
    except Exception as e:
        print(f"[WARNING] Error stopping Redmine auto-sync: {e}")
    try:
        stop_testrail_auto_sync()
        print("[OK] TestRail auto-sync stopped")
    except Exception as e:
        print(f"[WARNING] Error stopping TestRail auto-sync: {e}")
    try:
        stop_sheets_export_scheduler()
        print("[OK] Google Sheets Export stopped")
    except Exception as e:
        print(f"[WARNING] Error stopping Google Sheets Export: {e}")
    try:
        stop_monthly_report_scheduler()
        print("[OK] Monthly Team Report scheduler stopped")
    except Exception as e:
        print(f"[WARNING] Error stopping Monthly Team Report scheduler: {e}")


# ===== AUTH ENDPOINTS =====

def _login_impl(req: LoginRequest):
    """Authenticate with email and password. Returns JWT and user info."""
    db = SessionLocal()
    try:
        user = authenticate_user(db, req.email, req.password)
        if not user:
            raise HTTPException(status_code=401, detail="Invalid email or password")
        token = create_access_token({
            "sub": user["email"],
            "role": user["role"],
            "employee_id": user.get("employee_id"),
            "id": str(user.get("id", "")),
        })
        name = None
        designation = None
        if user.get("employee_id"):
            emp = db.query(Employee).filter(Employee.employee_id == user["employee_id"]).first()
            if emp:
                name = emp.name
                designation = emp.role
        elif user.get("role") == "CLIENT":
            client = db.query(ClientProfile).filter(ClientProfile.email == user["email"]).first()
            if client:
                name = client.name
                designation = "CLIENT"
        return {
            "access_token": token,
            "token_type": "bearer",
            "user": {
                "email": user["email"],
                "role": user["role"],
                "employee_id": user.get("employee_id"),
                "password_changed_at": user.get("password_changed_at"),
                "name": name,
                "designation": designation,
            },
        }
    finally:
        db.close()


@app.post("/auth/login")
def login(req: LoginRequest):
    """Authenticate with email and password. Returns JWT and user info."""
    return _login_impl(req)


@app.post("/login")
def login_alias(req: LoginRequest):
    """Alias for /auth/login for hosting/proxy setups that expect POST /login."""
    return _login_impl(req)


@app.get("/auth/me")
def auth_me(current_user: dict = Depends(get_current_user)):
    """Get current authenticated user info with permissions."""
    db = SessionLocal()
    try:
        name = None
        designation = None
        team = None
        password_changed_at = None
        allowed_modules = None
        user_role = current_user.get("role", "EMPLOYEE")
        employee_id = current_user.get("employee_id")
        
        planning_team = None  # QA or DEVELOPMENT for Task Planning tab visibility
        if employee_id:
            emp = db.query(Employee).filter(Employee.employee_id == employee_id).first()
            if emp:
                name = emp.name
                designation = emp.role
                team = emp.team  # Get employee's team (DEVELOPMENT or QA)
                # Derive planning_team from User.role (LEAD_QA, LEAD_DEV, etc.) then Employee.team
                ur = (user_role or "").upper()
                if "LEAD_QA" in ur or "MANAGER_QA" in ur:
                    planning_team = "QA"
                elif "LEAD_DEV" in ur or "MANAGER_DEV" in ur:
                    planning_team = "DEVELOPMENT"
                elif (emp.team or "").strip().upper() == "QA":
                    planning_team = "QA"
                elif (emp.team or "").strip().upper() == "DEVELOPMENT":
                    planning_team = "DEVELOPMENT"
                else:
                    planning_team = "DEVELOPMENT"
            # Get password_changed_at from User table
            user_record = db.query(User).filter(User.employee_id == employee_id).first()
            if user_record:
                password_changed_at = user_record.password_changed_at.isoformat() if user_record.password_changed_at else None
        elif user_role == "CLIENT":
            client = db.query(ClientProfile).filter(ClientProfile.email == current_user["email"]).first()
            if client:
                name = client.name
                designation = "CLIENT"
                allowed_modules = client.allowed_modules if isinstance(client.allowed_modules, list) else DEFAULT_CLIENT_ALLOWED_MODULES
            if allowed_modules is None:
                allowed_modules = DEFAULT_CLIENT_ALLOWED_MODULES
            user_record = db.query(User).filter(User.email == current_user["email"]).first()
            if user_record:
                password_changed_at = user_record.password_changed_at.isoformat() if user_record.password_changed_at else None
        
        # Determine permission levels
        is_admin = user_role == "ADMIN"
        is_manager = "MANAGER" in user_role
        is_lead = "LEAD" in user_role
        
        out = {
            "email": current_user["email"],
            "role": current_user["role"],
            "employee_id": employee_id,
            "name": name,
            "designation": designation,
            "team": team,  # Include team for calendar filtering
            "planning_team": planning_team,  # QA or DEVELOPMENT for Task Planning (Dev/QA tab visibility)
            "password_changed_at": password_changed_at,
            "permissions": {
                "can_view_all_data": True,  # All users can see all data (names, ticket info, etc.)
                "can_access_all_profiles": is_admin or is_manager,  # Admin/Manager can access all profiles
                "can_edit_all_profiles": is_admin or is_manager,  # Admin/Manager can edit all profiles
                "can_manage_all_tasks": is_admin or is_manager,  # Admin/Manager can manage all tasks
                "can_access_reportee_profiles": is_lead,  # Lead can access reportee profiles
                "can_edit_reportee_profiles": is_lead,  # Lead can edit reportee profiles
                "can_manage_team_tasks": is_lead,  # Lead can manage team tasks
                "can_access_calendar": True,  # Everyone can access calendar
                "can_access_reports": is_admin or is_manager or is_lead,  # Reports accessible to managers, leads, admins
                "can_change_user_roles": is_admin or is_manager,  # Admin/Manager can change user roles
                "can_view_all_teams_calendar": is_admin or is_manager,  # Admin/Manager can view all teams in calendar
            }
        }
        if allowed_modules is not None:
            out["allowed_modules"] = allowed_modules
        return out
    finally:
        db.close()


@app.get("/auth/check-profile-access/{employee_id}")
def check_profile_access(employee_id: str, current_user: dict = Depends(get_current_user)):
    """Check if current user can access a specific employee's profile."""
    db = SessionLocal()
    try:
        employee = db.query(Employee).filter(Employee.employee_id == employee_id).first()
        if not employee:
            return {"can_access": False, "can_edit": False, "reason": "Employee not found"}
        
        can_access = can_access_employee_profile(db, current_user, employee_id)
        can_edit = can_edit_employee_profile(db, current_user, employee_id)
        can_tasks = can_manage_tasks_for(db, current_user, employee_id)
        
        return {
            "can_access": can_access,
            "can_edit": can_edit,
            "can_manage_tasks": can_tasks,
            "employee_name": employee.name,
        }
    finally:
        db.close()


@app.post("/auth/change-password")
def change_password(
    req: ChangePasswordRequest,
    current_user: dict = Depends(get_current_user),
):
    """Change password (first-time or voluntary)."""
    db = SessionLocal()
    try:
        email = current_user["email"]
        role = current_user["role"]

        if role == "ADMIN":
            admin = db.query(AdminConfig).filter(AdminConfig.email == email).first()
            if not admin or not authenticate_user(db, email, req.current_password):
                raise HTTPException(status_code=400, detail="Current password is incorrect")
            admin.password_hash = hash_password(req.new_password)
        else:
            user = db.query(User).filter(User.email == email).first()
            if not user or not authenticate_user(db, email, req.current_password):
                raise HTTPException(status_code=400, detail="Current password is incorrect")
            user.password_hash = hash_password(req.new_password)
            user.password_changed_at = datetime.utcnow()

        db.commit()
        return {"message": "Password changed successfully"}
    finally:
        db.close()


# ===== ADMIN ENDPOINTS =====

class AdminConfigUpdate(BaseModel):
    email: Optional[str] = None
    new_password: Optional[str] = None


class ClientProfileCreate(BaseModel):
    name: str
    email: str
    is_active: Optional[bool] = True


class ClientProfileUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    is_active: Optional[bool] = None


class ClientProfileImportItem(BaseModel):
    name: str
    email: str


class ClientProfileImportBody(BaseModel):
    profiles: List[ClientProfileImportItem]


@app.get("/admin/users")
def admin_list_users(current_user: dict = Depends(require_role(["ADMIN", "MANAGER_DEV", "MANAGER_QA"]))):
    """List all users (admin and managers)."""
    db = SessionLocal()
    try:
        users = db.query(User).order_by(User.email).all()
        result = []
        for u in users:
            emp_name = None
            if u.employee_id:
                emp = db.query(Employee).filter(Employee.employee_id == u.employee_id).first()
                emp_name = emp.name if emp else None
            result.append({
                "id": u.id,
                "email": u.email,
                "role": u.role,
                "employee_id": u.employee_id,
                "employee_name": emp_name,
                "password_changed_at": u.password_changed_at.isoformat() if u.password_changed_at else None,
            })
        return {"users": result}
    finally:
        db.close()


@app.post("/admin/users/{user_id}/reset-password")
def admin_reset_user_password(
    user_id: int,
    current_user: dict = Depends(require_role(["ADMIN", "MANAGER_DEV", "MANAGER_QA"])),
):
    """Reset user password. CLIENT users reset to CLIENT_DEFAULT_PASSWORD; others to employee_id/email-prefix."""
    db = SessionLocal()
    try:
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        if user.role == "CLIENT":
            default_password = CLIENT_DEFAULT_PASSWORD
        else:
            default_password = user.employee_id or user.email.split("@")[0] or "changeme"
        user.password_hash = hash_password(default_password)
        user.password_changed_at = None
        db.commit()
        return {"message": "Password reset successfully", "default_password": default_password}
    finally:
        db.close()


@app.get("/admin/clients")
def admin_list_clients(current_user: dict = Depends(require_role(["ADMIN"]))):
    """List all client profiles (admin only)."""
    db = SessionLocal()
    try:
        clients = db.query(ClientProfile).order_by(ClientProfile.name.asc()).all()
        result = []
        for client in clients:
            linked_user = db.query(User).filter(User.email == client.email).first()
            allowed = client.allowed_modules if isinstance(client.allowed_modules, list) else None
            result.append({
                "id": client.id,
                "name": client.name,
                "email": client.email,
                "is_active": bool(client.is_active),
                "allowed_modules": allowed if allowed is not None else DEFAULT_CLIENT_ALLOWED_MODULES,
                "created_on": client.created_on.isoformat() if client.created_on else None,
                "updated_on": client.updated_on.isoformat() if client.updated_on else None,
                "password_changed_at": linked_user.password_changed_at.isoformat() if linked_user and linked_user.password_changed_at else None,
            })
        return {"clients": result}
    finally:
        db.close()


@app.post("/admin/clients")
def admin_create_client(
    body: ClientProfileCreate,
    current_user: dict = Depends(require_role(["ADMIN"])),
):
    """Create a client profile and corresponding CLIENT login."""
    db = SessionLocal()
    try:
        email = (body.email or "").strip().lower()
        name = (body.name or "").strip()
        if not email or not name:
            raise HTTPException(status_code=400, detail="Name and email are required")

        existing_client = db.query(ClientProfile).filter(ClientProfile.email == email).first()
        if existing_client:
            raise HTTPException(status_code=400, detail="Client email already exists")

        existing_user = db.query(User).filter(User.email == email).first()
        if existing_user:
            raise HTTPException(status_code=400, detail="A user with this email already exists")

        client = ClientProfile(
            name=name,
            email=email,
            is_active=bool(body.is_active if body.is_active is not None else True),
        )
        db.add(client)

        db.add(User(
            email=email,
            password_hash=hash_password(CLIENT_DEFAULT_PASSWORD),
            role="CLIENT",
            employee_id=None,
            password_changed_at=None,
        ))

        db.commit()
        db.refresh(client)
        return {
            "message": "Client profile created successfully",
            "client": {
                "id": client.id,
                "name": client.name,
                "email": client.email,
                "is_active": client.is_active,
            },
            "default_password": CLIENT_DEFAULT_PASSWORD,
        }
    finally:
        db.close()


@app.put("/admin/clients/{client_id}")
def admin_update_client(
    client_id: int,
    body: ClientProfileUpdate,
    current_user: dict = Depends(require_role(["ADMIN"])),
):
    """Update client profile details (admin only)."""
    db = SessionLocal()
    try:
        client = db.query(ClientProfile).filter(ClientProfile.id == client_id).first()
        if not client:
            raise HTTPException(status_code=404, detail="Client profile not found")

        update_data = body.dict(exclude_unset=True)
        new_email = None
        if "email" in update_data and update_data["email"] is not None:
            new_email = update_data["email"].strip().lower()
            if not new_email:
                raise HTTPException(status_code=400, detail="Email cannot be empty")
            exists = db.query(ClientProfile).filter(ClientProfile.email == new_email, ClientProfile.id != client_id).first()
            if exists:
                raise HTTPException(status_code=400, detail="Another client already uses this email")
            existing_user = db.query(User).filter(User.email == new_email).first()
            if existing_user and existing_user.email != client.email:
                raise HTTPException(status_code=400, detail="A user with this email already exists")

        old_email = client.email
        if "name" in update_data and update_data["name"] is not None:
            name = update_data["name"].strip()
            if not name:
                raise HTTPException(status_code=400, detail="Name cannot be empty")
            client.name = name
        if new_email is not None:
            client.email = new_email
        if "is_active" in update_data and update_data["is_active"] is not None:
            client.is_active = bool(update_data["is_active"])

        linked_user = db.query(User).filter(User.email == old_email).first()
        if linked_user:
            linked_user.email = client.email
            linked_user.role = "CLIENT"
            linked_user.employee_id = None

        db.commit()
        db.refresh(client)
        return {
            "message": "Client profile updated successfully",
            "client": {
                "id": client.id,
                "name": client.name,
                "email": client.email,
                "is_active": client.is_active,
                "password_changed_at": linked_user.password_changed_at.isoformat() if linked_user and linked_user.password_changed_at else None,
            },
        }
    finally:
        db.close()


@app.post("/admin/clients/{client_id}/reset-password")
def admin_reset_client_password(
    client_id: int,
    current_user: dict = Depends(require_role(["ADMIN"])),
):
    """Reset client password to default and require change on next login."""
    db = SessionLocal()
    try:
        client = db.query(ClientProfile).filter(ClientProfile.id == client_id).first()
        if not client:
            raise HTTPException(status_code=404, detail="Client profile not found")

        user = db.query(User).filter(User.email == client.email).first()
        if not user:
            user = User(
                email=client.email,
                password_hash=hash_password(CLIENT_DEFAULT_PASSWORD),
                role="CLIENT",
                employee_id=None,
                password_changed_at=None,
            )
            db.add(user)
        else:
            user.password_hash = hash_password(CLIENT_DEFAULT_PASSWORD)
            user.role = "CLIENT"
            user.employee_id = None
            user.password_changed_at = None

        db.commit()
        return {"message": "Client password reset successfully", "default_password": CLIENT_DEFAULT_PASSWORD}
    finally:
        db.close()


class ClientModulesUpdate(BaseModel):
    allowed_modules: List[str]


@app.put("/admin/clients/{client_id}/modules")
def admin_update_client_modules(
    client_id: int,
    body: ClientModulesUpdate,
    current_user: dict = Depends(require_role(["ADMIN"])),
):
    """Set which modules a client can access (admin only)."""
    db = SessionLocal()
    try:
        client = db.query(ClientProfile).filter(ClientProfile.id == client_id).first()
        if not client:
            raise HTTPException(status_code=404, detail="Client profile not found")
        client.allowed_modules = body.allowed_modules if body.allowed_modules else None
        db.commit()
        db.refresh(client)
        return {
            "message": "Client module access updated",
            "client_id": client.id,
            "allowed_modules": client.allowed_modules if client.allowed_modules is not None else DEFAULT_CLIENT_ALLOWED_MODULES,
        }
    finally:
        db.close()


@app.post("/admin/clients/import")
def admin_import_clients(
    body: ClientProfileImportBody,
    current_user: dict = Depends(require_role(["ADMIN"])),
):
    """Bulk create client profiles and CLIENT logins. Skips emails that already exist."""
    db = SessionLocal()
    created = 0
    skipped = []
    errors = []
    try:
        for item in body.profiles:
            email = (item.email or "").strip().lower()
            name = (item.name or "").strip()
            if not email or not name:
                errors.append(f"Skipped empty name/email: {name!r} / {email!r}")
                continue
            if db.query(ClientProfile).filter(ClientProfile.email == email).first():
                skipped.append(email)
                continue
            if db.query(User).filter(User.email == email).first():
                skipped.append(email)
                continue
            client = ClientProfile(
                name=name,
                email=email,
                is_active=True,
            )
            db.add(client)
            db.add(User(
                email=email,
                password_hash=hash_password(CLIENT_DEFAULT_PASSWORD),
                role="CLIENT",
                employee_id=None,
                password_changed_at=None,
            ))
            created += 1
        db.commit()
        return {
            "message": f"Import complete: {created} created, {len(skipped)} skipped (already exist)",
            "created": created,
            "skipped": skipped,
            "errors": errors[:20],
            "default_password": CLIENT_DEFAULT_PASSWORD,
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


@app.put("/admin/config")
def admin_update_config(
    body: AdminConfigUpdate,
    current_user: dict = Depends(require_role(["ADMIN"])),
):
    """Update admin email and/or password. Admin only."""
    db = SessionLocal()
    try:
        admin = db.query(AdminConfig).first()
        if not admin:
            raise HTTPException(status_code=404, detail="Admin config not found")
        if body.email:
            admin.email = body.email.strip().lower()
        if body.new_password:
            admin.password_hash = hash_password(body.new_password)
        admin.updated_at = datetime.utcnow()
        db.commit()
        return {"message": "Admin config updated successfully"}
    finally:
        db.close()


# ===== TEAM CLASSIFICATION HELPER =====

def get_team_classification(db: Session) -> dict:
    """
    Build a mapping of employee names to their team (DEV/QA).
    Names not in the employee database are classified as 'BIS Team' (client).
    """
    employees = db.query(Employee.name, Employee.team).filter(Employee.is_active == True).all()
    team_map = {}
    for emp in employees:
        if emp.name:
            raw_name = emp.name.strip()
            team_map[raw_name.lower()] = emp.team or "UNKNOWN"
            team_map[_normalize_person_name(raw_name)] = emp.team or "UNKNOWN"
            team_map[_compact_person_name(raw_name)] = emp.team or "UNKNOWN"
    return team_map


def classify_person(name: str, team_map: dict) -> str:
    """
    Classify a person's team based on the team_map.
    Returns 'DEV', 'QA', or 'BIS Team' (for client/external people).
    """
    if not name:
        return "Unknown"
    
    raw_name = name.strip().lower()
    norm_name = _normalize_person_name(name)
    compact_name = _compact_person_name(name)

    team = team_map.get(raw_name) or team_map.get(norm_name) or team_map.get(compact_name)
    if team:
        if team == "DEVELOPMENT":
            return "DEV"
        elif team == "QA":
            return "QA"
        return team
    
    # Not in employee database = BIS Team (client)
    return "BIS Team"


@app.get("/")
def root():
    return {
        "status": "ok",
        "message": "QA Dashboard API",
        "timesheet_health": "/timesheet/health",
        "docs": "OpenAPI at /docs",
    }

@app.get("/bugs")
def get_bugs(
    ticket_id: Optional[int] = Query(None),
    environment: str = Query("All"),
    platform: str = Query("All"),
    only_open: bool = Query(False)
):
    db: Session = SessionLocal()

    query = db.query(Bug)
    
    # Only filter by ticket_id if provided and not 0 (0 is used as placeholder for "all")
    if ticket_id is not None and ticket_id != 0:
        query = query.filter(Bug.ticket_id == ticket_id)

    if environment != "All":
        query = query.filter(Bug.environment == environment)

    if platform != "All":
        query = query.filter(Bug.platform == platform)

    if only_open:
        query = query.filter(
            Bug.status.in_(["New", "Reopened", "Fixed", "Assigned to Dev"])
        )

    bugs = query.all()
    db.close()
    return bugs

@app.get("/bugs/summary")
def bug_summary(
    ticket_id: int = Query(...),
    environment: str = Query("All"),
    platform: str = Query("All")
):
    db: Session = SessionLocal()

    query = db.query(Bug).filter(Bug.ticket_id == ticket_id)

    if environment != "All":
        query = query.filter(Bug.environment == environment)

    if platform != "All":
        query = query.filter(Bug.platform == platform)

    bugs = query.all()

    total = len(bugs)
    open_bugs = len([b for b in bugs if b.status in ["New", "Reopened", "Fixed", "Assigned to Dev"]])
    pending = len([b for b in bugs if b.status == "Released to QA"])
    closed = len([b for b in bugs if b.status == "Closed"])
    deferred = len([b for b in bugs if b.status == "Deferred"])
    rejected = len([b for b in bugs if b.status == "Rejected"])

    db.close()

    return {
        "ticket_id": ticket_id,
        "environment": environment,
        "total_bugs": total,
        "open_bugs": open_bugs,
        "pending_retest": pending,
        "closed_bugs": closed,
        "deferred_bugs": deferred,
        "rejected_bugs": rejected
    }


@app.get("/bugs/ticket-info")
def get_ticket_info(ticket_id: int = Query(...)):
    """Get ticket title and platform info. Prefer title from PM API (TicketTracking), else Bug subject."""
    db: Session = SessionLocal()
    try:
        tracking = db.query(TicketTracking).filter(TicketTracking.ticket_id == ticket_id).first()
        if tracking and getattr(tracking, 'title', None) and str(tracking.title).strip():
            return {
                "ticket_id": ticket_id,
                "ticket_title": str(tracking.title).strip(),
                "platform": "Web"
            }
        bug = db.query(Bug).filter(Bug.ticket_id == ticket_id).first()
        if bug and bug.subject:
            title = bug.subject.split(" - ")[0] if " - " in bug.subject else bug.subject
            return {
                "ticket_id": ticket_id,
                "ticket_title": title,
                "platform": bug.platform or "Web"
            }
        return {
            "ticket_id": ticket_id,
            "ticket_title": f"Ticket #{ticket_id}",
            "platform": "Web"
        }
    finally:
        db.close()


@app.get("/bugs/severity-breakdown")
def severity_breakdown(
    ticket_id: Optional[int] = Query(None),
    environment: str = Query("All"),
    platform: str = Query("All")
):
    """Get bug counts by status and severity for the bar chart"""
    db: Session = SessionLocal()

    query = db.query(Bug)
    
    if ticket_id is not None:
        query = query.filter(Bug.ticket_id == ticket_id)

    if environment != "All":
        query = query.filter(Bug.environment == environment)

    if platform != "All":
        query = query.filter(Bug.platform == platform)

    bugs = query.all()
    db.close()

    # Define statuses and severities
    statuses = ["New", "Assigned to Dev", "Fixed", "Released to QA", "Reopened", "Closed"]
    severities = ["Critical", "Major", "Minor", "Low Bug"]

    # Build matrix: for each status, count bugs by severity
    result = {}
    for status in statuses:
        result[status] = {}
        for severity in severities:
            count = len([b for b in bugs if b.status == status and b.severity == severity])
            result[status][severity] = count

    return {
        "statuses": statuses,
        "severities": severities,
        "data": result
    }


@app.get("/bugs/priority-breakdown")
def priority_breakdown(
    ticket_id: Optional[int] = Query(None),
    environment: str = Query("All"),
    platform: str = Query("All")
):
    """Get bug counts by priority for the pie chart"""
    db: Session = SessionLocal()

    query = db.query(Bug)
    
    if ticket_id is not None:
        query = query.filter(Bug.ticket_id == ticket_id)

    if environment != "All":
        query = query.filter(Bug.environment == environment)

    if platform != "All":
        query = query.filter(Bug.platform == platform)

    bugs = query.all()
    db.close()

    priorities = ["High", "Medium", "Low", "Low Bug"]
    result = {}
    for priority in priorities:
        result[priority] = len([b for b in bugs if b.priority == priority])

    return result


@app.get("/bugs/metrics")
def bug_metrics(
    ticket_id: Optional[int] = Query(None),
    environment: str = Query("All"),
    platform: str = Query("All")
):
    """Get closure rate and critical bugs percentage"""
    db: Session = SessionLocal()

    query = db.query(Bug)
    
    if ticket_id is not None:
        query = query.filter(Bug.ticket_id == ticket_id)

    if environment != "All":
        query = query.filter(Bug.environment == environment)

    bugs = query.all()
    db.close()

    total = len(bugs)
    closed = len([b for b in bugs if b.status == "Closed"])
    critical = len([b for b in bugs if b.severity == "Critical"])

    closure_rate = round((closed / total * 100), 1) if total > 0 else 0
    critical_percentage = round((critical / total * 100), 1) if total > 0 else 0

    return {
        "closure_rate": closure_rate,
        "critical_percentage": critical_percentage,
        "total_bugs": total,
        "closed_bugs": closed,
        "critical_bugs": critical
    }


@app.get("/bugs/all-summary")
def all_bugs_summary(environment: str = Query("All")):
    """Get summary for all bugs across all tickets"""
    db: Session = SessionLocal()

    query = db.query(Bug)

    if environment != "All":
        query = query.filter(Bug.environment == environment)

    bugs = query.all()
    db.close()

    total = len(bugs)
    open_bugs = len([b for b in bugs if b.status in ["New", "Reopened", "Fixed", "Assigned to Dev"]])
    pending = len([b for b in bugs if b.status == "Released to QA"])
    closed = len([b for b in bugs if b.status == "Closed"])
    deferred = len([b for b in bugs if b.status == "Deferred"])
    rejected = len([b for b in bugs if b.status == "Rejected"])

    return {
        "environment": environment,
        "total_bugs": total,
        "open_bugs": open_bugs,
        "pending_retest": pending,
        "closed_bugs": closed,
        "deferred_bugs": deferred,
        "rejected_bugs": rejected
    }


@app.get("/bugs/assignee-breakdown")
def assignee_breakdown(
    ticket_id: Optional[int] = Query(None),
    environment: str = Query("All"),
    platform: str = Query("All")
):
    """Get bug distribution by assignee with team classification"""
    db: Session = SessionLocal()

    try:
        # Get team classification map
        team_map = get_team_classification(db)
        
        query = db.query(Bug)

        if ticket_id is not None:
            query = query.filter(Bug.ticket_id == ticket_id)

        if environment != "All":
            query = query.filter(Bug.environment == environment)

        if platform != "All":
            query = query.filter(Bug.platform == platform)

        bugs = query.all()

        assignee_data = defaultdict(lambda: {"open": 0, "closed": 0, "total": 0, "team": "Unknown"})
        
        for bug in bugs:
            assignee = bug.assignee or "Unassigned"
            assignee_data[assignee]["total"] += 1
            assignee_data[assignee]["team"] = classify_person(assignee, team_map)
            if bug.status == "Closed":
                assignee_data[assignee]["closed"] += 1
            else:
                assignee_data[assignee]["open"] += 1

        result = {assignee: data for assignee, data in assignee_data.items()}
        return result
    finally:
        db.close()


@app.get("/bugs/author-breakdown")
def author_breakdown(
    ticket_id: Optional[int] = Query(None),
    environment: str = Query("All"),
    platform: str = Query("All")
):
    """Get bug distribution by author (who reported bugs) with team classification"""
    db: Session = SessionLocal()

    try:
        # Get team classification map
        team_map = get_team_classification(db)
        
        query = db.query(Bug)

        if ticket_id is not None:
            query = query.filter(Bug.ticket_id == ticket_id)

        if environment != "All":
            query = query.filter(Bug.environment == environment)

        if platform != "All":
            query = query.filter(Bug.platform == platform)

        bugs = query.all()

        author_data = defaultdict(lambda: {"total": 0, "by_severity": defaultdict(int), "team": "Unknown"})
        
        for bug in bugs:
            author = bug.author or "Unknown"
            author_data[author]["total"] += 1
            author_data[author]["team"] = classify_person(author, team_map)
            if bug.severity:
                author_data[author]["by_severity"][bug.severity] += 1

        result = {}
        for author, data in author_data.items():
            result[author] = {
                "total": data["total"],
                "by_severity": dict(data["by_severity"]),
                "team": data["team"]
            }
        
        return result
    finally:
        db.close()


@app.get("/bugs/team-summary")
def bug_team_summary(
    ticket_id: Optional[int] = Query(None),
    environment: str = Query("All")
):
    """Get bug summary grouped by team (DEV, QA, BIS Team)"""
    db: Session = SessionLocal()

    try:
        # Get team classification map
        team_map = get_team_classification(db)
        
        query = db.query(Bug)

        if ticket_id is not None and ticket_id != 0:
            query = query.filter(Bug.ticket_id == ticket_id)

        if environment != "All":
            query = query.filter(Bug.environment == environment)

        bugs = query.all()

        team_data = {
            "DEV": {"assignees": {}, "total_bugs": 0, "open": 0, "closed": 0},
            "QA": {"assignees": {}, "total_bugs": 0, "open": 0, "closed": 0},
            "BIS Team": {"assignees": {}, "total_bugs": 0, "open": 0, "closed": 0}
        }
        
        for bug in bugs:
            assignee = bug.assignee or "Unassigned"
            team = classify_person(assignee, team_map)
            
            if team not in team_data:
                team = "BIS Team"  # Default fallback
            
            team_data[team]["total_bugs"] += 1
            if bug.status == "Closed":
                team_data[team]["closed"] += 1
            else:
                team_data[team]["open"] += 1
            
            if assignee not in team_data[team]["assignees"]:
                team_data[team]["assignees"][assignee] = {"total": 0, "open": 0, "closed": 0}
            
            team_data[team]["assignees"][assignee]["total"] += 1
            if bug.status == "Closed":
                team_data[team]["assignees"][assignee]["closed"] += 1
            else:
                team_data[team]["assignees"][assignee]["open"] += 1

        return team_data
    finally:
        db.close()


@app.get("/bugs/module-breakdown")
def module_breakdown(
    ticket_id: Optional[int] = Query(None),
    environment: str = Query("All")
):
    """Get bug distribution by module"""
    db: Session = SessionLocal()

    query = db.query(Bug)

    if ticket_id is not None:
        query = query.filter(Bug.ticket_id == ticket_id)

    if environment != "All":
        query = query.filter(Bug.environment == environment)

    bugs = query.all()
    db.close()

    module_data = defaultdict(int)
    
    for bug in bugs:
        module = bug.module or "Unknown"
        module_data[module] += 1

    return dict(module_data)


@app.get("/bugs/feature-breakdown")
def feature_breakdown(
    ticket_id: Optional[int] = Query(None),
    environment: str = Query("All")
):
    """Get bug distribution by feature"""
    db: Session = SessionLocal()

    query = db.query(Bug)

    if ticket_id is not None:
        query = query.filter(Bug.ticket_id == ticket_id)

    if environment != "All":
        query = query.filter(Bug.environment == environment)

    bugs = query.all()
    db.close()

    feature_data = defaultdict(lambda: {"open": 0, "closed": 0, "total": 0})
    
    for bug in bugs:
        feature = bug.feature or "Unknown"
        feature_data[feature]["total"] += 1
        if bug.status == "Closed":
            feature_data[feature]["closed"] += 1
        else:
            feature_data[feature]["open"] += 1

    result = {feature: data for feature, data in feature_data.items()}
    return result


@app.get("/bugs/browser-os-breakdown")
def browser_os_breakdown(
    ticket_id: Optional[int] = Query(None),
    environment: str = Query("All")
):
    """Get bug distribution by browser and OS combinations"""
    db: Session = SessionLocal()

    query = db.query(Bug)

    if ticket_id is not None:
        query = query.filter(Bug.ticket_id == ticket_id)

    if environment != "All":
        query = query.filter(Bug.environment == environment)

    bugs = query.all()
    db.close()

    browser_os_data = defaultdict(int)
    
    for bug in bugs:
        browser = bug.browser or "Unknown"
        os = bug.os or "Unknown"
        key = f"{browser} / {os}"
        browser_os_data[key] += 1

    return dict(browser_os_data)


@app.get("/bugs/platform-breakdown")
def platform_breakdown(
    ticket_id: Optional[int] = Query(None),
    environment: str = Query("All")
):
    """Get bug distribution by platform"""
    db: Session = SessionLocal()

    query = db.query(Bug)

    if ticket_id is not None:
        query = query.filter(Bug.ticket_id == ticket_id)

    if environment != "All":
        query = query.filter(Bug.environment == environment)

    bugs = query.all()
    db.close()

    platform_data = defaultdict(lambda: {"open": 0, "closed": 0, "total": 0, "by_status": defaultdict(int)})
    
    for bug in bugs:
        platform = bug.platform or "Unknown"
        platform_data[platform]["total"] += 1
        platform_data[platform]["by_status"][bug.status or "Unknown"] += 1
        if bug.status == "Closed":
            platform_data[platform]["closed"] += 1
        else:
            platform_data[platform]["open"] += 1

    result = {}
    for platform, data in platform_data.items():
        result[platform] = {
            "total": data["total"],
            "open": data["open"],
            "closed": data["closed"],
            "by_status": dict(data["by_status"])
        }
    
    return result


@app.get("/bugs/age-analysis")
def age_analysis(
    ticket_id: Optional[int] = Query(None),
    environment: str = Query("All")
):
    """Get bug age metrics"""
    db: Session = SessionLocal()

    query = db.query(Bug)

    if ticket_id is not None:
        query = query.filter(Bug.ticket_id == ticket_id)

    if environment != "All":
        query = query.filter(Bug.environment == environment)

    bugs = query.all()
    db.close()

    now = datetime.now()
    open_bugs = [b for b in bugs if b.status not in ["Closed", "Deferred"]]
    
    ages = []
    age_buckets = {"0-7": 0, "7-30": 0, "30-60": 0, "60+": 0}
    oldest_age = 0
    
    for bug in open_bugs:
        if bug.created_on:
            age_days = (now - bug.created_on.replace(tzinfo=None) if bug.created_on.tzinfo else bug.created_on).days
            ages.append(age_days)
            oldest_age = max(oldest_age, age_days)
            
            if age_days <= 7:
                age_buckets["0-7"] += 1
            elif age_days <= 30:
                age_buckets["7-30"] += 1
            elif age_days <= 60:
                age_buckets["30-60"] += 1
            else:
                age_buckets["60+"] += 1

    avg_age = sum(ages) / len(ages) if ages else 0

    return {
        "average_age_days": round(avg_age, 1),
        "oldest_age_days": oldest_age,
        "total_open_bugs": len(open_bugs),
        "age_buckets": age_buckets
    }


@app.get("/bugs/resolution-time")
def resolution_time(
    ticket_id: Optional[int] = Query(None),
    environment: str = Query("All")
):
    """Get resolution time metrics"""
    db: Session = SessionLocal()

    query = db.query(Bug)

    if ticket_id is not None:
        query = query.filter(Bug.ticket_id == ticket_id)

    if environment != "All":
        query = query.filter(Bug.environment == environment)

    bugs = query.all()
    db.close()

    closed_bugs = [b for b in bugs if b.status == "Closed" and b.created_on and b.closed_on]
    
    resolution_times = []
    time_buckets = {"<1": 0, "1-3": 0, "3-7": 0, "7-30": 0, "30+": 0}
    
    for bug in closed_bugs:
        created = bug.created_on.replace(tzinfo=None) if bug.created_on.tzinfo else bug.created_on
        closed = bug.closed_on.replace(tzinfo=None) if bug.closed_on.tzinfo else bug.closed_on
        days = (closed - created).days
        resolution_times.append(days)
        
        if days < 1:
            time_buckets["<1"] += 1
        elif days <= 3:
            time_buckets["1-3"] += 1
        elif days <= 7:
            time_buckets["3-7"] += 1
        elif days <= 30:
            time_buckets["7-30"] += 1
        else:
            time_buckets["30+"] += 1

    if resolution_times:
        sorted_times = sorted(resolution_times)
        avg_time = sum(resolution_times) / len(resolution_times)
        median_time = sorted_times[len(sorted_times) // 2]
        fastest = min(resolution_times)
        slowest = max(resolution_times)
    else:
        avg_time = median_time = fastest = slowest = 0

    return {
        "average_days": round(avg_time, 1),
        "median_days": median_time,
        "fastest_days": fastest,
        "slowest_days": slowest,
        "total_resolved": len(closed_bugs),
        "time_buckets": time_buckets
    }


@app.get("/bugs/reopened-analysis")
def reopened_analysis(
    ticket_id: Optional[int] = Query(None),
    environment: str = Query("All")
):
    """Get reopened bugs analysis"""
    db: Session = SessionLocal()

    query = db.query(Bug)

    if ticket_id is not None:
        query = query.filter(Bug.ticket_id == ticket_id)

    if environment != "All":
        query = query.filter(Bug.environment == environment)

    bugs = query.all()
    db.close()

    reopened_bugs = [b for b in bugs if b.status == "Reopened"]
    total_bugs = len(bugs)
    
    reopened_by_severity = defaultdict(int)
    reopened_by_priority = defaultdict(int)
    
    for bug in reopened_bugs:
        if bug.severity:
            reopened_by_severity[bug.severity] += 1
        if bug.priority:
            reopened_by_priority[bug.priority] += 1

    reopened_percentage = round((len(reopened_bugs) / total_bugs * 100), 1) if total_bugs > 0 else 0

    return {
        "total_reopened": len(reopened_bugs),
        "reopened_percentage": reopened_percentage,
        "total_bugs": total_bugs,
        "by_severity": dict(reopened_by_severity),
        "by_priority": dict(reopened_by_priority)
    }


@app.get("/bugs/deferred-bugs")
def deferred_bugs(
    ticket_id: Optional[int] = Query(None),
    environment: str = Query("All")
):
    """Get deferred bugs with ageing information"""
    db: Session = SessionLocal()

    query = db.query(Bug)

    if ticket_id is not None:
        query = query.filter(Bug.ticket_id == ticket_id)

    if environment != "All":
        query = query.filter(Bug.environment == environment)

    query = query.filter(Bug.status == "Deferred")
    bugs = query.all()
    db.close()

    now = datetime.now()
    deferred_list = []
    
    for bug in bugs:
        age_days = 0
        if bug.created_on:
            created = bug.created_on.replace(tzinfo=None) if bug.created_on.tzinfo else bug.created_on
            age_days = (now - created).days
        
        deferred_list.append({
            "bug_id": bug.bug_id,
            "subject": bug.subject,
            "severity": bug.severity,
            "priority": bug.priority,
            "assignee": bug.assignee,
            "age_days": age_days,
            "created_on": bug.created_on.isoformat() if bug.created_on else None
        })
    
    # Sort by age (oldest first)
    deferred_list.sort(key=lambda x: x["age_days"], reverse=True)
    
    return deferred_list


@app.get("/bugs/time-tracking")
def bug_time_tracking(
    ticket_id: Optional[int] = Query(None),
    environment: str = Query("All")
):
    """Get estimate vs actual time comparison with variance analysis"""
    db: Session = SessionLocal()
    
    try:
        query = db.query(Bug)
        
        if ticket_id is not None and ticket_id != 0:
            query = query.filter(Bug.ticket_id == ticket_id)
        
        if environment != "All":
            query = query.filter(Bug.environment == environment)
        
        bugs = query.all()
        
        total_estimated = 0
        total_spent = 0
        estimated_count = 0
        not_estimated_count = 0
        bugs_with_variance = []
        
        for bug in bugs:
            estimated = bug.estimated_hours or 0
            spent = bug.spent_hours or 0
            
            if estimated > 0:
                estimated_count += 1
                total_estimated += estimated
                total_spent += spent
                
                variance_percent = ((spent - estimated) / estimated) * 100 if estimated > 0 else 0
                bugs_with_variance.append({
                    "bug_id": bug.bug_id,
                    "subject": bug.subject[:50] + "..." if len(bug.subject or "") > 50 else bug.subject,
                    "estimated_hours": estimated,
                    "spent_hours": spent,
                    "variance_percent": round(variance_percent, 1),
                    "variance_status": "green" if abs(variance_percent) < 10 else ("amber" if abs(variance_percent) < 30 else "red")
                })
            else:
                not_estimated_count += 1
        
        overall_variance = ((total_spent - total_estimated) / total_estimated * 100) if total_estimated > 0 else 0
        
        # Group by variance status
        variance_distribution = {
            "under_estimate": len([b for b in bugs_with_variance if b["variance_percent"] < -10]),
            "on_track": len([b for b in bugs_with_variance if -10 <= b["variance_percent"] <= 10]),
            "over_estimate": len([b for b in bugs_with_variance if b["variance_percent"] > 10])
        }
        
        return {
            "total_bugs": len(bugs),
            "estimated_count": estimated_count,
            "not_estimated_count": not_estimated_count,
            "not_estimated_percent": round((not_estimated_count / len(bugs) * 100) if bugs else 0, 1),
            "total_estimated_hours": round(total_estimated, 1),
            "total_spent_hours": round(total_spent, 1),
            "overall_variance_percent": round(overall_variance, 1),
            "variance_distribution": variance_distribution,
            "top_variances": sorted(bugs_with_variance, key=lambda x: abs(x["variance_percent"]), reverse=True)[:10]
        }
    finally:
        db.close()


@app.get("/bugs/sla-analysis")
def bug_sla_analysis(
    ticket_id: Optional[int] = Query(None),
    environment: str = Query("All")
):
    """Get due date/SLA tracking - overdue, on-time, no due date"""
    db: Session = SessionLocal()
    
    try:
        query = db.query(Bug)
        
        if ticket_id is not None and ticket_id != 0:
            query = query.filter(Bug.ticket_id == ticket_id)
        
        if environment != "All":
            query = query.filter(Bug.environment == environment)
        
        bugs = query.all()
        now = datetime.now()
        
        overdue = []
        on_time = []
        no_due_date = []
        completed_on_time = 0
        completed_late = 0
        
        for bug in bugs:
            if bug.due_date is None:
                no_due_date.append(bug.bug_id)
            else:
                due = bug.due_date.replace(tzinfo=None) if bug.due_date.tzinfo else bug.due_date
                
                if bug.status == "Closed" and bug.closed_on:
                    closed = bug.closed_on.replace(tzinfo=None) if bug.closed_on.tzinfo else bug.closed_on
                    if closed <= due:
                        completed_on_time += 1
                        on_time.append(bug.bug_id)
                    else:
                        completed_late += 1
                        overdue.append({
                            "bug_id": bug.bug_id,
                            "subject": bug.subject[:50] + "..." if len(bug.subject or "") > 50 else bug.subject,
                            "due_date": due.isoformat(),
                            "days_overdue": (closed - due).days,
                            "status": bug.status,
                            "severity": bug.severity
                        })
                elif bug.status != "Closed":
                    if now > due:
                        days_overdue = (now - due).days
                        overdue.append({
                            "bug_id": bug.bug_id,
                            "subject": bug.subject[:50] + "..." if len(bug.subject or "") > 50 else bug.subject,
                            "due_date": due.isoformat(),
                            "days_overdue": days_overdue,
                            "status": bug.status,
                            "severity": bug.severity
                        })
                    else:
                        on_time.append(bug.bug_id)
        
        # Sort overdue by days overdue (most overdue first)
        overdue_list = sorted(overdue, key=lambda x: x["days_overdue"], reverse=True)
        
        return {
            "total_bugs": len(bugs),
            "overdue_count": len(overdue),
            "on_time_count": len(on_time),
            "no_due_date_count": len(no_due_date),
            "completed_on_time": completed_on_time,
            "completed_late": completed_late,
            "sla_compliance_rate": round((len(on_time) / (len(on_time) + len(overdue)) * 100) if (len(on_time) + len(overdue)) > 0 else 0, 1),
            "overdue_bugs": overdue_list[:20],  # Top 20 overdue
            "distribution": {
                "overdue": len(overdue),
                "on_time": len(on_time),
                "no_due_date": len(no_due_date)
            }
        }
    finally:
        db.close()


@app.get("/bugs/lifecycle-analysis")
def bug_lifecycle_analysis(
    ticket_id: Optional[int] = Query(None),
    environment: str = Query("All")
):
    """Get bug lifecycle metrics - start to close timeline"""
    db: Session = SessionLocal()
    
    try:
        query = db.query(Bug)
        
        if ticket_id is not None and ticket_id != 0:
            query = query.filter(Bug.ticket_id == ticket_id)
        
        if environment != "All":
            query = query.filter(Bug.environment == environment)
        
        bugs = query.all()
        
        lifecycle_days = []
        creation_to_close = []
        
        for bug in bugs:
            # Calculate lifecycle from start_date to closed_on
            if bug.start_date and bug.closed_on:
                start = bug.start_date.replace(tzinfo=None) if bug.start_date.tzinfo else bug.start_date
                closed = bug.closed_on.replace(tzinfo=None) if bug.closed_on.tzinfo else bug.closed_on
                days = (closed - start).days
                if days >= 0:
                    lifecycle_days.append(days)
            
            # Also calculate from created_on to closed_on
            if bug.created_on and bug.closed_on:
                created = bug.created_on.replace(tzinfo=None) if bug.created_on.tzinfo else bug.created_on
                closed = bug.closed_on.replace(tzinfo=None) if bug.closed_on.tzinfo else bug.closed_on
                days = (closed - created).days
                if days >= 0:
                    creation_to_close.append(days)
        
        # Calculate distribution buckets
        def get_distribution(days_list):
            return {
                "0-1": len([d for d in days_list if d <= 1]),
                "2-3": len([d for d in days_list if 2 <= d <= 3]),
                "4-7": len([d for d in days_list if 4 <= d <= 7]),
                "8-14": len([d for d in days_list if 8 <= d <= 14]),
                "15-30": len([d for d in days_list if 15 <= d <= 30]),
                "30+": len([d for d in days_list if d > 30])
            }
        
        avg_lifecycle = sum(lifecycle_days) / len(lifecycle_days) if lifecycle_days else 0
        avg_creation_to_close = sum(creation_to_close) / len(creation_to_close) if creation_to_close else 0
        
        return {
            "total_closed_bugs": len(creation_to_close),
            "avg_lifecycle_days": round(avg_lifecycle, 1),
            "avg_creation_to_close_days": round(avg_creation_to_close, 1),
            "min_lifecycle_days": min(lifecycle_days) if lifecycle_days else 0,
            "max_lifecycle_days": max(lifecycle_days) if lifecycle_days else 0,
            "median_lifecycle_days": sorted(lifecycle_days)[len(lifecycle_days)//2] if lifecycle_days else 0,
            "lifecycle_distribution": get_distribution(lifecycle_days),
            "creation_close_distribution": get_distribution(creation_to_close)
        }
    finally:
        db.close()


@app.get("/bugs/completion-progress")
def bug_completion_progress(
    ticket_id: Optional[int] = Query(None),
    environment: str = Query("All")
):
    """Get done_ratio/completion progress distribution"""
    db: Session = SessionLocal()
    
    try:
        query = db.query(Bug)
        
        if ticket_id is not None and ticket_id != 0:
            query = query.filter(Bug.ticket_id == ticket_id)
        
        if environment != "All":
            query = query.filter(Bug.environment == environment)
        
        # Only get open bugs (not closed)
        query = query.filter(Bug.status != "Closed")
        
        bugs = query.all()
        
        completion_buckets = {
            "0%": 0,
            "1-25%": 0,
            "26-50%": 0,
            "51-75%": 0,
            "76-99%": 0,
            "100%": 0
        }
        
        total_done_ratio = 0
        bugs_with_progress = 0
        
        for bug in bugs:
            done = bug.done_ratio or 0
            total_done_ratio += done
            
            if done > 0:
                bugs_with_progress += 1
            
            if done == 0:
                completion_buckets["0%"] += 1
            elif done <= 25:
                completion_buckets["1-25%"] += 1
            elif done <= 50:
                completion_buckets["26-50%"] += 1
            elif done <= 75:
                completion_buckets["51-75%"] += 1
            elif done < 100:
                completion_buckets["76-99%"] += 1
            else:
                completion_buckets["100%"] += 1
        
        avg_completion = total_done_ratio / len(bugs) if bugs else 0
        
        return {
            "total_open_bugs": len(bugs),
            "bugs_with_progress": bugs_with_progress,
            "bugs_not_started": completion_buckets["0%"],
            "avg_completion_percent": round(avg_completion, 1),
            "completion_distribution": completion_buckets,
            "near_completion": completion_buckets["76-99%"] + completion_buckets["100%"]
        }
    finally:
        db.close()


# ===== TESTRAIL ENDPOINTS =====

@app.get("/testrail/summary")
def testrail_summary(ticket_id: int = Query(...)):
    """Get test case counts and status breakdown for a ticket"""
    db: Session = SessionLocal()
    
    try:
        # Get all test results for this ticket
        results = db.query(TestResult).filter(TestResult.ticket_id == ticket_id).all()
        
        total_tests = len(results)
        status_counts = {
            "Passed": 0,
            "Failed": 0,
            "Blocked": 0,
            "Retest": 0,
            "Untested": 0
        }
        
        for result in results:
            status = result.status_name or "Untested"
            if status in status_counts:
                status_counts[status] += 1
        
        # Get unique test cases count
        unique_cases = db.query(TestCase.case_id).filter(TestCase.ticket_id == ticket_id).distinct().count()
        plans_count = db.query(TestPlan).filter(TestPlan.ticket_id == ticket_id).count()
        runs_count = db.query(TestRun).filter(TestRun.ticket_id == ticket_id).count()
        
        # Get test plan name (most recent plan)
        test_plan = db.query(TestPlan).filter(TestPlan.ticket_id == ticket_id).order_by(TestPlan.created_on.desc()).first()
        plan_name = None
        if test_plan and test_plan.name:
            # Remove ticket_id_ prefix from plan name
            import re
            plan_name = re.sub(r'^\d+_', '', test_plan.name)
        
        return {
            "ticket_id": ticket_id,
            "total_test_cases": unique_cases,
            "total_test_results": total_tests,
            "status_counts": status_counts,
            "test_plans_count": plans_count,
            "test_runs_count": runs_count,
            "test_plan_name": plan_name
        }
    finally:
        db.close()


@app.get("/testrail/test-plans")
def testrail_test_plans(ticket_id: int = Query(...)):
    """Get all test plans for a ticket"""
    db: Session = SessionLocal()
    try:
        plans = db.query(TestPlan).filter(TestPlan.ticket_id == ticket_id).all()
        return [
            {
                "plan_id": plan.plan_id,
                "name": plan.name,
                "description": plan.description,
                "created_on": plan.created_on.isoformat() if plan.created_on else None,
                "updated_on": plan.updated_on.isoformat() if plan.updated_on else None,
                "custom_fields": plan.custom_fields
            }
            for plan in plans
        ]
    finally:
        db.close()


@app.get("/testrail/test-runs")
def testrail_test_runs(ticket_id: int = Query(...)):
    """Get all test runs for a ticket with their test results"""
    db: Session = SessionLocal()
    try:
        runs = db.query(TestRun).filter(TestRun.ticket_id == ticket_id).order_by(TestRun.created_on.desc()).all()
        result = []
        
        for run in runs:
            # Get all test results for this run
            results = db.query(TestResult).filter(TestResult.run_id == run.run_id).all()
            
            # Count statuses for this run
            status_counts = {
                "Passed": 0,
                "Failed": 0,
                "Blocked": 0,
                "Retest": 0,
                "Untested": 0
            }
            
            for res in results:
                status = res.status_name or "Untested"
                if status in status_counts:
                    status_counts[status] += 1
            
            # Get unique test cases in this run
            unique_cases = db.query(TestResult.case_id).filter(
                TestResult.run_id == run.run_id
            ).distinct().count()
            
            result.append({
                "run_id": run.run_id,
                "plan_id": run.plan_id,
                "name": run.name,
                "description": run.description,
                "status": run.status,
                "created_on": run.created_on.isoformat() if run.created_on else None,
                "updated_on": run.updated_on.isoformat() if run.updated_on else None,
                "total_tests": len(results),
                "unique_test_cases": unique_cases,
                "status_counts": status_counts,
                "custom_fields": run.custom_fields
            })
        
        return result
    finally:
        db.close()


@app.get("/testrail/test-cases")
def testrail_test_cases(ticket_id: int = Query(...)):
    """Get all test cases with results for a ticket"""
    db: Session = SessionLocal()
    try:
        # Get all test cases for this ticket
        cases = db.query(TestCase).filter(TestCase.ticket_id == ticket_id).all()
        
        # Get latest results for each case
        case_results = {}
        results = db.query(TestResult).filter(TestResult.ticket_id == ticket_id).all()
        
        for result in results:
            case_id = result.case_id
            if case_id not in case_results or (result.created_on and (
                not case_results[case_id].created_on or 
                result.created_on > case_results[case_id].created_on
            )):
                case_results[case_id] = result
        
        return [
            {
                "case_id": case.case_id,
                "run_id": case.run_id,
                "title": case.title,
                "section": case.section,
                "priority": case.priority,
                "type": case.type,
                "latest_status": case_results.get(case.case_id).status_name if case.case_id in case_results and case_results.get(case.case_id) else "Untested",
                "latest_result_id": case_results.get(case.case_id).test_id if case.case_id in case_results and case_results.get(case.case_id) else None,
                "custom_fields": case.custom_fields
            }
            for case in cases
        ]
    finally:
        db.close()


@app.get("/testrail/status-breakdown")
def testrail_status_breakdown(ticket_id: int = Query(...)):
    """Get test status distribution for a ticket"""
    db: Session = SessionLocal()
    try:
        results = db.query(TestResult).filter(TestResult.ticket_id == ticket_id).all()
        
        status_counts = defaultdict(int)
        for result in results:
            status = result.status_name or "Untested"
            status_counts[status] += 1
        
        total = len(results)
        
        return {
            "ticket_id": ticket_id,
            "total": total,
            "status_distribution": dict(status_counts),
            "percentages": {
                status: round((count / total * 100), 1) if total > 0 else 0
                for status, count in status_counts.items()
            }
        }
    finally:
        db.close()


# ===== AUTOMATION COVERAGE ENDPOINTS (TestRail Project 18) =====

def _dedupe_automation_cases(cases: List[AutomationTestCase]) -> List[AutomationTestCase]:
    """Return one record per logical test case for overall calculations.

    We prefer the latest synced row (higher DB id) when the same case appears
    in multiple runs.
    """
    deduped = {}
    for case in cases:
        key = case.case_id if case.case_id is not None else "test_{}".format(case.test_id)
        existing = deduped.get(key)
        if existing is None or (case.id or 0) > (existing.id or 0):
            deduped[key] = case
    return list(deduped.values())


@app.get("/automation/summary")
def automation_summary(
    ticket_id: Optional[int] = Query(None),
    run_id: Optional[str] = Query(None, description="Specific run ID or 'all' for combined view")
):
    """Get automation vs manual test execution summary for a ticket.
    Returns counts and percentages of automated vs manual test cases with pass/fail breakdown.
    Optionally filter by specific test run.
    """
    db: Session = SessionLocal()
    try:
        query = db.query(AutomationTestCase)
        if ticket_id is not None:
            query = query.filter(AutomationTestCase.ticket_id == ticket_id)
        
        # When specific run_id is provided, don't dedupe - show exact run data
        # When 'all' or no run_id, dedupe to get unique case counts
        if run_id and run_id != 'all':
            try:
                run_id_int = int(run_id)
                query = query.filter(AutomationTestCase.run_id == run_id_int)
                cases = query.all()  # No deduplication for specific run
            except ValueError:
                cases = _dedupe_automation_cases(query.all())
        else:
            cases = _dedupe_automation_cases(query.all())
        
        if not cases:
            return {
                "ticket_id": ticket_id,
                "total_cases": 0,
                "automated": {"count": 0, "percentage": 0, "passed": 0, "failed": 0, "blocked": 0, "retest": 0, "untested": 0},
                "manual": {"count": 0, "percentage": 0, "passed": 0, "failed": 0, "blocked": 0, "retest": 0, "untested": 0},
                "automation_coverage": 0,
                "planned_count": 0,
                "candidates_yes": 0,
                "candidates_no": 0,
                "candidates_none": 0,
            }
        
        total = len(cases)
        planned_count = 0
        candidates_yes = 0
        candidates_no = 0
        candidates_none = 0
        
        automated = {"count": 0, "passed": 0, "failed": 0, "blocked": 0, "retest": 0, "untested": 0}
        manual = {"count": 0, "passed": 0, "failed": 0, "blocked": 0, "retest": 0, "untested": 0}
        
        for case in cases:
            is_automated = (
                (case.automation_status and case.automation_status.lower() == "automated") or
                (case.execution_method and case.execution_method.lower() == "automated")
            )
            
            is_planned = (
                case.automation_status and case.automation_status.lower() == "planned"
            )
            
            target = automated if is_automated else manual
            target["count"] += 1
            
            status = (case.status_name or "Untested").lower()
            if status == "passed":
                target["passed"] += 1
            elif status == "failed":
                target["failed"] += 1
            elif status == "blocked":
                target["blocked"] += 1
            elif status == "retest":
                target["retest"] += 1
            else:
                target["untested"] += 1
            
            if is_planned:
                planned_count += 1
            
            # Count automation candidates
            candidate = (case.automation_candidate or "").strip()
            if candidate.lower() == "yes":
                candidates_yes += 1
            elif candidate.lower() == "no":
                candidates_no += 1
            else:
                candidates_none += 1
        
        automated["percentage"] = round((automated["count"] / total * 100), 1) if total > 0 else 0
        manual["percentage"] = round((manual["count"] / total * 100), 1) if total > 0 else 0
        
        return {
            "ticket_id": ticket_id,
            "total_cases": total,
            "automated": automated,
            "manual": manual,
            "automation_coverage": automated["percentage"],
            "planned_count": planned_count,
            "candidates_yes": candidates_yes,
            "candidates_no": candidates_no,
            "candidates_none": candidates_none,
        }
    finally:
        db.close()


@app.get("/automation/test-cases")
def automation_test_cases(
    ticket_id: Optional[int] = Query(None),
    run_id: Optional[str] = Query(None, description="Specific run ID or 'all' for combined view")
):
    """Get all test cases with automation details for a ticket"""
    db: Session = SessionLocal()
    try:
        query = db.query(AutomationTestCase)
        if ticket_id is not None:
            query = query.filter(AutomationTestCase.ticket_id == ticket_id)
        
        # When specific run_id is provided, don't dedupe - show all entries for that run
        if run_id and run_id != 'all':
            try:
                run_id_int = int(run_id)
                query = query.filter(AutomationTestCase.run_id == run_id_int)
                cases = query.all()  # No deduplication for specific run
            except ValueError:
                cases = _dedupe_automation_cases(query.all())
        else:
            cases = _dedupe_automation_cases(query.all())
        
        return [
            {
                "test_id": case.test_id,
                "case_id": case.case_id,
                "run_id": case.run_id,
                "title": case.title,
                "section": case.section,
                "priority": case.priority,
                "automation_status": case.automation_status,
                "execution_method": case.execution_method,
                "reusability_frequency": case.reusability_frequency,
                "automation_maintenance": case.automation_maintenance,
                "status_name": case.status_name,
                "status_id": case.status_id,
                "business_criticality": case.business_criticality,
                "functionality": case.functionality,
                "sub_functionality": case.sub_functionality,
                "life_cycle_status": case.life_cycle_status,
                "is_automated": (
                    (case.automation_status and case.automation_status.lower() == "automated") or
                    (case.execution_method and case.execution_method.lower() == "automated")
                )
            }
            for case in cases
        ]
    finally:
        db.close()


@app.get("/automation/effort")
def automation_effort(
    ticket_id: Optional[int] = Query(None),
    run_id: Optional[str] = Query(None, description="Specific run ID or 'all' for combined view")
):
    """Get automation effort hours and timeline for a ticket"""
    db: Session = SessionLocal()
    try:
        query = db.query(AutomationTestCase)
        if ticket_id is not None:
            query = query.filter(AutomationTestCase.ticket_id == ticket_id)
        if run_id and run_id != 'all':
            try:
                run_id_int = int(run_id)
                query = query.filter(AutomationTestCase.run_id == run_id_int)
                cases = query.all()  # No deduplication for specific run
            except ValueError:
                cases = _dedupe_automation_cases(query.all())
        else:
            cases = _dedupe_automation_cases(query.all())
        
        total_estimated = 0.0
        total_actual = 0.0
        cases_with_effort = []
        
        for case in cases:
            estimated = case.automation_estimated_hours or 0
            actual = case.automation_actual_hours or 0
            
            if estimated > 0 or actual > 0:
                total_estimated += estimated
                total_actual += actual
                cases_with_effort.append({
                    "case_id": case.case_id,
                    "run_id": case.run_id,
                    "title": case.title,
                    "automation_status": case.automation_status,
                    "estimated_hours": estimated,
                    "actual_hours": actual,
                    "variance": round(actual - estimated, 2) if estimated > 0 else None,
                    "planned_start": case.automation_planned_start.isoformat() if case.automation_planned_start else None,
                    "actual_start": case.automation_actual_start.isoformat() if case.automation_actual_start else None,
                    "actual_end": case.automation_actual_end.isoformat() if case.automation_actual_end else None
                })
        
        return {
            "ticket_id": ticket_id,
            "total_estimated_hours": round(total_estimated, 2),
            "total_actual_hours": round(total_actual, 2),
            "total_variance": round(total_actual - total_estimated, 2) if total_estimated > 0 else None,
            "efficiency": round((total_estimated / total_actual * 100), 1) if total_actual > 0 else None,
            "cases_count": len(cases_with_effort),
            "cases": cases_with_effort
        }
    finally:
        db.close()


@app.get("/automation/test-runs")
def automation_test_runs(ticket_id: Optional[int] = Query(None)):
    """Get all test runs for a ticket with automation breakdown"""
    db: Session = SessionLocal()
    try:
        query = db.query(AutomationTestRun)
        if ticket_id is not None:
            query = query.filter(AutomationTestRun.ticket_id == ticket_id)
        # Show runs in creation order (oldest -> newest)
        runs = query.order_by(AutomationTestRun.created_on.asc(), AutomationTestRun.run_id.asc()).all()
        
        result = []
        for run in runs:
            cases = db.query(AutomationTestCase).filter(
                AutomationTestCase.run_id == run.run_id
            ).all()
            
            automated_count = 0
            manual_count = 0
            passed = 0
            failed = 0
            
            for case in cases:
                is_automated = (
                    (case.automation_status and case.automation_status.lower() == "automated") or
                    (case.execution_method and case.execution_method.lower() == "automated")
                )
                if is_automated:
                    automated_count += 1
                else:
                    manual_count += 1
                
                if case.status_name == "Passed":
                    passed += 1
                elif case.status_name == "Failed":
                    failed += 1
            
            total = len(cases)
            result.append({
                "run_id": run.run_id,
                "plan_id": run.plan_id,
                "name": run.name,
                "status": run.status,
                "created_on": run.created_on.isoformat() if run.created_on else None,
                "total_cases": total,
                "automated_count": automated_count,
                "manual_count": manual_count,
                "automation_percentage": round((automated_count / total * 100), 1) if total > 0 else 0,
                "passed": passed,
                "failed": failed,
                "pass_rate": round((passed / total * 100), 1) if total > 0 else 0
            })
        
        return result
    finally:
        db.close()


@app.get("/automation/dashboard-metrics")
def automation_dashboard_metrics(ticket_id: Optional[int] = Query(None)):
    """Get automation coverage metrics, optionally filtered by ticket."""
    db: Session = SessionLocal()
    try:
        cases_query = db.query(AutomationTestCase)
        runs_query = db.query(AutomationTestRun)
        if ticket_id is not None:
            cases_query = cases_query.filter(AutomationTestCase.ticket_id == ticket_id)
            runs_query = runs_query.filter(AutomationTestRun.ticket_id == ticket_id)

        all_cases = cases_query.all()
        all_cases = _dedupe_automation_cases(all_cases)
        all_runs = runs_query.all()
        
        if not all_cases:
            return {
                "total_cases": 0,
                "total_automated": 0,
                "total_manual": 0,
                "total_planned": 0,
                "candidates_yes": 0,
                "candidates_no": 0,
                "candidates_none": 0,
                "overall_automation_percentage": 0,
                "tickets_with_automation": 0,
                "total_runs": 0,
                "runs_with_automated_cases": 0,
                "runs_with_manual_cases": 0,
                "reusability_breakdown": {},
                "status_breakdown": {"automated": {}, "manual": {}},
                "ticket_id": ticket_id,
            }
        
        total = len(all_cases)
        automated_count = 0
        manual_count = 0
        planned_count = 0
        candidates_yes = 0
        candidates_no = 0
        candidates_none = 0
        tickets = set()
        
        reusability = defaultdict(int)
        automated_status = defaultdict(int)
        manual_status = defaultdict(int)
        
        for case in all_cases:
            if case.ticket_id is not None:
                tickets.add(case.ticket_id)
            
            is_automated = (
                (case.automation_status and case.automation_status.lower() == "automated") or
                (case.execution_method and case.execution_method.lower() == "automated")
            )
            
            is_planned = (
                case.automation_status and case.automation_status.lower() == "planned"
            )
            
            if is_automated:
                automated_count += 1
                automated_status[case.status_name or "Untested"] += 1
            else:
                manual_count += 1
                manual_status[case.status_name or "Untested"] += 1
            
            if is_planned:
                planned_count += 1
            
            # Count automation candidates
            candidate = (case.automation_candidate or "").strip()
            if candidate.lower() == "yes":
                candidates_yes += 1
            elif candidate.lower() == "no":
                candidates_no += 1
            else:
                candidates_none += 1
            
            if case.reusability_frequency:
                reusability[case.reusability_frequency] += 1

        runs_with_automated_cases = 0
        runs_with_manual_cases = 0
        for run in all_runs:
            run_cases = db.query(AutomationTestCase).filter(
                AutomationTestCase.run_id == run.run_id
            ).all()
            if not run_cases:
                continue

            automated_in_run = 0
            manual_in_run = 0
            for case in run_cases:
                is_automated = (
                    (case.automation_status and case.automation_status.lower() == "automated") or
                    (case.execution_method and case.execution_method.lower() == "automated")
                )
                if is_automated:
                    automated_in_run += 1
                else:
                    manual_in_run += 1

            if automated_in_run > 0:
                runs_with_automated_cases += 1
            if manual_in_run > 0:
                runs_with_manual_cases += 1
        
        # Automation percentage based on candidates (Yes) - the target cases to automate
        remaining_to_automate = max(0, candidates_yes - automated_count)
        automation_percentage = round((automated_count / candidates_yes * 100), 1) if candidates_yes > 0 else 0
        
        return {
            "total_cases": total,
            "total_automated": automated_count,
            "total_manual": manual_count,
            "total_planned": planned_count,
            "candidates_yes": candidates_yes,
            "candidates_no": candidates_no,
            "candidates_none": candidates_none,
            "remaining_to_automate": remaining_to_automate,
            "overall_automation_percentage": automation_percentage,
            "tickets_with_automation": len(tickets),
            "total_runs": len(all_runs),
            "runs_with_automated_cases": runs_with_automated_cases,
            "runs_with_manual_cases": runs_with_manual_cases,
            "reusability_breakdown": dict(reusability),
            "status_breakdown": {
                "automated": dict(automated_status),
                "manual": dict(manual_status)
            },
            "ticket_id": ticket_id,
        }
    finally:
        db.close()


@app.get("/automation/overall-functionality")
def automation_overall_functionality():
    """Get overall automation status across all core cases, grouped by functionality."""
    db: Session = SessionLocal()
    try:
        all_cases = db.query(AutomationTestCase).all()
        all_cases = _dedupe_automation_cases(all_cases)
        if not all_cases:
            return {
                "overall": {
                    "total_core_cases": 0,
                    "automated_cases": 0,
                    "manual_cases": 0,
                    "planned_cases": 0,
                    "candidates_yes": 0,
                    "candidates_no": 0,
                    "candidates_none": 0,
                    "automation_percentage": 0,
                    "status_breakdown": {}
                },
                "by_functionality": [],
                "by_section": []
            }

        functionality = defaultdict(
            lambda: {
                "total_cases": 0,
                "automated_cases": 0,
                "manual_cases": 0,
                "status_breakdown": defaultdict(int)
            }
        )
        section = defaultdict(
            lambda: {
                "total_cases": 0,
                "automated_cases": 0,
                "manual_cases": 0,
                "status_breakdown": defaultdict(int)
            }
        )

        overall_status_breakdown = defaultdict(int)
        overall_automated = 0
        overall_planned = 0
        candidates_yes = 0
        candidates_no = 0
        candidates_none = 0

        for case in all_cases:
            is_automated = (
                (case.automation_status and case.automation_status.lower() == "automated") or
                (case.execution_method and case.execution_method.lower() == "automated")
            )
            
            is_planned = (
                case.automation_status and case.automation_status.lower() == "planned"
            )

            status_key = case.automation_status or "Not Set"
            overall_status_breakdown[status_key] += 1

            func_key = case.functionality or "Unknown"
            functionality[func_key]["total_cases"] += 1
            functionality[func_key]["status_breakdown"][status_key] += 1

            section_key = case.section or "Unknown"
            section[section_key]["total_cases"] += 1
            section[section_key]["status_breakdown"][status_key] += 1

            if is_automated:
                overall_automated += 1
                functionality[func_key]["automated_cases"] += 1
                section[section_key]["automated_cases"] += 1
            else:
                functionality[func_key]["manual_cases"] += 1
                section[section_key]["manual_cases"] += 1
            
            if is_planned:
                overall_planned += 1
            
            # Count automation candidates
            candidate = (case.automation_candidate or "").strip()
            if candidate.lower() == "yes":
                candidates_yes += 1
            elif candidate.lower() == "no":
                candidates_no += 1
            else:
                candidates_none += 1

        overall_total = len(all_cases)
        by_functionality = []
        for func_name, data in sorted(functionality.items(), key=lambda x: -x[1]["total_cases"]):
            auto_pct = round((data["automated_cases"] / data["total_cases"] * 100), 1) if data["total_cases"] else 0
            by_functionality.append({
                "functionality": str(func_name),
                "total_cases": data["total_cases"],
                "automated_cases": data["automated_cases"],
                "manual_cases": data["manual_cases"],
                "automation_percentage": auto_pct,
                "status_breakdown": dict(data["status_breakdown"])
            })

        by_section = []
        for section_name, data in sorted(section.items(), key=lambda x: -x[1]["total_cases"]):
            auto_pct = round((data["automated_cases"] / data["total_cases"] * 100), 1) if data["total_cases"] else 0
            by_section.append({
                "section": str(section_name),
                "total_cases": data["total_cases"],
                "automated_cases": data["automated_cases"],
                "manual_cases": data["manual_cases"],
                "automation_percentage": auto_pct,
                "status_breakdown": dict(data["status_breakdown"])
            })

        # Automation percentage based on candidates (Yes) - the target cases to automate
        # Remaining = Candidates (Yes) - Already Automated
        remaining_to_automate = max(0, candidates_yes - overall_automated)
        automation_percentage = round((overall_automated / candidates_yes * 100), 1) if candidates_yes > 0 else 0
        
        return {
            "overall": {
                "total_core_cases": overall_total,
                "automated_cases": overall_automated,
                "manual_cases": overall_total - overall_automated,
                "planned_cases": overall_planned,
                "candidates_yes": candidates_yes,
                "candidates_no": candidates_no,
                "candidates_none": candidates_none,
                "remaining_to_automate": remaining_to_automate,
                "automation_percentage": automation_percentage,
                "status_breakdown": dict(overall_status_breakdown)
            },
            "by_functionality": by_functionality,
            "by_section": by_section
        }
    finally:
        db.close()


@app.get("/automation/search-tickets")
def automation_search_tickets(query: str = Query("", description="Search query for ticket ID")):
    """Search tickets that have automation coverage data"""
    db: Session = SessionLocal()
    try:
        ticket_ids_query = db.query(AutomationTestCase.ticket_id).filter(
            AutomationTestCase.ticket_id.isnot(None)
        ).distinct()
        
        if query:
            query_str = query.strip()
            if query_str.isdigit():
                ticket_ids_query = ticket_ids_query.filter(
                    AutomationTestCase.ticket_id.cast(String).like(f"{query_str}%")
                )
        
        ticket_ids = [row[0] for row in ticket_ids_query.all()]
        
        results = []
        for ticket_id in ticket_ids[:50]:
            cases = db.query(AutomationTestCase).filter(
                AutomationTestCase.ticket_id == ticket_id
            ).all()
            cases = _dedupe_automation_cases(cases)
            
            # Count unique run IDs for this ticket
            run_ids = set(c.run_id for c in cases if c.run_id is not None)
            run_count = len(run_ids)
            
            total = len(cases)
            automated = sum(1 for c in cases if (
                (c.automation_status and c.automation_status.lower() == "automated") or
                (c.execution_method and c.execution_method.lower() == "automated")
            ))
            
            tracking = db.query(TicketTracking).filter(
                TicketTracking.ticket_id == ticket_id
            ).first()
            
            results.append({
                "ticket_id": ticket_id,
                "title": tracking.title if tracking else f"Ticket #{ticket_id}",
                "status": tracking.status if tracking else None,
                "total_cases": total,
                "automated_cases": automated,
                "manual_cases": total - automated,
                "automation_percentage": round((automated / total * 100), 1) if total > 0 else 0,
                "run_count": run_count
            })
        
        results.sort(key=lambda x: x["ticket_id"], reverse=True)
        return results
    finally:
        db.close()


@app.get("/automation/reusability-metrics")
def automation_reusability_metrics(
    ticket_id: Optional[int] = None,
    run_id: Optional[str] = Query(None, description="Specific run ID or 'all' for combined view")
):
    """Get reusability metrics for test cases, optionally filtered by ticket and run"""
    db: Session = SessionLocal()
    try:
        query = db.query(AutomationTestCase)
        if ticket_id:
            query = query.filter(AutomationTestCase.ticket_id == ticket_id)
        if run_id and run_id != 'all':
            try:
                run_id_int = int(run_id)
                query = query.filter(AutomationTestCase.run_id == run_id_int)
            except ValueError:
                pass
        
        cases = query.all()
        
        if not cases:
            return {
                "ticket_id": ticket_id,
                "total_cases": 0,
                "reusability_breakdown": {},
                "by_automation_status": {}
            }
        
        reusability = defaultdict(lambda: {"total": 0, "automated": 0, "manual": 0})
        
        for case in cases:
            freq = case.reusability_frequency or "Not Set"
            is_automated = (
                (case.automation_status and case.automation_status.lower() == "automated") or
                (case.execution_method and case.execution_method.lower() == "automated")
            )
            
            reusability[freq]["total"] += 1
            if is_automated:
                reusability[freq]["automated"] += 1
            else:
                reusability[freq]["manual"] += 1
        
        return {
            "ticket_id": ticket_id,
            "total_cases": len(cases),
            "reusability_breakdown": {
                freq: {
                    "total": data["total"],
                    "automated": data["automated"],
                    "manual": data["manual"],
                    "automation_percentage": round((data["automated"] / data["total"] * 100), 1) if data["total"] > 0 else 0
                }
                for freq, data in reusability.items()
            }
        }
    finally:
        db.close()


@app.get("/automation/progress")
def automation_progress(
    ticket_id: Optional[int] = Query(None),
    run_id: Optional[str] = Query(None, description="Specific run ID or 'all' for combined view")
):
    """Get automation progress metrics: overall percentage, maintenance breakdown, by section and functionality"""
    db: Session = SessionLocal()
    try:
        query = db.query(AutomationTestCase)
        if ticket_id is not None:
            query = query.filter(AutomationTestCase.ticket_id == ticket_id)
        if run_id and run_id != 'all':
            try:
                run_id_int = int(run_id)
                query = query.filter(AutomationTestCase.run_id == run_id_int)
                cases = query.all()  # No deduplication for specific run
            except ValueError:
                cases = _dedupe_automation_cases(query.all())
        else:
            cases = _dedupe_automation_cases(query.all())

        if not cases:
            return {
                "ticket_id": ticket_id,
                "overall": {
                    "total_cases": 0,
                    "automated_cases": 0,
                    "remaining_cases": 0,
                    "not_automatable": 0,
                    "automation_percentage": 0,
                    "candidates_yes": 0,
                    "candidates_no": 0,
                    "candidates_none": 0,
                    "planned_cases": 0,
                },
                "maintenance_breakdown": [],
                "by_section": [],
                "by_functionality": []
            }

        total_cases = len(cases)
        automated_cases = 0
        not_automatable = 0
        planned_cases = 0
        candidates_yes = 0
        candidates_no = 0
        candidates_none = 0
        
        maintenance_counts = defaultdict(int)
        section_data = defaultdict(lambda: {"total": 0, "automated": 0})
        functionality_data = defaultdict(lambda: {"total": 0, "automated": 0})

        for case in cases:
            is_automated = (
                (case.automation_status and case.automation_status.lower() == "automated") or
                (case.execution_method and case.execution_method.lower() == "automated")
            )
            
            is_not_automatable = (
                case.automation_status and case.automation_status.lower() == "not automatable"
            )
            
            is_planned = (
                case.automation_status and case.automation_status.lower() == "planned"
            )

            if is_automated:
                automated_cases += 1
                maint_status = case.automation_maintenance or "No Maintenance Required"
                maintenance_counts[maint_status] += 1
            
            if is_not_automatable:
                not_automatable += 1
            
            if is_planned:
                planned_cases += 1
            
            # Count automation candidates
            candidate = (case.automation_candidate or "").strip()
            if candidate.lower() == "yes":
                candidates_yes += 1
            elif candidate.lower() == "no":
                candidates_no += 1
            else:
                candidates_none += 1

            # Track by section
            section = case.section or "Unknown"
            section_data[section]["total"] += 1
            if is_automated:
                section_data[section]["automated"] += 1

            # Track by functionality
            func = case.functionality or "Unknown"
            functionality_data[func]["total"] += 1
            if is_automated:
                functionality_data[func]["automated"] += 1

        # Remaining to automate = Candidates (Yes) - Already Automated
        # Automation percentage is based on candidates that should be automated
        remaining_cases = max(0, candidates_yes - automated_cases)
        automation_percentage = round((automated_cases / candidates_yes * 100), 1) if candidates_yes > 0 else 0

        # Build maintenance breakdown
        maintenance_breakdown = []
        for status, count in sorted(maintenance_counts.items(), key=lambda x: -x[1]):
            pct = round((count / automated_cases * 100), 1) if automated_cases > 0 else 0
            maintenance_breakdown.append({
                "status": status,
                "count": count,
                "percentage": pct
            })

        # Build section breakdown
        by_section = []
        for section, data in sorted(section_data.items(), key=lambda x: -x[1]["total"]):
            remaining = data["total"] - data["automated"]
            pct = round((data["automated"] / data["total"] * 100), 1) if data["total"] > 0 else 0
            by_section.append({
                "section": section,
                "total": data["total"],
                "automated": data["automated"],
                "remaining": remaining,
                "percentage": pct
            })

        # Build functionality breakdown
        by_functionality = []
        for func, data in sorted(functionality_data.items(), key=lambda x: -x[1]["total"]):
            remaining = data["total"] - data["automated"]
            pct = round((data["automated"] / data["total"] * 100), 1) if data["total"] > 0 else 0
            by_functionality.append({
                "functionality": func,
                "total": data["total"],
                "automated": data["automated"],
                "remaining": remaining,
                "percentage": pct
            })

        return {
            "ticket_id": ticket_id,
            "overall": {
                "total_cases": total_cases,
                "automated_cases": automated_cases,
                "remaining_cases": remaining_cases,
                "not_automatable": not_automatable,
                "automation_percentage": automation_percentage,
                "candidates_yes": candidates_yes,
                "candidates_no": candidates_no,
                "candidates_none": candidates_none,
                "planned_cases": planned_cases,
            },
            "maintenance_breakdown": maintenance_breakdown,
            "by_section": by_section,
            "by_functionality": by_functionality
        }
    finally:
        db.close()


@app.get("/automation/planned-cases")
def automation_planned_cases(
    period: Optional[str] = Query(None, description="Filter period: day, week, month, quarter, year, all"),
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    automation_candidate: Optional[str] = Query(None, description="Filter by automation candidate: Yes, No, None"),
    ticket_id: Optional[int] = Query(None, description="Filter by ticket ID"),
):
    """Get planned automation cases with date filters and trend data."""
    db: Session = SessionLocal()
    try:
        query = db.query(AutomationTestCase).filter(
            AutomationTestCase.automation_status.ilike("planned")
        )
        
        if ticket_id is not None:
            query = query.filter(AutomationTestCase.ticket_id == ticket_id)
        
        if automation_candidate:
            query = query.filter(AutomationTestCase.automation_candidate == automation_candidate)
        
        # Date filtering based on planned_on
        now = datetime.now()
        filter_start = None
        filter_end = None
        
        if start_date and end_date:
            try:
                filter_start = datetime.strptime(start_date, "%Y-%m-%d")
                filter_end = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            except ValueError:
                pass
        elif period:
            if period == "day":
                filter_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
                filter_end = now
            elif period == "week":
                filter_start = now - timedelta(days=now.weekday())
                filter_start = filter_start.replace(hour=0, minute=0, second=0, microsecond=0)
                filter_end = now
            elif period == "month":
                filter_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                filter_end = now
            elif period == "quarter":
                quarter_start_month = ((now.month - 1) // 3) * 3 + 1
                filter_start = now.replace(month=quarter_start_month, day=1, hour=0, minute=0, second=0, microsecond=0)
                filter_end = now
            elif period == "year":
                filter_start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
                filter_end = now
        
        # Get all planned cases (for total count)
        all_planned_cases = _dedupe_automation_cases(query.all())
        total_planned = len(all_planned_cases)
        
        # Filter by date range if specified
        if filter_start and filter_end:
            period_cases = [c for c in all_planned_cases if c.planned_on and filter_start <= c.planned_on <= filter_end]
        else:
            period_cases = all_planned_cases
        
        planned_in_period = len(period_cases)
        
        # Calculate candidates pending planning
        candidates_query = db.query(AutomationTestCase).filter(
            AutomationTestCase.automation_candidate == "Yes",
            or_(
                AutomationTestCase.automation_status.is_(None),
                ~AutomationTestCase.automation_status.ilike("planned"),
                ~AutomationTestCase.automation_status.ilike("automated"),
                ~AutomationTestCase.automation_status.ilike("in progress")
            )
        )
        if ticket_id is not None:
            candidates_query = candidates_query.filter(AutomationTestCase.ticket_id == ticket_id)
        candidates_pending = len(_dedupe_automation_cases(candidates_query.all()))
        
        # Group by date for trend chart
        by_date = defaultdict(int)
        for case in period_cases:
            if case.planned_on:
                date_str = case.planned_on.strftime("%Y-%m-%d")
                by_date[date_str] += 1
        
        by_date_list = [{"date": d, "count": c} for d, c in sorted(by_date.items())]
        
        # Build case list
        cases_list = []
        for case in period_cases:
            cases_list.append({
                "case_id": case.case_id,
                "test_id": case.test_id,
                "title": case.title,
                "ticket_id": case.ticket_id,
                "section": case.section,
                "functionality": case.functionality,
                "automation_candidate": case.automation_candidate,
                "automation_status": case.automation_status,
                "planned_on": case.planned_on.isoformat() if case.planned_on else None,
                "business_criticality": case.business_criticality,
            })
        
        return {
            "summary": {
                "total_planned": total_planned,
                "planned_in_period": planned_in_period,
                "candidates_pending_planning": candidates_pending,
            },
            "by_date": by_date_list,
            "cases": cases_list,
            "filters": {
                "period": period,
                "start_date": start_date,
                "end_date": end_date,
                "automation_candidate": automation_candidate,
                "ticket_id": ticket_id,
            }
        }
    finally:
        db.close()


@app.get("/automation/automated-cases")
def automation_automated_cases(
    period: Optional[str] = Query(None, description="Filter period: day, week, month, quarter, year, all"),
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    automation_candidate: Optional[str] = Query(None, description="Filter by automation candidate: Yes, No, None"),
    ticket_id: Optional[int] = Query(None, description="Filter by ticket ID"),
):
    """Get automated cases with date filters and trend data."""
    db: Session = SessionLocal()
    try:
        query = db.query(AutomationTestCase).filter(
            or_(
                AutomationTestCase.automation_status.ilike("automated"),
                AutomationTestCase.execution_method.ilike("automated")
            )
        )
        
        if ticket_id is not None:
            query = query.filter(AutomationTestCase.ticket_id == ticket_id)
        
        if automation_candidate:
            query = query.filter(AutomationTestCase.automation_candidate == automation_candidate)
        
        # Date filtering based on automated_on
        now = datetime.now()
        filter_start = None
        filter_end = None
        
        if start_date and end_date:
            try:
                filter_start = datetime.strptime(start_date, "%Y-%m-%d")
                filter_end = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            except ValueError:
                pass
        elif period:
            if period == "day":
                filter_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
                filter_end = now
            elif period == "week":
                filter_start = now - timedelta(days=now.weekday())
                filter_start = filter_start.replace(hour=0, minute=0, second=0, microsecond=0)
                filter_end = now
            elif period == "month":
                filter_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                filter_end = now
            elif period == "quarter":
                quarter_start_month = ((now.month - 1) // 3) * 3 + 1
                filter_start = now.replace(month=quarter_start_month, day=1, hour=0, minute=0, second=0, microsecond=0)
                filter_end = now
            elif period == "year":
                filter_start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
                filter_end = now
        
        # Get all automated cases (for total count)
        all_automated_cases = _dedupe_automation_cases(query.all())
        total_automated = len(all_automated_cases)
        
        # Filter by date range if specified
        if filter_start and filter_end:
            period_cases = [c for c in all_automated_cases if c.automated_on and filter_start <= c.automated_on <= filter_end]
        else:
            period_cases = all_automated_cases
        
        automated_in_period = len(period_cases)
        
        # Calculate automation candidate stats
        candidates_yes = len([c for c in all_automated_cases if c.automation_candidate == "Yes"])
        candidates_no = len([c for c in all_automated_cases if c.automation_candidate == "No"])
        
        # Calculate pending automation (Candidate=Yes but not yet Automated)
        pending_query = db.query(AutomationTestCase).filter(
            AutomationTestCase.automation_candidate == "Yes",
            ~or_(
                AutomationTestCase.automation_status.ilike("automated"),
                AutomationTestCase.execution_method.ilike("automated")
            )
        )
        if ticket_id is not None:
            pending_query = pending_query.filter(AutomationTestCase.ticket_id == ticket_id)
        pending_automation = len(_dedupe_automation_cases(pending_query.all()))
        
        # Group by date for trend chart
        by_date = defaultdict(int)
        for case in period_cases:
            if case.automated_on:
                date_str = case.automated_on.strftime("%Y-%m-%d")
                by_date[date_str] += 1
        
        by_date_list = [{"date": d, "count": c} for d, c in sorted(by_date.items())]
        
        # Build case list
        cases_list = []
        for case in period_cases:
            cases_list.append({
                "case_id": case.case_id,
                "test_id": case.test_id,
                "title": case.title,
                "ticket_id": case.ticket_id,
                "section": case.section,
                "functionality": case.functionality,
                "automation_candidate": case.automation_candidate,
                "automation_status": case.automation_status,
                "automated_on": case.automated_on.isoformat() if case.automated_on else None,
                "business_criticality": case.business_criticality,
                "status_name": case.status_name,
            })
        
        return {
            "summary": {
                "total_automated": total_automated,
                "automated_in_period": automated_in_period,
                "automation_candidates_yes": candidates_yes,
                "automation_candidates_no": candidates_no,
                "pending_automation": pending_automation,
            },
            "by_date": by_date_list,
            "cases": cases_list,
            "filters": {
                "period": period,
                "start_date": start_date,
                "end_date": end_date,
                "automation_candidate": automation_candidate,
                "ticket_id": ticket_id,
            }
        }
    finally:
        db.close()


@app.get("/automation/workflow-summary")
def automation_workflow_summary(ticket_id: Optional[int] = Query(None)):
    """Get automation workflow summary: Candidate → Planned → Automated counts."""
    db: Session = SessionLocal()
    try:
        query = db.query(AutomationTestCase)
        if ticket_id is not None:
            query = query.filter(AutomationTestCase.ticket_id == ticket_id)
        
        all_cases = _dedupe_automation_cases(query.all())
        
        # Count by workflow stage
        candidates_yes = 0
        candidates_no = 0
        planned = 0
        automated = 0
        not_automatable = 0
        
        for case in all_cases:
            # Count candidates
            if case.automation_candidate == "Yes":
                candidates_yes += 1
            elif case.automation_candidate == "No":
                candidates_no += 1
            
            # Count by status
            status = (case.automation_status or "").lower()
            exec_method = (case.execution_method or "").lower()
            
            if status == "automated" or exec_method == "automated":
                automated += 1
            elif status == "planned":
                planned += 1
            elif status == "not automatable":
                not_automatable += 1
        
        # Pending = Candidate Yes but not Planned or Automated
        pending_planning = 0
        for case in all_cases:
            if case.automation_candidate == "Yes":
                status = (case.automation_status or "").lower()
                exec_method = (case.execution_method or "").lower()
                if status not in ["planned", "automated", "in progress"] and exec_method != "automated":
                    pending_planning += 1
        
        return {
            "workflow": {
                "candidates_yes": candidates_yes,
                "candidates_no": candidates_no,
                "pending_planning": pending_planning,
                "planned": planned,
                "automated": automated,
                "not_automatable": not_automatable,
            },
            "total_cases": len(all_cases),
            "ticket_id": ticket_id,
        }
    finally:
        db.close()


@app.get("/automation/sync-status")
def automation_sync_status():
    """Get the last sync status for TestRail automation data."""
    db: Session = SessionLocal()
    try:
        last_sync = db.query(SyncLog).filter(
            SyncLog.sync_source == "testrail_automation"
        ).order_by(SyncLog.completed_at.desc()).first()
        
        if not last_sync:
            return {
                "last_sync": None,
                "success": None,
                "message": "No sync has been performed yet"
            }
        
        return {
            "last_sync": last_sync.completed_at.isoformat() if last_sync.completed_at else None,
            "started_at": last_sync.started_at.isoformat() if last_sync.started_at else None,
            "success": last_sync.success,
            "message": last_sync.message,
            "records_processed": last_sync.total_records,
            "records_created": last_sync.records_added,
            "duration_seconds": last_sync.duration_seconds
        }
    finally:
        db.close()


@app.post("/automation/sync")
def automation_sync():
    """Trigger a sync of automation coverage data from TestRail Project 18.
    This runs the sync_automation_testrail script to fetch latest data.
    """
    import subprocess
    import sys
    
    try:
        script_path = os.path.join(os.path.dirname(__file__), "sync_automation_testrail.py")
        
        if not os.path.exists(script_path):
            raise HTTPException(status_code=500, detail="Sync script not found")
        
        result = subprocess.run(
            [sys.executable, script_path],
            capture_output=True,
            text=True,
            timeout=300,
            cwd=os.path.dirname(__file__)
        )
        
        if result.returncode != 0:
            return {
                "success": False,
                "message": "Sync failed",
                "error": result.stderr[-2000:] if result.stderr else "Unknown error",
                "output": result.stdout[-2000:] if result.stdout else ""
            }
        
        return {
            "success": True,
            "message": "Automation coverage data synced successfully",
            "output": result.stdout[-2000:] if result.stdout else ""
        }
        
    except subprocess.TimeoutExpired:
        raise HTTPException(status_code=504, detail="Sync timed out after 5 minutes")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/automation/sync-to-sheets")
def automation_sync_to_google_sheets(skip_testrail: bool = Query(False, description="Skip TestRail sync, only export existing data to Sheets")):
    """
    Full sync: Fetch TestRail data → Database → Google Sheets.
    This performs:
    1. Sync TestRail automation data to database (with 30s timeout, skippable)
    2. Export TestRail data to Google Sheets
    Returns the sync timestamp for display in the sheet.
    Even if TestRail sync fails, will still export existing data to Sheets.
    Set skip_testrail=true to only export existing data without fetching from TestRail.
    """
    import subprocess
    import sys
    from datetime import datetime
    
    results = {
        "testrail_sync": None,
        "sheets_export": None,
        "timestamp": None,
        "success": False
    }
    
    # Step 1: Sync TestRail to Database (with shorter timeout, continue on failure)
    if skip_testrail:
        results["testrail_sync"] = {
            "success": True,
            "skipped": True,
            "message": "TestRail sync skipped, using existing data"
        }
    else:
        try:
            script_path = os.path.join(os.path.dirname(__file__), "sync_automation_testrail.py")
            
            if os.path.exists(script_path):
                sync_result = subprocess.run(
                    [sys.executable, script_path],
                    capture_output=True,
                    text=True,
                    timeout=30,
                    cwd=os.path.dirname(__file__)
                )
                
                if sync_result.returncode != 0:
                    results["testrail_sync"] = {
                        "success": False,
                        "error": (sync_result.stderr[-500:] if sync_result.stderr else "") + 
                                 (sync_result.stdout[-500:] if sync_result.stdout else "") or "Unknown error"
                    }
                else:
                    results["testrail_sync"] = {"success": True}
            else:
                results["testrail_sync"] = {
                    "success": False,
                    "error": "TestRail sync script not found"
                }
        except subprocess.TimeoutExpired:
            results["testrail_sync"] = {
                "success": False,
                "error": "TestRail sync timed out (30s). Continuing with existing data..."
            }
        except Exception as e:
            results["testrail_sync"] = {
                "success": False,
                "error": f"TestRail sync error: {str(e)}"
            }
    
    # Step 2: Export to Google Sheets (always attempt this)
    try:
        exporter = get_sheets_exporter()
        if exporter:
            db = SessionLocal()
            try:
                runs_result = exporter.export_testrail_runs(db)
                cases_result = exporter.export_testrail_cases(db)
                
                results["sheets_export"] = {
                    "success": runs_result.get("success", False) and cases_result.get("success", False),
                    "runs_exported": runs_result.get("rows", 0),
                    "cases_exported": cases_result.get("rows", 0)
                }
            finally:
                db.close()
        else:
            results["sheets_export"] = {
                "success": False,
                "error": "Google Sheets export not configured"
            }
    except Exception as e:
        results["sheets_export"] = {
            "success": False,
            "error": str(e)
        }
    
    # Set timestamp
    results["timestamp"] = datetime.now().isoformat()
    
    # Success if sheets export worked (TestRail sync is optional since we may have existing data)
    results["success"] = results["sheets_export"].get("success", False)
    
    return results


@app.post("/automation/export-to-sheets")
def automation_export_to_google_sheets_only():
    """
    Export existing TestRail data from database to Google Sheets.
    Does NOT fetch new data from TestRail - uses existing database data.
    Use this when TestRail API is unreachable but you want to sync existing data to Sheets.
    """
    from datetime import datetime
    
    results = {
        "sheets_export": None,
        "timestamp": None,
        "success": False
    }
    
    try:
        exporter = get_sheets_exporter()
        if exporter:
            db = SessionLocal()
            try:
                runs_result = exporter.export_testrail_runs(db)
                cases_result = exporter.export_testrail_cases(db)
                
                results["sheets_export"] = {
                    "success": runs_result.get("success", False) and cases_result.get("success", False),
                    "runs_exported": runs_result.get("rows", 0),
                    "cases_exported": cases_result.get("rows", 0)
                }
            finally:
                db.close()
        else:
            results["sheets_export"] = {
                "success": False,
                "error": "Google Sheets export not configured"
            }
    except Exception as e:
        results["sheets_export"] = {
            "success": False,
            "error": str(e)
        }
    
    results["timestamp"] = datetime.now().isoformat()
    results["success"] = results["sheets_export"].get("success", False)
    
    return results


# ===== TICKET TRACKING ENDPOINTS =====

@app.get("/tickets/search")
def search_tickets(query: str = Query("", description="Search query for ticket ID or title")):
    """Search tickets for autocomplete - returns matching ticket IDs from PM tracker Excel import only"""
    db: Session = SessionLocal()
    try:
        # Get tickets ONLY from TicketTracking (PM tracker Excel import)
        tracking_tickets = db.query(TicketTracking).all()
        
        # Build a map of ticket_id -> first bug subject for titles (fallback when tracking has no title)
        ticket_id_to_title = {}
        if tracking_tickets:
            ticket_ids = [t.ticket_id for t in tracking_tickets]
            bugs = db.query(Bug.ticket_id, Bug.subject).filter(
                Bug.ticket_id.in_(ticket_ids)
            ).distinct(Bug.ticket_id).all()
            for bug in bugs:
                if bug.ticket_id and bug.subject:
                    title_parts = bug.subject.split(" - ")
                    ticket_id_to_title[bug.ticket_id] = title_parts[0] if title_parts else bug.subject
        
        # Build ticket list from TicketTracking; prefer title from PM API (tracking.title)
        tickets = []
        for t in tracking_tickets:
            title = (getattr(t, 'title', None) or '').strip() or ticket_id_to_title.get(t.ticket_id) or f"Ticket #{t.ticket_id}"
            ticket_data = {
                "ticket_id": t.ticket_id,
                "title": title,
                "status": t.status,
                "assignee": t.current_assignee
            }
            tickets.append(ticket_data)
        
        # Filter by query if provided
        if query:
            query_str = query.strip()
            
            # First, find tickets where ticket_id STARTS WITH the query
            starts_with = [
                t for t in tickets
                if str(t["ticket_id"]).startswith(query_str)
            ]
            
            # If we have matches that start with the query, return only those
            if starts_with:
                # Sort by ticket_id descending (most recent first)
                starts_with.sort(key=lambda x: x["ticket_id"], reverse=True)
                return starts_with[:50]
            
            # Otherwise, fall back to tickets that CONTAIN the query anywhere
            query_lower = query_str.lower()
            contains = [
                t for t in tickets
                if query_str in str(t["ticket_id"]) or query_lower in (t["title"] or "").lower()
            ]
            
            # Sort by ticket_id descending (most recent first)
            contains.sort(key=lambda x: x["ticket_id"], reverse=True)
            return contains[:50]
        
        # No query - return all tickets sorted by ticket_id descending
        tickets.sort(key=lambda x: x["ticket_id"], reverse=True)
        
        # Limit results for performance
        return tickets[:50]
    finally:
        db.close()


@app.get("/ticket-tracking/summary/all")
def get_ticket_tracking_summary():
    """Get overview metrics for all tracked tickets"""
    db: Session = SessionLocal()
    try:
        all_tracking = db.query(TicketTracking).all()
        
        if not all_tracking:
            return {
                "total_tickets": 0,
                "avg_dev_estimate": 0,
                "avg_dev_actual": 0,
                "avg_qa_estimate": 0,
                "avg_qa_actual": 0,
                "dev_efficiency": 0,
                "qa_efficiency": 0,
                "status_breakdown": {}
            }
        
        total = len(all_tracking)
        
        # Calculate averages
        dev_estimates = [t.dev_estimate_hours for t in all_tracking if t.dev_estimate_hours]
        dev_actuals = [t.actual_dev_hours for t in all_tracking if t.actual_dev_hours]
        qa_estimates = [t.qa_estimate_hours for t in all_tracking if t.qa_estimate_hours]
        qa_actuals = [t.actual_qa_hours for t in all_tracking if t.actual_qa_hours]
        
        avg_dev_estimate = sum(dev_estimates) / len(dev_estimates) if dev_estimates else 0
        avg_dev_actual = sum(dev_actuals) / len(dev_actuals) if dev_actuals else 0
        avg_qa_estimate = sum(qa_estimates) / len(qa_estimates) if qa_estimates else 0
        avg_qa_actual = sum(qa_actuals) / len(qa_actuals) if qa_actuals else 0
        
        # Calculate efficiency (how well estimates match actual)
        dev_efficiency = (avg_dev_estimate / avg_dev_actual * 100) if avg_dev_actual > 0 else 100
        qa_efficiency = (avg_qa_estimate / avg_qa_actual * 100) if avg_qa_actual > 0 else 100
        
        # Status breakdown
        status_counts = defaultdict(int)
        for t in all_tracking:
            status_counts[t.status or "Unknown"] += 1
        
        return {
            "total_tickets": total,
            "avg_dev_estimate": round(avg_dev_estimate, 1),
            "avg_dev_actual": round(avg_dev_actual, 1),
            "avg_qa_estimate": round(avg_qa_estimate, 1),
            "avg_qa_actual": round(avg_qa_actual, 1),
            "dev_efficiency": round(dev_efficiency, 1),
            "qa_efficiency": round(qa_efficiency, 1),
            "status_breakdown": dict(status_counts)
        }
    finally:
        db.close()


@app.get("/ticket-tracking/team-metrics")
def get_team_metrics():
    """Get developer/QC productivity metrics"""
    db: Session = SessionLocal()
    try:
        all_tracking = db.query(TicketTracking).all()
        
        if not all_tracking:
            return {
                "developers": {},
                "qc_testers": {}
            }
        
        # Developer metrics
        dev_metrics = defaultdict(lambda: {"tickets": 0, "total_hours": 0, "total_estimate": 0})
        qc_metrics = defaultdict(lambda: {"tickets": 0, "total_hours": 0, "total_estimate": 0})
        
        for t in all_tracking:
            # Backend developer
            if t.backend_developer:
                dev_metrics[t.backend_developer]["tickets"] += 1
                if t.actual_dev_hours:
                    dev_metrics[t.backend_developer]["total_hours"] += t.actual_dev_hours
                if t.dev_estimate_hours:
                    dev_metrics[t.backend_developer]["total_estimate"] += t.dev_estimate_hours
            
            # Frontend developer
            if t.frontend_developer:
                dev_metrics[t.frontend_developer]["tickets"] += 1
                if t.actual_dev_hours:
                    dev_metrics[t.frontend_developer]["total_hours"] += t.actual_dev_hours
                if t.dev_estimate_hours:
                    dev_metrics[t.frontend_developer]["total_estimate"] += t.dev_estimate_hours
            
            # QC Tester
            if t.qc_tester:
                qc_metrics[t.qc_tester]["tickets"] += 1
                if t.actual_qa_hours:
                    qc_metrics[t.qc_tester]["total_hours"] += t.actual_qa_hours
                if t.qa_estimate_hours:
                    qc_metrics[t.qc_tester]["total_estimate"] += t.qa_estimate_hours
        
        # Calculate efficiency for each person
        for dev, data in dev_metrics.items():
            if data["total_hours"] > 0:
                data["efficiency"] = round((data["total_estimate"] / data["total_hours"]) * 100, 1)
            else:
                data["efficiency"] = 100
            data["total_hours"] = round(data["total_hours"], 1)
            data["total_estimate"] = round(data["total_estimate"], 1)
        
        for qc, data in qc_metrics.items():
            if data["total_hours"] > 0:
                data["efficiency"] = round((data["total_estimate"] / data["total_hours"]) * 100, 1)
            else:
                data["efficiency"] = 100
            data["total_hours"] = round(data["total_hours"], 1)
            data["total_estimate"] = round(data["total_estimate"], 1)
        
        return {
            "developers": dict(dev_metrics),
            "qc_testers": dict(qc_metrics)
        }
    finally:
        db.close()


# ===== PM TRACKER SYNC (API only; ticket_id is the mapping key) =====

@app.post("/ticket-tracking/sync-latest")
def sync_latest_ticket_report():
    """
    Sync PM Tracker data from the API only.
    Tickets are upserted by ticket_id (main mapping key).
    
    Returns:
        {
            "success": bool,
            "message": str,
            "sync_source": "api",
            "records_added": int,
            "records_updated": int,
            "records_skipped": int,
            "errors": int,
            "duration_seconds": float
        }
    """
    db = SessionLocal()
    start_time = time.time()
    
    try:
        success, message, stats, sync_source, _, _ = _sync_from_api(db, start_time)
        duration_seconds = time.time() - start_time
        
        return {
            "success": success,
            "message": message,
            "sync_source": sync_source or "api",
            "records_added": stats.get('records_added', 0),
            "records_updated": stats.get('records_updated', 0),
            "records_skipped": stats.get('records_skipped', 0),
            "records_skipped_unchanged_closed": stats.get('records_skipped_unchanged_closed', 0),
            "errors": stats.get('errors', 0),
            "duration_seconds": round(duration_seconds, 2)
        }
    
    except Exception as e:
        logging.exception("Unexpected error in sync-latest endpoint")
        return {
            "success": False,
            "message": f"Unexpected error: {str(e)}",
            "sync_source": "api",
            "records_added": 0,
            "records_updated": 0,
            "records_skipped": 0,
            "records_skipped_unchanged_closed": 0,
            "errors": 1,
            "duration_seconds": round(time.time() - start_time, 2)
        }
    finally:
        db.close()


def _sync_from_api(
    db: Session,
    start_time: float,
    fallback_from: Optional[str] = None,
    fallback_reason: Optional[str] = None
) -> tuple:
    """
    Perform sync from PM Tracker API (delegates to pm_sync_runner).
    Returns:
        (success, message, stats, sync_source, fallback_occurred, fallback_reason_str)
    """
    success, message, stats, sync_source = run_pm_api_sync(db, start_time)
    return success, message, stats, sync_source, fallback_from is not None, fallback_reason


# ===== SYNC HEALTH ENDPOINTS =====

@app.get("/sync/health")
def get_sync_health():
    """
    Unified sync health status for all sync sources.
    Returns freshness (FRESH/STALE/CRITICAL), consecutive failures,
    pause status, and last sync details for PM Tracker, Redmine, and Google Sheets.
    """
    return sync_health.get_overall_health()


@app.get("/sync/health/{source_name}")
def get_sync_health_source(source_name: str):
    """Get health status for a specific sync source."""
    valid_sources = ["pm_tracker", "redmine", "google_sheets"]
    if source_name not in valid_sources:
        raise HTTPException(status_code=404, detail=f"Unknown sync source: {source_name}. Valid: {valid_sources}")
    return sync_health.get_source(source_name).get_status()


@app.post("/sync/health/{source_name}/unpause")
def unpause_sync_source(source_name: str):
    """
    Unpause a sync source that was auto-paused due to consecutive failures.
    Use this after fixing the underlying issue (e.g., re-authenticating MFA).
    """
    if source_name == "pm_tracker":
        return unpause_pm_sync()
    valid_sources = ["pm_tracker", "redmine", "google_sheets"]
    if source_name not in valid_sources:
        raise HTTPException(status_code=404, detail=f"Unknown sync source: {source_name}")
    health = sync_health.get_source(source_name)
    was_paused = health.is_paused
    health.unpause()
    return {
        "was_paused": was_paused,
        "is_paused": health.is_paused,
        "message": f"{source_name} sync unpaused." if was_paused else f"{source_name} was not paused.",
    }


@app.get("/ticket-tracking/sync-method")
def get_sync_method():
    """
    Get last PM Tracker API sync status (data is always synced from API; ticket_id is the mapping key).
    
    Returns:
        {
            "current_method": "api",
            "last_sync": { ... } | null
        }
    """
    db = SessionLocal()
    try:
        last_sync = get_last_sync_info(db)
        return {
            "current_method": "api",
            "last_sync": last_sync
        }
    finally:
        db.close()


@app.post("/ticket-tracking/test-api-connection")
def test_api_connection():
    """
    Test PM Tracker API connection and authentication
    
    Returns:
        {
            "success": bool,
            "message": str
        }
    """
    try:
        client = PMApiClient()
        success, message = client.test_connection()
        return {
            "success": success,
            "message": message
        }
    except Exception as e:
        return {
            "success": False,
            "message": f"Connection test failed: {str(e)}"
        }


@app.get("/pm-ticket-raw/{ticket_id}")
def get_pm_ticket_raw(ticket_id: int):
    """
    Fetch and return all raw data from PM Tracker API for a specific ticket.
    Useful for debugging field mapping and inspecting available fields.
    
    Returns:
        {
            "success": bool,
            "ticket_id": int,
            "raw": dict (all API fields as returned),
            "mapped": dict (after field mapping),
            "field_names": list (all API field names)
        }
    """
    try:
        client = PMApiClient()
        success, tickets, message = client.fetch_tickets()
        if not success or not tickets:
            return {
                "success": False,
                "ticket_id": ticket_id,
                "error": message,
                "raw": None,
                "mapped": None,
                "field_names": [],
            }
        # Find ticket by TicketNumber, ticket_id, id, or ticket_number (case-insensitive key lookup)
        ticket_id_str = str(ticket_id)
        raw_ticket = None
        for t in tickets:
            t_lower = {k.lower(): v for k, v in t.items()}
            tid = (
                t_lower.get("ticketnumber") or t_lower.get("ticket_id") or t_lower.get("id")
                or t_lower.get("ticket_number") or t_lower.get("ticketid")
            )
            if tid is not None and str(tid) == ticket_id_str:
                raw_ticket = t
                break
        if not raw_ticket:
            return {
                "success": True,
                "ticket_id": ticket_id,
                "error": f"Ticket {ticket_id} not found in PM API response ({len(tickets)} tickets returned)",
                "raw": None,
                "mapped": None,
                "field_names": list(tickets[0].keys()) if tickets else [],
            }
        mapped = client.map_api_fields([raw_ticket])
        return {
            "success": True,
            "ticket_id": ticket_id,
            "raw": raw_ticket,
            "mapped": mapped[0] if mapped else None,
            "field_names": list(raw_ticket.keys()),
        }
    except Exception as e:
        return {
            "success": False,
            "ticket_id": ticket_id,
            "error": str(e),
            "raw": None,
            "mapped": None,
            "field_names": [],
        }


@app.delete("/ticket-tracking/sync-history")
def clear_sync_history(days: int = 0):
    """
    Clear old sync history logs
    
    Args:
        days: If > 0, delete logs older than this many days. If 0, delete all logs.
        
    Returns:
        {
            "success": bool,
            "message": str,
            "records_deleted": int
        }
    """
    db = SessionLocal()
    try:
        if days > 0:
            from datetime import datetime, timedelta
            cutoff_date = datetime.utcnow() - timedelta(days=days)
            from models import SyncLog
            result = db.query(SyncLog).filter(
                SyncLog.started_at < cutoff_date
            ).delete(synchronize_session=False)
        else:
            from models import SyncLog
            result = db.query(SyncLog).delete(synchronize_session=False)
        
        db.commit()
        
        return {
            "success": True,
            "message": f"Deleted {result} sync log records",
            "records_deleted": result
        }
    except Exception as e:
        db.rollback()
        return {
            "success": False,
            "message": str(e),
            "records_deleted": 0
        }
    finally:
        db.close()


@app.get("/ticket-tracking/sync-history")
def get_sync_history(limit: int = Query(50, ge=1, le=500)):
    """
    Get sync operation history
    
    Args:
        limit: Maximum number of records to return (1-500)
        
    Returns:
        List of sync log records with most recent first
    """
    db = SessionLocal()
    try:
        logs = db.query(SyncLog).order_by(
            SyncLog.started_at.desc()
        ).limit(limit).all()
        
        return [
            {
                "id": log.id,
                "sync_source": log.sync_source,
                "success": log.success,
                "message": log.message,
                "records_added": log.records_added,
                "records_updated": log.records_updated,
                "records_skipped": log.records_skipped,
                "errors": log.errors,
                "duration_seconds": log.duration_seconds,
                "fallback_from": log.fallback_from,
                "fallback_reason": log.fallback_reason,
                "started_at": log.started_at.isoformat() if log.started_at else None,
                "completed_at": log.completed_at.isoformat() if log.completed_at else None,
            }
            for log in logs
        ]
    finally:
        db.close()


@app.get("/ticket-tracking/sync-status")
def get_ticket_sync_status():
    """Get status of last PM Tracker API sync and auto-sync scheduler (ticket_id is the mapping key)."""
    db = SessionLocal()
    try:
        last_updated = db.query(func.max(TicketTracking.updated_on)).scalar()
        return {
            "last_db_update": last_updated.isoformat() if last_updated else None,
            "source": "api",
            "auto_sync": get_pm_scheduler_status()
        }
    finally:
        db.close()


@app.get("/redmine/sync/status")
def get_redmine_sync_status():
    """Get status of Redmine auto-sync scheduler and last sync result."""
    return get_redmine_scheduler_status()


@app.get("/testrail/sync/status")
def get_testrail_sync_status():
    """Get status of TestRail auto-sync scheduler and last sync result."""
    return get_testrail_scheduler_status()


@app.post("/testrail/sync")
def trigger_testrail_sync():
    """Trigger TestRail sync immediately."""
    try:
        result = trigger_testrail_sync_now()
        status_code = 200 if result.get("success") else 500
        if status_code != 200:
            raise HTTPException(status_code=status_code, detail=result)
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"TestRail sync failed: {str(e)}")


@app.post("/redmine/sync")
def trigger_redmine_sync(all_bugs: bool = Query(True, description="Include all bugs (default True for accurate counts)")):
    """
    Trigger Redmine bug sync to database.
    By default fetches ALL bugs including closed ones to ensure accurate bug counts.
    Set all_bugs=false to use query_id=20 which may exclude some bugs.
    """
    try:
        processed, created, updated = sync_redmine_bugs(all_bugs=all_bugs)
        return {
            "success": True,
            "message": "Redmine sync completed",
            "processed": processed,
            "created": created,
            "updated": updated,
            "auto_sync": get_redmine_scheduler_status(),
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Redmine sync failed: {str(e)}")


@app.get("/ticket-tracking/{ticket_id}")
def get_ticket_tracking(ticket_id: int):
    """Get tracking data for a specific ticket, including developers from Redmine"""
    db: Session = SessionLocal()
    try:
        tracking = db.query(TicketTracking).filter(TicketTracking.ticket_id == ticket_id).first()
        
        # Get developers from Redmine bugs for this ticket
        bugs = db.query(Bug).filter(Bug.ticket_id == ticket_id).all()
        redmine_developers = set()
        for bug in bugs:
            if bug.assignee and bug.assignee.strip():
                redmine_developers.add(bug.assignee.strip())
        
        if not tracking:
            # Return just Redmine data if no tracking data
            times_moved_to_fail = get_qc_fail_count(db, ticket_id)
            if redmine_developers:
                return {
                    "ticket_id": ticket_id,
                    "status": None,
                    "developers": list(redmine_developers),
                    "qc_testers": [],
                    "eta": None,
                    "current_assignee": None,
                    "dev_estimate_hours": None,
                    "actual_dev_hours": None,
                    "qa_estimate_hours": None,
                    "actual_qa_hours": None,
                    "dev_deviation": None,
                    "qa_deviation": None,
                    "qa_vs_dev_ratio": None,
                    "updated_on": None,
                    "times_moved_to_fail": times_moved_to_fail,
                }
            return {"ticket_id": ticket_id, "times_moved_to_fail": times_moved_to_fail}
        
        # Collect all developers (from tracking + Redmine)
        developers = set()
        if tracking.backend_developer:
            developers.add(tracking.backend_developer.strip())
        if tracking.frontend_developer:
            developers.add(tracking.frontend_developer.strip())
        if tracking.developer_assigned:
            developers.add(tracking.developer_assigned.strip())
        developers.update(redmine_developers)
        developers = [d for d in developers if d]
        
        qc_testers = []
        if tracking.qc_tester:
            qc_testers = [t.strip() for t in tracking.qc_tester.split(',') if t.strip()]
        
        dev_deviation = None
        if tracking.dev_estimate_hours and tracking.actual_dev_hours:
            dev_deviation = round(tracking.actual_dev_hours - tracking.dev_estimate_hours, 1)
        
        qa_deviation = None
        if tracking.qa_estimate_hours and tracking.actual_qa_hours:
            qa_deviation = round(tracking.actual_qa_hours - tracking.qa_estimate_hours, 1)
        
        qa_vs_dev_ratio = None
        if tracking.actual_dev_hours and tracking.actual_qa_hours and tracking.actual_dev_hours > 0:
            qa_vs_dev_ratio = round((tracking.actual_qa_hours / tracking.actual_dev_hours) * 100, 1)

        # Ageing: created_on -> closed_on or today
        today = datetime.now().date()
        is_closed = (tracking.status or '').lower() in ['closed', 'moved to live', 'completed']
        _, ageing_days, days_to_close = _ticket_ageing(tracking, today, is_closed)

        # Priority change history
        priority_history = db.query(TicketPriorityHistory).filter(
            TicketPriorityHistory.ticket_id == ticket_id
        ).order_by(TicketPriorityHistory.changed_on.desc()).limit(50).all()
        priority_history_list = [
            {
                "previous_priority": h.previous_priority,
                "new_priority": h.new_priority,
                "changed_on": h.changed_on.isoformat() if h.changed_on else None,
            }
            for h in priority_history
        ]
        
        times_moved_to_fail = get_qc_fail_count(db, ticket_id)
        return {
            "ticket_id": tracking.ticket_id,
            "title": (getattr(tracking, 'title', None) or '').strip() or None,
            "priority": getattr(tracking, 'priority', None) or None,
            "status": tracking.status,
            "created_on": tracking.created_on.isoformat() if getattr(tracking, 'created_on', None) else None,
            "closed_on": tracking.closed_on.isoformat() if getattr(tracking, 'closed_on', None) else None,
            "ageing_days": ageing_days,
            "days_to_close": days_to_close,
            "priority_history": priority_history_list,
            "times_moved_to_fail": times_moved_to_fail,
            "developers": developers,
            "qc_testers": qc_testers,
            "eta": tracking.eta.isoformat() if tracking.eta else None,
            "current_assignee": tracking.current_assignee,
            "backend_developer": tracking.backend_developer,
            "frontend_developer": tracking.frontend_developer,
            "dev_estimate_hours": tracking.dev_estimate_hours,
            "actual_dev_hours": tracking.actual_dev_hours,
            "qa_estimate_hours": tracking.qa_estimate_hours,
            "actual_qa_hours": tracking.actual_qa_hours,
            "dev_deviation": dev_deviation,
            "qa_deviation": qa_deviation,
            "qa_vs_dev_ratio": qa_vs_dev_ratio,
            "updated_on": tracking.updated_on.isoformat() if tracking.updated_on else None
        }
    finally:
        db.close()


@app.get("/tickets-dashboard/ticket/{ticket_id}/priority-history")
def get_ticket_priority_history(ticket_id: int):
    """Get priority change history for a ticket (for dashboard and reports)."""
    db: Session = SessionLocal()
    try:
        history = db.query(TicketPriorityHistory).filter(
            TicketPriorityHistory.ticket_id == ticket_id
        ).order_by(TicketPriorityHistory.changed_on.desc()).limit(100).all()
        return [
            {
                "previous_priority": h.previous_priority,
                "new_priority": h.new_priority,
                "changed_on": h.changed_on.isoformat() if h.changed_on else None,
                "source": h.source,
            }
            for h in history
        ]
    finally:
        db.close()


# ===== TICKETS DASHBOARD ENDPOINTS =====

# Status to Team Mapping (same as frontend)
STATUS_TEAM_MAPPING = {
    'NEW': 'BIS',
    'Ready For Development': 'DEV',
    'Quote Required': 'BIS',
    'Closed': 'Completed',
    'Backlog—Unranked': 'BIS',
    'Moved to Live': 'BIS',
    'Technical Review': 'DEV',
    'Approved for Live': 'DEV',
    'Live - awaiting fixes': 'DEV',
    'Express Lane Review': 'DEV',
    'In Progress': 'DEV',
    'Start Code Review': 'DEV',
    'Quote': 'BIS',
    'QC Testing': 'QA',
    'Under Review': 'BIS',
    'Code Review Failed': 'DEV',
    'QC Review Fail': 'DEV',
    'Pending Quote Approval': 'BIS',
    'BIS Testing': 'BIS - QA',
    'Planning': 'BIS',
    'Testing In Progress': 'BIS - QA',
    'Code Review Passed': 'DEV',
    'QC Testing in Progress': 'QA',
    'QC Testing Hold': 'QA',
    'Hold/Pending': 'BIS',
    'Design Review': 'BIS',
    'Ready for Design': 'BIS',
    'Design In Progress': 'BIS',
    'Tested - Awaiting Fixes': 'DEV',
    'Re-opened': 'DEV',
    'Reopened': 'DEV'
}

# Statuses that mean "with QA" (newly released to QA when ticket moves to one of these)
QA_TEAM_STATUSES = [
    'QC Testing',
    'QC Testing in Progress',
    'QC Testing Hold'
]

# Priority name -> display order (1 = highest, lower number = higher priority)
PRIORITY_ORDER = {
    'URGENT': 1,
    'High (Bugs)': 2,
    'High (Billable)': 3,
    'EPIC!': 4,
    'Medium (Bugs)': 5,
    'High Level 1': 6,
    'High Level 2': 7,
    'High Level 3': 8,
    'High Level 4': 9,
    'Medium': 10,
    'Low': 11,
    'Quote': 12,
    'Suggestion': 13,
}

@app.get("/tickets-dashboard/priority-order")
def get_priority_order():
    """Get priority names and their display order (1 = highest). Used for ticket priority charts."""
    # Return list of { priority, order } sorted by order ascending
    return [
        {"priority": name, "order": order}
        for name, order in sorted(PRIORITY_ORDER.items(), key=lambda x: x[1])
    ]

def _ticket_ageing(ticket, today: date, is_closed: bool) -> tuple:
    """Return (age_days, ageing_days, days_to_close) for a ticket. ageing_days = created->closed or created->today; days_to_close = closed - created (closed only)."""
    created_dt = getattr(ticket, 'created_on', None)
    closed_dt = getattr(ticket, 'closed_on', None)
    created_date = created_dt.date() if created_dt and hasattr(created_dt, 'date') else (created_dt if isinstance(created_dt, date) else None)
    closed_date = closed_dt.date() if closed_dt and hasattr(closed_dt, 'date') else (closed_dt if isinstance(closed_dt, date) else None)
    ageing_days = None
    days_to_close = None
    if created_date:
        if is_closed and closed_date:
            ageing_days = (closed_date - created_date).days
            days_to_close = ageing_days
        else:
            ageing_days = (today - created_date).days
    age_days = ageing_days if ageing_days is not None else 0
    if age_days == 0 and ticket.updated_on:
        age_delta = today - (ticket.updated_on.date() if hasattr(ticket.updated_on, 'date') else ticket.updated_on)
        age_days = age_delta.days
    return age_days, ageing_days, days_to_close


def _priority_changes_count_map(db: Session, ticket_ids: List[int]) -> Dict[int, int]:
    """Return dict of ticket_id -> count of priority history entries for each ticket."""
    if not ticket_ids:
        return {}
    rows = db.query(TicketPriorityHistory.ticket_id, func.count(TicketPriorityHistory.id)).filter(
        TicketPriorityHistory.ticket_id.in_(ticket_ids)
    ).group_by(TicketPriorityHistory.ticket_id).all()
    return {ticket_id: count for ticket_id, count in rows}


@app.get("/tickets-dashboard/overview")
def get_tickets_overview():
    """Get overall tickets dashboard data with team breakdown"""
    db: Session = SessionLocal()
    try:
        all_tickets = db.query(TicketTracking).all()
        
        if not all_tickets:
            return {
                "total_tickets": 0,
                "by_status": {},
                "by_team": {},
                "by_assignee": {},
                "team_status_breakdown": {},
                "eta_analysis": {
                    "overdue": 0,
                    "due_this_week": 0,
                    "no_eta": 0,
                    "on_track": 0
                }
            }
        
        today = datetime.now().date()
        week_from_now = today + timedelta(days=7)
        
        # Initialize counters
        by_status = defaultdict(int)
        by_team = defaultdict(int)
        by_priority = defaultdict(int)  # active tickets by priority
        by_assignee = defaultdict(list)
        team_status_breakdown = defaultdict(lambda: defaultdict(int))
        team_tickets = defaultdict(list)
        
        eta_overdue = 0
        eta_due_this_week = 0
        eta_no_eta = 0
        eta_on_track = 0
        
        completed_count = 0
        completed_tickets = []
        priority_counts = _priority_changes_count_map(db, [t.ticket_id for t in all_tickets])
        
        for ticket in all_tickets:
            status = ticket.status or 'Unknown'
            team = STATUS_TEAM_MAPPING.get(status, 'Unknown')
            assignee = ticket.current_assignee or 'Unassigned'
            
            # Check if completed
            is_closed = status.lower() in ['closed', 'moved to live', 'completed']
            ticket_age, ageing_days, days_to_close = _ticket_ageing(ticket, today, is_closed)
            
            ticket_data = {
                "ticket_id": ticket.ticket_id,
                "title": (getattr(ticket, 'title', None) or '').strip() or None,
                "status": status,
                "priority": getattr(ticket, 'priority', None) or None,
                "priority_changes_count": priority_counts.get(ticket.ticket_id, 0),
                "team": team,
                "assignee": assignee,
                "eta": ticket.eta.isoformat() if ticket.eta else None,
                "age_days": ticket_age,
                "ageing_days": ageing_days,
                "days_to_close": days_to_close,
                "created_on": ticket.created_on.isoformat() if getattr(ticket, 'created_on', None) else None,
                "closed_on": ticket.closed_on.isoformat() if getattr(ticket, 'closed_on', None) else None,
                "dev_estimate": ticket.dev_estimate_hours,
                "dev_actual": ticket.actual_dev_hours,
                "qa_estimate": ticket.qa_estimate_hours,
                "qa_actual": ticket.actual_qa_hours,
                "backend_developer": ticket.backend_developer,
                "frontend_developer": ticket.frontend_developer,
                "qc_tester": ticket.qc_tester,
                "updated_on": ticket.updated_on.isoformat() if ticket.updated_on else None
            }
            
            if is_closed:
                completed_count += 1
                completed_tickets.append(ticket_data)
                continue  # Skip completed tickets from active tracking
            
            # Count by status (only active tickets)
            by_status[status] += 1
            
            # Count by team (only active tickets)
            by_team[team] += 1
            
            # Count by priority (only active tickets)
            priority_label = (getattr(ticket, 'priority', None) or '').strip() or 'Unspecified'
            by_priority[priority_label] += 1
            
            # Track team status breakdown (only active tickets)
            team_status_breakdown[team][status] += 1
            
            # Track tickets by assignee (only active tickets)
            by_assignee[assignee].append(ticket_data)
            
            # Track tickets by team (only active tickets)
            team_tickets[team].append(ticket_data)
            
            # ETA analysis (only active tickets)
            if not ticket.eta:
                eta_no_eta += 1
            else:
                eta_date = ticket.eta.date() if hasattr(ticket.eta, 'date') else ticket.eta
                if eta_date < today:
                    eta_overdue += 1
                elif eta_date <= week_from_now:
                    eta_due_this_week += 1
                else:
                    eta_on_track += 1
        
        return {
            "total_tickets": len(all_tickets),
            "completed_count": completed_count,
            "completed_tickets": completed_tickets,
            "active_tickets": len(all_tickets) - completed_count,
            "by_status": dict(by_status),
            "by_team": dict(by_team),
            "by_priority": dict(by_priority),
            "by_assignee": {k: {"count": len(v), "tickets": v} for k, v in by_assignee.items()},
            "team_status_breakdown": {k: dict(v) for k, v in team_status_breakdown.items()},
            "team_tickets": {k: v for k, v in team_tickets.items()},
            "eta_analysis": {
                "overdue": eta_overdue,
                "due_this_week": eta_due_this_week,
                "no_eta": eta_no_eta,
                "on_track": eta_on_track
            }
        }
    finally:
        db.close()


@app.get("/tickets-dashboard/team/{team_name}")
def get_team_tickets(team_name: str):
    """Get detailed tickets for a specific team"""
    db: Session = SessionLocal()
    try:
        all_tickets = db.query(TicketTracking).all()
        
        team_tickets = []
        status_breakdown = defaultdict(int)
        assignee_breakdown = defaultdict(list)
        
        today = datetime.now().date()
        team_ticket_ids = []
        for ticket in all_tickets:
            status = ticket.status or 'Unknown'
            team = STATUS_TEAM_MAPPING.get(status, 'Unknown')
            team_normalized = team.lower().replace(' ', '-').replace('/', '-')
            team_name_normalized = team_name.lower().replace(' ', '-').replace('/', '-')
            if team_normalized == team_name_normalized:
                team_ticket_ids.append(ticket.ticket_id)
        priority_counts = _priority_changes_count_map(db, team_ticket_ids)
        
        for ticket in all_tickets:
            status = ticket.status or 'Unknown'
            team = STATUS_TEAM_MAPPING.get(status, 'Unknown')
            
            # Match team (case-insensitive, handle variations)
            team_normalized = team.lower().replace(' ', '-').replace('/', '-')
            team_name_normalized = team_name.lower().replace(' ', '-').replace('/', '-')
            
            if team_normalized == team_name_normalized:
                assignee = ticket.current_assignee or 'Unassigned'
                is_closed = status.lower() in ['closed', 'moved to live', 'completed']
                ticket_age, ageing_days, days_to_close = _ticket_ageing(ticket, today, is_closed)
                
                ticket_data = {
                    "ticket_id": ticket.ticket_id,
                    "title": (getattr(ticket, 'title', None) or '').strip() or None,
                    "status": status,
                    "priority": getattr(ticket, 'priority', None) or None,
                    "priority_changes_count": priority_counts.get(ticket.ticket_id, 0),
                    "assignee": assignee,
                    "eta": ticket.eta.isoformat() if ticket.eta else None,
                    "age_days": ticket_age,
                    "ageing_days": ageing_days,
                    "days_to_close": days_to_close,
                    "created_on": ticket.created_on.isoformat() if getattr(ticket, 'created_on', None) else None,
                    "closed_on": ticket.closed_on.isoformat() if getattr(ticket, 'closed_on', None) else None,
                    "dev_estimate": ticket.dev_estimate_hours,
                    "dev_actual": ticket.actual_dev_hours,
                    "qa_estimate": ticket.qa_estimate_hours,
                    "qa_actual": ticket.actual_qa_hours,
                    "backend_developer": ticket.backend_developer,
                    "frontend_developer": ticket.frontend_developer,
                    "qc_tester": ticket.qc_tester,
                    "updated_on": ticket.updated_on.isoformat() if ticket.updated_on else None
                }
                
                team_tickets.append(ticket_data)
                status_breakdown[status] += 1
                assignee_breakdown[assignee].append(ticket_data)
        
        return {
            "team": team_name,
            "total_tickets": len(team_tickets),
            "tickets": team_tickets,
            "status_breakdown": dict(status_breakdown),
            "assignee_breakdown": {k: {"count": len(v), "tickets": v} for k, v in assignee_breakdown.items()}
        }
    finally:
        db.close()


@app.get("/tickets-dashboard/assignee/{assignee_name}")
def get_assignee_tickets(assignee_name: str):
    """Get tickets assigned to a specific person"""
    db: Session = SessionLocal()
    try:
        # Handle 'Unassigned' case
        if assignee_name.lower() == 'unassigned':
            tickets = db.query(TicketTracking).filter(
                (TicketTracking.current_assignee == None) | (TicketTracking.current_assignee == '')
            ).all()
        else:
            tickets = db.query(TicketTracking).filter(
                TicketTracking.current_assignee.ilike(f"%{assignee_name}%")
            ).all()
        
        result = []
        status_breakdown = defaultdict(int)
        team_breakdown = defaultdict(int)
        
        today = datetime.now().date()
        ticket_ids = [t.ticket_id for t in tickets]
        priority_counts = _priority_changes_count_map(db, ticket_ids)
        
        for ticket in tickets:
            status = ticket.status or 'Unknown'
            team = STATUS_TEAM_MAPPING.get(status, 'Unknown')
            is_closed = status.lower() in ['closed', 'moved to live', 'completed']
            ticket_age, ageing_days, days_to_close = _ticket_ageing(ticket, today, is_closed)
            
            result.append({
                "ticket_id": ticket.ticket_id,
                "title": (getattr(ticket, 'title', None) or '').strip() or None,
                "status": status,
                "priority": getattr(ticket, 'priority', None) or None,
                "priority_changes_count": priority_counts.get(ticket.ticket_id, 0),
                "team": team,
                "eta": ticket.eta.isoformat() if ticket.eta else None,
                "age_days": ticket_age,
                "ageing_days": ageing_days,
                "days_to_close": days_to_close,
                "created_on": ticket.created_on.isoformat() if getattr(ticket, 'created_on', None) else None,
                "closed_on": ticket.closed_on.isoformat() if getattr(ticket, 'closed_on', None) else None,
                "dev_estimate": ticket.dev_estimate_hours,
                "dev_actual": ticket.actual_dev_hours,
                "qa_estimate": ticket.qa_estimate_hours,
                "qa_actual": ticket.actual_qa_hours,
                "backend_developer": ticket.backend_developer,
                "frontend_developer": ticket.frontend_developer,
                "qc_tester": ticket.qc_tester,
                "updated_on": ticket.updated_on.isoformat() if ticket.updated_on else None
            })
            
            status_breakdown[status] += 1
            team_breakdown[team] += 1
        
        return {
            "assignee": assignee_name,
            "total_tickets": len(result),
            "tickets": result,
            "status_breakdown": dict(status_breakdown),
            "team_breakdown": dict(team_breakdown)
        }
    finally:
        db.close()


@app.get("/tickets-dashboard/status/{status_name}")
def get_status_tickets(status_name: str):
    """Get all tickets with a specific status"""
    db: Session = SessionLocal()
    try:
        tickets = db.query(TicketTracking).filter(
            TicketTracking.status.ilike(f"%{status_name}%")
        ).all()
        
        result = []
        assignee_breakdown = defaultdict(int)
        
        for ticket in tickets:
            status = ticket.status or 'Unknown'
            team = STATUS_TEAM_MAPPING.get(status, 'Unknown')
            assignee = ticket.current_assignee or 'Unassigned'
            
            result.append({
                "ticket_id": ticket.ticket_id,
                "status": status,
                "team": team,
                "assignee": assignee,
                "eta": ticket.eta.isoformat() if ticket.eta else None,
                "dev_estimate": ticket.dev_estimate_hours,
                "dev_actual": ticket.actual_dev_hours
            })
            
            assignee_breakdown[assignee] += 1
        
        return {
            "status": status_name,
            "team": STATUS_TEAM_MAPPING.get(status_name, 'Unknown'),
            "total_tickets": len(result),
            "tickets": result,
            "assignee_breakdown": dict(assignee_breakdown)
        }
    finally:
        db.close()


@app.get("/tickets-dashboard/eta-alerts")
def get_eta_alerts():
    """Get tickets with ETA concerns (overdue, due soon, no ETA)"""
    db: Session = SessionLocal()
    try:
        all_tickets = db.query(TicketTracking).all()
        
        today = datetime.now().date()
        week_from_now = today + timedelta(days=7)
        
        overdue = []
        due_this_week = []
        no_eta = []
        active_ticket_ids = [
            t.ticket_id for t in all_tickets
            if (t.status or '').lower() not in ['closed', 'moved to live', 'completed']
        ]
        priority_counts = _priority_changes_count_map(db, active_ticket_ids)
        
        for ticket in all_tickets:
            status = ticket.status or 'Unknown'
            is_closed = status.lower() in ['closed', 'moved to live', 'completed']
            
            if is_closed:
                continue
            
            team = STATUS_TEAM_MAPPING.get(status, 'Unknown')
            
            ticket_age, ageing_days, days_to_close = _ticket_ageing(ticket, today, False)
            ticket_data = {
                "ticket_id": ticket.ticket_id,
                "title": (getattr(ticket, 'title', None) or '').strip() or None,
                "status": status,
                "priority": getattr(ticket, 'priority', None) or None,
                "priority_changes_count": priority_counts.get(ticket.ticket_id, 0),
                "team": team,
                "assignee": ticket.current_assignee or 'Unassigned',
                "eta": ticket.eta.isoformat() if ticket.eta else None,
                "age_days": ticket_age,
                "ageing_days": ageing_days,
                "days_to_close": days_to_close,
                "created_on": ticket.created_on.isoformat() if getattr(ticket, 'created_on', None) else None,
                "dev_estimate": ticket.dev_estimate_hours,
                "dev_actual": ticket.actual_dev_hours,
                "qa_estimate": ticket.qa_estimate_hours,
                "qa_actual": ticket.actual_qa_hours,
            }
            
            if not ticket.eta:
                no_eta.append(ticket_data)
            else:
                eta_date = ticket.eta.date() if hasattr(ticket.eta, 'date') else ticket.eta
                if eta_date < today:
                    days_overdue = (today - eta_date).days
                    ticket_data["days_overdue"] = days_overdue
                    overdue.append(ticket_data)
                elif eta_date <= week_from_now:
                    days_until = (eta_date - today).days
                    ticket_data["days_until_eta"] = days_until
                    due_this_week.append(ticket_data)
        
        # Sort by urgency
        overdue.sort(key=lambda x: x.get("days_overdue", 0), reverse=True)
        due_this_week.sort(key=lambda x: x.get("days_until_eta", 7))
        
        return {
            "overdue": overdue,
            "due_this_week": due_this_week,
            "no_eta": no_eta,
            "summary": {
                "overdue_count": len(overdue),
                "due_this_week_count": len(due_this_week),
                "no_eta_count": len(no_eta)
            }
        }
    finally:
        db.close()


@app.get("/tickets-dashboard/newly-released-to-qa")
def get_newly_released_to_qa(
    days: int = Query(7, ge=1, le=90, description="Number of days to look back for tickets released to QA")
):
    """Get list of tickets newly released to QA (moved to QC Testing / QC Testing in Progress / QC Testing Hold) in the given period. Returns estimates, ETA, developer(s), current status, module, QC tester."""
    db: Session = SessionLocal()
    try:
        now = datetime.now()
        start = now - timedelta(days=days)
        end = now
        qc_newly_history = db.query(TicketStatusHistory).filter(
            TicketStatusHistory.new_status.in_(QA_TEAM_STATUSES),
            TicketStatusHistory.changed_on >= start,
            TicketStatusHistory.changed_on <= end
        ).order_by(TicketStatusHistory.changed_on.desc()).all()
        if not qc_newly_history:
            return {"tickets": [], "period_days": days, "from": start.isoformat(), "to": end.isoformat()}
        ticket_ids = list({h.ticket_id for h in qc_newly_history})
        tickets = db.query(TicketTracking).filter(TicketTracking.ticket_id.in_(ticket_ids)).all()
        ticket_by_id = {t.ticket_id: t for t in tickets}
        # Module from first Bug per ticket (optional)
        module_by_ticket = {}
        for tid in ticket_ids:
            bug = db.query(Bug).filter(Bug.ticket_id == tid).first()
            if bug and bug.module:
                module_by_ticket[tid] = bug.module
        moved_on_by_id = {}
        for h in qc_newly_history:
            if h.ticket_id not in moved_on_by_id:
                moved_on_by_id[h.ticket_id] = h.changed_on
        today = now.date()
        result = []
        for ticket in tickets:
            developers = []
            if ticket.backend_developer:
                developers.append(ticket.backend_developer)
            if ticket.frontend_developer:
                developers.append(ticket.frontend_developer)
            developers = list(set(developers))
            is_closed = (ticket.status or "").lower() in ["closed", "moved to live", "completed"]
            _, ageing_days, days_to_close = _ticket_ageing(ticket, today, is_closed)
            result.append({
                "ticket_id": ticket.ticket_id,
                "title": (getattr(ticket, "title", None) or "").strip() or None,
                "priority": getattr(ticket, "priority", None) or None,
                "status": ticket.status or "Unknown",
                "eta": ticket.eta.isoformat() if ticket.eta else None,
                "dev_estimate_hours": ticket.dev_estimate_hours,
                "qa_estimate_hours": ticket.qa_estimate_hours,
                "developers": developers,
                "developers_str": ", ".join(developers) if developers else "Not Assigned",
                "module": module_by_ticket.get(ticket.ticket_id) or "N/A",
                "qc_tester": ticket.qc_tester or "Not Assigned",
                "moved_to_qc_on": moved_on_by_id.get(ticket.ticket_id).isoformat() if moved_on_by_id.get(ticket.ticket_id) else None,
                "ageing_days": ageing_days,
                "days_to_close": days_to_close,
            })
        # Sort by moved_to_qc_on descending (newest first)
        result.sort(key=lambda x: x.get("moved_to_qc_on") or "", reverse=True)
        return {
            "tickets": result,
            "period_days": days,
            "from": start.isoformat(),
            "to": end.isoformat(),
        }
    finally:
        db.close()


# Dev statuses for automation planning analysis
DEV_TEAM_STATUSES = [
    'In Progress', 'Technical Review', 'Start Code Review', 'Code Review Passed',
    'Code Review Failed', 'Approved for Live', 'Ready For Development', 
    'Express Lane Review', 'QC Review Fail', 'Tested - Awaiting Fixes', 'Re-opened'
]

@app.get("/tickets-dashboard/automation-planning")
def get_automation_planning_analysis():
    """
    Analyze tickets in Dev status to identify modules/areas with high volume,
    helping prioritize automation efforts for upcoming QA work.
    """
    db: Session = SessionLocal()
    try:
        # Get all tickets in Dev team statuses
        dev_tickets = db.query(TicketTracking).filter(
            TicketTracking.status.in_(DEV_TEAM_STATUSES)
        ).all()
        
        if not dev_tickets:
            return {
                "total_dev_tickets": 0,
                "by_subdepartment": [],
                "by_priority": [],
                "by_module_keyword": [],
                "tickets": [],
                "recommendations": []
            }
        
        # Group by subdepartment
        subdepartment_counts = defaultdict(lambda: {"count": 0, "tickets": [], "priorities": defaultdict(int)})
        priority_counts = defaultdict(int)
        module_keywords = defaultdict(lambda: {"count": 0, "tickets": []})
        
        # Common module keywords to extract from titles
        KEYWORD_PATTERNS = [
            'Dashboard', 'Report', 'API', 'Login', 'Auth', 'User', 'Admin',
            'Payment', 'Invoice', 'Order', 'Cart', 'Checkout', 'Product',
            'Notification', 'Email', 'SMS', 'Calendar', 'Schedule', 'Booking',
            'Search', 'Filter', 'Export', 'Import', 'Upload', 'Download',
            'Settings', 'Config', 'Profile', 'Account', 'Registration',
            'Integration', 'Sync', 'Webhook', 'Mobile', 'App', 'Web',
            'Database', 'Migration', 'Performance', 'Security', 'Bug', 'Fix'
        ]
        
        today = datetime.now().date()
        all_ticket_data = []
        
        for ticket in dev_tickets:
            subdept = (ticket.subdepartment or 'Unknown').strip() or 'Unknown'
            priority = (ticket.priority or 'Unknown').strip() or 'Unknown'
            title = (ticket.title or '').strip()
            status = ticket.status or 'Unknown'
            
            # Calculate age
            created_date = None
            if ticket.created_on:
                created_date = ticket.created_on.date() if hasattr(ticket.created_on, 'date') else ticket.created_on
            age_days = (today - created_date).days if created_date else None
            
            ticket_data = {
                "ticket_id": ticket.ticket_id,
                "title": title or f"Ticket #{ticket.ticket_id}",
                "status": status,
                "priority": priority,
                "subdepartment": subdept,
                "assignee": ticket.current_assignee or 'Unassigned',
                "backend_developer": ticket.backend_developer,
                "frontend_developer": ticket.frontend_developer,
                "eta": ticket.eta.isoformat() if ticket.eta else None,
                "age_days": age_days,
                "dev_estimate": ticket.dev_estimate_hours,
                "dev_actual": ticket.actual_dev_hours,
                "created_on": ticket.created_on.isoformat() if ticket.created_on else None
            }
            
            all_ticket_data.append(ticket_data)
            
            # Count by subdepartment
            subdepartment_counts[subdept]["count"] += 1
            subdepartment_counts[subdept]["tickets"].append(ticket_data)
            subdepartment_counts[subdept]["priorities"][priority] += 1
            
            # Count by priority
            priority_counts[priority] += 1
            
            # Extract module keywords from title
            if title:
                title_lower = title.lower()
                for keyword in KEYWORD_PATTERNS:
                    if keyword.lower() in title_lower:
                        module_keywords[keyword]["count"] += 1
                        module_keywords[keyword]["tickets"].append(ticket_data)
        
        # Format subdepartment data
        by_subdepartment = []
        for subdept, data in sorted(subdepartment_counts.items(), key=lambda x: x[1]["count"], reverse=True):
            by_subdepartment.append({
                "subdepartment": subdept,
                "count": data["count"],
                "percentage": round(data["count"] / len(dev_tickets) * 100, 1),
                "priorities": dict(data["priorities"]),
                "ticket_ids": [t["ticket_id"] for t in data["tickets"]]
            })
        
        # Format priority data
        by_priority = []
        for priority, count in sorted(priority_counts.items(), key=lambda x: x[1], reverse=True):
            by_priority.append({
                "priority": priority,
                "count": count,
                "percentage": round(count / len(dev_tickets) * 100, 1)
            })
        
        # Format module keywords (top 15)
        by_module_keyword = []
        for keyword, data in sorted(module_keywords.items(), key=lambda x: x[1]["count"], reverse=True)[:15]:
            by_module_keyword.append({
                "keyword": keyword,
                "count": data["count"],
                "percentage": round(data["count"] / len(dev_tickets) * 100, 1),
                "ticket_ids": [t["ticket_id"] for t in data["tickets"]]
            })
        
        # Generate recommendations
        recommendations = []
        if by_subdepartment:
            top_subdept = by_subdepartment[0]
            if top_subdept["count"] >= 5:
                recommendations.append({
                    "type": "high_volume_subdepartment",
                    "message": f"Focus automation on '{top_subdept['subdepartment']}' - {top_subdept['count']} tickets ({top_subdept['percentage']}%) coming to QA",
                    "priority": "high"
                })
        
        if by_module_keyword:
            top_keywords = [k for k in by_module_keyword[:3] if k["count"] >= 3]
            if top_keywords:
                keyword_list = ", ".join([k["keyword"] for k in top_keywords])
                recommendations.append({
                    "type": "common_modules",
                    "message": f"Common areas: {keyword_list} - consider creating reusable test automation for these",
                    "priority": "medium"
                })
        
        # Check for urgent/high priority concentration
        urgent_high = sum(c for p, c in priority_counts.items() if 'urgent' in p.lower() or 'high' in p.lower())
        if urgent_high > len(dev_tickets) * 0.3:
            recommendations.append({
                "type": "priority_alert",
                "message": f"{urgent_high} tickets ({round(urgent_high/len(dev_tickets)*100)}%) are Urgent/High priority - prioritize automation for quick turnaround",
                "priority": "high"
            })
        
        return {
            "total_dev_tickets": len(dev_tickets),
            "by_subdepartment": by_subdepartment,
            "by_priority": by_priority,
            "by_module_keyword": by_module_keyword,
            "tickets": sorted(all_ticket_data, key=lambda x: (x["priority"] or "ZZZ", -(x["age_days"] or 0))),
            "recommendations": recommendations,
            "dev_statuses_included": DEV_TEAM_STATUSES
        }
    finally:
        db.close()


@app.get("/tickets-dashboard/time-analysis")
def get_time_analysis(
    period: str = Query("last_week", description="Time period: last_week, last_2_weeks, last_month, custom"),
    start_date: Optional[str] = Query(None, description="Start date for custom period (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date for custom period (YYYY-MM-DD)")
):
    """Get time-based analysis of ticket activity by team"""
    db: Session = SessionLocal()
    try:
        today = datetime.now().date()
        
        # Determine date range
        if period == "last_week":
            range_start = today - timedelta(days=7)
            range_end = today
        elif period == "last_2_weeks":
            range_start = today - timedelta(days=14)
            range_end = today
        elif period == "last_month":
            range_start = today - timedelta(days=30)
            range_end = today
        elif period == "custom" and start_date and end_date:
            range_start = datetime.strptime(start_date, "%Y-%m-%d").date()
            range_end = datetime.strptime(end_date, "%Y-%m-%d").date()
        else:
            range_start = today - timedelta(days=7)
            range_end = today
        
        all_tickets = db.query(TicketTracking).all()
        
        print(f"Time Analysis: period={period}, range={range_start} to {range_end}, total_tickets={len(all_tickets)}")
        
        # Filter tickets by update date within period
        period_tickets = []
        for ticket in all_tickets:
            if ticket.updated_on:
                update_date = ticket.updated_on.date() if hasattr(ticket.updated_on, 'date') else ticket.updated_on
                if range_start <= update_date <= range_end:
                    period_tickets.append(ticket)
        
        print(f"Period tickets found: {len(period_tickets)}")
        
        # Team-centric analysis structure
        teams_data = {
            'BIS': {
                'name': 'BIS',
                'description': 'Business Intelligence & Strategy',
                'members': defaultdict(lambda: {'tickets': [], 'statuses': defaultdict(int)}),
                'total_tickets': 0,
                'status_breakdown': defaultdict(int),
                'transitions': defaultdict(int)  # e.g., how many moved to Dev
            },
            'DEV': {
                'name': 'DEV',
                'description': 'Development Team',
                'members': defaultdict(lambda: {'tickets': [], 'statuses': defaultdict(int)}),
                'total_tickets': 0,
                'status_breakdown': defaultdict(int),
                'transitions': defaultdict(int)
            },
            'QA': {
                'name': 'QA',
                'description': 'Quality Assurance',
                'members': defaultdict(lambda: {'tickets': [], 'statuses': defaultdict(int)}),
                'total_tickets': 0,
                'status_breakdown': defaultdict(int),
                'transitions': defaultdict(int),
                'moved_to_bis_testing': 0,  # Special metric for QA
                'moved_to_dev': 0  # Tickets sent back to dev
            },
            'BIS - QA': {
                'name': 'BIS - QA',
                'description': 'BIS Quality Testing',
                'members': defaultdict(lambda: {'tickets': [], 'statuses': defaultdict(int)}),
                'total_tickets': 0,
                'status_breakdown': defaultdict(int),
                'transitions': defaultdict(int)
            }
        }
        
        # Track closed tickets separately
        closed_tickets_count = 0
        active_tickets_count = 0
        
        # Track achievements for each team
        achievements = {
            'DEV': {
                'moved_to_qc_testing': 0,
                'label': 'Moved to QC Testing'
            },
            'QA': {
                'moved_to_bis_testing': 0,
                'moved_to_closed': 0,
                'label_bis': 'Moved to BIS Testing',
                'label_closed': 'Moved to Closed'
            },
            'BIS - QA': {
                'approved_for_live': 0,
                'label': 'Approved for Live'
            }
        }
        
        # Process tickets
        for ticket in period_tickets:
            status = ticket.status or 'Unknown'
            team = STATUS_TEAM_MAPPING.get(status, 'Unknown')
            
            # Check if ticket is closed/completed
            is_closed = status.lower() in ['closed', 'moved to live', 'completed'] or team == 'Completed'
            
            # Track achievements based on current status (these are milestones reached)
            # DEV achievement: tickets that moved to QC Testing
            if status in ['QC Testing', 'QC Testing in Progress', 'QC Testing Hold']:
                achievements['DEV']['moved_to_qc_testing'] += 1
            
            # QA achievement: tickets moved to BIS Testing
            if status == 'BIS Testing':
                achievements['QA']['moved_to_bis_testing'] += 1
            
            # QA achievement: tickets moved to Closed
            if status.lower() in ['closed', 'moved to live']:
                achievements['QA']['moved_to_closed'] += 1
            
            # BIS-QA achievement: tickets approved for live
            if status in ['Approved for Live', 'Moved to Live']:
                achievements['BIS - QA']['approved_for_live'] += 1
            
            if is_closed:
                closed_tickets_count += 1
                continue  # Skip closed tickets from team analysis
            
            active_tickets_count += 1
            
            if team not in teams_data:
                teams_data[team] = {
                    'name': team,
                    'description': team,
                    'members': defaultdict(lambda: {'tickets': [], 'statuses': defaultdict(int)}),
                    'total_tickets': 0,
                    'status_breakdown': defaultdict(int),
                    'transitions': defaultdict(int)
                }
            
            ticket_data = {
                "ticket_id": ticket.ticket_id,
                "status": status,
                "team": team,
                "assignee": ticket.current_assignee or 'Unassigned',
                "updated_on": ticket.updated_on.isoformat() if ticket.updated_on else None,
                "eta": ticket.eta.isoformat() if ticket.eta else None,
                "dev_estimate": ticket.dev_estimate_hours,
                "dev_actual": ticket.actual_dev_hours,
                "qa_estimate": ticket.qa_estimate_hours,
                "qa_actual": ticket.actual_qa_hours,
                "qc_tester": ticket.qc_tester,
                "backend_developer": ticket.backend_developer,
                "frontend_developer": ticket.frontend_developer
            }
            
            teams_data[team]['total_tickets'] += 1
            teams_data[team]['status_breakdown'][status] += 1
            
            # Get the right member based on team
            if team == 'QA':
                member = ticket.qc_tester or ticket.current_assignee or 'Unassigned'
                # Track QA-specific metrics
                if status == 'BIS Testing':
                    teams_data['QA']['moved_to_bis_testing'] += 1
                elif status in ['Code Review Failed', 'QC Review Fail', 'Tested - Awaiting Fixes']:
                    teams_data['QA']['moved_to_dev'] += 1
            elif team == 'DEV':
                member = ticket.backend_developer or ticket.frontend_developer or ticket.current_assignee or 'Unassigned'
            else:
                member = ticket.current_assignee or 'Unassigned'
            
            teams_data[team]['members'][member]['tickets'].append(ticket_data)
            teams_data[team]['members'][member]['statuses'][status] += 1
        
        # Convert to serializable format
        result_teams = {}
        for team_key, team_data in teams_data.items():
            if team_data['total_tickets'] > 0:  # Only include teams with activity
                members_list = []
                for member_name, member_data in team_data['members'].items():
                    members_list.append({
                        'name': member_name,
                        'ticket_count': len(member_data['tickets']),
                        'tickets': member_data['tickets'],
                        'status_breakdown': dict(member_data['statuses'])
                    })
                # Sort members by ticket count
                members_list.sort(key=lambda x: x['ticket_count'], reverse=True)
                
                result_teams[team_key] = {
                    'name': team_data['name'],
                    'description': team_data['description'],
                    'total_tickets': team_data['total_tickets'],
                    'status_breakdown': dict(team_data['status_breakdown']),
                    'members': members_list
                }
                
                # Add QA-specific metrics
                if team_key == 'QA':
                    result_teams[team_key]['moved_to_bis_testing'] = team_data.get('moved_to_bis_testing', 0)
                    result_teams[team_key]['moved_to_dev'] = team_data.get('moved_to_dev', 0)
        
        return {
            "period": {
                "type": period,
                "start_date": range_start.isoformat(),
                "end_date": range_end.isoformat(),
                "days": (range_end - range_start).days
            },
            "summary": {
                "total_tickets_worked": len(period_tickets),
                "active_tickets": active_tickets_count,
                "closed_tickets": closed_tickets_count,
                "teams_active": len(result_teams)
            },
            "achievements": {
                "DEV": {
                    "count": achievements['DEV']['moved_to_qc_testing'],
                    "label": "Moved to QC Testing",
                    "icon": "🧪"
                },
                "QA": {
                    "bis_testing": {
                        "count": achievements['QA']['moved_to_bis_testing'],
                        "label": "Moved to BIS Testing",
                        "icon": "🔍"
                    },
                    "closed": {
                        "count": achievements['QA']['moved_to_closed'],
                        "label": "Moved to Closed",
                        "icon": "✅"
                    }
                },
                "BIS_QA": {
                    "count": achievements['BIS - QA']['approved_for_live'],
                    "label": "Approved for Live",
                    "icon": "🚀"
                }
            },
            # Debug info
            "_debug": {
                "period": period,
                "range_start": range_start.isoformat(),
                "range_end": range_end.isoformat(),
                "period_tickets_count": len(period_tickets),
                "achievements_raw": {
                    "dev_to_qc": achievements['DEV']['moved_to_qc_testing'],
                    "qa_to_bis": achievements['QA']['moved_to_bis_testing'],
                    "qa_to_closed": achievements['QA']['moved_to_closed'],
                    "bis_qa_approved": achievements['BIS - QA']['approved_for_live']
                }
            },
            "teams": result_teams
        }
    finally:
        db.close()


@app.get("/tickets-dashboard/user-performance")
def get_user_performance(
    user: str = Query(..., description="User name to get performance for"),
    period: str = Query("last_month", description="Time period")
):
    """Get detailed performance metrics for a specific user"""
    db: Session = SessionLocal()
    try:
        today = datetime.now().date()
        
        # Determine date range
        if period == "last_week":
            range_start = today - timedelta(days=7)
        elif period == "last_2_weeks":
            range_start = today - timedelta(days=14)
        elif period == "last_month":
            range_start = today - timedelta(days=30)
        else:
            range_start = today - timedelta(days=30)
        
        all_tickets = db.query(TicketTracking).all()
        
        user_lower = user.lower()
        user_tickets = []
        
        for ticket in all_tickets:
            assignee = (ticket.current_assignee or '').lower()
            backend_dev = (ticket.backend_developer or '').lower()
            frontend_dev = (ticket.frontend_developer or '').lower()
            qc_tester = (ticket.qc_tester or '').lower()
            
            is_user_ticket = user_lower in [assignee, backend_dev, frontend_dev, qc_tester]
            
            if is_user_ticket:
                status = ticket.status or 'Unknown'
                team = STATUS_TEAM_MAPPING.get(status, 'Unknown')
                
                # Check if updated within period
                in_period = False
                if ticket.updated_on:
                    update_date = ticket.updated_on.date() if hasattr(ticket.updated_on, 'date') else ticket.updated_on
                    in_period = update_date >= range_start
                
                user_tickets.append({
                    "ticket_id": ticket.ticket_id,
                    "status": status,
                    "team": team,
                    "role": "Assignee" if assignee == user_lower else 
                            "Backend Dev" if backend_dev == user_lower else
                            "Frontend Dev" if frontend_dev == user_lower else
                            "QC Tester",
                    "updated_on": ticket.updated_on.isoformat() if ticket.updated_on else None,
                    "eta": ticket.eta.isoformat() if ticket.eta else None,
                    "in_period": in_period,
                    "dev_estimate": ticket.dev_estimate_hours,
                    "dev_actual": ticket.actual_dev_hours,
                    "qa_estimate": ticket.qa_estimate_hours,
                    "qa_actual": ticket.actual_qa_hours
                })
        
        # Calculate metrics
        total_tickets = len(user_tickets)
        period_tickets = [t for t in user_tickets if t.get("in_period")]
        completed = [t for t in user_tickets if t["status"].lower() in ['closed', 'moved to live', 'completed']]
        
        # Status breakdown
        status_breakdown = defaultdict(int)
        team_breakdown = defaultdict(int)
        role_breakdown = defaultdict(int)
        
        for ticket in user_tickets:
            status_breakdown[ticket["status"]] += 1
            team_breakdown[ticket["team"]] += 1
            role_breakdown[ticket["role"]] += 1
        
        return {
            "user": user,
            "period": period,
            "metrics": {
                "total_tickets_assigned": total_tickets,
                "tickets_worked_in_period": len(period_tickets),
                "completed_tickets": len(completed),
                "completion_rate": round((len(completed) / total_tickets * 100), 1) if total_tickets > 0 else 0
            },
            "breakdown": {
                "by_status": dict(status_breakdown),
                "by_team": dict(team_breakdown),
                "by_role": dict(role_breakdown)
            },
            "tickets": user_tickets
        }
    finally:
        db.close()


# ===== EMPLOYEE MANAGEMENT ENDPOINTS =====

def get_date_range(period: str):
    """Get date range for a given period"""
    today = datetime.now()
    if period == "past_week":
        return today - timedelta(days=7), today
    elif period == "past_month":
        return today - timedelta(days=30), today
    elif period == "past_quarter":
        return today - timedelta(days=90), today
    elif period == "one_year":
        return today - timedelta(days=365), today
    else:  # overall
        return None, today


def get_period_range(kind: str, offset: int = 0):
    """Calendar-aligned date range for a month/quarter, offset periods back from now.

    offset=0 is the current period, 1 the previous one, etc. Returns
    (start_datetime, end_datetime, label) where end is the last day at 23:59:59.
    """
    now = datetime.now()
    if kind == "quarter":
        # Index quarters globally so subtracting offset rolls across years cleanly.
        q_index = (now.year * 4 + (now.month - 1) // 3) - offset
        year, q = divmod(q_index, 4)
        start_month = q * 3 + 1
        end_month = start_month + 2
        start = datetime(year, start_month, 1)
        last_day = calendar.monthrange(year, end_month)[1]
        end = datetime(year, end_month, last_day, 23, 59, 59)
        label = f"Q{q + 1} {year}"
    else:  # month
        m_index = (now.year * 12 + (now.month - 1)) - offset
        year, m0 = divmod(m_index, 12)
        month = m0 + 1
        start = datetime(year, month, 1)
        last_day = calendar.monthrange(year, month)[1]
        end = datetime(year, month, last_day, 23, 59, 59)
        label = start.strftime("%B %Y")
    return start, end, label


def calculate_experience_years(date_of_joining):
    """Calculate years of experience from joining date"""
    if not date_of_joining:
        return 0
    today = datetime.now()
    delta = today - date_of_joining
    return round(delta.days / 365.25, 1)  # One decimal place

def calculate_bis_experience(bis_introduced_date):
    """Calculate BIS experience from BIS introduced date"""
    if not bis_introduced_date:
        return None
    today = datetime.now()
    delta = today - bis_introduced_date
    return round(delta.days / 365.25, 1)  # One decimal place

def calculate_total_experience(date_of_joining, previous_experience):
    """Calculate total experience (Techversant + previous)"""
    techversant_exp = calculate_experience_years(date_of_joining)
    prev_exp = previous_experience or 0
    return round(techversant_exp + prev_exp, 1)  # One decimal place


def _resolve_lead_or_manager_name(db: Session, value: str) -> Optional[str]:
    """Resolve lead/manager string to canonical Employee.name (case-insensitive match)."""
    if not value or not str(value).strip():
        return None
    v = str(value).strip()
    emp = db.query(Employee).filter(func.lower(Employee.name) == func.lower(v)).first()
    return emp.name if emp else v


@app.get("/employees")
def list_employees(
    team: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    lead: Optional[str] = Query(None),
    is_active: Optional[bool] = Query(True),
    search: Optional[str] = Query(None),
    employment_status: Optional[str] = Query(None),
    archived: Optional[bool] = Query(False, description="If true, return only archived (resigned) employees"),
    serving_notice: Optional[bool] = Query(None, description="If true, return only employees serving notice period"),
    for_display: Optional[bool] = Query(False, description="If true, return all employees for name lookup in dashboards (Tickets, Bugs, etc.) - no role filtering"),
    for_lead_dropdown: Optional[bool] = Query(False, description="If true, return only employees eligible to be assigned as Lead (LEAD/MANAGER/ADMIN role or job title containing Lead)"),
    for_manager_dropdown: Optional[bool] = Query(False, description="If true, return only employees eligible to be assigned as Manager (MANAGER/ADMIN role or job title containing Manager)"),
    current_user: dict = Depends(get_current_user),
):
    """List all employees with optional filters. Filtered by role visibility unless for_display=True."""
    db: Session = SessionLocal()
    try:
        query = db.query(Employee)

        # Filter by archived status
        if archived is not None:
            query = query.filter(Employee.archived == archived)
        
        if is_active is not None:
            query = query.filter(Employee.is_active == is_active)
        if team:
            query = query.filter(Employee.team.ilike(f"%{team}%"))
        if category:
            query = query.filter(Employee.category.ilike(f"%{category}%"))
        if lead:
            query = query.filter(Employee.lead.ilike(f"%{lead}%"))
        if employment_status:
            query = query.filter(Employee.employment_status == employment_status)
        if serving_notice:
            query = query.filter(Employee.employment_status == "Serving Notice Period")
        if search:
            query = query.filter(
                or_(
                    Employee.name.ilike(f"%{search}%"),
                    Employee.employee_id.ilike(f"%{search}%"),
                    Employee.email.ilike(f"%{search}%")
                )
            )

        if for_lead_dropdown or for_manager_dropdown:
            sub = db.query(User.employee_id).filter(User.employee_id.isnot(None))
            if for_lead_dropdown and not for_manager_dropdown:
                sub = sub.filter(or_(
                    User.role.ilike("%LEAD%"),
                    User.role.ilike("%MANAGER%"),
                    User.role == "ADMIN",
                ))
            elif for_manager_dropdown:
                sub = sub.filter(or_(
                    User.role.ilike("%MANAGER%"),
                    User.role == "ADMIN",
                ))
            role_eligible_ids = [r[0] for r in sub.distinct().all()]
            if for_lead_dropdown and not for_manager_dropdown:
                job_title_cond = or_(
                    func.upper(Employee.role).like("%LEAD%"),
                    func.upper(Employee.role).like("%MANAGER%"),
                )
            elif for_manager_dropdown:
                job_title_cond = func.upper(Employee.role).like("%MANAGER%")
            else:
                job_title_cond = None
            if job_title_cond is not None:
                if role_eligible_ids:
                    query = query.filter(or_(Employee.employee_id.in_(role_eligible_ids), job_title_cond))
                else:
                    query = query.filter(job_title_cond)

        # For dashboard/ticket display (Tickets Dashboard, All Bugs, main Dashboard), return all employees
        # so developer/QA names can be resolved for display and linking - all users see all names
        if not for_display and not for_lead_dropdown and not for_manager_dropdown:
            visible = get_visible_employee_ids(db, current_user)
            if visible is not None:
                query = query.filter(Employee.employee_id.in_(visible))
        
        employees = query.order_by(Employee.name).all()
        
        result = []
        for emp in employees:
            # Calculate notice period days based on category
            notice_period_days = 90 if emp.category and emp.category.upper() == "BILLED" else 30
            
            # Calculate expected LWD if serving notice
            expected_lwd = None
            if emp.employment_status == "Serving Notice Period" and emp.resignation_date:
                expected_lwd = emp.expected_lwd or (emp.resignation_date + timedelta(days=notice_period_days))
            
            result.append({
                "id": emp.id,
                "employee_id": emp.employee_id,
                "name": emp.name,
                "email": emp.email,
                "role": emp.role,
                "designation": emp.designation,
                "location": emp.location,
                "mode_of_work": emp.mode_of_work or "Onsite",
                "date_of_joining": emp.date_of_joining.isoformat() if emp.date_of_joining else None,
                "team": emp.team,
                "category": emp.category,
                "employment_status": emp.employment_status or "Ongoing Employee",
                "lead": emp.lead,
                "manager": emp.manager,
                "photo_url": emp.photo_url,
                "experience_years": calculate_experience_years(emp.date_of_joining),
                "is_active": emp.is_active,
                "archived": emp.archived or False,
                "archived_on": emp.archived_on.isoformat() if emp.archived_on else None,
                "resignation_date": emp.resignation_date.isoformat() if emp.resignation_date else None,
                "expected_lwd": expected_lwd.isoformat() if expected_lwd else (emp.expected_lwd.isoformat() if emp.expected_lwd else None),
                "notice_period_days": notice_period_days if emp.employment_status == "Serving Notice Period" else None,
            })
        
        return result
    finally:
        db.close()


@app.get("/employees/filter-options")
def get_employee_filter_options(
    employment_status: Optional[str] = Query(None, description="Ongoing Employee or Resigned to match list view"),
    current_user: dict = Depends(get_current_user),
):
    """Return distinct teams, categories, and leads for filter dropdowns (employees list and calendar)."""
    db: Session = SessionLocal()
    try:
        query = db.query(Employee)
        if employment_status:
            query = query.filter(func.upper(Employee.employment_status) == employment_status.upper().strip())
        visible = get_visible_employee_ids(db, current_user)
        if visible is not None:
            query = query.filter(Employee.employee_id.in_(visible))
        employees = query.all()
        teams = sorted(set((e.team or "").strip() for e in employees if (e.team or "").strip()))
        categories = sorted(set((e.category or "").strip() for e in employees if (e.category or "").strip()))
        leads = sorted(set((e.lead or "").strip() for e in employees if (e.lead or "").strip()))
        return {"teams": teams, "categories": categories, "leads": leads}
    finally:
        db.close()


@app.get("/employees/export-all")
def export_all_employees(
    team: Optional[str] = Query(None, description="Filter by team"),
    category: Optional[str] = Query(None, description="Filter by category"),
    lead: Optional[str] = Query(None, description="Filter by lead"),
    search: Optional[str] = Query(None, description="Search by name, ID, or email"),
    employment_status: Optional[str] = Query(None, description="Filter by employment status"),
    current_user: dict = Depends(get_current_user),
):
    """Export all employees with basic profile details to Excel format. Filtered by role visibility."""
    db: Session = SessionLocal()
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        
        query = db.query(Employee)
        if team:
            query = query.filter(Employee.team.ilike(f"%{team}%"))
        if category:
            query = query.filter(Employee.category.ilike(f"%{category}%"))
        if lead:
            query = query.filter(Employee.lead.ilike(f"%{lead}%"))
        if search:
            query = query.filter(
                or_(
                    Employee.name.ilike(f"%{search}%"),
                    Employee.employee_id.ilike(f"%{search}%"),
                    Employee.email.ilike(f"%{search}%")
                )
            )
        if employment_status:
            query = query.filter(func.upper(Employee.employment_status) == employment_status.upper())
        visible = get_visible_employee_ids(db, current_user)
        if visible is not None:
            query = query.filter(Employee.employee_id.in_(visible))
        employees = query.order_by(Employee.name).all()
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Employee Profiles"
        
        # Define header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Base headers - fixed profile columns
        base_headers = [
            "Employee ID",
            "Name",
            "Email",
            "Role",
            "Designation",
            "Location",
            "Mode of Work",
            "Date of Joining",
            "Team",
            "Category",
            "Employment Status",
            "Resignation Date",
            "Expected LWD",
            "Reporting To (Lead)",
            "Reporting Manager",
            "Previous Experience",
            "BIS Introduced Date",
            "Experience (Years)",
            "Active Status",
            "System Role"
        ]
        
        # Build a map of employee_id -> system role from User table
        all_users = db.query(User).all()
        user_role_map = {u.employee_id: u.role for u in all_users if u.employee_id}
        
        # Collect all unique dynamic column names from all employees' mapping_data
        dynamic_columns = set()
        for emp in employees:
            if emp.mapping_data:
                for key in emp.mapping_data.keys():
                    dynamic_columns.add(key)
        
        # Sort dynamic columns for consistent ordering
        # Put standard columns first (Column 1-5, Notes), then any custom columns alphabetically
        standard_dynamic = ["Column 1", "Column 2", "Column 3", "Column 4", "Column 5", "Notes"]
        custom_columns = sorted([c for c in dynamic_columns if c not in standard_dynamic])
        
        # Build ordered list of dynamic columns
        ordered_dynamic_columns = []
        for col in standard_dynamic:
            if col in dynamic_columns:
                ordered_dynamic_columns.append(col)
        ordered_dynamic_columns.extend(custom_columns)
        
        # If no dynamic columns exist, add default empty columns for user to fill
        if not ordered_dynamic_columns:
            ordered_dynamic_columns = ["Column 1", "Column 2", "Column 3", "Notes"]
        
        # Combine headers
        headers = base_headers + ordered_dynamic_columns
        
        ws.append(headers)
        
        # Style header row
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_style
        
        # Number of base columns (for styling)
        num_base_cols = len(base_headers)
        
        # Add employee data
        for emp in employees:
            # Get existing mapping data if available
            mapping = emp.mapping_data or {}
            
            # Get system role from User table
            system_role = user_role_map.get(emp.employee_id, "EMPLOYEE")
            
            # Build base row data
            row = [
                emp.employee_id or "",
                emp.name or "",
                emp.email or "",
                emp.role or "",
                emp.designation or "",
                emp.location or "",
                emp.mode_of_work or "Onsite",
                emp.date_of_joining.strftime("%d-%b-%Y") if emp.date_of_joining else "",
                emp.team or "",
                emp.category or "",
                emp.employment_status or "Ongoing Employee",
                emp.resignation_date.strftime("%d-%b-%Y") if emp.resignation_date else "",
                emp.expected_lwd.strftime("%d-%b-%Y") if emp.expected_lwd else "",
                emp.lead or "",
                emp.manager or "",
                round(emp.previous_experience, 1) if emp.previous_experience is not None else "",
                emp.bis_introduced_date.strftime("%d-%b-%Y") if emp.bis_introduced_date else "",
                calculate_experience_years(emp.date_of_joining),
                "Active" if emp.is_active else "Inactive",
                system_role
            ]
            
            # Add dynamic column values
            for col_name in ordered_dynamic_columns:
                row.append(mapping.get(col_name, "") or "")
            
            ws.append(row)
            
            # Style data row
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=ws.max_row, column=col_idx)
                cell.border = border_style
                if col_idx == 6:  # Date column
                    cell.alignment = Alignment(horizontal='left')
                elif col_idx > num_base_cols:  # Dynamic/mapping columns
                    cell.fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")  # Light yellow
        
        # Adjust column widths - base columns
        base_column_widths = {
            'A': 15,  # Employee ID
            'B': 30,  # Name
            'C': 30,  # Email
            'D': 25,  # Role
            'E': 15,  # Location
            'F': 18,  # Date of Joining
            'G': 15,  # Team
            'H': 15,  # Category
            'I': 20,  # Employment Status
            'J': 25,  # Reporting To (Lead)
            'K': 25,  # Reporting Manager
            'L': 18,  # Previous Experience
            'M': 15,  # Experience (Years)
            'N': 15,  # Active Status
            'O': 18   # System Role
        }
        
        for col, width in base_column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Set width for dynamic columns (starting from column N onwards)
        from openpyxl.utils import get_column_letter
        for i, col_name in enumerate(ordered_dynamic_columns):
            col_letter = get_column_letter(num_base_cols + 1 + i)
            # Notes column gets extra width
            ws.column_dimensions[col_letter].width = 30 if col_name == "Notes" else 20
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Add filter to header row
        ws.auto_filter.ref = ws.dimensions
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        filter_parts = []
        if team:
            filter_parts.append(f"Team_{team}")
        if category:
            filter_parts.append(f"Category_{category}")
        if employment_status:
            filter_parts.append(f"Status_{employment_status}")
        
        filter_str = "_".join(filter_parts) if filter_parts else "All"
        filename = f"Employee_Profiles_Export_{filter_str}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error exporting employees: {str(e)}")
    finally:
        db.close()


@app.post("/employees/import-mapping")
def import_employee_mapping_data(
    file_path: Optional[str] = Query(None, description="Path to Excel file. If not provided, will look for latest in Downloads folder")
):
    """Import employee mapping data from Excel file (Column 1-5, Notes)"""
    db: Session = SessionLocal()
    try:
        import openpyxl
        from pathlib import Path
        import sys
        
        # Get Downloads folder path
        def get_downloads_folder():
            """Get the user's Downloads folder path"""
            if sys.platform == 'win32':
                import winreg
                try:
                    with winreg.OpenKey(winreg.HKEY_CURRENT_USER, 
                        r'SOFTWARE\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders') as key:
                        downloads = winreg.QueryValueEx(key, '{374DE290-123F-4565-9164-39C4925E467B}')[0]
                        return downloads
                except:
                    pass
            return os.path.join(Path.home(), 'Downloads')
        
        DOWNLOADS_FOLDER = get_downloads_folder()
        
        # Determine file path
        excel_path = file_path
        if not excel_path:
            # Find the most recent Employee_Profiles_Export_*.xlsx file in Downloads
            if not os.path.exists(DOWNLOADS_FOLDER):
                raise HTTPException(status_code=404, detail=f"Downloads folder not found: {DOWNLOADS_FOLDER}")
            
            import glob
            pattern = os.path.join(DOWNLOADS_FOLDER, "Employee_Profiles_Export_*.xlsx")
            matching_files = glob.glob(pattern)
            
            if not matching_files:
                raise HTTPException(
                    status_code=404, 
                    detail=f"No Employee_Profiles_Export_*.xlsx files found in {DOWNLOADS_FOLDER}"
                )
            
            # Sort by modification time (newest first)
            matching_files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
            excel_path = matching_files[0]
        
        if not os.path.exists(excel_path):
            raise HTTPException(status_code=404, detail=f"File not found: {excel_path}")
        
        # Load workbook
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        # Read headers to find column indices
        headers = [cell.value for cell in ws[1]]
        
        # Define base profile columns that should NOT be treated as mapping data
        base_profile_columns = {
            "employee id", "name", "email", "role", "location",
            "date of joining", "team", "category", "employment status",
            "reporting to (lead)", "reporting to", "lead", "reporting manager", "manager",
            "experience (years)", "experience", "active status", "active", "status",
            "user role", "user_role", "password", "login password",
            "system role", "system_role", "access role", "access_role", "previous experience", "previous_experience",
            "designation", "mode of work", "mode_of_work", "resignation date", "resignation_date",
            "expected lwd", "expected_lwd", "bis introduced date", "bis_introduced_date", "bis date"
        }
        
        # Find column indices
        col_indices = {}
        dynamic_columns = {}  # Store dynamic column name -> index mapping
        
        for idx, header in enumerate(headers, 1):
            header_str = str(header).strip() if header else ""
            header_lower = header_str.lower()
            
            if header_str == "Employee ID":
                col_indices["employee_id"] = idx
            elif header_lower in ["previous experience", "previous_experience", "prev experience", "prev exp"]:
                col_indices["previous_experience"] = idx
            elif header_lower in ["reporting to (lead)", "reporting to", "lead", "reporting lead"]:
                col_indices["lead"] = idx
            elif header_lower in ["reporting manager", "manager", "reporting to (manager)"]:
                col_indices["manager"] = idx
            elif header_lower in ["user role", "user_role", "access role", "system role", "system_role"]:
                col_indices["system_role"] = idx
            elif header_lower in ["password", "login password"]:
                col_indices["password"] = idx
            elif header_lower in ["designation", "job title", "title"]:
                col_indices["designation"] = idx
            elif header_lower in ["mode of work", "mode_of_work", "work mode"]:
                col_indices["mode_of_work"] = idx
            elif header_lower in ["resignation date", "resignation_date", "resigned date"]:
                col_indices["resignation_date"] = idx
            elif header_lower in ["employment status", "employment_status", "emp status"]:
                col_indices["employment_status"] = idx
            elif header_lower in ["bis introduced date", "bis_introduced_date", "bis date", "billing date"]:
                col_indices["bis_introduced_date"] = idx
            elif header_str and header_lower not in base_profile_columns:
                # This is a dynamic/mapping column - store with original name
                dynamic_columns[header_str] = idx
        
        if "employee_id" not in col_indices:
            raise HTTPException(status_code=400, detail="Employee ID column not found in Excel file")
        
        # Process rows
        updated_count = 0
        not_found = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            # Get employee ID
            emp_id_cell = row[col_indices["employee_id"] - 1]
            employee_id = str(emp_id_cell.value).strip() if emp_id_cell.value else None
            
            if not employee_id or employee_id.lower() in ['none', 'null', '']:
                continue
            
            # Find employee in database
            employee = db.query(Employee).filter(Employee.employee_id == employee_id).first()
            
            if not employee:
                not_found.append(employee_id)
                continue
            
            # Extract mapping data from all dynamic columns
            mapping_data = {}
            for col_name, col_idx in dynamic_columns.items():
                val = row[col_idx - 1].value
                if val is not None and str(val).strip():
                    mapping_data[col_name] = str(val).strip()
            
            # Update previous_experience if column exists
            if "previous_experience" in col_indices:
                val = row[col_indices["previous_experience"] - 1].value
                if val is not None:
                    try:
                        # Try to convert to float
                        prev_exp = float(val)
                        employee.previous_experience = prev_exp
                    except (ValueError, TypeError):
                        # If conversion fails, skip this value
                        pass
            
            # Update lead if column exists
            if "lead" in col_indices:
                val = row[col_indices["lead"] - 1].value
                if val is not None:
                    lead_value = str(val).strip()
                    if lead_value:
                        employee.lead = lead_value
                    else:
                        employee.lead = None
            
            # Update manager if column exists
            if "manager" in col_indices:
                val = row[col_indices["manager"] - 1].value
                if val is not None:
                    manager_value = str(val).strip()
                    if manager_value:
                        employee.manager = manager_value
                    else:
                        employee.manager = None
            
            # Update designation if column exists
            if "designation" in col_indices:
                val = row[col_indices["designation"] - 1].value
                if val is not None:
                    designation_value = str(val).strip()
                    if designation_value:
                        employee.designation = designation_value
            
            # Update mode_of_work if column exists
            if "mode_of_work" in col_indices:
                val = row[col_indices["mode_of_work"] - 1].value
                if val is not None:
                    mode_value = str(val).strip()
                    # Validate mode value
                    valid_modes = ["Onsite", "Remote", "Hybrid"]
                    # Try to match case-insensitively
                    for valid_mode in valid_modes:
                        if mode_value.lower() == valid_mode.lower():
                            employee.mode_of_work = valid_mode
                            break
            
            # Update employment_status if column exists
            if "employment_status" in col_indices:
                val = row[col_indices["employment_status"] - 1].value
                if val is not None:
                    status_value = str(val).strip()
                    valid_statuses = ["Ongoing Employee", "Serving Notice Period", "Resigned"]
                    # Try to match case-insensitively
                    for valid_status in valid_statuses:
                        if status_value.lower() == valid_status.lower():
                            employee.employment_status = valid_status
                            # Auto-archive if resigned
                            if valid_status == "Resigned" and not employee.archived:
                                employee.archived = True
                                employee.archived_on = datetime.utcnow()
                            break
            
            # Update resignation_date if column exists
            if "resignation_date" in col_indices:
                val = row[col_indices["resignation_date"] - 1].value
                if val is not None:
                    try:
                        if isinstance(val, datetime):
                            employee.resignation_date = val
                        elif isinstance(val, str) and val.strip():
                            # Try common date formats
                            for fmt in ["%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y"]:
                                try:
                                    employee.resignation_date = datetime.strptime(val.strip(), fmt)
                                    break
                                except ValueError:
                                    continue
                            # Auto-calculate expected LWD if not set
                            if employee.resignation_date and not employee.expected_lwd:
                                notice_days = 90 if employee.category and employee.category.upper() == "BILLED" else 30
                                employee.expected_lwd = employee.resignation_date + timedelta(days=notice_days)
                                employee.employment_status = "Serving Notice Period"
                    except Exception:
                        pass
            
            # Update bis_introduced_date if column exists
            if "bis_introduced_date" in col_indices:
                val = row[col_indices["bis_introduced_date"] - 1].value
                if val is not None:
                    try:
                        if isinstance(val, datetime):
                            employee.bis_introduced_date = val
                        elif isinstance(val, str) and val.strip():
                            # Try common date formats
                            for fmt in ["%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y"]:
                                try:
                                    employee.bis_introduced_date = datetime.strptime(val.strip(), fmt)
                                    break
                                except ValueError:
                                    continue
                    except Exception:
                        pass

            # Update system role if column exists
            if "system_role" in col_indices:
                val = row[col_indices["system_role"] - 1].value
                if val is not None:
                    role_value = str(val).strip().upper()
                    # Validate role value
                    valid_roles = ["ADMIN", "MANAGER_DEV", "MANAGER_QA", "LEAD_DEV", "LEAD_QA", "EMPLOYEE", "CLIENT"]
                    if role_value in valid_roles:
                        # Find or create User record for this employee
                        user_record = db.query(User).filter(User.employee_id == employee_id).first()
                        if user_record:
                            user_record.role = role_value
                        else:
                            # Create new user with default password (employee_id)
                            new_user = User(
                                email=employee.email,
                                employee_id=employee_id,
                                password_hash=hash_password(employee_id),
                                role=role_value,
                                password_changed_at=None
                            )
                            db.add(new_user)
            
            # Update employee mapping_data (set to None if empty dict to clear old data)
            employee.mapping_data = mapping_data if mapping_data else None
            updated_count += 1
        
        db.commit()
        
        return {
            "success": True,
            "message": f"Successfully imported mapping data for {updated_count} employees",
            "updated_count": updated_count,
            "not_found": not_found,
            "file_used": os.path.basename(excel_path)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error importing mapping data: {str(e)}")
    finally:
        db.close()


@app.get("/employees/team-overview")
def get_team_overview(current_user: dict = Depends(get_current_user)):
    """Get team-level summary for PM dashboard. Filtered by role visibility."""
    db: Session = SessionLocal()
    try:
        emp_query = db.query(Employee).filter(Employee.is_active == True)
        visible = get_visible_employee_ids(db, current_user)
        if visible is not None:
            emp_query = emp_query.filter(Employee.employee_id.in_(visible))
        employees = emp_query.all()
        
        team_stats = {
            "DEVELOPMENT": {"total": 0, "billed": 0, "unbilled": 0},
            "QA": {"total": 0, "billed": 0, "unbilled": 0}
        }
        
        leads = defaultdict(lambda: {"total": 0, "dev": 0, "qa": 0})
        
        for emp in employees:
            team = emp.team or "Unknown"
            if team not in team_stats:
                team_stats[team] = {"total": 0, "billed": 0, "unbilled": 0}
            
            team_stats[team]["total"] += 1
            if emp.category and "BILLED" in emp.category.upper():
                if "UN" in emp.category.upper():
                    team_stats[team]["unbilled"] += 1
                else:
                    team_stats[team]["billed"] += 1
            
            if emp.lead:
                leads[emp.lead]["total"] += 1
                if team == "DEVELOPMENT":
                    leads[emp.lead]["dev"] += 1
                elif team == "QA":
                    leads[emp.lead]["qa"] += 1
        
        return {
            "total_employees": len(employees),
            "team_breakdown": team_stats,
            "leads": dict(leads)
        }
    finally:
        db.close()


@app.get("/employees/{employee_id}")
def get_employee(employee_id: str, current_user: dict = Depends(get_current_user)):
    """Get single employee details. Requires profile access permission."""
    db: Session = SessionLocal()
    try:
        employee = db.query(Employee).filter(
            or_(
                Employee.employee_id == employee_id,
                Employee.id == int(employee_id) if employee_id.isdigit() else False
            )
        ).first()
        
        if not employee:
            raise HTTPException(status_code=404, detail="Employee not found")
        if not can_access_employee_profile(db, current_user, employee.employee_id):
            raise HTTPException(status_code=403, detail="Access denied to this employee profile")

        # Calculate experience metrics
        techversant_exp = calculate_experience_years(employee.date_of_joining)
        bis_exp = calculate_bis_experience(employee.bis_introduced_date) if employee.category == "BILLED" else None
        total_exp = calculate_total_experience(employee.date_of_joining, employee.previous_experience)
        
        # Determine if employee is a fresher (no previous experience)
        is_fresher = (employee.previous_experience or 0) == 0
        
        # Calculate notice period based on category
        notice_period_days = 90 if employee.category and employee.category.upper() == "BILLED" else 30
        
        # Calculate expected LWD if serving notice
        expected_lwd = None
        days_remaining = None
        if employee.employment_status == "Serving Notice Period" and employee.resignation_date:
            expected_lwd = employee.expected_lwd or (employee.resignation_date + timedelta(days=notice_period_days))
            days_remaining = (expected_lwd - datetime.utcnow()).days if expected_lwd > datetime.utcnow() else 0
        
        # Get user's access_role if they have a User account
        user = db.query(User).filter(User.employee_id == employee.employee_id).first()
        access_role = user.role if user else None
        
        # Check edit permissions for the current user
        can_edit = can_edit_employee_profile(db, current_user, employee.employee_id)
        can_tasks = can_manage_tasks_for(db, current_user, employee.employee_id)
        
        return {
            "id": employee.id,
            "employee_id": employee.employee_id,
            "name": employee.name,
            "email": employee.email,
            "role": employee.role,  # Job title
            "designation": employee.designation,  # Designation (can be different from role)
            "access_role": access_role,  # Access role (MANAGER_QA, LEAD_DEV, EMPLOYEE, etc.)
            "location": employee.location,
            "mode_of_work": employee.mode_of_work or "Onsite",
            "date_of_joining": employee.date_of_joining.isoformat() if employee.date_of_joining else None,
            "team": employee.team,
            "category": employee.category,
            "employment_status": employee.employment_status or "Ongoing Employee",
            "lead": employee.lead,
            "manager": employee.manager,
            "previous_experience": round(float(employee.previous_experience), 1) if employee.previous_experience is not None else None,
            "is_fresher": is_fresher,
            "bis_introduced_date": employee.bis_introduced_date.isoformat() if employee.bis_introduced_date else None,
            "techversant_experience": techversant_exp,
            "bis_experience": bis_exp,
            "total_experience": total_exp,
            "bis_status": "Un-Billed" if employee.category != "BILLED" else "Billed",
            "platform": employee.platform,
            "photo_url": employee.photo_url,
            "experience_years": techversant_exp,  # Keep for backward compatibility
            "is_active": employee.is_active,
            "mapping_data": employee.mapping_data or {},
            # Notice period and resignation tracking
            "resignation_date": employee.resignation_date.isoformat() if employee.resignation_date else None,
            "expected_lwd": expected_lwd.isoformat() if expected_lwd else (employee.expected_lwd.isoformat() if employee.expected_lwd else None),
            "notice_period_days": notice_period_days if employee.employment_status == "Serving Notice Period" else None,
            "days_remaining": days_remaining,
            # Archive status
            "archived": employee.archived or False,
            "archived_on": employee.archived_on.isoformat() if employee.archived_on else None,
            # Timestamps
            "created_on": employee.created_on.isoformat() if employee.created_on else None,
            "updated_on": employee.updated_on.isoformat() if employee.updated_on else None,
            "_permissions": {
                "can_edit": can_edit,
                "can_manage_tasks": can_tasks,
            }
        }
    finally:
        db.close()


@app.get("/employees/{employee_id}/export")
def export_employee_profile(employee_id: str):
    """Export employee profile data to Excel format"""
    db: Session = SessionLocal()
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        
        # Find employee
        employee = db.query(Employee).filter(
            or_(
                Employee.employee_id == employee_id,
                Employee.id == int(employee_id) if employee_id.isdigit() else False
            )
        ).first()
        
        if not employee:
            raise HTTPException(status_code=404, detail="Employee not found")

        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # ===== Sheet 1: Basic Information =====
        ws_basic = wb.create_sheet("Basic Information", 0)
        ws_basic.append(["Field", "Value"])
        
        basic_data = [
            ["Employee ID", employee.employee_id],
            ["Name", employee.name],
            ["Email", employee.email],
            ["Role", employee.role or "N/A"],
            ["Location", employee.location or "N/A"],
            ["Date of Joining", employee.date_of_joining.strftime("%d-%b-%Y") if employee.date_of_joining else "N/A"],
            ["Team", employee.team or "N/A"],
            ["Category", employee.category or "N/A"],
            ["Employment Status", employee.employment_status or "Ongoing Employee"],
            ["Reporting To (Lead)", employee.lead or "N/A"],
            ["Experience (Years)", calculate_experience_years(employee.date_of_joining)],
            ["Active Status", "Active" if employee.is_active else "Inactive"],
        ]
        
        for row in basic_data:
            ws_basic.append(row)
        
        # Style header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        ws_basic['A1'].fill = header_fill
        ws_basic['A1'].font = header_font
        ws_basic['B1'].fill = header_fill
        ws_basic['B1'].font = header_font
        
        # Adjust column widths
        ws_basic.column_dimensions['A'].width = 25
        ws_basic.column_dimensions['B'].width = 40
        
        # ===== Sheet 2: Performance Metrics =====
        ws_perf = wb.create_sheet("Performance Metrics")
        ws_perf.append(["Metric", "Value", "Period"])
        
        # Get performance data directly from database
        try:
            # Get latest RAG status
            latest_review = db.query(EmployeeReview).filter(
                EmployeeReview.employee_id == employee.employee_id
            ).order_by(EmployeeReview.review_date.desc()).first()
            
            perf_rows = []
            if latest_review:
                perf_rows.extend([
                    ["RAG Status", latest_review.rag_status or "N/A", latest_review.review_period or "Overall"],
                    ["RAG Score", latest_review.rag_score or 0, latest_review.review_period or "Overall"],
                    ["Overall Rating", latest_review.overall_rating or 0, latest_review.review_period or "Overall"],
                ])
            
            # Get bug/ticket counts
            if employee.team == "DEVELOPMENT":
                bugs_resolved = db.query(Bug).filter(
                    Bug.assignee == employee.name,
                    Bug.status == "Resolved"
                ).count()
                bugs_created = db.query(Bug).filter(
                    Bug.author == employee.name
                ).count()
                tickets_completed = db.query(TicketTracking).filter(
                    TicketTracking.developer_assigned == employee.name,
                    TicketTracking.status == "Completed"
                ).count()
                tickets_in_progress = db.query(TicketTracking).filter(
                    TicketTracking.developer_assigned == employee.name,
                    TicketTracking.status.in_(["In Progress", "In Development"])
                ).count()
                
                perf_rows.extend([
                    ["Bugs Resolved", bugs_resolved, "Overall"],
                    ["Bugs Created", bugs_created, "Overall"],
                    ["Tickets Completed", tickets_completed, "Overall"],
                    ["Tickets In Progress", tickets_in_progress, "Overall"],
                ])
            else:
                bugs_found = db.query(Bug).filter(
                    Bug.author == employee.name
                ).count()
                bugs_resolved = db.query(Bug).filter(
                    Bug.assignee == employee.name,
                    Bug.status == "Resolved"
                ).count()
                test_cases_executed = db.query(TestResult).filter(
                    TestResult.executed_by == employee.name
                ).count()
                test_cases_passed = db.query(TestResult).filter(
                    TestResult.executed_by == employee.name,
                    TestResult.status == "Passed"
                ).count()
                
                perf_rows.extend([
                    ["Bugs Found", bugs_found, "Overall"],
                    ["Bugs Resolved", bugs_resolved, "Overall"],
                    ["Test Cases Executed", test_cases_executed, "Overall"],
                    ["Test Cases Passed", test_cases_passed, "Overall"],
                ])
            
            # Get timesheet summary (last 30 days)
            thirty_days_ago = date.today() - timedelta(days=30)
            timesheet_entries = db.query(EnhancedTimesheet).filter(
                EnhancedTimesheet.employee_name == employee.name,
                EnhancedTimesheet.date >= thirty_days_ago
            ).all()
            
            total_hours = sum(e.hours_logged or 0 for e in timesheet_entries)
            total_productive = sum(e.productive_hours or 0 for e in timesheet_entries)
            working_days = len(set(e.date for e in timesheet_entries))
            avg_daily = total_hours / working_days if working_days > 0 else 0
            
            perf_rows.extend([
                ["Total Hours Logged (30 days)", f"{total_hours:.1f}h", "Last 30 Days"],
                ["Total Productive Hours (30 days)", f"{total_productive:.1f}h", "Last 30 Days"],
                ["Working Days (30 days)", working_days, "Last 30 Days"],
                ["Avg Daily Hours (30 days)", f"{avg_daily:.1f}h", "Last 30 Days"],
            ])
            
            for row in perf_rows:
                ws_perf.append(row)
        except Exception as e:
            ws_perf.append(["Error", f"Could not fetch performance data: {str(e)}", "N/A"])
        
        # Style header
        ws_perf['A1'].fill = header_fill
        ws_perf['A1'].font = header_font
        ws_perf['B1'].fill = header_fill
        ws_perf['B1'].font = header_font
        ws_perf['C1'].fill = header_fill
        ws_perf['C1'].font = header_font
        
        ws_perf.column_dimensions['A'].width = 25
        ws_perf.column_dimensions['B'].width = 20
        ws_perf.column_dimensions['C'].width = 15
        
        # ===== Sheet 3: Goals =====
        ws_goals = wb.create_sheet("Goals & Development")
        ws_goals.append(["Type", "Title", "Description", "Status", "Progress %", "Target Date", "Created By"])
        
        try:
            goals = db.query(EmployeeGoal).filter(
                EmployeeGoal.employee_id == employee.employee_id
            ).order_by(EmployeeGoal.goal_type, EmployeeGoal.created_on.desc()).all()
            
            for goal in goals:
                goal_type_label = "Goal"
                if goal.goal_type == "strength":
                    goal_type_label = "Strength"
                elif goal.goal_type == "improvement":
                    goal_type_label = "Area of Improvement"
                
                ws_goals.append([
                    goal_type_label,
                    goal.title or "",
                    goal.description or "",
                    goal.status or "",
                    goal.progress or 0,
                    goal.target_date.strftime("%d-%b-%Y") if goal.target_date else "",
                    goal.created_by or ""
                ])
        except Exception as e:
            ws_goals.append(["Error", f"Could not fetch goals data: {str(e)}", "", "", "", "", ""])
        
        # Style header
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws_goals[f'{col}1'].fill = header_fill
            ws_goals[f'{col}1'].font = header_font
            ws_goals.column_dimensions[col].width = 20
        
        # ===== Sheet 4: Performance Reviews =====
        ws_reviews = wb.create_sheet("Performance Reviews")
        ws_reviews.append(["Review Period", "Review Date", "RAG Status", "RAG Score", "Overall Rating", 
                          "Technical", "Productivity", "Quality", "Communication", "Recommendation", "Reviewed By"])
        
        try:
            reviews = db.query(EmployeeReview).filter(
                EmployeeReview.employee_id == employee.employee_id
            ).order_by(EmployeeReview.review_date.desc()).all()
            
            for review in reviews:
                ws_reviews.append([
                    review.review_period or "",
                    review.review_date.strftime("%d-%b-%Y") if review.review_date else "",
                    review.rag_status or "",
                    review.rag_score or 0,
                    review.overall_rating or 0,
                    review.technical_rating or 0,
                    review.productivity_rating or 0,
                    review.quality_rating or 0,
                    review.communication_rating or 0,
                    review.recommendation or "",
                    review.reviewed_by or ""
                ])
        except Exception as e:
            ws_reviews.append(["Error", f"Could not fetch reviews data: {str(e)}", "", "", "", "", "", "", "", "", ""])
        
        # Style header
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            ws_reviews[f'{col}1'].fill = header_fill
            ws_reviews[f'{col}1'].font = header_font
            ws_reviews.column_dimensions[col].width = 15
        
        # ===== Sheet 5: KPI Ratings =====
        ws_kpi = wb.create_sheet("KPI Ratings")
        ws_kpi.append(["Quarter", "KPI Name", "Category", "Manager Rating", "Manager Comments", "Self Rating", "Self Comments"])
        
        try:
            # Get KPI ratings grouped by quarter
            kpi_ratings = db.query(KPIRating).filter(
                KPIRating.employee_id == employee.employee_id
            ).order_by(KPIRating.year.desc(), KPIRating.quarter_number.desc(), KPIRating.kpi_id).all()
            
            current_quarter = None
            for rating in kpi_ratings:
                quarter_str = f"{rating.year}-Q{rating.quarter_number}"
                if quarter_str != current_quarter:
                    current_quarter = quarter_str
                    ws_kpi.append([quarter_str, "", "", "", "", "", ""])  # Quarter header
                
                # Get KPI details
                kpi = db.query(KPI).filter(KPI.id == rating.kpi_id).first()
                kpi_name = kpi.kpi_name if kpi else f"KPI ID: {rating.kpi_id}"
                kpi_category = kpi.category if kpi else ""
                
                ws_kpi.append([
                    "",
                    kpi_name,
                    kpi_category,
                    rating.manager_rating or "",
                    rating.manager_comments or "",
                    rating.self_rating or "",
                    rating.self_comments or ""
                ])
        except Exception as e:
            ws_kpi.append(["Error", f"Could not fetch KPI data: {str(e)}", "", "", "", "", ""])
        
        # Style header
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws_kpi[f'{col}1'].fill = header_fill
            ws_kpi[f'{col}1'].font = header_font
            ws_kpi.column_dimensions[col].width = 20
        
        # ===== Sheet 6: Recent Timesheet Summary =====
        ws_timesheet = wb.create_sheet("Timesheet Summary")
        ws_timesheet.append(["Date", "Ticket ID", "Task Description", "Hours Logged", "Productive Hours", "Project Name", "Team"])
        
        try:
            # Get last 30 days of timesheet entries
            thirty_days_ago = date.today() - timedelta(days=30)
            entries = db.query(EnhancedTimesheet).filter(
                EnhancedTimesheet.employee_name == employee.name,
                EnhancedTimesheet.date >= thirty_days_ago
            ).order_by(EnhancedTimesheet.date.desc()).limit(100).all()
            
            for entry in entries:
                ws_timesheet.append([
                    entry.date.strftime("%d-%b-%Y") if entry.date else "",
                    entry.ticket_id or "",
                    entry.task_description or "",
                    entry.hours_logged or 0,
                    entry.productive_hours or 0,
                    entry.project_name or "",
                    entry.team or ""
                ])
        except Exception as e:
            ws_timesheet.append(["Error", f"Could not fetch timesheet data: {str(e)}", "", "", "", "", ""])
        
        # Style header
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws_timesheet[f'{col}1'].fill = header_fill
            ws_timesheet[f'{col}1'].font = header_font
            ws_timesheet.column_dimensions[col].width = 20
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        filename = f"Employee_Profile_{employee.employee_id}_{employee.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error exporting employee profile: {str(e)}")
    finally:
        db.close()


@app.post("/employees")
def create_employee(
    employee: EmployeeCreate,
    current_user: dict = Depends(require_role(["ADMIN"])),
):
    """Create a new employee"""
    db: Session = SessionLocal()
    try:
        # Check if employee_id or email already exists
        existing = db.query(Employee).filter(
            or_(
                Employee.employee_id == employee.employee_id,
                Employee.email == employee.email
            )
        ).first()
        
        if existing:
            raise HTTPException(status_code=400, detail="Employee ID or email already exists")
        
        new_employee = Employee(
            employee_id=employee.employee_id,
            name=employee.name,
            email=employee.email,
            role=employee.role,
            location=employee.location,
            date_of_joining=employee.date_of_joining,
            team=employee.team.upper() if employee.team else None,
            category=employee.category,
            employment_status=employee.employment_status or "Ongoing Employee",
            lead=employee.lead,
            is_active=True,
            created_on=datetime.utcnow()
        )
        
        db.add(new_employee)
        db.commit()
        db.refresh(new_employee)
        
        return {"message": "Employee created successfully", "id": new_employee.id, "employee_id": new_employee.employee_id}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


@app.put("/employees/{employee_id}")
def update_employee(
    employee_id: str,
    updates: EmployeeUpdate,
    current_user: dict = Depends(get_current_user),
):
    """Update an employee and cascade updates to related records"""
    db: Session = SessionLocal()
    try:
        employee = db.query(Employee).filter(
            or_(
                Employee.employee_id == employee_id,
                Employee.id == int(employee_id) if employee_id.isdigit() else False
            )
        ).first()
        
        if not employee:
            raise HTTPException(status_code=404, detail="Employee not found")
        if not can_edit_employee_profile(db, current_user, employee.employee_id):
            raise HTTPException(status_code=403, detail="Access denied - insufficient permissions to edit this profile")
        
        # Handle access_role update (requires ADMIN or MANAGER role)
        update_data = updates.dict(exclude_unset=True)
        if 'access_role' in update_data and update_data['access_role']:
            user_role = current_user.get("role", "")
            if user_role != "ADMIN" and "MANAGER" not in user_role:
                raise HTTPException(status_code=403, detail="Only Admin or Manager can change access roles")
            
            # Valid access roles
            valid_roles = ["ADMIN", "MANAGER_DEV", "MANAGER_QA", "LEAD_DEV", "LEAD_QA", "EMPLOYEE", "CLIENT"]
            new_access_role = update_data['access_role'].upper()
            if new_access_role not in valid_roles:
                raise HTTPException(status_code=400, detail=f"Invalid access role. Must be one of: {', '.join(valid_roles)}")
            
            # Update or create User record
            user = db.query(User).filter(User.employee_id == employee.employee_id).first()
            if user:
                user.role = new_access_role
            else:
                # Create new user with default password (employee_id)
                new_user = User(
                    email=employee.email.lower(),
                    password_hash=hash_password(employee.employee_id),
                    role=new_access_role,
                    employee_id=employee.employee_id,
                )
                db.add(new_user)
            
            # Remove access_role from update_data as it's not an Employee field
            del update_data['access_role']
        
        # Store old values for cascading updates
        old_name = employee.name
        old_lead = employee.lead
        old_manager = employee.manager
        
        update_data = updates.dict(exclude_unset=True)
        for field, value in update_data.items():
            if value is not None:
                if field == 'team' and value:
                    value = value.upper()
                if field == 'lead' and value:
                    value = _resolve_lead_or_manager_name(db, value) or value
                if field == 'manager' and value:
                    value = _resolve_lead_or_manager_name(db, value) or value
                if field == 'mode_of_work' and value:
                    # Validate mode_of_work
                    valid_modes = ["Onsite", "Remote", "Hybrid"]
                    if value not in valid_modes:
                        raise HTTPException(status_code=400, detail=f"Invalid mode_of_work. Must be one of: {', '.join(valid_modes)}")
                setattr(employee, field, value)
        
        # Auto-calculate expected_lwd when resignation_date is set and employment_status is "Serving Notice Period"
        if 'resignation_date' in update_data or 'employment_status' in update_data:
            if employee.employment_status == "Serving Notice Period" and employee.resignation_date:
                # Calculate notice period based on category (BILLED = 90 days, UN-BILLED = 30 days)
                notice_period_days = 90 if employee.category and employee.category.upper() == "BILLED" else 30
                # Only auto-calculate if expected_lwd is not explicitly provided
                if 'expected_lwd' not in update_data or update_data.get('expected_lwd') is None:
                    employee.expected_lwd = employee.resignation_date + timedelta(days=notice_period_days)
        
        new_name = employee.name
        new_lead = employee.lead
        new_manager = employee.manager
        
        # Cascade updates to related records
        update_count = 0
        
        # If employee name changed, update all employees who have this person as lead or manager (exact match, case-insensitive)
        if 'name' in update_data and old_name and new_name and old_name.strip().lower() != new_name.strip().lower():
            old_lower = old_name.strip().lower()
            lead_reportees = db.query(Employee).filter(Employee.lead.ilike(f"%{old_name}%")).all()
            for reportee in lead_reportees:
                if (reportee.lead or "").strip().lower() == old_lower:
                    reportee.lead = new_name
                    update_count += 1
            manager_reportees = db.query(Employee).filter(Employee.manager.ilike(f"%{old_name}%")).all()
            for reportee in manager_reportees:
                if (reportee.manager or "").strip().lower() == old_lower:
                    reportee.manager = new_name
                    update_count += 1
            
            # Update WeeklyPlan.planned_by
            weekly_plans = db.query(WeeklyPlan).filter(
                WeeklyPlan.planned_by.ilike(f"%{old_name}%")
            ).all()
            for plan in weekly_plans:
                if plan.planned_by and old_name in plan.planned_by:
                    plan.planned_by = plan.planned_by.replace(old_name, new_name)
                    update_count += 1
            
            # Update PlannedTask.assigned_by
            planned_tasks = db.query(PlannedTask).filter(
                PlannedTask.assigned_by.ilike(f"%{old_name}%")
            ).all()
            for task in planned_tasks:
                if task.assigned_by and old_name in task.assigned_by:
                    task.assigned_by = task.assigned_by.replace(old_name, new_name)
                    update_count += 1
        
        # Do NOT cascade lead/manager changes to other employees. Changing one employee's
        # lead (e.g. Amala -> Aravind) must only update that employee; others who report
        # to Amala must keep Amala as lead. Name-change cascade above handles correcting
        # a person's name when they are lead/manager (e.g. "Amala" -> "Amala R").

        employee.updated_on = datetime.utcnow()
        db.commit()
        
        message = f"Employee updated successfully"
        if update_count > 0:
            message += f". Updated {update_count} related record(s)."
        
        return {"message": message, "related_records_updated": update_count}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


@app.post("/employees/{employee_id}/photo")
async def upload_employee_photo(
    employee_id: str,
    request: Request,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
):
    """Upload and save employee profile photo."""
    db: Session = SessionLocal()
    try:
        employee = db.query(Employee).filter(Employee.employee_id == employee_id).first()
        if not employee:
            raise HTTPException(status_code=404, detail="Employee not found")
        if not can_access_employee(db, current_user, employee.employee_id):
            raise HTTPException(status_code=403, detail="Access denied")
        if not file.content_type or not file.content_type.startswith("image/"):
            raise HTTPException(status_code=400, detail="Invalid file type. Please upload an image.")

        filename = file.filename or ""
        ext = os.path.splitext(filename)[1].lower()
        allowed_exts = {".jpg", ".jpeg", ".png", ".webp", ".gif"}
        if ext and ext not in allowed_exts:
            raise HTTPException(status_code=400, detail="Unsupported image format.")

        if not ext:
            content_map = {
                "image/jpeg": ".jpg",
                "image/png": ".png",
                "image/webp": ".webp",
                "image/gif": ".gif"
            }
            ext = content_map.get(file.content_type, ".jpg")

        timestamp = int(datetime.utcnow().timestamp())
        safe_filename = f"{employee_id}_{timestamp}{ext}"
        file_path = os.path.join(PROFILE_PHOTO_DIR, safe_filename)

        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        base_url = str(request.base_url).rstrip("/")
        photo_url = f"{base_url}/uploads/profile_photos/{safe_filename}"

        employee.photo_url = photo_url
        employee.updated_on = datetime.utcnow()
        db.commit()

        return {"photo_url": photo_url}
    finally:
        db.close()


@app.delete("/employees/{employee_id}")
def delete_employee(
    employee_id: str,
    current_user: dict = Depends(require_role(["ADMIN"])),
):
    """Soft delete an employee (set is_active=False)"""
    db: Session = SessionLocal()
    try:
        employee = db.query(Employee).filter(
            or_(
                Employee.employee_id == employee_id,
                Employee.id == int(employee_id) if employee_id.isdigit() else False
            )
        ).first()
        
        if not employee:
            raise HTTPException(status_code=404, detail="Employee not found")
        
        employee.is_active = False
        employee.updated_on = datetime.utcnow()
        db.commit()
        
        return {"message": "Employee deactivated successfully"}
    finally:
        db.close()


@app.post("/employees/import")
async def import_employees(
    file: UploadFile = File(...),
    current_user: dict = Depends(require_role(["ADMIN"])),
):
    """Import employees from Excel file"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")
    
    # Save to temp file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name
    
    try:
        from sync_employees_to_db import import_employees as do_import
        success, imported, updated = do_import(tmp_path)
        
        if success:
            return {
                "success": True,
                "message": f"Import completed: {imported} new, {updated} updated",
                "imported": imported,
                "updated": updated
            }
        else:
            raise HTTPException(status_code=500, detail="Import failed")
    finally:
        os.unlink(tmp_path)


# ===== EMPLOYEE PERFORMANCE ENDPOINTS =====

def _build_employee_metrics(db, employee, start_date, end_date, is_dev, is_manager_role,
                            mode_of_work=None, *, tickets=None, bugs=None, tests=None,
                            timesheet_entries=None):
    """Build the per-employee `metrics` dict (tickets/bugs/tests/timesheet).

    When pre-fetched, pre-bucketed lists are passed (tickets/bugs/tests/timesheet_entries)
    they are used directly instead of querying — this lets the leaderboard avoid per-employee
    queries. With all list args left as None the behavior matches the original inline logic in
    get_employee_performance exactly.
    """
    employee_name = employee.name
    metrics = {}

    # ===== TICKET METRICS (from ticket_tracking) =====
    ticket_ids = []
    if not is_manager_role:
        if tickets is None:
            ticket_query = db.query(TicketTracking)
            if is_dev:
                ticket_query = ticket_query.filter(
                    or_(
                        TicketTracking.backend_developer.ilike(f"%{employee_name}%"),
                        TicketTracking.frontend_developer.ilike(f"%{employee_name}%")
                    )
                )
            else:  # QA
                ticket_query = ticket_query.filter(
                    TicketTracking.qc_tester.ilike(f"%{employee_name}%")
                )
            if start_date:
                ticket_query = ticket_query.filter(TicketTracking.updated_on >= start_date)
            tickets = ticket_query.all()
        ticket_ids = [t.ticket_id for t in tickets]

        # Calculate estimate vs actual
        total_estimate = sum(t.dev_estimate_hours or 0 for t in tickets) if is_dev else sum(t.qa_estimate_hours or 0 for t in tickets)
        total_actual = sum(t.actual_dev_hours or 0 for t in tickets) if is_dev else sum(t.actual_qa_hours or 0 for t in tickets)

        metrics["tickets"] = {
            "count": len(tickets),
            "ticket_ids": ticket_ids[:50],  # Limit to 50
            "estimate_hours": round(total_estimate, 1),
            "actual_hours": round(total_actual, 1),
            "estimate_accuracy": round((total_estimate / total_actual * 100), 1) if total_actual > 0 else 100
        }
    else:
        metrics["tickets"] = {
            "count": 0,
            "ticket_ids": [],
            "estimate_hours": 0,
            "actual_hours": 0,
            "estimate_accuracy": None,
        }

    # ===== BUG METRICS (from bugs) =====
    total_bugs = 0
    if not is_manager_role:
        if bugs is None:
            bug_query = db.query(Bug)
            if is_dev:
                bug_query = bug_query.filter(Bug.assignee.ilike(f"%{employee_name}%"))
            else:  # QA - bugs reported by this person
                bug_query = bug_query.filter(Bug.author.ilike(f"%{employee_name}%"))
            if start_date:
                bug_query = bug_query.filter(Bug.created_on >= start_date)
            bugs = bug_query.all()
        total_bugs = len(bugs)

        if total_bugs > 0:
            # Status breakdown
            closed_bugs = len([b for b in bugs if b.status == "Closed"])
            reopened_bugs = len([b for b in bugs if b.status == "Reopened"])
            rejected_bugs = len([b for b in bugs if b.status == "Rejected"])

            # Severity breakdown
            critical_bugs = len([b for b in bugs if b.severity == "Critical"])
            major_bugs = len([b for b in bugs if b.severity == "Major"])
            minor_bugs = len([b for b in bugs if b.severity == "Minor"])

            # Environment breakdown
            live_bugs = len([b for b in bugs if b.environment == "Live"])
            pre_bugs = len([b for b in bugs if b.environment == "Pre"])
            staging_bugs = len([b for b in bugs if b.environment == "Staging"])

            # Bug ageing (for open bugs)
            open_bugs = [b for b in bugs if b.status not in ["Closed", "Rejected"]]
            ages = []
            for bug in open_bugs:
                if bug.created_on:
                    age = (datetime.now() - bug.created_on).days
                    ages.append(age)
            avg_ageing = round(sum(ages) / len(ages), 1) if ages else 0

            # Resolution time (for closed bugs)
            resolution_times = []
            for bug in bugs:
                if bug.status == "Closed" and bug.created_on and bug.closed_on:
                    days = (bug.closed_on - bug.created_on).days
                    resolution_times.append(days)
            avg_resolution = round(sum(resolution_times) / len(resolution_times), 1) if resolution_times else 0

            # Modules expertise
            modules = list(set(b.module for b in bugs if b.module))

            # Bug types
            bug_types = defaultdict(int)
            for bug in bugs:
                tracker = bug.tracker or "Unknown"
                bug_types[tracker] += 1

            metrics["bugs"] = {
                "total": total_bugs,
                "closed": closed_bugs,
                "reopened": reopened_bugs,
                "rejected": rejected_bugs,
                "closure_rate": round((closed_bugs / total_bugs * 100), 1),
                "reopened_percent": round((reopened_bugs / total_bugs * 100), 1),
                "rejected_percent": round((rejected_bugs / total_bugs * 100), 1),
                "severity": {
                    "critical": critical_bugs,
                    "critical_percent": round((critical_bugs / total_bugs * 100), 1),
                    "major": major_bugs,
                    "minor": minor_bugs
                },
                "environment": {
                    "live": live_bugs,
                    "live_percent": round((live_bugs / total_bugs * 100), 1),
                    "pre": pre_bugs,
                    "pre_percent": round((pre_bugs / total_bugs * 100), 1),
                    "staging": staging_bugs,
                    "staging_percent": round((staging_bugs / total_bugs * 100), 1)
                },
                "avg_ageing_days": avg_ageing,
                "avg_resolution_days": avg_resolution,
                "modules_expertise": modules[:15],
                "bug_types": dict(bug_types)
            }
        else:
            metrics["bugs"] = {"total": 0}
    else:
        metrics["bugs"] = {"total": 0, "disabled": True}

    # ===== TESTRAIL METRICS (QA only) =====
    if not is_dev and not is_manager_role:
        if tests is None:
            test_query = db.query(TestResult).filter(
                TestResult.assigned_to.ilike(f"%{employee_name}%")
            )
            if start_date:
                test_query = test_query.filter(TestResult.created_on >= start_date)
            tests = test_query.all()
        test_results = tests
        total_tests = len(test_results)

        if total_tests > 0:
            passed = len([t for t in test_results if t.status_name == "Passed"])
            failed = len([t for t in test_results if t.status_name == "Failed"])
            blocked = len([t for t in test_results if t.status_name == "Blocked"])

            # Unique test runs
            unique_runs = len(set(t.run_id for t in test_results if t.run_id))

            metrics["tests"] = {
                "total_executed": total_tests,
                "passed": passed,
                "failed": failed,
                "blocked": blocked,
                "pass_rate": round((passed / total_tests * 100), 1),
                "fail_rate": round((failed / total_tests * 100), 1),
                "blocked_percent": round((blocked / total_tests * 100), 1),
                "test_runs_participated": unique_runs
            }

            # Bugs per ticket
            if len(ticket_ids) > 0:
                metrics["bugs_per_ticket"] = round(total_bugs / len(ticket_ids), 1)
        else:
            metrics["tests"] = {"total_executed": 0}
    elif is_manager_role:
        metrics["tests"] = {"total_executed": 0, "disabled": True}

    # ===== TIMESHEET METRICS =====
    timesheet_team = "DEV" if is_dev else "QA"
    if timesheet_entries is None:
        enhanced_query = db.query(EnhancedTimesheet).filter(
            EnhancedTimesheet.employee_name.ilike(f"%{employee_name}%"),
            EnhancedTimesheet.team == timesheet_team
        )
        if start_date:
            enhanced_query = enhanced_query.filter(EnhancedTimesheet.date >= start_date.date())
        enhanced_entries = enhanced_query.all()
        if enhanced_entries:
            total_hours = round(sum(e.hours_logged or 0 for e in enhanced_entries), 1)
            timesheet_entries_count = len(enhanced_entries)
        else:
            timesheet_query = db.query(Timesheet).filter(
                Timesheet.employee_name.ilike(f"%{employee_name}%")
            )
            if start_date:
                timesheet_query = timesheet_query.filter(Timesheet.date >= start_date.date())
            timesheets = timesheet_query.all()
            total_minutes = sum(t.time_logged_minutes or 0 for t in timesheets)
            total_hours = round(total_minutes / 60, 1)
            timesheet_entries_count = len(timesheets)
    else:
        total_hours = round(sum(e.hours_logged or 0 for e in timesheet_entries), 1)
        timesheet_entries_count = len(timesheet_entries)

    # Calculate working days in period
    if start_date:
        working_days = sum(1 for i in range((end_date - start_date).days + 1)
                         if (start_date + timedelta(days=i)).weekday() < 5)
    else:
        working_days = 250  # Approximate yearly working days

    expected_hours_per_day = 8
    mode_lower = (mode_of_work or "").lower()
    if "part" in mode_lower or "half" in mode_lower:
        expected_hours_per_day = 4
    elif "intern" in mode_lower:
        expected_hours_per_day = 6
    expected_hours = working_days * expected_hours_per_day

    metrics["timesheet"] = {
        "total_hours": total_hours,
        "expected_hours": expected_hours,
        "utilization_percent": round((total_hours / expected_hours * 100), 1) if expected_hours > 0 else 0,
        "avg_daily_hours": round(total_hours / working_days, 1) if working_days > 0 else 0,
        "entries_count": timesheet_entries_count
    }

    return metrics


@app.get("/employees/{employee_id}/performance")
def get_employee_performance(
    employee_id: str,
    period: str = Query("overall", description="past_week, past_month, past_quarter, one_year, overall"),
    current_user: dict = Depends(get_current_user),
):
    """Get comprehensive performance metrics for an employee"""
    db: Session = SessionLocal()
    try:
        employee = db.query(Employee).filter(
            or_(
                Employee.employee_id == employee_id,
                Employee.id == int(employee_id) if employee_id.isdigit() else False
            )
        ).first()
        if not employee:
            raise HTTPException(status_code=404, detail="Employee not found")
        if not can_access_employee(db, current_user, employee.employee_id):
            raise HTTPException(status_code=403, detail="Access denied")

        start_date, end_date = get_date_range(period)
        employee_name = employee.name
        team_upper = (employee.team or "").upper()
        is_dev = "DEV" in team_upper or team_upper == "DEVELOPMENT"

        # Access role and role context (manager roles have different evaluation)
        user_account = db.query(User).filter(User.employee_id == employee.employee_id).first()
        access_role = user_account.role if user_account else None
        is_manager_role = (access_role and "MANAGER" in access_role.upper()) or ("MANAGER" in (employee.role or "").upper())

        # Mode of work (if available in mapping_data)
        mode_of_work = None
        mapping = employee.mapping_data or {}
        for key in ("mode_of_work", "work_mode", "Mode of Work", "Work Mode", "Mode", "ModeOfWork"):
            if mapping.get(key):
                mode_of_work = mapping.get(key)
                break
        
        # Build base response
        result = {
            "employee": {
                "id": employee.id,
                "employee_id": employee.employee_id,
                "name": employee.name,
                "team": employee.team,
                "role": employee.role,
                "category": employee.category,
                "lead": employee.lead,
                "experience_years": calculate_experience_years(employee.date_of_joining)
            },
            "role_context": {
                "access_role": access_role,
                "designation": employee.role,
                "team": employee.team,
                "mode_of_work": mode_of_work,
                "is_manager": bool(is_manager_role),
                "role_type": "manager" if is_manager_role else ("dev" if is_dev else "qa"),
            },
            "period": period,
            "metrics": {}
        }

        result["metrics"] = _build_employee_metrics(
            db, employee, start_date, end_date, is_dev, is_manager_role, mode_of_work
        )

        # ===== PLANNING TIMESHEET SUMMARY (for rating context) =====
        plan_ts_summary = _get_planning_timesheet_summary(db, employee, weeks=5)
        if plan_ts_summary:
            result["planning_timesheet"] = plan_ts_summary

        # ===== RAG SCORE CALCULATION =====
        rag_score = calculate_rag_score(result["metrics"], is_dev, planning_timesheet=plan_ts_summary, role_context=result.get("role_context"))
        result["rag_status"] = {
            "score": rag_score,
            "status": "GREEN" if rag_score >= 70 else "AMBER" if rag_score >= 50 else "RED"
        }
        
        return result
    finally:
        db.close()


# Leaderboard weights (sum = 100). Tune here.
# Presence (attendance) is a major billing criterion; quality stays high; throughput credits
# work awaiting external review so raw closure count is not the deciding factor.
LEADERBOARD_WEIGHTS = {"presence": 25, "throughput": 20, "output": 12, "quality": 30, "efficiency": 13}

# Leave is a billing loss: deduct this many composite points per leave day taken, capped.
LEAVE_PENALTY_PER_DAY = 1.5
LEAVE_PENALTY_CAP = 15
# Leave types that count as a billing-loss absence (WFH = still working; Holiday = company-wide).
ABSENCE_LEAVE_TYPES = ("leave", "sick leave", "casual leave", "half day", "earned leave", "lop")

# Statuses where the person's work is done but the ticket is held for EXTERNAL review/deploy
# (waiting in BIS, or approved and awaiting go-live). Not the person's delay — credited so they
# aren't penalized for a low closed-count caused by tickets sitting in review.
AWAITING_REVIEW_STATUSES = ("BIS Testing", "Approved for Live")
AWAITING_CREDIT = 0.7  # partial credit vs a fully shipped ticket

# QA testers on the mobile/sprint model — excluded from the QA board. (Now empty: all QA team
# members are considered, including the mobile testers Anjaly/Arya.)
MOBILE_QA_EXCLUDE = ()

# Specific people excluded from the leaderboard. Compact-name matches disambiguate similar names
# ("Vishnu V S" excluded but "Vishnu CS" kept); token matches use a unique first name.
PERFORMANCE_EXCLUDE_COMPACT = ("vishnuvs",)
PERFORMANCE_EXCLUDE_TOKENS = ("vivek", "varsha")  # Vivek V Nair, Varsha Dcruz P


def _ticket_complexity(t, is_dev):
    """Per-ticket complexity by DEPTH of work done: priority weight (1.0-3.0) scaled by the
    ACTUAL hours worked on the ticket (effort/depth), falling back to the estimate when actuals
    are missing. Deeper, harder tickets (more hours on a higher-priority item) score more than
    shallow ones, so volume alone doesn't dominate."""
    rank = PRIORITY_ORDER.get((t.priority or "").strip(), 13)
    priority_weight = 1.0 + (13 - rank) / 12 * 2.0  # URGENT≈3.0 … Suggestion≈1.0
    actual = (t.actual_dev_hours if is_dev else t.actual_qa_hours) or 0
    est = (t.dev_estimate_hours if is_dev else t.qa_estimate_hours) or 0
    depth_hours = actual or est  # prefer real effort; fall back to estimate
    return priority_weight * (1 + min(depth_hours, 40) / 40)


def _strip_paren(name):
    """Drop a trailing '(login)' so PM names match bug/timesheet names.

    PM ticket developer names look like 'Abhijai Kp (AKP916)' while bugs/timesheets store
    plain 'Abhijai K P'. Removing the parenthetical lets the compacted-name match succeed.
    """
    return re.sub(r"\([^)]*\)", " ", name or "")


def _employee_mode_of_work(emp):
    mp = emp.mapping_data or {}
    for k in ("mode_of_work", "work_mode", "Mode of Work", "Work Mode", "Mode", "ModeOfWork"):
        if mp.get(k):
            return mp.get(k)
    return None


@app.get("/employees/performance/leaderboard")
def get_performance_leaderboard(
    period: str = Query("month", regex="^(month|quarter)$"),
    offset: int = Query(0, ge=0, le=240),
    team: str = Query("all", regex="^(qa|dev|all)$"),
):
    """Per-team performance leaderboards (balanced composite score) for a calendar month/quarter.

    Combines throughput (ticket volume × complexity), output (bugs/test execution), quality
    (reused RAG score), and efficiency (estimate accuracy + utilization). Volume sub-scores are
    normalized against the team max so higher output genuinely ranks higher at equal quality.
    Returns the full ranked list per team (frontend shows top 3 + expandable rest).
    """
    db: Session = SessionLocal()
    try:
        start_date, end_date, label = get_period_range(period, offset)

        # Once a period has ended its leaderboard is frozen: computed once, then served unchanged
        # so history never shifts as ticket/bug data drifts afterwards.
        period_ended = end_date.date() < datetime.now().date()
        period_key = f"{period}:{label}:{team}"
        if period_ended:
            snap = db.query(PerformanceSnapshot).filter_by(period_key=period_key).first()
            if snap and snap.frozen and snap.payload:
                return snap.payload

        # Delivered-in-period tickets: real close/ship date in the window.
        # (TicketTracking.updated_on is only the last sync time, so it cannot define a period.)
        closed_tickets = db.query(TicketTracking).filter(
            TicketTracking.closed_on >= start_date,
            TicketTracking.closed_on <= end_date,
        ).all()

        # ---- Build the roster ----
        # Prefer the employees table; where it is empty/partial (some deployments do not sync it),
        # fall back to the people who actually appear on delivered tickets: QC testers => QA,
        # backend/frontend developers => Dev. Lightweight shim objects stand in for missing rows.
        user_roles = {u.employee_id: (u.role or "") for u in db.query(User).all()}

        def _is_manager(emp):
            return ("MANAGER" in (emp.role or "").upper()) or \
                   ("MANAGER" in user_roles.get(emp.employee_id, "").upper())

        class _LbPerson:
            __slots__ = ("name", "employee_id", "team", "role", "mapping_data")

            def __init__(self, name, team):
                self.name = name
                self.employee_id = "name:" + _compact_person_name(name)
                self.team = team
                self.role = team
                self.mapping_data = {}

        qa_by_id, dev_by_id = {}, {}
        name_to_employee_id = {}

        def _excluded_qa(name):
            # Mobile/sprint-model QA testers are not ranked on this delivery board.
            toks = _normalize_person_name(name).split()
            return any(x in toks for x in MOBILE_QA_EXCLUDE)

        def _excluded(name):
            # Specific people excluded from both boards (compact-name match, or unique first-name token).
            if _compact_person_name(name) in PERFORMANCE_EXCLUDE_COMPACT:
                return True
            toks = _normalize_person_name(name).split()
            return any(x in toks for x in PERFORMANCE_EXCLUDE_TOKENS)

        def _register(person, team_ids):
            team_ids[person.employee_id] = person
            if person.name:
                for variant in (person.name.strip(), _strip_paren(person.name).strip()):
                    if not variant:
                        continue
                    name_to_employee_id.setdefault(variant, person.employee_id)
                    name_to_employee_id.setdefault(_normalize_person_name(variant), person.employee_id)
                    name_to_employee_id.setdefault(_compact_person_name(variant), person.employee_id)

        for emp in db.query(Employee).filter(
            Employee.is_active == True, Employee.archived == False,
        ).all():
            if _is_manager(emp) or _excluded(emp.name):
                continue
            tu = (emp.team or "").upper()
            if "DEV" in tu:
                _register(emp, dev_by_id)
            elif "QA" in tu and not _excluded_qa(emp.name):
                _register(emp, qa_by_id)

        # EmployeeNameMapping aliases (only for already-registered employees).
        known_ids = set(qa_by_id) | set(dev_by_id)
        try:
            for m in db.query(EmployeeNameMapping).filter(EmployeeNameMapping.is_active == True).all():
                if m.employee_id and m.employee_id in known_ids and m.alternate_name:
                    alt = m.alternate_name.strip()
                    name_to_employee_id.setdefault(alt, m.employee_id)
                    name_to_employee_id.setdefault(_normalize_person_name(alt), m.employee_id)
                    name_to_employee_id.setdefault(_compact_person_name(alt), m.employee_id)
        except Exception:
            pass

        # First+last-name index over EMPLOYEES only, kept only when unambiguous. Lets a short PM
        # name ("Gautham Krishna") resolve to the fuller employee record ("Gautham Krishna KP")
        # instead of spawning a duplicate, without merging distinct people (e.g. the various Vishnus).
        _f2_sets = defaultdict(set)
        for _eid, _p in list(qa_by_id.items()) + list(dev_by_id.items()):
            _tk = _normalize_person_name(_p.name).split()
            if len(_tk) >= 2:
                _f2_sets[(_tk[0], _tk[1])].add(_eid)
        first2_to_eid = {k: next(iter(v)) for k, v in _f2_sets.items() if len(v) == 1}

        def _resolve(raw):
            if not raw:
                return None
            for variant in (raw.strip(), _strip_paren(raw).strip()):
                eid = (name_to_employee_id.get(variant)
                       or name_to_employee_id.get(_normalize_person_name(variant))
                       or name_to_employee_id.get(_compact_person_name(variant)))
                if eid:
                    return eid
            toks = _normalize_person_name(_strip_paren(raw)).split()
            if len(toks) >= 2:
                return first2_to_eid.get((toks[0], toks[1]))
            return None

        # Add ticket-derived persons for any names not already mapped to a real employee.
        qa_name_counts, dev_name_counts = defaultdict(int), defaultdict(int)
        for t in closed_tickets:
            qc = _strip_paren(t.qc_tester or "").strip()
            if qc:
                qa_name_counts[qc] += 1
            for nm in (t.backend_developer, t.frontend_developer):
                nm = _strip_paren(nm or "").strip()
                if nm:
                    dev_name_counts[nm] += 1
        # Overlaps go to the role where the name appears more often (tie => QA).
        for nm, qc in qa_name_counts.items():
            if not _resolve(nm) and not _excluded_qa(nm) and not _excluded(nm) and qc >= dev_name_counts.get(nm, 0):
                _register(_LbPerson(nm, "QA"), qa_by_id)
        for nm in dev_name_counts:
            if not _resolve(nm) and not _excluded(nm):
                _register(_LbPerson(nm, "DEV"), dev_by_id)

        qa_emps = list(qa_by_id.values())
        dev_emps = list(dev_by_id.values())
        qa_ids = set(qa_by_id)
        dev_ids = set(dev_by_id)

        # ---- Bucket delivered tickets by person ----
        tickets_by_emp = defaultdict(list)
        seen_ticket_pairs = set()

        def _add_ticket(eid, t):
            key = (eid, t.id)
            if eid and key not in seen_ticket_pairs:
                seen_ticket_pairs.add(key)
                tickets_by_emp[eid].append(t)

        for t in closed_tickets:
            qa_eid = _resolve(t.qc_tester)
            if qa_eid in qa_ids:
                _add_ticket(qa_eid, t)
            for nm in (t.backend_developer, t.frontend_developer):
                d_eid = _resolve(nm)
                if d_eid in dev_ids:
                    _add_ticket(d_eid, t)

        # Tickets the person handed off that are now awaiting external review (BIS) / go-live.
        # Current-status based, so only meaningful for the live (current) period.
        awaiting_by_emp = defaultdict(list)
        if end_date.date() >= datetime.now().date():
            seen_await = set()

            def _add_await(eid, t):
                key = (eid, t.id)
                if eid and key not in seen_await:
                    seen_await.add(key)
                    awaiting_by_emp[eid].append(t)

            for t in db.query(TicketTracking).filter(
                TicketTracking.status.in_(AWAITING_REVIEW_STATUSES)
            ).all():
                qa_eid = _resolve(t.qc_tester)
                if qa_eid in qa_ids:
                    _add_await(qa_eid, t)
                for nm in (t.backend_developer, t.frontend_developer):
                    d_eid = _resolve(nm)
                    if d_eid in dev_ids:
                        _add_await(d_eid, t)

        bugs_by_emp = defaultdict(list)
        for b in db.query(Bug).filter(
            Bug.created_on >= start_date,
            Bug.created_on <= end_date,
        ).all():
            qa_eid = _resolve(b.author)        # QA reported the bug
            if qa_eid in qa_ids:
                bugs_by_emp[qa_eid].append(b)
            dev_eid = _resolve(b.assignee)     # Dev assigned to fix it
            if dev_eid in dev_ids:
                bugs_by_emp[dev_eid].append(b)

        tests_by_emp = defaultdict(list)
        for r in db.query(TestResult).filter(
            TestResult.created_on >= start_date,
            TestResult.created_on <= end_date,
        ).all():
            eid = _resolve(r.assigned_to)
            if eid in qa_ids:
                tests_by_emp[eid].append(r)

        ts_by_emp = defaultdict(list)
        for e in db.query(EnhancedTimesheet).filter(
            EnhancedTimesheet.date >= start_date.date(),
            EnhancedTimesheet.date <= end_date.date(),
        ).all():
            eid = _resolve(e.employee_name)
            if eid:
                ts_by_emp[eid].append(e)

        # Leave days taken in the period (billing loss), counted exactly like the Calendar module:
        # sum(leave hours)/8 over actual-leave entries (any "Leave" type; WFH/Holiday excluded).
        leave_days_by_emp = defaultdict(float)
        for lv in db.query(LeaveEntry).filter(
            LeaveEntry.date >= start_date.date(),
            LeaveEntry.date <= end_date.date(),
        ).all():
            lt = (lv.leave_type or "").strip().lower()
            if "leave" not in lt:  # actual leave only — excludes WFH / Holiday
                continue
            eid = _resolve(lv.employee_name)
            if eid:
                leave_days_by_emp[eid] += (lv.hours or 8) / 8.0

        # Working days in the period (Mon–Fri) for attendance/presence.
        period_working_days = sum(
            1 for i in range((end_date.date() - start_date.date()).days + 1)
            if (start_date.date() + timedelta(days=i)).weekday() < 5
        ) or 1

        def _build_team(emps, is_dev):
            rows = []
            for emp in emps:
                etickets = tickets_by_emp.get(emp.employee_id, [])
                ebugs = bugs_by_emp.get(emp.employee_id, [])
                etests = None if is_dev else tests_by_emp.get(emp.employee_id, [])
                ets = ts_by_emp.get(emp.employee_id, [])
                metrics = _build_employee_metrics(
                    db, emp, start_date, end_date, is_dev, False,
                    _employee_mode_of_work(emp),
                    tickets=etickets, bugs=ebugs, tests=etests, timesheet_entries=ets,
                )
                eawait = awaiting_by_emp.get(emp.employee_id, [])
                tcount = metrics["tickets"]["count"]
                bcount = metrics["bugs"].get("total", 0)
                testcount = metrics.get("tests", {}).get("total_executed", 0)
                awaiting_n = len(eawait)
                if tcount == 0 and bcount == 0 and testcount == 0 and awaiting_n == 0:
                    continue  # no activity this period → off the board
                delivered_cwv = sum(_ticket_complexity(t, is_dev) for t in etickets)
                awaiting_cwv = sum(_ticket_complexity(t, is_dev) for t in eawait)
                # Throughput credits shipped work fully and handed-off/awaiting-review work partially,
                # so a low closed-count caused by tickets sitting in BIS review isn't penalized.
                throughput_base = delivered_cwv + AWAITING_CREDIT * awaiting_cwv
                if is_dev:
                    output_raw = metrics["bugs"].get("closed", 0) + 0.5 * bcount
                else:
                    output_raw = bcount + testcount
                rag = calculate_rag_score(metrics, is_dev, planning_timesheet=None,
                                          role_context={"is_manager": False})
                # Presence / attendance (billing): days physically logged + productive hours.
                daily_hours = defaultdict(float)
                for e in ets:
                    daily_hours[e.date] += (e.hours_logged or 0)
                present_days = len([d for d, h in daily_hours.items() if h > 0])
                days_under_8 = len([d for d, h in daily_hours.items() if 0 < h < 8])
                days_over_8 = len([d for d, h in daily_hours.items() if h > 8])
                total_logged = round(sum(daily_hours.values()), 1)
                avg_hours_per_day = round(total_logged / present_days, 1) if present_days else 0
                productive_hours = round(sum((e.productive_hours or e.hours_logged or 0) for e in ets), 1)
                expected_hours = period_working_days * 8
                attendance_ratio = present_days / period_working_days
                prod_ratio = (productive_hours / expected_hours) if expected_hours else 0
                presence = round(100 * min(1.0, 0.6 * attendance_ratio + 0.4 * min(1.0, prod_ratio)), 1)
                leave_days = round(leave_days_by_emp.get(emp.employee_id, 0), 1)
                # Time overruns: tickets where ACTUAL exceeded ESTIMATE (genuine overrun — scope
                # changes get re-estimated so actual≈estimate and don't count). Per-ticket so an
                # underrun elsewhere can't mask it.
                estd = overrun_tickets = 0
                overrun_hours = 0.0
                for t in etickets:
                    est = (t.dev_estimate_hours if is_dev else t.qa_estimate_hours) or 0
                    act = (t.actual_dev_hours if is_dev else t.actual_qa_hours) or 0
                    if est > 0:
                        estd += 1
                        if act > est:
                            overrun_tickets += 1
                            overrun_hours += (act - est)
                on_time_rate = round(100 * (estd - overrun_tickets) / estd, 1) if estd else 100.0
                rows.append({"emp": emp, "metrics": metrics,
                             "delivered_cwv": round(delivered_cwv, 1),
                             "awaiting_cwv": round(awaiting_cwv, 1),
                             "throughput_base": throughput_base,
                             "awaiting_n": awaiting_n,
                             "output_raw": output_raw, "rag": rag,
                             "presence": presence, "present_days": present_days,
                             "productive_hours": productive_hours, "leave_days": leave_days,
                             "avg_hours_per_day": avg_hours_per_day, "days_under_8": days_under_8,
                             "days_over_8": days_over_8, "total_logged": total_logged,
                             "on_time_rate": on_time_rate, "overrun_tickets": overrun_tickets,
                             "overrun_hours": round(overrun_hours, 1), "estimated_tickets": estd})

            max_cwv = max((r["throughput_base"] for r in rows), default=0) or 0
            max_out = max((r["output_raw"] for r in rows), default=0) or 0
            scored = []
            for r in rows:
                m = r["metrics"]
                throughput = 100 * (r["throughput_base"] / max_cwv) ** 0.5 if max_cwv > 0 else 0
                output = 100 * (r["output_raw"] / max_out) ** 0.5 if max_out > 0 else 0
                quality = r["rag"]
                est_acc = m["tickets"].get("estimate_accuracy")
                est_acc = 100 if est_acc is None else est_acc
                util = m["timesheet"].get("utilization_percent", 0)
                # Efficiency blends on-time delivery (per-ticket: did actual stay within estimate —
                # genuine overruns lower this; re-estimated scope changes stay on-time) with how close
                # the aggregate estimate matched actual, plus utilization (real logged effort).
                on_time = r["on_time_rate"]
                est_close = max(0, 100 - abs(100 - est_acc))
                efficiency = round(0.45 * on_time + 0.25 * est_close + 0.30 * min(100, util), 1)
                presence = r["presence"]
                sub = {"presence": presence, "throughput": round(throughput, 1),
                       "output": round(output, 1), "quality": round(quality, 1),
                       "efficiency": round(efficiency, 1)}
                contrib = {k: round(sub[k] * LEADERBOARD_WEIGHTS[k] / 100, 1) for k in sub}
                # Leave is a billing loss — deduct from the composite.
                leave_penalty = round(min(LEAVE_PENALTY_CAP, r["leave_days"] * LEAVE_PENALTY_PER_DAY), 1)
                composite = round(max(0, sum(contrib.values()) - leave_penalty), 1)
                delivered = m["tickets"]["count"]
                bugs = m["bugs"].get("total", 0)
                tests = m.get("tests", {}).get("total_executed", 0)
                hours = m["timesheet"]["total_hours"]

                # Human-readable "how they earned this" summary.
                summary_lines = [f"{r['present_days']}/{period_working_days} days present · "
                                 f"avg {r['avg_hours_per_day']}h/day · {r['days_under_8']} day(s) under 8h · "
                                 f"{r['productive_hours']}h productive (attendance {presence})"]
                if r["leave_days"]:
                    summary_lines.append(f"{r['leave_days']} leave day(s) taken "
                                         f"(−{leave_penalty} billing-loss penalty)")
                summary_lines.append(f"Delivered {delivered} ticket(s) to live "
                                     f"({r['delivered_cwv']} depth pts)")
                if r["awaiting_n"]:
                    summary_lines.append(f"{r['awaiting_n']} ticket(s) handed off, awaiting BIS "
                                         f"review/go-live (credited)")
                if bugs:
                    summary_lines.append(f"{'Found' if not is_dev else 'Handled'} {bugs} bug(s)")
                if tests:
                    summary_lines.append(f"{tests} test result(s) executed")
                summary_lines.append(f"{round(quality, 1)}% quality · {est_acc}% estimate accuracy")
                if r["overrun_tickets"]:
                    summary_lines.append(f"{r['overrun_tickets']} ticket(s) over estimate "
                                         f"(+{r['overrun_hours']}h overrun) · on-time {on_time}%")

                scored.append({
                    "employee_id": r["emp"].employee_id,
                    "name": r["emp"].name,
                    "team": r["emp"].team,
                    "role": r["emp"].role,
                    "composite_score": composite,
                    "sub_scores": sub,
                    "weighted_contributions": contrib,
                    "leave_penalty": leave_penalty,
                    "summary_lines": summary_lines,
                    "raw_metrics": {
                        "tickets": delivered,
                        "delivered_to_live": delivered,            # shipped (closed) in-period
                        "awaiting_review": r["awaiting_n"],        # handed off, awaiting BIS/go-live
                        "complexity_weighted_volume": r["delivered_cwv"],
                        "awaiting_complexity": r["awaiting_cwv"],
                        "bugs": bugs,
                        "test_results_executed": tests,
                        "hours": hours,
                        "present_days": r["present_days"],
                        "working_days": period_working_days,
                        "productive_hours": r["productive_hours"],
                        "avg_hours_per_day": r["avg_hours_per_day"],
                        "days_under_8": r["days_under_8"],
                        "days_over_8": r["days_over_8"],
                        "total_logged_hours": r["total_logged"],
                        "leave_days": r["leave_days"],
                        "quality_percent": round(quality, 1),
                        "estimate_accuracy": est_acc,
                        "on_time_rate": r["on_time_rate"],
                        "overrun_tickets": r["overrun_tickets"],
                        "overrun_hours": r["overrun_hours"],
                        "utilization_percent": util,
                        "rag_score": round(r["rag"], 1),
                    },
                    "_tb": r["throughput_base"],
                })
            scored.sort(key=lambda x: (x["composite_score"], x["_tb"]), reverse=True)
            for i, s in enumerate(scored, 1):
                s["rank"] = i
                s.pop("_tb", None)
            return scored

        def _team_summary(lst):
            if not lst:
                return {"delivered_total": 0, "bugs_total": 0, "avg_quality": 0, "members": 0}
            return {
                "delivered_total": sum(s["raw_metrics"]["delivered_to_live"] for s in lst),
                "bugs_total": sum(s["raw_metrics"]["bugs"] for s in lst),
                "avg_quality": round(sum(s["raw_metrics"]["quality_percent"] for s in lst) / len(lst), 1),
                "members": len(lst),
            }

        resp = {"period": {"kind": period, "offset": offset, "label": label,
                           "start": start_date.isoformat(), "end": end_date.isoformat(),
                           "ended": period_ended, "frozen": period_ended}}
        if team in ("qa", "all"):
            resp["qa"] = _build_team(qa_emps, False)
        if team in ("dev", "all"):
            resp["dev"] = _build_team(dev_emps, True)
        resp["summary"] = {"qa": _team_summary(resp.get("qa", [])),
                           "dev": _team_summary(resp.get("dev", []))}

        # Freeze ended periods on first computation (immutable thereafter).
        if period_ended:
            try:
                db.add(PerformanceSnapshot(
                    period_key=period_key, period_kind=period, period_label=label,
                    team=team, payload=resp, frozen=True,
                ))
                db.commit()
            except Exception:
                db.rollback()
        return resp
    finally:
        db.close()


# QA testing statuses — a ticket "enters QA" when it moves into one of these from a non-QA status.
QA_TESTING_STATUSES = ("QC Testing", "QC Testing in Progress", "QC Testing Hold", "Testing In Progress")


def _compute_qa_flow_month(db, start_date, end_date):
    """QA flow for one month: fresh received into QA, handed to BIS, and closed — with module &
    QC-tester breakdowns. Closed comes from ticket_tracking.closed_on (full history); fresh/BIS come
    from TicketStatusHistory transitions (only available from when history capture began)."""
    closed = db.query(TicketTracking).filter(
        TicketTracking.closed_on >= start_date, TicketTracking.closed_on <= end_date,
    ).all()
    trans = db.query(TicketStatusHistory).filter(
        TicketStatusHistory.changed_on >= start_date, TicketStatusHistory.changed_on <= end_date,
    ).order_by(TicketStatusHistory.changed_on).all()

    tids = {t.ticket_id for t in closed} | {h.ticket_id for h in trans}
    mod_map = {}
    if tids:
        for tid, sub in db.query(TicketTracking.ticket_id, TicketTracking.subdepartment).filter(
            TicketTracking.ticket_id.in_(tids)
        ).all():
            mod_map[tid] = (sub or "").strip() or "Unassigned"

    def modname(tid):
        return mod_map.get(tid, "Unassigned")

    # per-module / per-tester {fresh, bis, closed}
    by_module = defaultdict(lambda: {"fresh": 0, "bis": 0, "closed": 0})
    by_tester = defaultdict(lambda: {"fresh": 0, "bis": 0, "closed": 0})

    fresh_tids, bis_tids = set(), set()
    for h in trans:
        ns, ps = h.new_status, h.previous_status
        if ns in QA_TESTING_STATUSES and ps not in QA_TESTING_STATUSES and h.ticket_id not in fresh_tids:
            fresh_tids.add(h.ticket_id)
            by_module[modname(h.ticket_id)]["fresh"] += 1
            by_tester[(h.qc_tester or "Unassigned")]["fresh"] += 1
        if ns == "BIS Testing" and h.ticket_id not in bis_tids:
            bis_tids.add(h.ticket_id)
            by_module[modname(h.ticket_id)]["bis"] += 1
            by_tester[(h.qc_tester or "Unassigned")]["bis"] += 1
    for t in closed:
        by_module[modname(t.ticket_id)]["closed"] += 1
        by_tester[(t.qc_tester or "Unassigned")]["closed"] += 1

    def _rows(d, key):
        rows = [{key: k, **v} for k, v in d.items()]
        rows.sort(key=lambda x: (x["fresh"] + x["bis"] + x["closed"]), reverse=True)
        return rows

    return {
        "fresh_received": len(fresh_tids),
        "handed_to_bis": len(bis_tids),
        "closed": len(closed),
        "by_module": _rows(by_module, "module"),
        "by_qc_tester": _rows(by_tester, "qc_tester"),
    }


@app.get("/qa-flow")
def get_qa_flow(offset: int = Query(0, ge=0, le=240), trend: int = Query(6, ge=1, le=24)):
    """Monthly QA flow (fresh received / handed to BIS / closed) with module + QC-tester detail for
    the selected month, plus a trailing trend series for the comparison graph. Ended months are
    frozen (snapshot) so history is preserved as it accumulates."""
    db: Session = SessionLocal()
    try:
        today = datetime.now().date()

        def _month(off):
            s, e, lbl = get_period_range("month", off)
            ended = e.date() < today
            if ended:
                snap = db.query(QAFlowSnapshot).filter_by(period_label=lbl).first()
                if snap and snap.frozen and snap.payload:
                    return lbl, s, e, ended, snap.payload
            data = _compute_qa_flow_month(db, s, e)
            if ended:
                try:
                    db.add(QAFlowSnapshot(period_label=lbl, payload=data, frozen=True))
                    db.commit()
                except Exception:
                    db.rollback()
            return lbl, s, e, ended, data

        label, start, end, ended, detail = _month(offset)
        result = dict(detail)
        result["period"] = {"label": label, "start": start.isoformat(),
                            "end": end.isoformat(), "frozen": ended}

        trend_series = []
        for k in range(trend - 1, -1, -1):
            lbl, _s, _e, _ed, d = _month(offset + k)
            trend_series.append({"label": lbl, "fresh_received": d["fresh_received"],
                                 "handed_to_bis": d["handed_to_bis"], "closed": d["closed"]})
        result["trend"] = trend_series
        return result
    finally:
        db.close()


def _compute_ticket_movement(db, start_date, end_date):
    """Per-month ticket movement: tickets that came to QC as NEW vs for REFIX (with refix count),
    were delivered to BIS, approved for live, and closed — each as a list with module/QC-tester
    detail. New/refix/BIS/approved use TicketStatusHistory transitions; closed uses closed_on."""
    # All transitions up to end of month, in order, so refix (re-entry) counts are correct.
    trans = db.query(TicketStatusHistory).filter(
        TicketStatusHistory.changed_on <= end_date,
    ).order_by(TicketStatusHistory.changed_on).all()
    closed = db.query(TicketTracking).filter(
        TicketTracking.closed_on >= start_date, TicketTracking.closed_on <= end_date,
    ).all()

    # Authoritative fail-loop (QC cycle) count per ticket, from the cycle tracker.
    cycle_count = {}
    try:
        import json as _json
        _p = os.path.join(os.path.dirname(__file__), "data", "qc_cycle_tracker.json")
        if os.path.exists(_p):
            with open(_p, encoding="utf-8") as _f:
                for k, v in _json.load(_f).items():
                    try:
                        cycle_count[int(k)] = (v or {}).get("cycle_count", 0) or 0
                    except Exception:
                        pass
    except Exception:
        pass

    tids = {h.ticket_id for h in trans} | {t.ticket_id for t in closed}
    meta = {}
    if tids:
        for (tid, title, sub, prio, qc, bd, fd, st, cr, eta, qae, qaa) in db.query(
            TicketTracking.ticket_id, TicketTracking.title, TicketTracking.subdepartment,
            TicketTracking.priority, TicketTracking.qc_tester, TicketTracking.backend_developer,
            TicketTracking.frontend_developer, TicketTracking.status, TicketTracking.created_on,
            TicketTracking.eta, TicketTracking.qa_estimate_hours, TicketTracking.actual_qa_hours,
        ).filter(TicketTracking.ticket_id.in_(tids)).all():
            meta[tid] = {
                "title": title, "module": (sub or "").strip() or "Unassigned",
                "priority": prio, "qc_tester": qc,
                "developer": _strip_paren(bd or fd or "").strip() or "Unassigned",
                "current_status": st or "",
                "created_on": cr.isoformat() if cr else None,
                "eta": eta.isoformat() if eta else None,
                "qa_estimate_hours": qae, "qa_actual_hours": qaa,
            }

    def item(tid, when, qc=None, extra=None):
        m = meta.get(tid, {})
        d = {
            "ticket_id": tid, "title": m.get("title") or "", "module": m.get("module", "Unassigned"),
            "priority": m.get("priority") or "", "qc_tester": qc or m.get("qc_tester") or "Unassigned",
            "developer": m.get("developer", "Unassigned"), "current_status": m.get("current_status", ""),
            "created_on": m.get("created_on"), "eta": m.get("eta"),
            "qa_estimate_hours": m.get("qa_estimate_hours"), "qa_actual_hours": m.get("qa_actual_hours"),
            "fail_loops": cycle_count.get(tid, 0), "date": when.isoformat() if when else None,
        }
        if extra:
            d.update(extra)
        return d

    def in_month(d):
        return d and start_date <= d <= end_date

    # A QC entry is a RETEST if it re-enters QC after a prior entry, OR comes back from a rework/fail
    # stage (works immediately from `previous_status`, even before deep history accumulates).
    REWORK = {"QC Review Fail", "Code Review Failed", "Tested - Awaiting Fixes", "BIS Testing"}
    new_list, refix_list, bis_list, appr_list = [], [], [], []
    seen_bis, seen_appr = set(), set()
    qc_entries = defaultdict(int)

    for h in trans:
        ns, ps, tid = h.new_status, h.previous_status, h.ticket_id
        if ns in QA_TESTING_STATUSES and ps not in QA_TESTING_STATUSES:
            prior = qc_entries[tid]
            qc_entries[tid] += 1
            if in_month(h.changed_on):
                if prior > 0 or ps in REWORK:
                    loops = max(prior, cycle_count.get(tid, 0), 1)
                    refix_list.append(item(tid, h.changed_on, h.qc_tester,
                                           {"refix_count": loops, "from_status": ps or ""}))
                else:
                    new_list.append(item(tid, h.changed_on, h.qc_tester))
        if ns == "BIS Testing" and tid not in seen_bis and in_month(h.changed_on):
            seen_bis.add(tid)
            bis_list.append(item(tid, h.changed_on, h.qc_tester))
        if ns == "Approved for Live" and tid not in seen_appr and in_month(h.changed_on):
            seen_appr.add(tid)
            appr_list.append(item(tid, h.changed_on, h.qc_tester))

    closed_list = [item(t.ticket_id, t.closed_on, t.qc_tester) for t in closed]
    return {
        "new_to_qc": {"count": len(new_list), "tickets": new_list},
        "refix_to_qc": {"count": len(refix_list), "tickets": refix_list},
        "to_bis": {"count": len(bis_list), "tickets": bis_list},
        "approved_for_live": {"count": len(appr_list), "tickets": appr_list},
        "closed": {"count": len(closed_list), "tickets": closed_list},
    }


@app.get("/ticket-movement")
def get_ticket_movement(offset: int = Query(0, ge=0, le=240), trend: int = Query(12, ge=1, le=24)):
    """Monthly ticket movement for the Ticket Movement Calendar. Ended months are frozen so the
    lists don't change after the month closes (tickets counted up to 23:59 of the last day).
    Also returns a trailing `trend` series (default 12 months) of per-category counts for the
    month-to-month comparison chart."""
    db: Session = SessionLocal()
    try:
        today = datetime.now().date()
        CATS = ("new_to_qc", "refix_to_qc", "to_bis", "approved_for_live", "closed")

        def _month(off):
            s, e, lbl = get_period_range("month", off)
            ended = e.date() < today
            if ended:
                snap = db.query(TicketMovementSnapshot).filter_by(period_label=lbl).first()
                if snap and snap.frozen and snap.payload:
                    return lbl, s, e, ended, snap.payload
            d = _compute_ticket_movement(db, s, e)
            d["period"] = {"label": lbl, "start": s.isoformat(), "end": e.isoformat(), "frozen": ended}
            if ended:
                try:
                    db.add(TicketMovementSnapshot(period_label=lbl, payload=d, frozen=True))
                    db.commit()
                except Exception:
                    db.rollback()
            return lbl, s, e, ended, d

        _, _, _, _, data = _month(offset)
        series = []
        for k in range(trend - 1, -1, -1):
            lbl, _s, _e, _ended, d = _month(offset + k)
            row = {"label": lbl}
            for c in CATS:
                row[c] = (d.get(c) or {}).get("count", 0)
            series.append(row)
        data = dict(data)
        data["trend"] = series
        return data
    finally:
        db.close()


def _get_planning_timesheet_summary(db: Session, employee: Employee, weeks: int = 5) -> Optional[dict]:
    """Get plan vs actual summary for an employee (used in performance and ratings)."""
    try:
        team_upper = (employee.team or "").upper()
        if "DEV" in team_upper or team_upper == "DEVELOPMENT":
            team_lower = "dev"
            timesheet_team = "DEV"
        else:
            team_lower = "qa"
            timesheet_team = "QA"

        week_start, _ = get_planning_week_dates(date.today())
        total_planned = 0.0
        total_actual = 0.0

        for w in range(weeks):
            ws = week_start - timedelta(days=7 * w)
            we = ws + timedelta(days=4)
            alloc_rows = []
            if team_lower == "dev":
                pw = db.query(DevPlanningWeek).filter(DevPlanningWeek.week_start == ws).first()
                if pw:
                    alloc_rows = (
                        db.query(DevPlannedTask, DevPlannedAllocation)
                        .join(DevPlannedAllocation, DevPlannedAllocation.task_id == DevPlannedTask.id)
                        .filter(
                            DevPlannedTask.planning_week_id == pw.id,
                            DevPlannedTask.status == "active",
                            DevPlannedTask.employee_name == employee.name,
                            DevPlannedAllocation.allocation_date >= ws,
                            DevPlannedAllocation.allocation_date <= we,
                        )
                    ).all()
            else:
                pw = db.query(QAPlanningWeek).filter(QAPlanningWeek.week_start == ws).first()
                if pw:
                    alloc_rows = (
                        db.query(QAPlannedTask, QAPlannedAllocation)
                        .join(QAPlannedAllocation, QAPlannedAllocation.task_id == QAPlannedTask.id)
                        .filter(
                            QAPlannedTask.planning_week_id == pw.id,
                            QAPlannedTask.status == "active",
                            QAPlannedTask.employee_name == employee.name,
                            QAPlannedAllocation.allocation_date >= ws,
                            QAPlannedAllocation.allocation_date <= we,
                        )
                    ).all()

            planned_hours = sum(float(a[1].hours or 0) for a in alloc_rows)
            actual_entries = db.query(EnhancedTimesheet).filter(
                EnhancedTimesheet.employee_name == employee.name,
                EnhancedTimesheet.team == timesheet_team,
                EnhancedTimesheet.date >= ws,
                EnhancedTimesheet.date <= we,
            ).all()
            actual_hours = sum(float(e.hours_logged or 0) for e in actual_entries)
            total_planned += planned_hours
            total_actual += actual_hours

        if total_planned == 0 and total_actual == 0:
            return None
        overall_variance = round(total_actual - total_planned, 2)
        overall_variance_pct = round((overall_variance / total_planned * 100), 1) if total_planned > 0 else 0
        overall_accuracy = round(100 - abs(overall_variance_pct), 1) if total_planned > 0 else None
        return {
            "total_planned_hours": round(total_planned, 2),
            "total_actual_hours": round(total_actual, 2),
            "total_variance": overall_variance,
            "variance_percent": overall_variance_pct,
            "estimation_accuracy": overall_accuracy,
        }
    except Exception:
        return None


def calculate_rag_score(metrics, is_dev, planning_timesheet: Optional[dict] = None, role_context: Optional[dict] = None):
    """Calculate RAG score based on metrics"""
    role_context = role_context or {}
    if role_context.get("is_manager"):
        score = 0
        weights_used = 0
        timesheet = metrics.get("timesheet", {})

        # Utilization (40%)
        if timesheet.get("expected_hours", 0) > 0:
            utilization = min(100, timesheet.get("utilization_percent", 0))
            score += (utilization / 100) * 40
            weights_used += 40

        # Plan vs Actual accuracy (40%)
        if planning_timesheet and planning_timesheet.get("estimation_accuracy") is not None:
            acc = planning_timesheet.get("estimation_accuracy", 0)
            score += (acc / 100) * 40
            weights_used += 40

        # Avg daily hours (20%)
        if timesheet.get("avg_daily_hours") is not None:
            daily = min(100, (timesheet.get("avg_daily_hours", 0) / 8) * 100)
            score += (daily / 100) * 20
            weights_used += 20

        base_score = (score / weights_used * 100) if weights_used > 0 else 0
        return round(base_score, 1)
    score = 0
    weights_used = 0
    
    bugs = metrics.get("bugs", {})
    timesheet = metrics.get("timesheet", {})
    tickets = metrics.get("tickets", {})
    
    if is_dev:
        # Closure rate (25%)
        if bugs.get("total", 0) > 0:
            closure_rate = bugs.get("closure_rate", 0)
            score += (closure_rate / 100) * 25
            weights_used += 25
        
        # Re-opened % inverse (20%)
        if bugs.get("total", 0) > 0:
            reopened_pct = bugs.get("reopened_percent", 0)
            reopened_score = max(0, 100 - (reopened_pct * 5))  # Penalize heavily
            score += (reopened_score / 100) * 20
            weights_used += 20
        
        # Estimate accuracy (20%) - prefer planning/timesheet accuracy if available
        if planning_timesheet and planning_timesheet.get("estimation_accuracy") is not None:
            accuracy = planning_timesheet.get("estimation_accuracy", 100)
            accuracy_score = 100 - abs(100 - accuracy)
            score += max(0, accuracy_score / 100) * 20
            weights_used += 20
        elif tickets.get("actual_hours", 0) > 0:
            accuracy = tickets.get("estimate_accuracy", 100)
            accuracy_score = 100 - abs(100 - accuracy)  # Closer to 100% is better
            score += max(0, accuracy_score / 100) * 20
            weights_used += 20
        
        # Utilization (20%)
        if timesheet.get("expected_hours", 0) > 0:
            utilization = min(100, timesheet.get("utilization_percent", 0))
            score += (utilization / 100) * 20
            weights_used += 20
        
        # Resolution time (15%) - lower is better
        if bugs.get("avg_resolution_days", 0) > 0:
            res_time = bugs.get("avg_resolution_days", 0)
            res_score = max(0, 100 - (res_time * 2))  # 50 days = 0 score
            score += (res_score / 100) * 15
            weights_used += 15
    else:  # QA
        tests = metrics.get("tests", {})
        
        # Pass rate (20%)
        if tests.get("total_executed", 0) > 0:
            pass_rate = tests.get("pass_rate", 0)
            score += (pass_rate / 100) * 20
            weights_used += 20
        
        # Bugs per ticket (25%) - higher is generally better for QA
        bugs_per_ticket = metrics.get("bugs_per_ticket", 0)
        if bugs_per_ticket > 0:
            bpt_score = min(100, bugs_per_ticket * 20)  # 5+ bugs/ticket = 100
            score += (bpt_score / 100) * 25
            weights_used += 25
        
        # Rejected % inverse (15%)
        if bugs.get("total", 0) > 0:
            rejected_pct = bugs.get("rejected_percent", 0)
            rejected_score = max(0, 100 - (rejected_pct * 5))
            score += (rejected_score / 100) * 15
            weights_used += 15
        
        # Utilization (20%)
        if timesheet.get("expected_hours", 0) > 0:
            utilization = min(100, timesheet.get("utilization_percent", 0))
            score += (utilization / 100) * 20
            weights_used += 20
        
        # Critical bugs found (20%) - higher is better for QA
        if bugs.get("total", 0) > 0:
            critical_pct = bugs.get("severity", {}).get("critical_percent", 0)
            critical_score = min(100, critical_pct * 5)  # Finding critical bugs is good
            score += (critical_score / 100) * 20
            weights_used += 20

    # Plan vs Actual bonus (up to 5 points) - estimation accuracy from planning/timesheet
    if planning_timesheet and planning_timesheet.get("estimation_accuracy") is not None:
        acc = planning_timesheet["estimation_accuracy"]
        bonus = min(5, max(0, acc / 20))  # 100% accuracy = 5 pts, 0% = 0 pts
        base_score = (score / weights_used * 100) if weights_used > 0 else 0
        score = min(100, base_score + bonus)
        return round(score, 1)
    
    # Normalize to 100 if not all weights were used
    if weights_used > 0:
        score = (score / weights_used) * 100
    
    return round(score, 1)


@app.get("/employees/{employee_id}/timesheet-summary")
def get_employee_timesheet_summary(
    employee_id: str,
    period: str = Query("past_month"),
    current_user: dict = Depends(get_current_user),
):
    """Get detailed timesheet summary for an employee"""
    db: Session = SessionLocal()
    try:
        employee = db.query(Employee).filter(
            or_(
                Employee.employee_id == employee_id,
                Employee.id == int(employee_id) if employee_id.isdigit() else False
            )
        ).first()
        if not employee:
            raise HTTPException(status_code=404, detail="Employee not found")
        if not can_access_employee(db, current_user, employee.employee_id):
            raise HTTPException(status_code=403, detail="Access denied")

        start_date, end_date = get_date_range(period)
        
        query = db.query(Timesheet).filter(
            Timesheet.employee_name.ilike(f"%{employee.name}%")
        )
        
        if start_date:
            query = query.filter(Timesheet.date >= start_date.date())
        
        timesheets = query.order_by(Timesheet.date.desc()).all()
        
        # Daily breakdown
        daily_data = defaultdict(int)
        ticket_hours = defaultdict(int)
        
        for ts in timesheets:
            day_key = ts.date.isoformat() if ts.date else "unknown"
            daily_data[day_key] += ts.time_logged_minutes or 0
            if ts.ticket_id:
                ticket_hours[ts.ticket_id] += ts.time_logged_minutes or 0
        
        # Convert to hours
        daily_hours = {k: round(v / 60, 2) for k, v in daily_data.items()}
        ticket_hours_formatted = {k: round(v / 60, 2) for k, v in ticket_hours.items()}
        
        total_minutes = sum(daily_data.values())
        
        return {
            "employee_name": employee.name,
            "period": period,
            "total_hours": round(total_minutes / 60, 1),
            "total_entries": len(timesheets),
            "unique_tickets": len(ticket_hours),
            "daily_hours": dict(sorted(daily_hours.items(), reverse=True)[:30]),
            "ticket_hours": dict(sorted(ticket_hours_formatted.items(), key=lambda x: x[1], reverse=True)[:20])
        }
    finally:
        db.close()


@app.get("/employees/{employee_id}/planning-timesheet")
def get_employee_planning_timesheet(
    employee_id: str,
    weeks: int = Query(5, description="Number of weeks to include (current + past)"),
    current_user: dict = Depends(get_current_user),
):
    """
    Get plan vs actual, timesheet, and planning data for an employee.
    """
    db = SessionLocal()
    try:
        employee = db.query(Employee).filter(
            or_(
                Employee.employee_id == employee_id,
                Employee.id == int(employee_id) if employee_id.isdigit() else False
            )
        ).first()
        if not employee:
            raise HTTPException(status_code=404, detail="Employee not found")
        if not can_access_employee(db, current_user, employee.employee_id):
            raise HTTPException(status_code=403, detail="Access denied")

        # Map Employee.team to planning team (dev/qa)
        team_upper = (employee.team or "").upper()
        if "DEV" in team_upper or team_upper == "DEVELOPMENT":
            team_lower = "dev"
            timesheet_team = "DEV"
        else:
            team_lower = "qa"
            timesheet_team = "QA"

        week_start, _ = get_planning_week_dates(date.today())
        weekly_summaries = []
        all_planned_tasks = []
        all_actual_entries = []

        for w in range(weeks):
            ws = week_start - timedelta(days=7 * w)
            we = ws + timedelta(days=4)

            # Planned: Dev or QA
            alloc_rows = []
            if team_lower == "dev":
                pw = db.query(DevPlanningWeek).filter(DevPlanningWeek.week_start == ws).first()
                if pw:
                    alloc_rows = (
                        db.query(DevPlannedTask, DevPlannedAllocation)
                        .join(DevPlannedAllocation, DevPlannedAllocation.task_id == DevPlannedTask.id)
                        .filter(
                            DevPlannedTask.planning_week_id == pw.id,
                            DevPlannedTask.status == "active",
                            DevPlannedTask.employee_name == employee.name,
                            DevPlannedAllocation.allocation_date >= ws,
                            DevPlannedAllocation.allocation_date <= we,
                        )
                    ).all()
                    for row in alloc_rows:
                        task, alloc = row[0], row[1]
                        h = round(float(alloc.hours or 0), 2)
                        all_planned_tasks.append({
                            "week_start": ws.isoformat(),
                            "date": alloc.allocation_date.isoformat(),
                            "ticket_id": task.ticket_id,
                            "ticket_title": getattr(task, "ticket_title", None) or "",
                            "activity_description": getattr(task, "activity_description", None) or "",
                            "generic_category": getattr(task, "generic_category", None),
                            "hours": h,
                        })
            else:
                pw = db.query(QAPlanningWeek).filter(QAPlanningWeek.week_start == ws).first()
                if pw:
                    alloc_rows = (
                        db.query(QAPlannedTask, QAPlannedAllocation)
                        .join(QAPlannedAllocation, QAPlannedAllocation.task_id == QAPlannedTask.id)
                        .filter(
                            QAPlannedTask.planning_week_id == pw.id,
                            QAPlannedTask.status == "active",
                            QAPlannedTask.employee_name == employee.name,
                            QAPlannedAllocation.allocation_date >= ws,
                            QAPlannedAllocation.allocation_date <= we,
                        )
                    ).all()
                    for row in alloc_rows:
                        task, alloc = row[0], row[1]
                        h = round(float(alloc.hours or 0), 2)
                        all_planned_tasks.append({
                            "week_start": ws.isoformat(),
                            "date": alloc.allocation_date.isoformat(),
                            "ticket_id": task.ticket_id,
                            "ticket_title": getattr(task, "ticket_title", None) or "",
                            "activity_description": getattr(task, "activity_description", None) or "",
                            "generic_category": getattr(task, "generic_category", None),
                            "hours": h,
                        })
                else:
                    alloc_rows = []

            planned_hours = round(sum(float(a[1].hours or 0) for a in alloc_rows), 2)

            # Actual: EnhancedTimesheet
            actual_entries = db.query(EnhancedTimesheet).filter(
                EnhancedTimesheet.employee_name == employee.name,
                EnhancedTimesheet.team == timesheet_team,
                EnhancedTimesheet.date >= ws,
                EnhancedTimesheet.date <= we,
            ).order_by(EnhancedTimesheet.date.desc()).all()

            actual_hours = round(sum(float(e.hours_logged or 0) for e in actual_entries), 2)
            for e in actual_entries:
                all_actual_entries.append({
                    "date": e.date.isoformat(),
                    "ticket_id": e.ticket_id,
                    "task_description": e.task_description or "",
                    "project_name": e.project_name or "",
                    "hours": round(float(e.hours_logged or 0), 2),
                })

            variance = round(actual_hours - planned_hours, 2)
            variance_pct = round((variance / planned_hours * 100), 1) if planned_hours > 0 else (0 if actual_hours == 0 else None)
            estimation_accuracy = round(100 - abs(variance_pct), 1) if planned_hours > 0 and variance_pct is not None else None

            weekly_summaries.append({
                "week_start": ws.isoformat(),
                "week_end": we.isoformat(),
                "planned_hours": planned_hours,
                "actual_hours": actual_hours,
                "variance": variance,
                "variance_percent": variance_pct,
                "estimation_accuracy": estimation_accuracy,
                "planned_task_count": len(alloc_rows),
                "actual_entry_count": len(actual_entries),
            })

        # Aggregate summary
        total_planned = sum(s["planned_hours"] for s in weekly_summaries)
        total_actual = sum(s["actual_hours"] for s in weekly_summaries)
        overall_variance = round(total_actual - total_planned, 2)
        overall_variance_pct = round((overall_variance / total_planned * 100), 1) if total_planned > 0 else 0
        overall_accuracy = round(100 - abs(overall_variance_pct), 1) if total_planned > 0 else None

        return {
            "employee_id": employee.employee_id,
            "employee_name": employee.name,
            "team": team_lower,
            "summary": {
                "total_planned_hours": round(total_planned, 2),
                "total_actual_hours": round(total_actual, 2),
                "total_variance": overall_variance,
                "variance_percent": overall_variance_pct,
                "estimation_accuracy": overall_accuracy,
                "total_planned_tasks": len(all_planned_tasks),
                "total_actual_entries": len(all_actual_entries),
                "unique_tickets_worked": len(set(e["ticket_id"] for e in all_actual_entries if e.get("ticket_id"))),
            },
            "weekly_summaries": weekly_summaries,
            "recent_planned_tasks": sorted(all_planned_tasks, key=lambda x: (x["date"], x.get("ticket_id") or 0), reverse=True)[:30],
            "recent_timesheet_entries": sorted(all_actual_entries, key=lambda x: x["date"], reverse=True)[:50],
            "planning_module_link": f"/planning?module=comparison",
            "planning_employee_filter": employee.name,
        }
    finally:
        db.close()


@app.get("/employees/{employee_id}/times-failed")
def get_employee_times_failed(
    employee_id: str,
    period: str = Query("past_quarter", description="past_week, past_month, past_quarter, one_year, overall"),
    start_date: Optional[str] = Query(None, description="Start date YYYY-MM-DD (optional, for custom range)"),
    end_date: Optional[str] = Query(None, description="End date YYYY-MM-DD (optional)"),
    current_user: dict = Depends(get_current_user),
):
    """
    Get count of times tickets (assigned to this resource as backend/frontend developer)
    were moved to QC Review Fail (or Tested - Awaiting Fixes, Code Review Failed) in the selected period.
    """
    db: Session = SessionLocal()
    try:
        employee = db.query(Employee).filter(
            or_(
                Employee.employee_id == employee_id,
                Employee.id == int(employee_id) if employee_id.isdigit() else False
            )
        ).first()
        if not employee:
            raise HTTPException(status_code=404, detail="Employee not found")
        if not can_access_employee_profile(db, current_user, employee.employee_id):
            raise HTTPException(status_code=403, detail="Access denied to this employee profile")

        if start_date and end_date:
            try:
                range_start = datetime.strptime(start_date, "%Y-%m-%d").date()
                range_end = datetime.strptime(end_date, "%Y-%m-%d").date()
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid start_date or end_date; use YYYY-MM-DD")
        else:
            range_start, range_end = get_date_range(period)
            if range_start is None:
                range_start = datetime.now() - timedelta(days=365 * 2)
            if range_end is None:
                range_end = datetime.now()

        start_dt = range_start if isinstance(range_start, datetime) else datetime.combine(range_start, datetime.min.time())
        end_dt = range_end if isinstance(range_end, datetime) else datetime.combine(range_end, datetime.max.time())
        range_start_date = start_dt.date() if hasattr(start_dt, 'date') else range_start
        range_end_date = end_dt.date() if hasattr(end_dt, 'date') else range_end

        employee_name = (employee.name or "").strip()
        ticket_query = db.query(TicketTracking.ticket_id).filter(
            or_(
                TicketTracking.backend_developer.ilike(f"%{employee_name}%"),
                TicketTracking.frontend_developer.ilike(f"%{employee_name}%")
            )
        ).distinct()
        ticket_ids = [r[0] for r in ticket_query.all()]

        if not ticket_ids:
            return {
                "employee_id": employee.employee_id,
                "employee_name": employee.name,
                "period": period,
                "start_date": range_start_date.isoformat() if hasattr(range_start_date, 'isoformat') else str(range_start_date),
                "end_date": range_end_date.isoformat() if hasattr(range_end_date, 'isoformat') else str(range_end_date),
                "total_times_failed": 0,
                "by_ticket": [],
            }

        from sqlalchemy import func
        rows = (
            db.query(TicketStatusHistory.ticket_id, func.count(TicketStatusHistory.id).label("cnt"))
            .filter(
                TicketStatusHistory.ticket_id.in_(ticket_ids),
                TicketStatusHistory.new_status.in_(QC_FAIL_STATUSES),
                TicketStatusHistory.changed_on >= start_dt,
                TicketStatusHistory.changed_on <= end_dt,
            )
            .group_by(TicketStatusHistory.ticket_id)
            .all()
        )
        by_ticket = [{"ticket_id": r.ticket_id, "times_failed_in_period": r.cnt} for r in rows]
        total_times_failed = sum(r.cnt for r in rows)

        return {
            "employee_id": employee.employee_id,
            "employee_name": employee.name,
            "period": period,
            "start_date": range_start_date.isoformat() if hasattr(range_start_date, 'isoformat') else str(range_start_date),
            "end_date": range_end_date.isoformat() if hasattr(range_end_date, 'isoformat') else str(range_end_date),
            "total_times_failed": total_times_failed,
            "by_ticket": by_ticket,
        }
    finally:
        db.close()


@app.get("/my-tasks")
def get_my_tasks(
    view: str = Query("week", description="week | month | all"),
    date_str: Optional[str] = Query(None, description="Reference date YYYY-MM-DD. Default: today. Ignored when view=all or custom range"),
    start_date_str: Optional[str] = Query(None, description="Start date YYYY-MM-DD for custom range"),
    end_date_str: Optional[str] = Query(None, description="End date YYYY-MM-DD for custom range"),
    current_user: dict = Depends(get_current_user),
):
    """
    Get current user's tasks: ongoing, future assigned, and completed (past).
    Supports weekly, monthly, all-data views, or custom date range via start_date_str/end_date_str.
    """
    db: Session = SessionLocal()
    try:
        employee_id = current_user.get("employee_id")
        if not employee_id:
            raise HTTPException(status_code=403, detail="Employee account required")
        employee = db.query(Employee).filter(Employee.employee_id == employee_id).first()
        if not employee:
            raise HTTPException(status_code=404, detail="Employee not found")

        ref = datetime.strptime(date_str or date.today().isoformat(), "%Y-%m-%d").date()
        team_upper = (employee.team or "").upper()
        is_dev = "DEV" in team_upper or team_upper == "DEVELOPMENT"
        timesheet_team = "DEV" if is_dev else "QA"

        # Custom date range takes precedence
        if start_date_str and end_date_str:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
            if start_date > end_date:
                start_date, end_date = end_date, start_date
        elif view == "all":
            start_date = None
            end_date = None
        elif view == "month":
            start_date = ref.replace(day=1)
            if ref.month == 12:
                end_date = date(ref.year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(ref.year, ref.month + 1, 1) - timedelta(days=1)
        else:
            # Week: Monday to Friday
            start_date = ref - timedelta(days=ref.weekday())
            end_date = start_date + timedelta(days=4)

        today = date.today()
        ongoing_tasks = []
        future_tasks = []
        completed_planned = []
        completed_work = []

        if is_dev:
            tasks_filter = [
                DevPlannedTask.employee_name == employee.name,
                DevPlannedTask.status == "active",
            ]
            if start_date is not None and end_date is not None:
                tasks_filter.append(DevPlannedTask.start_date <= end_date)
                tasks_filter.append(or_(
                    DevPlannedTask.end_date.is_(None),
                    DevPlannedTask.end_date >= start_date,
                ))
            tasks_query = (
                db.query(DevPlannedTask)
                .filter(*tasks_filter)
                .order_by(DevPlannedTask.start_date)
            )
            dev_tasks_list = tasks_query.all()
            dev_ticket_ids = [t.ticket_id for t in dev_tasks_list if t.ticket_id]
            ticket_priority_map = {}
            if dev_ticket_ids:
                tkts = db.query(TicketTracking.ticket_id, TicketTracking.priority).filter(TicketTracking.ticket_id.in_(dev_ticket_ids)).all()
                ticket_priority_map = {tk.ticket_id: tk.priority for tk in tkts if tk.priority}
            for t in dev_tasks_list:
                task_end = t.end_date or t.start_date
                task_start = t.start_date
                item = {
                    "id": t.id,
                    "type": "dev",
                    "ticket_id": t.ticket_id,
                    "ticket_title": t.ticket_title,
                    "ticket_priority": ticket_priority_map.get(t.ticket_id) if t.ticket_id else None,
                    "generic_category": t.generic_category,
                    "activity_description": t.activity_description or "",
                    "start_date": task_start.isoformat(),
                    "end_date": task_end.isoformat() if task_end else None,
                    "total_hours": round(float(t.total_planned_hours or 0), 1),
                    "created_by": t.created_by,
                }
                if task_end < today:
                    completed_planned.append(item)
                elif task_start > today:
                    future_tasks.append(item)
                else:
                    ongoing_tasks.append(item)
        else:
            qa_filter = [
                QAPlannedTask.employee_name == employee.name,
                QAPlannedTask.status == "active",
            ]
            if start_date is not None and end_date is not None:
                qa_filter.append(QAPlannedTask.start_date <= end_date)
                qa_filter.append(or_(
                    QAPlannedTask.end_date.is_(None),
                    QAPlannedTask.end_date >= start_date,
                ))
            tasks_query = (
                db.query(QAPlannedTask)
                .filter(*qa_filter)
                .order_by(QAPlannedTask.start_date)
            )
            for t in tasks_query.all():
                task_end = t.end_date or t.start_date
                task_start = t.start_date
                item = {
                    "id": t.id,
                    "type": "qa",
                    "ticket_id": t.ticket_id,
                    "ticket_title": t.ticket_title,
                    "ticket_priority": t.ticket_priority,
                    "generic_category": t.generic_category,
                    "task_type": getattr(t, "task_type", None),
                    "activity_description": t.activity_description or "",
                    "start_date": task_start.isoformat(),
                    "end_date": task_end.isoformat() if task_end else None,
                    "total_hours": round(float(t.total_planned_hours or 0), 1),
                    "created_by": t.created_by,
                }
                if task_end < today:
                    completed_planned.append(item)
                elif task_start > today:
                    future_tasks.append(item)
                else:
                    ongoing_tasks.append(item)

        # Completed work from timesheet (actual logged hours)
        timesheet_filter = [
            EnhancedTimesheet.employee_name == employee.name,
            EnhancedTimesheet.team == timesheet_team,
        ]
        if start_date is not None and end_date is not None:
            timesheet_filter.append(EnhancedTimesheet.date >= start_date)
            timesheet_filter.append(EnhancedTimesheet.date <= end_date)
        timesheet_entries = (
            db.query(EnhancedTimesheet)
            .filter(*timesheet_filter)
            .order_by(EnhancedTimesheet.date.desc())
            .all()
        )
        for e in timesheet_entries:
            completed_work.append({
                "date": e.date.isoformat(),
                "ticket_id": e.ticket_id,
                "task_description": e.task_description or "",
                "project_name": e.project_name or "",
                "hours": round(float(e.hours_logged or 0), 2),
                "leave_type": e.leave_type,
            })

        return {
            "employee_id": employee.employee_id,
            "employee_name": employee.name,
            "team": "dev" if is_dev else "qa",
            "view": view,
            "start_date": start_date.isoformat() if start_date else None,
            "end_date": end_date.isoformat() if end_date else None,
            "ongoing_tasks": ongoing_tasks,
            "future_tasks": future_tasks,
            "completed_planned": completed_planned,
            "completed_work": completed_work,
        }
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    finally:
        db.close()


def _get_reportee_names_for_user(db: Session, current_user: dict) -> Optional[Tuple[List[str], int]]:
    """Return (list of reportee names, count) for lead/manager, or (None, 0) if no reportees."""
    role = current_user.get("role", "")
    if role != "ADMIN" and "MANAGER" not in role and "LEAD" not in role:
        return None, 0
    employee_id = current_user.get("employee_id")
    if not employee_id:
        return None, 0
    employee = db.query(Employee).filter(Employee.employee_id == employee_id).first()
    if not employee:
        return None, 0
    names = set()
    # Direct reportees (lead)
    direct = db.query(Employee).filter(
        Employee.lead.ilike(f"%{employee.name}%"),
        Employee.is_active == True,
        Employee.employee_id != employee_id,
    ).all()
    names.update((e.name for e in direct if e.name))
    # For manager: indirect reportees
    if "MANAGER" in role or role == "ADMIN":
        indirect = db.query(Employee).filter(
            Employee.manager.ilike(f"%{employee.name}%"),
            Employee.is_active == True,
            Employee.employee_id != employee_id,
        ).all()
        names.update((e.name for e in indirect if e.name))
    lst = list(names) if names else None
    return lst, len(lst) if lst else 0


@app.get("/my-tasks/team/check")
def get_my_tasks_team_check(current_user: dict = Depends(get_current_user)):
    """Check if current user can see My Team tab. True for LEADs and MANAGERs with reportees."""
    db: Session = SessionLocal()
    try:
        _, count = _get_reportee_names_for_user(db, current_user)
        return {"has_reportees": count > 0, "reportee_count": count}
    finally:
        db.close()


@app.get("/my-tasks/team")
def get_my_tasks_team(
    view: str = Query("week", description="week | month | all"),
    date_str: Optional[str] = Query(None, description="Reference date YYYY-MM-DD"),
    current_user: dict = Depends(get_current_user),
):
    """
    Get team metrics for leads and managers with reportees (direct and indirect for managers).
    Returns ticket counts by status for the period, current status, and per-member activity.
    """
    db: Session = SessionLocal()
    try:
        reportee_names, reportee_count = _get_reportee_names_for_user(db, current_user)
        if not reportee_names or reportee_count == 0:
            raise HTTPException(status_code=403, detail="My Team is available only for users with reportees")

        ref = datetime.strptime(date_str or date.today().isoformat(), "%Y-%m-%d").date()
        if view == "all":
            start_date = None
            end_date = None
        elif view == "month":
            start_date = ref.replace(day=1)
            end_date = date(ref.year, ref.month + 1, 1) - timedelta(days=1) if ref.month < 12 else date(ref.year + 1, 1, 1) - timedelta(days=1)
        else:
            start_date = ref - timedelta(days=ref.weekday())
            end_date = start_date + timedelta(days=4)

        reportee_lower = {n.strip().lower() for n in reportee_names if n}

        def _assigned_to_team(ticket, field: str) -> bool:
            val = (getattr(ticket, field, None) or "").strip()
            return val and val.lower() in reportee_lower

        def _qa_assigned(ticket) -> bool:
            return _assigned_to_team(ticket, "qc_tester")

        def _dev_assigned(ticket) -> bool:
            return _assigned_to_team(ticket, "backend_developer") or _assigned_to_team(ticket, "frontend_developer") or _assigned_to_team(ticket, "current_assignee")

        QA_PENDING = "QC Testing"
        QA_IN_PROGRESS = "QC Testing in Progress"
        QA_HOLD = "QC Testing Hold"
        BIS_STATUSES = ["BIS Testing", "Testing In Progress"]
        CLOSED_STATUSES = ["Closed", "Moved to Live", "Completed"]

        all_tickets = db.query(TicketTracking).all()

        # Get latest status change for each ticket to calculate time in current status
        # Query: get most recent status change per ticket
        from sqlalchemy import func as sqlfunc
        subq_latest = db.query(
            TicketStatusHistory.ticket_id,
            sqlfunc.max(TicketStatusHistory.changed_on).label("last_change")
        ).group_by(TicketStatusHistory.ticket_id).subquery()
        
        latest_status_changes = db.query(TicketStatusHistory).join(
            subq_latest,
            (TicketStatusHistory.ticket_id == subq_latest.c.ticket_id) &
            (TicketStatusHistory.changed_on == subq_latest.c.last_change)
        ).all()
        
        status_change_map = {h.ticket_id: h for h in latest_status_changes}
        today_dt = datetime.now()

        def _calc_status_duration(ticket_id):
            """Calculate days in current status."""
            hist = status_change_map.get(ticket_id)
            if hist and hist.changed_on:
                delta = today_dt - hist.changed_on
                return round(delta.total_seconds() / 86400, 1)  # Days
            return None

        def _calc_time_status(t, is_qa=True):
            """Calculate time tracking status: on_track, exceeded, or at_risk."""
            if is_qa:
                est = t.qa_estimate_hours or 0
                actual = t.actual_qa_hours or 0
            else:
                est = t.dev_estimate_hours or 0
                actual = t.actual_dev_hours or 0
            if est <= 0:
                return None  # No estimate
            ratio = actual / est if est > 0 else 0
            if ratio <= 0.8:
                return "on_track"
            elif ratio <= 1.0:
                return "at_risk"
            else:
                return "exceeded"

        def _build_ticket_info(t, assignee_field, is_qa=True):
            """Build enriched ticket info with priority, ETA, time tracking."""
            eta = getattr(t, "eta", None)
            eta_str = eta.strftime("%Y-%m-%d") if eta and hasattr(eta, "strftime") else (str(eta)[:10] if eta else None)
            eta_status = None
            if eta:
                eta_date = eta.date() if hasattr(eta, "date") else eta
                if isinstance(eta_date, date):
                    days_to_eta = (eta_date - date.today()).days
                    if days_to_eta < 0:
                        eta_status = "overdue"
                    elif days_to_eta <= 2:
                        eta_status = "due_soon"
                    else:
                        eta_status = "on_track"
            est_hours = t.qa_estimate_hours if is_qa else t.dev_estimate_hours
            actual_hours = t.actual_qa_hours if is_qa else t.actual_dev_hours
            return {
                "ticket_id": t.ticket_id,
                "title": (t.title or "")[:80],
                "priority": (t.priority or "").strip() or None,
                "eta": eta_str,
                "eta_status": eta_status,
                "estimate_hours": round(est_hours, 1) if est_hours else None,
                "actual_hours": round(actual_hours, 1) if actual_hours else None,
                "time_status": _calc_time_status(t, is_qa),
                "days_in_status": _calc_status_duration(t.ticket_id),
                "status": (t.status or "").strip(),
                assignee_field: getattr(t, assignee_field, None) or (t.current_assignee if not is_qa else t.qc_tester),
            }

        # Period metrics (for selected time range)
        period_qa = {"completed": 0, "in_progress": 0, "on_hold": 0, "moved_to_bis": 0}
        period_dev = {"completed": 0, "in_progress": 0, "ready_for_qc": 0}

        # Current status (regardless of period) - now with enriched info
        current_qa = {"pending": [], "in_progress": [], "on_hold": [], "bis_testing": []}
        current_dev = {"in_progress": [], "ready_for_qc": [], "other": []}

        # Moved to BIS in period - from status history
        bis_in_period_ticket_ids = set()
        if start_date and end_date:
            bis_changes = db.query(TicketStatusHistory).filter(
                TicketStatusHistory.new_status.in_(BIS_STATUSES),
                TicketStatusHistory.changed_on >= datetime.combine(start_date, datetime.min.time()),
                TicketStatusHistory.changed_on <= datetime.combine(end_date, datetime.max.time()),
            ).all()
            bis_in_period_ticket_ids = {h.ticket_id for h in bis_changes}

        for t in all_tickets:
            status = (t.status or "").strip()
            is_closed = status in CLOSED_STATUSES or (status and status.lower() in ["closed", "moved to live", "completed"])
            closed_on = getattr(t, "closed_on", None)

            # QA tickets
            if _qa_assigned(t):
                if is_closed and start_date and end_date and closed_on:
                    cd = closed_on.date() if hasattr(closed_on, "date") else closed_on
                    if start_date <= cd <= end_date:
                        period_qa["completed"] += 1
                elif status == QA_IN_PROGRESS:
                    period_qa["in_progress"] += 1
                    if not is_closed:
                        current_qa["in_progress"].append(_build_ticket_info(t, "qc_tester", is_qa=True))
                elif status == QA_HOLD:
                    period_qa["on_hold"] += 1
                    if not is_closed:
                        current_qa["on_hold"].append(_build_ticket_info(t, "qc_tester", is_qa=True))
                elif status == QA_PENDING:
                    if not is_closed:
                        current_qa["pending"].append(_build_ticket_info(t, "qc_tester", is_qa=True))
                elif status in BIS_STATUSES:
                    if t.ticket_id in bis_in_period_ticket_ids:
                        period_qa["moved_to_bis"] += 1
                    if not is_closed:
                        current_qa["bis_testing"].append(_build_ticket_info(t, "qc_tester", is_qa=True))

            # DEV tickets
            if _dev_assigned(t):
                if is_closed and start_date and end_date and closed_on:
                    cd = closed_on.date() if hasattr(closed_on, "date") else closed_on
                    if start_date <= cd <= end_date:
                        period_dev["completed"] += 1
                elif status == "In Progress":
                    period_dev["in_progress"] += 1
                    if not is_closed:
                        current_dev["in_progress"].append(_build_ticket_info(t, "assignee", is_qa=False))
                elif status in ["Code Review Passed", "Approved for Live"]:
                    period_dev["ready_for_qc"] += 1
                    if not is_closed:
                        current_dev["ready_for_qc"].append(_build_ticket_info(t, "assignee", is_qa=False))
                elif status not in CLOSED_STATUSES and status not in BIS_STATUSES:
                    if not is_closed and status in ["Ready For Development", "Technical Review", "Start Code Review", "Code Review Failed", "QC Review Fail", "Tested - Awaiting Fixes"]:
                        current_dev["other"].append(_build_ticket_info(t, "assignee", is_qa=False))

        # Planned tasks for reportees (current week)
        from dev_planning import get_planning_week_dates
        today = date.today()
        week_start, week_end = get_planning_week_dates(today)
        qa_planned = []
        dev_planned = []
        pw_qa = get_qa_planning_week(week_start, db)
        pw_dev = get_planning_week(week_start, db)
        
        # Get ticket info for planned tasks to include priority
        def _get_ticket_priority(ticket_id):
            if not ticket_id:
                return None
            for t in all_tickets:
                if t.ticket_id == ticket_id:
                    return (t.priority or "").strip() or None
            return None
        
        def _get_ticket_eta(ticket_id):
            if not ticket_id:
                return None
            for t in all_tickets:
                if t.ticket_id == ticket_id:
                    eta = t.eta
                    if eta and hasattr(eta, "strftime"):
                        return eta.strftime("%Y-%m-%d")
                    elif eta:
                        return str(eta)[:10]
            return None
        
        if pw_qa:
            qa_tasks = db.query(QAPlannedTask).filter(
                QAPlannedTask.planning_week_id == pw_qa.id,
                QAPlannedTask.status == "active",
                QAPlannedTask.employee_name.in_(reportee_names),
            ).all()
            qa_planned = [{
                "employee_name": t.employee_name,
                "ticket_id": t.ticket_id,
                "activity_description": (t.activity_description or "")[:100],
                "total_hours": float(t.total_planned_hours or 0),
                "priority": t.ticket_priority or _get_ticket_priority(t.ticket_id),
                "start_date": t.start_date.isoformat() if t.start_date else None,
                "end_date": t.end_date.isoformat() if t.end_date else None,
                "eta": _get_ticket_eta(t.ticket_id),
                "generic_category": t.generic_category,
            } for t in qa_tasks]
        if pw_dev:
            dev_tasks = db.query(DevPlannedTask).filter(
                DevPlannedTask.planning_week_id == pw_dev.id,
                DevPlannedTask.status == "active",
                DevPlannedTask.employee_name.in_(reportee_names),
            ).all()
            dev_planned = [{
                "employee_name": t.employee_name,
                "ticket_id": t.ticket_id,
                "activity_description": (t.activity_description or "")[:100],
                "total_hours": float(t.total_planned_hours or 0),
                "priority": _get_ticket_priority(t.ticket_id),
                "start_date": t.start_date.isoformat() if t.start_date else None,
                "end_date": t.end_date.isoformat() if t.end_date else None,
                "eta": _get_ticket_eta(t.ticket_id),
                "generic_category": t.generic_category,
            } for t in dev_tasks]

        # Per-member activity (for lead to view each team member)
        employee = db.query(Employee).filter(Employee.employee_id == current_user.get("employee_id")).first()
        direct_reportees = db.query(Employee).filter(
            Employee.lead.ilike(f"%{employee.name}%"),
            Employee.is_active == True,
            Employee.employee_id != employee.employee_id,
        ).order_by(Employee.name).all() if employee else []
        member_activity = []
        for emp in direct_reportees:
            name = emp.name
            if not name:
                continue
            name_lower = name.strip().lower()
            qa_tickets = [t for t in (current_qa.get("pending", []) + current_qa.get("in_progress", []) + current_qa.get("on_hold", []) + current_qa.get("bis_testing", [])) if (t.get("qc_tester") or "").strip().lower() == name_lower]
            dev_tickets = [t for t in (current_dev.get("in_progress", []) + current_dev.get("ready_for_qc", []) + current_dev.get("other", [])) if (t.get("assignee") or "").strip().lower() == name_lower or (t.get("backend_developer") or "").strip().lower() == name_lower or (t.get("frontend_developer") or "").strip().lower() == name_lower]
            planned_qa = [p for p in qa_planned if (p.get("employee_name") or "").strip().lower() == name_lower]
            planned_dev = [p for p in dev_planned if (p.get("employee_name") or "").strip().lower() == name_lower]
            
            # Compute summary stats for member
            all_member_tickets = qa_tickets + dev_tickets
            overdue_count = sum(1 for t in all_member_tickets if t.get("eta_status") == "overdue")
            exceeded_count = sum(1 for t in all_member_tickets if t.get("time_status") == "exceeded")
            at_risk_count = sum(1 for t in all_member_tickets if t.get("time_status") == "at_risk")
            on_hold_count = sum(1 for t in all_member_tickets if "hold" in (t.get("status") or "").lower())
            urgent_high_count = sum(1 for t in all_member_tickets if t.get("priority") in ["URGENT", "High (Bugs)", "High"])
            
            # Find the max days on hold
            hold_tickets = [t for t in all_member_tickets if "hold" in (t.get("status") or "").lower()]
            max_hold_days = max((t.get("days_in_status") or 0 for t in hold_tickets), default=0)
            
            member_activity.append({
                "employee_id": emp.employee_id,
                "name": name,
                "team": (emp.team or "").strip() or "Unknown",
                "qa_tickets": qa_tickets[:15],
                "dev_tickets": dev_tickets[:15],
                "planned_qa": planned_qa[:10],
                "planned_dev": planned_dev[:10],
                "qa_count": len(qa_tickets),
                "dev_count": len(dev_tickets),
                "planned_hours": round(sum(p.get("total_hours", 0) for p in planned_qa + planned_dev), 1),
                # New enriched stats
                "overdue_count": overdue_count,
                "exceeded_count": exceeded_count,
                "at_risk_count": at_risk_count,
                "on_hold_count": on_hold_count,
                "urgent_high_count": urgent_high_count,
                "max_hold_days": round(max_hold_days, 1) if max_hold_days else 0,
            })

        return {
            "reportees": reportee_names,
            "member_activity": member_activity,
            "view": view,
            "start_date": start_date.isoformat() if start_date else None,
            "end_date": end_date.isoformat() if end_date else None,
            "period": {
                "qa": period_qa,
                "dev": period_dev,
            },
            "current": {
                "qa": {k: v[:20] for k, v in current_qa.items()},
                "dev": {k: v[:20] for k, v in current_dev.items()},
            },
            "planned_this_week": {
                "qa": qa_planned[:30],
                "dev": dev_planned[:30],
            },
        }
    except HTTPException:
        raise
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    finally:
        db.close()


@app.get("/employees/{employee_id}/rag-history")
def get_employee_rag_history(employee_id: str, current_user: dict = Depends(get_current_user)):
    """Get historical RAG scores for an employee across different time periods."""
    db: Session = SessionLocal()
    try:
        employee = db.query(Employee).filter(
            or_(
                Employee.employee_id == employee_id,
                Employee.id == int(employee_id) if employee_id.isdigit() else False
            )
        ).first()
        if not employee:
            raise HTTPException(status_code=404, detail="Employee not found")
        if not can_access_employee(db, current_user, employee.employee_id):
            raise HTTPException(status_code=403, detail="Access denied")

        is_dev = employee.team == "DEVELOPMENT"
        employee_name = employee.name
        
        # Calculate RAG scores for different periods
        periods = ["past_week", "past_month", "past_quarter", "one_year"]
        period_labels = {
            "past_week": "Past Week",
            "past_month": "Past Month", 
            "past_quarter": "Past Quarter",
            "one_year": "Past Year"
        }
        
        rag_history = []
        
        for period in periods:
            start_date, end_date = get_date_range(period)
            
            # Get bugs for this period
            if is_dev:
                # DEV: bugs assigned to them
                bugs_query = db.query(Bug).filter(
                    Bug.assignee.ilike(f"%{employee_name}%")
                )
            else:
                # QA: bugs reported by them
                bugs_query = db.query(Bug).filter(
                    Bug.author.ilike(f"%{employee_name}%")
                )
            if start_date:
                bugs_query = bugs_query.filter(Bug.created_on >= start_date)
            bugs = bugs_query.all()
            
            # Get test results for QA
            tests = []
            if not is_dev:
                tests_query = db.query(TestResult).filter(
                    TestResult.assigned_to.ilike(f"%{employee_name}%")
                )
                if start_date:
                    tests_query = tests_query.filter(TestResult.created_on >= start_date)
                tests = tests_query.all()
            
            # Get timesheets
            ts_query = db.query(Timesheet).filter(
                Timesheet.employee_name.ilike(f"%{employee_name}%")
            )
            if start_date:
                ts_query = ts_query.filter(Timesheet.date >= start_date.date())
            timesheets = ts_query.all()
            
            # Build simplified metrics
            total_bugs = len(bugs)
            closed_bugs = len([b for b in bugs if b.status == "Closed"])
            reopened = len([b for b in bugs if b.status == "Reopened"])
            
            metrics = {
                "bugs": {
                    "total": total_bugs,
                    "closure_rate": round((closed_bugs / total_bugs * 100) if total_bugs > 0 else 0, 1),
                    "reopened_percent": round((reopened / total_bugs * 100) if total_bugs > 0 else 0, 1),
                    "rejected_percent": round((len([b for b in bugs if b.status == "Rejected"]) / total_bugs * 100) if total_bugs > 0 else 0, 1),
                    "severity": {
                        "critical_percent": round((len([b for b in bugs if b.severity == "Critical"]) / total_bugs * 100) if total_bugs > 0 else 0, 1)
                    }
                },
                "tickets": {
                    "actual_hours": 0,
                    "estimate_accuracy": 100
                },
                "timesheet": {
                    "expected_hours": 40 if period == "past_week" else 160 if period == "past_month" else 480 if period == "past_quarter" else 2000,
                    "utilization_percent": 0
                },
                "tests": {
                    "total_executed": len(tests),
                    "pass_rate": round((len([t for t in tests if t.status_name == "Passed"]) / len(tests) * 100) if tests else 0, 1)
                },
                "bugs_per_ticket": 0
            }
            
            # Calculate timesheet utilization
            total_minutes = sum(t.time_logged_minutes or 0 for t in timesheets)
            total_hours = round(total_minutes / 60, 1)
            if metrics["timesheet"]["expected_hours"] > 0:
                metrics["timesheet"]["utilization_percent"] = round(
                    (total_hours / metrics["timesheet"]["expected_hours"] * 100), 1
                )
            
            # Calculate RAG score
            rag_score = calculate_rag_score(metrics, is_dev)
            rag_status = "GREEN" if rag_score >= 70 else "AMBER" if rag_score >= 50 else "RED"
            
            rag_history.append({
                "period": period,
                "label": period_labels[period],
                "score": rag_score,
                "status": rag_status,
                "bugs_count": total_bugs,
                "tests_count": len(tests) if not is_dev else None
            })
        
        # Also get saved reviews for historical context
        reviews = db.query(EmployeeReview).filter(
            EmployeeReview.employee_id == employee_id
        ).order_by(EmployeeReview.review_date.desc()).limit(5).all()
        
        review_history = []
        for review in reviews:
            review_history.append({
                "period": review.review_period,
                "date": review.review_date.isoformat() if review.review_date else None,
                "score": review.rag_score,
                "status": review.rag_status,
                "overall_rating": review.overall_rating
            })
        
        return {
            "employee_id": employee_id,
            "employee_name": employee_name,
            "team": employee.team,
            "current_rag": rag_history[0] if rag_history else None,
            "rag_trend": rag_history,
            "review_history": review_history
        }
    finally:
        db.close()


# ===== GOALS ENDPOINTS =====

@app.get("/employees/{employee_id}/goals")
def get_employee_goals(employee_id: str, current_user: dict = Depends(get_current_user)):
    """Get goals, strengths, and improvements for an employee"""
    db: Session = SessionLocal()
    try:
        employee = db.query(Employee).filter(Employee.employee_id == employee_id).first()
        if not employee:
            raise HTTPException(status_code=404, detail="Employee not found")
        if not can_access_employee(db, current_user, employee.employee_id):
            raise HTTPException(status_code=403, detail="Access denied")
        goals = db.query(EmployeeGoal).filter(
            EmployeeGoal.employee_id == employee_id
        ).order_by(EmployeeGoal.created_on.desc()).all()
        
        result = {
            "goals": [],
            "strengths": [],
            "improvements": []
        }
        
        for goal in goals:
            goal_data = {
                "id": goal.id,
                "title": goal.title,
                "description": goal.description,
                "target_date": goal.target_date.isoformat() if goal.target_date else None,
                "status": goal.status,
                "progress": goal.progress,
                "created_by": goal.created_by,
                "created_on": goal.created_on.isoformat() if goal.created_on else None
            }
            
            if goal.goal_type == "goal":
                result["goals"].append(goal_data)
            elif goal.goal_type == "strength":
                result["strengths"].append(goal_data)
            elif goal.goal_type == "improvement":
                result["improvements"].append(goal_data)
        
        return result
    finally:
        db.close()


@app.post("/employees/{employee_id}/goals")
def create_employee_goal(employee_id: str, goal: GoalCreate):
    """Create a new goal, strength, or improvement"""
    db: Session = SessionLocal()
    try:
        employee = db.query(Employee).filter(Employee.employee_id == employee_id).first()
        if not employee:
            raise HTTPException(status_code=404, detail="Employee not found")
        new_goal = EmployeeGoal(
            employee_id=employee_id,
            goal_type=goal.goal_type,
            title=goal.title,
            description=goal.description,
            target_date=goal.target_date,
            status="active",
            progress=0,
            created_by=goal.created_by,
            created_on=datetime.utcnow()
        )
        
        db.add(new_goal)
        db.commit()
        db.refresh(new_goal)
        
        return {"message": "Goal created successfully", "id": new_goal.id}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


@app.put("/goals/{goal_id}")
def update_goal(goal_id: int, updates: GoalUpdate):
    """Update a goal"""
    db: Session = SessionLocal()
    try:
        goal = db.query(EmployeeGoal).filter(EmployeeGoal.id == goal_id).first()
        
        if not goal:
            raise HTTPException(status_code=404, detail="Goal not found")
        
        update_data = updates.dict(exclude_unset=True)
        for field, value in update_data.items():
            if value is not None:
                setattr(goal, field, value)
        
        goal.updated_on = datetime.utcnow()
        db.commit()
        
        return {"message": "Goal updated successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


@app.delete("/goals/{goal_id}")
def delete_goal(goal_id: int):
    """Delete a goal"""
    db: Session = SessionLocal()
    try:
        goal = db.query(EmployeeGoal).filter(EmployeeGoal.id == goal_id).first()
        
        if not goal:
            raise HTTPException(status_code=404, detail="Goal not found")
        
        db.delete(goal)
        db.commit()
        
        return {"message": "Goal deleted successfully"}
    finally:
        db.close()


# ===== EMPLOYEE SKILLS ENDPOINTS =====

class SkillCreate(BaseModel):
    skill_name: str
    proficiency_level: int = Field(..., ge=1, le=5, description="Proficiency level 1-5")
    years_of_experience: Optional[float] = None


class SkillUpdate(BaseModel):
    skill_name: Optional[str] = None
    proficiency_level: Optional[int] = Field(None, ge=1, le=5)
    years_of_experience: Optional[float] = None


# Predefined skill suggestions
SKILL_SUGGESTIONS = {
    "QA": [
        "Manual Testing", "Automation Testing", "Selenium", "Cypress", "Playwright",
        "API Testing", "Postman", "REST Assured", "JMeter", "Performance Testing",
        "Mobile Testing", "Appium", "TestRail", "JIRA", "Test Case Design",
        "SQL", "Agile/Scrum", "Bug Reporting", "Regression Testing", "Smoke Testing",
        "UAT", "Load Testing", "Security Testing", "BDD/Cucumber", "Jenkins"
    ],
    "DEVELOPMENT": [
        "JavaScript", "TypeScript", "React", "Angular", "Vue.js", "Node.js",
        "Python", "Django", "FastAPI", "Flask", "Java", "Spring Boot",
        "C#", ".NET", "PHP", "Laravel", "Ruby", "Rails", "Go", "Rust",
        "SQL", "PostgreSQL", "MySQL", "MongoDB", "Redis", "Docker",
        "Kubernetes", "AWS", "Azure", "GCP", "Git", "CI/CD", "REST API",
        "GraphQL", "HTML/CSS", "SASS", "Tailwind", "Agile/Scrum"
    ]
}


@app.get("/employees/{employee_id}/skills")
def get_employee_skills(employee_id: str, current_user: dict = Depends(get_current_user)):
    """Get skills for an employee"""
    db: Session = SessionLocal()
    try:
        employee = db.query(Employee).filter(
            or_(
                Employee.employee_id == employee_id,
                Employee.id == int(employee_id) if employee_id.isdigit() else False
            )
        ).first()
        
        if not employee:
            raise HTTPException(status_code=404, detail="Employee not found")
        
        skills = db.query(EmployeeSkill).filter(
            EmployeeSkill.employee_id == employee.employee_id
        ).order_by(EmployeeSkill.proficiency_level.desc(), EmployeeSkill.skill_name).all()
        
        proficiency_labels = {1: "Beginner", 2: "Intermediate", 3: "Advanced", 4: "Expert", 5: "Master"}
        
        return {
            "skills": [
                {
                    "id": s.id,
                    "skill_name": s.skill_name,
                    "proficiency_level": s.proficiency_level,
                    "proficiency_label": proficiency_labels.get(s.proficiency_level, "Unknown"),
                    "years_of_experience": s.years_of_experience,
                    "created_on": s.created_on.isoformat() if s.created_on else None,
                    "updated_on": s.updated_on.isoformat() if s.updated_on else None,
                }
                for s in skills
            ],
            "suggestions": SKILL_SUGGESTIONS.get(employee.team, SKILL_SUGGESTIONS["DEVELOPMENT"])
        }
    finally:
        db.close()


@app.post("/employees/{employee_id}/skills")
def create_employee_skill(employee_id: str, skill: SkillCreate, current_user: dict = Depends(get_current_user)):
    """Add a skill to an employee's profile. Employee can add their own skills, or admin/manager can add."""
    db: Session = SessionLocal()
    try:
        employee = db.query(Employee).filter(
            or_(
                Employee.employee_id == employee_id,
                Employee.id == int(employee_id) if employee_id.isdigit() else False
            )
        ).first()
        
        if not employee:
            raise HTTPException(status_code=404, detail="Employee not found")
        
        # Check permission: employee can edit own skills, or user with edit permissions
        is_own_profile = current_user.get("employee_id") == employee.employee_id
        can_edit = can_edit_employee_profile(db, current_user, employee.employee_id)
        
        if not is_own_profile and not can_edit:
            raise HTTPException(status_code=403, detail="You don't have permission to add skills for this employee")
        
        # Check for duplicate skill
        existing = db.query(EmployeeSkill).filter(
            EmployeeSkill.employee_id == employee.employee_id,
            func.lower(EmployeeSkill.skill_name) == skill.skill_name.lower()
        ).first()
        
        if existing:
            raise HTTPException(status_code=400, detail="This skill already exists for this employee")
        
        new_skill = EmployeeSkill(
            employee_id=employee.employee_id,
            skill_name=skill.skill_name.strip(),
            proficiency_level=skill.proficiency_level,
            years_of_experience=skill.years_of_experience
        )
        
        db.add(new_skill)
        db.commit()
        db.refresh(new_skill)
        
        proficiency_labels = {1: "Beginner", 2: "Intermediate", 3: "Advanced", 4: "Expert", 5: "Master"}
        
        return {
            "id": new_skill.id,
            "skill_name": new_skill.skill_name,
            "proficiency_level": new_skill.proficiency_level,
            "proficiency_label": proficiency_labels.get(new_skill.proficiency_level, "Unknown"),
            "years_of_experience": new_skill.years_of_experience,
            "message": "Skill added successfully"
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


@app.put("/employees/{employee_id}/skills/{skill_id}")
def update_employee_skill(employee_id: str, skill_id: int, skill: SkillUpdate, current_user: dict = Depends(get_current_user)):
    """Update an employee skill"""
    db: Session = SessionLocal()
    try:
        employee = db.query(Employee).filter(
            or_(
                Employee.employee_id == employee_id,
                Employee.id == int(employee_id) if employee_id.isdigit() else False
            )
        ).first()
        
        if not employee:
            raise HTTPException(status_code=404, detail="Employee not found")
        
        existing_skill = db.query(EmployeeSkill).filter(
            EmployeeSkill.id == skill_id,
            EmployeeSkill.employee_id == employee.employee_id
        ).first()
        
        if not existing_skill:
            raise HTTPException(status_code=404, detail="Skill not found")
        
        # Check permission
        is_own_profile = current_user.get("employee_id") == employee.employee_id
        can_edit = can_edit_employee_profile(db, current_user, employee.employee_id)
        
        if not is_own_profile and not can_edit:
            raise HTTPException(status_code=403, detail="You don't have permission to update skills for this employee")
        
        # Update fields
        if skill.skill_name is not None:
            existing_skill.skill_name = skill.skill_name.strip()
        if skill.proficiency_level is not None:
            existing_skill.proficiency_level = skill.proficiency_level
        if skill.years_of_experience is not None:
            existing_skill.years_of_experience = skill.years_of_experience
        
        existing_skill.updated_on = datetime.utcnow()
        db.commit()
        
        return {"message": "Skill updated successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


@app.delete("/employees/{employee_id}/skills/{skill_id}")
def delete_employee_skill(employee_id: str, skill_id: int, current_user: dict = Depends(get_current_user)):
    """Delete an employee skill"""
    db: Session = SessionLocal()
    try:
        employee = db.query(Employee).filter(
            or_(
                Employee.employee_id == employee_id,
                Employee.id == int(employee_id) if employee_id.isdigit() else False
            )
        ).first()
        
        if not employee:
            raise HTTPException(status_code=404, detail="Employee not found")
        
        existing_skill = db.query(EmployeeSkill).filter(
            EmployeeSkill.id == skill_id,
            EmployeeSkill.employee_id == employee.employee_id
        ).first()
        
        if not existing_skill:
            raise HTTPException(status_code=404, detail="Skill not found")
        
        # Check permission
        is_own_profile = current_user.get("employee_id") == employee.employee_id
        can_edit = can_edit_employee_profile(db, current_user, employee.employee_id)
        
        if not is_own_profile and not can_edit:
            raise HTTPException(status_code=403, detail="You don't have permission to delete skills for this employee")
        
        db.delete(existing_skill)
        db.commit()
        
        return {"message": "Skill deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


# ===== EMPLOYEE ARCHIVE/RESTORE ENDPOINTS =====

@app.post("/employees/{employee_id}/archive")
def archive_employee(employee_id: str, current_user: dict = Depends(get_current_user)):
    """Archive a resigned employee. Only admin/managers can archive."""
    db: Session = SessionLocal()
    try:
        # Check if user is admin or manager
        user_role = current_user.get("role", "")
        if user_role not in ["ADMIN", "MANAGER_QA", "MANAGER_DEV", "MANAGER"]:
            raise HTTPException(status_code=403, detail="Only admins and managers can archive employees")
        
        employee = db.query(Employee).filter(
            or_(
                Employee.employee_id == employee_id,
                Employee.id == int(employee_id) if employee_id.isdigit() else False
            )
        ).first()
        
        if not employee:
            raise HTTPException(status_code=404, detail="Employee not found")
        
        if employee.archived:
            raise HTTPException(status_code=400, detail="Employee is already archived")
        
        # Archive the employee
        employee.archived = True
        employee.archived_on = datetime.utcnow()
        employee.is_active = False
        employee.employment_status = "Resigned"
        
        db.commit()
        
        return {"message": "Employee archived successfully", "archived_on": employee.archived_on.isoformat()}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


@app.post("/employees/{employee_id}/restore")
def restore_employee(employee_id: str, current_user: dict = Depends(get_current_user)):
    """Restore an archived employee. Only admin/managers can restore."""
    db: Session = SessionLocal()
    try:
        # Check if user is admin or manager
        user_role = current_user.get("role", "")
        if user_role not in ["ADMIN", "MANAGER_QA", "MANAGER_DEV", "MANAGER"]:
            raise HTTPException(status_code=403, detail="Only admins and managers can restore employees")
        
        employee = db.query(Employee).filter(
            or_(
                Employee.employee_id == employee_id,
                Employee.id == int(employee_id) if employee_id.isdigit() else False
            )
        ).first()
        
        if not employee:
            raise HTTPException(status_code=404, detail="Employee not found")
        
        if not employee.archived:
            raise HTTPException(status_code=400, detail="Employee is not archived")
        
        # Restore the employee
        employee.archived = False
        employee.archived_on = None
        employee.is_active = True
        employee.employment_status = "Ongoing Employee"
        employee.resignation_date = None
        employee.expected_lwd = None
        
        db.commit()
        
        return {"message": "Employee restored successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


# ===== REVIEW ENDPOINTS =====

@app.get("/employees/{employee_id}/reportees")
def get_employee_reportees(employee_id: str, current_user: dict = Depends(get_current_user)):
    """Get direct and indirect reportees for a lead/manager"""
    db: Session = SessionLocal()
    try:
        employee = db.query(Employee).filter(
            or_(
                Employee.employee_id == employee_id,
                Employee.id == int(employee_id) if employee_id.isdigit() else False
            )
        ).first()
        if not employee:
            raise HTTPException(status_code=404, detail="Employee not found")
        if not can_access_employee(db, current_user, employee.employee_id):
            raise HTTPException(status_code=403, detail="Access denied")

        # Find direct reportees - employees where this person is the lead
        direct_reportees = db.query(Employee).filter(
            Employee.lead.ilike(f"%{employee.name}%"),
            Employee.is_active == True,
            Employee.employee_id != employee.employee_id  # Exclude self
        ).order_by(Employee.name).all()
        
        # Find indirect reportees - employees where this person is the manager but NOT the lead
        indirect_reportees = db.query(Employee).filter(
            Employee.manager.ilike(f"%{employee.name}%"),
            ~Employee.lead.ilike(f"%{employee.name}%"),  # Not already a direct reportee
            Employee.is_active == True,
            Employee.employee_id != employee.employee_id  # Exclude self
        ).order_by(Employee.name).all()
        
        # Also get employees reporting to the direct reportees (for managers)
        # These are people whose lead reports to this manager
        manager_indirect = []
        for direct in direct_reportees:
            # Find people who report to this direct reportee
            sub_reportees = db.query(Employee).filter(
                Employee.lead.ilike(f"%{direct.name}%"),
                Employee.is_active == True,
                Employee.employee_id != direct.employee_id
            ).all()
            for sub in sub_reportees:
                if sub.employee_id not in [d.employee_id for d in direct_reportees]:
                    if sub.employee_id not in [m.employee_id for m in manager_indirect]:
                        manager_indirect.append(sub)
        
        # Calculate team breakdown
        all_direct = direct_reportees
        all_indirect = indirect_reportees + manager_indirect
        
        dev_direct = [e for e in all_direct if (e.team or '').upper() == 'DEVELOPMENT']
        qa_direct = [e for e in all_direct if (e.team or '').upper() == 'QA']
        dev_indirect = [e for e in all_indirect if (e.team or '').upper() == 'DEVELOPMENT']
        qa_indirect = [e for e in all_indirect if (e.team or '').upper() == 'QA']
        
        return {
            "direct_reportees": [{
                "employee_id": emp.employee_id,
                "name": emp.name,
                "role": emp.role,
                "team": emp.team,
                "email": emp.email,
                "category": emp.category
            } for emp in direct_reportees],
            "indirect_reportees": [{
                "employee_id": emp.employee_id,
                "name": emp.name,
                "role": emp.role,
                "team": emp.team,
                "email": emp.email,
                "category": emp.category,
                "reports_to": emp.lead
            } for emp in indirect_reportees + manager_indirect],
            "total_direct": len(direct_reportees),
            "total_indirect": len(indirect_reportees) + len(manager_indirect),
            "team_breakdown": {
                "development": {
                    "direct": len(dev_direct),
                    "indirect": len(dev_indirect),
                    "total": len(dev_direct) + len(dev_indirect)
                },
                "qa": {
                    "direct": len(qa_direct),
                    "indirect": len(qa_indirect),
                    "total": len(qa_direct) + len(qa_indirect)
                }
            }
        }
    finally:
        db.close()


@app.get("/team-leads")
def get_team_leads():
    """Get DEV Lead and QA Lead information"""
    db: Session = SessionLocal()
    try:
        # Find DEV Lead (role contains LEAD and team is DEVELOPMENT)
        dev_lead = db.query(Employee).filter(
            func.upper(Employee.role).like("%LEAD%"),
            func.upper(Employee.team) == "DEVELOPMENT",
            Employee.is_active == True
        ).first()
        
        # Find QA Lead/Manager (role contains QA and (MANAGER or LEAD) and team is QA)
        qa_lead = db.query(Employee).filter(
            or_(
                func.upper(Employee.role).like("%QA%MANAGER%"),
                func.upper(Employee.role).like("%QA%LEAD%")
            ),
            func.upper(Employee.team) == "QA",
            Employee.is_active == True
        ).first()
        
        result = {}
        
        if dev_lead:
            result["dev_lead"] = {
                "employee_id": dev_lead.employee_id,
                "name": dev_lead.name,
                "email": dev_lead.email,
                "role": dev_lead.role
            }
        else:
            result["dev_lead"] = None
            
        if qa_lead:
            result["qa_lead"] = {
                "employee_id": qa_lead.employee_id,
                "name": qa_lead.name,
                "email": qa_lead.email,
                "role": qa_lead.role
            }
        else:
            result["qa_lead"] = None
        
        return result
    finally:
        db.close()


@app.get("/employees/{employee_id}/reviews")
def get_employee_reviews(employee_id: str):
    """Get all reviews for an employee"""
    db: Session = SessionLocal()
    try:
        employee = db.query(Employee).filter(Employee.employee_id == employee_id).first()
        if not employee:
            raise HTTPException(status_code=404, detail="Employee not found")

        reviews = db.query(EmployeeReview).filter(
            EmployeeReview.employee_id == employee_id
        ).order_by(EmployeeReview.review_date.desc()).all()
        
        return [
            {
                "id": r.id,
                "review_period": r.review_period,
                "review_date": r.review_date.isoformat() if r.review_date else None,
                "rag_status": r.rag_status,
                "rag_score": r.rag_score,
                "technical_rating": r.technical_rating,
                "productivity_rating": r.productivity_rating,
                "quality_rating": r.quality_rating,
                "communication_rating": r.communication_rating,
                "overall_rating": r.overall_rating,
                "strengths_summary": r.strengths_summary,
                "improvements_summary": r.improvements_summary,
                "manager_comments": r.manager_comments,
                "recommendation": r.recommendation,
                "salary_hike_percent": r.salary_hike_percent,
                "reviewed_by": r.reviewed_by
            }
            for r in reviews
        ]
    finally:
        db.close()


@app.post("/employees/{employee_id}/reviews")
def create_employee_review(employee_id: str, review: ReviewCreate):
    """Create a new performance review"""
    db: Session = SessionLocal()
    try:
        # Calculate overall rating
        overall = (review.technical_rating + review.productivity_rating + 
                   review.quality_rating + review.communication_rating) / 4
        
        # Get current RAG score
        employee = db.query(Employee).filter(Employee.employee_id == employee_id).first()
        rag_score = 0
        rag_status = "AMBER"
        
        if employee:
            # Calculate RAG from performance metrics
            try:
                perf = get_employee_performance(employee_id, "one_year")
                rag_score = perf.get("rag_status", {}).get("score", 0)
                rag_status = perf.get("rag_status", {}).get("status", "AMBER")
            except:
                pass
        
        new_review = EmployeeReview(
            employee_id=employee_id,
            review_period=review.review_period,
            review_date=review.review_date,
            rag_status=rag_status,
            rag_score=rag_score,
            technical_rating=review.technical_rating,
            productivity_rating=review.productivity_rating,
            quality_rating=review.quality_rating,
            communication_rating=review.communication_rating,
            overall_rating=round(overall, 1),
            strengths_summary=review.strengths_summary,
            improvements_summary=review.improvements_summary,
            manager_comments=review.manager_comments,
            recommendation=review.recommendation,
            salary_hike_percent=review.salary_hike_percent,
            reviewed_by=review.reviewed_by,
            created_on=datetime.utcnow()
        )
        
        db.add(new_review)
        db.commit()
        db.refresh(new_review)
        
        return {"message": "Review created successfully", "id": new_review.id}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


@app.put("/reviews/{review_id}")
def update_review(review_id: int, review: ReviewCreate):
    """Update a performance review"""
    db: Session = SessionLocal()
    try:
        existing = db.query(EmployeeReview).filter(EmployeeReview.id == review_id).first()
        
        if not existing:
            raise HTTPException(status_code=404, detail="Review not found")
        
        overall = (review.technical_rating + review.productivity_rating + 
                   review.quality_rating + review.communication_rating) / 4
        
        existing.review_period = review.review_period
        existing.review_date = review.review_date
        existing.technical_rating = review.technical_rating
        existing.productivity_rating = review.productivity_rating
        existing.quality_rating = review.quality_rating
        existing.communication_rating = review.communication_rating
        existing.overall_rating = round(overall, 1)
        existing.strengths_summary = review.strengths_summary
        existing.improvements_summary = review.improvements_summary
        existing.manager_comments = review.manager_comments
        existing.recommendation = review.recommendation
        existing.salary_hike_percent = review.salary_hike_percent
        existing.reviewed_by = review.reviewed_by
        existing.updated_on = datetime.utcnow()
        
        db.commit()
        
        return {"message": "Review updated successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


# ===== KPI MANAGEMENT ENDPOINTS =====

@app.post("/kpis/import")
async def import_kpi_matrix(file: UploadFile = File(...)):
    """Import KPI matrix from Excel file with multiple sheets (one per role)"""
    db: Session = SessionLocal()
    try:
        import openpyxl
        import re
        
        # Save uploaded file temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            content = await file.read()
            tmp_file.write(content)
            tmp_path = tmp_file.name
        
        try:
            workbook = openpyxl.load_workbook(tmp_path)
            
            # Map sheet names to role names (normalize to match database)
            role_mapping = {
                'Software Engineer': 'SOFTWARE ENGINEER',
                'Lead': 'LEAD',
                'Project Manager': 'PROJECT MANAGER',
                'Department Heads': 'DEPARTMENT HEAD',
                'QA Engineer': 'QA ENGINEER',
                'QA Manager': 'QA MANAGER'
            }
            
            # Determine team from role
            def get_team_from_role(role_name):
                if 'QA' in role_name.upper():
                    return 'QA'
                elif 'SOFTWARE ENGINEER' in role_name.upper() or 'LEAD' in role_name.upper():
                    return 'DEVELOPMENT'
                else:
                    return None  # For PM, Department Heads, etc.
            
            total_imported = 0
            total_updated = 0
            sheet_summary = []
            
            # Process each sheet (each sheet represents a role)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                role_name = role_mapping.get(sheet_name, sheet_name.upper())
                team = get_team_from_role(role_name)
                
                imported_count = 0
                updated_count = 0
                current_kra_group = None
                
                # Process rows starting from row 3 (row 1 is empty, row 2 might be header or first data)
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    # Skip if KPI name (column B) is empty
                    if not row[1] or not str(row[1]).strip():
                        # If KRA Group (column A) has value, update current KRA
                        if row[0] and str(row[0]).strip():
                            current_kra_group = str(row[0]).strip()
                        continue
                    
                    # Extract data from columns
                    # Column A: KRA Group (category)
                    # Column B: KPI Name
                    # Column C: Weight %
                    # Column F: Evaluation Guideline (description)
                    
                    kpi_name = str(row[1]).strip()
                    if not kpi_name or kpi_name.lower() in ['kpi', 'none']:
                        continue
                    
                    # Use KRA Group from column A, or current_kra_group, or None
                    category = None
                    if row[0] and str(row[0]).strip():
                        category = str(row[0]).strip()
                        current_kra_group = category
                    elif current_kra_group:
                        category = current_kra_group
                    
                    # Weight from column C (convert percentage to decimal if needed)
                    weight_value = row[2]
                    if weight_value is not None:
                        try:
                            weight = float(weight_value)
                            # If weight is > 1, assume it's a percentage, convert to decimal
                            if weight > 1:
                                weight = weight / 100.0
                        except (ValueError, TypeError):
                            weight = 1.0
                    else:
                        weight = 1.0
                    
                    # Description from column F (Evaluation Guideline)
                    description = None
                    if len(row) > 5 and row[5]:
                        description = str(row[5]).strip()
                    
                    # Generate KPI code from name (sanitize and make unique)
                    role_prefix = role_name.replace(' ', '_')[:30]
                    kpi_code_base = re.sub(r'[^a-zA-Z0-9]', '_', kpi_name.upper())[:65]
                    kpi_code = f"{role_prefix}_{kpi_code_base}"[:100]  # Ensure total length <= 100
                    
                    # Check if KPI already exists
                    existing = db.query(KPI).filter(
                        KPI.kpi_code == kpi_code
                    ).first()
                    
                    if existing:
                        # Update existing
                        existing.kpi_name = kpi_name
                        existing.description = description
                        existing.role = role_name
                        existing.team = team
                        existing.category = category
                        existing.weight = weight
                        updated_count += 1
                    else:
                        # Create new
                        new_kpi = KPI(
                            kpi_code=kpi_code,
                            kpi_name=kpi_name,
                            description=description,
                            role=role_name,
                            team=team,
                            category=category,
                            weight=weight
                        )
                        db.add(new_kpi)
                        db.flush()  # Flush to avoid bulk insert conflicts
                        imported_count += 1
                
                total_imported += imported_count
                total_updated += updated_count
                sheet_summary.append({
                    "sheet": sheet_name,
                    "role": role_name,
                    "imported": imported_count,
                    "updated": updated_count
                })
            
            db.commit()
            
            return {
                "message": "KPI matrix imported successfully",
                "imported": total_imported,
                "updated": total_updated,
                "total": total_imported + total_updated,
                "sheets_processed": len(workbook.sheetnames),
                "sheet_details": sheet_summary
            }
        finally:
            os.unlink(tmp_path)
            
    except Exception as e:
        db.rollback()
        import traceback
        error_detail = f"Error importing KPI matrix: {str(e)}\n{traceback.format_exc()}"
        raise HTTPException(status_code=500, detail=error_detail)
    finally:
        db.close()


@app.get("/kpis")
def list_kpis(role: Optional[str] = None, team: Optional[str] = None):
    """Get all KPIs, optionally filtered by role and team"""
    db: Session = SessionLocal()
    try:
        query = db.query(KPI).filter(KPI.is_active == True)
        
        if role:
            query = query.filter(KPI.role == role)
        if team:
            query = query.filter(KPI.team == team)
        
        kpis = query.order_by(KPI.category, KPI.kpi_name).all()
        
        return [{
            "id": k.id,
            "kpi_code": k.kpi_code,
            "kpi_name": k.kpi_name,
            "description": k.description,
            "role": k.role,
            "team": k.team,
            "category": k.category,
            "weight": k.weight
        } for k in kpis]
    finally:
        db.close()


@app.get("/employees/{employee_id}/kpis")
def get_employee_kpis(employee_id: str):
    """Get all KPIs applicable to an employee based on their role and team"""
    db: Session = SessionLocal()
    try:
        employee = db.query(Employee).filter(
            or_(
                Employee.employee_id == employee_id,
                Employee.id == int(employee_id) if employee_id.isdigit() else False
            )
        ).first()
        
        if not employee:
            raise HTTPException(status_code=404, detail="Employee not found")

        if not employee.role:
            return []  # No role means no KPIs
        
        # Normalize role for comparison (uppercase)
        employee_role = employee.role.upper().strip()
        
        # Get KPIs that match the employee's role exactly (case-insensitive)
        query = db.query(KPI).filter(
            KPI.is_active == True,
            func.upper(func.trim(KPI.role)) == employee_role
        )
        
        # Filter by team: match exact team OR if KPI.team is None (applies to all teams)
        if employee.team:
            # Normalize team for comparison (uppercase)
            employee_team = employee.team.upper().strip()
            query = query.filter(
                or_(
                    func.upper(func.trim(KPI.team)) == employee_team,
                    KPI.team.is_(None)
                )
            )
        else:
            # If employee has no team, only show KPIs with no team specified
            query = query.filter(KPI.team.is_(None))
        
        kpis = query.order_by(KPI.category, KPI.kpi_name).all()
        
        return [{
            "id": k.id,
            "kpi_code": k.kpi_code,
            "kpi_name": k.kpi_name,
            "description": k.description,
            "category": k.category,
            "weight": k.weight
        } for k in kpis]
    finally:
        db.close()


@app.get("/employees/{employee_id}/kpi-ratings")
def get_employee_kpi_ratings(employee_id: str, quarter: Optional[str] = None):
    """Get KPI ratings for an employee, optionally filtered by quarter"""
    db: Session = SessionLocal()
    try:
        employee = db.query(Employee).filter(
            or_(
                Employee.employee_id == employee_id,
                Employee.id == int(employee_id) if employee_id.isdigit() else False
            )
        ).first()
        
        if not employee:
            raise HTTPException(status_code=404, detail="Employee not found")

        # Get current year and quarter if not specified
        if not quarter:
            now = datetime.now()
            year = now.year
            quarter_num = (now.month - 1) // 3 + 1
            quarter = f"{year}-Q{quarter_num}"
        
        # Get all KPIs for this employee - match role and team exactly
        if not employee.role:
            return {"kpis": [], "quarter": quarter}
        
        # Normalize role for comparison (uppercase)
        employee_role = employee.role.upper().strip()
        
        query = db.query(KPI).filter(
            KPI.is_active == True,
            func.upper(func.trim(KPI.role)) == employee_role
        )
        
        # Filter by team: match exact team OR if KPI.team is None (applies to all teams)
        if employee.team:
            # Normalize team for comparison (uppercase)
            employee_team = employee.team.upper().strip()
            query = query.filter(
                or_(
                    func.upper(func.trim(KPI.team)) == employee_team,
                    KPI.team.is_(None)
                )
            )
        else:
            # If employee has no team, only show KPIs with no team specified
            query = query.filter(KPI.team.is_(None))
        
        kpis = query.order_by(KPI.category, KPI.kpi_name).all()
        
        # Get ratings for this quarter
        ratings_query = db.query(KPIRating).filter(
            KPIRating.employee_id == employee.employee_id,
            KPIRating.quarter == quarter
        )
        ratings = {r.kpi_id: r for r in ratings_query.all()}
        
        # Calculate overall weighted score
        total_weighted_score = 0.0
        total_weight = 0.0
        rated_kpis_count = 0
        
        result = []
        for kpi in kpis:
            rating = ratings.get(kpi.id)
            
            # Calculate score for this KPI (use manager_rating if available, else performance_score, else 0)
            kpi_score = 0.0
            if rating:
                if rating.manager_rating is not None:
                    # Convert 1-5 scale to percentage (1=20%, 5=100%)
                    kpi_score = (rating.manager_rating / 5.0) * 100
                elif rating.performance_score is not None:
                    kpi_score = rating.performance_score
                elif rating.final_score is not None:
                    kpi_score = (rating.final_score / 5.0) * 100 if rating.final_score <= 5 else rating.final_score
                
                if kpi_score > 0:
                    total_weighted_score += kpi_score * kpi.weight
                    total_weight += kpi.weight
                    rated_kpis_count += 1
            
            result.append({
                "kpi_id": kpi.id,
                "kpi_code": kpi.kpi_code,
                "kpi_name": kpi.kpi_name,
                "description": kpi.description,
                "category": kpi.category,
                "weight": kpi.weight,
                "rating": rating.rating if rating else None,
                "self_rating": rating.self_rating if rating else None,
                "lead_rating": rating.lead_rating if rating else None,
                "manager_rating": rating.manager_rating if rating else None,
                "performance_score": rating.performance_score if rating else None,
                "performance_percentage": rating.performance_percentage if rating else None,
                "final_score": rating.final_score if rating else None,
                "self_comments": rating.self_comments if rating else None,
                "lead_comments": rating.lead_comments if rating else None,
                "manager_comments": rating.manager_comments if rating else None,
                "rated_by": rating.rated_by if rating else None,
                "rated_on": rating.rated_on.isoformat() if rating and rating.rated_on else None
            })
        
        # Calculate overall weighted average
        overall_score = 0.0
        overall_rating_label = "Not Rated"
        
        if total_weight > 0 and rated_kpis_count > 0:
            overall_score = total_weighted_score / total_weight
            
            # Determine rating label based on score
            if overall_score >= 90:
                overall_rating_label = "Outstanding"
            elif overall_score >= 80:
                overall_rating_label = "Excellent"
            elif overall_score >= 70:
                overall_rating_label = "Good"
            elif overall_score >= 60:
                overall_rating_label = "Satisfactory"
            elif overall_score >= 50:
                overall_rating_label = "Needs Improvement"
            else:
                overall_rating_label = "Poor"
        
        return {
            "employee_id": employee.employee_id,
            "employee_name": employee.name,
            "role": employee.role,
            "quarter": quarter,
            "kpis": result,
            "overall_score": round(overall_score, 2),
            "overall_rating": overall_rating_label,
            "rated_kpis_count": rated_kpis_count,
            "total_kpis_count": len(kpis)
        }
    finally:
        db.close()


@app.post("/employees/{employee_id}/kpi-ratings")
def submit_kpi_ratings(
    employee_id: str,
    ratings: List[KPIRatingCreate]
):
    """Submit KPI ratings for an employee for a quarter"""
    db: Session = SessionLocal()
    try:
        employee = db.query(Employee).filter(
            or_(
                Employee.employee_id == employee_id,
                Employee.id == int(employee_id) if employee_id.isdigit() else False
            )
        ).first()
        
        if not employee:
            raise HTTPException(status_code=404, detail="Employee not found")

        # Parse quarter to get year and quarter number
        quarter = ratings[0].quarter if ratings else None
        if not quarter:
            raise HTTPException(status_code=400, detail="Quarter is required")
        
        year = int(quarter.split('-')[0])
        quarter_num = int(quarter.split('-Q')[1])
        
        submitted_count = 0
        
        for rating_data in ratings:
            # Verify KPI exists
            kpi = db.query(KPI).filter(KPI.id == rating_data.kpi_id).first()
            if not kpi:
                continue
            
            # Check if rating already exists
            existing = db.query(KPIRating).filter(
                KPIRating.employee_id == employee.employee_id,
                KPIRating.kpi_id == rating_data.kpi_id,
                KPIRating.quarter == quarter
            ).first()
            
            # Calculate performance score from actual metrics
            performance_score = calculate_kpi_performance_score(
                db, employee, kpi, quarter
            )
            
            # Calculate final score (use manager rating if provided, otherwise performance score)
            final_score = None
            if rating_data.manager_rating is not None:
                final_score = rating_data.manager_rating
            elif performance_score is not None:
                final_score = performance_score
            
            # Determine if lead and manager are the same person
            is_lead_manager_same = employee.lead and employee.manager and employee.lead.strip().upper() == employee.manager.strip().upper()
            
            # Update or create rating based on who is rating
            if existing:
                # Update existing - only update the field for the current rater
                if rating_data.rated_by == "self":
                    existing.self_rating = rating_data.self_rating
                    existing.self_comments = rating_data.self_comments
                elif rating_data.rated_by == "lead":
                    existing.lead_rating = rating_data.lead_rating
                    existing.lead_comments = rating_data.lead_comments
                    # If lead and manager are same, also update manager fields
                    if is_lead_manager_same:
                        existing.manager_rating = rating_data.lead_rating
                        existing.manager_comments = rating_data.lead_comments
                elif rating_data.rated_by == "manager":
                    existing.manager_rating = rating_data.manager_rating
                    existing.manager_comments = rating_data.manager_comments
                    # If lead and manager are same, also update lead fields
                    if is_lead_manager_same:
                        existing.lead_rating = rating_data.manager_rating
                        existing.lead_comments = rating_data.manager_comments
                
                # Keep backward compatibility
                if rating_data.rating is not None:
                    existing.rating = rating_data.rating
                if rating_data.manager_rating is not None:
                    existing.manager_rating = rating_data.manager_rating
                if rating_data.manager_comments is not None:
                    existing.manager_comments = rating_data.manager_comments
                
                existing.performance_score = performance_score
                existing.final_score = final_score
                existing.rated_by = rating_data.rated_by
                existing.rated_on = datetime.now()
            else:
                # Create new
                new_rating = KPIRating(
                    employee_id=employee.employee_id,
                    kpi_id=rating_data.kpi_id,
                    quarter=quarter,
                    year=year,
                    quarter_number=quarter_num,
                    rating=rating_data.rating,  # Backward compatibility
                    self_rating=rating_data.self_rating if rating_data.rated_by == "self" else None,
                    lead_rating=rating_data.lead_rating if rating_data.rated_by == "lead" else None,
                    manager_rating=rating_data.manager_rating if rating_data.rated_by == "manager" else None,
                    self_comments=rating_data.self_comments if rating_data.rated_by == "self" else None,
                    lead_comments=rating_data.lead_comments if rating_data.rated_by == "lead" else None,
                    manager_comments=rating_data.manager_comments if rating_data.rated_by == "manager" else None,
                    performance_score=performance_score,
                    final_score=final_score,
                    rated_by=rating_data.rated_by
                )
                # If lead and manager are same, copy lead rating to manager
                if is_lead_manager_same and rating_data.rated_by == "lead":
                    new_rating.manager_rating = rating_data.lead_rating
                    new_rating.manager_comments = rating_data.lead_comments
                elif is_lead_manager_same and rating_data.rated_by == "manager":
                    new_rating.lead_rating = rating_data.manager_rating
                    new_rating.lead_comments = rating_data.manager_comments
                
                db.add(new_rating)
            
            submitted_count += 1
        
        db.commit()
        
        return {
            "message": f"Successfully submitted {submitted_count} KPI ratings",
            "quarter": quarter,
            "count": submitted_count
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


def calculate_kpi_performance_score(db: Session, employee: Employee, kpi: KPI, quarter: str) -> Optional[float]:
    """Calculate performance score for a KPI based on actual metrics"""
    # Parse quarter to get date range
    year = int(quarter.split('-')[0])
    quarter_num = int(quarter.split('-Q')[1])
    start_month = (quarter_num - 1) * 3 + 1
    start_date = datetime(year, start_month, 1)
    
    if quarter_num == 4:
        end_date = datetime(year + 1, 1, 1)
    else:
        end_date = datetime(year, start_month + 3, 1)
    
    # This is a placeholder - implement actual calculation based on KPI code
    # You can implement specific calculations based on kpi.kpi_code
    # Example: If KPI is about bug closure rate, calculate from actual bugs
    
    return None


# ===== STATUS HISTORY ENDPOINTS =====

@app.get("/status-history/tickets")
def get_ticket_status_history(
    ticket_id: Optional[int] = Query(None, description="Filter by specific ticket ID"),
    status: Optional[str] = Query(None, description="Filter by status (new_status)"),
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    limit: int = Query(100, description="Maximum records to return")
):
    """Get ticket status change history"""
    db: Session = SessionLocal()
    try:
        query = db.query(TicketStatusHistory)
        
        if ticket_id:
            query = query.filter(TicketStatusHistory.ticket_id == ticket_id)
        
        if status:
            query = query.filter(TicketStatusHistory.new_status == status)
        
        if start_date:
            start = datetime.strptime(start_date, "%Y-%m-%d")
            query = query.filter(TicketStatusHistory.changed_on >= start)
        
        if end_date:
            end = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            query = query.filter(TicketStatusHistory.changed_on <= end)
        
        history = query.order_by(TicketStatusHistory.changed_on.desc()).limit(limit).all()
        
        return [
            {
                "id": h.id,
                "ticket_id": h.ticket_id,
                "previous_status": h.previous_status,
                "new_status": h.new_status,
                "changed_on": h.changed_on.isoformat() if h.changed_on else None,
                "current_assignee": h.current_assignee,
                "qc_tester": h.qc_tester,
                "duration_in_previous_status": h.duration_in_previous_status,
                "source": h.source
            }
            for h in history
        ]
    finally:
        db.close()


# QC statuses: first time ticket hits one of these we consider it "released to QC"
QC_RELEASE_STATUSES = ['QC Testing', 'QC Testing in Progress', 'QC Testing Hold']


@app.get("/tickets/{ticket_id}/status-history-after-qc")
def get_ticket_status_history_after_qc(
    ticket_id: int,
    current_user: dict = Depends(get_current_user),
):
    """
    Status change history for a ticket from the moment it was first released to QC Testing onward.
    Used in ETA calendar ticket details to show every status change after QC release.
    """
    db: Session = SessionLocal()
    try:
        all_history = (
            db.query(TicketStatusHistory)
            .filter(TicketStatusHistory.ticket_id == ticket_id)
            .order_by(TicketStatusHistory.changed_on.asc())
            .all()
        )
        qc_release_cutoff = None
        for h in all_history:
            if h.new_status and h.new_status in QC_RELEASE_STATUSES:
                qc_release_cutoff = h.changed_on
                break
        if qc_release_cutoff is None:
            return {"ticket_id": ticket_id, "released_to_qc_on": None, "history": []}
        after_qc = [
            h for h in all_history
            if h.changed_on and h.changed_on >= qc_release_cutoff
        ]
        return {
            "ticket_id": ticket_id,
            "released_to_qc_on": qc_release_cutoff.isoformat() if qc_release_cutoff else None,
            "history": [
                {
                    "id": h.id,
                    "previous_status": h.previous_status,
                    "new_status": h.new_status,
                    "changed_on": h.changed_on.isoformat() if h.changed_on else None,
                    "current_assignee": h.current_assignee,
                    "qc_tester": h.qc_tester,
                }
                for h in reversed(after_qc)
            ],
        }
    finally:
        db.close()


@app.get("/status-history/tickets/moved-to")
def get_tickets_moved_to_status(
    status: str = Query(..., description="Target status (e.g., 'BIS Testing')"),
    start_date: str = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: str = Query(..., description="End date (YYYY-MM-DD)")
):
    """Get tickets that moved to a specific status during a date range"""
    db: Session = SessionLocal()
    try:
        start = datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
        
        # Find status changes to the target status within the date range
        history = db.query(TicketStatusHistory).filter(
            TicketStatusHistory.new_status == status,
            TicketStatusHistory.changed_on >= start,
            TicketStatusHistory.changed_on <= end
        ).order_by(TicketStatusHistory.changed_on.desc()).all()
        
        # Get unique ticket IDs
        ticket_ids = list(set(h.ticket_id for h in history))
        
        # Get ticket details
        tickets = db.query(TicketTracking).filter(
            TicketTracking.ticket_id.in_(ticket_ids)
        ).all() if ticket_ids else []
        
        ticket_map = {t.ticket_id: t for t in tickets}
        
        result = []
        for h in history:
            ticket = ticket_map.get(h.ticket_id)
            result.append({
                "ticket_id": h.ticket_id,
                "moved_from": h.previous_status,
                "moved_to": h.new_status,
                "moved_on": h.changed_on.isoformat() if h.changed_on else None,
                "current_status": ticket.status if ticket else None,
                "qc_tester": ticket.qc_tester if ticket else h.qc_tester,
                "duration_in_previous_status_hours": h.duration_in_previous_status
            })
        
        return {
            "status": status,
            "date_range": {"start": start_date, "end": end_date},
            "total_count": len(ticket_ids),
            "tickets": result
        }
    finally:
        db.close()


@app.get("/status-history/bugs")
def get_bug_status_history(
    bug_id: Optional[int] = Query(None, description="Filter by specific bug ID"),
    ticket_id: Optional[int] = Query(None, description="Filter by ticket ID"),
    status: Optional[str] = Query(None, description="Filter by status"),
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    limit: int = Query(100, description="Maximum records to return")
):
    """Get bug status change history"""
    db: Session = SessionLocal()
    try:
        query = db.query(BugStatusHistory)
        
        if bug_id:
            query = query.filter(BugStatusHistory.bug_id == bug_id)
        
        if ticket_id:
            query = query.filter(BugStatusHistory.ticket_id == ticket_id)
        
        if status:
            query = query.filter(BugStatusHistory.new_status == status)
        
        if start_date:
            start = datetime.strptime(start_date, "%Y-%m-%d")
            query = query.filter(BugStatusHistory.changed_on >= start)
        
        if end_date:
            end = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            query = query.filter(BugStatusHistory.changed_on <= end)
        
        history = query.order_by(BugStatusHistory.changed_on.desc()).limit(limit).all()
        
        return [
            {
                "id": h.id,
                "bug_id": h.bug_id,
                "ticket_id": h.ticket_id,
                "previous_status": h.previous_status,
                "new_status": h.new_status,
                "changed_on": h.changed_on.isoformat() if h.changed_on else None,
                "assignee": h.assignee,
                "duration_in_previous_status": h.duration_in_previous_status,
                "source": h.source
            }
            for h in history
        ]
    finally:
        db.close()


@app.get("/status-history/summary")
def get_status_history_summary(
    start_date: str = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: str = Query(..., description="End date (YYYY-MM-DD)")
):
    """Get summary of status changes during a date range"""
    db: Session = SessionLocal()
    try:
        start = datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
        
        # Ticket status changes
        ticket_changes = db.query(TicketStatusHistory).filter(
            TicketStatusHistory.changed_on >= start,
            TicketStatusHistory.changed_on <= end
        ).all()
        
        # Bug status changes
        bug_changes = db.query(BugStatusHistory).filter(
            BugStatusHistory.changed_on >= start,
            BugStatusHistory.changed_on <= end
        ).all()
        
        # Aggregate ticket changes by status
        ticket_moved_to = defaultdict(int)
        ticket_moved_from = defaultdict(int)
        for tc in ticket_changes:
            if tc.new_status:
                ticket_moved_to[tc.new_status] += 1
            if tc.previous_status:
                ticket_moved_from[tc.previous_status] += 1
        
        # Aggregate bug changes by status
        bug_moved_to = defaultdict(int)
        bug_moved_from = defaultdict(int)
        for bc in bug_changes:
            if bc.new_status:
                bug_moved_to[bc.new_status] += 1
            if bc.previous_status:
                bug_moved_from[bc.previous_status] += 1
        
        return {
            "date_range": {"start": start_date, "end": end_date},
            "tickets": {
                "total_changes": len(ticket_changes),
                "unique_tickets": len(set(tc.ticket_id for tc in ticket_changes)),
                "moved_to": dict(ticket_moved_to),
                "moved_from": dict(ticket_moved_from)
            },
            "bugs": {
                "total_changes": len(bug_changes),
                "unique_bugs": len(set(bc.bug_id for bc in bug_changes)),
                "moved_to": dict(bug_moved_to),
                "moved_from": dict(bug_moved_from)
            }
        }
    finally:
        db.close()


# ===== WEEKLY REPORT ENDPOINTS =====

@app.get("/reports/weekly")
def generate_weekly_report(
    date: str = Query(None, description="Reference date (YYYY-MM-DD) for the week. Defaults to current week."),
    download: bool = Query(True, description="If true, returns the PDF file. If false, returns report data as JSON."),
    current_user: dict = Depends(require_reports_access),
):
    """Generate weekly QA report for the specified week"""
    from weekly_report import get_week_dates, get_weekly_data, generate_pdf_report
    import os
    
    try:
        # Get week dates
        week_start, week_end = get_week_dates(date)
        
        # Fetch data
        data = get_weekly_data(week_start, week_end)
        
        if not download:
            # Return JSON summary
            return {
                "week_start": week_start.strftime("%Y-%m-%d"),
                "week_end": week_end.strftime("%Y-%m-%d"),
                "summary": data['summary'],
                "tickets_bis_testing_count": len(data['tickets_bis_testing']),
                "tickets_closed_count": len(data['tickets_closed']),
                "tickets_in_progress_count": len(data['tickets_in_progress']),
                "next_week_plan_count": len(data['next_week_plan'])
            }
        
        # Generate PDF
        reports_folder = os.path.join(os.path.dirname(__file__), "reports")
        os.makedirs(reports_folder, exist_ok=True)
        
        output_path = os.path.join(
            reports_folder,
            f"QA_Weekly_Report_{week_start.strftime('%Y%m%d')}_{week_end.strftime('%Y%m%d')}.pdf"
        )
        
        generate_pdf_report(data, output_path)
        
        return FileResponse(
            output_path,
            media_type="application/pdf",
            filename=os.path.basename(output_path)
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/reports/weekly/preview")
def preview_weekly_report(
    date: str = Query(None, description="Reference date (YYYY-MM-DD) for the week"),
    current_user: dict = Depends(require_reports_access),
):
    """Get preview data for weekly report without generating PDF"""
    from weekly_report import get_week_dates, get_weekly_data
    
    try:
        week_start, week_end = get_week_dates(date)
        data = get_weekly_data(week_start, week_end)
        
        return {
            "week_start": week_start.strftime("%Y-%m-%d"),
            "week_end": week_end.strftime("%Y-%m-%d"),
            "summary": data['summary'],
            "tickets_bis_testing": data['tickets_bis_testing'],
            "tickets_closed": data['tickets_closed'],
            "tickets_in_progress": data['tickets_in_progress'][:20],  # Limit for preview
            "next_week_plan": data['next_week_plan']
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/reports/ticket/{ticket_id}")
def generate_ticket_report_endpoint(
    ticket_id: int,
    current_user: dict = Depends(require_reports_access),
):
    """Generate PDF report for a specific ticket with all its data"""
    from ticket_report import get_ticket_data, generate_ticket_pdf
    import os
    
    try:
        # Fetch data
        data = get_ticket_data(ticket_id)
        
        if not data:
            raise HTTPException(status_code=404, detail=f"Ticket #{ticket_id} not found")
        
        # Generate PDF
        reports_folder = os.path.join(os.path.dirname(__file__), "reports")
        os.makedirs(reports_folder, exist_ok=True)
        
        output_path = os.path.join(
            reports_folder,
            f"Ticket_Report_{ticket_id}.pdf"
        )
        
        generate_ticket_pdf(data, output_path)
        
        return FileResponse(
            output_path,
            media_type="application/pdf",
            filename=f"Ticket_Report_{ticket_id}.pdf"
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/reports/developer/{employee_id}")
def download_developer_report(
    employee_id: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_reports_access),
):
    """Generate and download a developer performance PDF report."""
    from developer_report import generate_developer_report_pdf

    try:
        output_path, filename = generate_developer_report_pdf(db, employee_id)
        return FileResponse(
            output_path,
            media_type="application/pdf",
            filename=filename,
        )
    except ValueError as e:
        message = str(e)
        if "not found" in message.lower():
            raise HTTPException(status_code=404, detail=message)
        raise HTTPException(status_code=400, detail=message)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/reports/developer/team/{lead_name}")
def download_team_developer_reports(
    lead_name: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_reports_access),
):
    """Generate and download ZIP of developer reports for all team members reporting to a lead."""
    from developer_report import generate_team_reports_zip

    try:
        zip_path, zip_filename, employee_names = generate_team_reports_zip(db, lead_name)
        return FileResponse(
            zip_path,
            media_type="application/zip",
            filename=zip_filename,
            headers={"X-Employees-Included": ", ".join(employee_names)},
        )
    except ValueError as e:
        message = str(e)
        if "not found" in message.lower() or "no development" in message.lower():
            raise HTTPException(status_code=404, detail=message)
        raise HTTPException(status_code=400, detail=message)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/reports/developer/bulk")
def download_bulk_developer_reports(
    employee_ids: str = Query(..., description="Comma-separated employee IDs"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_reports_access),
):
    """Generate and download ZIP of developer reports for specified employees."""
    from developer_report import generate_bulk_reports_zip

    try:
        ids = [eid.strip() for eid in employee_ids.split(",") if eid.strip()]
        if not ids:
            raise HTTPException(status_code=400, detail="No employee IDs provided")
        
        zip_path, zip_filename = generate_bulk_reports_zip(db, ids)
        return FileResponse(
            zip_path,
            media_type="application/zip",
            filename=zip_filename,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/reports/open-bugs")
def download_open_bugs_report(
    sort_by: str = Query("ageing", description="Sort by: ageing, bug_id, ticket_id, developer, severity"),
    sort_order: str = Query("desc", description="Sort order: asc or desc"),
    developer: str = Query(None, description="Filter by developer name (optional)"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_reports_access),
):
    """Generate and download a PDF report of all currently open bugs."""
    from open_bugs_report import generate_open_bugs_report_pdf

    try:
        output_path, filename = generate_open_bugs_report_pdf(
            db, sort_by=sort_by, sort_order=sort_order, developer_filter=developer
        )
        return FileResponse(
            output_path,
            media_type="application/pdf",
            filename=filename,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/reports/open-bugs/preview")
def preview_open_bugs_report(
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_reports_access),
):
    """Get preview data for open bugs report."""
    from open_bugs_report import get_open_bugs_preview

    try:
        return get_open_bugs_preview(db)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/reports/weekly-v2")
def generate_weekly_report_v2(
    date: str = Query(None, description="Reference date (YYYY-MM-DD) for the week"),
    project: str = Query(None, description="Project/Client name for the cover page"),
    last7days: bool = Query(True, description="If true, show last 7 days. If false, show Mon-Fri week."),
    current_user: dict = Depends(require_reports_access),
):
    """Generate comprehensive multi-page QA weekly report (V2)"""
    from qa_weekly_report_v2 import get_week_dates, get_comprehensive_data, generate_comprehensive_report
    import os
    
    try:
        # Get week dates - use last 7 days by default
        week_start, week_end = get_week_dates(date, use_last_7_days=last7days)
        
        # Fetch comprehensive data
        data = get_comprehensive_data(week_start, week_end)
        
        # Generate PDF
        reports_folder = os.path.join(os.path.dirname(__file__), "reports")
        os.makedirs(reports_folder, exist_ok=True)
        
        output_path = os.path.join(
            reports_folder,
            f"QA_Weekly_Report_V2_{week_start.strftime('%Y%m%d')}_{week_end.strftime('%Y%m%d')}.pdf"
        )
        
        generate_comprehensive_report(data, output_path, project)
        
        return FileResponse(
            output_path,
            media_type="application/pdf",
            filename=os.path.basename(output_path)
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/reports/weekly-v2/preview")
def preview_weekly_report_v2(
    date: str = Query(None, description="Reference date (YYYY-MM-DD) for the week"),
    last7days: bool = Query(True, description="If true, show last 7 days. If false, show Mon-Fri week."),
    current_user: dict = Depends(require_reports_access),
):
    """Get preview data for the comprehensive weekly report"""
    from qa_weekly_report_v2 import get_week_dates, get_comprehensive_data
    
    try:
        # Use last 7 days by default
        week_start, week_end = get_week_dates(date, use_last_7_days=last7days)
        data = get_comprehensive_data(week_start, week_end)
        
        # Build BIS tickets with open/deferred bug lists for report
        def _bis_ticket_for_preview(t):
            out = {
                "ticket_id": t['ticket_id'],
                "title": t['title'],
                "status": t['status'],
                "priority": t.get('priority'),
                "qa_tester": t.get('qa_tester'),
                "developers_str": t.get('developers_str'),
                "bugs_total": t['bugs_total'],
                "bugs_open": t['bugs_open'],
                "bugs_closed": t.get('bugs_closed', 0),
                "bugs_deferred": t.get('bugs_deferred', 0),
                "tests_total": t['tests_total'],
                "tests_passed": t.get('tests_passed', 0),
                "pass_rate": t['pass_rate'],
            }
            if t.get('bug_details'):
                open_bugs = [b for b in t['bug_details'] if (b.get('status') or '').lower() not in ('closed', 'resolved', 'verified', 'fixed', 'reject', 'deferred', 'wont fix', 'duplicate')]
                deferred_bugs = [b for b in t['bug_details'] if (b.get('status') or '').lower() in ('deferred', 'wont fix', 'duplicate')]
                out["open_bugs"] = open_bugs
                out["deferred_bugs"] = deferred_bugs
            else:
                out["open_bugs"] = []
                out["deferred_bugs"] = []
            return out

        return {
            "week_start": week_start.strftime("%Y-%m-%d"),
            "week_end": week_end.strftime("%Y-%m-%d"),
            "current_week": {
                "qa_tickets_count": len(data['current_week']['qa_tickets']),
                "qc_newly_added_count": len(data['current_week'].get('qc_testing_newly_added', [])),
                "bis_testing_count": len(data['current_week']['bis_testing_moved']),
                "closed_count": len(data['current_week']['closed_moved']),
                "in_progress_count": len(data['current_week']['in_progress']),
            },
            "qa_pending_breakdown": data.get('qa_pending_breakdown', {}),
            "previous_week": data['previous_week'],
            "variance": data.get('variance', {}),
            "metrics": data['metrics'],
            "breakdowns": {
                "by_status": dict(data['breakdowns']['by_status']),
                "by_module": dict(data['breakdowns']['by_module']),
            },
            "tickets_worked_on_this_week": data.get('tickets_worked_on_this_week', []),
            "qc_testing_newly_added": data['current_week'].get('qc_testing_newly_added', []),
            "on_hold_this_week": data.get('on_hold_this_week', []),
            "qa_failed_this_week": data.get('qa_failed_this_week', []),
            "bis_testing_moved": [_bis_ticket_for_preview(t) for t in data['current_week']['bis_testing_moved']],
            "next_week_plan": data.get('next_week_plan', []),
            "next_week_eta_calendar": data.get('next_week_eta_calendar', []),
            "next_week_plan_count": len(data['next_week_plan']),
            "bis_testing_tickets": [
                {
                    "ticket_id": t['ticket_id'],
                    "title": t['title'],
                    "status": t['status'],
                    "bugs_total": t['bugs_total'],
                    "bugs_open": t['bugs_open'],
                    "tests_total": t['tests_total'],
                    "pass_rate": t['pass_rate']
                }
                for t in data['current_week']['bis_testing_moved']
            ],
            "closed_tickets": [
                {
                    "ticket_id": t['ticket_id'],
                    "title": t['title'],
                    "bugs_closed": t['bugs_closed'],
                    "tests_passed": t['tests_passed']
                }
                for t in data['current_week']['closed_moved']
            ]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ===== QA METRICS DASHBOARD (4 KEY METRICS) =====

def _get_qa_metrics_data(period: str, db: Session, start_date: str = None, end_date: str = None):
    """
    Internal function to get QA metrics data.
    Used by both authenticated and public endpoints.
    """
    from pathlib import Path
    from collections import defaultdict, Counter
    import json
    
    # Period definitions
    PERIOD_DAYS = {
        'past_week': 7,
        'past_month': 30,
        'past_quarter': 90,
        'past_year': 365,
        'overall': None,
    }
    
    # Status definitions
    QC_TESTING = 'QC Testing'
    QC_TESTING_IN_PROGRESS = 'QC Testing in Progress'
    QC_REVIEW_FAIL = 'QC Review Fail'
    QC_TESTING_HOLD = 'QC Testing Hold'
    QC_TESTING_ON_HOLD = 'QC Testing On-hold'
    BIS_TESTING = 'BIS Testing'
    QC_FAIL_STATUSES = {QC_REVIEW_FAIL, 'Tested - Awaiting Fixes'}
    
    # Find the latest PM Activity Export file
    reports_dir = Path("reports")
    export_files = list(reports_dir.glob("PM_Activity_Export_*.csv"))
    if not export_files:
        raise HTTPException(status_code=404, detail="No PM Activity Export file found")
    
    latest_export = max(export_files, key=lambda f: f.stat().st_mtime)
    
    # Load data
    try:
        with open(latest_export, 'r', encoding='utf-8') as f:
            raw_data = json.load(f)
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=500, detail=f"Failed to parse data: {e}")
    
    # Group by ticket
    ticket_history = defaultdict(list)
    for record in raw_data:
        ticket_id = record.get('ticketId')
        if ticket_id:
            try:
                ticket_id = int(ticket_id)
            except (ValueError, TypeError):
                pass
            change_date = datetime.strptime(record['statusChangeDate'], '%Y-%m-%d %H:%M:%S')
            ticket_history[ticket_id].append({
                'date': change_date,
                'old_status': record.get('oldStatus'),
                'new_status': record.get('newStatus'),
            })
    
    # Sort histories
    for tid in ticket_history:
        ticket_history[tid].sort(key=lambda x: x['date'])
    
    # Optional explicit date-range filter (YYYY-MM-DD)
    range_start = None
    range_end = None
    if start_date:
        try:
            range_start = datetime.strptime(start_date, "%Y-%m-%d")
        except ValueError:
            range_start = None
    if end_date:
        try:
            range_end = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
        except ValueError:
            range_end = None

    # Filter by period and only tickets that entered QA
    days = PERIOD_DAYS.get(period)
    cutoff_date = datetime.now() - timedelta(days=days) if days else None
    
    def ticket_entered_qa(history):
        for h in history:
            if h.get('new_status') in (QC_TESTING, QC_TESTING_IN_PROGRESS):
                return True
        return False
    
    def calculate_business_days(start, end):
        if not start or not end or end <= start:
            return 0.0
        total = 0
        current = start
        while current < end:
            if current.weekday() < 5:
                total += 1
            current += timedelta(days=1)
        return total
    
    # Calculate metrics for filtered tickets
    tickets_data = []
    all_test_cycles = []
    all_waiting_times = []
    cycle_distribution = Counter()
    daily_stats = defaultdict(list)
    
    ticket_lookup = {t.ticket_id: t for t in db.query(TicketTracking).all()}
    
    for ticket_id, history in ticket_history.items():
        if not ticket_entered_qa(history):
            continue
        
        # Check period/date-range filter
        if range_start or range_end:
            has_activity_in_period = any(
                (range_start is None or h['date'] >= range_start) and
                (range_end is None or h['date'] <= range_end)
                for h in history
            )
            if not has_activity_in_period:
                continue
        elif cutoff_date:
            has_activity_in_period = any(h['date'] >= cutoff_date for h in history)
            if not has_activity_in_period:
                continue
        
        # Calculate metrics
        first_qc_testing = None
        first_bis_testing = None
        test_cycles = []
        waiting_times = []
        in_qc_testing = False
        in_progress = False
        in_dev_hold = False
        in_qa_hold = False
        qc_testing_entry = None
        progress_entry = None
        dev_hold_entry = None
        qa_hold_entry = None
        
        # Time tracking
        total_active_testing_days = 0.0
        total_waiting_in_queue_days = 0.0
        total_dev_hold_days = 0.0
        total_qa_hold_days = 0.0
        
        for h in history:
            new_status = h.get('new_status', '')
            change_date = h['date']
            
            # First QC Testing
            if new_status == QC_TESTING and first_qc_testing is None:
                first_qc_testing = change_date
                qc_testing_entry = change_date
                in_qc_testing = True
            elif new_status == QC_TESTING and not in_qc_testing:
                qc_testing_entry = change_date
                in_qc_testing = True
            
            # First BIS Testing
            if new_status == BIS_TESTING and first_bis_testing is None:
                first_bis_testing = change_date
            
            # Track QC Testing in Progress (active testing time)
            if new_status == QC_TESTING_IN_PROGRESS:
                if in_qc_testing and qc_testing_entry:
                    wait_days = calculate_business_days(qc_testing_entry, change_date)
                    waiting_times.append(wait_days)
                    total_waiting_in_queue_days += wait_days
                    all_waiting_times.append(wait_days)
                    in_qc_testing = False
                
                # Close any hold periods
                if in_dev_hold and dev_hold_entry:
                    total_dev_hold_days += calculate_business_days(dev_hold_entry, change_date)
                    in_dev_hold = False
                if in_qa_hold and qa_hold_entry:
                    total_qa_hold_days += calculate_business_days(qa_hold_entry, change_date)
                    in_qa_hold = False
                
                progress_entry = change_date
                in_progress = True
            
            # Track QA Hold (QC Testing Hold / On-hold)
            if new_status in (QC_TESTING_HOLD, QC_TESTING_ON_HOLD):
                if in_qc_testing and qc_testing_entry:
                    wait_days = calculate_business_days(qc_testing_entry, change_date)
                    waiting_times.append(wait_days)
                    total_waiting_in_queue_days += wait_days
                    all_waiting_times.append(wait_days)
                    in_qc_testing = False
                
                if in_progress and progress_entry:
                    total_active_testing_days += calculate_business_days(progress_entry, change_date)
                    in_progress = False
                
                qa_hold_entry = change_date
                in_qa_hold = True
            
            # Track Dev Hold (QC Review Fail / Tested - Awaiting Fixes)
            if new_status in QC_FAIL_STATUSES:
                if in_progress and progress_entry:
                    cycle_days = calculate_business_days(progress_entry, change_date)
                    total_active_testing_days += cycle_days
                    test_cycles.append({'days': cycle_days, 'result': 'Fail'})
                    all_test_cycles.append({'days': cycle_days, 'result': 'Fail'})
                    in_progress = False
                    progress_entry = None
                
                dev_hold_entry = change_date
                in_dev_hold = True
            
            # BIS Testing - close all periods
            if new_status == BIS_TESTING:
                if in_progress and progress_entry:
                    cycle_days = calculate_business_days(progress_entry, change_date)
                    total_active_testing_days += cycle_days
                    test_cycles.append({'days': cycle_days, 'result': 'Pass'})
                    all_test_cycles.append({'days': cycle_days, 'result': 'Pass'})
                    in_progress = False
                
                if in_dev_hold and dev_hold_entry:
                    total_dev_hold_days += calculate_business_days(dev_hold_entry, change_date)
                    in_dev_hold = False
                if in_qa_hold and qa_hold_entry:
                    total_qa_hold_days += calculate_business_days(qa_hold_entry, change_date)
                    in_qa_hold = False
        
        # QC Cycle Time
        qc_cycle_days = None
        if first_qc_testing and first_bis_testing:
            qc_cycle_days = calculate_business_days(first_qc_testing, first_bis_testing)
            daily_stats[first_bis_testing.strftime('%Y-%m-%d')].append(qc_cycle_days)
        
        # Cycle distribution
        num_cycles = len(test_cycles)
        if num_cycles > 0:
            if num_cycles == 1:
                cycle_distribution['1'] += 1
            elif num_cycles == 2:
                cycle_distribution['2'] += 1
            else:
                cycle_distribution['3+'] += 1
        
        ticket = ticket_lookup.get(ticket_id)
        tickets_data.append({
            'ticket_id': ticket_id,
            'current_status': history[-1]['new_status'] if history else '',
            'priority': getattr(ticket, 'priority', '') if ticket else '',
            'platform': getattr(ticket, 'subdepartment', '') if ticket else '',
            'qc_tester': getattr(ticket, 'qc_tester', '') if ticket else '',
            'first_qc_testing': first_qc_testing.isoformat() if first_qc_testing else None,
            'first_bis_testing': first_bis_testing.isoformat() if first_bis_testing else None,
            'qc_cycle_days': round(qc_cycle_days, 1) if qc_cycle_days else None,
            'test_cycles': num_cycles,
            'active_testing_days': round(total_active_testing_days, 1) if total_active_testing_days > 0 else None,
            'waiting_in_queue_days': round(total_waiting_in_queue_days, 1) if total_waiting_in_queue_days > 0 else None,
            'dev_hold_days': round(total_dev_hold_days, 1) if total_dev_hold_days > 0 else None,
            'qa_hold_days': round(total_qa_hold_days, 1) if total_qa_hold_days > 0 else None,
            'waiting_events': len(waiting_times),
            'avg_test_cycle_days': round(sum(c['days'] for c in test_cycles) / len(test_cycles), 1) if test_cycles else None,
            'avg_waiting_days': round(sum(waiting_times) / len(waiting_times), 1) if waiting_times else None,
        })
    
    # Aggregate metrics
    completed = [t for t in tickets_data if t['qc_cycle_days'] is not None]
    cycle_days_list = [t['qc_cycle_days'] for t in completed]
    tickets_with_cycles = [t for t in tickets_data if t['test_cycles'] > 0]
    one_cycle = len([t for t in tickets_with_cycles if t['test_cycles'] == 1])
    
    # Daily trend (last 14 days)
    daily_trend = []
    sorted_dates = sorted(daily_stats.keys())[-14:]
    for date in sorted_dates:
        vals = daily_stats[date]
        daily_trend.append({
            'date': date,
            'avg_days': round(sum(vals) / len(vals), 1) if vals else 0,
            'count': len(vals),
        })
    
    return {
        'metrics': {
            'total_tickets': len(tickets_data),
            'completed_tickets': len(completed),
            'avg_qc_cycle_days': round(sum(cycle_days_list) / len(cycle_days_list), 1) if cycle_days_list else 0,
            'median_qc_cycle_days': round(sorted(cycle_days_list)[len(cycle_days_list)//2], 1) if cycle_days_list else 0,
            'total_test_cycles': len(all_test_cycles),
            'avg_test_cycle_days': round(sum(c['days'] for c in all_test_cycles) / len(all_test_cycles), 1) if all_test_cycles else 0,
            'pass_cycles': len([c for c in all_test_cycles if c['result'] == 'Pass']),
            'fail_cycles': len([c for c in all_test_cycles if c['result'] == 'Fail']),
            'first_pass_rate': round(one_cycle / len(tickets_with_cycles) * 100, 1) if tickets_with_cycles else 0,
            'one_cycle_count': one_cycle,
            'multi_cycle_count': len(tickets_with_cycles) - one_cycle,
            'total_waiting_events': len(all_waiting_times),
            'avg_waiting_days': round(sum(all_waiting_times) / len(all_waiting_times), 1) if all_waiting_times else 0,
            'max_waiting_days': round(max(all_waiting_times), 1) if all_waiting_times else 0,
            'cycle_distribution': dict(cycle_distribution),
            'daily_trend': daily_trend,
        },
        'tickets': tickets_data,
    }


@app.get("/api/qa-metrics")
async def get_qa_metrics(
    period: str = Query("past_month", description="Time period"),
    start_date: str = Query(None, description="Start date YYYY-MM-DD"),
    end_date: str = Query(None, description="End date YYYY-MM-DD"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_reports_access)
):
    """Get QA metrics (authenticated)."""
    return _get_qa_metrics_data(period, db, start_date=start_date, end_date=end_date)


@app.get("/api/public/qa-metrics")
async def get_qa_metrics_public(
    period: str = Query("past_month", description="Time period"),
    start_date: str = Query(None, description="Start date YYYY-MM-DD"),
    end_date: str = Query(None, description="End date YYYY-MM-DD"),
    db: Session = Depends(get_db)
):
    """Get QA metrics (public - no auth required)."""
    return _get_qa_metrics_data(period, db, start_date=start_date, end_date=end_date)


def _export_qa_metrics_tickets_excel(period: str, db: Session, start_date: str = None, end_date: str = None):
    """Create and return Excel export path for QA metrics tickets."""
    from pathlib import Path
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    data = _get_qa_metrics_data(period, db, start_date=start_date, end_date=end_date)
    tickets = data.get("tickets", [])
    metrics = data.get("metrics", {})

    wb = Workbook()
    # Set calculation mode to automatic
    wb.calculation.calcMode = "auto"
    
    ws_summary = wb.active
    ws_summary.title = "Summary"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    title = "QA Metrics Ticket Export"
    ws_summary["A1"] = title
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    if start_date or end_date:
        ws_summary["A4"] = f"Range: {start_date or '...'} to {end_date or '...'}"
    else:
        ws_summary["A4"] = f"Period: {period.replace('_', ' ').title()}"

    # Summary metrics with formulas referencing Tickets sheet
    # Note: Tickets data starts at row 7, columns: A=ID, F=QC Start, G=BIS Date, H=Total QC Days, 
    # I=Active Testing, J=Queue Wait, K=Dev Hold, L=QA Hold, M=Test Cycles, N=Avg Test Cycle, O=Waiting Events, P=Avg Waiting
    
    ws_summary["A6"] = "Metric"
    ws_summary["B6"] = "Value"
    ws_summary["C6"] = "Formula"
    ws_summary["A6"].fill = header_fill
    ws_summary["B6"].fill = header_fill
    ws_summary["C6"].fill = header_fill
    ws_summary["A6"].font = header_font
    ws_summary["B6"].font = header_font
    ws_summary["C6"].font = header_font
    
    # Calculate data range for Tickets sheet
    first_row = 7
    last_row = first_row + len(tickets) - 1 if tickets else first_row
    
    # Row 7: Total Tickets
    ws_summary["A7"] = "Total Tickets"
    ws_summary["B7"] = f"=COUNTA(Tickets!A{first_row}:A{last_row})" if tickets else 0
    ws_summary["C7"] = f"COUNTA(Tickets!A{first_row}:A{last_row})"
    
    # Row 8: Completed Tickets (those with BIS Testing date in column G)
    ws_summary["A8"] = "Completed Tickets"
    ws_summary["B8"] = f"=COUNTA(Tickets!G{first_row}:G{last_row})" if tickets else 0
    ws_summary["C8"] = f"COUNTA(Tickets!G{first_row}:G{last_row})"
    
    # Row 9: Avg QC Cycle Days (column H)
    ws_summary["A9"] = "Avg QC Cycle Days"
    ws_summary["B9"] = f'=IF(COUNTA(Tickets!H{first_row}:H{last_row})>0,ROUND(AVERAGE(Tickets!H{first_row}:H{last_row}),1),0)' if tickets else 0
    ws_summary["C9"] = f"ROUND(AVERAGE(Tickets!H{first_row}:H{last_row}),1)"
    
    # Row 10: Median QC Cycle Days (column H)
    ws_summary["A10"] = "Median QC Cycle Days"
    ws_summary["B10"] = f'=IF(COUNTA(Tickets!H{first_row}:H{last_row})>0,MEDIAN(Tickets!H{first_row}:H{last_row}),0)' if tickets else 0
    ws_summary["C10"] = f"MEDIAN(Tickets!H{first_row}:H{last_row})"
    
    # Row 11: Total Test Cycles (sum of column M)
    ws_summary["A11"] = "Total Test Cycles"
    ws_summary["B11"] = f"=SUM(Tickets!M{first_row}:M{last_row})" if tickets else 0
    ws_summary["C11"] = f"SUM(Tickets!M{first_row}:M{last_row})"
    
    # Row 12: Avg Test Cycles per Ticket
    ws_summary["A12"] = "Avg Test Cycles per Ticket"
    ws_summary["B12"] = f'=IF(COUNTA(Tickets!M{first_row}:M{last_row})>0,ROUND(AVERAGE(Tickets!M{first_row}:M{last_row}),1),0)' if tickets else 0
    ws_summary["C12"] = f"ROUND(AVERAGE(Tickets!M{first_row}:M{last_row}),1)"
    
    # Row 13: First Pass Rate % (tickets with exactly 1 test cycle)
    ws_summary["A13"] = "First Pass Rate %"
    ws_summary["B13"] = f'=IF(COUNTA(Tickets!M{first_row}:M{last_row})>0,ROUND(COUNTIF(Tickets!M{first_row}:M{last_row},1)/COUNTA(Tickets!M{first_row}:M{last_row})*100,1),0)' if tickets else 0
    ws_summary["C13"] = f"COUNTIF(M=1) / COUNT(M) * 100"
    
    # Row 14: Avg Active Testing Days (column I)
    ws_summary["A14"] = "Avg Active Testing Days"
    ws_summary["B14"] = f'=IF(COUNTA(Tickets!I{first_row}:I{last_row})>0,ROUND(AVERAGE(Tickets!I{first_row}:I{last_row}),1),0)' if tickets else 0
    ws_summary["C14"] = f"ROUND(AVERAGE(Tickets!I{first_row}:I{last_row}),1)"
    
    # Row 15: Avg Queue Wait Days (column J)
    ws_summary["A15"] = "Avg Queue Wait Days"
    ws_summary["B15"] = f'=IF(COUNTA(Tickets!J{first_row}:J{last_row})>0,ROUND(AVERAGE(Tickets!J{first_row}:J{last_row}),1),0)' if tickets else 0
    ws_summary["C15"] = f"ROUND(AVERAGE(Tickets!J{first_row}:J{last_row}),1)"
    
    # Row 16: Avg Dev Hold Days (column K)
    ws_summary["A16"] = "Avg Dev Hold Days"
    ws_summary["B16"] = f'=IF(COUNTA(Tickets!K{first_row}:K{last_row})>0,ROUND(AVERAGE(Tickets!K{first_row}:K{last_row}),1),0)' if tickets else 0
    ws_summary["C16"] = f"ROUND(AVERAGE(Tickets!K{first_row}:K{last_row}),1)"
    
    # Row 17: Avg QA Hold Days (column L)
    ws_summary["A17"] = "Avg QA Hold Days"
    ws_summary["B17"] = f'=IF(COUNTA(Tickets!L{first_row}:L{last_row})>0,ROUND(AVERAGE(Tickets!L{first_row}:L{last_row}),1),0)' if tickets else 0
    ws_summary["C17"] = f"ROUND(AVERAGE(Tickets!L{first_row}:L{last_row}),1)"
    
    # Row 18: Max QC Cycle Days
    ws_summary["A18"] = "Max QC Cycle Days"
    ws_summary["B18"] = f'=IF(COUNTA(Tickets!H{first_row}:H{last_row})>0,MAX(Tickets!H{first_row}:H{last_row}),0)' if tickets else 0
    ws_summary["C18"] = f"MAX(Tickets!H{first_row}:H{last_row})"
    
    # Row 19: Min QC Cycle Days
    ws_summary["A19"] = "Min QC Cycle Days"
    ws_summary["B19"] = f'=IF(COUNTA(Tickets!H{first_row}:H{last_row})>0,MIN(Tickets!H{first_row}:H{last_row}),0)' if tickets else 0
    ws_summary["C19"] = f"MIN(Tickets!H{first_row}:H{last_row})"
    
    ws_summary.column_dimensions["A"].width = 28
    ws_summary.column_dimensions["B"].width = 16
    ws_summary.column_dimensions["C"].width = 45

    ws = wb.create_sheet("Tickets")
    
    headers = [
        "Ticket ID",
        "Current Status",
        "Priority",
        "Platform",
        "QC Tester",
        "First QC Testing",
        "First BIS Testing",
        "Total QC Days",
        "Active Testing Days",
        "Queue Wait Days",
        "Dev Hold Days",
        "QA Hold Days",
        "Test Cycles",
        "Avg Test Cycle Days",
        "Waiting Events",
        "Avg Waiting Days",
    ]

    header_row = 6
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.fill = header_fill
        cell.font = header_font

    first_data_row = header_row + 1
    last_data_row = header_row + len(tickets)

    # We'll add raw duration columns (Q-T) for formula-based calculations
    # Q = Raw Active Testing Days, R = Raw Queue Wait Days, S = Raw Dev Hold Days, T = Raw QA Hold Days
    # U = Raw Test Cycles, V = Raw Waiting Events
    extra_headers = ["Raw Active", "Raw Queue", "Raw Dev Hold", "Raw QA Hold", "Raw Cycles", "Raw Wait Events"]
    for col_offset, label in enumerate(extra_headers, start=17):
        cell = ws.cell(row=header_row, column=col_offset, value=label)
        cell.fill = header_fill
        cell.font = header_font
    
    for idx, t in enumerate(tickets, start=first_data_row):
        ws.cell(row=idx, column=1, value=t.get("ticket_id"))
        ws.cell(row=idx, column=2, value=t.get("current_status"))
        ws.cell(row=idx, column=3, value=t.get("priority"))
        ws.cell(row=idx, column=4, value=t.get("platform"))
        ws.cell(row=idx, column=5, value=t.get("qc_tester"))
        start_val = t.get("first_qc_testing")
        end_val = t.get("first_bis_testing")
        start_dt = datetime.fromisoformat(start_val) if start_val else None
        end_dt = datetime.fromisoformat(end_val) if end_val else None
        ws.cell(row=idx, column=6, value=start_dt)  # F - First QC Testing
        ws.cell(row=idx, column=7, value=end_dt)    # G - First BIS Testing
        
        # H - Total QC Days (formula)
        ws.cell(row=idx, column=8, value=f'=IF(AND(F{idx}<>"",G{idx}<>""),NETWORKDAYS(F{idx},G{idx}),"")')
        
        # I - Active Testing Days (formula referencing raw data in Q)
        ws.cell(row=idx, column=9, value=f'=IF(Q{idx}<>"",Q{idx},"")')
        
        # J - Queue Wait Days (formula referencing raw data in R)
        ws.cell(row=idx, column=10, value=f'=IF(R{idx}<>"",R{idx},"")')
        
        # K - Dev Hold Days (formula referencing raw data in S)
        ws.cell(row=idx, column=11, value=f'=IF(S{idx}<>"",S{idx},"")')
        
        # L - QA Hold Days (formula referencing raw data in T)
        ws.cell(row=idx, column=12, value=f'=IF(T{idx}<>"",T{idx},"")')
        
        # M - Test Cycles (formula referencing raw data in U)
        ws.cell(row=idx, column=13, value=f'=IF(U{idx}<>"",U{idx},"")')
        
        # N - Avg Test Cycle Days (formula: Active Testing / Cycles)
        ws.cell(row=idx, column=14, value=f'=IF(AND(U{idx}<>"",U{idx}>0),ROUND(Q{idx}/U{idx},1),"")')
        
        # O - Waiting Events (formula referencing raw data in V)
        ws.cell(row=idx, column=15, value=f'=IF(V{idx}<>"",V{idx},"")')
        
        # P - Avg Waiting Days (formula: Queue Wait / Waiting Events)
        ws.cell(row=idx, column=16, value=f'=IF(AND(V{idx}<>"",V{idx}>0),ROUND(R{idx}/V{idx},1),"")')
        
        # Raw data columns (Q-V) - these hold the actual calculated values
        ws.cell(row=idx, column=17, value=t.get("active_testing_days") or "")     # Q - Raw Active
        ws.cell(row=idx, column=18, value=t.get("waiting_in_queue_days") or "")   # R - Raw Queue
        ws.cell(row=idx, column=19, value=t.get("dev_hold_days") or "")           # S - Raw Dev Hold
        ws.cell(row=idx, column=20, value=t.get("qa_hold_days") or "")            # T - Raw QA Hold
        ws.cell(row=idx, column=21, value=t.get("test_cycles") or "")             # U - Raw Cycles
        ws.cell(row=idx, column=22, value=t.get("waiting_events") or "")

    for col, width in {
        "A": 12, "B": 24, "C": 14, "D": 18, "E": 18,
        "F": 22, "G": 22, "H": 14, "I": 18, "J": 16,
        "K": 14, "L": 14, "M": 12, "N": 18, "O": 14, "P": 18,
        "Q": 12, "R": 12, "S": 14, "T": 14, "U": 12, "V": 14
    }.items():
        ws.column_dimensions[col].width = width
    for row in range(first_data_row, last_data_row + 1):
        ws.cell(row=row, column=6).number_format = "YYYY-MM-DD HH:MM"
        ws.cell(row=row, column=7).number_format = "YYYY-MM-DD HH:MM"

    ws["R2"] = "Stats Formulas"
    ws["R3"] = "Avg QC Days (col H)"
    if tickets:
        ws["S3"] = f'=IF(COUNTA(H{first_data_row}:H{last_data_row})>0,AVERAGE(H{first_data_row}:H{last_data_row}),0)'
    else:
        ws["S3"] = 0
    ws["R4"] = "Median QC Days (col H)"
    if tickets:
        ws["S4"] = f'=IF(COUNTA(H{first_data_row}:H{last_data_row})>0,MEDIAN(H{first_data_row}:H{last_data_row}),0)'
    else:
        ws["S4"] = 0
    ws["R5"] = "First Pass Rate % (col M)"
    if tickets:
        ws["S5"] = f'=IF(COUNTA(M{first_data_row}:M{last_data_row})>0,COUNTIF(M{first_data_row}:M{last_data_row},1)/COUNTA(M{first_data_row}:M{last_data_row})*100,0)'
    else:
        ws["S5"] = 0

    output_path = Path("reports") / f"QA_Metrics_Tickets_{period}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path.parent.mkdir(exist_ok=True)
    wb.save(output_path)
    return output_path


@app.get("/api/qa-metrics/export")
async def export_qa_metrics(
    period: str = Query("past_month", description="Time period"),
    start_date: str = Query(None, description="Start date YYYY-MM-DD"),
    end_date: str = Query(None, description="End date YYYY-MM-DD"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_reports_access)
):
    """Export QA metrics ticket list to Excel (authenticated)."""
    from fastapi.responses import FileResponse

    file_path = _export_qa_metrics_tickets_excel(period, db, start_date=start_date, end_date=end_date)
    return FileResponse(
        path=str(file_path),
        filename=file_path.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/api/public/qa-metrics/export")
async def export_qa_metrics_public(
    period: str = Query("past_month", description="Time period"),
    start_date: str = Query(None, description="Start date YYYY-MM-DD"),
    end_date: str = Query(None, description="End date YYYY-MM-DD"),
    db: Session = Depends(get_db)
):
    """Export QA metrics ticket list to Excel (public - no auth required)."""
    from fastapi.responses import FileResponse

    file_path = _export_qa_metrics_tickets_excel(period, db, start_date=start_date, end_date=end_date)
    return FileResponse(
        path=str(file_path),
        filename=file_path.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _get_qa_metrics_comparison_data(db: Session):
    """Internal function for QA metrics comparison."""
    from pathlib import Path
    from collections import defaultdict
    import json
    
    PERIOD_DAYS = {
        'past_week': 7,
        'past_month': 30,
        'past_quarter': 90,
        'past_year': 365,
        'overall': None,
    }
    
    QC_TESTING = 'QC Testing'
    QC_TESTING_IN_PROGRESS = 'QC Testing in Progress'
    BIS_TESTING = 'BIS Testing'
    
    reports_dir = Path("reports")
    export_files = list(reports_dir.glob("PM_Activity_Export_*.csv"))
    if not export_files:
        return {}
    
    latest_export = max(export_files, key=lambda f: f.stat().st_mtime)
    
    try:
        with open(latest_export, 'r', encoding='utf-8') as f:
            raw_data = json.load(f)
    except:
        return {}
    
    ticket_history = defaultdict(list)
    for record in raw_data:
        ticket_id = record.get('ticketId')
        if ticket_id:
            try:
                ticket_id = int(ticket_id)
            except:
                pass
            change_date = datetime.strptime(record['statusChangeDate'], '%Y-%m-%d %H:%M:%S')
            ticket_history[ticket_id].append({
                'date': change_date,
                'new_status': record.get('newStatus'),
            })
    
    for tid in ticket_history:
        ticket_history[tid].sort(key=lambda x: x['date'])
    
    def calculate_business_days(start, end):
        if not start or not end or end <= start:
            return 0.0
        total = 0
        current = start
        while current < end:
            if current.weekday() < 5:
                total += 1
            current += timedelta(days=1)
        return total
    
    def ticket_entered_qa(history):
        for h in history:
            if h.get('new_status') in (QC_TESTING, QC_TESTING_IN_PROGRESS):
                return True
        return False
    
    result = {}
    now = datetime.now()
    
    for period_key, days in PERIOD_DAYS.items():
        cutoff = now - timedelta(days=days) if days else None
        
        total = 0
        completed = 0
        cycle_days_list = []
        first_pass = 0
        tickets_with_cycles = 0
        
        for ticket_id, history in ticket_history.items():
            if not ticket_entered_qa(history):
                continue
            
            if cutoff:
                if not any(h['date'] >= cutoff for h in history):
                    continue
            
            total += 1
            
            first_qc = None
            first_bis = None
            cycle_count = 0
            
            for h in history:
                if h['new_status'] == QC_TESTING and first_qc is None:
                    first_qc = h['date']
                if h['new_status'] == BIS_TESTING and first_bis is None:
                    first_bis = h['date']
                if h['new_status'] == QC_TESTING_IN_PROGRESS:
                    cycle_count += 1
            
            if first_qc and first_bis:
                completed += 1
                days_val = calculate_business_days(first_qc, first_bis)
                cycle_days_list.append(days_val)
            
            if cycle_count > 0:
                tickets_with_cycles += 1
                if cycle_count == 1:
                    first_pass += 1
        
        result[period_key] = {
            'total_tickets': total,
            'completed_tickets': completed,
            'avg_qc_cycle_days': round(sum(cycle_days_list) / len(cycle_days_list), 1) if cycle_days_list else 0,
            'first_pass_rate': round(first_pass / tickets_with_cycles * 100, 1) if tickets_with_cycles else 0,
        }
    
    return result


@app.get("/api/qa-metrics/comparison")
async def get_qa_metrics_comparison(
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_reports_access)
):
    """Get QA metrics comparison (authenticated)."""
    return _get_qa_metrics_comparison_data(db)


@app.get("/api/public/qa-metrics/comparison")
async def get_qa_metrics_comparison_public(
    db: Session = Depends(get_db)
):
    """Get QA metrics comparison (public - no auth required)."""
    return _get_qa_metrics_comparison_data(db)


# ===== QA CYCLE TIME DASHBOARD =====

@app.get("/reports/qa-metrics/download")
async def download_qa_metrics_excel(
    start_date: str = Query(None, description="Filter start date (YYYY-MM-DD)"),
    end_date: str = Query(None, description="Filter end date (YYYY-MM-DD)"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_reports_access)
):
    """
    Generate and download QA Metrics Excel with the 4 key metrics:
    1. QC Cycle Time (Overall) - Days from QC Testing to BIS Testing
    2. Test Cycle Time - Days per test cycle
    3. Number of Testing Cycles - Count of loops
    4. QC Waiting Time - Days waiting in queue
    
    All calculated columns use Excel formulas (highlighted in yellow).
    """
    from fastapi.responses import FileResponse
    from pathlib import Path
    from collections import defaultdict
    import json
    from qa_metrics_excel import generate_qa_metrics_excel
    
    # Parse date filters
    filter_start = None
    filter_end = None
    if start_date:
        try:
            filter_start = datetime.strptime(start_date, '%Y-%m-%d')
        except ValueError:
            pass
    if end_date:
        try:
            filter_end = datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
        except ValueError:
            pass

    # Find the latest PM Activity Export file
    reports_dir = Path("reports")
    export_files = list(reports_dir.glob("PM_Activity_Export_*.csv"))
    if not export_files:
        raise HTTPException(status_code=404, detail="No PM Activity Export file found. Please run the data fetch script first.")
    
    latest_export = max(export_files, key=lambda f: f.stat().st_mtime)
    logger.info(f"Generating QA Metrics Excel from: {latest_export}")
    
    # Load JSON data from the export file
    try:
        with open(latest_export, 'r', encoding='utf-8') as f:
            raw_data = json.load(f)
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=500, detail=f"Failed to parse PM Activity Export file: {e}")
    
    # Group by ticket (with optional date filtering)
    ticket_history = defaultdict(list)
    for record in raw_data:
        ticket_id = record.get('ticketId')
        if ticket_id:
            try:
                ticket_id = int(ticket_id)
            except (ValueError, TypeError):
                pass
            
            change_date = datetime.strptime(record['statusChangeDate'], '%Y-%m-%d %H:%M:%S')
            
            # Apply date filter
            if filter_start and change_date < filter_start:
                continue
            if filter_end and change_date > filter_end:
                continue
                
            ticket_history[ticket_id].append({
                'date': change_date,
                'old_status': record.get('oldStatus'),
                'new_status': record.get('newStatus'),
            })
    
    # Sort each ticket's history
    for tid in ticket_history:
        ticket_history[tid].sort(key=lambda x: x['date'])
    
    # Get ticket metadata from database for enrichment
    tickets = db.query(TicketTracking).all()
    ticket_lookup = {t.ticket_id: t for t in tickets}
    
    # Generate Excel with 4 key metrics
    date_suffix = ""
    if filter_start or filter_end:
        if start_date and end_date:
            date_suffix = f"_{start_date}_to_{end_date}"
        elif start_date:
            date_suffix = f"_from_{start_date}"
        elif end_date:
            date_suffix = f"_until_{end_date}"
    output_path = Path("reports") / f"QA_Metrics_4Key{date_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    try:
        generated_path = generate_qa_metrics_excel(
            ticket_history=ticket_history,
            ticket_lookup=ticket_lookup,
            output_path=output_path
        )
        
        return FileResponse(
            path=str(generated_path),
            filename=generated_path.name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        logger.error(f"Failed to generate QA Metrics Excel: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to generate metrics: {str(e)}")


@app.get("/reports/qa-dashboard/download")
async def download_qa_dashboard(
    start_date: str = Query(None, description="Filter start date (YYYY-MM-DD)"),
    end_date: str = Query(None, description="Filter end date (YYYY-MM-DD)"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_reports_access)
):
    """
    Generate and download QA Dashboard Excel file with embedded formulas.
    Reads from PM Activity Export JSON file for status history.
    Supports optional date range filtering.
    """
    from fastapi.responses import FileResponse
    from pathlib import Path
    from collections import defaultdict
    import json
    from qa_dashboard_excel import generate_qa_dashboard_excel
    
    # Parse date filters
    filter_start = None
    filter_end = None
    if start_date:
        try:
            filter_start = datetime.strptime(start_date, '%Y-%m-%d')
        except ValueError:
            pass
    if end_date:
        try:
            filter_end = datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
        except ValueError:
            pass

    # Find the latest PM Activity Export file
    reports_dir = Path("reports")
    export_files = list(reports_dir.glob("PM_Activity_Export_*.csv"))
    if not export_files:
        raise HTTPException(status_code=404, detail="No PM Activity Export file found. Please run the data fetch script first.")
    
    latest_export = max(export_files, key=lambda f: f.stat().st_mtime)
    logger.info(f"Generating QA Dashboard Excel from: {latest_export}")
    
    # Load JSON data from the export file
    try:
        with open(latest_export, 'r', encoding='utf-8') as f:
            raw_data = json.load(f)
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=500, detail=f"Failed to parse PM Activity Export file: {e}")
    
    # Group by ticket (with optional date filtering)
    ticket_history = defaultdict(list)
    for record in raw_data:
        ticket_id = record.get('ticketId')
        if ticket_id:
            try:
                ticket_id = int(ticket_id)
            except (ValueError, TypeError):
                pass
            
            change_date = datetime.strptime(record['statusChangeDate'], '%Y-%m-%d %H:%M:%S')
            
            # Apply date filter
            if filter_start and change_date < filter_start:
                continue
            if filter_end and change_date > filter_end:
                continue
                
            ticket_history[ticket_id].append({
                'date': change_date,
                'old_status': record.get('oldStatus'),
                'new_status': record.get('newStatus'),
            })
    
    # Sort each ticket's history
    for tid in ticket_history:
        ticket_history[tid].sort(key=lambda x: x['date'])
    
    # Get ticket metadata from database for enrichment
    tickets = db.query(TicketTracking).all()
    ticket_lookup = {t.ticket_id: t for t in tickets}
    
    # Generate Excel with formulas (include date range in filename if filtered)
    date_suffix = ""
    if filter_start or filter_end:
        if start_date and end_date:
            date_suffix = f"_{start_date}_to_{end_date}"
        elif start_date:
            date_suffix = f"_from_{start_date}"
        elif end_date:
            date_suffix = f"_until_{end_date}"
    output_path = Path("reports") / f"QA_Dashboard{date_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    try:
        generated_path = generate_qa_dashboard_excel(
            ticket_history=ticket_history,
            ticket_lookup=ticket_lookup,
            output_path=output_path
        )
        
        return FileResponse(
            path=str(generated_path),
            filename=generated_path.name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        logger.error(f"Failed to generate QA Dashboard Excel: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to generate dashboard: {str(e)}")


@app.get("/reports/production-bugs/download")
async def download_production_bugs_report(
    format: str = Query("excel", description="Report format: 'excel' or 'pdf'"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_reports_access)
):
    """
    Generate and download Production & Pre-Production Bugs Report.
    
    This report includes:
    - Bugs found in Production, Pre-production, and BIS Testing environments
    - Ticket-wise details with developers, testers, and time tracking
    - Statistics by environment, severity, and module
    
    Parameters:
    - format: 'excel' for XLSX file, 'pdf' for PDF file
    """
    from fastapi.responses import FileResponse
    from pathlib import Path
    import subprocess
    import sys
    
    try:
        # Run the report generation script
        script_path = Path(__file__).parent / "scripts" / "generate_prod_bugs_report.py"
        
        if not script_path.exists():
            raise HTTPException(status_code=500, detail="Report generation script not found")
        
        # Execute the script
        result = subprocess.run(
            [sys.executable, str(script_path)],
            capture_output=True,
            text=True,
            cwd=str(Path(__file__).parent)
        )
        
        if result.returncode != 0:
            logger.error(f"Report generation failed: {result.stderr}")
            raise HTTPException(status_code=500, detail=f"Report generation failed: {result.stderr[:200]}")
        
        # Parse output to get file paths
        output_lines = result.stdout.strip().split('\n')
        excel_file = None
        pdf_file = None
        
        for line in output_lines:
            if 'Excel:' in line or 'Excel Report:' in line:
                excel_file = line.split(':', 1)[-1].strip()
            elif 'PDF:' in line or 'PDF Report:' in line:
                pdf_file = line.split(':', 1)[-1].strip()
        
        # Select file based on format
        if format.lower() == 'pdf':
            if not pdf_file or not Path(pdf_file).exists():
                raise HTTPException(status_code=500, detail="PDF report was not generated")
            file_path = Path(pdf_file)
            media_type = "application/pdf"
        else:
            if not excel_file or not Path(excel_file).exists():
                raise HTTPException(status_code=500, detail="Excel report was not generated")
            file_path = Path(excel_file)
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        
        return FileResponse(
            path=str(file_path),
            filename=file_path.name,
            media_type=media_type
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to generate Production Bugs report: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to generate report: {str(e)}")


@app.get("/reports/production-bugs/preview")
async def preview_production_bugs_report(
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_reports_access)
):
    """
    Get preview/summary data for Production & Pre-Production Bugs Report.
    Returns statistics without generating the full report files.
    """
    from collections import defaultdict
    
    REPORT_ENVIRONMENTS = ['Production', 'Pre-production', 'BIS Testing (Pre)']
    
    # Query bugs in Production and Pre-Production environments
    bugs = db.query(Bug).filter(
        Bug.environment.in_(REPORT_ENVIRONMENTS)
    ).all()
    
    # Get unique ticket IDs
    ticket_ids = list(set(b.ticket_id for b in bugs if b.ticket_id))
    
    # Get ticket details
    tickets = {}
    if ticket_ids:
        ticket_records = db.query(TicketTracking).filter(
            TicketTracking.ticket_id.in_(ticket_ids)
        ).all()
        tickets = {t.ticket_id: t for t in ticket_records}
    
    # Filter out tickets with N/A titles
    def is_valid_ticket(ticket):
        if not ticket or not ticket.title:
            return False
        title_lower = ticket.title.strip().lower()
        return title_lower not in ['n/a', 'na', '-', '']
    
    valid_ticket_ids = [tid for tid in ticket_ids if is_valid_ticket(tickets.get(tid))]
    valid_bugs = [b for b in bugs if b.ticket_id in valid_ticket_ids]
    
    # Calculate statistics
    env_stats = defaultdict(lambda: {"total": 0, "open": 0, "closed": 0})
    sev_stats = defaultdict(lambda: {"prod": 0, "preprod": 0})
    
    for bug in valid_bugs:
        env = bug.environment or "Unknown"
        env_stats[env]["total"] += 1
        if bug.status and bug.status.lower() in ['closed', 'resolved', 'rejected']:
            env_stats[env]["closed"] += 1
        else:
            env_stats[env]["open"] += 1
        
        sev = bug.severity or "Unknown"
        if bug.environment == 'Production':
            sev_stats[sev]["prod"] += 1
        else:
            sev_stats[sev]["preprod"] += 1
    
    return {
        "total_bugs": len(valid_bugs),
        "tickets_affected": len(valid_ticket_ids),
        "excluded_tickets": len(ticket_ids) - len(valid_ticket_ids),
        "environment_breakdown": dict(env_stats),
        "severity_breakdown": dict(sev_stats),
        "generated_at": datetime.now().isoformat()
    }


@app.get("/api/qa-dashboard/metrics")
def get_qa_dashboard_metrics(
    start_date: str = Query(None, description="Filter start date (YYYY-MM-DD)"),
    end_date: str = Query(None, description="Filter end date (YYYY-MM-DD)"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_reports_access)
):
    """
    Get QA cycle time metrics for the dashboard.
    Reads from PM Activity Export JSON file for status history,
    and enriches with ticket metadata from TicketTracking table.
    Supports optional date range filtering.
    """
    from collections import Counter, defaultdict
    from datetime import datetime, timedelta
    
    # Parse date filters
    filter_start = None
    filter_end = None
    if start_date:
        try:
            filter_start = datetime.strptime(start_date, '%Y-%m-%d')
        except ValueError:
            pass
    if end_date:
        try:
            filter_end = datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
        except ValueError:
            pass
    from pathlib import Path
    import json
    
    # Status definitions - tracking all existing PM statuses
    # QA-related statuses (tickets actively in QA)
    QA_STATUSES = {'QC Testing', 'QC Testing in Progress', 'QC Review Fail', 'QC Testing On-hold', 'QC Testing Hold', 'Tested - Awaiting Fixes'}
    QA_START_STATUSES = {'QC Testing', 'QC Testing in Progress'}  # When ticket enters QA
    QA_END_STATUSES = {'BIS Testing', 'Closed', 'Approved for Live', 'Moved to Live'}  # When ticket exits QA successfully
    QA_HOLD_STATUSES = {'QC Testing On-hold', 'QC Testing Hold', 'Hold/Pending'}  # Hold statuses (time not counted)
    QA_FAIL_STATUSES = {'QC Review Fail', 'Tested - Awaiting Fixes'}  # Failed QA review
    
    # Find the latest PM Activity Export file
    reports_dir = Path("reports")
    export_files = list(reports_dir.glob("PM_Activity_Export_*.csv"))
    if not export_files:
        raise HTTPException(status_code=404, detail="No PM Activity Export file found. Please run the data fetch script first.")
    
    latest_export = max(export_files, key=lambda f: f.stat().st_mtime)
    logger.info(f"Loading QA dashboard data from: {latest_export}")
    
    # Load JSON data from the export file
    try:
        with open(latest_export, 'r', encoding='utf-8') as f:
            raw_data = json.load(f)
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=500, detail=f"Failed to parse PM Activity Export file: {e}")
    
    # Group by ticket (with optional date filtering)
    ticket_history = defaultdict(list)
    for record in raw_data:
        ticket_id = record.get('ticketId')
        if ticket_id:
            try:
                ticket_id = int(ticket_id)
            except (ValueError, TypeError):
                pass
            
            change_date = datetime.strptime(record['statusChangeDate'], '%Y-%m-%d %H:%M:%S')
            
            # Apply date filter - include ticket if any activity is in range
            if filter_start and change_date < filter_start:
                continue
            if filter_end and change_date > filter_end:
                continue
                
            ticket_history[ticket_id].append({
                'date': change_date,
                'old_status': record.get('oldStatus'),
                'new_status': record.get('newStatus'),
            })
    
    # Sort each ticket's history by date
    for tid in ticket_history:
        ticket_history[tid].sort(key=lambda x: x['date'])
    
    # Get ticket metadata from database for enrichment
    tickets = db.query(TicketTracking).all()
    ticket_lookup = {t.ticket_id: t for t in tickets}
    
    # Calculate metrics
    results = []
    status_counts = Counter()
    monthly_qa_times = defaultdict(list)
    platform_qa_times = defaultdict(list)
    priority_qa_times = defaultdict(list)
    cycle_distribution = defaultdict(list)
    qc_tester_qa_times = defaultdict(list)
    developer_qa_times = defaultdict(list)
    
    for ticket_id, history in ticket_history.items():
        if not history:
            continue
            
        current_status = history[-1]['new_status'] if history else None
        status_counts[current_status] += 1
        
        qa_start = None
        qa_end = None
        qa_cycles = 0
        fail_count = 0
        total_hold_hours = 0.0
        hold_start = None
        
        for h in history:
            if h['new_status'] in QA_START_STATUSES:
                if qa_start is None:
                    qa_start = h['date']
                qa_cycles += 1
            
            if h['new_status'] in QA_END_STATUSES and qa_start is not None:
                qa_end = h['date']
            
            if h['new_status'] in QA_HOLD_STATUSES:
                hold_start = h['date']
            elif hold_start is not None and h['new_status'] not in QA_HOLD_STATUSES:
                hold_duration = (h['date'] - hold_start).total_seconds() / 3600
                total_hold_hours += hold_duration
                hold_start = None
            
            if h['new_status'] in QA_FAIL_STATUSES:
                fail_count += 1
        
        # Calculate QA business days
        qa_business_days = None
        if qa_start and qa_end:
            gross_hours = (qa_end - qa_start).total_seconds() / 3600
            net_hours = max(gross_hours - total_hold_hours, 0)
            qa_business_days = net_hours / 8
            
            # Track by month
            month_key = qa_end.strftime('%Y-%m')
            monthly_qa_times[month_key].append(qa_business_days)
            
            # Get enrichment from ticket_lookup
            ticket = ticket_lookup.get(ticket_id)
            if ticket:
                platform = ticket.subdepartment or 'Unknown'
                priority = ticket.priority or 'Unknown'
                platform_qa_times[platform].append(qa_business_days)
                priority_qa_times[priority].append(qa_business_days)
                
                # Track by QC tester
                qc_tester = ticket.qc_tester or 'Unassigned'
                qc_tester_qa_times[qc_tester].append(qa_business_days)
                
                # Track by developer
                backend_dev = ticket.backend_developer or ''
                frontend_dev = ticket.frontend_developer or ''
                if backend_dev:
                    developer_qa_times[backend_dev].append(qa_business_days)
                if frontend_dev and frontend_dev != backend_dev:
                    developer_qa_times[frontend_dev].append(qa_business_days)
            
            # Track by cycle count
            cycle_key = '1 cycle' if qa_cycles == 1 else ('2 cycles' if qa_cycles == 2 else '3+ cycles')
            cycle_distribution[cycle_key].append(qa_business_days)
        
        ticket = ticket_lookup.get(ticket_id)
        results.append({
            'ticket_id': ticket_id,
            'title': ticket.title if ticket else '',
            'current_status': current_status,
            'priority': ticket.priority if ticket else '',
            'subdepartment': ticket.subdepartment if ticket else '',
            'qc_tester': ticket.qc_tester if ticket else '',
            'qa_business_days': round(qa_business_days, 2) if qa_business_days else None,
            'qa_cycles': qa_cycles,
            'qa_fail_count': fail_count,
            'qa_hold_hours': round(total_hold_hours, 2),
        })
    
    # Calculate summary stats
    qa_completed = [r for r in results if r['qa_business_days'] is not None]
    avg_qa_days = sum(r['qa_business_days'] for r in qa_completed) / len(qa_completed) if qa_completed else 0
    
    sorted_days = sorted([r['qa_business_days'] for r in qa_completed])
    median_qa_days = sorted_days[len(sorted_days)//2] if sorted_days else 0
    
    total_fails = sum(r['qa_fail_count'] for r in results)
    avg_cycles = sum(r['qa_cycles'] for r in results) / len(results) if results else 0
    first_pass = len([r for r in qa_completed if r['qa_cycles'] == 1])
    first_pass_rate = (first_pass / len(qa_completed) * 100) if qa_completed else 0
    
    # Platform breakdown
    platform_breakdown = []
    for platform, times in sorted(platform_qa_times.items(), key=lambda x: -len(x[1])):
        if times:
            platform_breakdown.append({
                'platform': platform,
                'tickets': len(times),
                'avg_days': round(sum(times) / len(times), 1),
                'total_days': round(sum(times), 1),
            })
    
    # Priority breakdown
    priority_breakdown = []
    priority_order = ['URGENT', 'High (Bugs)', 'High', 'Medium', 'Low']
    for priority, times in sorted(priority_qa_times.items(), 
                                   key=lambda x: priority_order.index(x[0]) if x[0] in priority_order else 99):
        if times:
            priority_breakdown.append({
                'priority': priority,
                'tickets': len(times),
                'avg_days': round(sum(times) / len(times), 1),
                'total_days': round(sum(times), 1),
            })
    
    # Cycle distribution
    cycle_breakdown = []
    total_completed = len(qa_completed)
    total_qa_time = sum(r['qa_business_days'] for r in qa_completed)
    
    for cycle_key in ['1 cycle', '2 cycles', '3+ cycles']:
        times = cycle_distribution.get(cycle_key, [])
        if times:
            cycle_breakdown.append({
                'cycles': cycle_key,
                'tickets': len(times),
                'avg_days': round(sum(times) / len(times), 1),
                'total_days': round(sum(times), 1),
                'pct_tickets': round(len(times) / total_completed * 100, 1) if total_completed else 0,
                'pct_time': round(sum(times) / total_qa_time * 100, 1) if total_qa_time else 0,
            })
    
    # Monthly trend
    monthly_trend = []
    for month, times in sorted(monthly_qa_times.items())[-12:]:
        monthly_trend.append({
            'month': month,
            'avg_days': round(sum(times) / len(times), 1) if times else 0,
            'tickets': len(times),
        })
    
    # Status distribution
    status_distribution = [
        {'status': status or 'Unknown', 'count': count}
        for status, count in status_counts.most_common(10)
    ]
    
    # In QA now (all QA-related statuses)
    in_qa_now = sum(status_counts.get(s, 0) for s in QA_STATUSES)
    
    # Total hold hours
    total_hold_hours = sum(r['qa_hold_hours'] for r in results if r['qa_hold_hours'])
    
    # QC Tester breakdown
    qc_tester_breakdown = []
    for tester, times in sorted(qc_tester_qa_times.items(), key=lambda x: -len(x[1])):
        if times:
            qc_tester_breakdown.append({
                'name': tester,
                'tickets': len(times),
                'avg_days': round(sum(times) / len(times), 1),
                'total_days': round(sum(times), 1),
            })
    
    # Developer breakdown
    developer_breakdown = []
    for dev, times in sorted(developer_qa_times.items(), key=lambda x: -len(x[1])):
        if times:
            developer_breakdown.append({
                'name': dev,
                'tickets': len(times),
                'avg_days': round(sum(times) / len(times), 1),
                'total_days': round(sum(times), 1),
            })
    
    return {
        'summary': {
            'total_tickets': len(results),
            'qa_completed': len(qa_completed),
            'avg_qa_days': round(avg_qa_days, 2),
            'median_qa_days': round(median_qa_days, 2),
            'first_pass_rate': round(first_pass_rate, 1),
            'avg_cycles': round(avg_cycles, 2),
            'total_fails': total_fails,
            'in_qa_now': in_qa_now,
            'total_hold_hours': round(total_hold_hours, 1),
        },
        'platform_breakdown': platform_breakdown[:10],
        'priority_breakdown': priority_breakdown[:10],
        'cycle_breakdown': cycle_breakdown,
        'monthly_trend': monthly_trend,
        'qc_tester_breakdown': qc_tester_breakdown[:15],
        'developer_breakdown': developer_breakdown[:15],
        'status_distribution': status_distribution,
        'reduction_targets': {
            'baseline': round(avg_qa_days, 2),
            'target_10': round(avg_qa_days * 0.9, 2),
            'target_20': round(avg_qa_days * 0.8, 2),
            'target_30': round(avg_qa_days * 0.7, 2),
            'target_50': round(avg_qa_days * 0.5, 2),
        },
        'generated_at': datetime.now().isoformat(),
        'date_filter': {
            'start_date': start_date,
            'end_date': end_date,
            'filtered': bool(filter_start or filter_end),
        },
    }


@app.get("/api/qa-dashboard/tickets")
def get_qa_dashboard_tickets(
    limit: int = Query(100, ge=1, le=1000),
    offset: int = Query(0, ge=0),
    platform: str = Query(None),
    priority: str = Query(None),
    min_cycles: int = Query(None, ge=1),
    start_date: str = Query(None, description="Filter start date (YYYY-MM-DD)"),
    end_date: str = Query(None, description="Filter end date (YYYY-MM-DD)"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_reports_access)
):
    """
    Get ticket-level QA cycle time data for the dashboard table.
    Reads from PM Activity Export JSON file.
    Supports pagination, filtering, and date range filtering.
    """
    from collections import defaultdict
    from pathlib import Path
    import json
    
    # Parse date filters
    filter_start = None
    filter_end = None
    if start_date:
        try:
            filter_start = datetime.strptime(start_date, '%Y-%m-%d')
        except ValueError:
            pass
    if end_date:
        try:
            filter_end = datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
        except ValueError:
            pass
    
    # Status definitions - tracking all existing PM statuses
    # QA-related statuses (tickets actively in QA)
    QA_STATUSES = {'QC Testing', 'QC Testing in Progress', 'QC Review Fail', 'QC Testing On-hold', 'QC Testing Hold', 'Tested - Awaiting Fixes'}
    QA_START_STATUSES = {'QC Testing', 'QC Testing in Progress'}  # When ticket enters QA
    QA_END_STATUSES = {'BIS Testing', 'Closed', 'Approved for Live', 'Moved to Live'}  # When ticket exits QA successfully
    QA_HOLD_STATUSES = {'QC Testing On-hold', 'QC Testing Hold', 'Hold/Pending'}  # Hold statuses (time not counted)
    QA_FAIL_STATUSES = {'QC Review Fail', 'Tested - Awaiting Fixes'}  # Failed QA review
    
    # Find the latest PM Activity Export file
    reports_dir = Path("reports")
    export_files = list(reports_dir.glob("PM_Activity_Export_*.csv"))
    if not export_files:
        raise HTTPException(status_code=404, detail="No PM Activity Export file found.")
    
    latest_export = max(export_files, key=lambda f: f.stat().st_mtime)
    
    # Load JSON data
    try:
        with open(latest_export, 'r', encoding='utf-8') as f:
            raw_data = json.load(f)
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=500, detail=f"Failed to parse PM Activity Export file: {e}")
    
    # Group by ticket (with optional date filtering)
    ticket_history = defaultdict(list)
    for record in raw_data:
        ticket_id = record.get('ticketId')
        if ticket_id:
            try:
                ticket_id = int(ticket_id)
            except (ValueError, TypeError):
                pass
            
            change_date = datetime.strptime(record['statusChangeDate'], '%Y-%m-%d %H:%M:%S')
            
            # Apply date filter
            if filter_start and change_date < filter_start:
                continue
            if filter_end and change_date > filter_end:
                continue
                
            ticket_history[ticket_id].append({
                'date': change_date,
                'old_status': record.get('oldStatus'),
                'new_status': record.get('newStatus'),
            })
    
    # Sort each ticket's history
    for tid in ticket_history:
        ticket_history[tid].sort(key=lambda x: x['date'])
    
    # Get ticket metadata from database for enrichment
    tickets = db.query(TicketTracking).all()
    ticket_lookup = {t.ticket_id: t for t in tickets}
    
    # Calculate metrics
    results = []
    
    for ticket_id, history in ticket_history.items():
        if not history:
            continue
            
        current_status = history[-1]['new_status'] if history else None
        
        qa_start = None
        qa_end = None
        qa_cycles = 0
        fail_count = 0
        total_hold_hours = 0.0
        hold_start = None
        
        for h in history:
            if h['new_status'] in QA_START_STATUSES:
                if qa_start is None:
                    qa_start = h['date']
                qa_cycles += 1
            
            if h['new_status'] in QA_END_STATUSES and qa_start is not None:
                qa_end = h['date']
            
            if h['new_status'] in QA_HOLD_STATUSES:
                hold_start = h['date']
            elif hold_start is not None:
                hold_duration = (h['date'] - hold_start).total_seconds() / 3600
                total_hold_hours += hold_duration
                hold_start = None
            
            if h['new_status'] in QA_FAIL_STATUSES:
                fail_count += 1
        
        qa_business_days = None
        if qa_start and qa_end:
            gross_hours = (qa_end - qa_start).total_seconds() / 3600
            net_hours = max(gross_hours - total_hold_hours, 0)
            qa_business_days = net_hours / 8
        
        ticket = ticket_lookup.get(ticket_id)
        
        # Apply filters
        if platform and ticket and ticket.subdepartment != platform:
            continue
        if priority and ticket and ticket.priority != priority:
            continue
        if min_cycles and qa_cycles < min_cycles:
            continue
        
        results.append({
            'ticket_id': ticket_id,
            'title': ticket.title if ticket else '',
            'current_status': current_status,
            'priority': ticket.priority if ticket else '',
            'subdepartment': ticket.subdepartment if ticket else '',
            'qc_tester': ticket.qc_tester if ticket else '',
            'backend_dev': ticket.backend_developer if ticket else '',
            'frontend_dev': ticket.frontend_developer if ticket else '',
            'current_assignee': ticket.current_assignee if ticket else '',
            'qa_start': qa_start.isoformat() if qa_start else None,
            'qa_end': qa_end.isoformat() if qa_end else None,
            'qa_business_days': round(qa_business_days, 2) if qa_business_days else None,
            'qa_cycles': qa_cycles,
            'qa_fail_count': fail_count,
            'qa_hold_hours': round(total_hold_hours, 2),
        })
    
    # Sort by ticket_id descending
    results.sort(key=lambda x: x['ticket_id'], reverse=True)
    
    total = len(results)
    paginated = results[offset:offset + limit]
    
    return {
        'total': total,
        'offset': offset,
        'limit': limit,
        'tickets': paginated,
    }


# ===== CALENDAR AND TASK PLANNING PYDANTIC MODELS =====

class PlannedTaskCreate(BaseModel):
    employee_id: Optional[str] = None
    employee_name: str
    ticket_id: str
    task_title: str
    task_description: Optional[str] = None
    project_name: Optional[str] = None
    planned_date: date
    planned_hours: float
    priority: Optional[str] = "medium"
    team: str
    assigned_by: str

class PlannedTaskUpdate(BaseModel):
    task_title: Optional[str] = None
    task_description: Optional[str] = None
    planned_hours: Optional[float] = None
    priority: Optional[str] = None
    status: Optional[str] = None
    actual_hours: Optional[float] = None

class WeeklyPlanCreate(BaseModel):
    employee_id: Optional[str] = None
    employee_name: str
    week_start: date
    assigned_tickets: List[dict]  # [{"ticket_id": "123", "priority": "high", "estimated_hours": 20}]
    notes: Optional[str] = None
    team: str
    planned_by: str

class WeeklyPlanUpdate(BaseModel):
    assigned_tickets: Optional[List[dict]] = None
    notes: Optional[str] = None
    status: Optional[str] = None


# ===== DEVELOPMENT TASK PLANNING PYDANTIC MODELS =====

class DevPlanningTaskCreate(BaseModel):
    employee_id: Optional[str] = None
    employee_name: str
    task_category: str  # Ticket | Team Meetings | Customer Support | Training | KT | Miscellaneous
    ticket_id: Optional[int] = None  # Required when task_category is Ticket
    activity_description: str
    start_date: date
    end_date: Optional[date] = None
    total_hours: Optional[float] = None  # Duration in hours (1-40). When provided, used directly.
    max_hours_per_day: Optional[float] = None  # 0.5–8, max hours for this task per day (dropdown)
    generic_category: Optional[str] = None  # Required when task_category is not Ticket
    justification: Optional[str] = None  # Required for generic tasks


class DevPlanningTaskUpdate(BaseModel):
    activity_description: Optional[str] = None
    start_date: Optional[date] = None
    end_date: Optional[date] = None
    allocation_pct: Optional[int] = None


class DevPlanningWeekStateUpdate(BaseModel):
    state: str  # submitted | approved | locked | draft (unlock)


# ===== GOOGLE SHEETS SYNC ENDPOINTS =====

@app.get("/sync/google-sheets/status")
def get_google_sheets_status():
    """Get the current status of Google Sheets sync configuration and scheduler."""
    config_status = get_sheets_sync_status()
    scheduler = get_scheduler()
    scheduler_status = scheduler.get_status()
    
    return {
        **config_status,
        "scheduler": scheduler_status
    }

# ===== MONTHLY TEAM REPORT ENDPOINTS =====

@app.get("/reports/monthly-team/status")
def monthly_team_report_status(current_user: dict = Depends(get_current_user)):
    """Get the monthly team report scheduler status (next/last run)."""
    return get_monthly_report_status()


@app.post("/reports/monthly-team/generate")
def trigger_monthly_team_report(
    month: Optional[str] = Query(
        None, description="Month in YYYY-MM. Defaults to the current month."
    ),
    teams: Optional[str] = Query(
        None,
        description="Comma-separated team keys to generate (QA,DEV). Default: both.",
    ),
    send_email: Optional[bool] = Query(
        None,
        description=(
            "Whether to email the PDFs after generation. "
            "If omitted, follows MONTHLY_REPORT_EMAIL_ENABLED env var. "
            "Pass true/false to override."
        ),
    ),
    current_user: dict = Depends(get_current_user),
):
    """Manually generate the per-team monthly PDF reports (one PDF per team)."""
    teams_list = [t.strip().upper() for t in teams.split(",")] if teams else None
    result = get_monthly_report_scheduler().trigger_manual(
        month, send_email=send_email, teams=teams_list
    )
    if not result.get("success"):
        raise HTTPException(
            status_code=500, detail=result.get("error", "Report generation failed")
        )
    return result


@app.get("/reports/monthly-team/download")
def download_monthly_team_report(
    team: str = Query("QA", description="Team key: QA or DEV"),
    month: Optional[str] = Query(
        None, description="Month in YYYY-MM. Defaults to the latest available report for the team."
    ),
    current_user: dict = Depends(get_current_user),
):
    """Download a previously generated monthly team report PDF."""
    team_key = team.upper()
    if team_key not in {"QA", "DEV"}:
        raise HTTPException(status_code=400, detail="team must be QA or DEV")
    reports_dir = os.path.join(os.path.dirname(__file__), "reports")
    prefix = f"Monthly_Team_Report_{team_key}_"
    if month:
        try:
            year, mon = map(int, month.split("-"))
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid month format. Use YYYY-MM")
        filename = f"{prefix}{year:04d}-{mon:02d}.pdf"
    else:
        try:
            candidates = [
                f for f in os.listdir(reports_dir)
                if f.startswith(prefix) and f.endswith(".pdf")
            ]
        except FileNotFoundError:
            candidates = []
        if not candidates:
            raise HTTPException(
                status_code=404,
                detail=f"No monthly {team_key} report has been generated yet.",
            )
        filename = sorted(candidates)[-1]
    path = os.path.join(reports_dir, filename)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail=f"Report not found: {filename}")
    return FileResponse(
        path,
        media_type="application/pdf",
        filename=filename,
    )


@app.post("/sync/google-sheets")
def trigger_google_sheets_sync(team: Optional[str] = Query(None, description="Team to sync: QA, DEV, or leave empty for all")):
    """Trigger a manual sync from Google Sheets."""
    try:
        sync = GoogleSheetsSync()
        if team:
            result = sync.sync_team(team.upper())
        else:
            result = sync.sync_all()
        return {"success": True, "result": result}
    except FileNotFoundError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Sync failed: {str(e)}")

@app.post("/sync/google-sheets/start")
def start_auto_sync_endpoint(
    interval_minutes: Optional[int] = Query(None, description="Sync interval in minutes (ignored if realtime=true)"),
    teams: Optional[str] = Query(None, description="Comma-separated teams: QA,DEV"),
    realtime: bool = Query(True, description="Enable real-time sync (2-minute intervals)")
):
    """Start automatic syncing of Google Sheets."""
    try:
        scheduler = get_scheduler()
        teams_list = [t.strip().upper() for t in teams.split(',')] if teams else None
        scheduler.start(sync_interval_minutes=interval_minutes, teams=teams_list, realtime=realtime)
        
        mode = "real-time (2-minute intervals)" if realtime else f"{interval_minutes or 5} minute intervals"
        return {
            "success": True,
            "message": f"Auto-sync started in {mode}",
            "status": scheduler.get_status()
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to start auto-sync: {str(e)}")

@app.post("/sync/google-sheets/stop")
def stop_auto_sync_endpoint():
    """Stop automatic syncing of Google Sheets."""
    try:
        scheduler = get_scheduler()
        scheduler.stop()
        return {
            "success": True,
            "message": "Auto-sync stopped"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to stop auto-sync: {str(e)}")

@app.post("/sync/google-sheets/trigger")
def trigger_scheduled_sync(teams: Optional[str] = Query(None, description="Comma-separated teams: QA,DEV")):
    """Manually trigger the scheduled sync job."""
    try:
        scheduler = get_scheduler()
        teams_list = [t.strip().upper() for t in teams.split(',')] if teams else None
        result = scheduler.trigger_manual_sync(teams=teams_list)
        return {
            "success": True,
            "message": "Manual sync triggered",
            "result": result
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to trigger sync: {str(e)}")


# ===== GOOGLE SHEETS EXPORT ENDPOINTS =====

@app.get("/sync/sheets-export/status")
def get_sheets_export_status():
    """Get Google Sheets export configuration status."""
    import os
    credentials_configured = bool(os.getenv("SHEETS_EXPORT_CREDENTIALS_FILE"))
    spreadsheet_configured = bool(os.getenv("SHEETS_EXPORT_SPREADSHEET_ID"))
    auto_sync_enabled = os.getenv("SHEETS_EXPORT_AUTO_SYNC", "false").lower() == "true"
    
    return {
        "configured": credentials_configured and spreadsheet_configured,
        "credentials_file": credentials_configured,
        "spreadsheet_id": spreadsheet_configured,
        "auto_sync_enabled": auto_sync_enabled,
        "google_api_available": GOOGLE_EXPORT_AVAILABLE,
        "spreadsheet_url": f"https://docs.google.com/spreadsheets/d/{os.getenv('SHEETS_EXPORT_SPREADSHEET_ID', '')}" if spreadsheet_configured else None
    }


@app.post("/sync/sheets-export/trigger")
def trigger_sheets_export():
    """
    Manually trigger an export of all data to Google Sheets.
    
    This exports:
    - PM_Tickets: All tickets from PM Tool
    - PM_Status_History: Ticket status change history
    - TestRail_Runs: Test runs from TestRail
    - TestRail_Cases: Test cases with automation status
    - TestRail_Bugs: Bug tracking data
    - TestRail_Results: Test execution results
    """
    try:
        result = trigger_manual_export()
        if result.get("success"):
            return {
                "success": True,
                "message": "Export completed successfully",
                "details": result
            }
        else:
            raise HTTPException(status_code=500, detail=result.get("error", "Export failed"))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


# ===== CALENDAR API ENDPOINTS =====

def is_weekend(check_date: date) -> bool:
    """Check if a date is a weekend (Saturday or Sunday)."""
    return check_date.weekday() >= 5  # 5 = Saturday, 6 = Sunday


def is_holiday(check_date: date, db: Session, include_optional: bool = False) -> Optional[Holiday]:
    """
    Check if a date is a holiday.
    Returns the Holiday object if found, None otherwise.
    If include_optional is False, only checks for regular holidays.
    """
    query = db.query(Holiday).filter(
        Holiday.holiday_date == check_date,
        Holiday.is_active == True
    )
    
    if not include_optional:
        query = query.filter(Holiday.category == 'Holiday')
    
    return query.first()


def is_working_day(check_date: date, db: Session, include_optional_holidays: bool = False) -> bool:
    """
    Check if a date is a working day (not weekend and not a holiday).
    If include_optional_holidays is True, optional holidays are also considered non-working days.
    """
    if is_weekend(check_date):
        return False
    
    holiday = is_holiday(check_date, db, include_optional=include_optional_holidays)
    if holiday:
        return False
    
    return True


def get_working_days_in_range(start_date: date, end_date: date, db: Session, include_optional_holidays: bool = False) -> int:
    """Count working days (excluding weekends and holidays) in a date range."""
    working_days = 0
    current_date = start_date
    
    while current_date <= end_date:
        if is_working_day(current_date, db, include_optional_holidays):
            working_days += 1
        current_date += timedelta(days=1)
    
    return working_days


@app.get("/calendar/holidays")
def get_holidays(
    year: Optional[int] = Query(None, description="Year. Defaults to current year."),
    category: Optional[str] = Query(None, description="Filter by category: 'Holiday' or 'Optional Holiday'")
):
    """Get list of holidays for a given year."""
    db = SessionLocal()
    try:
        if not year:
            year = date.today().year
        
        query = db.query(Holiday).filter(
            Holiday.year == year,
            Holiday.is_active == True
        )
        
        if category:
            query = query.filter(Holiday.category == category)
        
        holidays = query.order_by(Holiday.holiday_date).all()
        
        return {
            "year": year,
            "holidays": [
                {
                    "id": h.id,
                    "name": h.holiday_name,
                    "date": h.holiday_date.isoformat(),
                    "day_name": h.day_name,
                    "category": h.category
                }
                for h in holidays
            ]
        }
    finally:
        db.close()


@app.get("/calendar/weekly")
def get_weekly_calendar(
    team: str = Query("ALL", description="Team: QA, DEV, or ALL"),
    date_str: str = Query(None, description="Any date in the week (YYYY-MM-DD). Defaults to current week."),
    category: str = Query("ALL", description="Category: BILLED, UN-BILLED, or ALL"),
):
    """
    Get weekly calendar view showing daily time entries per employee.
    Filtered by role visibility.
    """
    db = SessionLocal()
    try:
        # Parse date and calculate week boundaries (Monday to Sunday)
        if date_str:
            try:
                target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
        else:
            target_date = date.today()
        
        # Calculate week start (Monday) and end (Sunday)
        week_start = target_date - timedelta(days=target_date.weekday())
        week_end = week_start + timedelta(days=6)
        
        # Get holidays for this week
        week_holidays = {}
        holidays_query = db.query(Holiday).filter(
            Holiday.holiday_date >= week_start,
            Holiday.holiday_date <= week_end,
            Holiday.is_active == True
        ).all()
        
        for holiday in holidays_query:
            week_holidays[holiday.holiday_date.isoformat()] = {
                "name": holiday.holiday_name,
                "category": holiday.category,
                "day_name": holiday.day_name
            }
        
        # Get list of employees (filtered by team and category)
        emp_query = db.query(Employee).filter(Employee.is_active == True)
        if team.upper() != "ALL":
            # Map "DEV" to "DEVELOPMENT" for Employee table (Employee.team uses "DEVELOPMENT", not "DEV")
            employee_team_filter = team.upper()
            if employee_team_filter == "DEV":
                employee_team_filter = "DEVELOPMENT"
            emp_query = emp_query.filter(Employee.team == employee_team_filter)
        if category.upper() != "ALL":
            # Use case-insensitive exact match for category
            # Match both "BILLED" and "UN-BILLED" (with or without hyphen)
            category_upper = category.upper()
            if category_upper == "UN-BILLED" or category_upper == "UNBILLED":
                emp_query = emp_query.filter(
                    or_(
                        func.upper(Employee.category) == "UN-BILLED",
                        func.upper(Employee.category) == "UNBILLED"
                    )
                )
            else:
                emp_query = emp_query.filter(func.upper(Employee.category) == category_upper)
        employees = emp_query.all()
        # No auth — show all employees

        # Get employee names for filtering timesheet data
        employee_names = [emp.name for emp in employees]
        # Map employee_id -> employee for calendar (one row per person, avoids duplicate names)
        employee_ids = {emp.employee_id for emp in employees}
        name_to_employee_id = {}
        for emp in employees:
            if not emp.name:
                continue
            raw_name = emp.name.strip()
            name_to_employee_id[raw_name] = emp.employee_id
            name_to_employee_id[_normalize_person_name(raw_name)] = emp.employee_id
            name_to_employee_id[_compact_person_name(raw_name)] = emp.employee_id
        # Load name mappings so alternate spellings (e.g. "Gautham Krishna KP") map to same employee_id
        names_for_query = set(employee_names)
        try:
            name_mappings = db.query(EmployeeNameMapping).filter(
                EmployeeNameMapping.is_active == True
            ).all()
            for m in name_mappings:
                if m.employee_id and m.employee_id in employee_ids and m.alternate_name:
                    alt = m.alternate_name.strip()
                    name_to_employee_id[alt] = m.employee_id
                    name_to_employee_id[_normalize_person_name(alt)] = m.employee_id
                    name_to_employee_id[_compact_person_name(alt)] = m.employee_id
                    names_for_query.add(alt)
        except Exception:
            pass
        
        # Map team for timesheet filter: EnhancedTimesheet uses "DEV", Employee uses "DEVELOPMENT"
        timesheet_team_filter = team.upper()
        if timesheet_team_filter == "DEVELOPMENT":
            timesheet_team_filter = "DEV"
        
        # Query timesheet data (include alternate names so we fetch entries for merged employees)
        query = db.query(EnhancedTimesheet).filter(
            EnhancedTimesheet.date >= week_start,
            EnhancedTimesheet.date <= week_end
        )
        
        if team.upper() != "ALL":
            query = query.filter(EnhancedTimesheet.team == timesheet_team_filter)
        if category.upper() != "ALL" and names_for_query:
            query = query.filter(EnhancedTimesheet.employee_name.in_(names_for_query))
        
        entries = query.order_by(
            EnhancedTimesheet.employee_name,
            EnhancedTimesheet.date
        ).all()
        
        # Also get leave entries (use same team mapping: DEVELOPMENT -> DEV)
        leave_query = db.query(LeaveEntry).filter(
            LeaveEntry.date >= week_start,
            LeaveEntry.date <= week_end
        )
        if team.upper() != "ALL":
            leave_query = leave_query.filter(LeaveEntry.team == timesheet_team_filter)
        if category.upper() != "ALL" and names_for_query:
            leave_query = leave_query.filter(LeaveEntry.employee_name.in_(names_for_query))
        leaves = leave_query.all()
        
        # Build employee calendar data keyed by employee_id (one row per person, no duplicates)
        employee_data = {}
        
        # Initialize all employees
        for emp in employees:
            employee_data[emp.employee_id] = {
                "employee_id": emp.employee_id,
                "employee_name": emp.name,
                "team": emp.team,
                "days": {}
            }
            # Initialize all days
            for i in range(7):
                day = week_start + timedelta(days=i)
                day_key = day.isoformat()
                is_weekend_day = is_weekend(day)
                holiday_info = week_holidays.get(day_key)
                
                employee_data[emp.employee_id]["days"][day_key] = {
                    "date": day_key,
                    "entries": [],
                    "total_hours": 0,
                    "productive_hours": 0,
                    "hours_logged": 0,
                    "ticket_hours": 0,
                    "non_ticket_hours": 0,
                    "leave_type": None,
                    "is_weekend": is_weekend_day,
                    "is_holiday": holiday_info is not None,
                    "holiday_name": holiday_info["name"] if holiday_info else None,
                    "holiday_category": holiday_info["category"] if holiday_info else None,
                    "is_working_day": not is_weekend_day and holiday_info is None
                }
        
        # Resolve timesheet entry to employee_id (by id or by name/mapping)
        def _resolve_employee_id(entry_employee_id, entry_employee_name):
            if entry_employee_id and entry_employee_id in employee_data:
                return entry_employee_id
            name = (entry_employee_name or "").strip()
            return (
                name_to_employee_id.get(name)
                or name_to_employee_id.get(_normalize_person_name(name))
                or name_to_employee_id.get(_compact_person_name(name))
            )
        
        # Classify ticket vs non-ticket: numeric ticket_id = ticket task
        def _is_ticket_task(tid):
            if not tid:
                return False
            tid_str = str(tid).strip()
            return tid_str.isdigit() and int(tid_str) > 0

        # Add timesheet entries only to existing employee rows (no new rows from name variants)
        for entry in entries:
            eid = _resolve_employee_id(entry.employee_id, entry.employee_name)
            if eid is None:
                continue
            day_key = entry.date.isoformat()
            if day_key in employee_data[eid]["days"]:
                # Get hours - use productive_hours if available, otherwise hours_logged
                productive = entry.productive_hours or 0
                hours_logged = entry.hours_logged or 0
                display_hours = productive if productive > 0 else hours_logged
                is_ticket = _is_ticket_task(entry.ticket_id)

                employee_data[eid]["days"][day_key]["entries"].append({
                    "ticket_id": entry.ticket_id,
                    "hours": display_hours,
                    "productive_hours": productive,
                    "hours_logged": hours_logged,
                    "task_description": entry.task_description,
                    "project_name": entry.project_name,
                    "is_ticket_task": is_ticket
                })
                employee_data[eid]["days"][day_key]["total_hours"] += display_hours
                employee_data[eid]["days"][day_key]["productive_hours"] += productive
                employee_data[eid]["days"][day_key]["hours_logged"] = employee_data[eid]["days"][day_key].get("hours_logged", 0) + hours_logged
                if is_ticket:
                    employee_data[eid]["days"][day_key]["ticket_hours"] = employee_data[eid]["days"][day_key].get("ticket_hours", 0) + display_hours
                else:
                    employee_data[eid]["days"][day_key]["non_ticket_hours"] = employee_data[eid]["days"][day_key].get("non_ticket_hours", 0) + display_hours
                if entry.leave_type:
                    employee_data[eid]["days"][day_key]["leave_type"] = entry.leave_type
        
        # Add leave entries (resolve to employee_id so same person not duplicated)
        for leave in leaves:
            eid = _resolve_employee_id(leave.employee_id, leave.employee_name)
            if eid is not None:
                day_key = leave.date.isoformat()
                if day_key in employee_data[eid]["days"]:
                    employee_data[eid]["days"][day_key]["leave_type"] = leave.leave_type
        
        # Calculate totals per employee
        for eid, data in employee_data.items():
            total = sum(d["total_hours"] for d in data["days"].values())
            productive = sum(d["productive_hours"] for d in data["days"].values())
            ticket_hrs = sum(d.get("ticket_hours", 0) for d in data["days"].values())
            non_ticket_hrs = sum(d.get("non_ticket_hours", 0) for d in data["days"].values())
            data["weekly_total_hours"] = total
            data["weekly_productive_hours"] = productive
            data["weekly_ticket_hours"] = ticket_hrs
            data["weekly_non_ticket_hours"] = non_ticket_hrs
            # Expected hours: 8 per working day
            working_day_count = sum(1 for d in data["days"].values() if d["is_working_day"])
            data["expected_hours"] = working_day_count * 8

        # Calculate working days in the week (excluding weekends and holidays)
        working_days = get_working_days_in_range(week_start, week_end, db, include_optional_holidays=False)

        return {
            "week_start": week_start.isoformat(),
            "week_end": week_end.isoformat(),
            "team": team,
            "working_days": working_days,
            "holidays": list(week_holidays.values()),
            "employees": sorted(employee_data.values(), key=lambda e: (e.get("employee_name") or "").lower())
        }
    finally:
        db.close()


@app.get("/calendar/monthly")
def get_monthly_calendar(
    team: str = Query("ALL", description="Team: QA, DEV, or ALL"),
    month: str = Query(None, description="Month (YYYY-MM). Defaults to current month."),
    category: str = Query("ALL", description="Category: BILLED, UN-BILLED, or ALL"),
):
    """
    Get monthly calendar view showing summary per employee.
    Filtered by role visibility.
    """
    db = SessionLocal()
    try:
        # Parse month
        if month:
            try:
                year, mon = map(int, month.split("-"))
                month_start = date(year, mon, 1)
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid month format. Use YYYY-MM")
        else:
            today = date.today()
            month_start = date(today.year, today.month, 1)
        
        # Calculate month end
        if month_start.month == 12:
            month_end = date(month_start.year + 1, 1, 1) - timedelta(days=1)
        else:
            month_end = date(month_start.year, month_start.month + 1, 1) - timedelta(days=1)
        
        # Get holidays for this month
        month_holidays = {}
        holidays_query = db.query(Holiday).filter(
            Holiday.holiday_date >= month_start,
            Holiday.holiday_date <= month_end,
            Holiday.is_active == True
        ).all()
        
        for holiday in holidays_query:
            month_holidays[holiday.holiday_date.isoformat()] = {
                "date": holiday.holiday_date.isoformat(),
                "name": holiday.holiday_name,
                "category": holiday.category,
                "day_name": holiday.day_name
            }
        
        # Get all active employees from the Employee master table (filtered by team and category)
        emp_query = db.query(Employee).filter(Employee.is_active == True)
        if team.upper() != "ALL":
            # Map "DEV" to "DEVELOPMENT" for Employee table (Employee.team uses "DEVELOPMENT", not "DEV")
            employee_team_filter = team.upper()
            if employee_team_filter == "DEV":
                employee_team_filter = "DEVELOPMENT"
            emp_query = emp_query.filter(Employee.team == employee_team_filter)
        if category.upper() != "ALL":
            # Use case-insensitive exact match for category
            # Match both "BILLED" and "UN-BILLED" (with or without hyphen)
            category_upper = category.upper()
            if category_upper == "UN-BILLED" or category_upper == "UNBILLED":
                emp_query = emp_query.filter(
                    or_(
                        func.upper(Employee.category) == "UN-BILLED",
                        func.upper(Employee.category) == "UNBILLED"
                    )
                )
            else:
                emp_query = emp_query.filter(func.upper(Employee.category) == category_upper)
        all_employees = emp_query.all()
        # No auth — show all employees

        # Get employee names for filtering timesheet data
        employee_names = [emp.name for emp in all_employees]
        employee_ids = {emp.employee_id for emp in all_employees}
        name_to_employee_id = {}
        for emp in all_employees:
            if not emp.name:
                continue
            raw_name = emp.name.strip()
            name_to_employee_id[raw_name] = emp.employee_id
            name_to_employee_id[_normalize_person_name(raw_name)] = emp.employee_id
            name_to_employee_id[_compact_person_name(raw_name)] = emp.employee_id
        names_for_query = set(employee_names)
        try:
            name_mappings = db.query(EmployeeNameMapping).filter(
                EmployeeNameMapping.is_active == True
            ).all()
            for m in name_mappings:
                if m.employee_id and m.employee_id in employee_ids and m.alternate_name:
                    alt = m.alternate_name.strip()
                    name_to_employee_id[alt] = m.employee_id
                    name_to_employee_id[_normalize_person_name(alt)] = m.employee_id
                    name_to_employee_id[_compact_person_name(alt)] = m.employee_id
                    names_for_query.add(alt)
        except Exception:
            pass
        
        # Map team for timesheet filter: EnhancedTimesheet uses "DEV", Employee uses "DEVELOPMENT"
        timesheet_team_filter = team.upper()
        if timesheet_team_filter == "DEVELOPMENT":
            timesheet_team_filter = "DEV"
        
        # Query timesheet data (include alternate names for merged employees)
        query = db.query(EnhancedTimesheet).filter(
            EnhancedTimesheet.date >= month_start,
            EnhancedTimesheet.date <= month_end
        )
        if team.upper() != "ALL":
            query = query.filter(EnhancedTimesheet.team == timesheet_team_filter)
        if category.upper() != "ALL" and names_for_query:
            query = query.filter(EnhancedTimesheet.employee_name.in_(names_for_query))
        entries = query.all()
        
        # Query leaves (include alternate names for merged employees)
        leave_query = db.query(LeaveEntry).filter(
            LeaveEntry.date >= month_start,
            LeaveEntry.date <= month_end
        )
        if team.upper() != "ALL":
            leave_query = leave_query.filter(LeaveEntry.team == timesheet_team_filter)
        if category.upper() != "ALL" and names_for_query:
            leave_query = leave_query.filter(LeaveEntry.employee_name.in_(names_for_query))
        leaves = leave_query.all()
        
        def _resolve_eid(entry_employee_id, entry_employee_name):
            if entry_employee_id and entry_employee_id in employee_ids:
                return entry_employee_id
            raw_name = (entry_employee_name or "").strip()
            return (
                name_to_employee_id.get(raw_name)
                or name_to_employee_id.get(_normalize_person_name(raw_name))
                or name_to_employee_id.get(_compact_person_name(raw_name))
            )
        
        # Classify ticket vs non-ticket
        def _is_ticket_task(tid):
            if not tid:
                return False
            tid_str = str(tid).strip()
            return tid_str.isdigit() and int(tid_str) > 0

        # Build employee data keyed by employee_id (one row per person, no duplicates)
        employee_data = defaultdict(lambda: {
            "days": defaultdict(lambda: {
                "hours": 0,
                "productive_hours": 0,
                "hours_logged": 0,
                "ticket_hours": 0,
                "non_ticket_hours": 0,
                "leave_type": None,
                "entries": []
            }),
            "total_hours": 0,
            "total_productive_hours": 0,
            "total_ticket_hours": 0,
            "total_non_ticket_hours": 0,
            "total_leave_days": 0,
            "total_leave_hours": 0,
            "expected_hours": 0,
            "working_days": 0
        })
        
        # Initialize all active employees (even those with no entries)
        for emp in all_employees:
            employee_data[emp.employee_id]["employee_id"] = emp.employee_id
            employee_data[emp.employee_id]["employee_name"] = emp.name
            employee_data[emp.employee_id]["team"] = emp.team
        
        for entry in entries:
            eid = _resolve_eid(entry.employee_id, entry.employee_name)
            if eid is None:
                continue
            day = entry.date.isoformat()
            productive = entry.productive_hours if entry.productive_hours is not None else None
            time_spent = entry.hours_logged or 0
            display_hours = productive if productive is not None else time_spent
            is_ticket = _is_ticket_task(entry.ticket_id)

            employee_data[eid]["days"][day]["productive_hours"] += productive if productive is not None else 0
            employee_data[eid]["days"][day]["hours_logged"] += time_spent
            employee_data[eid]["days"][day]["hours"] += display_hours
            employee_data[eid]["days"][day]["entries"].append({
                "ticket_id": entry.ticket_id,
                "hours": display_hours,
                "task_description": entry.task_description,
                "is_ticket_task": is_ticket
            })
            if is_ticket:
                employee_data[eid]["days"][day]["ticket_hours"] += display_hours
                employee_data[eid]["total_ticket_hours"] += display_hours
            else:
                employee_data[eid]["days"][day]["non_ticket_hours"] += display_hours
                employee_data[eid]["total_non_ticket_hours"] += display_hours
            if entry.leave_type:
                employee_data[eid]["days"][day]["leave_type"] = entry.leave_type
            employee_data[eid]["total_hours"] += display_hours
            employee_data[eid]["total_productive_hours"] += productive if productive is not None else 0
        
        for leave in leaves:
            eid = _resolve_eid(leave.employee_id, leave.employee_name)
            if eid is not None:
                day = leave.date.isoformat()
                employee_data[eid]["days"][day]["leave_type"] = leave.leave_type
                employee_data[eid]["total_leave_days"] += 1
                employee_data[eid]["total_leave_hours"] += float(leave.hours or 0)
        
        # Calculate working days and average productive hours
        today = date.today()
        for name, data in employee_data.items():
            # Add holiday/weekend information to each day
            for day_key in list(data["days"].keys()):
                day_date = datetime.strptime(day_key, "%Y-%m-%d").date()
                is_weekend_day = is_weekend(day_date)
                holiday_info = month_holidays.get(day_key)
                
                data["days"][day_key]["is_weekend"] = is_weekend_day
                data["days"][day_key]["is_holiday"] = holiday_info is not None
                data["days"][day_key]["holiday_name"] = holiday_info["name"] if holiday_info else None
                data["days"][day_key]["holiday_category"] = holiday_info["category"] if holiday_info else None
                data["days"][day_key]["is_working_day"] = not is_weekend_day and holiday_info is None
            
            # Only count past working days (excluding weekends and holidays)
            past_working_days = [
                day_key for day_key, d in data["days"].items() 
                if datetime.strptime(day_key, "%Y-%m-%d").date() <= today 
                and d.get("is_working_day", True)
            ]
            data["working_days"] = len(past_working_days)
            
            # Calculate average productive hours (only for past working days, excluding leave days)
            past_productive_days = [
                d for day_key, d in data["days"].items() 
                if datetime.strptime(day_key, "%Y-%m-%d").date() <= today 
                and d.get("is_working_day", True)
                and (d.get("productive_hours", 0) > 0 or d.get("hours_logged", 0) > 0) 
                and not d.get("leave_type")  # Exclude leave days
            ]
            if past_productive_days:
                total_productive = sum(d.get("productive_hours") or d.get("hours_logged", 0) for d in past_productive_days)
                data["avg_productive_hours"] = round(total_productive / len(past_productive_days), 1)
            else:
                data["avg_productive_hours"] = 0
            
            # Convert defaultdict to regular dict for JSON serialization
            data["days"] = dict(data["days"])
            for day in data["days"]:
                data["days"][day] = dict(data["days"][day])
        
        # Calculate total working days in the month
        total_working_days = get_working_days_in_range(month_start, month_end, db, include_optional_holidays=False)
        for eid, data in employee_data.items():
            expected = (total_working_days * 8) - float(data.get("total_leave_hours", 0) or 0)
            employee_data[eid]["expected_hours"] = round(max(expected, 0), 1)

        monthly_totals = {
            "productive_hours": round(sum(float(v.get("total_productive_hours", 0) or 0) for v in employee_data.values()), 1),
            "leave_hours": round(sum(float(v.get("total_leave_hours", 0) or 0) for v in employee_data.values()), 1),
            "leave_days": round(sum(float(v.get("total_leave_days", 0) or 0) for v in employee_data.values()), 1),
            "expected_hours": round(sum(float(v.get("expected_hours", 0) or 0) for v in employee_data.values()), 1),
        }
        
        return {
            "month": month_start.strftime("%Y-%m"),
            "month_start": month_start.isoformat(),
            "month_end": month_end.isoformat(),
            "team": team,
            "working_days": total_working_days,
            "holidays": list(month_holidays.values()),
            "monthly_totals": monthly_totals,
            "employees": sorted([dict(v) for v in employee_data.values()], key=lambda e: (e.get("employee_name") or "").lower())
        }
    finally:
        db.close()


@app.get("/calendar/employee/{employee_id}")
def get_employee_calendar(
    employee_id: str,
    period: str = Query("week", description="Period: week or month"),
    date_str: str = Query(None, description="Reference date (YYYY-MM-DD)"),
    current_user: dict = Depends(get_current_user),
):
    """Get calendar data for a specific employee."""
    db = SessionLocal()
    try:
        employee = db.query(Employee).filter(Employee.employee_id == employee_id).first()
        if not employee:
            raise HTTPException(status_code=404, detail="Employee not found")
        if not can_access_employee(db, current_user, employee.employee_id):
            raise HTTPException(status_code=403, detail="Access denied")
        
        # Parse date
        if date_str:
            try:
                target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
        else:
            target_date = date.today()
        
        # Calculate period boundaries
        if period == "week":
            start_date = target_date - timedelta(days=target_date.weekday())
            end_date = start_date + timedelta(days=6)
        else:  # month
            start_date = date(target_date.year, target_date.month, 1)
            if start_date.month == 12:
                end_date = date(start_date.year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(start_date.year, start_date.month + 1, 1) - timedelta(days=1)
        
        # Get holidays for this period
        period_holidays = {}
        holidays_query = db.query(Holiday).filter(
            Holiday.holiday_date >= start_date,
            Holiday.holiday_date <= end_date,
            Holiday.is_active == True
        ).all()
        
        for holiday in holidays_query:
            period_holidays[holiday.holiday_date.isoformat()] = {
                "name": holiday.holiday_name,
                "category": holiday.category,
                "day_name": holiday.day_name
            }
        
        # Query timesheet entries
        entries = db.query(EnhancedTimesheet).filter(
            EnhancedTimesheet.employee_name == employee.name,
            EnhancedTimesheet.date >= start_date,
            EnhancedTimesheet.date <= end_date
        ).order_by(EnhancedTimesheet.date).all()
        
        # Query leaves
        leaves = db.query(LeaveEntry).filter(
            LeaveEntry.employee_name == employee.name,
            LeaveEntry.date >= start_date,
            LeaveEntry.date <= end_date
        ).all()
        leave_map = {l.date.isoformat(): l.leave_type for l in leaves}
        
        # Query planned tasks
        planned = db.query(PlannedTask).filter(
            PlannedTask.employee_name == employee.name,
            PlannedTask.planned_date >= start_date,
            PlannedTask.planned_date <= end_date
        ).order_by(PlannedTask.planned_date).all()
        
        # Build day-by-day data
        days = {}
        current = start_date
        while current <= end_date:
            day_key = current.isoformat()
            is_weekend_day = is_weekend(current)
            holiday_info = period_holidays.get(day_key)
            
            days[day_key] = {
                "date": day_key,
                "actual_entries": [],
                "planned_tasks": [],
                "total_actual_hours": 0,
                "total_productive_hours": 0,
                "hours_logged": 0,
                "total_planned_hours": 0,
                "leave_type": leave_map.get(day_key),
                "is_weekend": is_weekend_day,
                "is_holiday": holiday_info is not None,
                "holiday_name": holiday_info["name"] if holiday_info else None,
                "holiday_category": holiday_info["category"] if holiday_info else None,
                "is_working_day": not is_weekend_day and holiday_info is None
            }
            current += timedelta(days=1)
        
        # Add actual entries - use productive_hours if available, otherwise hours_logged
        for entry in entries:
            day_key = entry.date.isoformat()
            if day_key in days:
                productive = entry.productive_hours or 0
                hours_logged = entry.hours_logged or 0
                display_hours = productive if productive > 0 else hours_logged
                
                days[day_key]["actual_entries"].append({
                    "ticket_id": entry.ticket_id,
                    "hours": display_hours,
                    "productive_hours": productive,
                    "hours_logged": hours_logged,
                    "task_description": entry.task_description,
                    "project_name": entry.project_name
                })
                days[day_key]["total_actual_hours"] += display_hours
                days[day_key]["total_productive_hours"] += productive
                days[day_key]["hours_logged"] += hours_logged
        
        # Add planned tasks
        for task in planned:
            day_key = task.planned_date.isoformat()
            if day_key in days:
                days[day_key]["planned_tasks"].append({
                    "id": task.id,
                    "ticket_id": task.ticket_id,
                    "task_title": task.task_title,
                    "planned_hours": task.planned_hours,
                    "priority": task.priority,
                    "status": task.status
                })
                days[day_key]["total_planned_hours"] += task.planned_hours or 0
        
        # Calculate working days
        working_days = get_working_days_in_range(start_date, end_date, db, include_optional_holidays=False)
        
        # Calculate summary
        total_actual = sum(d["total_actual_hours"] for d in days.values())
        total_productive = sum(d["total_productive_hours"] for d in days.values())
        total_hours_logged = sum(d["hours_logged"] for d in days.values())
        
        return {
            "employee_id": employee.employee_id,
            "employee_name": employee.name,
            "team": employee.team,
            "period": period,
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "working_days": working_days,
            "holidays": list(period_holidays.values()),
            "days": days,
            "summary": {
                "total_actual_hours": total_actual,
                "total_productive_hours": total_productive,
                "total_hours_logged": total_hours_logged,
                "total_planned_hours": sum(d["total_planned_hours"] for d in days.values()),
                "leave_days": len([d for d in days.values() if d["leave_type"]]),
                "working_days": working_days
            }
        }
    finally:
        db.close()


@app.get("/calendar/ticket/{ticket_id}/timesheet")
def get_ticket_timesheet_entries(ticket_id: str):
    """
    Get all timesheet entries for a specific ticket.
    Returns entries from all employees who worked on this ticket.
    """
    db = SessionLocal()
    try:
        # Query timesheet entries for this ticket
        entries = db.query(EnhancedTimesheet).filter(
            EnhancedTimesheet.ticket_id == ticket_id
        ).order_by(
            EnhancedTimesheet.date.desc(),
            EnhancedTimesheet.employee_name
        ).all()
        
        # Calculate summary
        total_hours = sum(e.hours_logged or 0 for e in entries)
        unique_employees = set(e.employee_name for e in entries)
        unique_dates = set(e.date for e in entries)
        
        return {
            "ticket_id": ticket_id,
            "entries": [
                {
                    "id": entry.id,
                    "date": entry.date.isoformat(),
                    "employee_id": entry.employee_id,
                    "employee_name": entry.employee_name,
                    "team": entry.team,
                    "hours_logged": entry.hours_logged,
                    "task_description": entry.task_description,
                    "project_name": entry.project_name,
                    "leave_type": entry.leave_type
                }
                for entry in entries
            ],
            "summary": {
                "total_hours": total_hours,
                "total_entries": len(entries),
                "unique_contributors": len(unique_employees),
                "days_worked": len(unique_dates),
                "contributors": list(unique_employees)
            }
        }
    finally:
        db.close()


# ===== TIMESHEET SUBMISSION & APPROVAL API =====

class TimeEntryCreate(BaseModel):
    activity_type: str
    date: date
    hours: float
    task_category: Optional[str] = None  # Ticket | Team Meetings | Customer Support | Training | KT | Leave | Miscellaneous
    ticket_id: Optional[str] = None
    task_description: Optional[str] = None
    project_name: Optional[str] = None
    planned_task_id: Optional[int] = None
    planned_task_source: Optional[str] = None
    variance_notes: Optional[str] = None  # required when hours differ from planned
    variance_reason_type: Optional[str] = None  # unplanned_task | estimate_ineffective | other
    employee_id: Optional[str] = None  # optional for leads/managers

class SubmitRequest(BaseModel):
    week_ending: date
    notes: Optional[str] = None

class EntryReview(BaseModel):
    entry_source: str
    entry_id: int
    status: str  # approved | revision_required | rejected
    productive_hours: Optional[float] = None
    notes: Optional[str] = None

class ApprovalRequest(BaseModel):
    notes: Optional[str] = None
    entry_reviews: Optional[List[EntryReview]] = None


class BulkApprovalRequest(BaseModel):
    submission_ids: List[int]
    action: str  # approve | reject | revision
    notes: Optional[str] = None


def _get_week_boundaries(target_date: date):
    weekday = target_date.weekday()  # Monday=0
    start = target_date - timedelta(days=weekday)
    end = start + timedelta(days=6)
    return start, end


def _next_working_day(start_date: date, db: Session) -> date:
    d = start_date + timedelta(days=1)
    while not is_working_day(d, db, include_optional_holidays=False):
        d += timedelta(days=1)
    return d


def _submission_due_after_week(week_end: date, db: Session) -> date:
    """Return submission due date: next Tuesday after week end (second working day after week_end)."""
    first_wd = _next_working_day(week_end, db)  # Monday
    return _next_working_day(first_wd, db)  # Tuesday


def _get_planned_hours_for_entry(db: Session, entry_date: date, planned_task_id: int, planned_task_source: Optional[str]) -> Optional[float]:
    """Return planned hours for the given task on the given date, or None if not found."""
    if planned_task_source == "dev":
        alloc = db.query(DevPlannedAllocation).filter(
            DevPlannedAllocation.task_id == planned_task_id,
            DevPlannedAllocation.allocation_date == entry_date,
        ).first()
        return round(float(alloc.hours or 0), 1) if alloc else None
    if planned_task_source == "qa":
        alloc = db.query(QAPlannedAllocation).filter(
            QAPlannedAllocation.task_id == planned_task_id,
            QAPlannedAllocation.allocation_date == entry_date,
        ).first()
        return round(float(alloc.hours or 0), 1) if alloc else None
    return None


def _norm_text(v: Optional[str]) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _normalize_person_name(name: Optional[str]) -> str:
    text = str(name or "").lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _compact_person_name(name: Optional[str]) -> str:
    return _normalize_person_name(name).replace(" ", "")


def _same_entry_identity(entry: TimeSheetEntry, req: TimeEntryCreate) -> bool:
    return (
        _norm_text(entry.activity_type) == _norm_text(req.activity_type) and
        _norm_text(entry.task_category) == _norm_text(req.task_category) and
        _norm_text(entry.ticket_id) == _norm_text(req.ticket_id) and
        _norm_text(entry.description) == _norm_text(req.task_description) and
        _norm_text(entry.project_name) == _norm_text(req.project_name) and
        entry.planned_task_id == req.planned_task_id and
        _norm_text(entry.planned_task_source) == _norm_text(req.planned_task_source)
    )


def _apply_entry_reviews(db: Session, submission: TimeSheetSubmission, reviews: Optional[List[EntryReview]], current_user: dict):
    if not reviews:
        return
    ws, we = submission.week_start, submission.week_end
    for r in reviews:
        if r.entry_source not in ["sync", "manual"]:
            raise HTTPException(status_code=400, detail="Invalid entry_source")
        if r.status not in ["approved", "revision_required", "rejected"]:
            raise HTTPException(status_code=400, detail="Invalid entry review status")
        if r.status in ["revision_required", "rejected"] and not (r.notes and str(r.notes).strip()):
            raise HTTPException(status_code=400, detail="Comment is required for rejected/revision-required entry reviews")

        if r.entry_source == "sync":
            entry = db.query(EnhancedTimesheet).filter(EnhancedTimesheet.id == r.entry_id).first()
            if not entry:
                raise HTTPException(status_code=404, detail="Timesheet entry not found")
            if entry.employee_id != submission.employee_id or entry.date < ws or entry.date > we:
                raise HTTPException(status_code=400, detail="Entry does not belong to this submission")
            if r.status == "approved":
                productive = r.productive_hours if r.productive_hours is not None else (entry.hours_logged or 0)
                entry.productive_hours = productive
        else:
            entry = db.query(TimeSheetEntry).filter(TimeSheetEntry.id == r.entry_id).first()
            if not entry:
                raise HTTPException(status_code=404, detail="Manual entry not found")
            if entry.employee_id != submission.employee_id or entry.date < ws or entry.date > we:
                raise HTTPException(status_code=400, detail="Entry does not belong to this submission")
            if r.status == "approved":
                productive = r.productive_hours if r.productive_hours is not None else (entry.hours or 0)
                entry.productive_hours = productive

        review = db.query(TimeSheetEntryReview).filter(
            TimeSheetEntryReview.submission_id == submission.id,
            TimeSheetEntryReview.entry_source == r.entry_source,
            TimeSheetEntryReview.entry_id == r.entry_id,
            TimeSheetEntryReview.reviewed_role == current_user.get("role"),
        ).first()
        if review:
            review.status = r.status
            review.notes = r.notes
            review.productive_hours = r.productive_hours
            review.reviewed_on = datetime.utcnow()
            review.reviewed_by = current_user.get("employee_id") or current_user.get("email")
        else:
            db.add(TimeSheetEntryReview(
                submission_id=submission.id,
                entry_source=r.entry_source,
                entry_id=r.entry_id,
                status=r.status,
                productive_hours=r.productive_hours,
                notes=r.notes,
                reviewed_by=current_user.get("employee_id") or current_user.get("email"),
                reviewed_role=current_user.get("role"),
            ))


def _has_non_approved_entry_reviews(reviews: Optional[List[EntryReview]]) -> bool:
    if not reviews:
        return False
    return any((r.status or "").strip().lower() != "approved" for r in reviews)


def _validate_lead_team_scope(db: Session, submission: TimeSheetSubmission, current_user: dict):
    role = current_user.get("role") or ""
    if "LEAD" not in role or "MANAGER" in role or role == "ADMIN":
        return
    lead_emp = db.query(Employee).filter(Employee.employee_id == current_user.get("employee_id")).first()
    sub_emp = db.query(Employee).filter(Employee.employee_id == submission.employee_id).first()
    if not lead_emp or not sub_emp or not lead_emp.team or not sub_emp.team or lead_emp.team != sub_emp.team:
        raise HTTPException(status_code=403, detail="Insufficient permissions to approve/review this submission")


@app.get("/timesheet")
def timesheet_root():
    """No-auth; confirms timesheet API is available. Use /timesheet/health for health check."""
    return {"message": "Timesheet API", "health": "/timesheet/health", "docs": "Use /timesheet/week (auth required), /timesheet/health, etc."}


@app.get("/timesheet/health")
def timesheet_health():
    """No-auth health check for timesheet API. Returns 200 if timesheet routes are deployed."""
    return {"status": "ok", "message": "Timesheet API is available"}


@app.get("/timesheet/week")
def get_timesheet_week(
    date: Optional[str] = Query(None, description="A date within the week to fetch (YYYY-MM-DD). Defaults to today."),
    employee_id: Optional[str] = Query(None),
    current_user: dict = Depends(get_current_user),
):
    """Return timesheet entries and aggregated hours for the week."""
    db: Session = SessionLocal()
    try:
        # Parse date safely (YYYY-MM-DD); parameter 'date' is the query string
        if date and str(date).strip():
            try:
                target_date = datetime.fromisoformat(str(date).strip()[:10]).date()
            except (ValueError, TypeError):
                target_date = datetime.utcnow().date()
        else:
            target_date = datetime.utcnow().date()
        ws, we = _get_week_boundaries(target_date)

        # Determine which employee to fetch for
        if employee_id:
            # Only leads/managers/admin can fetch other employees
            if not ("LEAD" in current_user.get("role", "") or "MANAGER" in current_user.get("role", "") or current_user.get("role") == "ADMIN"):
                raise HTTPException(status_code=403, detail="Insufficient permissions to view other employees' timesheets")
            target_employee_id = employee_id
        else:
            target_employee_id = current_user.get("employee_id")

        if not target_employee_id:
            raise HTTPException(
                status_code=400,
                detail="Your account is not linked to an employee. Ask an admin to set your employee ID in Settings, or use a manager/lead account linked to an employee to view timesheets."
            )

        # NOTE: Enhanced entries are NOT included - timesheet module uses only manual entries
        enhanced_entries = []  # Empty - timesheet module uses only manual entries

        manual_entries = db.query(TimeSheetEntry).filter(
            TimeSheetEntry.employee_id == target_employee_id,
            TimeSheetEntry.date >= ws,
            TimeSheetEntry.date <= we,
        ).all()

        leave_entries = db.query(LeaveEntry).filter(
            LeaveEntry.employee_id == target_employee_id,
            LeaveEntry.date >= ws,
            LeaveEntry.date <= we,
            LeaveEntry.status == 'approved'
        ).all()

        hours_logged = sum((m.hours or 0) for m in manual_entries)
        leave_hours = sum((l.hours or 0) for l in leave_entries)

        # Check if a submission exists for this week
        submission = db.query(TimeSheetSubmission).filter(
            TimeSheetSubmission.employee_id == target_employee_id,
            TimeSheetSubmission.week_start == ws
        ).first()

        review_map = {}
        if submission:
            reviews = (
                db.query(TimeSheetEntryReview)
                .filter(TimeSheetEntryReview.submission_id == submission.id)
                .order_by(TimeSheetEntryReview.reviewed_on.desc())
                .all()
            )
            for r in reviews:
                key = (r.entry_source, r.entry_id)
                if key not in review_map:
                    review_map[key] = {
                        "status": r.status,
                        "notes": r.notes,
                        "productive_hours": r.productive_hours,
                        "reviewed_on": r.reviewed_on.isoformat() if r.reviewed_on else None,
                        "reviewed_role": r.reviewed_role,
                    }

        # Build entries list
        entries = []
        for e in enhanced_entries:
            rkey = ("sync", e.id)
            time_spent_hours = e.hours_logged or 0
            entries.append({
                "source": "sync",
                "id": e.id,
                "date": e.date.isoformat(),
                "activity_type": e.ticket_id or e.leave_type or 'Work',
                "hours": time_spent_hours,
                "time_spent_hours": time_spent_hours,
                "planned_hours": None,
                "productive_hours": e.productive_hours,
                "ticket_id": e.ticket_id,
                "task_description": e.task_description,
                "project_name": e.project_name,
                "team": e.team,
                "review_status": review_map.get(rkey, {}).get("status"),
                "review_notes": review_map.get(rkey, {}).get("notes"),
                "review_productive_hours": review_map.get(rkey, {}).get("productive_hours"),
            })
        for m in manual_entries:
            rkey = ("manual", m.id)
            planned_hours = None
            if m.planned_task_id and m.planned_task_source:
                planned_hours = _get_planned_hours_for_entry(db, m.date, m.planned_task_id, m.planned_task_source)
            time_spent_hours = m.hours or 0
            entries.append({
                "source": "manual",
                "id": m.id,
                "date": m.date.isoformat(),
                "activity_type": m.activity_type,
                "task_category": getattr(m, "task_category", None),
                "hours": time_spent_hours,
                "time_spent_hours": time_spent_hours,
                "planned_hours": planned_hours,
                "productive_hours": m.productive_hours,
                "ticket_id": m.ticket_id,
                "task_description": m.description,
                "project_name": m.project_name,
                "planned_task_id": m.planned_task_id,
                "planned_task_source": m.planned_task_source,
                "variance_notes": getattr(m, "variance_notes", None),
                "variance_reason_type": getattr(m, "variance_reason_type", None),
                "review_status": review_map.get(rkey, {}).get("status"),
                "review_notes": review_map.get(rkey, {}).get("notes"),
                "review_productive_hours": review_map.get(rkey, {}).get("productive_hours"),
            })

        entries.sort(key=lambda x: (x["date"], x.get("ticket_id") or ""))

        return {
            "employee_id": target_employee_id,
            "week_start": ws.isoformat(),
            "week_end": we.isoformat(),
            "hours_logged": round(hours_logged, 2),
            "leave_hours": round(leave_hours, 2),
            "total_hours": round(hours_logged + leave_hours, 2),
            "submission": {
                "id": submission.id,
                "status": submission.status,
                "submitted_on": submission.submitted_on.isoformat() if submission else None,
            } if submission else None,
            "entries": entries,
        }
    except HTTPException:
        raise
    except Exception as e:
        import logging
        logging.exception("Timesheet week failed")
        raise HTTPException(status_code=500, detail=f"Timesheet error: {str(e)}. Ensure DB tables exist: run backend create_tables.py (includes timesheet_entry_reviews).")
    finally:
        db.close()


@app.get("/timesheet/team-weekly")
def get_team_timesheet_weekly(
    team: str = Query("QA", description="Team: QA or DEVELOPMENT"),
    date_str: str = Query(None, description="Any date in the week (YYYY-MM-DD). Defaults to current week."),
    current_user: dict = Depends(get_current_user),
):
    """
    Get weekly timesheet view showing daily time entries per employee for a team.
    Returns employees as rows with days as columns. Accessible to all authenticated users.
    """
    db = SessionLocal()
    try:
        # Parse date and calculate week boundaries (Monday to Sunday)
        if date_str:
            try:
                target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
        else:
            target_date = date.today()
        
        # Calculate week start (Monday) and end (Sunday)
        week_start = target_date - timedelta(days=target_date.weekday())
        week_end = week_start + timedelta(days=6)
        
        # Map team filter: "DEV" -> "DEVELOPMENT" for Employee table
        employee_team_filter = team.upper()
        if employee_team_filter == "DEV":
            employee_team_filter = "DEVELOPMENT"
        
        # Map team for timesheet filter: EnhancedTimesheet uses "DEV", Employee uses "DEVELOPMENT"
        timesheet_team_filter = team.upper()
        if timesheet_team_filter == "DEVELOPMENT":
            timesheet_team_filter = "DEV"
        
        # Get employees for the team
        emp_query = db.query(Employee).filter(
            Employee.is_active == True,
            Employee.team == employee_team_filter
        )
        employees = emp_query.all()
        
        # Get employee IDs and names
        employee_ids = [emp.employee_id for emp in employees]
        employee_names = [emp.name for emp in employees]
        
        # NOTE: Enhanced timesheet entries (synced from Google Sheets) are NOT included
        # in the Timesheet Module. They are only used by the Calendar Module.
        # The Timesheet Module only shows manual entries and leave entries.
        enhanced_entries = []  # Empty - timesheet module uses only manual entries
        
        # Query manual timesheet entries
        manual_entries = db.query(TimeSheetEntry).filter(
            TimeSheetEntry.date >= week_start,
            TimeSheetEntry.date <= week_end,
            TimeSheetEntry.employee_id.in_(employee_ids)
        ).all()
        
        # Query leave entries
        leave_entries = db.query(LeaveEntry).filter(
            LeaveEntry.date >= week_start,
            LeaveEntry.date <= week_end,
            LeaveEntry.team == timesheet_team_filter,
            LeaveEntry.status == 'approved'
        ).all()
        
        # Build employee data structure
        employees_data = []
        for emp in employees:
            emp_data = {
                "employee_id": emp.employee_id,
                "employee_name": emp.name,
                "team": emp.team,
                "days": {}
            }
            
            # Initialize all days in the week
            for i in range(7):
                day_date = week_start + timedelta(days=i)
                day_key = day_date.isoformat()
                emp_data["days"][day_key] = {
                    "entries": [],
                    "total_hours": 0,
                    "leave_hours": 0
                }
            
            # Add enhanced entries
            for entry in enhanced_entries:
                if entry.employee_id == emp.employee_id:
                    day_key = entry.date.isoformat()
                    if day_key in emp_data["days"]:
                        emp_data["days"][day_key]["entries"].append({
                            "source": "sync",
                            "id": entry.id,
                            "date": day_key,
                            "activity_type": entry.ticket_id or entry.leave_type or 'Work',
                            "hours": entry.hours_logged or 0,
                            "ticket_id": entry.ticket_id,
                            "task_description": entry.task_description,
                            "project_name": entry.project_name,
                        })
                        emp_data["days"][day_key]["total_hours"] += entry.hours_logged or 0
            
            # Add manual entries
            for entry in manual_entries:
                if entry.employee_id == emp.employee_id:
                    day_key = entry.date.isoformat()
                    if day_key in emp_data["days"]:
                        emp_data["days"][day_key]["entries"].append({
                            "source": "manual",
                            "id": entry.id,
                            "date": day_key,
                            "activity_type": entry.activity_type,
                            "task_category": getattr(entry, "task_category", None),
                            "hours": entry.hours or 0,
                            "ticket_id": entry.ticket_id,
                            "task_description": entry.description,
                            "project_name": entry.project_name,
                            "variance_notes": getattr(entry, "variance_notes", None),
                            "variance_reason_type": getattr(entry, "variance_reason_type", None),
                        })
                        emp_data["days"][day_key]["total_hours"] += entry.hours or 0
            
            # Add leave entries
            for leave in leave_entries:
                if leave.employee_id == emp.employee_id:
                    day_key = leave.date.isoformat()
                    if day_key in emp_data["days"]:
                        emp_data["days"][day_key]["entries"].append({
                            "source": "leave",
                            "id": leave.id,
                            "date": day_key,
                            "activity_type": leave.leave_type or "Leave",
                            "task_category": leave.leave_type or "Leave",
                            "hours": leave.hours or 0,
                            "ticket_id": None,
                            "task_description": leave.leave_type or "Leave",
                            "project_name": None,
                        })
                        emp_data["days"][day_key]["leave_hours"] += leave.hours or 0
            
            # Calculate weekly total
            weekly_total = sum(
                day["total_hours"] + day["leave_hours"]
                for day in emp_data["days"].values()
            )
            emp_data["weekly_total"] = round(weekly_total, 2)
            
            employees_data.append(emp_data)
        
        # Managers see all team rows.
        # Leads see self + direct reportees for the selected team.
        # Regular employees see only their own row.
        role = (current_user.get("role") or "").upper()
        is_manager = role == "ADMIN" or "MANAGER" in role
        is_lead = "LEAD" in role and not is_manager
        viewer_team = team
        if not is_manager:
            viewer_employee_id = current_user.get("employee_id")
            if is_lead:
                visible_ids = get_visible_employee_ids(db, current_user) or set()
                employees_data = [e for e in employees_data if e.get("employee_id") in visible_ids]
                # Keep team selection stable for lead views even when no rows in chosen team.
                if employees_data:
                    viewer_team = team
            elif viewer_employee_id:
                employees_data = [e for e in employees_data if e.get("employee_id") == viewer_employee_id]
                if employees_data:
                    viewer_team = employees_data[0].get("team") or team
                else:
                    # Viewer is in the other team: build their row from their actual team
                    viewer_emp = db.query(Employee).filter(
                        Employee.employee_id == viewer_employee_id,
                        Employee.is_active == True,
                    ).first()
                    if viewer_emp:
                        viewer_team = viewer_emp.team or "QA"
                        v_team_ts = "DEV" if (viewer_team or "").upper() in ("DEV", "DEVELOPMENT") else "QA"
                        # NOTE: Enhanced entries are NOT included - timesheet module uses only manual entries
                        enhanced_v = []  # Empty - timesheet module uses only manual entries
                        manual_v = db.query(TimeSheetEntry).filter(
                            TimeSheetEntry.employee_id == viewer_employee_id,
                            TimeSheetEntry.date >= week_start,
                            TimeSheetEntry.date <= week_end,
                        ).all()
                        leave_v = db.query(LeaveEntry).filter(
                            LeaveEntry.employee_id == viewer_employee_id,
                            LeaveEntry.date >= week_start,
                            LeaveEntry.date <= week_end,
                            LeaveEntry.status == "approved",
                        ).all()
                        emp_data = {
                            "employee_id": viewer_emp.employee_id,
                            "employee_name": viewer_emp.name,
                            "team": viewer_emp.team,
                            "days": {},
                        }
                        for i in range(7):
                            day_date = week_start + timedelta(days=i)
                            day_key = day_date.isoformat()
                            emp_data["days"][day_key] = {"entries": [], "total_hours": 0, "leave_hours": 0}
                        for entry in enhanced_v:
                            day_key = entry.date.isoformat()
                            if day_key in emp_data["days"]:
                                emp_data["days"][day_key]["entries"].append({
                                    "source": "sync", "id": entry.id, "date": day_key,
                                    "activity_type": entry.ticket_id or entry.leave_type or "Work",
                                    "hours": entry.hours_logged or 0, "ticket_id": entry.ticket_id,
                                    "task_description": entry.task_description, "project_name": entry.project_name,
                                })
                                emp_data["days"][day_key]["total_hours"] += entry.hours_logged or 0
                        for entry in manual_v:
                            day_key = entry.date.isoformat()
                            if day_key in emp_data["days"]:
                                emp_data["days"][day_key]["entries"].append({
                                    "source": "manual", "id": entry.id, "date": day_key,
                                    "activity_type": entry.activity_type, "task_category": getattr(entry, "task_category", None),
                                    "hours": entry.hours or 0, "ticket_id": entry.ticket_id,
                                    "task_description": entry.description, "project_name": entry.project_name,
                                    "variance_notes": getattr(entry, "variance_notes", None),
                                    "variance_reason_type": getattr(entry, "variance_reason_type", None),
                                })
                                emp_data["days"][day_key]["total_hours"] += entry.hours or 0
                        for leave in leave_v:
                            day_key = leave.date.isoformat()
                            if day_key in emp_data["days"]:
                                emp_data["days"][day_key]["entries"].append({
                                    "source": "leave", "id": leave.id, "date": day_key,
                                    "activity_type": leave.leave_type or "Leave",
                                    "task_category": leave.leave_type or "Leave",
                                    "hours": leave.hours or 0, "ticket_id": None,
                                    "task_description": leave.leave_type or "Leave", "project_name": None,
                                })
                                emp_data["days"][day_key]["leave_hours"] += leave.hours or 0
                        emp_data["weekly_total"] = round(sum(
                            d["total_hours"] + d["leave_hours"] for d in emp_data["days"].values()
                        ), 2)
                        employees_data = [emp_data]
            else:
                employees_data = []
        
        return {
            "team": viewer_team,
            "week_start": week_start.isoformat(),
            "week_end": week_end.isoformat(),
            "employees": employees_data,
            "viewer_can_see_all": is_manager,
        }
    except HTTPException:
        raise
    except Exception as e:
        import logging
        logging.exception("Team timesheet weekly failed")
        raise HTTPException(status_code=500, detail=f"Team timesheet error: {str(e)}")
    finally:
        db.close()


@app.get("/timesheet/month")
def get_timesheet_month(
    month: str = Query(..., description="Month YYYY-MM"),
    employee_id: Optional[str] = Query(None),
    current_user: dict = Depends(get_current_user),
):
    """Return timesheet entries for the given month (for calendar-style monthly view)."""
    db: Session = SessionLocal()
    try:
        try:
            year, mon = map(int, month.split("-"))
            month_start = date(year, mon, 1)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid month. Use YYYY-MM")
        if mon == 12:
            month_end = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            month_end = date(year, mon + 1, 1) - timedelta(days=1)

        target_employee_id = employee_id or current_user.get("employee_id")
        if not target_employee_id:
            raise HTTPException(status_code=400, detail="employee_id is required")
        if employee_id and not ("LEAD" in current_user.get("role", "") or "MANAGER" in current_user.get("role", "") or current_user.get("role") == "ADMIN"):
            raise HTTPException(status_code=403, detail="Insufficient permissions")

        # NOTE: Enhanced entries are NOT included - timesheet module uses only manual entries
        enhanced_entries = []  # Empty - timesheet module uses only manual entries
        manual_entries = db.query(TimeSheetEntry).filter(
            TimeSheetEntry.employee_id == target_employee_id,
            TimeSheetEntry.date >= month_start,
            TimeSheetEntry.date <= month_end,
        ).all()
        leave_entries = db.query(LeaveEntry).filter(
            LeaveEntry.employee_id == target_employee_id,
            LeaveEntry.date >= month_start,
            LeaveEntry.date <= month_end,
            LeaveEntry.status == "approved",
        ).all()

        days_map = {}
        for e in enhanced_entries:
            key = e.date.isoformat()
            if key not in days_map:
                days_map[key] = {"entries": [], "total_hours": 0, "leave_hours": 0}
            days_map[key]["entries"].append({
                "source": "sync", "id": e.id, "date": key, "activity_type": e.ticket_id or e.leave_type or "Work",
                "hours": e.hours_logged or 0, "ticket_id": e.ticket_id, "task_description": e.task_description,
                "project_name": e.project_name, "team": e.team,
            })
            days_map[key]["total_hours"] += e.hours_logged or 0
        for m in manual_entries:
            key = m.date.isoformat()
            if key not in days_map:
                days_map[key] = {"entries": [], "total_hours": 0, "leave_hours": 0}
            days_map[key]["entries"].append({
                "source": "manual", "id": m.id, "date": key, "activity_type": m.activity_type,
                "task_category": getattr(m, "task_category", None), "hours": m.hours or 0,
                "ticket_id": m.ticket_id, "task_description": m.description, "project_name": m.project_name,
            })
            days_map[key]["total_hours"] += m.hours or 0
        for l in leave_entries:
            key = l.date.isoformat()
            if key not in days_map:
                days_map[key] = {"entries": [], "total_hours": 0, "leave_hours": 0}
            days_map[key]["leave_hours"] += l.hours or 0

        return {
            "employee_id": target_employee_id,
            "month": month,
            "month_start": month_start.isoformat(),
            "month_end": month_end.isoformat(),
            "days": days_map,
        }
    finally:
        db.close()


@app.get("/timesheet/planned-tasks")
def get_timesheet_planned_tasks(
    date_str: str = Query(..., description="Date (YYYY-MM-DD)"),
    employee_id: Optional[str] = Query(None),
    current_user: dict = Depends(get_current_user),
):
    """Return planned tasks for a given employee and date."""
    db: Session = SessionLocal()
    try:
        try:
            target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

        target_employee_id = employee_id or current_user.get("employee_id")
        if not target_employee_id:
            raise HTTPException(status_code=400, detail="employee_id is required")

        if employee_id and not can_access_employee(db, current_user, employee_id):
            raise HTTPException(status_code=403, detail="Access denied")

        emp = db.query(Employee).filter(Employee.employee_id == target_employee_id).first()
        if not emp:
            raise HTTPException(status_code=404, detail="Employee not found")

        team_upper = (emp.team or "").upper()
        employee_names_for_match = set()
        if emp.name:
            employee_names_for_match.add(emp.name.strip())

        # Include alternate/canonical names mapped to this employee so planned tasks
        # from sheets with variant names still appear in timesheet modal/calendar flows.
        try:
            name_mappings = db.query(EmployeeNameMapping).filter(
                EmployeeNameMapping.is_active == True,
                or_(
                    EmployeeNameMapping.employee_id == emp.employee_id,
                    EmployeeNameMapping.canonical_name == emp.name
                )
            ).all()
            for m in name_mappings:
                if m.alternate_name:
                    employee_names_for_match.add(m.alternate_name.strip())
                if m.canonical_name:
                    employee_names_for_match.add(m.canonical_name.strip())
        except Exception:
            pass

        planned_tasks = []
        if "DEV" in team_upper or team_upper == "DEVELOPMENT":
            rows = db.query(DevPlannedTask, DevPlannedAllocation).join(
                DevPlannedAllocation, DevPlannedAllocation.task_id == DevPlannedTask.id
            ).filter(
                DevPlannedTask.status == "active",
                DevPlannedTask.employee_name.in_(employee_names_for_match),
                DevPlannedAllocation.allocation_date == target_date,
            ).all()
            for t, a in rows:
                planned_tasks.append({
                    "id": t.id,
                    "ticket_id": t.ticket_id,
                    "ticket_title": t.ticket_title,
                    "activity_description": t.activity_description,
                    "generic_category": t.generic_category,
                    "task_type": None,
                    "hours": round(float(a.hours or 0), 1),
                    "planned_date": target_date.isoformat(),
                    "category": "Ticket" if t.ticket_id else (t.generic_category or "Miscellaneous"),
                    "source": "dev",
                })
        else:
            rows = db.query(QAPlannedTask, QAPlannedAllocation).join(
                QAPlannedAllocation, QAPlannedAllocation.task_id == QAPlannedTask.id
            ).filter(
                QAPlannedTask.status == "active",
                QAPlannedTask.employee_name.in_(employee_names_for_match),
                QAPlannedAllocation.allocation_date == target_date,
            ).all()
            for t, a in rows:
                planned_tasks.append({
                    "id": t.id,
                    "ticket_id": t.ticket_id,
                    "ticket_title": t.ticket_title,
                    "activity_description": t.activity_description,
                    "generic_category": t.generic_category,
                    "task_type": t.task_type,
                    "hours": round(float(a.hours or 0), 1),
                    "planned_date": target_date.isoformat(),
                    "category": "Ticket" if t.ticket_id else (t.generic_category or "Miscellaneous"),
                    "source": "qa",
                })

        return {
            "employee_id": emp.employee_id,
            "employee_name": emp.name,
            "team": emp.team,
            "date": target_date.isoformat(),
            "planned_tasks": planned_tasks,
        }
    finally:
        db.close()


@app.post("/timesheet/entry")
def add_timesheet_entry(
    req: TimeEntryCreate,
    current_user: dict = Depends(get_current_user),
):
    """Add manual timesheet entry. Employees create for themselves; leads/managers may add for others."""
    db: Session = SessionLocal()
    try:
        actor_emp = current_user.get("employee_id")
        actor_role = current_user.get("role")
        target_employee = req.employee_id or actor_emp
        if not target_employee:
            raise HTTPException(status_code=400, detail="employee_id is required")

        # Permission check
        if target_employee != actor_emp and not ("LEAD" in actor_role or "MANAGER" in actor_role or actor_role == "ADMIN"):
            raise HTTPException(status_code=403, detail="Insufficient permissions to add entries for other employees")

        # Insert manual entry
        emp = db.query(Employee).filter(Employee.employee_id == target_employee).first()
        if not emp:
            raise HTTPException(status_code=404, detail="Employee not found")

        # Allow changes until final approval; lock only once fully approved.
        ws, _ = _get_week_boundaries(req.date)
        existing_submission = db.query(TimeSheetSubmission).filter(
            TimeSheetSubmission.employee_id == target_employee,
            TimeSheetSubmission.week_start == ws
        ).first()
        if existing_submission and existing_submission.status in ["Approved"]:
            raise HTTPException(status_code=400, detail="Timesheet already approved for this week. No further edits allowed.")

        if req.task_category == "Ticket" and not (req.ticket_id and str(req.ticket_id).strip()):
            raise HTTPException(status_code=400, detail="Ticket ID is required when task category is Ticket")

        # When linked to a planned task and hours differ from planned, require variance_notes
        if req.planned_task_id and req.planned_task_source:
            planned_hours = _get_planned_hours_for_entry(db, req.date, req.planned_task_id, req.planned_task_source)
            if planned_hours is not None and abs(req.hours - planned_hours) > 0.01:
                if not (req.variance_notes and str(req.variance_notes).strip()):
                    raise HTTPException(
                        status_code=400,
                        detail="Comment (variance notes) is required when entered hours differ from planned hours",
                    )

        # Duplicate-safe save: if a logically identical entry already exists for this day,
        # update it instead of creating a duplicate row.
        same_day_entries = db.query(TimeSheetEntry).filter(
            TimeSheetEntry.employee_id == target_employee,
            TimeSheetEntry.date == req.date,
        ).all()
        duplicate_entry = next((e for e in same_day_entries if _same_entry_identity(e, req)), None)
        if duplicate_entry:
            duplicate_entry.hours = req.hours
            duplicate_entry.variance_notes = req.variance_notes.strip() if req.variance_notes and str(req.variance_notes).strip() else None
            duplicate_entry.variance_reason_type = req.variance_reason_type.strip() if req.variance_reason_type and str(req.variance_reason_type).strip() else None
            db.commit()
            return {"ok": True, "entry_id": duplicate_entry.id, "deduplicated": True}

        entry = TimeSheetEntry(
            employee_id=target_employee,
            employee_name=emp.name,
            date=req.date,
            activity_type=req.activity_type,
            task_category=req.task_category,
            hours=req.hours,
            productive_hours=None,
            ticket_id=req.ticket_id,
            description=req.task_description,
            project_name=req.project_name,
            planned_task_id=req.planned_task_id,
            planned_task_source=req.planned_task_source,
            variance_notes=req.variance_notes.strip() if req.variance_notes and str(req.variance_notes).strip() else None,
            variance_reason_type=req.variance_reason_type.strip() if req.variance_reason_type and str(req.variance_reason_type).strip() else None,
        )
        db.add(entry)
        db.commit()
        db.refresh(entry)

        return {"ok": True, "entry_id": entry.id}
    finally:
        db.close()


@app.delete("/timesheet/entry/{entry_id}")
def delete_timesheet_entry(entry_id: int, current_user: dict = Depends(get_current_user)):
    db: Session = SessionLocal()
    try:
        entry = db.query(TimeSheetEntry).filter(TimeSheetEntry.id == entry_id).first()
        if not entry:
            raise HTTPException(status_code=404, detail="Entry not found")
        actor_emp = current_user.get("employee_id")
        actor_role = current_user.get("role")
        if entry.employee_id != actor_emp and not ("LEAD" in actor_role or "MANAGER" in actor_role or actor_role == "ADMIN"):
            raise HTTPException(status_code=403, detail="Insufficient permissions to delete this entry")
        ws, _ = _get_week_boundaries(entry.date)
        existing_submission = db.query(TimeSheetSubmission).filter(
            TimeSheetSubmission.employee_id == entry.employee_id,
            TimeSheetSubmission.week_start == ws
        ).first()
        db.delete(entry)
        db.commit()
        return {"ok": True}
    finally:
        db.close()


@app.delete("/timesheet/leave-entry/{leave_id}")
def delete_timesheet_leave_entry(leave_id: int, current_user: dict = Depends(get_current_user)):
    db: Session = SessionLocal()
    try:
        leave = db.query(LeaveEntry).filter(LeaveEntry.id == leave_id).first()
        if not leave:
            raise HTTPException(status_code=404, detail="Leave entry not found")
        actor_emp = current_user.get("employee_id")
        actor_role = current_user.get("role")
        if leave.employee_id != actor_emp and not ("LEAD" in actor_role or "MANAGER" in actor_role or actor_role == "ADMIN"):
            raise HTTPException(status_code=403, detail="Insufficient permissions to delete this leave entry")
        db.delete(leave)
        db.commit()
        return {"ok": True}
    finally:
        db.close()


@app.put("/timesheet/entry/{entry_id}")
def update_timesheet_entry(
    entry_id: int,
    req: TimeEntryCreate,
    current_user: dict = Depends(get_current_user),
):
    db: Session = SessionLocal()
    try:
        entry = db.query(TimeSheetEntry).filter(TimeSheetEntry.id == entry_id).first()
        if not entry:
            raise HTTPException(status_code=404, detail="Entry not found")
        actor_emp = current_user.get("employee_id")
        actor_role = current_user.get("role")
        if entry.employee_id != actor_emp and not ("LEAD" in actor_role or "MANAGER" in actor_role or actor_role == "ADMIN"):
            raise HTTPException(status_code=403, detail="Insufficient permissions to edit this entry")

        ws, _ = _get_week_boundaries(req.date)
        existing_submission = db.query(TimeSheetSubmission).filter(
            TimeSheetSubmission.employee_id == entry.employee_id,
            TimeSheetSubmission.week_start == ws
        ).first()
        if existing_submission and existing_submission.status in ["Approved"]:
            raise HTTPException(status_code=400, detail="Timesheet already approved for this week. No further edits allowed.")

        if req.task_category == "Ticket" and not (req.ticket_id and str(req.ticket_id).strip()):
            raise HTTPException(status_code=400, detail="Ticket ID is required when task category is Ticket")

        # When linked to a planned task and hours differ from planned, require variance_notes
        if req.planned_task_id and req.planned_task_source:
            planned_hours = _get_planned_hours_for_entry(db, req.date, req.planned_task_id, req.planned_task_source)
            if planned_hours is not None and abs(req.hours - planned_hours) > 0.01:
                if not (req.variance_notes and str(req.variance_notes).strip()):
                    raise HTTPException(
                        status_code=400,
                        detail="Comment (variance notes) is required when entered hours differ from planned hours",
                    )

        # If edit changes this entry to match another existing row, merge into that row
        # to avoid duplicate entries after save.
        same_day_entries = db.query(TimeSheetEntry).filter(
            TimeSheetEntry.employee_id == entry.employee_id,
            TimeSheetEntry.date == req.date,
            TimeSheetEntry.id != entry.id,
        ).all()
        duplicate_entry = next((e for e in same_day_entries if _same_entry_identity(e, req)), None)
        if duplicate_entry:
            duplicate_entry.hours = req.hours
            duplicate_entry.variance_notes = req.variance_notes.strip() if req.variance_notes and str(req.variance_notes).strip() else None
            duplicate_entry.variance_reason_type = req.variance_reason_type.strip() if req.variance_reason_type and str(req.variance_reason_type).strip() else None
            db.delete(entry)
            db.commit()
            return {"ok": True, "merged_into_entry_id": duplicate_entry.id}

        entry.date = req.date
        entry.activity_type = req.activity_type
        entry.task_category = req.task_category
        entry.hours = req.hours
        entry.ticket_id = req.ticket_id
        entry.description = req.task_description
        entry.project_name = req.project_name
        entry.planned_task_id = req.planned_task_id
        entry.planned_task_source = req.planned_task_source
        entry.variance_notes = req.variance_notes.strip() if req.variance_notes and str(req.variance_notes).strip() else None
        entry.variance_reason_type = req.variance_reason_type.strip() if req.variance_reason_type and str(req.variance_reason_type).strip() else None
        db.commit()
        return {"ok": True}
    finally:
        db.close()


@app.post("/timesheet/submit")
def submit_timesheet(req: SubmitRequest, current_user: dict = Depends(get_current_user)):
    """Submit weekly timesheet for approval. Enforces minimum 40-hour rule."""
    db: Session = SessionLocal()
    try:
        employee_id = current_user.get("employee_id")
        if not employee_id:
            raise HTTPException(status_code=400, detail="Employee account required to submit timesheet")

        week_end = req.week_ending
        ws, we = _get_week_boundaries(week_end)

        # Aggregate hours and leave
        # NOTE: Enhanced entries are NOT included - timesheet module uses only manual entries
        manual_entries = db.query(TimeSheetEntry).filter(
            TimeSheetEntry.employee_id == employee_id,
            TimeSheetEntry.date >= ws,
            TimeSheetEntry.date <= we,
        ).all()
        leave_entries = db.query(LeaveEntry).filter(
            LeaveEntry.employee_id == employee_id,
            LeaveEntry.date >= ws,
            LeaveEntry.date <= we,
            LeaveEntry.status == 'approved'
        ).all()

        hours_logged = sum((m.hours or 0) for m in manual_entries)
        leave_hours = sum((l.hours or 0) for l in leave_entries)
        total = hours_logged + leave_hours

        # Enforce minimum hours (40)
        MIN_HOURS_REQUIRED = 40
        if total < MIN_HOURS_REQUIRED:
            raise HTTPException(status_code=400, detail=f"Cannot submit: total hours ({total:.1f}) is less than required {MIN_HOURS_REQUIRED} hours")

        emp = db.query(Employee).filter(Employee.employee_id == employee_id).first()

        # If submission exists, allow resubmission only for rejected/revision-required
        existing = db.query(TimeSheetSubmission).filter(
            TimeSheetSubmission.employee_id == employee_id,
            TimeSheetSubmission.week_start == ws
        ).first()
        if existing:
            if existing.status not in ["Rejected", "Revision Required"]:
                raise HTTPException(status_code=400, detail="Timesheet already submitted for this week")
            existing.status = "Pending"
            existing.submitted_on = datetime.utcnow()
            existing.total_hours_logged = round(hours_logged, 2)
            existing.leave_hours = round(leave_hours, 2)
            existing.notes = req.notes
            existing.lead_id = None
            existing.manager_id = None
            existing.lead_approved_on = None
            existing.manager_approved_on = None
            db.query(TimeSheetEntryReview).filter(TimeSheetEntryReview.submission_id == existing.id).delete()
            db.commit()
            return {"ok": True, "submission_id": existing.id}

        # Create new submission
        submission = TimeSheetSubmission(
            employee_id=employee_id,
            employee_name=emp.name if emp else None,
            week_start=ws,
            week_end=we,
            status='Pending',
            submitted_on=datetime.utcnow(),
            total_hours_logged=round(hours_logged, 2),
            leave_hours=round(leave_hours, 2),
            notes=req.notes,
        )
        db.add(submission)
        db.commit()
        db.refresh(submission)

        return {"ok": True, "submission_id": submission.id}
    finally:
        db.close()


@app.get("/timesheet/pending-approvals")
def get_pending_approvals(team: Optional[str] = Query(None), current_user: dict = Depends(get_current_user)):
    """Return pending submissions that the current user can approve."""
    db: Session = SessionLocal()
    try:
        role = current_user.get("role")
        if role == "ADMIN" or "MANAGER" in role:
            status_filter = ["Lead Approved"]
        elif "LEAD" in role:
            status_filter = ["Pending"]
        else:
            raise HTTPException(status_code=403, detail="Insufficient permissions")

        q = db.query(TimeSheetSubmission).filter(TimeSheetSubmission.status.in_(status_filter))

        if "LEAD" in role and not ("MANAGER" in role or role == 'ADMIN'):
            # Filter to submissions from employees in the lead's team
            employee = db.query(Employee).filter(Employee.employee_id == current_user.get("employee_id")).first()
            if employee and employee.team:
                team_name = employee.team
                q = q.join(Employee, Employee.employee_id == TimeSheetSubmission.employee_id).filter(Employee.team == team_name)
        elif team:
            # Managers/admin can filter by team param
            q = q.join(Employee, Employee.employee_id == TimeSheetSubmission.employee_id).filter(Employee.team == team)

        subs = q.order_by(TimeSheetSubmission.week_start.desc()).all()
        result = []
        for s in subs:
            result.append({
                "id": s.id,
                "employee_id": s.employee_id,
                "employee_name": s.employee_name,
                "week_start": s.week_start.isoformat(),
                "week_end": s.week_end.isoformat(),
                "status": s.status,
                "submitted_on": s.submitted_on.isoformat() if s.submitted_on else None,
                "total_hours": s.total_hours_logged,
                "leave_hours": s.leave_hours,
            })
        return {"pending_timesheets": result}
    finally:
        db.close()


@app.get("/timesheet/completed-approvals")
def get_completed_approvals(team: Optional[str] = Query(None), current_user: dict = Depends(get_current_user)):
    """Return submissions that have been acted on (completed) by the current approver: lead or manager."""
    db: Session = SessionLocal()
    try:
        role = current_user.get("role")
        if role == "ADMIN" or "MANAGER" in role:
            status_filter = ["Approved", "Rejected", "Revision Required"]
        elif "LEAD" in role:
            status_filter = ["Lead Approved", "Approved", "Rejected", "Revision Required"]
        else:
            raise HTTPException(status_code=403, detail="Insufficient permissions")

        q = db.query(TimeSheetSubmission).filter(TimeSheetSubmission.status.in_(status_filter))

        if "LEAD" in role and not ("MANAGER" in role or role == 'ADMIN'):
            employee = db.query(Employee).filter(Employee.employee_id == current_user.get("employee_id")).first()
            if employee and employee.team:
                q = q.join(Employee, Employee.employee_id == TimeSheetSubmission.employee_id).filter(Employee.team == employee.team)
        elif team:
            q = q.join(Employee, Employee.employee_id == TimeSheetSubmission.employee_id).filter(Employee.team == team)

        subs = q.order_by(TimeSheetSubmission.week_start.desc()).limit(100).all()
        result = []
        for s in subs:
            result.append({
                "id": s.id,
                "employee_id": s.employee_id,
                "employee_name": s.employee_name,
                "week_start": s.week_start.isoformat(),
                "week_end": s.week_end.isoformat(),
                "status": s.status,
                "submitted_on": s.submitted_on.isoformat() if s.submitted_on else None,
                "lead_approved_on": s.lead_approved_on.isoformat() if getattr(s, 'lead_approved_on', None) else None,
                "manager_approved_on": s.manager_approved_on.isoformat() if getattr(s, 'manager_approved_on', None) else None,
                "total_hours": s.total_hours_logged,
                "leave_hours": s.leave_hours,
            })
        return {"completed_timesheets": result}
    finally:
        db.close()


@app.get("/timesheet/manager-summary")
def get_timesheet_manager_summary(
    period: str = Query("week", description="week | month"),
    date_str: Optional[str] = Query(None, description="Reference date for weekly summary (YYYY-MM-DD)"),
    month: Optional[str] = Query(None, description="Month for monthly summary (YYYY-MM)"),
    team: str = Query("ALL", description="QA | DEV | DEVELOPMENT | ALL"),
    category: str = Query("ALL", description="BILLED | UN-BILLED | ALL"),
    current_user: dict = Depends(get_current_user),
):
    db: Session = SessionLocal()
    try:
        role = current_user.get("role") or ""
        if not ("LEAD" in role or "MANAGER" in role or role == "ADMIN"):
            raise HTTPException(status_code=403, detail="Insufficient permissions")

        period_l = (period or "week").strip().lower()
        if period_l == "month":
            if month:
                try:
                    year, mon = map(int, month.split("-"))
                    range_start = date(year, mon, 1)
                except ValueError:
                    raise HTTPException(status_code=400, detail="Invalid month format. Use YYYY-MM")
            else:
                today = date.today()
                range_start = date(today.year, today.month, 1)
            if range_start.month == 12:
                range_end = date(range_start.year + 1, 1, 1) - timedelta(days=1)
            else:
                range_end = date(range_start.year, range_start.month + 1, 1) - timedelta(days=1)
        else:
            try:
                ref_date = datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else date.today()
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
            range_start, range_end = _get_week_boundaries(ref_date)

        visible_ids = get_visible_employee_ids(db, current_user)
        emp_query = db.query(Employee).filter(Employee.is_active == True)
        if visible_ids is not None:
            emp_query = emp_query.filter(Employee.employee_id.in_(list(visible_ids)))

        # Leads can only see their own team in summary
        if "LEAD" in role and not ("MANAGER" in role or role == "ADMIN"):
            lead_emp = db.query(Employee).filter(Employee.employee_id == current_user.get("employee_id")).first()
            if lead_emp and lead_emp.team:
                emp_query = emp_query.filter(Employee.team == lead_emp.team)

        team_upper = (team or "ALL").upper()
        if team_upper != "ALL":
            team_filter = "DEVELOPMENT" if team_upper == "DEV" else team_upper
            emp_query = emp_query.filter(Employee.team == team_filter)

        category_upper = (category or "ALL").upper()
        if category_upper != "ALL":
            if category_upper in ["UN-BILLED", "UNBILLED"]:
                emp_query = emp_query.filter(
                    or_(func.upper(Employee.category) == "UN-BILLED", func.upper(Employee.category) == "UNBILLED")
                )
            else:
                emp_query = emp_query.filter(func.upper(Employee.category) == category_upper)

        employees = emp_query.all()
        if not employees:
            return {
                "period": period_l,
                "range_start": range_start.isoformat(),
                "range_end": range_end.isoformat(),
                "totals": {"productive_hours": 0, "leave_hours": 0, "leave_days": 0, "working_days": 0, "expected_hours": 0},
                "by_team": {},
                "by_category": {},
            }

        employee_ids = [e.employee_id for e in employees]
        emp_meta = {
            e.employee_id: {
                "team": (e.team or "").upper(),
                "category": (e.category or "UNSPECIFIED").upper(),
                "productive_hours": 0.0,
                "leave_hours": 0.0,
                "leave_days": 0.0,
                "working_days": 0,
                "expected_hours": 0.0,
            }
            for e in employees
        }

        entries = db.query(TimeSheetEntry).filter(
            TimeSheetEntry.employee_id.in_(employee_ids),
            TimeSheetEntry.date >= range_start,
            TimeSheetEntry.date <= range_end,
        ).all()
        for e in entries:
            if e.employee_id not in emp_meta:
                continue
            productive = e.productive_hours if e.productive_hours is not None else (e.hours or 0.0)
            emp_meta[e.employee_id]["productive_hours"] += float(productive or 0.0)

        leaves = db.query(LeaveEntry).filter(
            LeaveEntry.employee_id.in_(employee_ids),
            LeaveEntry.date >= range_start,
            LeaveEntry.date <= range_end,
            LeaveEntry.status == "approved",
        ).all()
        for l in leaves:
            if l.employee_id not in emp_meta:
                continue
            leave_hours = float(l.hours or 0.0)
            emp_meta[l.employee_id]["leave_hours"] += leave_hours
            emp_meta[l.employee_id]["leave_days"] += round(leave_hours / 8.0, 2) if leave_hours > 0 else 0

        working_days = get_working_days_in_range(range_start, range_end, db, include_optional_holidays=False)
        for emp_id in emp_meta:
            emp_meta[emp_id]["working_days"] = working_days
            expected = (working_days * 8.0) - emp_meta[emp_id]["leave_hours"]
            emp_meta[emp_id]["expected_hours"] = round(max(expected, 0.0), 2)

        def _empty_bucket():
            return {"productive_hours": 0.0, "leave_hours": 0.0, "leave_days": 0.0, "working_days": 0.0, "expected_hours": 0.0, "employees": 0}

        totals = _empty_bucket()
        by_team: Dict[str, Dict[str, Any]] = defaultdict(_empty_bucket)
        by_category: Dict[str, Dict[str, Any]] = defaultdict(_empty_bucket)

        for meta in emp_meta.values():
            for bucket in (totals, by_team[meta["team"]], by_category[meta["category"]]):
                bucket["productive_hours"] += meta["productive_hours"]
                bucket["leave_hours"] += meta["leave_hours"]
                bucket["leave_days"] += meta["leave_days"]
                bucket["working_days"] += meta["working_days"]
                bucket["expected_hours"] += meta["expected_hours"]
                bucket["employees"] += 1

        return {
            "period": period_l,
            "range_start": range_start.isoformat(),
            "range_end": range_end.isoformat(),
            "totals": {k: round(v, 2) if isinstance(v, float) else v for k, v in totals.items()},
            "by_team": {k: {ik: round(iv, 2) if isinstance(iv, float) else iv for ik, iv in v.items()} for k, v in by_team.items()},
            "by_category": {k: {ik: round(iv, 2) if isinstance(iv, float) else iv for ik, iv in v.items()} for k, v in by_category.items()},
        }
    finally:
        db.close()


@app.get("/timesheet/submission/{submission_id}")
def get_timesheet_submission(submission_id: int, current_user: dict = Depends(get_current_user)):
    db: Session = SessionLocal()
    try:
        s = db.query(TimeSheetSubmission).filter(TimeSheetSubmission.id == submission_id).first()
        if not s:
            raise HTTPException(status_code=404, detail="Submission not found")

        if not can_access_employee(db, current_user, s.employee_id):
            raise HTTPException(status_code=403, detail="Access denied")

        ws, we = s.week_start, s.week_end
        # NOTE: Enhanced entries are NOT included - timesheet module uses only manual entries
        enhanced_entries = []  # Empty - timesheet module uses only manual entries
        manual_entries = db.query(TimeSheetEntry).filter(
            TimeSheetEntry.employee_id == s.employee_id,
            TimeSheetEntry.date >= ws,
            TimeSheetEntry.date <= we,
        ).all()

        review_map = {}
        reviews = (
            db.query(TimeSheetEntryReview)
            .filter(TimeSheetEntryReview.submission_id == s.id)
            .order_by(TimeSheetEntryReview.reviewed_on.desc())
            .all()
        )
        for r in reviews:
            key = (r.entry_source, r.entry_id)
            if key not in review_map:
                review_map[key] = {
                    "status": r.status,
                    "notes": r.notes,
                    "productive_hours": r.productive_hours,
                    "reviewed_on": r.reviewed_on.isoformat() if r.reviewed_on else None,
                    "reviewed_role": r.reviewed_role,
                }

        entries = []
        for e in enhanced_entries:
            rkey = ("sync", e.id)
            entries.append({
                "source": "sync",
                "id": e.id,
                "date": e.date.isoformat(),
                "activity_type": e.ticket_id or e.leave_type or 'Work',
                "hours": e.hours_logged or 0,
                "productive_hours": e.productive_hours,
                "ticket_id": e.ticket_id,
                "task_description": e.task_description,
                "project_name": e.project_name,
                "team": e.team,
                "review_status": review_map.get(rkey, {}).get("status"),
                "review_notes": review_map.get(rkey, {}).get("notes"),
                "review_productive_hours": review_map.get(rkey, {}).get("productive_hours"),
            })
        for m in manual_entries:
            rkey = ("manual", m.id)
            entries.append({
                "source": "manual",
                "id": m.id,
                "date": m.date.isoformat(),
                "activity_type": m.activity_type,
                "task_category": getattr(m, "task_category", None),
                "hours": m.hours or 0,
                "productive_hours": m.productive_hours,
                "ticket_id": m.ticket_id,
                "task_description": m.description,
                "project_name": m.project_name,
                "planned_task_id": m.planned_task_id,
                "planned_task_source": m.planned_task_source,
                "variance_notes": getattr(m, "variance_notes", None),
                "variance_reason_type": getattr(m, "variance_reason_type", None),
                "review_status": review_map.get(rkey, {}).get("status"),
                "review_notes": review_map.get(rkey, {}).get("notes"),
                "review_productive_hours": review_map.get(rkey, {}).get("productive_hours"),
            })

        entries.sort(key=lambda x: (x["date"], x.get("ticket_id") or ""))

        return {
            "submission": {
                "id": s.id,
                "employee_id": s.employee_id,
                "employee_name": s.employee_name,
                "week_start": ws.isoformat(),
                "week_end": we.isoformat(),
                "status": s.status,
                "submitted_on": s.submitted_on.isoformat() if s.submitted_on else None,
                "total_hours": s.total_hours_logged,
                "leave_hours": s.leave_hours,
                "notes": s.notes,
            },
            "entries": entries,
        }
    finally:
        db.close()


@app.post("/timesheet/approve/{submission_id}")
def approve_submission(submission_id: int, req: ApprovalRequest, current_user: dict = Depends(get_current_user)):
    db: Session = SessionLocal()
    try:
        s = db.query(TimeSheetSubmission).filter(TimeSheetSubmission.id == submission_id).first()
        if not s:
            raise HTTPException(status_code=404, detail="Submission not found")

        role = current_user.get("role") or ""
        user_emp = current_user.get("employee_id")
        _validate_lead_team_scope(db, s, current_user)
        has_non_approved_reviews = _has_non_approved_entry_reviews(req.entry_reviews)
        if "LEAD" in role:
            if s.status not in ["Pending", "Revision Required"]:
                raise HTTPException(status_code=400, detail="Submission not pending lead approval")
            if has_non_approved_reviews:
                s.status = "Revision Required"
                action = "revision_requested"
            else:
                s.status = 'Lead Approved'
                s.lead_id = user_emp
                s.lead_approved_on = datetime.utcnow()
                action = 'lead_approved'
        elif "MANAGER" in role or role == 'ADMIN':
            if s.status not in ["Lead Approved"]:
                raise HTTPException(status_code=400, detail="Submission not pending manager approval")
            if has_non_approved_reviews:
                s.status = "Revision Required"
                action = "revision_requested"
            else:
                s.status = 'Approved'
                s.manager_id = user_emp
                s.manager_approved_on = datetime.utcnow()
                action = 'approved'
        else:
            raise HTTPException(status_code=403, detail="Insufficient permissions to approve")

        _apply_entry_reviews(db, s, req.entry_reviews, current_user)

        log = TimeSheetApprovalLog(
            submission_id=s.id,
            approver_id=user_emp,
            approver_role=role,
            action=action,
            notes=req.notes,
        )
        db.add(log)
        db.commit()
        return {"ok": True}
    finally:
        db.close()


@app.post("/timesheet/reject/{submission_id}")
def reject_submission(submission_id: int, req: ApprovalRequest, current_user: dict = Depends(get_current_user)):
    db: Session = SessionLocal()
    try:
        s = db.query(TimeSheetSubmission).filter(TimeSheetSubmission.id == submission_id).first()
        if not s:
            raise HTTPException(status_code=404, detail="Submission not found")

        role = current_user.get("role") or ""
        user_emp = current_user.get("employee_id")
        _validate_lead_team_scope(db, s, current_user)
        if not ("LEAD" in role or "MANAGER" in role or role == 'ADMIN'):
            raise HTTPException(status_code=403, detail="Insufficient permissions to reject")
        if s.status == "Approved":
            raise HTTPException(status_code=400, detail="Cannot reject an approved submission")

        s.status = 'Rejected'
        _apply_entry_reviews(db, s, req.entry_reviews, current_user)
        log = TimeSheetApprovalLog(
            submission_id=s.id,
            approver_id=user_emp,
            approver_role=role,
            action='rejected',
            notes=req.notes,
        )
        db.add(log)
        db.commit()
        return {"ok": True}
    finally:
        db.close()


@app.post("/timesheet/request-revision/{submission_id}")
def request_revision(submission_id: int, req: ApprovalRequest, current_user: dict = Depends(get_current_user)):
    db: Session = SessionLocal()
    try:
        s = db.query(TimeSheetSubmission).filter(TimeSheetSubmission.id == submission_id).first()
        if not s:
            raise HTTPException(status_code=404, detail="Submission not found")

        role = current_user.get("role") or ""
        user_emp = current_user.get("employee_id")
        _validate_lead_team_scope(db, s, current_user)
        if not ("LEAD" in role or "MANAGER" in role or role == 'ADMIN'):
            raise HTTPException(status_code=403, detail="Insufficient permissions to request revision")
        if s.status == "Approved":
            raise HTTPException(status_code=400, detail="Cannot request revision for an approved submission")

        s.status = 'Revision Required'
        _apply_entry_reviews(db, s, req.entry_reviews, current_user)
        log = TimeSheetApprovalLog(
            submission_id=s.id,
            approver_id=user_emp,
            approver_role=role,
            action='revision_requested',
            notes=req.notes,
        )
        db.add(log)
        db.commit()
        return {"ok": True}
    finally:
        db.close()


@app.post("/timesheet/approvals/bulk")
def bulk_approval_action(req: BulkApprovalRequest, current_user: dict = Depends(get_current_user)):
    db: Session = SessionLocal()
    try:
        role = current_user.get("role") or ""
        if not ("LEAD" in role or "MANAGER" in role or role == "ADMIN"):
            raise HTTPException(status_code=403, detail="Insufficient permissions to process bulk approvals")
        action = (req.action or "").strip().lower()
        if action not in ["approve", "reject", "revision"]:
            raise HTTPException(status_code=400, detail="Invalid action. Use approve, reject, or revision")
        submission_ids = [int(sid) for sid in (req.submission_ids or []) if sid is not None]
        if not submission_ids:
            raise HTTPException(status_code=400, detail="At least one submission id is required")

        results = []
        for sid in submission_ids:
            s = db.query(TimeSheetSubmission).filter(TimeSheetSubmission.id == sid).first()
            if not s:
                results.append({"submission_id": sid, "ok": False, "error": "Submission not found"})
                continue
            try:
                _validate_lead_team_scope(db, s, current_user)
                user_emp = current_user.get("employee_id")
                if action == "approve":
                    if "LEAD" in role:
                        if s.status not in ["Pending", "Revision Required"]:
                            raise HTTPException(status_code=400, detail="Submission not pending lead approval")
                        s.status = "Lead Approved"
                        s.lead_id = user_emp
                        s.lead_approved_on = datetime.utcnow()
                        log_action = "lead_approved"
                    elif "MANAGER" in role or role == "ADMIN":
                        if s.status not in ["Lead Approved"]:
                            raise HTTPException(status_code=400, detail="Submission not pending manager approval")
                        s.status = "Approved"
                        s.manager_id = user_emp
                        s.manager_approved_on = datetime.utcnow()
                        log_action = "approved"
                    else:
                        raise HTTPException(status_code=403, detail="Insufficient permissions to approve")
                elif action == "reject":
                    if s.status == "Approved":
                        raise HTTPException(status_code=400, detail="Cannot reject an approved submission")
                    s.status = "Rejected"
                    log_action = "rejected"
                else:
                    if s.status == "Approved":
                        raise HTTPException(status_code=400, detail="Cannot request revision for an approved submission")
                    s.status = "Revision Required"
                    log_action = "revision_requested"

                db.add(TimeSheetApprovalLog(
                    submission_id=s.id,
                    approver_id=current_user.get("employee_id"),
                    approver_role=role,
                    action=log_action,
                    notes=req.notes,
                ))
                results.append({"submission_id": sid, "ok": True, "status": s.status})
            except HTTPException as e:
                results.append({"submission_id": sid, "ok": False, "error": e.detail})

        db.commit()
        success_count = sum(1 for r in results if r.get("ok"))
        return {"ok": True, "processed": len(results), "successful": success_count, "results": results}
    finally:
        db.close()


@app.get("/timesheet/my-week-summary")
def get_my_week_summary(
    date_str: str = Query(..., description="Any date in week (YYYY-MM-DD)"),
    current_user: dict = Depends(get_current_user),
):
    """Return current user's week summary: total_hours, leave_hours, by_category, ticket_count for submit 40h check and summary display."""
    db: Session = SessionLocal()
    try:
        employee_id = current_user.get("employee_id")
        if not employee_id:
            raise HTTPException(status_code=400, detail="Employee account required")
        try:
            target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
        ws, we = _get_week_boundaries(target_date)

        manual_entries = db.query(TimeSheetEntry).filter(
            TimeSheetEntry.employee_id == employee_id,
            TimeSheetEntry.date >= ws,
            TimeSheetEntry.date <= we,
        ).all()
        leave_entries = db.query(LeaveEntry).filter(
            LeaveEntry.employee_id == employee_id,
            LeaveEntry.date >= ws,
            LeaveEntry.date <= we,
            LeaveEntry.status == "approved",
        ).all()
        # NOTE: Enhanced entries are NOT included - timesheet module uses only manual entries

        hours_logged = sum(e.hours or 0 for e in manual_entries)
        leave_hours = sum(l.hours or 0 for l in leave_entries)
        total_hours = round(hours_logged + leave_hours, 2)

        by_category = {}
        for e in manual_entries:
            cat = getattr(e, "task_category", None) or "Miscellaneous"
            by_category[cat] = by_category.get(cat, 0) + (e.hours or 0)

        ticket_ids = set()
        for e in manual_entries:
            if e.ticket_id and str(e.ticket_id).strip():
                ticket_ids.add(str(e.ticket_id).strip())

        return {
            "week_start": ws.isoformat(),
            "week_end": we.isoformat(),
            "hours_logged": round(hours_logged, 2),
            "leave_hours": round(leave_hours, 2),
            "total_hours": total_hours,
            "by_category": by_category,
            "ticket_count": len(ticket_ids),
        }
    finally:
        db.close()


@app.get("/timesheet/my-submissions")
def get_my_submissions(current_user: dict = Depends(get_current_user)):
    """Return current user's own timesheet submissions (for My submissions tab)."""
    db: Session = SessionLocal()
    try:
        employee_id = current_user.get("employee_id")
        if not employee_id:
            raise HTTPException(status_code=400, detail="Employee account required")

        subs = (
            db.query(TimeSheetSubmission)
            .filter(TimeSheetSubmission.employee_id == employee_id)
            .order_by(TimeSheetSubmission.week_start.desc())
            .all()
        )
        return [
            {
                "id": s.id,
                "week_start": s.week_start.isoformat(),
                "week_end": s.week_end.isoformat(),
                "status": s.status,
                "submitted_on": s.submitted_on.isoformat() if s.submitted_on else None,
                "total_hours_logged": s.total_hours_logged,
                "leave_hours": s.leave_hours,
                "lead_approved_on": s.lead_approved_on.isoformat() if s.lead_approved_on else None,
                "manager_approved_on": s.manager_approved_on.isoformat() if s.manager_approved_on else None,
            }
            for s in subs
        ]
    finally:
        db.close()


@app.get("/timesheet/lock-status")
def get_timesheet_lock_status(current_user: dict = Depends(get_current_user)):
    """
    Enforce timesheet submission and approval deadlines.
    - Employees must submit last week's timesheet by the next working day.
    - Leads/Managers must clear pending approvals by the following working day.
    """
    db: Session = SessionLocal()
    try:
        today = date.today()
        role = current_user.get("role", "")
        employee_id = current_user.get("employee_id")

        # Check missing submission for last week (due: next Tuesday after week end)
        if employee_id:
            last_week_ref = today - timedelta(days=7)
            last_ws, last_we = _get_week_boundaries(last_week_ref)
            submission_due = _submission_due_after_week(last_we, db)

            if today >= submission_due:
                submission = db.query(TimeSheetSubmission).filter(
                    TimeSheetSubmission.employee_id == employee_id,
                    TimeSheetSubmission.week_start == last_ws
                ).first()
                if not submission or submission.status in ["Rejected", "Revision Required"]:
                    return {
                        "locked": True,
                        "reason": "submission_overdue",
                        "message": "Submit last week's timesheet to continue.",
                        "week_start": last_ws.isoformat(),
                        "week_end": last_we.isoformat(),
                        "due_date": submission_due.isoformat(),
                    }

        # Check pending approvals for leads/managers after approval day
        if "LEAD" in role or "MANAGER" in role or role == "ADMIN":
            last_week_ref = today - timedelta(days=7)
            last_ws, last_we = _get_week_boundaries(last_week_ref)
            approval_day = _next_working_day(last_we, db)
            lock_day = _next_working_day(approval_day, db)

            if today >= lock_day:
                if role == "ADMIN" or "MANAGER" in role:
                    status_filter = ["Lead Approved"]
                else:
                    status_filter = ["Pending"]

                q = db.query(TimeSheetSubmission).filter(TimeSheetSubmission.status.in_(status_filter))

                # Scope to team for leads
                if "LEAD" in role and "MANAGER" not in role and role != "ADMIN":
                    emp = db.query(Employee).filter(Employee.employee_id == employee_id).first() if employee_id else None
                    if emp and emp.team:
                        q = q.join(Employee, Employee.employee_id == TimeSheetSubmission.employee_id).filter(
                            Employee.team == emp.team
                        )

                pending_count = q.count()
                if pending_count > 0:
                    return {
                        "locked": True,
                        "reason": "pending_approvals",
                        "message": "Pending timesheet approvals must be completed to continue.",
                        "week_start": last_ws.isoformat(),
                        "week_end": last_we.isoformat(),
                        "due_date": approval_day.isoformat(),
                        "pending_count": pending_count,
                    }

        return {"locked": False}
    finally:
        db.close()


@app.get("/calendar/leaves")
def get_team_leaves(
    team: str = Query("ALL", description="Team: QA, DEV, or ALL"),
    month: str = Query(None, description="Month (YYYY-MM). Defaults to current month.")
):
    """
    Get leave entries for a team in a given month.
    """
    db = SessionLocal()
    try:
        # Parse month
        if month:
            try:
                year, mon = map(int, month.split("-"))
                month_start = date(year, mon, 1)
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid month format. Use YYYY-MM")
        else:
            today = date.today()
            month_start = date(today.year, today.month, 1)
        
        # Calculate month end
        if month_start.month == 12:
            month_end = date(month_start.year + 1, 1, 1) - timedelta(days=1)
        else:
            month_end = date(month_start.year, month_start.month + 1, 1) - timedelta(days=1)

        # Query leaves
        query = db.query(LeaveEntry).filter(
            LeaveEntry.date >= month_start,
            LeaveEntry.date <= month_end
        )
        if team.upper() != "ALL":
            query = query.filter(LeaveEntry.team == team.upper())
        
        leaves = query.order_by(LeaveEntry.date, LeaveEntry.employee_name).all()
        
        return {
            "month": month_start.strftime("%Y-%m"),
            "team": team,
            "leaves": [
                {
                    "id": l.id,
                    "employee_id": l.employee_id,
                    "employee_name": l.employee_name,
                    "date": l.date.isoformat(),
                    "leave_type": l.leave_type,
                    "hours": l.hours,
                    "status": l.status
                }
                for l in leaves
            ],
            "summary": {
                "total_leave_entries": len(leaves),
                "by_type": dict(defaultdict(int, {l.leave_type: sum(1 for x in leaves if x.leave_type == l.leave_type) for l in leaves}))
            }
        }
    finally:
        db.close()


# ===== TASK PLANNING API ENDPOINTS =====

@app.get("/planning/weekly")
def get_weekly_plan(
    team: str = Query("ALL", description="Team: QA, DEV, or ALL"),
    week_start: str = Query(None, description="Week start date (YYYY-MM-DD). Defaults to current week.")
):
    """
    Get weekly task planning for a team.
    Returns planned tasks for each employee in the week.
    """
    db = SessionLocal()
    try:
        # Parse week start
        if week_start:
            try:
                start_date = datetime.strptime(week_start, "%Y-%m-%d").date()
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
        else:
            today = date.today()
            start_date = today - timedelta(days=today.weekday())
        
        week_end = start_date + timedelta(days=6)
        
        # Query weekly plans
        plan_query = db.query(WeeklyPlan).filter(
            WeeklyPlan.week_start == start_date
        )
        if team.upper() != "ALL":
            plan_query = plan_query.filter(WeeklyPlan.team == team.upper())
        weekly_plans = plan_query.all()
        
        # Query individual planned tasks
        task_query = db.query(PlannedTask).filter(
            PlannedTask.planned_date >= start_date,
            PlannedTask.planned_date <= week_end
        )
        if team.upper() != "ALL":
            task_query = task_query.filter(PlannedTask.team == team.upper())
        tasks = task_query.order_by(PlannedTask.employee_name, PlannedTask.planned_date).all()
        
        # Get employees
        emp_query = db.query(Employee).filter(Employee.is_active == True)
        if team.upper() != "ALL":
            # Map "DEV" to "DEVELOPMENT" for Employee table (Employee.team uses "DEVELOPMENT", not "DEV")
            employee_team_filter = team.upper()
            if employee_team_filter == "DEV":
                employee_team_filter = "DEVELOPMENT"
            emp_query = emp_query.filter(Employee.team == employee_team_filter)
        employees = emp_query.all()
        
        # Build response
        employee_plans = {}
        
        for emp in employees:
            employee_plans[emp.name] = {
                "employee_id": emp.employee_id,
                "employee_name": emp.name,
                "team": emp.team,
                "weekly_plan": None,
                "daily_tasks": {}
            }
            # Initialize days
            for i in range(7):
                day = start_date + timedelta(days=i)
                employee_plans[emp.name]["daily_tasks"][day.isoformat()] = []
        
        # Add weekly plans
        for plan in weekly_plans:
            name = plan.employee_name
            if name in employee_plans:
                employee_plans[name]["weekly_plan"] = {
                    "id": plan.id,
                    "assigned_tickets": plan.assigned_tickets,
                    "total_planned_hours": plan.total_planned_hours,
                    "notes": plan.notes,
                    "status": plan.status
                }
        
        # Add daily tasks
        for task in tasks:
            name = task.employee_name
            if name not in employee_plans:
                employee_plans[name] = {
                    "employee_id": task.employee_id,
                    "employee_name": name,
                    "team": task.team,
                    "weekly_plan": None,
                    "daily_tasks": {}
                }
                for i in range(7):
                    day = start_date + timedelta(days=i)
                    employee_plans[name]["daily_tasks"][day.isoformat()] = []
            
            day_key = task.planned_date.isoformat()
            if day_key in employee_plans[name]["daily_tasks"]:
                employee_plans[name]["daily_tasks"][day_key].append({
                    "id": task.id,
                    "ticket_id": task.ticket_id,
                    "task_title": task.task_title,
                    "planned_hours": task.planned_hours,
                    "priority": task.priority,
                    "status": task.status,
                    "project_name": task.project_name
                })
        
        return {
            "week_start": start_date.isoformat(),
            "week_end": week_end.isoformat(),
            "team": team,
            "employees": list(employee_plans.values())
        }
    finally:
        db.close()


@app.post("/planning/task")
def create_planned_task(task: PlannedTaskCreate):
    """
    Create a new planned task for an employee.
    """
    db = SessionLocal()
    try:
        # Look up employee ID if not provided
        employee_id = task.employee_id
        if not employee_id:
            employee = db.query(Employee).filter(
                Employee.name.ilike(f"%{task.employee_name}%")
            ).first()
            if employee:
                employee_id = employee.employee_id

        new_task = PlannedTask(
            employee_id=employee_id,
            employee_name=task.employee_name,
            ticket_id=task.ticket_id,
            task_title=task.task_title,
            task_description=task.task_description,
            project_name=task.project_name,
            planned_date=task.planned_date,
            planned_hours=task.planned_hours,
            priority=task.priority,
            team=task.team,
            assigned_by=task.assigned_by,
            status='planned'
        )
        
        db.add(new_task)
        db.commit()
        db.refresh(new_task)
        
        return {
            "success": True,
            "task": {
                "id": new_task.id,
                "employee_name": new_task.employee_name,
                "ticket_id": new_task.ticket_id,
                "task_title": new_task.task_title,
                "planned_date": new_task.planned_date.isoformat(),
                "planned_hours": new_task.planned_hours,
                "priority": new_task.priority,
                "status": new_task.status
            }
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


@app.put("/planning/task/{task_id}")
def update_planned_task(task_id: int, updates: PlannedTaskUpdate):
    """
    Update an existing planned task.
    """
    db = SessionLocal()
    try:
        task = db.query(PlannedTask).filter(PlannedTask.id == task_id).first()
        if not task:
            raise HTTPException(status_code=404, detail="Task not found")
        
        if updates.task_title is not None:
            task.task_title = updates.task_title
        if updates.task_description is not None:
            task.task_description = updates.task_description
        if updates.planned_hours is not None:
            task.planned_hours = updates.planned_hours
        if updates.priority is not None:
            task.priority = updates.priority
        if updates.status is not None:
            task.status = updates.status
            if updates.status == 'completed':
                task.completed_on = datetime.utcnow()
        if updates.actual_hours is not None:
            task.actual_hours = updates.actual_hours
        
        db.commit()
        db.refresh(task)
        
        return {
            "success": True,
            "task": {
                "id": task.id,
                "employee_name": task.employee_name,
                "ticket_id": task.ticket_id,
                "task_title": task.task_title,
                "planned_date": task.planned_date.isoformat(),
                "planned_hours": task.planned_hours,
                "actual_hours": task.actual_hours,
                "priority": task.priority,
                "status": task.status
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


@app.delete("/planning/task/{task_id}")
def delete_planned_task(task_id: int):
    """
    Delete a planned task.
    """
    db = SessionLocal()
    try:
        task = db.query(PlannedTask).filter(PlannedTask.id == task_id).first()
        if not task:
            raise HTTPException(status_code=404, detail="Task not found")
        
        db.delete(task)
        db.commit()
        
        return {"success": True, "message": f"Task {task_id} deleted"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


@app.post("/planning/weekly-plan")
def create_weekly_plan(plan: WeeklyPlanCreate):
    """
    Create or update a weekly plan for an employee.
    """
    db = SessionLocal()
    try:
        # Calculate week details
        week_start = plan.week_start
        # Ensure it's a Monday
        if week_start.weekday() != 0:
            week_start = week_start - timedelta(days=week_start.weekday())
        week_end = week_start + timedelta(days=6)
        year = week_start.isocalendar()[0]
        week_number = week_start.isocalendar()[1]
        
        # Look up employee ID if not provided
        employee_id = plan.employee_id
        if not employee_id:
            employee = db.query(Employee).filter(
                Employee.name.ilike(f"%{plan.employee_name}%")
            ).first()
            if employee:
                employee_id = employee.employee_id
        
        # Calculate total planned hours from tickets
        total_hours = sum(t.get('estimated_hours', 0) for t in plan.assigned_tickets)
        
        # Check if plan exists
        existing = db.query(WeeklyPlan).filter(
            WeeklyPlan.employee_name == plan.employee_name,
            WeeklyPlan.week_start == week_start
        ).first()
        
        if existing:
            # Update existing plan
            existing.assigned_tickets = plan.assigned_tickets
            existing.total_planned_hours = total_hours
            existing.notes = plan.notes
            db.commit()
            db.refresh(existing)
            weekly_plan = existing
        else:
            # Create new plan
            weekly_plan = WeeklyPlan(
                employee_id=employee_id,
                employee_name=plan.employee_name,
                week_start=week_start,
                week_end=week_end,
                year=year,
                week_number=week_number,
                assigned_tickets=plan.assigned_tickets,
                total_planned_hours=total_hours,
                notes=plan.notes,
                team=plan.team,
                planned_by=plan.planned_by,
                status='draft'
            )
            db.add(weekly_plan)
            db.commit()
            db.refresh(weekly_plan)
        
        return {
            "success": True,
            "plan": {
                "id": weekly_plan.id,
                "employee_name": weekly_plan.employee_name,
                "week_start": weekly_plan.week_start.isoformat(),
                "week_end": weekly_plan.week_end.isoformat(),
                "assigned_tickets": weekly_plan.assigned_tickets,
                "total_planned_hours": weekly_plan.total_planned_hours,
                "status": weekly_plan.status
            }
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


@app.put("/planning/weekly-plan/{plan_id}")
def update_weekly_plan(plan_id: int, updates: WeeklyPlanUpdate):
    """
    Update a weekly plan.
    """
    db = SessionLocal()
    try:
        plan = db.query(WeeklyPlan).filter(WeeklyPlan.id == plan_id).first()
        if not plan:
            raise HTTPException(status_code=404, detail="Weekly plan not found")
        
        if updates.assigned_tickets is not None:
            plan.assigned_tickets = updates.assigned_tickets
            plan.total_planned_hours = sum(t.get('estimated_hours', 0) for t in updates.assigned_tickets)
        if updates.notes is not None:
            plan.notes = updates.notes
        if updates.status is not None:
            plan.status = updates.status
        
        db.commit()
        db.refresh(plan)
        
        return {
            "success": True,
            "plan": {
                "id": plan.id,
                "employee_name": plan.employee_name,
                "week_start": plan.week_start.isoformat(),
                "assigned_tickets": plan.assigned_tickets,
                "total_planned_hours": plan.total_planned_hours,
                "status": plan.status
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


# ===== DEVELOPMENT TASK PLANNING API =====

def _planning_can_edit(role: str) -> bool:
    """Only Admin, Manager or Lead can create/edit plans."""
    if not role:
        return False
    return role == "ADMIN" or "MANAGER" in role or "LEAD" in role


def _get_user_display_name(db: Session, current_user: dict) -> str:
    """Get display name for current user (employee name or email)."""
    emp_id = current_user.get("employee_id")
    if emp_id:
        emp = db.query(Employee).filter(Employee.employee_id == emp_id).first()
        if emp and emp.name:
            return emp.name.strip()
    return current_user.get("email", "User") or "User"


@app.get("/dev-planning")
def dev_planning_health():
    """Health check for Development Task Planning module. Returns 200 if the module is loaded."""
    return {"status": "ok", "module": "dev-planning"}


@app.get("/dev-planning/overview")
def dev_planning_overview(current_user: dict = Depends(get_current_user)):
    """
    Get Dev Task Planning overview with categorized ticket data.
    
    Returns:
    - dev_tickets: Tickets currently with DEV team, categorized by:
      - by_priority: {URGENT: [...], High: [...], Medium: [...], Low: [...]}
      - to_be_assigned: Tickets without developer assigned
      - already_assigned: Tickets with developer but not In Progress
      - in_progress: Tickets with 'In Progress' status
      - ready_for_qc: Tickets ready to move to QC Testing (Code Review Passed, etc.)
    - qa_tickets: Tickets currently with QA team, categorized by:
      - pending: QC Testing (not started)
      - in_progress: QC Testing in Progress
      - bis_testing: Moved to BIS Testing
      - on_hold: QC Testing Hold
    """
    db: Session = SessionLocal()
    try:
        # Only include tickets that are still in PM Tracker (excludes stale/deleted tickets)
        all_tickets = db.query(TicketTracking).filter(
            TicketTracking.in_pm_tracker == True
        ).all()
        today = datetime.now().date()
        
        # DEV team statuses
        DEV_STATUSES = [
            'Ready For Development', 'Technical Review', 'Approved for Live',
            'Live - awaiting fixes', 'Express Lane Review', 'In Progress',
            'Start Code Review', 'Code Review Failed', 'QC Review Fail',
            'Code Review Passed', 'Tested - Awaiting Fixes', 'Re-opened', 'Reopened'
        ]
        
        # QA team statuses
        QA_STATUSES = ['QC Testing', 'QC Testing in Progress', 'QC Testing Hold']
        BIS_TESTING_STATUSES = ['BIS Testing', 'Testing In Progress']
        
        # Ready for QC (dev work done, ready to hand off)
        READY_FOR_QC_STATUSES = ['Code Review Passed', 'Approved for Live']
        
        # QC Review Failed (returned from QA with issues)
        QC_REVIEW_FAIL_STATUSES = ['QC Review Fail', 'Tested - Awaiting Fixes']
        
        # Containers for categorized tickets
        dev_by_priority = {'URGENT': [], 'High (Bugs)': [], 'High': [], 'Medium': [], 'Low': [], 'Unspecified': []}
        dev_to_be_assigned = []
        dev_already_assigned = []
        dev_in_progress = []
        dev_ready_for_qc = []
        dev_qc_review_failed = []
        dev_all = []
        
        qa_pending = []
        qa_in_progress = []
        qa_bis_testing = []
        qa_on_hold = []
        qa_all = []
        
        priority_counts_map = _priority_changes_count_map(db, [t.ticket_id for t in all_tickets])
        # Build assignee name -> employee_id map for links
        assignee_to_id = {}
        for emp in db.query(Employee.employee_id, Employee.name).filter(Employee.is_active == True).all():
            if emp.name and emp.name.strip():
                assignee_to_id[emp.name.strip().lower()] = emp.employee_id

        for ticket in all_tickets:
            status = ticket.status or 'Unknown'
            is_closed = status.lower() in ['closed', 'moved to live', 'completed']
            if is_closed:
                continue
            
            ticket_age, ageing_days, days_to_close = _ticket_ageing(ticket, today, False)
            priority = (ticket.priority or '').strip() or 'Unspecified'
            backend_dev = (ticket.backend_developer or '').strip()
            frontend_dev = (ticket.frontend_developer or '').strip()
            has_dev_assigned = bool(backend_dev or frontend_dev)
            assignee = ticket.current_assignee or 'Unassigned'
            assignee_employee_id = assignee_to_id.get((assignee or "").strip().lower()) if assignee and assignee != "Unassigned" else None

            ticket_data = {
                "ticket_id": ticket.ticket_id,
                "title": (getattr(ticket, 'title', None) or '').strip() or f"Ticket #{ticket.ticket_id}",
                "status": status,
                "priority": priority,
                "priority_changes_count": priority_counts_map.get(ticket.ticket_id, 0),
                "assignee": assignee,
                "assignee_employee_id": assignee_employee_id,
                "backend_developer": backend_dev or None,
                "frontend_developer": frontend_dev or None,
                "qc_tester": (ticket.qc_tester or '').strip() or None,
                "eta": ticket.eta.isoformat() if ticket.eta else None,
                "age_days": ticket_age,
                "ageing_days": ageing_days,
                "dev_estimate": ticket.dev_estimate_hours,
                "dev_actual": ticket.actual_dev_hours,
                "qa_estimate": ticket.qa_estimate_hours,
                "qa_actual": ticket.actual_qa_hours,
                "created_on": ticket.created_on.isoformat() if getattr(ticket, 'created_on', None) else None,
                "updated_on": ticket.updated_on.isoformat() if ticket.updated_on else None,
            }
            
            # Categorize DEV tickets
            if status in DEV_STATUSES:
                dev_all.append(ticket_data)
                
                # By priority
                priority_key = priority if priority in dev_by_priority else 'Unspecified'
                dev_by_priority[priority_key].append(ticket_data)
                
                # Assignment status
                if status in QC_REVIEW_FAIL_STATUSES:
                    dev_qc_review_failed.append(ticket_data)
                elif not has_dev_assigned:
                    dev_to_be_assigned.append(ticket_data)
                elif status == 'In Progress':
                    dev_in_progress.append(ticket_data)
                elif status in READY_FOR_QC_STATUSES:
                    dev_ready_for_qc.append(ticket_data)
                else:
                    dev_already_assigned.append(ticket_data)
            
            # Categorize QA tickets
            if status in QA_STATUSES:
                qa_all.append(ticket_data)
                if status == 'QC Testing':
                    qa_pending.append(ticket_data)
                elif status == 'QC Testing in Progress':
                    qa_in_progress.append(ticket_data)
                elif status == 'QC Testing Hold':
                    qa_on_hold.append(ticket_data)
            
            if status in BIS_TESTING_STATUSES:
                qa_all.append(ticket_data)
                qa_bis_testing.append(ticket_data)
        
        # By assignee (all active dev + qa tickets; tickets are in one list only)
        by_assignee = defaultdict(list)
        for t in dev_all:
            by_assignee[t["assignee"]].append(t)
        for t in qa_all:
            by_assignee[t["assignee"]].append(t)
        
        # Sort by priority (URGENT first) then by age
        priority_order = {'URGENT': 0, 'High (Bugs)': 1, 'High': 2, 'Medium': 3, 'Low': 4, 'Unspecified': 5}
        sort_key = lambda t: (priority_order.get(t['priority'], 5), -(t['age_days'] or 0))
        
        for lst in [dev_to_be_assigned, dev_already_assigned, dev_in_progress, dev_ready_for_qc,
                    dev_qc_review_failed, qa_pending, qa_in_progress, qa_bis_testing, qa_on_hold]:
            lst.sort(key=sort_key)
        
        return {
            "by_assignee": {k: {"count": len(v), "tickets": v} for k, v in by_assignee.items()},
            "dev_tickets": {
                "total": len(dev_all),
                "by_priority": {k: {"count": len(v), "tickets": v} for k, v in dev_by_priority.items()},
                "to_be_assigned": {"count": len(dev_to_be_assigned), "tickets": dev_to_be_assigned},
                "already_assigned": {"count": len(dev_already_assigned), "tickets": dev_already_assigned},
                "in_progress": {"count": len(dev_in_progress), "tickets": dev_in_progress},
                "ready_for_qc": {"count": len(dev_ready_for_qc), "tickets": dev_ready_for_qc},
                "qc_review_failed": {"count": len(dev_qc_review_failed), "tickets": dev_qc_review_failed},
            },
            "qa_tickets": {
                "total": len(qa_all),
                "pending": {"count": len(qa_pending), "tickets": qa_pending},
                "in_progress": {"count": len(qa_in_progress), "tickets": qa_in_progress},
                "bis_testing": {"count": len(qa_bis_testing), "tickets": qa_bis_testing},
                "on_hold": {"count": len(qa_on_hold), "tickets": qa_on_hold},
            },
            "summary": {
                "dev_total": len(dev_all),
                "qa_total": len(qa_all),
                "dev_unassigned": len(dev_to_be_assigned),
                "dev_in_progress": len(dev_in_progress),
                "dev_qc_review_failed": len(dev_qc_review_failed),
                "qa_pending": len(qa_pending),
                "qa_in_progress": len(qa_in_progress),
            }
        }
    finally:
        db.close()


# ETA calendar: completed-status keywords (match frontend ETACalendar)
ETA_CALENDAR_COMPLETED_KEYWORDS = ['complete', 'completed', 'closed', 'done', 'resolved', 'moved to live']


def _is_completed_status_for_eta(status: Optional[str]) -> bool:
    if not status:
        return False
    n = str(status).strip().lower()
    return any(k in n for k in ETA_CALENDAR_COMPLETED_KEYWORDS)


@app.get("/eta-calendar/tickets")
def eta_calendar_tickets(current_user: dict = Depends(get_current_user)):
    """
    All tickets that have an ETA (any status). For ETA calendar: show every ticket with ETA.
    Returns completed_within_eta (true if closed on or before ETA) and eta_rescheduled (true if ETA was changed).
    """
    db: Session = SessionLocal()
    try:
        tickets = (
            db.query(TicketTracking)
            .filter(TicketTracking.eta.isnot(None))
            .all()
        )
        ticket_ids = [t.ticket_id for t in tickets]
        open_bugs_map = {}
        if ticket_ids:
            open_statuses = ["New", "Reopened", "Fixed", "Assigned to Dev"]
            rows = (
                db.query(Bug.ticket_id, func.count(Bug.id))
                .filter(Bug.ticket_id.in_(ticket_ids), Bug.status.in_(open_statuses))
                .group_by(Bug.ticket_id)
                .all()
            )
            open_bugs_map = {tid: c for tid, c in rows}
        result = []
        for t in tickets:
            eta_dt = t.eta
            closed_dt = getattr(t, 'closed_on', None)
            status = (t.status or '').strip()
            is_completed = _is_completed_status_for_eta(status)
            completed_within_eta = False
            if is_completed and closed_dt and eta_dt:
                closed_d = closed_dt.date() if hasattr(closed_dt, 'date') else closed_dt
                eta_d = eta_dt.date() if hasattr(eta_dt, 'date') else eta_dt
                completed_within_eta = closed_d <= eta_d
            developers = []
            if t.backend_developer:
                developers.append(t.backend_developer)
            if t.frontend_developer:
                developers.append(t.frontend_developer)
            developers = list(set(developers))
            result.append({
                'ticket_id': t.ticket_id,
                'title': (getattr(t, 'title', None) or '').strip() or f"Ticket #{t.ticket_id}",
                'status': status,
                'priority': (t.priority or '').strip() or 'Unspecified',
                'eta': eta_dt.isoformat() if eta_dt else None,
                'closed_on': closed_dt.isoformat() if closed_dt else None,
                'qc_tester': (t.qc_tester or '').strip() or None,
                'developers_str': ', '.join(developers) if developers else 'Not Assigned',
                'developers': developers,
                'completed_within_eta': completed_within_eta,
                'times_moved_to_fail': get_qc_fail_count(db, t.ticket_id),
                'open_bugs_count': open_bugs_map.get(t.ticket_id, 0),
            })
        return {'tickets': result}
    finally:
        db.close()


# ===== LIVE PM DATA ENDPOINTS (new modules) =====
from pm_live_data import (
    get_live_qc_queue, get_live_team_board, get_live_activity_summary, get_live_bis_tracking,
    fetch_live_tickets,
    load_module_ownership, save_module_ownership, auto_detect_modules_and_members,
    get_live_resource_occupancy, get_live_assignment_suggestions, get_live_module_ownership_matrix,
    get_live_team_queue, get_live_automation_utilization, get_live_build_quality, get_live_dev_dashboard,
)


@app.post("/live/refresh")
def force_refresh_pm_data():
    """Force refresh PM data from API, clearing all caches."""
    from pm_live_data import clear_response_cache
    clear_response_cache()
    success, tickets, msg = fetch_live_tickets(force_refresh=True)
    return {'success': success, 'ticket_count': len(tickets) if tickets else 0, 'message': msg}


@app.get("/live/qc-queue")
def live_qc_queue():
    """Live QC queue fetched directly from PM API, enriched with planning status:
    a QC-Testing ticket with no qc_tester is 'in_planning' when its PM Assign-To (current_assignee)
    is a QA-team member (the owner planning it), else truly 'unassigned'."""
    data = get_live_qc_queue()

    # QA-team name lookup (normalized + compacted + paren-stripped) for the Assign-To check.
    qa_lookup = set()
    db: Session = SessionLocal()
    try:
        for emp in db.query(Employee).filter(
            Employee.is_active == True, Employee.archived == False,
        ).all():
            if "QA" in (emp.team or "").upper() and emp.name:
                qa_lookup.add(_normalize_person_name(emp.name))
                qa_lookup.add(_compact_person_name(emp.name))
    finally:
        db.close()

    def _qa_member(name):
        if not name:
            return None
        for v in (name.strip(), _strip_paren(name).strip()):
            if _normalize_person_name(v) in qa_lookup or _compact_person_name(v) in qa_lookup:
                return name.strip()
        return None

    def _enrich_planning(t):
        if (t.get("status") or "") not in ("QC Testing", "QC Testing in Progress", "QC Testing Hold"):
            return
        if t.get("qc_tester"):
            t["planning_status"] = "assigned"
            return
        planner = _qa_member(t.get("current_assignee"))
        if planner:
            t["planning_status"] = "in_planning"
            t["planner"] = planner
        else:
            t["planning_status"] = "unassigned"

    for key in ("queue", "qc_failed", "bis_testing", "approved_for_live", "no_qa_estimate"):
        section = data.get(key)
        tickets = section.get("tickets") if isinstance(section, dict) else section
        for t in (tickets or []):
            _enrich_planning(t)
    return data


@app.get("/live/team-board")
def live_team_board():
    """Live team board fetched directly from PM API."""
    return get_live_team_board()


@app.get("/live/team-board/activity-distribution")
def live_activity_distribution():
    """Live activity distribution."""
    board = get_live_team_board()
    dist = {}
    for m in board.get('members', []):
        a = m['activity']
        dist[a] = dist.get(a, 0) + 1
    plat = {}
    for m in board.get('members', []):
        if m['activity'] != 'idle':
            p = m.get('platform', 'Web')
            plat[p] = plat.get(p, 0) + 1
    return {'activity_distribution': dist, 'platform_distribution': plat, 'summary': board.get('summary', {})}


@app.get("/live/team-board/member/{name}")
def live_team_board_member(name: str):
    """Live member detail."""
    board = get_live_team_board()
    for m in board.get('members', []):
        if m['name'].lower() == name.lower() or m['employee_id'] == name:
            return m
    raise HTTPException(status_code=404, detail="Member not found")


@app.get("/live/qa-activity-summary")
def live_activity_summary(
    period: str = Query('past_5_days', regex='^(past_5_days|current_month|custom)$'),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
):
    """Live QA activity summary from PM API."""
    sd = None
    ed = None
    if period == 'custom' and start_date and end_date:
        sd = date.fromisoformat(start_date)
        ed = date.fromisoformat(end_date)
    return get_live_activity_summary(period=period, start_override=sd, end_override=ed)


@app.get("/live/bis-to-closed")
def live_bis_to_closed():
    """Live BIS tracking from PM API."""
    return get_live_bis_tracking()


@app.get("/live/qc-review-fail")
def live_qc_review_fail():
    """Live QC review fail tickets from PM API."""
    data = get_live_qc_queue()
    return data.get('qc_failed', {'tickets': [], 'total': 0})


@app.get("/live/reports/weekly")
def generate_weekly_report(start_date: Optional[str] = Query(None), end_date: Optional[str] = Query(None)):
    """Generate and download QA Report Excel."""
    import subprocess
    script = os.path.join(os.path.dirname(__file__), 'generate_qa_reports.py')
    cmd = ['python', script]
    if start_date and end_date:
        cmd += [f'--start={start_date}', f'--end={end_date}']
    subprocess.run(cmd, cwd=os.path.dirname(__file__), timeout=30)
    from datetime import date as d
    report_date = end_date if end_date else d.today().strftime("%Y-%m-%d")
    path = os.path.join(os.path.dirname(__file__), 'reports', f'QA_Report_{d.fromisoformat(report_date).strftime("%Y%m%d")}.xlsx')
    if os.path.exists(path):
        return FileResponse(path, filename=os.path.basename(path), media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    raise HTTPException(status_code=500, detail="Report generation failed")


@app.get("/live/reports/monthly")
def generate_monthly_report(start_date: Optional[str] = Query(None), end_date: Optional[str] = Query(None)):
    """Generate and download QA Monthly Report Excel."""
    import subprocess
    script = os.path.join(os.path.dirname(__file__), 'generate_qa_reports.py')
    cmd = ['python', script]
    if start_date and end_date:
        cmd += [f'--start={start_date}', f'--end={end_date}']
    subprocess.run(cmd, cwd=os.path.dirname(__file__), timeout=30)
    from datetime import date as d
    report_date = end_date if end_date else d.today().strftime("%Y-%m-%d")
    path = os.path.join(os.path.dirname(__file__), 'reports', f'QA_Report_{d.fromisoformat(report_date).strftime("%Y%m%d")}.xlsx')
    if os.path.exists(path):
        return FileResponse(path, filename=os.path.basename(path), media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    raise HTTPException(status_code=500, detail="Report generation failed")


@app.get("/live/reports/automation-weekly")
def generate_automation_report(start_date: Optional[str] = Query(None), end_date: Optional[str] = Query(None)):
    """Generate and download Automation Utilization Report PDF."""
    import subprocess
    script = os.path.join(os.path.dirname(__file__), 'generate_automation_pdf_report.py')
    cmd = ['python', script]
    if start_date and end_date:
        cmd += [f'--start={start_date}', f'--end={end_date}']
    subprocess.run(cmd, cwd=os.path.dirname(__file__), timeout=120)
    from datetime import date as d
    report_date = end_date if end_date else d.today().strftime("%Y-%m-%d")
    path = os.path.join(os.path.dirname(__file__), 'reports', f'Automation_Utilization_Report_{d.fromisoformat(report_date).strftime("%Y%m%d")}.pdf')
    if os.path.exists(path):
        return FileResponse(path, filename=os.path.basename(path), media_type='application/pdf')
    raise HTTPException(status_code=500, detail="Report generation failed")


@app.get("/live/reports/module-ownership")
def generate_module_report():
    """Download Module Ownership Report Excel."""
    path = os.path.join(os.path.dirname(__file__), 'reports', 'Module_Ownership_Report.xlsx')
    if os.path.exists(path):
        return FileResponse(path, filename=os.path.basename(path), media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    raise HTTPException(status_code=404, detail="Report not found. Generate it first.")


@app.get("/live/reports/dev-weekly")
def generate_dev_weekly_report(start_date: Optional[str] = Query(None), end_date: Optional[str] = Query(None)):
    """Generate and download Dev Report Excel."""
    import subprocess
    script = os.path.join(os.path.dirname(__file__), 'generate_dev_report.py')
    cmd = ['python', script]
    if start_date and end_date:
        cmd += [f'--start={start_date}', f'--end={end_date}']
    subprocess.run(cmd, cwd=os.path.dirname(__file__), timeout=30)
    from datetime import date as d
    path = os.path.join(os.path.dirname(__file__), 'reports', f'Dev_Report_{d.today().strftime("%Y%m%d")}.xlsx')
    if os.path.exists(path):
        return FileResponse(path, filename=os.path.basename(path), media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    raise HTTPException(status_code=500, detail="Report generation failed")


@app.get("/live/reports/dev-monthly")
def generate_dev_monthly_report(start_date: Optional[str] = Query(None), end_date: Optional[str] = Query(None)):
    """Generate and download Dev Report Excel."""
    import subprocess
    script = os.path.join(os.path.dirname(__file__), 'generate_dev_report.py')
    cmd = ['python', script]
    if start_date and end_date:
        cmd += [f'--start={start_date}', f'--end={end_date}']
    subprocess.run(cmd, cwd=os.path.dirname(__file__), timeout=30)
    from datetime import date as d
    path = os.path.join(os.path.dirname(__file__), 'reports', f'Dev_Report_{d.today().strftime("%Y%m%d")}.xlsx')
    if os.path.exists(path):
        return FileResponse(path, filename=os.path.basename(path), media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    raise HTTPException(status_code=500, detail="Report generation failed")


# ===== MODULE OWNERSHIP & RESOURCE PLANNING =====

@app.get("/live/module-ownership")
def get_module_ownership_config():
    return load_module_ownership()


@app.put("/live/module-ownership")
def update_module_ownership_config(body: dict = Body(...)):
    data = load_module_ownership()
    if 'modules' in body:
        data['modules'] = body['modules']
    if 'team_members' in body:
        data['team_members'] = body['team_members']
    save_module_ownership(data)
    return {'success': True, 'data': data}


@app.post("/live/module-ownership/auto-detect")
def auto_detect_ownership():
    return auto_detect_modules_and_members()


@app.post("/live/export-tickets")
def export_tickets_to_excel(body: dict = Body(...)):
    """Generate formatted Excel from ticket list with hyperlinks."""
    import tempfile
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    tickets = body.get('tickets', [])
    filename = body.get('filename', 'Tickets')

    wb = Workbook()
    ws = wb.active
    ws.title = filename[:31]  # Excel tab name limit

    # Styles
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
    link_font = Font(color='0563C1', underline='single', size=9)
    data_font = Font(size=9)
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC')
    )
    alt_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')

    # Status colors
    status_fills = {
        'QC Testing': PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid'),
        'QC Testing in Progress': PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid'),
        'QC Testing Hold': PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid'),
        'QC Review Fail': PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid'),
        'BIS Testing': PatternFill(start_color='EDE7F6', end_color='EDE7F6', fill_type='solid'),
        'Approved for Live': PatternFill(start_color='E0F7FA', end_color='E0F7FA', fill_type='solid'),
        'In Progress': PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid'),
        'Code Review Passed': PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid'),
        'Start Code Review': PatternFill(start_color='BBDEFB', end_color='BBDEFB', fill_type='solid'),
    }

    headers = ['Ticket', 'Title', 'Status', 'Priority', 'Platform', 'Module', 'QC Tester', 'Developer', 'Est Hrs', 'Actual Hrs', 'ETA']
    col_widths = [12, 50, 22, 18, 10, 22, 20, 20, 10, 10, 14]

    # Title row
    ws.merge_cells('A1:K1')
    title_cell = ws['A1']
    title_cell.value = f'{filename} — {len(tickets)} tickets'
    title_cell.font = Font(bold=True, size=13, color='1a1a2e')
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 28

    # Headers
    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[chr(64 + c) if c <= 26 else 'A'].width = w

    PM_URL = 'https://www.bissafety.app/pm/tickets#!/'

    # Data rows
    for i, t in enumerate(tickets, 4):
        tid = t.get('ticket_id', '')
        row_fill = alt_fill if (i % 2 == 0) else None

        # Ticket ID with hyperlink
        cell = ws.cell(row=i, column=1, value=f'#{tid}')
        cell.hyperlink = f'{PM_URL}{tid}'
        cell.font = link_font
        cell.alignment = center
        cell.border = border
        if row_fill: cell.fill = row_fill

        vals = [
            t.get('title', ''), t.get('status', ''), t.get('priority', ''),
            t.get('platform', ''), t.get('module', ''),
            t.get('qc_tester', ''), t.get('developers_str', '') or t.get('developer', ''),
            t.get('qa_estimate_hours') or t.get('dev_estimate_hours') or '',
            t.get('qa_actual_hours') or t.get('actual_dev_hours') or '',
            t.get('eta', ''),
        ]
        for c, v in enumerate(vals, 2):
            cell = ws.cell(row=i, column=c, value=v if v else '')
            cell.font = data_font
            cell.alignment = left if c == 2 else center
            cell.border = border
            if row_fill: cell.fill = row_fill
            # Color status cell
            if c == 3 and v in status_fills:
                cell.fill = status_fills[v]

    # Freeze header
    ws.freeze_panes = 'A4'
    # Auto-filter
    ws.auto_filter.ref = f'A3:K{3 + len(tickets)}'

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(tmp.name)
    tmp.close()

    safe_filename = filename.replace(' ', '_').replace('/', '-')
    return FileResponse(tmp.name, filename=f'{safe_filename}.xlsx',
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.get("/live/assign-to-summary")
def live_assign_to_summary():
    """All tickets grouped by Assign To field with dev/qa/bis status breakdown."""
    from pm_live_data import fetch_live_tickets, load_module_ownership
    success, all_tickets, _ = fetch_live_tickets()
    if not success:
        return {'persons': [], 'total': 0}

    ownership = load_module_ownership()
    qa_team = set(ownership.get('team_members', []))
    # Dev team from Employee table
    try:
        from models import Employee
        db = SessionLocal()
        dev_employees = db.query(Employee).filter(Employee.is_active == True, Employee.team == 'DEVELOPMENT').all()
        dev_team = set(e.name for e in dev_employees)
        db.close()
    except Exception:
        dev_team = set()

    DEV_STATUSES = {'Ready For Development', 'In Progress', 'Hold/Pending', 'Start Code Review',
        'Code Review Failed', 'Code Review Passed', 'Express Lane Review'}
    QA_STATUSES = {'QC Testing', 'QC Testing in Progress', 'QC Testing Hold', 'QC Review Fail'}
    BIS_STATUSES = {'BIS Testing', 'Approved for Live', 'Moved to Live'}

    CLOSED_STATUSES = {'Closed', 'Moved to Live', 'Cancelled', 'Rejected', 'Duplicate'}
    persons = {}
    for t in all_tickets:
        if t['status'] in CLOSED_STATUSES:
            continue
        assignee = (t.get('current_assignee') or '').strip()
        if not assignee:
            continue
        if assignee not in persons:
            # Determine team
            team = 'Dev' if assignee in dev_team or any(assignee.lower() in d.lower() for d in dev_team) else \
                   'QA' if assignee in qa_team or any(assignee.lower() in q.lower() for q in qa_team) else 'BIS'
            persons[assignee] = {'name': assignee, 'team': team, 'dev': 0, 'qa': 0, 'bis': 0, 'other': 0, 'total': 0, 'tickets': []}

        s = t['status']
        persons[assignee]['total'] += 1
        if s in DEV_STATUSES:
            persons[assignee]['dev'] += 1
        elif s in QA_STATUSES:
            persons[assignee]['qa'] += 1
        elif s in BIS_STATUSES:
            persons[assignee]['bis'] += 1
        else:
            persons[assignee]['other'] += 1

        persons[assignee]['tickets'].append({
            'ticket_id': t['ticket_id'], 'title': t['title'], 'status': s,
            'priority': t['priority'], 'module': t.get('module', ''),
            'platform': t.get('platform', ''), 'qc_tester': t.get('qc_tester', ''),
            'developers_str': t.get('developers_str', ''),
            'qa_estimate_hours': t.get('qa_estimate_hours', 0),
            'dev_estimate_hours': t.get('dev_estimate_hours', 0),
        })

    person_list = sorted(persons.values(), key=lambda p: -p['total'])
    return {'persons': person_list, 'total': len(person_list)}


@app.get("/live/ticket-calendar")
def live_ticket_calendar():
    """Monthly calendar showing ticket movement per day per status."""
    from pm_live_data import _load_ageing_tracker, fetch_live_tickets
    from collections import defaultdict

    ageing = _load_ageing_tracker()
    success, all_tickets, _ = fetch_live_tickets()
    ticket_map = {str(t['ticket_id']): t for t in all_tickets} if success else {}

    # Group by first_seen date and status
    daily = defaultdict(lambda: defaultdict(list))
    for tid, entry in ageing.items():
        fs = entry.get('first_seen', '')
        status = entry.get('status', '')
        if fs and status:
            t = ticket_map.get(tid, {})
            daily[fs][status].append({
                'ticket_id': int(tid) if tid.isdigit() else tid,
                'title': t.get('title', ''),
                'status': status,
                'priority': t.get('priority', ''),
                'module': t.get('module', ''),
                'developers_str': t.get('developers_str', ''),
                'qc_tester': t.get('qc_tester', ''),
            })

    # Convert to list format
    calendar = []
    for day in sorted(daily.keys()):
        statuses = {}
        total = 0
        for status, tickets in daily[day].items():
            statuses[status] = {'count': len(tickets), 'tickets': tickets}
            total += len(tickets)
        calendar.append({'date': day, 'total': total, 'statuses': statuses})

    return {'calendar': calendar}


@app.get("/live/build-quality")
def live_build_quality():
    return get_live_build_quality()


@app.get("/live/automation-utilization")
def live_automation_utilization():
    return get_live_automation_utilization()


@app.get("/live/automation-team")
def live_automation_team():
    from automation_team_tracker import get_team_stats
    return get_team_stats()


@app.get("/live/automation-team/{person}")
def live_automation_member_weekly(person: str):
    from automation_team_tracker import get_member_weekly_activity
    return get_member_weekly_activity(person)


@app.post("/live/automation-team/log")
def automation_team_manual_entry(body: dict = Body(...)):
    from automation_team_tracker import add_manual_entry
    return add_manual_entry(
        person=body.get('person', ''),
        entry_date=body.get('date', ''),
        module=body.get('module', ''),
        cases_scripted=body.get('cases_scripted', 0),
        cases_executed=body.get('cases_executed', 0),
        activity=body.get('activity', 'scripting'),
        notes=body.get('notes', ''),
    )


@app.post("/live/automation-team/import-excel")
def automation_team_import():
    from automation_team_tracker import import_from_excel
    excel_path = os.environ.get('AUTOMATION_DASHBOARD_EXCEL', r'D:\Vishnu VS\bis-automation\automation_dashboard.xlsx')
    return import_from_excel(excel_path, person='Vishnu VS')


@app.post("/live/automation-team/sync-git")
def automation_team_sync_git():
    from automation_team_tracker import sync_git_activity
    return sync_git_activity()


@app.get("/live/team-queue")
def live_team_queue_endpoint():
    return get_live_team_queue()


@app.get("/live/resource-occupancy")
def live_resource_occupancy_endpoint():
    return get_live_resource_occupancy()


@app.get("/live/assignment-suggestions")
def live_assignment_suggestions_endpoint():
    return get_live_assignment_suggestions()


@app.get("/live/module-ownership-matrix")
def live_module_matrix():
    return get_live_module_ownership_matrix()


@app.get("/live/dev-dashboard")
def live_dev_dashboard():
    """Dev team pipeline insights: resource-wise, ticket-wise, module-wise."""
    return get_live_dev_dashboard()


@app.get("/live/module-tickets/{module_name}")
def live_module_tickets(module_name: str, status_group: str = Query('all'), platform: Optional[str] = Query(None)):
    """Get tickets for a specific module, optionally filtered by status group and platform."""
    success, all_tickets, _ = fetch_live_tickets()
    if not success:
        return {'tickets': [], 'count': 0}

    DEV_NEAR_QC = {'Code Review Passed'}
    DEV_CODE_REVIEW = {'Start Code Review', 'Code Review Failed'}
    DEV_IN_PROGRESS = {'In Progress', 'Hold/Pending'}
    DEV_EARLY = {'Planning', 'Ready For Development', 'NEW', 'DRAFT', 'Ready for Design',
                 'Technical Review', 'Design Review', 'Design In Progress'}
    QC_STATUSES_SET = {'QC Testing', 'QC Testing in Progress', 'QC Testing Hold'}

    filtered = [t for t in all_tickets if t.get('module') == module_name]
    if platform:
        filtered = [t for t in filtered if t.get('platform') == platform]

    QC_SET = {'QC Testing', 'QC Testing in Progress', 'QC Testing Hold'}
    DEV_CR_PASSED = {'Code Review Passed'}
    DEV_CR = {'Start Code Review', 'Code Review Failed'}
    DEV_WIP = {'In Progress', 'Hold/Pending'}
    DEV_EARLY_SET = {'Planning', 'Ready For Development', 'NEW', 'DRAFT', 'Ready for Design',
                     'Technical Review', 'Design Review', 'Design In Progress'}
    DEV_ALL = DEV_CR_PASSED | DEV_CR | DEV_WIP | DEV_EARLY_SET

    if status_group == 'qc_active':
        filtered = [t for t in filtered if t['status'] in QC_SET]
    elif status_group == 'qc_testing':
        filtered = [t for t in filtered if t['status'] == 'QC Testing']
    elif status_group == 'qc_hold':
        filtered = [t for t in filtered if t['status'] == 'QC Testing Hold']
    elif status_group == 'in_progress':
        filtered = [t for t in filtered if t['status'] == 'QC Testing in Progress']
    elif status_group == 'qc_failed':
        filtered = [t for t in filtered if t['status'] == 'QC Review Fail']
    elif status_group == 'bis':
        filtered = [t for t in filtered if t['status'] == 'BIS Testing']
    elif status_group == 'approved':
        filtered = [t for t in filtered if t['status'] == 'Approved for Live']
    elif status_group == 'dev_pipeline':
        filtered = [t for t in filtered if t['status'] in DEV_ALL]
    elif status_group == 'cr_passed':
        filtered = [t for t in filtered if t['status'] in DEV_CR_PASSED]
    elif status_group == 'dev_in_progress':
        filtered = [t for t in filtered if t['status'] in DEV_WIP and not t.get('qc_tester')]
    elif status_group == 'dev_refix':
        filtered = [t for t in filtered if t['status'] in DEV_ALL and t.get('qc_tester')]

    # Load ageing tracker for days-in-status
    from pm_live_data import _load_ageing_tracker, load_module_ownership, _parse_date
    ageing = _load_ageing_tracker()
    today_date = date.today()

    # For unassigned tickets: suggest assignee
    ownership = load_module_ownership()
    mod_config = ownership.get('modules', {}).get(module_name, {})
    primary_owners = mod_config.get('primary_owners', [])
    support_owners = mod_config.get('support_owners', [])

    result = []
    for t in filtered:
        tid = str(t['ticket_id'])
        age_entry = ageing.get(tid, {})
        days_in_qc = 0
        if age_entry.get('status') == t['status']:
            first_seen = _parse_date(age_entry.get('first_seen'))
            if first_seen:
                days_in_qc = max(0, (today_date - first_seen).days)

        item = {
            'ticket_id': t['ticket_id'], 'title': t['title'], 'status': t['status'],
            'priority': t['priority'], 'qc_tester': t.get('qc_tester') or '-',
            'developers_str': t.get('developers_str', '-'),
            'qa_estimate_hours': t.get('qa_estimate_hours', 0),
            'qa_actual_hours': t.get('qa_actual_hours', 0),
            'eta': t.get('eta'), 'platform': t.get('platform', 'Web'),
            'days_in_qc': days_in_qc,
        }

        # Suggest assignee for unassigned tickets
        if not t.get('qc_tester') and t['status'] in QC_SET:
            suggested = primary_owners[0] if primary_owners else (support_owners[0] if support_owners else '')
            item['suggested_assignee'] = suggested

        result.append(item)

    return {'tickets': result, 'count': len(result), 'module': module_name, 'status_group': status_group}


# ===== QC QUEUE & AGEING ANALYTICS (old - database backed) =====

@app.get("/qc-queue")
def get_qc_queue():
    """
    Smart-prioritized QC testing queue. Each ticket scored 0-100 based on
    priority, ageing in QC, re-entry after fail, ETA urgency, and ticket type.
    Higher score = should be picked up first. Tickets tested by dev are separated.
    """
    db: Session = SessionLocal()
    try:
        data = get_qa_overview_data(db)
        queue = data.get('queue', [])

        scored_queue = []
        dev_tested = []
        for t in queue:
            scoring = calculate_qc_priority_score(t)
            t['priority_score'] = scoring['score']
            t['score_breakdown'] = scoring['breakdown']
            if t.get('tested_by_dev'):
                dev_tested.append(t)
            else:
                scored_queue.append(t)

        # Sort by score descending
        scored_queue.sort(key=lambda t: (-t['priority_score'], t['ticket_id']))
        dev_tested.sort(key=lambda t: (-t['priority_score'], t['ticket_id']))

        return {
            'queue': scored_queue,
            'dev_tested': dev_tested,
            'total': len(scored_queue),
            'dev_tested_count': len(dev_tested),
            'status_cards': data.get('status_cards', {}),
        }
    finally:
        db.close()


@app.get("/qc-queue/scoring/{ticket_id}")
def get_qc_queue_scoring(ticket_id: int):
    """Score breakdown for a specific ticket (tooltip data)."""
    db: Session = SessionLocal()
    try:
        from qa_planning import get_moved_to_qc_date, is_retesting_after_failure, get_retest_cycle_count, get_module_for_ticket, get_platform_for_ticket
        ticket = db.query(TicketTracking).filter(TicketTracking.ticket_id == ticket_id).first()
        if not ticket:
            raise HTTPException(status_code=404, detail="Ticket not found")

        today = date.today()
        moved_qc = get_moved_to_qc_date(db, ticket_id)
        moved_qc_date = moved_qc.date() if moved_qc and hasattr(moved_qc, 'date') else None
        days_in_qc = (today - moved_qc_date).days if moved_qc_date else 0

        t = {
            'ticket_id': ticket_id,
            'priority': (ticket.priority or '').strip(),
            'days_in_qc': days_in_qc,
            'retest_cycle_count': get_retest_cycle_count(db, ticket_id),
            'eta': ticket.eta.isoformat() if ticket.eta else None,
        }
        return calculate_qc_priority_score(t, today)
    finally:
        db.close()


@app.get("/tickets/{ticket_id}/status-durations")
def get_ticket_status_durations(ticket_id: int):
    """Time spent in each status (business days) for a ticket."""
    db: Session = SessionLocal()
    try:
        return get_status_durations(db, ticket_id)
    finally:
        db.close()


@app.get("/tickets/{ticket_id}/qc-cycles")
def get_ticket_qc_cycles(ticket_id: int):
    """Cycle-by-cycle QC breakdown: how many times ticket went through QC testing."""
    db: Session = SessionLocal()
    try:
        return get_qc_cycle_details(db, ticket_id)
    finally:
        db.close()


# Pipeline phase grouping for movement/speed analysis.
TICKET_PHASES = [
    ("Dev", {"Ready For Development", "In Progress", "Hold/Pending", "NEW", "DRAFT"}),
    ("Code Review", {"Start Code Review", "Code Review Failed", "Code Review Passed", "Express Lane Review"}),
    ("QC", {"QC Testing", "QC Testing in Progress", "QC Testing Hold", "QC Review Fail",
            "Tested - Awaiting Fixes", "Testing In Progress"}),
    ("BIS", {"BIS Testing"}),
    ("Approved", {"Approved for Live"}),
    ("Closed", {"Moved to Live", "Closed"}),
]
_QC_ENTRY_STATUSES = ("QC Testing", "QC Testing in Progress", "QC Testing Hold", "Testing In Progress")


def _phase_of(status):
    for name, sset in TICKET_PHASES:
        if status in sset:
            return name
    return "Other"


def _build_ticket_journey(history, ticket, today):
    """Ordered status legs for a ticket: each {status, phase, entered/exited, days, hours, current}."""
    legs = []

    def _leg(status, entered, exited, assignee, tester, current=False):
        if not entered:
            return
        end = exited or datetime.combine(today, datetime.min.time())
        ed = entered.date() if isinstance(entered, datetime) else entered
        xd = end.date() if isinstance(end, datetime) else end
        days = max(0, (xd - ed).days)
        hours = None
        if isinstance(entered, datetime) and isinstance(end, datetime) and end > entered:
            hours = round((end - entered).total_seconds() / 3600, 1)
        legs.append({
            "status": status, "phase": _phase_of(status),
            "entered_on": entered.isoformat() if hasattr(entered, "isoformat") else None,
            "exited_on": exited.isoformat() if exited and hasattr(exited, "isoformat") else None,
            "days": days, "hours": hours, "assignee": assignee, "qc_tester": tester,
            "is_current": current,
        })

    if history:
        for i, h in enumerate(history):
            status = h.previous_status or "Created"
            entered = history[i - 1].changed_on if i > 0 else ticket.created_on
            _leg(status, entered, h.changed_on, h.current_assignee, h.qc_tester)
        last = history[-1]
        _leg(last.new_status, last.changed_on, None, last.current_assignee, last.qc_tester,
             current=(ticket.closed_on is None))
    elif ticket.created_on:
        _leg(ticket.status, ticket.created_on, ticket.closed_on, ticket.current_assignee,
             ticket.qc_tester, current=(ticket.closed_on is None))
    return legs


def _ticket_speed_summary(db, ticket, history, today):
    """Speed metrics for one ticket from its (pre-fetched) status history."""
    durations = get_status_durations(db, ticket.ticket_id, today, history=history, created_on=ticket.created_on)
    cycles = get_qc_cycle_details(db, ticket.ticket_id, today, history=history)
    created = ticket.created_on
    end_dt = ticket.closed_on or datetime.combine(today, datetime.min.time())
    lead_time_days = max(0, (end_dt.date() - created.date()).days) if created else None
    first_qc = next((h.changed_on for h in history if h.new_status in _QC_ENTRY_STATUSES), None)
    first_bis = next((h.changed_on for h in history if h.new_status == "BIS Testing"), None)
    dev_to_qc_days = max(0, (first_qc.date() - created.date()).days) if (first_qc and created) else None
    qc_to_bis_days = max(0, (first_bis.date() - first_qc.date()).days) if (first_qc and first_bis) else None
    last_change = history[-1].changed_on if history else ticket.created_on
    current_stage_age = max(0, (today - last_change.date()).days) if last_change else 0
    phase_days = {}
    for st, d in (durations.get("durations") or {}).items():
        phase_days[_phase_of(st)] = phase_days.get(_phase_of(st), 0) + d
    return {
        "ticket_id": ticket.ticket_id, "title": ticket.title or "",
        "module": (ticket.subdepartment or "").strip() or "Unassigned",
        "qc_tester": ticket.qc_tester or "Unassigned",
        "developer": _strip_paren(ticket.backend_developer or ticket.frontend_developer or "").strip() or "Unassigned",
        "priority": ticket.priority or "",
        "current_status": durations.get("current_status") or ticket.status,
        "created_on": created.isoformat() if created else None,
        "closed_on": ticket.closed_on.isoformat() if ticket.closed_on else None,
        "lead_time_days": lead_time_days,
        "qc_days": durations.get("total_qc_days", 0),
        "hold_days": durations.get("total_hold_days", 0),
        "transitions": durations.get("transitions", 0),
        "cycles": cycles.get("total_cycles", 0),
        "first_pass": cycles.get("first_pass", False),
        "dev_to_qc_days": dev_to_qc_days, "qc_to_bis_days": qc_to_bis_days,
        "current_stage_age": current_stage_age,
        "phase_days": phase_days,
    }, durations, cycles


@app.get("/tickets/{ticket_id}/movement")
def get_ticket_movement_detail(ticket_id: int):
    """Full movement & speed for one ticket: status journey, per-status durations, QC cycles,
    phase times, and summary speed metrics."""
    db: Session = SessionLocal()
    try:
        ticket = db.query(TicketTracking).filter(TicketTracking.ticket_id == ticket_id).first()
        if not ticket:
            raise HTTPException(status_code=404, detail="Ticket not found")
        today = date.today()
        history = (
            db.query(TicketStatusHistory)
            .filter(TicketStatusHistory.ticket_id == ticket_id)
            .order_by(TicketStatusHistory.changed_on.asc())
            .all()
        )
        summary, durations, cycles = _ticket_speed_summary(db, ticket, history, today)
        summary["eta"] = ticket.eta.isoformat() if ticket.eta else None
        summary["qa_estimate_hours"] = ticket.qa_estimate_hours
        summary["qa_actual_hours"] = ticket.actual_qa_hours
        return {
            "summary": summary,
            "journey": _build_ticket_journey(history, ticket, today),
            "durations": durations.get("durations", {}),
            "phase_days": summary["phase_days"],
            "cycles": cycles,
        }
    finally:
        db.close()


@app.get("/ticket-speed")
def get_ticket_speed(
    period: str = Query("month", regex="^(month|quarter|all)$"),
    offset: int = Query(0, ge=0, le=240),
    scope: str = Query("closed", regex="^(closed|active|all)$"),
):
    """Per-ticket movement speed across a scope of tickets, with summary and per-module / per-QC-tester
    breakdowns. scope=closed → tickets closed in the period; active → currently open; all → both."""
    db: Session = SessionLocal()
    try:
        today = date.today()
        start = end = None
        label = "All time"
        if period != "all":
            start, end, label = get_period_range(period, offset)

        tickets, seen = [], set()

        def _add(rows):
            for t in rows:
                if t.ticket_id not in seen:
                    seen.add(t.ticket_id)
                    tickets.append(t)

        if scope in ("closed", "all"):
            cq = db.query(TicketTracking).filter(TicketTracking.closed_on.isnot(None))
            if start:
                cq = cq.filter(TicketTracking.closed_on >= start, TicketTracking.closed_on <= end)
            _add(cq.all())
        if scope in ("active", "all"):
            _add(db.query(TicketTracking).filter(TicketTracking.closed_on.is_(None)).all())
        truncated = len(tickets) > 2000
        tickets = tickets[:2000]

        # Bulk-load all status history for these tickets (avoids N+1).
        hist_by_tid = defaultdict(list)
        tids = [t.ticket_id for t in tickets]
        if tids:
            for h in (
                db.query(TicketStatusHistory)
                .filter(TicketStatusHistory.ticket_id.in_(tids))
                .order_by(TicketStatusHistory.changed_on.asc())
                .all()
            ):
                hist_by_tid[h.ticket_id].append(h)

        rows = []
        for t in tickets:
            row, _, _ = _ticket_speed_summary(db, t, hist_by_tid.get(t.ticket_id, []), today)
            row.pop("phase_days", None)
            rows.append(row)

        def _avg(vals):
            vals = [v for v in vals if v is not None]
            return round(sum(vals) / len(vals), 1) if vals else 0

        def _median(vals):
            vals = sorted(v for v in vals if v is not None)
            if not vals:
                return 0
            n = len(vals)
            return vals[n // 2] if n % 2 else round((vals[n // 2 - 1] + vals[n // 2]) / 2, 1)

        lead = [r["lead_time_days"] for r in rows]
        cycle_rows = [r for r in rows if r["cycles"] > 0]
        summary = {
            "period": {"kind": period, "offset": offset, "label": label},
            "scope": scope, "tickets": len(rows), "truncated": truncated,
            "avg_lead_time_days": _avg(lead), "median_lead_time_days": _median(lead),
            "avg_qc_days": _avg([r["qc_days"] for r in rows]),
            "avg_cycles": _avg([r["cycles"] for r in cycle_rows]),
            "first_pass_rate": round(100 * sum(1 for r in cycle_rows if r["first_pass"]) / len(cycle_rows), 1) if cycle_rows else 0,
        }

        def _group(key):
            g = defaultdict(list)
            for r in rows:
                g[r[key]].append(r)
            out = []
            for k, rs in g.items():
                out.append({
                    key: k, "tickets": len(rs),
                    "avg_lead_time_days": _avg([r["lead_time_days"] for r in rs]),
                    "avg_qc_days": _avg([r["qc_days"] for r in rs]),
                    "avg_cycles": _avg([r["cycles"] for r in rs if r["cycles"] > 0]),
                })
            out.sort(key=lambda x: x["tickets"], reverse=True)
            return out

        return {"summary": summary, "rows": rows,
                "by_module": _group("module"), "by_qc_tester": _group("qc_tester")}
    finally:
        db.close()


@app.get("/qc-cycles/summary")
def qc_cycles_summary():
    """
    Aggregate QC cycle stats: avg cycles per ticket, first-pass rate,
    cycle distribution, top cycling tickets.
    """
    db: Session = SessionLocal()
    try:
        return get_qc_cycles_summary(db)
    finally:
        db.close()


@app.get("/ageing/overview")
def ageing_overview():
    """Team-wide ageing: tickets by age bucket (0-3d, 3-7d, 7-15d, 15+d), avg ageing."""
    db: Session = SessionLocal()
    try:
        return get_ageing_overview(db)
    finally:
        db.close()


@app.get("/ageing/bottlenecks")
def ageing_bottlenecks(
    limit: int = Query(20, ge=1, le=100),
):
    """Tickets with longest QC wait, with per-status duration breakdown."""
    db: Session = SessionLocal()
    try:
        return get_ageing_bottlenecks(db, limit=limit)
    finally:
        db.close()


@app.get("/analytics/ticket-flow")
def ticket_flow_rate(
    weeks: int = Query(8, ge=1, le=52),
):
    """
    Rate of tickets entering and exiting QA per week.
    Shows weekly in/out counts and net change for throughput analysis.
    """
    db: Session = SessionLocal()
    try:
        return get_ticket_flow_rate(db, weeks=weeks)
    finally:
        db.close()


@app.get("/analytics/bis-to-closed")
def bis_to_closed():
    """
    Track duration from BIS Testing to Closed/Moved to Live for each ticket.
    Shows avg days, per-ticket breakdown, and tickets still pending in BIS.
    """
    db: Session = SessionLocal()
    try:
        return get_bis_to_closed_tracking(db)
    finally:
        db.close()


@app.get("/analytics/qa-activity-summary")
def qa_activity_summary(
    period: str = Query('past_5_days', regex='^(past_5_days|current_month|custom)$'),
    start_date: Optional[str] = Query(None, description="Custom start date YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="Custom end date YYYY-MM-DD"),
):
    """
    QA team activity summary for selected period.
    Per-member story: tickets worked on, status transitions, holds,
    priority details, current state. Designed for sharing with management.
    Supports custom date range with period=custom&start_date=...&end_date=...
    """
    db: Session = SessionLocal()
    try:
        sd_override = None
        ed_override = None
        if period == 'custom' and start_date and end_date:
            from datetime import date as date_type
            sd_override = date_type.fromisoformat(start_date)
            ed_override = date_type.fromisoformat(end_date)
        return get_qa_activity_summary(db, period=period, start_date_override=sd_override, end_date_override=ed_override)
    finally:
        db.close()


# ===== TEAM ACTIVITY BOARD (read-only, computed from synced PM data) =====

@app.get("/team-board")
def get_team_board():
    """
    All QA team members with their current ticket assignments, activity type,
    ETA, hours, workload, and idle status. Read-only view of synced PM data.
    """
    db: Session = SessionLocal()
    try:
        today = date.today()
        qa_employees = get_qa_employees(db)
        overview = get_qa_overview_data(db, today)
        queue = overview.get('queue', [])

        # Build tester → tickets mapping from active QC tickets
        tester_tickets = {}
        for t in queue:
            tester = (t.get('qc_tester') or '').strip()
            if tester:
                for name in (n.strip() for n in tester.split(',') if n.strip()):
                    name_lower = name.lower()
                    if name_lower not in tester_tickets:
                        tester_tickets[name_lower] = {'name': name, 'tickets': []}
                    tester_tickets[name_lower]['tickets'].append(t)

        members = []
        busy_count = 0
        idle_count = 0
        total_days = 0

        for emp in qa_employees:
            emp_name = (emp.name or '').strip()
            emp_lower = emp_name.lower()
            assigned = tester_tickets.get(emp_lower, {}).get('tickets', [])

            # Determine activity status
            if not assigned:
                activity = 'idle'
                idle_count += 1
            else:
                statuses = [t.get('activity_type', '') for t in assigned]
                if 'on_hold' in statuses and len(set(statuses)) == 1:
                    activity = 'on_hold'
                elif 'in_progress' in statuses or 'pending_retest' in statuses:
                    activity = 'active'
                    busy_count += 1
                else:
                    activity = 'assigned'
                    busy_count += 1

            # Total ageing across assigned tickets
            member_days = sum(t.get('days_in_qc', 0) for t in assigned)
            total_days += member_days

            # Primary ticket (highest score)
            primary_ticket = None
            if assigned:
                scored = []
                for t in assigned:
                    s = calculate_qc_priority_score(t, today)
                    scored.append((s['score'], t))
                scored.sort(key=lambda x: -x[0])
                pt = scored[0][1]
                primary_ticket = {
                    'ticket_id': pt.get('ticket_id'),
                    'title': pt.get('title', ''),
                    'status': pt.get('status', ''),
                    'priority': pt.get('priority', ''),
                    'days_in_qc': pt.get('days_in_qc', 0),
                    'eta': pt.get('eta'),
                    'activity_label': pt.get('activity_label', ''),
                    'module': pt.get('module', ''),
                    'retest_cycle_count': pt.get('retest_cycle_count', 0),
                    'tested_by_dev': pt.get('tested_by_dev', False),
                }

            members.append({
                'employee_id': emp.employee_id,
                'name': emp_name,
                'designation': emp.designation,
                'platform': emp.platform or 'Web',
                'activity': activity,
                'ticket_count': len(assigned),
                'primary_ticket': primary_ticket,
                'all_tickets': [
                    {
                        'ticket_id': t.get('ticket_id'),
                        'title': t.get('title', ''),
                        'status': t.get('status', ''),
                        'priority': t.get('priority', ''),
                        'days_in_qc': t.get('days_in_qc', 0),
                        'activity_label': t.get('activity_label', ''),
                        'eta': t.get('eta'),
                        'module': t.get('module', ''),
                        'tested_by_dev': t.get('tested_by_dev', False),
                    }
                    for t in assigned
                ],
                'total_qa_estimate_hours': sum(t.get('qa_estimate_hours') or 0 for t in assigned),
                'total_qa_actual_hours': sum(t.get('qa_actual_hours') or 0 for t in assigned),
            })

        # Sort: active first, then assigned, then on_hold, then idle
        activity_order = {'active': 0, 'assigned': 1, 'on_hold': 2, 'idle': 3}
        members.sort(key=lambda m: (activity_order.get(m['activity'], 9), m['name']))

        return {
            'members': members,
            'summary': {
                'total_members': len(members),
                'busy': busy_count,
                'idle': idle_count,
                'on_hold': sum(1 for m in members if m['activity'] == 'on_hold'),
                'total_qc_tickets': overview.get('total', 0),
                'avg_ageing': round(total_days / len(queue), 1) if queue else 0,
            },
            'status_cards': overview.get('status_cards', {}),
        }
    finally:
        db.close()


@app.get("/team-board/member/{employee_id}")
def get_team_board_member(employee_id: str):
    """Single member deep-dive: all assigned tickets with status durations."""
    db: Session = SessionLocal()
    try:
        emp = db.query(Employee).filter(Employee.employee_id == employee_id).first()
        if not emp:
            raise HTTPException(status_code=404, detail="Employee not found")

        today = date.today()
        overview = get_qa_overview_data(db, today)
        queue = overview.get('queue', [])
        emp_lower = (emp.name or '').strip().lower()

        assigned = []
        for t in queue:
            tester = (t.get('qc_tester') or '').strip().lower()
            if emp_lower and emp_lower in tester:
                durations = get_status_durations(db, t['ticket_id'], today)
                cycles = get_qc_cycle_details(db, t['ticket_id'], today)
                scoring = calculate_qc_priority_score(t, today)
                assigned.append({
                    **t,
                    'priority_score': scoring['score'],
                    'status_durations': durations.get('durations', {}),
                    'total_hold_days': durations.get('total_hold_days', 0),
                    'cycle_details': cycles,
                })

        assigned.sort(key=lambda t: -(t.get('priority_score', 0)))

        return {
            'employee_id': emp.employee_id,
            'name': emp.name,
            'designation': emp.designation,
            'platform': emp.platform or 'Web',
            'ticket_count': len(assigned),
            'tickets': assigned,
        }
    finally:
        db.close()


@app.get("/team-board/idle-members")
def get_idle_members():
    """QA members with no active QC ticket — available for assignment in PM tool."""
    db: Session = SessionLocal()
    try:
        board = get_team_board()
        idle = [m for m in board['members'] if m['activity'] == 'idle']
        return {'idle_members': idle, 'count': len(idle)}
    finally:
        db.close()


@app.get("/team-board/activity-distribution")
def get_activity_distribution():
    """Breakdown: how many members on each activity type."""
    db: Session = SessionLocal()
    try:
        board = get_team_board()
        dist = {}
        for m in board['members']:
            activity = m['activity']
            dist[activity] = dist.get(activity, 0) + 1

        # Also break down by platform
        platform_dist = {}
        for m in board['members']:
            if m['activity'] != 'idle':
                platform = m.get('platform', 'Web')
                platform_dist[platform] = platform_dist.get(platform, 0) + 1

        return {
            'activity_distribution': dist,
            'platform_distribution': platform_dist,
            'summary': board['summary'],
        }
    finally:
        db.close()


@app.get("/qa-planning/overview")
def qa_planning_overview(current_user: dict = Depends(get_current_user)):
    """
    Get QA Task Planning overview: active QC tickets (excludes BIS Testing).
    Returns status cards, priority-ordered queue, ageing, activity types, sub-department grouping.
    """
    db: Session = SessionLocal()
    try:
        # Full QA for planning leads (User.role or Employee job title/reportees); else self only
        if is_planning_lead(db, current_user):
            visible = None
        else:
            visible = get_visible_employee_ids(db, current_user)
        data = get_qa_overview_data(db)
        employees = get_qa_employees(db, visible)
        data['qa_employees'] = [
            {'employee_id': e.employee_id, 'name': e.name}
            for e in employees
        ]
        data['priority_order'] = list(QA_PRIORITY_ORDER.keys())
        return data
    finally:
        db.close()


@app.get("/qa-planning/qc-review-fail")
def qa_planning_qc_review_fail():
    """
    List tickets in QC Review Fail status (and Tested - Awaiting Fixes, Code Review Failed).
    Returns tickets with full details for the QC Review Fail tab.
    """
    db = SessionLocal()
    try:
        data = get_qa_qc_review_fail_data(db)
        return data
    finally:
        db.close()


class QATicketTestedByDevUpdate(BaseModel):
    tested_by_dev: bool


@app.patch("/qa-planning/ticket/{ticket_id}/tested-by-dev")
def qa_planning_set_tested_by_dev(
    ticket_id: int,
    body: QATicketTestedByDevUpdate,
    current_user: dict = Depends(get_current_user),
):
    """Set or unset the 'Tested By Dev' flag for a ticket (pending QA). Only applies to tickets in QC statuses."""
    db = SessionLocal()
    try:
        t = db.query(TicketTracking).filter(
            TicketTracking.ticket_id == ticket_id,
            TicketTracking.status.in_(QA_QC_STATUSES),
        ).first()
        if not t:
            raise HTTPException(status_code=404, detail="Ticket not found or not in QA QC statuses")
        flag = db.query(QATicketFlag).filter(QATicketFlag.ticket_id == ticket_id).first()
        if flag is None:
            flag = QATicketFlag(ticket_id=ticket_id, tested_by_dev=body.tested_by_dev)
            db.add(flag)
        else:
            flag.tested_by_dev = body.tested_by_dev
        db.commit()
        db.refresh(flag)
        return {"ticket_id": ticket_id, "tested_by_dev": flag.tested_by_dev}
    finally:
        db.close()


class QAReleaseResourceBody(BaseModel):
    release_date: Optional[str] = None  # YYYY-MM-DD; default today


class QAHoldTaskBody(BaseModel):
    """Request body for holding a task."""
    hold_type: str  # 'full' for entire task, 'day' for specific day
    hold_reason: str  # Required reason for holding
    hold_date: Optional[str] = None  # Required if hold_type='day', format YYYY-MM-DD


class QAResumeTaskBody(BaseModel):
    """Request body for resuming a held task."""
    resume_reason: Optional[str] = None  # Optional reason for resuming


@app.post("/qa-planning/task/{task_id}/release-resource")
def qa_planning_release_resource(
    task_id: int,
    body: Optional[QAReleaseResourceBody] = None,
    current_user: dict = Depends(get_current_user),
):
    """Mark QA resource as free from a given date (e.g. task failed early). From that date onward the task no longer blocks the resource. Optional; use when task fails and another task can be assigned."""
    db = SessionLocal()
    try:
        task = db.query(QAPlannedTask).filter(
            QAPlannedTask.id == task_id,
            QAPlannedTask.status == "active",
        ).first()
        if not task:
            raise HTTPException(status_code=404, detail="Task not found or not active")
        if not can_manage_tasks_for(db, current_user, task.employee_id):
            raise HTTPException(status_code=403, detail="Not allowed to manage this task")
        date_str = (body and body.release_date) or date.today().isoformat()
        try:
            release_date = datetime.strptime(date_str.strip(), "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="release_date must be YYYY-MM-DD")
        # Set resource_released_at to start of that day (so allocations on that date and later are excluded)
        task.resource_released_at = datetime.combine(release_date, datetime.min.time())
        db.commit()
        db.refresh(task)
        return {
            "task_id": task_id,
            "resource_released_at": task.resource_released_at.isoformat() if task.resource_released_at else None,
            "message": "QA resource is free from this date; another task can be assigned (not mandatory).",
        }
    finally:
        db.close()


@app.post("/qa-planning/task/{task_id}/hold")
def qa_planning_hold_task(
    task_id: int,
    body: QAHoldTaskBody,
    current_user: dict = Depends(get_current_user),
):
    """
    Put a QA planned task on hold. Requires PM Tracker verification.
    - hold_type: 'full' for entire task, 'day' for specific day
    - hold_reason: Required reason for holding (for reporting)
    - hold_date: Required if hold_type='day', format YYYY-MM-DD
    """
    from models import QATaskHoldHistory, QAPlannedAllocation
    
    if not body.hold_reason or not body.hold_reason.strip():
        raise HTTPException(status_code=400, detail="Hold reason is required")
    
    if body.hold_type not in ('full', 'day'):
        raise HTTPException(status_code=400, detail="hold_type must be 'full' or 'day'")
    
    hold_date = None
    if body.hold_type == 'day':
        if not body.hold_date:
            raise HTTPException(status_code=400, detail="hold_date is required for day-level hold")
        try:
            hold_date = datetime.strptime(body.hold_date.strip(), "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="hold_date must be YYYY-MM-DD")
    
    db = SessionLocal()
    try:
        task = db.query(QAPlannedTask).filter(
            QAPlannedTask.id == task_id,
            QAPlannedTask.status == "active",
        ).first()
        if not task:
            raise HTTPException(status_code=404, detail="Task not found or not active")
        
        if not can_manage_tasks_for(db, current_user, task.employee_id):
            raise HTTPException(status_code=403, detail="Not allowed to manage this task")
        
        # Check if task is already on hold
        if task.is_on_hold and body.hold_type == 'full':
            raise HTTPException(status_code=400, detail="Task is already on hold")
        
        # Verify PM Tracker status if it's a ticket task
        pm_status = None
        pm_verified = False
        if task.ticket_id:
            ticket = db.query(TicketTracking).filter(TicketTracking.ticket_id == task.ticket_id).first()
            if ticket:
                pm_status = ticket.status
                # Verify the ticket is in "QC Testing Hold" status in PM Tracker
                if pm_status == "QC Testing Hold":
                    pm_verified = True
                else:
                    raise HTTPException(
                        status_code=400, 
                        detail=f"Cannot put task on hold. Ticket status in PM Tracker is '{pm_status}'. "
                               f"Please update the ticket to 'QC Testing Hold' in PM Tracker first, "
                               f"then use the Refresh button to sync, and try again."
                    )
        
        now = datetime.utcnow()
        
        if body.hold_type == 'full':
            # Full task hold
            task.is_on_hold = True
            task.hold_reason = body.hold_reason.strip()
            task.hold_started_at = now
            task.hold_type = 'full'
            task.hold_date = None
            task.hold_ended_at = None
            task.updated_by = current_user.get("email")
        else:
            # Day-level hold - mark specific allocation as on hold
            alloc = db.query(QAPlannedAllocation).filter(
                QAPlannedAllocation.task_id == task_id,
                QAPlannedAllocation.allocation_date == hold_date,
            ).first()
            if not alloc:
                raise HTTPException(status_code=404, detail=f"No allocation found for {body.hold_date}")
            alloc.is_on_hold = True
        
        # Create hold history record
        hold_history = QATaskHoldHistory(
            task_id=task_id,
            ticket_id=task.ticket_id,
            employee_id=task.employee_id,
            employee_name=task.employee_name,
            hold_type=body.hold_type,
            hold_date=hold_date,
            hold_reason=body.hold_reason.strip(),
            pm_tracker_status=pm_status,
            pm_tracker_verified=pm_verified,
            hold_started_at=now,
            created_by=current_user.get("email"),
        )
        db.add(hold_history)
        
        db.commit()
        db.refresh(task)
        
        return {
            "task_id": task_id,
            "is_on_hold": task.is_on_hold,
            "hold_type": body.hold_type,
            "hold_date": body.hold_date if body.hold_type == 'day' else None,
            "hold_reason": body.hold_reason.strip(),
            "pm_tracker_status": pm_status,
            "pm_tracker_verified": pm_verified,
            "message": f"Task {'put on hold' if body.hold_type == 'full' else f'day {body.hold_date} put on hold'} successfully.",
        }
    finally:
        db.close()


@app.post("/qa-planning/task/{task_id}/resume")
def qa_planning_resume_task(
    task_id: int,
    body: Optional[QAResumeTaskBody] = None,
    current_user: dict = Depends(get_current_user),
):
    """
    Resume a held QA planned task.
    """
    from models import QATaskHoldHistory, QAPlannedAllocation
    
    db = SessionLocal()
    try:
        task = db.query(QAPlannedTask).filter(
            QAPlannedTask.id == task_id,
            QAPlannedTask.status == "active",
        ).first()
        if not task:
            raise HTTPException(status_code=404, detail="Task not found or not active")
        
        if not can_manage_tasks_for(db, current_user, task.employee_id):
            raise HTTPException(status_code=403, detail="Not allowed to manage this task")
        
        if not task.is_on_hold:
            raise HTTPException(status_code=400, detail="Task is not on hold")
        
        now = datetime.utcnow()
        resume_reason = (body and body.resume_reason and body.resume_reason.strip()) or None
        
        # Find the latest active hold history record and close it
        latest_hold = db.query(QATaskHoldHistory).filter(
            QATaskHoldHistory.task_id == task_id,
            QATaskHoldHistory.hold_ended_at == None,
        ).order_by(QATaskHoldHistory.hold_started_at.desc()).first()
        
        if latest_hold:
            latest_hold.hold_ended_at = now
            latest_hold.resumed_reason = resume_reason
        
        # Clear hold from task
        task.is_on_hold = False
        task.hold_ended_at = now
        task.updated_by = current_user.get("email")
        
        # Also clear any day-level holds on allocations
        db.query(QAPlannedAllocation).filter(
            QAPlannedAllocation.task_id == task_id,
            QAPlannedAllocation.is_on_hold == True,
        ).update({"is_on_hold": False})
        
        db.commit()
        db.refresh(task)
        
        return {
            "task_id": task_id,
            "is_on_hold": False,
            "message": "Task resumed successfully.",
        }
    finally:
        db.close()


@app.post("/qa-planning/refresh-pm-tracker")
def qa_planning_refresh_pm_tracker(
    ticket_id: Optional[int] = Query(None, description="Specific ticket ID to refresh, or None for full sync"),
    current_user: dict = Depends(get_current_user),
):
    """
    Refresh PM Tracker data for QA planning. 
    - If ticket_id provided: return latest status for that ticket after sync
    - If no ticket_id: trigger full sync and return summary
    """
    db = SessionLocal()
    start_time = time.time()
    
    try:
        # Trigger PM Tracker sync
        success, message, stats, sync_source = run_pm_api_sync(db, start_time)
        duration_seconds = time.time() - start_time
        
        result = {
            "success": success,
            "message": message,
            "sync_source": sync_source or "api",
            "records_updated": stats.get('records_updated', 0) + stats.get('records_added', 0),
            "duration_seconds": round(duration_seconds, 2),
        }
        
        # If specific ticket requested, return its latest status
        if ticket_id:
            ticket = db.query(TicketTracking).filter(TicketTracking.ticket_id == ticket_id).first()
            if ticket:
                result["ticket"] = {
                    "ticket_id": ticket.ticket_id,
                    "title": ticket.title,
                    "status": ticket.status,
                    "priority": ticket.priority,
                    "qc_tester": ticket.qc_tester,
                    "is_hold_status": ticket.status == "QC Testing Hold",
                }
            else:
                result["ticket"] = None
                result["ticket_message"] = f"Ticket #{ticket_id} not found"
        
        return result
    
    except Exception as e:
        logging.exception("Error refreshing PM Tracker for QA planning")
        return {
            "success": False,
            "message": f"Error: {str(e)}",
            "duration_seconds": round(time.time() - start_time, 2),
        }
    finally:
        db.close()


@app.get("/qa-planning/task/{task_id}/hold-history")
def qa_planning_task_hold_history(
    task_id: int,
    current_user: dict = Depends(get_current_user),
):
    """Get hold history for a task."""
    from models import QATaskHoldHistory
    
    db = SessionLocal()
    try:
        history = db.query(QATaskHoldHistory).filter(
            QATaskHoldHistory.task_id == task_id,
        ).order_by(QATaskHoldHistory.hold_started_at.desc()).all()
        
        return {
            "task_id": task_id,
            "history": [
                {
                    "id": h.id,
                    "hold_type": h.hold_type,
                    "hold_date": h.hold_date.isoformat() if h.hold_date else None,
                    "hold_reason": h.hold_reason,
                    "pm_tracker_status": h.pm_tracker_status,
                    "pm_tracker_verified": h.pm_tracker_verified,
                    "hold_started_at": h.hold_started_at.isoformat() if h.hold_started_at else None,
                    "hold_ended_at": h.hold_ended_at.isoformat() if h.hold_ended_at else None,
                    "resumed_reason": h.resumed_reason,
                    "created_by": h.created_by,
                }
                for h in history
            ],
        }
    finally:
        db.close()


@app.get("/qa-planning/overview/export-excel")
def qa_planning_export_excel(
    search: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    tester: Optional[str] = Query(None),
    module: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    platform: Optional[str] = Query(None),
    planning: Optional[str] = Query(None, description="planned | not_planned"),
    current_user: dict = Depends(get_current_user),
):
    """Export QA active tickets to Excel. Applies same filters as overview UI."""
    import io
    import pandas as pd
    db = SessionLocal()
    try:
        data = get_qa_overview_data(db)
        queue = data.get("queue", [])
        if search and search.strip():
            q = search.strip().lower()
            queue = [t for t in queue if (
                str(t.get("ticket_id", "")).lower().find(q) >= 0 or
                (t.get("title") or "").lower().find(q) >= 0 or
                (t.get("qc_tester") or "").lower().find(q) >= 0 or
                (t.get("module") or "").lower().find(q) >= 0
            )]
        if priority:
            queue = [t for t in queue if t.get("priority") == priority]
        if tester:
            queue = [t for t in queue if t.get("qc_tester") == tester]
        if module:
            queue = [t for t in queue if t.get("module") == module]
        if status:
            queue = [t for t in queue if t.get("status") == status]
        if platform:
            queue = [t for t in queue if (t.get("platform") or "Web") == platform]
        if planning == "planned":
            queue = [t for t in queue if t.get("qa_estimate_hours") and t.get("qa_estimate_hours", 0) > 0]
        elif planning == "not_planned":
            queue = [t for t in queue if not t.get("qa_estimate_hours") or t.get("qa_estimate_hours", 0) <= 0]
        if not queue:
            raise HTTPException(status_code=404, detail="No active QA tickets to export")
        rows = []
        for t in queue:
            rows.append({
                "Ticket ID": t.get("ticket_id"),
                "Title": t.get("title"),
                "Status": t.get("status"),
                "Priority": t.get("priority"),
                "Platform": t.get("platform", "Web"),
                "Module": t.get("module"),
                "QC Tester": t.get("qc_tester"),
                "QA Lead": t.get("qa_lead"),
                "Suggested QA": t.get("suggested_qa"),
                "Developers": t.get("developers_str"),
                "QA Estimate (h)": t.get("qa_estimate_hours"),
                "Actual QA (h)": t.get("qa_actual_hours"),
                "Dev Estimate (h)": t.get("dev_estimate_hours"),
                "Days in QC": t.get("days_in_qc"),
                "Days on Hold": t.get("days_on_hold"),
                "Activity Type": t.get("activity_label"),
                "Moved to QC On": t.get("moved_to_qc_on"),
                "ETA": t.get("eta"),
                "Next in Queue": "Yes" if t.get("is_next_in_queue") else "",
            })
        df = pd.DataFrame(rows)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="QA Active Tickets", index=False)
            from openpyxl.utils import get_column_letter
            ws = writer.sheets["QA Active Tickets"]
            for idx, col_name in enumerate(df.columns, 1):
                try:
                    col_max = df[col_name].fillna("").astype(str).str.len().max()
                except Exception:
                    col_max = 0
                max_len = max(col_max if len(df) > 0 else 0, len(str(col_name)))
                ws.column_dimensions[get_column_letter(idx)].width = min(int(max_len) + 2, 50)
        output.seek(0)
        filename = f"QA_Active_Tickets_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except HTTPException:
        raise
    finally:
        db.close()


class QAPlanningTaskCreate(BaseModel):
    employee_id: Optional[str] = None
    employee_name: str
    task_category: str  # Ticket | Team Meetings | Customer Support | Training | KT | Leave | Miscellaneous
    ticket_id: Optional[int] = None
    task_type: Optional[str] = None  # Manual Testing, Automation Testing, API Testing, Non-Functional Testing
    activity_description: str
    start_date: date
    end_date: Optional[date] = None
    total_hours: Optional[float] = None
    max_hours_per_day: Optional[float] = None
    generic_category: Optional[str] = None
    justification: Optional[str] = None


class QAPlanningTaskUpdate(BaseModel):
    start_date: Optional[date] = None
    total_hours: Optional[float] = None
    max_hours_per_day: Optional[float] = None


@app.get("/qa-planning/week/{week_start_str}")
def qa_planning_get_week(
    week_start_str: str,
    current_user: dict = Depends(get_current_user),
):
    """Get QA planning week with tasks, allocations, employee summary."""
    db = SessionLocal()
    try:
        week_start = datetime.strptime(week_start_str, "%Y-%m-%d").date()
        if week_start.weekday() != 0:
            raise HTTPException(status_code=400, detail="week_start must be a Monday")
        week_end = week_start + timedelta(days=4)

        # Week (planner + Resource Blocked Until): planning lead sees full QA dept; else self only
        if is_planning_lead(db, current_user):
            visible = None
        else:
            visible = get_visible_employee_ids(db, current_user)
        pw = get_qa_planning_week(week_start, db)
        lead_groups = get_qa_employees_for_planner(db, visible)
        
        emp_names = [e.name for _, members in lead_groups for e in members]

        leave_map = get_leave_hours_for_employees(emp_names, week_start, week_end, db)
        planning_week_id = pw.id if pw else None
        alloc_map = get_qa_allocated_hours_for_week(week_start, week_end, db, planning_week_id)

        employee_summary = []
        employee_groups = []
        for lead_name, members in lead_groups:
            group_members = []
            for e in members:
                name = e.name
                alloc_total = sum(alloc_map.get(name, {}).values())
                leave_total = sum(leave_map.get(name, {}).values())
                total_used = alloc_total + leave_total
                remaining = max(0, HOURS_PER_WEEK - total_used)
                status = "Fully Allocated" if remaining <= 0 else ("Partially Allocated" if alloc_total > 0 else "Available")
                can_manage = can_manage_tasks_for(db, current_user, e.employee_id)
                emp_data = {
                    "employee_id": e.employee_id,
                    "employee_name": name,
                    "role": getattr(e, "role", None) or "QA",
                    "allocated_hours": round(alloc_total, 1),
                    "leave_hours": round(leave_total, 1),
                    "remaining_hours": round(remaining, 1),
                    "allocation_status": status,
                    "lead_name": lead_name if lead_name != "_unassigned" else None,
                    "can_manage_tasks": can_manage,
                }
                employee_summary.append(emp_data)
                group_members.append(emp_data)
            employee_groups.append({
                "lead_name": None if lead_name == "_unassigned" else lead_name,
                "members": group_members,
            })

        tasks_list = []
        seen_task_ids = set()

        def append_task(t, allocs_to_show):
            alloc_sum = sum(a.hours or 0 for a in allocs_to_show)
            display_hours = (t.total_planned_hours or 0) if (t.total_planned_hours and t.total_planned_hours > 0) else alloc_sum
            # Use allocation date bounds as source of truth for display range.
            # This prevents stale task.start_date/end_date from showing incorrect same-day ranges.
            min_alloc_date, max_alloc_date = db.query(
                func.min(QAPlannedAllocation.allocation_date),
                func.max(QAPlannedAllocation.allocation_date),
            ).filter(QAPlannedAllocation.task_id == t.id).first()
            display_start = min_alloc_date or t.start_date
            display_end = max_alloc_date or t.end_date or display_start
            tasks_list.append({
                "id": t.id,
                "employee_name": t.employee_name,
                "employee_id": t.employee_id,
                "ticket_id": t.ticket_id,
                "ticket_title": t.ticket_title,
                "ticket_priority": t.ticket_priority,
                "generic_category": t.generic_category,
                "activity_description": t.activity_description,
                "start_date": display_start.isoformat() if display_start else None,
                "end_date": display_end.isoformat() if display_end else None,
                "total_planned_hours": round(display_hours, 1),
                "created_by": t.created_by,
                "allocations": [{"date": a.allocation_date.isoformat(), "hours": a.hours, "is_on_hold": getattr(a, 'is_on_hold', False)} for a in allocs_to_show],
                "spillover": (pw is None) or (t.planning_week_id != pw.id),
                "resource_released_at": t.resource_released_at.isoformat() if getattr(t, "resource_released_at", None) else None,
                # Hold-related fields
                "is_on_hold": getattr(t, 'is_on_hold', False) or False,
                "hold_reason": getattr(t, 'hold_reason', None),
                "hold_type": getattr(t, 'hold_type', None),
                "hold_started_at": t.hold_started_at.isoformat() if getattr(t, 'hold_started_at', None) else None,
            })

        if pw:
            tasks = db.query(QAPlannedTask).filter(
                QAPlannedTask.planning_week_id == pw.id,
                QAPlannedTask.status == "active",
            ).order_by(QAPlannedTask.employee_name, QAPlannedTask.start_date).all()
            for t in tasks:
                seen_task_ids.add(t.id)
                allocs = db.query(QAPlannedAllocation).filter(QAPlannedAllocation.task_id == t.id).order_by(QAPlannedAllocation.allocation_date).all()
                append_task(t, allocs)

        # Include spillover tasks: tasks that have allocations in this week but belong to another planning week (active only)
        spillover_allocs = (
            db.query(QAPlannedAllocation.task_id)
            .join(QAPlannedTask, QAPlannedTask.id == QAPlannedAllocation.task_id)
            .filter(
                QAPlannedTask.status == "active",
                QAPlannedAllocation.allocation_date >= week_start,
                QAPlannedAllocation.allocation_date <= week_end,
            )
            .distinct()
            .all()
        )
        spillover_task_ids = [tid for (tid,) in spillover_allocs if tid not in seen_task_ids]
        if spillover_task_ids:
            spillover_tasks = db.query(QAPlannedTask).filter(
                QAPlannedTask.id.in_(spillover_task_ids),
                QAPlannedTask.status == "active",
            ).all()
            for t in spillover_tasks:
                allocs_in_week = db.query(QAPlannedAllocation).filter(
                    QAPlannedAllocation.task_id == t.id,
                    QAPlannedAllocation.allocation_date >= week_start,
                    QAPlannedAllocation.allocation_date <= week_end,
                ).order_by(QAPlannedAllocation.allocation_date).all()
                if allocs_in_week:
                    seen_task_ids.add(t.id)
                    append_task(t, allocs_in_week)

        return {
            "week_start": week_start.isoformat(),
            "week_end": week_end.isoformat(),
            "state": pw.state if pw else "draft",
            "planning_week_id": pw.id if pw else None,
            "employees": employee_summary,
            "employee_groups": employee_groups,
            "tasks": tasks_list,
        }
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date. Use YYYY-MM-DD")
    finally:
        db.close()


@app.post("/qa-planning/week")
def qa_planning_create_week(
    week_start_str: str = Query(..., alias="week_start"),
    current_user: dict = Depends(get_current_user),
):
    """Create or get QA planning week (draft)."""
    db = SessionLocal()
    try:
        role = current_user.get("role", "")
        if not _planning_can_edit(role):
            raise HTTPException(status_code=403, detail="Only Manager or Lead can create plans")
        user_name = _get_user_display_name(db, current_user)
        week_start = datetime.strptime(week_start_str, "%Y-%m-%d").date()
        if week_start.weekday() != 0:
            raise HTTPException(status_code=400, detail="week_start must be a Monday")
        pw = get_or_create_qa_planning_week(week_start, db, user_name)
        db.commit()
        return {"planning_week_id": pw.id, "week_start": pw.week_start.isoformat(), "state": pw.state}
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date. Use YYYY-MM-DD")
    finally:
        db.close()


@app.patch("/qa-planning/week/{week_start_str}")
def qa_planning_update_week_state(
    week_start_str: str,
    body: DevPlanningWeekStateUpdate,
    current_user: dict = Depends(get_current_user),
):
    """Update QA planning week state: submit, approve, lock, or unlock (draft). Only Manager/Lead."""
    db = SessionLocal()
    try:
        role = current_user.get("role", "")
        if not _planning_can_edit(role):
            raise HTTPException(status_code=403, detail="Only Manager or Lead can change plan state")
        user_name = _get_user_display_name(db, current_user)
        week_start = datetime.strptime(week_start_str, "%Y-%m-%d").date()
        pw = get_qa_planning_week(week_start, db)
        if not pw:
            raise HTTPException(status_code=404, detail="Planning week not found")
        if body.state not in PLANNING_STATES:
            raise HTTPException(status_code=400, detail=f"state must be one of {PLANNING_STATES}")

        now = datetime.utcnow()
        if body.state == "submitted":
            pw.state = "submitted"
            pw.submitted_at = now
            pw.submitted_by = user_name
        elif body.state == "approved":
            pw.state = "approved"
            pw.approved_at = now
            pw.approved_by = user_name
        elif body.state == "locked":
            pw.state = "locked"
            pw.locked_at = now
            pw.locked_by = user_name
        elif body.state == "draft":
            if pw.state == "locked":
                pw.unlocked_at = now
                pw.unlocked_by = user_name
            pw.state = "draft"

        db.commit()
        db.refresh(pw)
        return {"planning_week_id": pw.id, "state": pw.state}
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date. Use YYYY-MM-DD")
    finally:
        db.close()


@app.get("/qa-planning/ticket-suggestions")
def qa_planning_ticket_suggestions(
    assignee: Optional[str] = Query(None, description="Tester being assigned - prioritizes their tickets"),
):
    """Categorized ticket suggestions for create-task: next in queue, on hold, for retesting, ageing."""
    db = SessionLocal()
    try:
        data = get_qa_ticket_suggestions(db, assignee=assignee)
        return data
    finally:
        db.close()


def _ticket_to_list_item(t):
    """Single ticket to list payload for /qa-planning/tickets response."""
    return {
        "ticket_id": t.ticket_id,
        "title": t.title,
        "status": t.status,
        "priority": t.priority,
        "qc_tester": t.qc_tester,
        "qa_estimate_hours": t.qa_estimate_hours,
        "dev_estimate_hours": t.dev_estimate_hours,
    }


@app.get("/qa-planning/tickets")
def qa_planning_tickets(
    search: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    assignee: Optional[str] = Query(None, description="Filter/sort for tester being assigned"),
):
    """List tickets for task planner. Shows tickets of all statuses (excluding closed/completed) so any ticket can be selected when adding a task."""
    db = SessionLocal()
    try:
        # All non-closed tickets (all statuses) for task planner
        q = db.query(TicketTracking).filter(
            TicketTracking.status.isnot(None),
            ~TicketTracking.status.in_(CLOSED_STATUSES),
        )
        if search:
            s = f"%{search}%"
            q = q.filter(
                or_(
                    TicketTracking.ticket_id.cast(String).like(s),
                    TicketTracking.title.ilike(s),
                    TicketTracking.qc_tester.ilike(s),
                )
            )
        if status:
            q = q.filter(TicketTracking.status == status)
        if priority:
            q = q.filter(TicketTracking.priority == priority)
        tickets = q.order_by(TicketTracking.ticket_id.desc()).limit(200).all()
        # If search is a single ticket ID, include that ticket even when not in QC status (any status / any team)
        ticket_ids_seen = {t.ticket_id for t in tickets}
        if search and search.strip().isdigit():
            tid = int(search.strip())
            if tid not in ticket_ids_seen:
                t_any = db.query(TicketTracking).filter(TicketTracking.ticket_id == tid).first()
                if t_any:
                    tickets = [t_any] + tickets
                    ticket_ids_seen.add(tid)
        if assignee and assignee.strip():
            assignee_lower = assignee.strip().lower()
            def sort_key(t):
                qc = (t.qc_tester or "").strip().lower()
                is_unassigned = not qc
                is_for_tester = assignee_lower in qc or qc in assignee_lower
                is_hold = (t.status or "").lower().find("hold") >= 0
                if is_unassigned:
                    return (0, t.ticket_id)
                if is_for_tester:
                    return (1, -t.ticket_id)
                if is_hold:
                    return (2, -t.ticket_id)
                return (3, t.ticket_id)
            tickets = sorted(tickets, key=sort_key)
        return {"tickets": [_ticket_to_list_item(t) for t in tickets]}
    finally:
        db.close()


def _qa_planning_ticket_payload(t):
    """Build the JSON payload for a ticket in QA add-task (shared by get and refresh)."""
    if not t:
        return None
    qa_est = t.qa_estimate_hours or 0
    qa_actual = t.actual_qa_hours or 0
    remaining = max(0, qa_est - qa_actual) if qa_est else None
    eta_val = None
    if getattr(t, "eta", None):
        eta_val = t.eta.isoformat() if hasattr(t.eta, "isoformat") else str(t.eta)
    return {
        "ticket_id": t.ticket_id,
        "title": t.title,
        "status": t.status,
        "priority": t.priority,
        "qc_tester": t.qc_tester,
        "qa_estimate_hours": t.qa_estimate_hours,
        "actual_qa_hours": t.actual_qa_hours,
        "remaining_qa_hours": remaining,
        "dev_estimate_hours": t.dev_estimate_hours,
        "eta": eta_val,
    }


@app.get("/qa-planning/ticket/{ticket_id}")
def qa_planning_get_ticket(ticket_id: int):
    """Get ticket details for QA add-task. Returns any ticket by ID (any status) so users can create tasks at their discretion."""
    db = SessionLocal()
    try:
        t = db.query(TicketTracking).filter(TicketTracking.ticket_id == ticket_id).first()
        if not t:
            return None
        payload = _qa_planning_ticket_payload(t)
        if payload:
            payload["in_qc_status"] = (t.status or "") in QA_QC_STATUSES
        return payload
    finally:
        db.close()


@app.post("/qa-planning/ticket/{ticket_id}/refresh")
def qa_planning_refresh_ticket(ticket_id: str):
    """Fetch this ticket from PM API only, upsert into DB, return ticket details for QA add-task. Fast single-ticket refresh.
    Returns ticket data even when not in QC status (with in_qc_status=false) so Add Task can show details and warnings."""
    try:
        tid = int(ticket_id)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="Invalid ticket ID")
    db = SessionLocal()
    try:
        client = PMApiClient()
        success, tickets, message = client.fetch_tickets(ticket_id=tid)
        if success and tickets:
            try:
                mapped = client.map_api_fields(tickets)
                upsert_tickets(db, mapped, sync_source="api")
            except Exception as e:
                logger.exception(f"Refresh ticket {tid}: map/upsert failed")
                raise HTTPException(status_code=500, detail=f"Failed to update ticket from PM: {str(e)}")
        elif not success:
            logger.warning(f"Refresh ticket {tid}: API returned success=False - {message}")
        # Prefer ticket in QC statuses; if not found, return ticket by id anyway so UI can show data and status warning
        t_qc = db.query(TicketTracking).filter(
            TicketTracking.ticket_id == tid,
            TicketTracking.status.in_(QA_QC_STATUSES),
        ).first()
        if t_qc:
            payload = _qa_planning_ticket_payload(t_qc)
            if payload:
                payload["in_qc_status"] = True
            return payload
        t_any = db.query(TicketTracking).filter(TicketTracking.ticket_id == tid).first()
        if t_any:
            payload = _qa_planning_ticket_payload(t_any)
            if payload:
                payload["in_qc_status"] = False
            return payload
        return None
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Refresh ticket {tid} failed")
        raise HTTPException(status_code=500, detail=f"Refresh failed: {str(e)}")
    finally:
        db.close()


@app.get("/qa-planning/available-hours")
def qa_planning_available_hours(
    employee_name: str = Query(...),
    date: str = Query(..., description="YYYY-MM-DD"),
    current_user: dict = Depends(get_current_user),
):
    """Get available hours for QA employee on date."""
    db = SessionLocal()
    try:
        emp = db.query(Employee).filter(Employee.name.ilike(f"%{employee_name}%")).first()
        if not emp:
            raise HTTPException(status_code=404, detail="Employee not found")
        if not can_access_employee(db, current_user, emp.employee_id):
            raise HTTPException(status_code=403, detail="Access denied")
        d = datetime.strptime(date, "%Y-%m-%d").date()
        avail = get_qa_available_hours_on_date(employee_name, d, db)
        return {"available_hours": avail}
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date")
    finally:
        db.close()


@app.get("/qa-planning/next-available-date")
def qa_planning_next_available_date(
    employee_name: str = Query(...),
    from_date: Optional[str] = Query(None, description="YYYY-MM-DD; default today"),
    current_user: dict = Depends(get_current_user),
):
    """First date on or after from_date (or today) where the QA employee has available hours."""
    db = SessionLocal()
    try:
        emp = db.query(Employee).filter(Employee.name.ilike(f"%{employee_name}%")).first()
        if not emp:
            raise HTTPException(status_code=404, detail="Employee not found")
        if not can_access_employee(db, current_user, emp.employee_id):
            raise HTTPException(status_code=403, detail="Access denied")
        start = datetime.strptime(from_date or datetime.now().strftime("%Y-%m-%d"), "%Y-%m-%d").date()
        next_d = get_qa_next_available_date(employee_name, start, db)
        return {"date": next_d.isoformat()}
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date. Use YYYY-MM-DD")
    finally:
        db.close()


@app.get("/qa-planning/availability-summary")
def qa_planning_availability_summary(
    employee_name: str = Query(...),
    week_start: str = Query(..., description="Monday of week (YYYY-MM-DD)"),
    current_user: dict = Depends(get_current_user),
):
    """Fully available from date and partially available days this week for the QA employee."""
    db = SessionLocal()
    try:
        emp = db.query(Employee).filter(Employee.name.ilike(f"%{employee_name}%")).first()
        if not emp:
            raise HTTPException(status_code=404, detail="Employee not found")
        if not can_access_employee(db, current_user, emp.employee_id):
            raise HTTPException(status_code=403, detail="Access denied")
        wstart = datetime.strptime(week_start, "%Y-%m-%d").date()
        return get_qa_availability_summary(employee_name, wstart, db)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date. Use YYYY-MM-DD")
    finally:
        db.close()


@app.get("/qa-planning/allocation-preview")
def qa_planning_allocation_preview(
    employee_name: str = Query(...),
    start_date: str = Query(...),
    total_hours: float = Query(...),
    max_hours_per_day: float = Query(8.0),
    week_start: str = Query(...),
    current_user: dict = Depends(get_current_user),
):
    """Preview allocation distribution for QA task."""
    db = SessionLocal()
    try:
        emp = db.query(Employee).filter(Employee.name.ilike(f"%{employee_name}%")).first()
        if not emp:
            raise HTTPException(status_code=404, detail="Employee not found")
        if not can_access_employee(db, current_user, emp.employee_id):
            raise HTTPException(status_code=403, detail="Access denied")
        start = datetime.strptime(start_date, "%Y-%m-%d").date()
        wstart = datetime.strptime(week_start, "%Y-%m-%d").date()
        wend = wstart + timedelta(days=4)
        dist = simulate_qa_allocation_distribution(employee_name, start, total_hours, wstart, wend, db, max_hours_per_day=max_hours_per_day)
        return {
            "distribution": [{"date": d.isoformat(), "hours": h} for d, h in dist],
            "total": sum(h for _, h in dist),
        }
    except ValueError as e:
        return {"error": str(e)}
    finally:
        db.close()


@app.post("/qa-planning/tasks")
def qa_planning_add_task(
    body: QAPlanningTaskCreate,
    week_start_str: str = Query(..., alias="week_start"),
    current_user: dict = Depends(get_current_user),
):
    """Add a QA planned task. Only Manager/Lead."""
    db = SessionLocal()
    try:
        role = current_user.get("role", "")
        if not _planning_can_edit(role):
            raise HTTPException(status_code=403, detail="Only Manager or Lead can add tasks")
        user_name = _get_user_display_name(db, current_user)
        week_start = datetime.strptime(week_start_str, "%Y-%m-%d").date()
        if week_start.weekday() != 0:
            raise HTTPException(status_code=400, detail="week_start must be a Monday")
        week_end = week_start + timedelta(days=4)

        pw = get_or_create_qa_planning_week(week_start, db, user_name)
        db.commit()

        TASK_CATEGORIES = [
            "Ticket",
            "Team Meetings",
            "Customer Support",
            "Training",
            "KT",
            "Leave",
            "Half Day Leave",
            "Miscellaneous",
            "Generic Task",
            "Regression",
            "Live Testing",
        ]
        GENERIC_CATEGORIES = [
            "Team Meetings",
            "Customer Support",
            "Training",
            "KT",
            "Leave",
            "Half Day Leave",
            "Miscellaneous",
            "Generic Task",
            "Regression",
            "Live Testing",
        ]

        raw_task_category = (body.task_category or body.generic_category or "").strip()
        task_category_map = {c.lower(): c for c in TASK_CATEGORIES}
        normalized_task_category = task_category_map.get(raw_task_category.lower())
        if not normalized_task_category:
            raise HTTPException(status_code=400, detail="Invalid task category. Use: " + ", ".join(TASK_CATEGORIES))
        body.task_category = normalized_task_category
        max_hours_per_day = body.max_hours_per_day if body.max_hours_per_day is not None else 8.0
        is_half_day_leave = body.task_category == "Half Day Leave"
        is_leave_category = body.task_category in ["Leave", "Half Day Leave"]
        if is_half_day_leave:
            max_hours_per_day = 4.0

        ticket_id = body.ticket_id
        ticket_title = None
        ticket_priority = None
        if body.task_category == "Ticket":
            if not ticket_id:
                raise HTTPException(status_code=400, detail="Ticket ID is required for Ticket category")
            tkt = db.query(TicketTracking).filter(TicketTracking.ticket_id == ticket_id).first()
            if not tkt:
                raise HTTPException(status_code=404, detail="Ticket not found")
            in_qc = (tkt.status or "") in QA_QC_STATUSES
            if in_qc:
                if not tkt.qa_estimate_hours or tkt.qa_estimate_hours <= 0:
                    raise HTTPException(status_code=400, detail="QA Estimate is required in PM Tracker. Please add it and refresh.")
                if not (tkt.qc_tester or "").strip():
                    raise HTTPException(status_code=400, detail="QC Tester is required in PM Tracker. Please assign and refresh.")
            ticket_title = tkt.title
            ticket_priority = tkt.priority

        if body.task_category != "Ticket":
            raw_generic_category = (body.generic_category or body.task_category or "").strip()
            generic_category_map = {c.lower(): c for c in GENERIC_CATEGORIES}
            normalized_generic_category = generic_category_map.get(raw_generic_category.lower())
            if not normalized_generic_category:
                raise HTTPException(status_code=400, detail="Task category must be one of: " + ", ".join(GENERIC_CATEGORIES))
            body.generic_category = normalized_generic_category

        emp = db.query(Employee).filter(Employee.name.ilike(f"%{body.employee_name}%")).first()
        if not emp:
            raise HTTPException(status_code=404, detail="Employee not found")
        if not can_access_employee(db, current_user, emp.employee_id):
            raise HTTPException(status_code=403, detail="Access denied to this employee")

        if is_half_day_leave:
            total_hours = 4.0
        else:
            total_hours = body.total_hours if body.total_hours is not None else (max_hours_per_day if is_leave_category else 8.0)
        if body.task_category == "Ticket" and tkt and tkt.qa_estimate_hours and (body.total_hours is None or body.total_hours <= 0):
            total_hours = float(tkt.qa_estimate_hours)
        if total_hours is None or total_hours < 0.5:
            raise HTTPException(status_code=400, detail="Duration must be at least 0.5 hours")

        try:
            simulate_qa_allocation_distribution(
                emp.name, body.start_date, total_hours, week_start, week_end, db, pw.id, max_hours_per_day=max_hours_per_day
            )
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))

        # Duplicate detection: check for existing identical task (prevents double-submit on network issues)
        existing_task = db.query(QAPlannedTask).filter(
            QAPlannedTask.planning_week_id == pw.id,
            QAPlannedTask.employee_name == emp.name,
            QAPlannedTask.status == "active",
            QAPlannedTask.start_date == body.start_date,
            QAPlannedTask.total_planned_hours == round(total_hours, 1),
        )
        if body.task_category == "Ticket" and ticket_id:
            existing_task = existing_task.filter(QAPlannedTask.ticket_id == ticket_id)
        else:
            existing_task = existing_task.filter(
                QAPlannedTask.generic_category == body.generic_category,
                QAPlannedTask.activity_description == body.activity_description.strip(),
            )
        existing_task = existing_task.first()
        if existing_task:
            # Return existing task instead of creating duplicate
            allocs = db.query(QAPlannedAllocation).filter(QAPlannedAllocation.task_id == existing_task.id).order_by(QAPlannedAllocation.allocation_date).all()
            return {
                "id": existing_task.id,
                "employee_name": existing_task.employee_name,
                "ticket_id": existing_task.ticket_id,
                "ticket_title": existing_task.ticket_title,
                "activity_description": existing_task.activity_description,
                "start_date": existing_task.start_date.isoformat() if existing_task.start_date else None,
                "end_date": existing_task.end_date.isoformat() if existing_task.end_date else None,
                "total_planned_hours": float(existing_task.total_planned_hours or 0),
                "allocations": [{"date": a.allocation_date.isoformat(), "hours": float(a.hours or 0)} for a in allocs],
                "duplicate_prevented": True,
            }

        task = QAPlannedTask(
            planning_week_id=pw.id,
            employee_id=emp.employee_id,
            employee_name=emp.name,
            ticket_id=ticket_id,
            ticket_title=ticket_title,
            ticket_priority=ticket_priority,
            generic_category=body.generic_category if body.task_category != "Ticket" else None,
            task_type=body.task_type if body.task_category == "Ticket" else None,
            activity_description=body.activity_description.strip(),
            start_date=body.start_date,
            end_date=body.end_date or body.start_date,
            total_planned_hours=round(total_hours, 1),
            status="active",
            created_by=user_name,
        )
        db.add(task)
        db.commit()
        db.refresh(task)

        create_qa_allocations_for_task(task.id, emp.name, body.start_date, task.total_planned_hours, db, max_hours_per_day)
        alloc_sum = db.query(func.coalesce(func.sum(QAPlannedAllocation.hours), 0)).filter(QAPlannedAllocation.task_id == task.id).scalar()
        last_date = db.query(func.max(QAPlannedAllocation.allocation_date)).filter(QAPlannedAllocation.task_id == task.id).scalar()
        task.total_planned_hours = round(float(alloc_sum or 0), 1)
        if last_date:
            task.end_date = last_date
        db.commit()

        allocs = db.query(QAPlannedAllocation).filter(QAPlannedAllocation.task_id == task.id).order_by(QAPlannedAllocation.allocation_date).all()
        return {
            "success": True,
            "task": {
                "id": task.id,
                "employee_name": task.employee_name,
                "ticket_id": task.ticket_id,
                "ticket_priority": task.ticket_priority,
                "activity_description": task.activity_description,
                "start_date": task.start_date.isoformat(),
                "end_date": task.end_date.isoformat() if task.end_date else None,
                "total_planned_hours": task.total_planned_hours,
                "allocations": [{"date": a.allocation_date.isoformat(), "hours": a.hours} for a in allocs],
            }
        }
    except HTTPException:
        raise
    finally:
        db.close()


@app.delete("/qa-planning/tasks/{task_id}")
def qa_planning_delete_task(
    task_id: int,
    current_user: dict = Depends(get_current_user),
):
    """Delete a QA planned task. Only Manager/Lead."""
    db = SessionLocal()
    try:
        role = current_user.get("role", "")
        if not _planning_can_edit(role):
            raise HTTPException(status_code=403, detail="Only Manager or Lead can delete tasks")
        user_name = _get_user_display_name(db, current_user)
        user_email = (current_user.get("email") or "").strip().lower()
        task = db.query(QAPlannedTask).filter(QAPlannedTask.id == task_id).first()
        if not task:
            raise HTTPException(status_code=404, detail="Task not found")
        _emp = db.query(Employee).filter(Employee.name == task.employee_name).first()
        task_employee_id = (task.employee_id or (_emp.employee_id if _emp else None)) or ""
        task_created_by = (getattr(task, "created_by", None) or "").strip().lower()
        can_manage_employee = can_manage_tasks_for(db, current_user, task_employee_id)
        is_task_creator = bool(task_created_by) and task_created_by in {
            (user_name or "").strip().lower(),
            user_email,
        }
        if not can_manage_employee and not is_task_creator:
            raise HTTPException(status_code=403, detail="You can only remove tasks for employees you manage or tasks you created")
        pw = db.query(QAPlanningWeek).filter(QAPlanningWeek.id == task.planning_week_id).first()
        if pw and getattr(pw, "state", None) in ("approved", "locked"):
            raise HTTPException(status_code=403, detail="Cannot delete task; plan is approved or locked")
        db.query(QAPlannedAllocation).filter(QAPlannedAllocation.task_id == task_id).delete()
        task.status = "removed"
        db.commit()
        return {"success": True}
    finally:
        db.close()


def _qa_planning_update_task_impl(task_id: int, body: QAPlanningTaskUpdate, current_user: dict, db: Session):
    """Shared logic for PATCH/PUT update of a QA planned task."""
    role = current_user.get("role", "")
    if not _planning_can_edit(role):
        raise HTTPException(status_code=403, detail="Only Manager or Lead can edit tasks")
    task = db.query(QAPlannedTask).filter(QAPlannedTask.id == task_id, QAPlannedTask.status == "active").first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    _emp = db.query(Employee).filter(Employee.name == task.employee_name).first()
    task_employee_id = (task.employee_id or (_emp.employee_id if _emp else None)) or ""
    if not can_manage_tasks_for(db, current_user, task_employee_id):
        raise HTTPException(status_code=403, detail="You can only edit tasks for employees you manage")
    pw = db.query(QAPlanningWeek).filter(QAPlanningWeek.id == task.planning_week_id).first()
    if not pw:
        raise HTTPException(status_code=400, detail="Planning week not found")
    if getattr(pw, "state", None) in ("approved", "locked"):
        raise HTTPException(status_code=403, detail="Cannot edit task; plan is approved or locked")
    week_start = pw.week_start
    week_end = week_start + timedelta(days=4)
    start_date = body.start_date if body.start_date is not None else task.start_date
    total_hours = body.total_hours if body.total_hours is not None else task.total_planned_hours
    max_hours_per_day = body.max_hours_per_day if body.max_hours_per_day is not None else 8.0
    if total_hours is None or total_hours < 0.5:
        raise HTTPException(status_code=400, detail="Duration must be at least 0.5 hours")
    # Delete existing allocations first so simulate sees correct free capacity (otherwise current task's hours are double-counted)
    db.query(QAPlannedAllocation).filter(QAPlannedAllocation.task_id == task_id).delete()
    db.flush()
    try:
        simulate_qa_allocation_distribution(
            task.employee_name, start_date, total_hours, week_start, week_end, db,
            planning_week_id=pw.id, exclude_task_id=task_id, max_hours_per_day=max_hours_per_day
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    task.start_date = start_date
    task.total_planned_hours = round(total_hours, 1)
    create_qa_allocations_for_task(task.id, task.employee_name, start_date, task.total_planned_hours, db, max_hours_per_day)
    alloc_sum = db.query(func.coalesce(func.sum(QAPlannedAllocation.hours), 0)).filter(QAPlannedAllocation.task_id == task.id).scalar()
    last_date = db.query(func.max(QAPlannedAllocation.allocation_date)).filter(QAPlannedAllocation.task_id == task.id).scalar()
    task.total_planned_hours = round(float(alloc_sum or 0), 1)
    if last_date:
        task.end_date = last_date
    task.updated_by = _get_user_display_name(db, current_user)
    db.commit()
    db.refresh(task)
    allocs = db.query(QAPlannedAllocation).filter(QAPlannedAllocation.task_id == task.id).order_by(QAPlannedAllocation.allocation_date).all()
    return {
        "success": True,
        "task": {
            "id": task.id,
            "employee_name": task.employee_name,
            "start_date": task.start_date.isoformat() if task.start_date else None,
            "end_date": task.end_date.isoformat() if task.end_date else None,
            "total_planned_hours": task.total_planned_hours,
            "allocations": [{"date": a.allocation_date.isoformat(), "hours": a.hours} for a in allocs],
        }
    }


@app.patch("/qa-planning/tasks/{task_id}")
def qa_planning_update_task_patch(
    task_id: int,
    body: QAPlanningTaskUpdate,
    current_user: dict = Depends(get_current_user),
):
    """Update a QA planned task (start date, total hours). Recreates allocations. Only Manager/Lead. Blocked if week approved/locked."""
    db = SessionLocal()
    try:
        return _qa_planning_update_task_impl(task_id, body, current_user, db)
    finally:
        db.close()


@app.put("/qa-planning/tasks/{task_id}")
def qa_planning_update_task_put(
    task_id: int,
    body: QAPlanningTaskUpdate,
    current_user: dict = Depends(get_current_user),
):
    """Update a QA planned task (same as PATCH). PUT supported for compatibility where PATCH is blocked."""
    db = SessionLocal()
    try:
        return _qa_planning_update_task_impl(task_id, body, current_user, db)
    finally:
        db.close()


@app.get("/qa-planning/calendar")
def qa_planning_calendar(
    view: str = Query("weekly", description="weekly | monthly"),
    date_str: Optional[str] = Query(None),
    month_str: Optional[str] = Query(None, description="YYYY-MM for monthly"),
    current_user: dict = Depends(get_current_user),
):
    """Calendar view: employees x days with allocated hours and task labels. Same format as dev-planning/calendar."""
    db = SessionLocal()
    try:
        from dev_planning import get_planning_week_dates
        # Calendar: planning lead sees full QA dept; else self only
        if is_planning_lead(db, current_user):
            visible = None
        else:
            visible = get_visible_employee_ids(db, current_user)
        lead_groups = get_qa_employees_for_planner(db, visible)
        ref = (date_str or month_str or "").strip() or date.today().isoformat()
        if len(ref) == 7:  # YYYY-MM
            ref = ref + "-01"
        try:
            ref_date = datetime.strptime(ref, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid date format. Use YYYY-MM-DD or YYYY-MM, got: {ref!r}")
        week_start, week_end = get_planning_week_dates(ref_date)

        if view == "monthly" and month_str:
            y, m = map(int, month_str.split("-"))
            month_start = date(y, m, 1)
            month_end = (date(y, m + 1, 1) - timedelta(days=1)) if m < 12 else (date(y + 1, 1, 1) - timedelta(days=1))
        else:
            month_start = week_start
            month_end = week_end

        start_range = month_start
        end_range = month_end

        employees = [e for _, members in lead_groups for e in members]
        emp_names = [e.name for e in employees]
        leave_map = get_leave_hours_for_employees(emp_names, start_range, end_range, db)
        alloc_map = get_qa_allocated_hours_for_week(start_range, end_range, db, None)

        tasks_by_emp_date = defaultdict(lambda: defaultdict(lambda: {"hours": 0, "items": []}))
        seen_keys = defaultdict(set)
        tasks = db.query(QAPlannedTask, QAPlannedAllocation).join(
            QAPlannedAllocation, QAPlannedAllocation.task_id == QAPlannedTask.id
        ).filter(
            QAPlannedTask.status == "active",
            QAPlannedAllocation.allocation_date >= start_range,
            QAPlannedAllocation.allocation_date <= end_range,
            _allocation_not_released(),
        ).all()
        # Resolve priority from TicketTracking when task.ticket_priority is null (e.g. older tasks)
        # Only integer ticket_ids: TicketTracking.ticket_id is integer; planned tasks use int, timesheet may have strings
        def _int_ticket_ids(ids):
            out = []
            for x in ids:
                if x is None:
                    continue
                try:
                    out.append(int(x) if not isinstance(x, int) else x)
                except (TypeError, ValueError):
                    continue
            return list(set(out))
        ticket_ids_qa = _int_ticket_ids([t.ticket_id for t, _ in tasks if t.ticket_id])
        ticket_priority_map = {}
        if ticket_ids_qa:
            tkts = db.query(TicketTracking).filter(TicketTracking.ticket_id.in_(ticket_ids_qa)).all()
            ticket_priority_map = {tk.ticket_id: tk.priority for tk in tkts if getattr(tk, "priority", None)}
        for t, a in tasks:
            key = (t.ticket_id and f"#{t.ticket_id}") or (t.generic_category or "")
            category = "Ticket" if t.ticket_id else (t.generic_category or "Miscellaneous")
            priority = t.ticket_priority or (ticket_priority_map.get(t.ticket_id) if t.ticket_id else None)
            tasks_by_emp_date[t.employee_name][a.allocation_date]["hours"] += a.hours
            cell_key = (t.employee_name, a.allocation_date)
            if key and key not in seen_keys[cell_key]:
                seen_keys[cell_key].add(key)
                tasks_by_emp_date[t.employee_name][a.allocation_date]["items"].append({
                    "text": key,
                    "ticket_id": t.ticket_id,
                    "ticket_priority": priority,
                    "category": category,
                    "over_estimate": False,
                })

        # Planner Calendar shows only planning-module data (no EnhancedTimesheet / Google Sheet actuals here)
        times_by_emp_date = defaultdict(lambda: defaultdict(lambda: {"hours": 0, "items": []}))

        working_days_list = get_working_days_list(start_range, end_range, db)
        capacity_hours = len(working_days_list) * 8

        rows = []
        for e in employees:
            alloc_total = sum(alloc_map.get(e.name, {}).values())
            leave_total = sum(leave_map.get(e.name, {}).values())
            total_used = alloc_total + leave_total
            remaining = max(0, capacity_hours - total_used)
            status = "Fully Allocated" if remaining <= 0 else ("Partially Allocated" if alloc_total > 0 else "Available")
            can_manage = can_manage_tasks_for(db, current_user, e.employee_id)

            row = {
                "employee_id": e.employee_id,
                "employee_name": e.name,
                "lead_name": (e.lead or "").strip() or None,
                "allocated_hours": round(alloc_total, 1),
                "leave_hours": round(leave_total, 1),
                "remaining_hours": round(remaining, 1),
                "allocation_status": status,
                "can_manage_tasks": can_manage,
                "days": {},
            }
            d = start_range
            while d <= end_range:
                leave_h = leave_map.get(e.name, {}).get(d, 0)
                alloc_h = alloc_map.get(e.name, {}).get(d, 0)
                # Use same source as items (tasks_by_emp_date) so cell hours match displayed tasks
                planned_h = tasks_by_emp_date[e.name][d]["hours"]
                cell_hours = round(planned_h, 1) if planned_h else round(alloc_h, 1)
                total = leave_h + cell_hours
                # Planned hours and items
                cell = {"hours": cell_hours, "leave_hours": round(leave_h, 1), "total": round(total, 1)}
                cell["items"] = tasks_by_emp_date[e.name][d]["items"][:5]
                # Actual (timesheet) hours and ticket items for the date (if available)
                actual_h = times_by_emp_date.get(e.name, {}).get(d, {}).get("hours", 0)
                raw_actual = times_by_emp_date.get(e.name, {}).get(d, {}).get("items", [])[:5]
                # Enrich actual items with ticket_priority and text for calendar display
                actual_items = []
                for it in raw_actual:
                    tid = it.get("ticket_id")
                    try:
                        tid_int = int(tid) if tid is not None and (isinstance(tid, int) or (isinstance(tid, str) and str(tid).strip().isdigit())) else None
                    except (TypeError, ValueError):
                        tid_int = None
                    priority = ticket_priority_map.get(tid_int) if tid_int is not None else None
                    item = dict(it)
                    if tid:
                        item["text"] = f"#{tid}" if tid_int is not None else str(tid)[:30]
                        item["ticket_priority"] = priority
                    actual_items.append(item)
                cell["actual_hours"] = round(actual_h, 1)
                cell["actual_items"] = actual_items
                row["days"][d.isoformat()] = cell
                d += timedelta(days=1)
            rows.append(row)
        return {"view": view, "start": start_range.isoformat(), "end": end_range.isoformat(), "employees": rows}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"QA calendar error: {str(e)}")
    finally:
        db.close()


@app.get("/qa-planning/day-details")
def qa_planning_day_details(
    employee_name: str = Query(...),
    date_str: str = Query(...),
    current_user: dict = Depends(get_current_user),
):
    """Get detailed task list for an employee on a specific date."""
    db = SessionLocal()
    try:
        emp = db.query(Employee).filter(Employee.name.ilike(f"%{employee_name}%")).first()
        if not emp:
            raise HTTPException(status_code=404, detail="Employee not found")
        if not can_access_employee(db, current_user, emp.employee_id):
            raise HTTPException(status_code=403, detail="Access denied")
        alloc_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        allocs = db.query(QAPlannedTask, QAPlannedAllocation).join(
            QAPlannedAllocation, QAPlannedAllocation.task_id == QAPlannedTask.id
        ).filter(
            QAPlannedTask.status == "active",
            QAPlannedTask.employee_name == employee_name,
            QAPlannedAllocation.allocation_date == alloc_date,
        ).all()
        planned_tasks = []
        for t, a in allocs:
            planned_tasks.append({
                "task_id": t.id,
                "ticket_id": t.ticket_id,
                "ticket_priority": t.ticket_priority,
                "generic_category": t.generic_category,
                "activity_description": t.activity_description or (f"#{t.ticket_id}" if t.ticket_id else t.generic_category),
                "hours": round(a.hours, 1),
                "total_planned_hours": t.total_planned_hours,
                "category": "Ticket" if t.ticket_id else (t.generic_category or "Miscellaneous"),
                "is_planned": True,
            })
        
        # Also get actual timesheet entries for the date and employee
        actual_entries = []
        try:
            ts_data = db.query(EnhancedTimesheet).filter(
                EnhancedTimesheet.employee_name == employee_name,
                EnhancedTimesheet.date == alloc_date,
            ).all()
            for ts in ts_data:
                hours = ts.productive_hours if ts.productive_hours and ts.productive_hours > 0 else (ts.hours_logged or 0)
                actual_entries.append({
                    "ticket_id": ts.ticket_id,
                    "hours": round(hours, 1),
                    "task_description": ts.task_description,
                    "project_name": ts.project_name,
                    "is_actual": True,
                })
        except Exception:
            pass
        
        # Combine: planned tasks first, then actual entries with different entries
        all_tasks = planned_tasks + actual_entries
        return {"employee_name": employee_name, "date": date_str, "tasks": all_tasks}
    finally:
        db.close()


@app.get("/dev-planning/weeks")
def dev_planning_list_weeks(
    year: Optional[int] = Query(None, description="Filter by year. Default: current year."),
    current_user: dict = Depends(get_current_user),
):
    """List planning weeks (optionally by year)."""
    db = SessionLocal()
    try:
        y = year or date.today().year
        start = date(y, 1, 1)
        end = date(y, 12, 31)
        weeks = (
            db.query(DevPlanningWeek)
            .filter(DevPlanningWeek.week_start >= start, DevPlanningWeek.week_start <= end)
            .order_by(DevPlanningWeek.week_start.desc())
            .all()
        )
        return {
            "weeks": [
                {
                    "id": w.id,
                    "week_start": w.week_start.isoformat(),
                    "week_end": w.week_end.isoformat(),
                    "state": w.state,
                    "created_by": w.created_by,
                }
            ]
            for w in weeks
        }
    finally:
        db.close()


@app.get("/dev-planning/working-days")
def dev_planning_working_days(
    week_start_str: str = Query(..., alias="week_start", description="Monday of week (YYYY-MM-DD)"),
    current_user: dict = Depends(get_current_user),
):
    """Get list of working days (Mon–Fri excluding holidays) for the planning week."""
    db = SessionLocal()
    try:
        week_start = datetime.strptime(week_start_str, "%Y-%m-%d").date()
        week_end = week_start + timedelta(days=4)
        days = get_working_days_list(week_start, week_end, db)
        return {"week_start": week_start.isoformat(), "week_end": week_end.isoformat(), "working_days": [d.isoformat() for d in days]}
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date. Use YYYY-MM-DD")
    finally:
        db.close()


@app.get("/dev-planning/available-hours")
def dev_planning_available_hours(
    employee_name: str = Query(..., description="Employee name"),
    date_str: str = Query(..., alias="date", description="Date (YYYY-MM-DD)"),
    current_user: dict = Depends(get_current_user),
):
    """Get available hours for an employee on a specific date."""
    db = SessionLocal()
    try:
        emp = db.query(Employee).filter(Employee.name.ilike(f"%{employee_name}%")).first()
        if not emp:
            raise HTTPException(status_code=404, detail="Employee not found")
        if not can_access_employee(db, current_user, emp.employee_id):
            raise HTTPException(status_code=403, detail="Access denied")
        target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        available = get_available_hours_on_date(employee_name, target_date, db)
        return {
            "date": target_date.isoformat(),
            "employee_name": employee_name,
            "available_hours": round(available, 1),
            "max_per_day": HOURS_PER_DAY,
        }
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date. Use YYYY-MM-DD")
    finally:
        db.close()


@app.get("/dev-planning/next-available-date")
def dev_planning_next_available_date(
    employee_name: str = Query(...),
    from_date: Optional[str] = Query(None, description="YYYY-MM-DD; default today"),
    current_user: dict = Depends(get_current_user),
):
    """First date on or after from_date (or today) where the employee has available hours."""
    db = SessionLocal()
    try:
        emp = db.query(Employee).filter(Employee.name.ilike(f"%{employee_name}%")).first()
        if not emp:
            raise HTTPException(status_code=404, detail="Employee not found")
        if not can_access_employee(db, current_user, emp.employee_id):
            raise HTTPException(status_code=403, detail="Access denied")
        start = datetime.strptime(from_date or datetime.now().strftime("%Y-%m-%d"), "%Y-%m-%d").date()
        next_d = get_next_available_date(employee_name, start, db)
        return {"date": next_d.isoformat()}
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date. Use YYYY-MM-DD")
    finally:
        db.close()


@app.get("/dev-planning/availability-summary")
def dev_planning_availability_summary(
    employee_name: str = Query(...),
    week_start: str = Query(..., description="Monday of week (YYYY-MM-DD)"),
    current_user: dict = Depends(get_current_user),
):
    """Fully available from date and partially available days this week for the employee."""
    db = SessionLocal()
    try:
        emp = db.query(Employee).filter(Employee.name.ilike(f"%{employee_name}%")).first()
        if not emp:
            raise HTTPException(status_code=404, detail="Employee not found")
        if not can_access_employee(db, current_user, emp.employee_id):
            raise HTTPException(status_code=403, detail="Access denied")
        wstart = datetime.strptime(week_start, "%Y-%m-%d").date()
        return get_availability_summary(employee_name, wstart, db)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date. Use YYYY-MM-DD")
    finally:
        db.close()


@app.get("/dev-planning/allocation-preview")
def dev_planning_allocation_preview(
    employee_name: str = Query(..., description="Employee name"),
    start_date_str: str = Query(..., alias="start_date", description="Start date (YYYY-MM-DD)"),
    total_hours: float = Query(..., ge=0.5, le=40, description="Duration in hours"),
    max_hours_per_day: float = Query(8, ge=0.5, le=8, description="Max hours per day for this task"),
    week_start_str: str = Query(..., alias="week_start", description="Monday of week (YYYY-MM-DD)"),
    ticket_id: Optional[int] = Query(None, description="Ticket ID for duplicate check"),
    generic_category: Optional[str] = Query(None, description="Generic category for duplicate check"),
    current_user: dict = Depends(get_current_user),
):
    """Preview how hours would be distributed (max_hours_per_day per day). Returns distribution or error."""
    db = SessionLocal()
    try:
        emp = db.query(Employee).filter(Employee.name.ilike(f"%{employee_name}%")).first()
        if not emp:
            raise HTTPException(status_code=404, detail="Employee not found")
        if not can_access_employee(db, current_user, emp.employee_id):
            raise HTTPException(status_code=403, detail="Access denied")
        week_start = datetime.strptime(week_start_str, "%Y-%m-%d").date()
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        week_end = week_start + timedelta(days=4)
        pw = get_planning_week(week_start, db)
        planning_week_id = pw.id if pw else None
        # Get max available on start date for suggestion
        max_available_start = get_available_hours_on_date(employee_name, start_date, db)

        dist = simulate_allocation_distribution(
            employee_name, start_date, total_hours, week_start, week_end, db, planning_week_id,
            max_hours_per_day=max_hours_per_day,
        )
        proposed_dates = [d for d, h in dist]
        dup_error = check_duplicate_task(employee_name, ticket_id, generic_category, proposed_dates, db)
        if dup_error:
            raise HTTPException(status_code=400, detail=dup_error)
        return {
            "distribution": [{"date": d.isoformat(), "hours": round(h, 1)} for d, h in dist],
            "total": round(sum(h for _, h in dist), 1),
            "max_available_on_start_date": round(max_available_start, 1),
        }
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid date. Use YYYY-MM-DD")
    finally:
        db.close()


@app.get("/dev-planning/week/{week_start_str}")
def dev_planning_get_week(
    week_start_str: str,
    current_user: dict = Depends(get_current_user),
):
    """Get one planning week with state, tasks, allocations, and employee summary."""
    db = SessionLocal()
    try:
        week_start = datetime.strptime(week_start_str, "%Y-%m-%d").date()
        if week_start.weekday() != 0:
            raise HTTPException(status_code=400, detail="week_start must be a Monday (YYYY-MM-DD)")
        week_end = week_start + timedelta(days=4)

        # Week (planner + Resource Blocked Until): planning lead sees full Dev dept; else self only
        if is_planning_lead(db, current_user):
            visible = None
        else:
            visible = get_visible_employee_ids(db, current_user)
        employees = get_development_employees(db, visible)

        emp_names = [e.name for e in employees]
        leave_map = get_leave_hours_for_employees(emp_names, week_start, week_end, db)

        pw = get_planning_week(week_start, db)
        planning_week_id = pw.id if pw else None
        alloc_map = get_allocated_hours_for_week(week_start, week_end, db, planning_week_id)

        # Employee summary: allocated, leave, remaining; can_manage_tasks for lead view vs assign
        employee_summary = []
        for e in employees:
            name = e.name
            alloc_total = sum(alloc_map.get(name, {}).values())
            leave_total = sum(leave_map.get(name, {}).values())
            total_used = alloc_total + leave_total
            remaining = max(0, HOURS_PER_WEEK - total_used)
            status = "Fully Allocated" if remaining <= 0 else ("Partially Allocated" if alloc_total > 0 else "Available")
            can_manage = can_manage_tasks_for(db, current_user, e.employee_id)
            employee_summary.append({
                "employee_id": e.employee_id,
                "employee_name": name,
                "lead_name": (e.lead or "").strip() or None,
                "role": getattr(e, "role", None) or "Developer",
                "allocated_hours": round(alloc_total, 1),
                "leave_hours": round(leave_total, 1),
                "remaining_hours": round(remaining, 1),
                "allocation_status": status,
                "can_manage_tasks": can_manage,
            })

        # Tasks and allocations for the week (include spillover: tasks from other weeks with allocations in this week)
        tasks_list = []
        seen_task_ids = set()

        def append_dev_task(t, allocs_to_show):
            # Use allocation date bounds as source of truth for display range.
            # This prevents stale task.start_date/end_date from showing incorrect same-day ranges.
            min_alloc_date, max_alloc_date = db.query(
                func.min(DevPlannedAllocation.allocation_date),
                func.max(DevPlannedAllocation.allocation_date),
            ).filter(DevPlannedAllocation.task_id == t.id).first()
            display_start = min_alloc_date or t.start_date
            display_end = max_alloc_date or t.end_date or display_start
            tasks_list.append({
                "id": t.id,
                "employee_name": t.employee_name,
                "employee_id": t.employee_id,
                "ticket_id": t.ticket_id,
                "ticket_title": t.ticket_title,
                "ticket_priority": ticket_priority_map.get(t.ticket_id) if t.ticket_id else None,
                "generic_category": t.generic_category,
                "activity_description": t.activity_description,
                "start_date": display_start.isoformat() if display_start else None,
                "end_date": display_end.isoformat() if display_end else None,
                "allocation_pct": t.allocation_pct,
                "total_planned_hours": t.total_planned_hours,
                "created_by": t.created_by,
                "allocations": [{"date": a.allocation_date.isoformat(), "hours": a.hours} for a in allocs_to_show],
                "spillover": (pw is None) or (t.planning_week_id != pw.id),
            })

        week_tasks = []
        ticket_ids = []
        if pw:
            week_tasks = db.query(DevPlannedTask).filter(
                DevPlannedTask.planning_week_id == pw.id,
                DevPlannedTask.status == "active",
            ).order_by(DevPlannedTask.employee_name, DevPlannedTask.start_date).all()
            for t in week_tasks:
                seen_task_ids.add(t.id)
            ticket_ids = [t.ticket_id for t in week_tasks if t.ticket_id]

        # Spillover: tasks that have allocations in this week but belong to another planning week (active only)
        spillover_allocs = (
            db.query(DevPlannedAllocation.task_id)
            .join(DevPlannedTask, DevPlannedTask.id == DevPlannedAllocation.task_id)
            .filter(
                DevPlannedTask.status == "active",
                DevPlannedAllocation.allocation_date >= week_start,
                DevPlannedAllocation.allocation_date <= week_end,
            )
            .distinct()
            .all()
        )
        spillover_task_ids = [tid for (tid,) in spillover_allocs if tid not in seen_task_ids]
        spillover_tasks = []
        if spillover_task_ids:
            spillover_tasks = db.query(DevPlannedTask).filter(
                DevPlannedTask.id.in_(spillover_task_ids),
                DevPlannedTask.status == "active",
            ).all()
            for t in spillover_tasks:
                seen_task_ids.add(t.id)
            ticket_ids = list(set(ticket_ids) | {t.ticket_id for t in spillover_tasks if t.ticket_id})

        ticket_priority_map = {}
        if ticket_ids:
            tkts = db.query(TicketTracking.ticket_id, TicketTracking.priority).filter(TicketTracking.ticket_id.in_(ticket_ids)).all()
            ticket_priority_map = {tk.ticket_id: tk.priority for tk in tkts if tk.priority}

        for t in week_tasks:
            allocs = db.query(DevPlannedAllocation).filter(DevPlannedAllocation.task_id == t.id).order_by(DevPlannedAllocation.allocation_date).all()
            append_dev_task(t, allocs)
        for t in spillover_tasks:
            allocs_in_week = db.query(DevPlannedAllocation).filter(
                DevPlannedAllocation.task_id == t.id,
                DevPlannedAllocation.allocation_date >= week_start,
                DevPlannedAllocation.allocation_date <= week_end,
            ).order_by(DevPlannedAllocation.allocation_date).all()
            if allocs_in_week:
                append_dev_task(t, allocs_in_week)

        return {
            "week_start": week_start.isoformat(),
            "week_end": week_end.isoformat(),
            "state": pw.state if pw else "draft",
            "planning_week_id": pw.id if pw else None,
            "employees": employee_summary,
            "tasks": tasks_list,
        }
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date. Use YYYY-MM-DD")
    finally:
        db.close()


@app.post("/dev-planning/week")
def dev_planning_create_week(
    week_start_str: str = Query(..., alias="week_start", description="Monday of week (YYYY-MM-DD)"),
    current_user: dict = Depends(get_current_user),
):
    """Create or get planning week (draft). Only Manager/Lead."""
    db = SessionLocal()
    try:
        role = current_user.get("role", "")
        if not _planning_can_edit(role):
            raise HTTPException(status_code=403, detail="Only Manager or Lead can create plans")
        user_name = _get_user_display_name(db, current_user)
        week_start = datetime.strptime(week_start_str, "%Y-%m-%d").date()
        if week_start.weekday() != 0:
            raise HTTPException(status_code=400, detail="week_start must be a Monday")
        pw = get_or_create_planning_week(week_start, db, user_name)
        log_audit(db, pw.id, "create_week", "week", pw.id, None, {"state": pw.state}, user_name)
        db.commit()
        return {"planning_week_id": pw.id, "week_start": pw.week_start.isoformat(), "state": pw.state}
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date. Use YYYY-MM-DD")
    finally:
        db.close()


@app.patch("/dev-planning/week/{week_start_str}")
def dev_planning_update_week_state(
    week_start_str: str,
    body: DevPlanningWeekStateUpdate,
    current_user: dict = Depends(get_current_user),
):
    """Update planning week state: submit, approve, lock, or unlock (draft). Only Manager/Lead."""
    db = SessionLocal()
    try:
        role = current_user.get("role", "")
        if not _planning_can_edit(role):
            raise HTTPException(status_code=403, detail="Only Manager or Lead can change plan state")
        user_name = _get_user_display_name(db, current_user)
        week_start = datetime.strptime(week_start_str, "%Y-%m-%d").date()
        pw = get_planning_week(week_start, db)
        if not pw:
            raise HTTPException(status_code=404, detail="Planning week not found")
        if body.state not in PLANNING_STATES:
            raise HTTPException(status_code=400, detail=f"state must be one of {PLANNING_STATES}")

        old_state = pw.state
        now = datetime.utcnow()
        if body.state == "submitted":
            pw.state = "submitted"
            pw.submitted_at = now
            pw.submitted_by = user_name
        elif body.state == "approved":
            pw.state = "approved"
            pw.approved_at = now
            pw.approved_by = user_name
        elif body.state == "locked":
            pw.state = "locked"
            pw.locked_at = now
            pw.locked_by = user_name
        elif body.state == "draft":
            if old_state == "locked":
                pw.unlocked_at = now
                pw.unlocked_by = user_name
            pw.state = "draft"
        log_audit(db, pw.id, "update_week_state", "week", pw.id, {"state": old_state}, {"state": pw.state}, user_name)
        db.commit()
        db.refresh(pw)
        return {"planning_week_id": pw.id, "state": pw.state}
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date. Use YYYY-MM-DD")
    finally:
        db.close()


# Dev applicable statuses (active, non-closed) - used for dev-planning tickets list
DEV_APPLICABLE_STATUSES = [
    'Ready For Development', 'Technical Review', 'Approved for Live',
    'Live - awaiting fixes', 'Express Lane Review', 'In Progress',
    'Start Code Review', 'Code Review Failed', 'QC Review Fail',
    'Code Review Passed', 'Tested - Awaiting Fixes', 'Re-opened', 'Reopened'
]


@app.get("/dev-planning/tickets/filter-options")
def dev_planning_tickets_filter_options(current_user: dict = Depends(get_current_user)):
    """Return distinct statuses, priorities, and assignees for Dev tickets filter dropdowns."""
    db = SessionLocal()
    try:
        base = db.query(TicketTracking).filter(
            TicketTracking.ticket_id.isnot(None),
            TicketTracking.status.in_(DEV_APPLICABLE_STATUSES),
        )
        statuses = [r[0] for r in base.with_entities(TicketTracking.status).distinct().all() if r[0]]
        priorities = [r[0] for r in base.with_entities(TicketTracking.priority).distinct().all() if r[0]]
        assignees = set()
        for col in (TicketTracking.backend_developer, TicketTracking.frontend_developer, TicketTracking.current_assignee):
            for r in base.with_entities(col).distinct().all():
                s = (r[0] or "").strip()
                if s:
                    assignees.add(s)
        assignees = sorted(assignees)
        return {"statuses": sorted(statuses), "priorities": sorted(priorities), "assignees": assignees}
    finally:
        db.close()


@app.get("/dev-planning/ticket/{ticket_id}")
def dev_planning_get_ticket(
    ticket_id: int,
    current_user: dict = Depends(get_current_user),
):
    """Get ticket details for add-task lookup. Returns null if not found or not in DEV applicable statuses."""
    db = SessionLocal()
    try:
        t = db.query(TicketTracking).filter(
            TicketTracking.ticket_id == ticket_id,
            TicketTracking.status.in_(DEV_APPLICABLE_STATUSES),
        ).first()
        if not t:
            return None
        estimate = float(t.dev_estimate_hours) if t.dev_estimate_hours is not None else None
        utilised = float(t.actual_dev_hours) if t.actual_dev_hours is not None else 0
        remaining = (estimate - utilised) if estimate is not None else None
        return {
            "ticket_id": t.ticket_id,
            "title": t.title,
            "status": t.status,
            "priority": t.priority,
            "dev_estimate_hours": estimate,
            "actual_dev_hours": utilised,
            "remaining_dev_hours": remaining,
            "qa_estimate_hours": t.qa_estimate_hours,
            "assignee": t.current_assignee or t.backend_developer or t.frontend_developer,
        }
    finally:
        db.close()


@app.get("/dev-planning/tickets")
def dev_planning_tickets(
    current_user: dict = Depends(get_current_user),
    search: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    assignee: Optional[str] = Query(None),
    unassigned: Optional[bool] = Query(None, description="Only tickets with no developer assigned"),
    has_estimate: Optional[bool] = Query(None, description="Filter tickets with dev_estimate_hours"),
):
    """List PM Tracker tickets for left panel. Only active Dev tickets in applicable statuses."""
    db = SessionLocal()
    try:
        q = db.query(TicketTracking).filter(TicketTracking.ticket_id.isnot(None))
        # Only active Dev tickets in applicable statuses (excludes closed, QA, BIS, etc.)
        q = q.filter(TicketTracking.status.in_(DEV_APPLICABLE_STATUSES))
        if search:
            search_term = search.strip()
            if search_term:
                # Search across ticket_id (partial), title, and assignee fields
                search_pattern = f"%{search_term}%"
                ticket_id_match = func.cast(TicketTracking.ticket_id, String).ilike(search_pattern)
                title_match = TicketTracking.title.ilike(search_pattern)
                backend_match = TicketTracking.backend_developer.ilike(search_pattern)
                frontend_match = TicketTracking.frontend_developer.ilike(search_pattern)
                assignee_match = TicketTracking.current_assignee.ilike(search_pattern)
                q = q.filter(or_(
                    ticket_id_match,
                    title_match,
                    backend_match,
                    frontend_match,
                    assignee_match,
                ))
        if status:
            q = q.filter(TicketTracking.status.ilike(f"%{status}%"))
        if priority:
            q = q.filter(TicketTracking.priority.ilike(f"%{priority}%"))
        if assignee:
            q = q.filter(or_(
                TicketTracking.backend_developer.ilike(f"%{assignee}%"),
                TicketTracking.frontend_developer.ilike(f"%{assignee}%"),
                TicketTracking.current_assignee.ilike(f"%{assignee}%"),
            ))
        if unassigned:
            # No backend AND no frontend developer assigned
            q = q.filter(
                func.coalesce(func.trim(TicketTracking.backend_developer), "") == "",
                func.coalesce(func.trim(TicketTracking.frontend_developer), "") == "",
            )
        if has_estimate is not None:
            if has_estimate:
                q = q.filter(TicketTracking.dev_estimate_hours.isnot(None), TicketTracking.dev_estimate_hours > 0)
            else:
                q = q.filter(or_(TicketTracking.dev_estimate_hours.is_(None), TicketTracking.dev_estimate_hours == 0))
        tickets = q.order_by(TicketTracking.updated_on.desc()).limit(200).all()
        return {
            "tickets": [
                {
                    "ticket_id": t.ticket_id,
                    "title": t.title,
                    "status": t.status,
                    "priority": t.priority,
                    "assignee": t.current_assignee or t.backend_developer or t.frontend_developer,
                    "dev_estimate_hours": t.dev_estimate_hours,
                    "qa_estimate_hours": t.qa_estimate_hours,
                    "eta": t.eta.isoformat() if t.eta else None,
                }
                for t in tickets
            ]
        }
    finally:
        db.close()


@app.post("/dev-planning/tasks")
def dev_planning_add_task(
    body: DevPlanningTaskCreate,
    week_start_str: str = Query(..., alias="week_start", description="Monday of planning week (YYYY-MM-DD)"),
    current_user: dict = Depends(get_current_user),
):
    """Add a planned task with allocations. Validates 8h/day, 40h/week. Only Manager/Lead."""
    db = SessionLocal()
    try:
        role = current_user.get("role", "")
        if not _planning_can_edit(role):
            raise HTTPException(status_code=403, detail="Only Manager or Lead can add tasks")
        user_name = _get_user_display_name(db, current_user)
        week_start = datetime.strptime(week_start_str, "%Y-%m-%d").date()
        if week_start.weekday() != 0:
            raise HTTPException(status_code=400, detail="week_start must be a Monday")
        week_end = week_start + timedelta(days=4)

        pw = get_or_create_planning_week(week_start, db, user_name)
        db.commit()

        # Validations
        raw_task_category = (body.task_category or body.generic_category or "").strip()
        task_category_map = {c.lower(): c for c in TASK_CATEGORIES}
        normalized_task_category = task_category_map.get(raw_task_category.lower())
        if not normalized_task_category:
            raise HTTPException(status_code=400, detail="Invalid task category. Use: " + ", ".join(TASK_CATEGORIES))
        body.task_category = normalized_task_category
        if body.start_date < week_start or body.start_date > week_end:
            raise HTTPException(status_code=400, detail="Start date must be within the planning week")
        # Note: Past date validation removed - users can create tasks for past dates in the current week
        if body.end_date and body.end_date < body.start_date:
            raise HTTPException(status_code=400, detail="End date cannot be before start date")
        max_hours_per_day = body.max_hours_per_day if body.max_hours_per_day is not None else 8.0
        is_half_day_leave = body.task_category == "Half Day Leave"
        if is_half_day_leave:
            max_hours_per_day = 4.0
        if max_hours_per_day < 0.5 or max_hours_per_day > 8:
            raise HTTPException(status_code=400, detail="Max hours per day must be 0.5–8")

        ticket_id = body.ticket_id
        ticket_title = None
        if body.task_category == "Ticket":
            if not ticket_id:
                raise HTTPException(status_code=400, detail="Ticket ID is required when Task Category is Ticket")
            tkt = db.query(TicketTracking).filter(TicketTracking.ticket_id == ticket_id).first()
            if not tkt:
                raise HTTPException(status_code=404, detail="Ticket not found")
            # Require dev estimate only when total_hours not provided (we'd use ticket estimate)
            if body.total_hours is None and (tkt.dev_estimate_hours is None or tkt.dev_estimate_hours <= 0):
                raise HTTPException(status_code=400, detail="Ticket has no dev estimate; enter Duration (hours) to allocate")
            # Validate: allocated hours must not exceed remaining (estimate - utilised)
            if body.total_hours is not None and tkt.dev_estimate_hours is not None:
                utilised = float(tkt.actual_dev_hours or 0)
                remaining = float(tkt.dev_estimate_hours) - utilised
                if remaining >= 0 and body.total_hours > remaining:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Cannot allocate {body.total_hours}h; only {remaining:.1f}h remaining (estimate {tkt.dev_estimate_hours}h − utilised {utilised}h)"
                    )
            ticket_title = tkt.title
        else:
            raw_generic_category = (body.generic_category or body.task_category or "").strip()
            generic_category_map = {c.lower(): c for c in GENERIC_CATEGORIES}
            normalized_generic_category = generic_category_map.get(raw_generic_category.lower())
            if not normalized_generic_category:
                raise HTTPException(status_code=400, detail="Task category must be one of: " + ", ".join(GENERIC_CATEGORIES))
            body.generic_category = normalized_generic_category
            # Justification is optional for non-ticket tasks

        # Resolve employee
        emp = db.query(Employee).filter(Employee.name.ilike(f"%{body.employee_name}%")).first()
        if not emp:
            raise HTTPException(status_code=404, detail="Employee not found")
        if not can_access_employee(db, current_user, emp.employee_id):
            raise HTTPException(status_code=403, detail="Access denied to this employee")
        employee_id = emp.employee_id
        employee_name = emp.name

        # Total hours: for Half Day Leave always 4h, else use total_hours if provided (1-40), else derive from end_date or ticket
        if is_half_day_leave:
            total_hours = 4.0
            body.end_date = body.start_date
        elif body.total_hours is not None:
            total_hours = min(max(float(body.total_hours), 0.5), HOURS_PER_WEEK)
            body.end_date = body.start_date  # spillover will extend
        elif body.end_date:
            working_days = get_working_days_list(body.start_date, body.end_date, db)
            num_days = len(working_days)
            total_hours = num_days * max_hours_per_day
        elif ticket_id and ticket_title:
            total_hours = float(tkt.dev_estimate_hours)
            body.end_date = body.start_date
        else:
            total_hours = max_hours_per_day
            body.end_date = body.start_date

        # Pre-check: can we distribute total_hours from start_date (max_hours_per_day/day, no over-allocation)?
        try:
            proposed_distribution = simulate_allocation_distribution(
                employee_name, body.start_date, total_hours, week_start, week_end, db, pw.id, max_hours_per_day=max_hours_per_day
            )
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))

        # Duplicate detection: check for existing identical task (prevents double-submit on network issues)
        existing_task = db.query(DevPlannedTask).filter(
            DevPlannedTask.planning_week_id == pw.id,
            DevPlannedTask.employee_name == employee_name,
            DevPlannedTask.status == "active",
            DevPlannedTask.start_date == body.start_date,
            DevPlannedTask.total_planned_hours == round(total_hours, 1),
        )
        if body.task_category == "Ticket" and ticket_id:
            existing_task = existing_task.filter(DevPlannedTask.ticket_id == ticket_id)
        else:
            existing_task = existing_task.filter(
                DevPlannedTask.generic_category == body.generic_category,
                DevPlannedTask.activity_description == body.activity_description.strip(),
            )
        existing_task = existing_task.first()
        if existing_task:
            # Return existing task instead of creating duplicate
            allocs = db.query(DevPlannedAllocation).filter(DevPlannedAllocation.task_id == existing_task.id).order_by(DevPlannedAllocation.allocation_date).all()
            return {
                "success": True,
                "task": {
                    "id": existing_task.id,
                    "employee_name": existing_task.employee_name,
                    "ticket_id": existing_task.ticket_id,
                    "activity_description": existing_task.activity_description,
                    "start_date": existing_task.start_date.isoformat() if existing_task.start_date else None,
                    "end_date": existing_task.end_date.isoformat() if existing_task.end_date else None,
                    "total_planned_hours": float(existing_task.total_planned_hours or 0),
                    "allocations": [{"date": a.allocation_date.isoformat(), "hours": float(a.hours or 0)} for a in allocs],
                },
                "duplicate_prevented": True,
            }

        # Check for duplicate task (same ticket/category on same days) - still show warning for different hours/dates
        proposed_dates = [d for d, h in proposed_distribution]
        dup_error = check_duplicate_task(
            employee_name, ticket_id, body.generic_category, proposed_dates, db
        )
        if dup_error:
            raise HTTPException(status_code=400, detail=dup_error)

        allocation_pct = max(10, min(100, round((max_hours_per_day / 8) * 100)))
        task = DevPlannedTask(
            planning_week_id=pw.id,
            employee_id=employee_id,
            employee_name=employee_name,
            ticket_id=ticket_id,
            ticket_title=ticket_title,
            generic_category=body.generic_category,
            justification=body.justification,
            activity_description=body.activity_description.strip(),
            start_date=body.start_date,
            end_date=body.end_date,
            allocation_pct=allocation_pct,
            total_planned_hours=round(total_hours, 1),
            status="active",
            created_by=user_name,
        )
        db.add(task)
        db.commit()
        db.refresh(task)

        # Create allocation rows (spillover over working days, max max_hours_per_day per day)
        create_allocations_for_task(
            task.id, employee_name, body.start_date, task.total_planned_hours,
            week_start, week_end, db, pw.id, max_hours_per_day=max_hours_per_day,
        )
        # Sync total_planned_hours and end_date from allocations
        alloc_sum = db.query(func.coalesce(func.sum(DevPlannedAllocation.hours), 0)).filter(DevPlannedAllocation.task_id == task.id).scalar()
        last_date = db.query(func.max(DevPlannedAllocation.allocation_date)).filter(DevPlannedAllocation.task_id == task.id).scalar()
        task.total_planned_hours = round(float(alloc_sum or 0), 1)
        if last_date:
            task.end_date = last_date
        db.commit()

        log_audit(db, pw.id, "add_task", "task", task.id, None, {"employee_name": employee_name, "total_hours": task.total_planned_hours}, user_name)
        allocs = db.query(DevPlannedAllocation).filter(DevPlannedAllocation.task_id == task.id).order_by(DevPlannedAllocation.allocation_date).all()
        return {
            "success": True,
            "task": {
                "id": task.id,
                "employee_name": task.employee_name,
                "ticket_id": task.ticket_id,
                "activity_description": task.activity_description,
                "start_date": task.start_date.isoformat(),
                "end_date": task.end_date.isoformat() if task.end_date else None,
                "total_planned_hours": task.total_planned_hours,
                "allocations": [{"date": a.allocation_date.isoformat(), "hours": a.hours} for a in allocs],
            }
        }
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        db.close()


@app.patch("/dev-planning/tasks/{task_id}")
def dev_planning_update_task(
    task_id: int,
    body: DevPlanningTaskUpdate,
    current_user: dict = Depends(get_current_user),
):
    """Update a planned task; recalculates allocations. Only Manager/Lead. Blocked if week approved/locked."""
    db = SessionLocal()
    try:
        role = current_user.get("role", "")
        if not _planning_can_edit(role):
            raise HTTPException(status_code=403, detail="Only Manager or Lead can edit tasks")
        user_name = _get_user_display_name(db, current_user)
        task = db.query(DevPlannedTask).filter(DevPlannedTask.id == task_id).first()
        if not task:
            raise HTTPException(status_code=404, detail="Task not found")
        _emp = db.query(Employee).filter(Employee.name == task.employee_name).first()
        task_employee_id = (task.employee_id or (_emp.employee_id if _emp else None)) or ""
        if not can_manage_tasks_for(db, current_user, task_employee_id):
            raise HTTPException(status_code=403, detail="You can only edit tasks for employees you manage")
        pw = db.query(DevPlanningWeek).filter(DevPlanningWeek.id == task.planning_week_id).first()
        if pw and pw.state in ("approved", "locked"):
            raise HTTPException(status_code=403, detail="Cannot edit task; plan is approved or locked")
        week_start, week_end = task.planning_week_id and pw.week_start or get_planning_week_dates(task.start_date), (pw.week_end if pw else task.start_date + timedelta(days=4))

        # Apply updates
        if body.activity_description is not None:
            task.activity_description = body.activity_description
        if body.start_date is not None:
            task.start_date = body.start_date
        if body.end_date is not None:
            task.end_date = body.end_date
        if body.allocation_pct is not None:
            if body.allocation_pct not in ALLOCATION_PCT_VALID:
                raise HTTPException(status_code=400, detail="Allocation percentage must be multiple of 10 (10-100)")
            task.allocation_pct = body.allocation_pct

        # Recreate allocations: delete existing and re-run create_allocations_for_task
        db.query(DevPlannedAllocation).filter(DevPlannedAllocation.task_id == task_id).delete()
        if pw:
            create_allocations_for_task(task.id, task.employee_name, task.start_date, task.total_planned_hours, pw.week_start, pw.week_end, db, pw.id)
        total_now = db.query(func.coalesce(func.sum(DevPlannedAllocation.hours), 0)).filter(DevPlannedAllocation.task_id == task_id).scalar()
        task.total_planned_hours = round(float(total_now or 0), 1)
        db.commit()
        db.refresh(task)
        log_audit(db, task.planning_week_id, "edit_task", "task", task.id, None, {"total_planned_hours": task.total_planned_hours}, user_name)
        return {"success": True, "task": {"id": task.id, "total_planned_hours": task.total_planned_hours}}
    except HTTPException:
        raise
    finally:
        db.close()


@app.delete("/dev-planning/tasks/{task_id}")
def dev_planning_delete_task(
    task_id: int,
    current_user: dict = Depends(get_current_user),
):
    """Delete a planned task. Only Manager/Lead. Blocked if week approved/locked."""
    db = SessionLocal()
    try:
        role = current_user.get("role", "")
        if not _planning_can_edit(role):
            raise HTTPException(status_code=403, detail="Only Manager or Lead can delete tasks")
        user_name = _get_user_display_name(db, current_user)
        user_email = (current_user.get("email") or "").strip().lower()
        task = db.query(DevPlannedTask).filter(DevPlannedTask.id == task_id).first()
        if not task:
            raise HTTPException(status_code=404, detail="Task not found")
        _emp = db.query(Employee).filter(Employee.name == task.employee_name).first()
        task_employee_id = (task.employee_id or (_emp.employee_id if _emp else None)) or ""
        task_created_by = (getattr(task, "created_by", None) or "").strip().lower()
        can_manage_employee = can_manage_tasks_for(db, current_user, task_employee_id)
        is_task_creator = bool(task_created_by) and task_created_by in {
            (user_name or "").strip().lower(),
            user_email,
        }
        if not can_manage_employee and not is_task_creator:
            raise HTTPException(status_code=403, detail="You can only delete tasks for employees you manage or tasks you created")
        today = date.today()
        # Block delete only if task is entirely in the past (no allocation on or after today); allow spillover tasks with future work
        if task.start_date and task.start_date < today:
            has_future_allocation = db.query(DevPlannedAllocation).filter(
                DevPlannedAllocation.task_id == task_id,
                DevPlannedAllocation.allocation_date >= today,
            ).first() is not None
            if not has_future_allocation:
                raise HTTPException(status_code=403, detail="Cannot delete task; past tasks cannot be edited")
        pw = db.query(DevPlanningWeek).filter(DevPlanningWeek.id == task.planning_week_id).first()
        if pw and pw.state in ("approved", "locked"):
            raise HTTPException(status_code=403, detail="Cannot delete task; plan is approved or locked")
        db.query(DevPlannedAllocation).filter(DevPlannedAllocation.task_id == task_id).delete()
        db.delete(task)
        log_audit(db, task.planning_week_id, "delete_task", "task", task_id, {"employee_name": task.employee_name}, None, user_name)
        db.commit()
        return {"success": True, "message": "Task deleted"}
    except HTTPException:
        raise
    finally:
        db.close()


class AllocationUpdate(BaseModel):
    date: str
    hours: float


class TaskAllocationsUpdate(BaseModel):
    allocations: List[AllocationUpdate]
    spillover_hours: Optional[float] = 0  # Hours to redistribute to next available days


@app.put("/dev-planning/tasks/{task_id}/allocations")
def dev_planning_update_task_allocations(
    task_id: int,
    body: TaskAllocationsUpdate,
    current_user: dict = Depends(get_current_user),
):
    """Update task allocations with spillover support. Removed hours spill to next available days."""
    db = SessionLocal()
    try:
        role = current_user.get("role", "")
        if not _planning_can_edit(role):
            raise HTTPException(status_code=403, detail="Only Manager or Lead can edit tasks")
        user_name = _get_user_display_name(db, current_user)

        task = db.query(DevPlannedTask).filter(DevPlannedTask.id == task_id).first()
        if not task:
            raise HTTPException(status_code=404, detail="Task not found")

        pw = db.query(DevPlanningWeek).filter(DevPlanningWeek.id == task.planning_week_id).first()
        if pw and pw.state in ("approved", "locked"):
            raise HTTPException(status_code=403, detail="Cannot edit task; plan is approved or locked")

        today = date.today()
        new_allocs = []
        for a in body.allocations:
            alloc_date = datetime.strptime(a.date, "%Y-%m-%d").date()
            # Skip past dates that were removed (keep them as-is)
            if a.hours > 0:
                new_allocs.append({"date": alloc_date, "hours": min(8, max(0, a.hours))})

        # Get current allocations
        current_allocs = db.query(DevPlannedAllocation).filter(DevPlannedAllocation.task_id == task_id).all()
        current_total = sum(a.hours for a in current_allocs)
        new_total = sum(a["hours"] for a in new_allocs)
        spillover = body.spillover_hours or 0

        # If there's spillover, find next available days and add allocations
        if spillover > 0:
            # Find the last date in new allocations
            if new_allocs:
                last_date = max(a["date"] for a in new_allocs)
            else:
                last_date = task.start_date

            # Get working days after last_date
            range_end = last_date + timedelta(days=60)
            from dev_planning import get_working_days_list, get_leave_hours_for_employees, get_allocated_hours_for_week, HOURS_PER_DAY
            working_days = get_working_days_list(last_date + timedelta(days=1), range_end, db)

            # Get existing allocations for this employee
            alloc_map = get_allocated_hours_for_week(last_date, range_end, db, None)
            leave_map = get_leave_hours_for_employees([task.employee_name], last_date, range_end, db)

            remaining = spillover
            for d in working_days:
                if remaining <= 0:
                    break
                existing = alloc_map.get(task.employee_name, {}).get(d, 0)
                leave_hours = leave_map.get(task.employee_name, {}).get(d, 0)
                available = HOURS_PER_DAY - existing - leave_hours
                if available <= 0:
                    continue
                hours_this_day = min(remaining, available, 8)
                new_allocs.append({"date": d, "hours": hours_this_day})
                remaining -= hours_this_day

        # Delete old allocations and create new ones
        db.query(DevPlannedAllocation).filter(DevPlannedAllocation.task_id == task_id).delete()

        for a in new_allocs:
            db.add(DevPlannedAllocation(task_id=task_id, allocation_date=a["date"], hours=a["hours"]))

        # Update task total and end date
        final_total = sum(a["hours"] for a in new_allocs)
        task.total_planned_hours = round(final_total, 1)
        if new_allocs:
            task.end_date = max(a["date"] for a in new_allocs)

        # If no allocations left, mark task as removed
        if len(new_allocs) == 0:
            task.status = "removed"

        db.commit()
        log_audit(db, task.planning_week_id, "update_allocations", "task", task_id, 
                  {"original_total": current_total}, {"new_total": final_total, "spillover": spillover}, user_name)

        return {
            "success": True,
            "task": {
                "id": task.id,
                "total_planned_hours": task.total_planned_hours,
                "allocations": [{"date": a["date"].isoformat(), "hours": a["hours"]} for a in new_allocs],
            }
        }
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        db.close()


@app.get("/dev-planning/calendar")
def dev_planning_calendar(
    view: str = Query("weekly", description="weekly | monthly"),
    date_str: Optional[str] = Query(None),
    month_str: Optional[str] = Query(None, description="YYYY-MM for monthly"),
    current_user: dict = Depends(get_current_user),
):
    """Calendar view: employees x days with allocated hours and task labels."""
    db = SessionLocal()
    try:
        # Calendar: planning lead sees full Dev dept; else self only
        if is_planning_lead(db, current_user):
            visible = None
        else:
            visible = get_visible_employee_ids(db, current_user)
        employees = get_development_employees(db, visible)
        ref = datetime.strptime(date_str or date.today().isoformat(), "%Y-%m-%d").date() if date_str else date.today()
        week_start, week_end = get_planning_week_dates(ref)
        if view == "monthly" and month_str:
            y, m = map(int, month_str.split("-"))
            month_start = date(y, m, 1)
            month_end = (date(y, m + 1, 1) - timedelta(days=1)) if m < 12 else (date(y + 1, 1, 1) - timedelta(days=1))
        else:
            month_start = week_start
            month_end = week_end

        start_range = month_start
        end_range = month_end
        emp_names = [e.name for e in employees]
        leave_map = get_leave_hours_for_employees(emp_names, start_range, end_range, db)
        alloc_map = get_allocated_hours_for_week(start_range, end_range, db, None)

        tasks_by_emp_date = defaultdict(lambda: defaultdict(lambda: {"hours": 0, "items": []}))
        seen_keys = defaultdict(set)
        tasks = db.query(DevPlannedTask, DevPlannedAllocation).join(
            DevPlannedAllocation, DevPlannedAllocation.task_id == DevPlannedTask.id
        ).filter(
            DevPlannedTask.status == "active",
            DevPlannedAllocation.allocation_date >= start_range,
            DevPlannedAllocation.allocation_date <= end_range,
        ).all()
        ticket_over = {}
        ticket_priority_map = {}
        if tasks:
            ticket_ids = [t.ticket_id for t, _ in tasks if t.ticket_id]
            if ticket_ids:
                tkts = db.query(TicketTracking).filter(TicketTracking.ticket_id.in_(ticket_ids)).all()
                for tk in tkts:
                    est = float(tk.dev_estimate_hours or 0)
                    act = float(tk.actual_dev_hours or 0)
                    ticket_over[tk.ticket_id] = est > 0 and act > est
                    if getattr(tk, "priority", None):
                        ticket_priority_map[tk.ticket_id] = tk.priority
        for t, a in tasks:
            key = (t.ticket_id and f"#{t.ticket_id}") or (t.generic_category or "")
            category = "Ticket" if t.ticket_id else (t.generic_category or "Miscellaneous")
            over_estimate = bool(t.ticket_id and ticket_over.get(t.ticket_id, False))
            ticket_priority = ticket_priority_map.get(t.ticket_id) if t.ticket_id else None
            tasks_by_emp_date[t.employee_name][a.allocation_date]["hours"] += a.hours
            cell_key = (t.employee_name, a.allocation_date)
            if key and key not in seen_keys[cell_key]:
                seen_keys[cell_key].add(key)
                item = {"text": key, "ticket_id": t.ticket_id, "ticket_priority": ticket_priority, "category": category, "over_estimate": over_estimate}
                tasks_by_emp_date[t.employee_name][a.allocation_date]["items"].append(item)

        # Also gather actual timesheet entries (to show plan vs actual for past days)
        times_by_emp_date = defaultdict(lambda: defaultdict(lambda: {"hours": 0, "items": []}))
        try:
            if emp_names:
                ts_entries = db.query(EnhancedTimesheet).filter(
                    EnhancedTimesheet.date >= start_range,
                    EnhancedTimesheet.date <= end_range,
                    EnhancedTimesheet.employee_name.in_(emp_names)
                ).all()
            else:
                ts_entries = []
            for ent in ts_entries:
                name = ent.employee_name
                ddate = ent.date
                hours = ent.productive_hours if ent.productive_hours and ent.productive_hours > 0 else (ent.hours_logged or 0)
                times_by_emp_date[name][ddate]["hours"] += hours
                # Include ticket references where present
                times_by_emp_date[name][ddate]["items"].append({
                    "ticket_id": ent.ticket_id,
                    "hours": hours,
                    "task_description": ent.task_description,
                    "project_name": ent.project_name,
                })
        except Exception:
            # If EnhancedTimesheet table isn't populated or other issues occur, continue without actuals
            times_by_emp_date = defaultdict(lambda: defaultdict(lambda: {"hours": 0, "items": []}))

        working_days_list = get_working_days_list(start_range, end_range, db)
        capacity_hours = len(working_days_list) * 8

        rows = []
        for e in employees:
            alloc_total = sum(alloc_map.get(e.name, {}).values())
            leave_total = sum(leave_map.get(e.name, {}).values())
            total_used = alloc_total + leave_total
            remaining = max(0, capacity_hours - total_used)
            status = "Fully Allocated" if remaining <= 0 else ("Partially Allocated" if alloc_total > 0 else "Available")
            can_manage = can_manage_tasks_for(db, current_user, e.employee_id)

            row = {
                "employee_id": e.employee_id,
                "employee_name": e.name,
                "lead_name": (e.lead or "").strip() or None,
                "allocated_hours": round(alloc_total, 1),
                "leave_hours": round(leave_total, 1),
                "remaining_hours": round(remaining, 1),
                "allocation_status": status,
                "can_manage_tasks": can_manage,
                "days": {},
            }
            d = start_range
            while d <= end_range:
                leave_h = leave_map.get(e.name, {}).get(d, 0)
                alloc_h = alloc_map.get(e.name, {}).get(d, 0)
                total = leave_h + alloc_h
                # Planned hours and items
                cell = {"hours": round(alloc_h, 1), "leave_hours": round(leave_h, 1), "total": round(total, 1)}
                cell["items"] = tasks_by_emp_date[e.name][d]["items"][:5]
                # Actual (timesheet) hours and ticket items for the date (if available)
                actual_h = times_by_emp_date.get(e.name, {}).get(d, {}).get("hours", 0)
                actual_items = times_by_emp_date.get(e.name, {}).get(d, {}).get("items", [])[:5]
                cell["actual_hours"] = round(actual_h, 1)
                cell["actual_items"] = actual_items
                row["days"][d.isoformat()] = cell
                d += timedelta(days=1)
            rows.append(row)
        return {"view": view, "start": start_range.isoformat(), "end": end_range.isoformat(), "employees": rows}
    finally:
        db.close()


@app.get("/dev-planning/day-details")
def dev_planning_day_details(
    employee_name: str = Query(...),
    date_str: str = Query(...),
):
    """Get detailed task list for an employee on a specific date."""
    db = SessionLocal()
    try:
        alloc_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        allocs = db.query(DevPlannedTask, DevPlannedAllocation).join(
            DevPlannedAllocation, DevPlannedAllocation.task_id == DevPlannedTask.id
        ).filter(
            DevPlannedTask.status == "active",
            DevPlannedTask.employee_name == employee_name,
            DevPlannedAllocation.allocation_date == alloc_date,
        ).all()
        tasks = []
        ticket_ids = [t.ticket_id for t, _ in allocs if t.ticket_id]
        ticket_info = {}
        if ticket_ids:
            tkts = db.query(TicketTracking).filter(TicketTracking.ticket_id.in_(ticket_ids)).all()
            for tk in tkts:
                est = float(tk.dev_estimate_hours or 0)
                act = float(tk.actual_dev_hours or 0)
                ticket_info[tk.ticket_id] = {
                    "dev_estimate_hours": est,
                    "actual_dev_hours": act,
                    "remaining_dev_hours": (est - act) if est else None,
                    "over_estimate": est > 0 and act > est,
                    "ticket_priority": getattr(tk, "priority", None),
                }
        for t, a in allocs:
            ti = ticket_info.get(t.ticket_id, {}) if t.ticket_id else {}
            tasks.append({
                "task_id": t.id,
                "ticket_id": t.ticket_id,
                "ticket_priority": ti.get("ticket_priority"),
                "generic_category": t.generic_category,
                "activity_description": t.activity_description,
                "hours": round(a.hours, 1),
                "total_planned_hours": t.total_planned_hours,
                "category": "Ticket" if t.ticket_id else (t.generic_category or "Miscellaneous"),
                "over_estimate": ti.get("over_estimate", False),
                "dev_estimate_hours": ti.get("dev_estimate_hours"),
                "actual_dev_hours": ti.get("actual_dev_hours"),
                "remaining_dev_hours": ti.get("remaining_dev_hours"),
                "is_planned": True,
            })
        
        # Also get actual timesheet entries for the date and employee
        actual_entries = []
        try:
            ts_data = db.query(EnhancedTimesheet).filter(
                EnhancedTimesheet.employee_name == employee_name,
                EnhancedTimesheet.date == alloc_date,
            ).all()
            for ts in ts_data:
                hours = ts.productive_hours if ts.productive_hours and ts.productive_hours > 0 else (ts.hours_logged or 0)
                actual_entries.append({
                    "ticket_id": ts.ticket_id,
                    "hours": round(hours, 1),
                    "task_description": ts.task_description,
                    "project_name": ts.project_name,
                    "is_actual": True,
                })
        except Exception:
            pass
        
        all_tasks = tasks + actual_entries
        return {"employee_name": employee_name, "date": date_str, "tasks": all_tasks}
    finally:
        db.close()


@app.post("/dev-planning/week/{week_start_str}/copy-from/{source_week_str}")
def dev_planning_copy_week(
    week_start_str: str,
    source_week_str: str,
    current_user: dict = Depends(get_current_user),
):
    """Copy previous week's tasks into this week (Draft only). Only Manager/Lead."""
    db = SessionLocal()
    try:
        role = current_user.get("role", "")
        if not _planning_can_edit(role):
            raise HTTPException(status_code=403, detail="Only Manager or Lead can copy plans")
        user_name = _get_user_display_name(db, current_user)
        week_start = datetime.strptime(week_start_str, "%Y-%m-%d").date()
        source_start = datetime.strptime(source_week_str, "%Y-%m-%d").date()
        if week_start.weekday() != 0 or source_start.weekday() != 0:
            raise HTTPException(status_code=400, detail="Both dates must be Mondays")
        pw_target = get_or_create_planning_week(week_start, db, user_name)
        if pw_target.state != "draft":
            raise HTTPException(status_code=400, detail="Target week must be in draft to copy")
        pw_source = get_planning_week(source_start, db)
        if not pw_source:
            return {"success": True, "message": "No source week found; nothing to copy"}
        source_tasks = db.query(DevPlannedTask).filter(
            DevPlannedTask.planning_week_id == pw_source.id,
            DevPlannedTask.status == "active",
        ).all()
        # Simple copy: same employees/tasks, new start_date = week_start, same total hours
        delta_days = (week_start - source_start).days
        for t in source_tasks:
            new_task = DevPlannedTask(
                planning_week_id=pw_target.id,
                employee_id=t.employee_id,
                employee_name=t.employee_name,
                ticket_id=t.ticket_id,
                ticket_title=t.ticket_title,
                generic_category=t.generic_category,
                justification=t.justification,
                activity_description=t.activity_description,
                start_date=t.start_date + timedelta(days=delta_days),
                end_date=(t.end_date + timedelta(days=delta_days)) if t.end_date else None,
                allocation_pct=t.allocation_pct,
                total_planned_hours=t.total_planned_hours,
                status="active",
                created_by=user_name,
            )
            db.add(new_task)
            db.flush()
            for a in db.query(DevPlannedAllocation).filter(DevPlannedAllocation.task_id == t.id).all():
                db.add(DevPlannedAllocation(task_id=new_task.id, allocation_date=a.allocation_date + timedelta(days=delta_days), hours=a.hours))
        db.commit()
        return {"success": True, "message": f"Copied {len(source_tasks)} tasks from {source_week_str}"}
    except HTTPException:
        raise
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date. Use YYYY-MM-DD")
    finally:
        db.close()


# ===== PLAN VS ACTUAL COMPARISON API =====

@app.get("/planning/comparison")
def get_plan_vs_actual(
    employee_id: str = Query(None, description="Employee ID (optional, for individual comparison)"),
    team: str = Query("ALL", description="Team: QA, DEV, or ALL"),
    period: str = Query("week", description="Period: week or month"),
    date_str: str = Query(None, description="Reference date (YYYY-MM-DD)")
):
    """
    Get plan vs actual comparison showing planned tasks against actual timesheet entries.
    Returns variance analysis and metrics.
    """
    db = SessionLocal()
    try:
        # Parse date
        if date_str:
            try:
                target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
        else:
            target_date = date.today()
        
        # Calculate period boundaries
        if period == "week":
            start_date = target_date - timedelta(days=target_date.weekday())
            end_date = start_date + timedelta(days=6)
        else:  # month
            start_date = date(target_date.year, target_date.month, 1)
            if start_date.month == 12:
                end_date = date(start_date.year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(start_date.year, start_date.month + 1, 1) - timedelta(days=1)
        
        # Build base queries
        planned_query = db.query(PlannedTask).filter(
            PlannedTask.planned_date >= start_date,
            PlannedTask.planned_date <= end_date
        )
        actual_query = db.query(EnhancedTimesheet).filter(
            EnhancedTimesheet.date >= start_date,
            EnhancedTimesheet.date <= end_date
        )
        
        # Filter by employee if specified
        if employee_id:
            employee = db.query(Employee).filter(Employee.employee_id == employee_id).first()
            if not employee:
                raise HTTPException(status_code=404, detail="Employee not found")
            planned_query = planned_query.filter(PlannedTask.employee_name == employee.name)
            actual_query = actual_query.filter(EnhancedTimesheet.employee_name == employee.name)
        elif team.upper() != "ALL":
            planned_query = planned_query.filter(PlannedTask.team == team.upper())
            actual_query = actual_query.filter(EnhancedTimesheet.team == team.upper())
        
        planned_tasks = planned_query.all()
        actual_entries = actual_query.all()
        
        # Build comparison data by employee
        employee_comparison = defaultdict(lambda: {
            "planned_hours": 0,
            "actual_hours": 0,
            "variance": 0,
            "variance_percent": 0,
            "planned_tasks": [],
            "actual_entries": [],
            "by_ticket": {}
        })
        
        # Process planned tasks
        for task in planned_tasks:
            name = task.employee_name
            employee_comparison[name]["employee_id"] = task.employee_id
            employee_comparison[name]["employee_name"] = name
            employee_comparison[name]["team"] = task.team
            employee_comparison[name]["planned_hours"] += task.planned_hours or 0
            employee_comparison[name]["planned_tasks"].append({
                "id": task.id,
                "ticket_id": task.ticket_id,
                "task_title": task.task_title,
                "planned_hours": task.planned_hours,
                "actual_hours": task.actual_hours,
                "priority": task.priority,
                "status": task.status,
                "date": task.planned_date.isoformat()
            })
            
            # Track by ticket
            ticket = task.ticket_id
            if ticket not in employee_comparison[name]["by_ticket"]:
                employee_comparison[name]["by_ticket"][ticket] = {
                    "planned_hours": 0,
                    "actual_hours": 0
                }
            employee_comparison[name]["by_ticket"][ticket]["planned_hours"] += task.planned_hours or 0
        
        # Process actual entries
        for entry in actual_entries:
            name = entry.employee_name
            if name not in employee_comparison:
                employee_comparison[name]["employee_id"] = entry.employee_id
                employee_comparison[name]["employee_name"] = name
                employee_comparison[name]["team"] = entry.team
            
            employee_comparison[name]["actual_hours"] += entry.hours_logged or 0
            employee_comparison[name]["actual_entries"].append({
                "ticket_id": entry.ticket_id,
                "hours": entry.hours_logged,
                "task_description": entry.task_description,
                "date": entry.date.isoformat()
            })
            
            # Track by ticket
            ticket = entry.ticket_id
            if ticket not in employee_comparison[name]["by_ticket"]:
                employee_comparison[name]["by_ticket"][ticket] = {
                    "planned_hours": 0,
                    "actual_hours": 0
                }
            employee_comparison[name]["by_ticket"][ticket]["actual_hours"] += entry.hours_logged or 0
        
        # Calculate variances
        total_planned = 0
        total_actual = 0
        
        for name, data in employee_comparison.items():
            planned = data["planned_hours"]
            actual = data["actual_hours"]
            variance = actual - planned
            variance_percent = (variance / planned * 100) if planned > 0 else 0
            
            data["variance"] = round(variance, 2)
            data["variance_percent"] = round(variance_percent, 1)
            data["estimation_accuracy"] = round(100 - abs(variance_percent), 1) if planned > 0 else None
            
            # Calculate ticket-level variance
            for ticket, ticket_data in data["by_ticket"].items():
                ticket_planned = ticket_data["planned_hours"]
                ticket_actual = ticket_data["actual_hours"]
                ticket_data["variance"] = round(ticket_actual - ticket_planned, 2)
            
            data["by_ticket"] = dict(data["by_ticket"])
            
            total_planned += planned
            total_actual += actual
        
        # Calculate overall metrics
        overall_variance = total_actual - total_planned
        overall_variance_percent = (overall_variance / total_planned * 100) if total_planned > 0 else 0
        
        return {
            "period": period,
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "team": team,
            "employee_id": employee_id,
            "employees": [dict(v) for v in employee_comparison.values()],
            "summary": {
                "total_planned_hours": round(total_planned, 2),
                "total_actual_hours": round(total_actual, 2),
                "total_variance": round(overall_variance, 2),
                "variance_percent": round(overall_variance_percent, 1),
                "estimation_accuracy": round(100 - abs(overall_variance_percent), 1) if total_planned > 0 else None,
                "over_estimation": overall_variance < 0,
                "employee_count": len(employee_comparison)
            }
        }
    finally:
        db.close()


@app.get("/planning/comparison/planning")
def get_plan_vs_actual_planning(
    team: str = Query(..., description="Team: dev or qa"),
    week_start_str: str = Query(None, description="Week start (YYYY-MM-DD, Monday). Default: current week"),
    employee_name: str = Query(None, description="Filter by employee name (optional)")
):
    """
    Plan vs Actual comparison using Dev/QA planning tables vs EnhancedTimesheet.
    Planned data: DevPlannedTask/QAPlannedTask + allocations.
    Actual data: EnhancedTimesheet (synced from Excel/Google Sheets).
    """
    db = SessionLocal()
    try:
        # Parse week
        if week_start_str:
            try:
                week_start = datetime.strptime(week_start_str, "%Y-%m-%d").date()
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid week_start. Use YYYY-MM-DD")
        else:
            week_start, _ = get_planning_week_dates(date.today())
        week_end = week_start + timedelta(days=4)  # Mon–Fri

        team_lower = team.lower()
        if team_lower not in ("dev", "qa"):
            raise HTTPException(status_code=400, detail="team must be 'dev' or 'qa'")

        # Map team for timesheet: EnhancedTimesheet uses "DEV", "QA"
        timesheet_team = "DEV" if team_lower == "dev" else "QA"

        # Get planned tasks and allocations
        if team_lower == "dev":
            pw = db.query(DevPlanningWeek).filter(DevPlanningWeek.week_start == week_start).first()
            if not pw:
                planned_tasks = []
                alloc_rows = []
            else:
                planned_tasks = db.query(DevPlannedTask).filter(
                    DevPlannedTask.planning_week_id == pw.id,
                    DevPlannedTask.status == "active",
                ).order_by(DevPlannedTask.employee_name, DevPlannedTask.start_date).all()
                alloc_rows = (
                    db.query(DevPlannedTask, DevPlannedAllocation)
                    .join(DevPlannedAllocation, DevPlannedAllocation.task_id == DevPlannedTask.id)
                    .filter(
                        DevPlannedTask.planning_week_id == pw.id,
                        DevPlannedTask.status == "active",
                        DevPlannedAllocation.allocation_date >= week_start,
                        DevPlannedAllocation.allocation_date <= week_end,
                    )
                ).all()
        else:
            pw = db.query(QAPlanningWeek).filter(QAPlanningWeek.week_start == week_start).first()
            if not pw:
                planned_tasks = []
                alloc_rows = []
            else:
                planned_tasks = db.query(QAPlannedTask).filter(
                    QAPlannedTask.planning_week_id == pw.id,
                    QAPlannedTask.status == "active",
                ).order_by(QAPlannedTask.employee_name, QAPlannedTask.start_date).all()
                alloc_rows = (
                    db.query(QAPlannedTask, QAPlannedAllocation)
                    .join(QAPlannedAllocation, QAPlannedAllocation.task_id == QAPlannedTask.id)
                    .filter(
                        QAPlannedTask.planning_week_id == pw.id,
                        QAPlannedTask.status == "active",
                        QAPlannedAllocation.allocation_date >= week_start,
                        QAPlannedAllocation.allocation_date <= week_end,
                    )
                ).all()

        # Build planned hours by employee and by task from allocations
        planned_by_employee = defaultdict(lambda: {"hours": 0, "tasks": []})
        task_hours_agg = defaultdict(float)
        for row in alloc_rows:
            task, alloc = row[0], row[1]
            name = task.employee_name
            if employee_name and name != employee_name:
                continue
            h = float(alloc.hours or 0)
            task_hours_agg[task.id] += h
            planned_by_employee[name]["hours"] += h

        for task in planned_tasks:
            name = task.employee_name
            if employee_name and name != employee_name:
                continue
            tid = task.id
            th = round(task_hours_agg[tid], 2)
            planned_by_employee[name]["tasks"].append({
                "task_id": tid,
                "ticket_id": task.ticket_id,
                "ticket_title": getattr(task, "ticket_title", None) or "",
                "activity_description": getattr(task, "activity_description", None) or "",
                "generic_category": getattr(task, "generic_category", None),
                "planned_hours": th,
            })
        for name in planned_by_employee:
            planned_by_employee[name]["hours"] = round(planned_by_employee[name]["hours"], 2)

        # Get actual timesheet entries (Enhanced + manual TimeSheetEntry for source of truth)
        actual_query = db.query(EnhancedTimesheet).filter(
            EnhancedTimesheet.date >= week_start,
            EnhancedTimesheet.date <= week_end,
            EnhancedTimesheet.team == timesheet_team,
        )
        if employee_name:
            actual_query = actual_query.filter(EnhancedTimesheet.employee_name == employee_name)
        actual_entries = actual_query.order_by(EnhancedTimesheet.employee_name, EnhancedTimesheet.date).all()

        # Also get manual timesheet entries for the same week/team
        if team_lower == "dev":
            active_employees = get_development_employees(db)
        else:
            active_employees = get_qa_employees(db)
        employee_ids_planning = [e.employee_id for e in active_employees if e.employee_id]
        if employee_name:
            emp_lookup_plan = {e.name: e for e in active_employees if e.name}
            employee_ids_planning = [emp_lookup_plan[employee_name].employee_id] if employee_name in emp_lookup_plan else []
        manual_actuals = []
        if employee_ids_planning:
            manual_actuals = db.query(TimeSheetEntry).filter(
                TimeSheetEntry.date >= week_start,
                TimeSheetEntry.date <= week_end,
                TimeSheetEntry.employee_id.in_(employee_ids_planning),
            ).order_by(TimeSheetEntry.employee_name, TimeSheetEntry.date).all()

        # Build actual by employee, by ticket, and by day
        actual_by_employee = defaultdict(lambda: {"hours": 0, "entries": [], "by_ticket": defaultdict(float), "by_day": defaultdict(float)})
        for e in actual_entries:
            name = e.employee_name
            h = float(e.hours_logged or 0)
            actual_by_employee[name]["hours"] += h
            ticket_key = str(e.ticket_id) if e.ticket_id is not None else "(no ticket)"
            actual_by_employee[name]["by_ticket"][ticket_key] += h
            actual_by_employee[name]["by_day"][e.date.isoformat()] += h
            actual_by_employee[name]["entries"].append({
                "source": "sync",
                "ticket_id": e.ticket_id,
                "hours": round(h, 2),
                "task_description": e.task_description,
                "date": e.date.isoformat(),
            })
        for e in manual_actuals:
            name = e.employee_name
            h = float(e.hours or 0)
            actual_by_employee[name]["hours"] += h
            ticket_key = str(e.ticket_id) if e.ticket_id is not None else "(no ticket)"
            actual_by_employee[name]["by_ticket"][ticket_key] += h
            actual_by_employee[name]["by_day"][e.date.isoformat()] += h
            actual_by_employee[name]["entries"].append({
                "source": "manual",
                "ticket_id": e.ticket_id,
                "hours": round(h, 2),
                "task_description": e.description,
                "date": e.date.isoformat(),
                "variance_notes": getattr(e, "variance_notes", None),
                "variance_reason_type": getattr(e, "variance_reason_type", None),
            })
        for name in actual_by_employee:
            actual_by_employee[name]["hours"] = round(actual_by_employee[name]["hours"], 2)
            actual_by_employee[name]["by_ticket"] = {
                k: round(v, 2) for k, v in dict(actual_by_employee[name]["by_ticket"]).items()
            }
            actual_by_employee[name]["by_day"] = {
                k: round(v, 2) for k, v in dict(actual_by_employee[name]["by_day"]).items()
            }

        # Build planned by_day from allocations
        planned_by_day = defaultdict(lambda: defaultdict(float))
        for row in alloc_rows:
            task, alloc = row[0], row[1]
            name = task.employee_name
            if employee_name and name != employee_name:
                continue
            h = float(alloc.hours or 0)
            planned_by_day[name][alloc.allocation_date.isoformat()] += h

        # Get ALL active employees for the team (from Employee table)
        if team_lower == "dev":
            active_employees = get_development_employees(db)
        else:
            active_employees = get_qa_employees(db)
        all_names = {e.name for e in active_employees if e.name}
        if employee_name:
            all_names = {n for n in all_names if n == employee_name}

        employees = []
        total_planned = 0
        total_actual = 0
        employees_with_planned = 0
        employees_with_actual = 0
        employees_with_no_timesheet = 0
        emp_lookup = {e.name: e for e in active_employees if e.name}
        for name in sorted(all_names):
            planned = planned_by_employee[name]["hours"]
            actual = actual_by_employee[name]["hours"]
            variance = round(actual - planned, 2)
            variance_pct = round((variance / planned * 100), 1) if planned > 0 else (0 if actual == 0 else None)
            total_planned += planned
            total_actual += actual
            if planned > 0:
                employees_with_planned += 1
            if actual > 0:
                employees_with_actual += 1
            else:
                employees_with_no_timesheet += 1
            emp_obj = emp_lookup.get(name)
            # Count tasks and tickets
            planned_task_count = len(planned_by_employee[name]["tasks"])
            actual_entries_list = actual_by_employee[name].get("entries", [])
            actual_entry_count = len(actual_entries_list)
            # Count unique tickets from actual entries
            unique_tickets = set(e.get("ticket_id") for e in actual_entries_list if e.get("ticket_id") is not None)
            actual_ticket_count = len(unique_tickets)
            employees.append({
                "employee_id": emp_obj.employee_id if emp_obj else None,
                "employee_name": name,
                "role": emp_obj.role if emp_obj else None,
                "lead": emp_obj.lead if emp_obj else None,
                "planned_hours": planned,
                "actual_hours": actual,
                "variance": variance,
                "variance_percent": variance_pct,
                "planned_task_count": planned_task_count,
                "actual_entry_count": actual_entry_count,
                "actual_ticket_count": actual_ticket_count,
                "planned_tasks": planned_by_employee[name]["tasks"],
                "actual_entries": actual_entries_list,
                "by_ticket": actual_by_employee[name].get("by_ticket", {}),
                "by_day_planned": {k: round(v, 2) for k, v in dict(planned_by_day[name]).items()},
                "by_day_actual": actual_by_employee[name].get("by_day", {}),
            })

        overall_variance = round(total_actual - total_planned, 2)
        overall_variance_pct = round((overall_variance / total_planned * 100), 1) if total_planned > 0 else 0

        # Build by_day summary for the week (Mon–Fri)
        week_days = [(week_start + timedelta(days=i)).isoformat() for i in range(5)]
        by_day_summary = []
        for d in week_days:
            day_planned = sum(planned_by_day[n].get(d, 0) for n in all_names)
            day_actual = sum(actual_by_employee[n].get("by_day", {}).get(d, 0) for n in all_names)
            by_day_summary.append({
                "date": d,
                "planned_hours": round(day_planned, 2),
                "actual_hours": round(day_actual, 2),
                "variance": round(day_actual - day_planned, 2),
            })

        # Build daily_view: per-day planned tasks and actual entries with full details
        planned_by_day_detail = defaultdict(list)
        for row in alloc_rows:
            task, alloc = row[0], row[1]
            name = task.employee_name
            if employee_name and name != employee_name:
                continue
            d_str = alloc.allocation_date.isoformat()
            h = round(float(alloc.hours or 0), 2)
            planned_by_day_detail[d_str].append({
                "employee_name": name,
                "task_id": task.id,
                "ticket_id": task.ticket_id,
                "ticket_title": getattr(task, "ticket_title", None) or "",
                "activity_description": getattr(task, "activity_description", None) or "",
                "generic_category": getattr(task, "generic_category", None),
                "hours": h,
            })

        actual_by_day_detail = defaultdict(list)
        for e in actual_entries:
            name = e.employee_name
            if employee_name and name != employee_name:
                continue
            d_str = e.date.isoformat()
            h = round(float(e.hours_logged or 0), 2)
            actual_by_day_detail[d_str].append({
                "employee_name": name,
                "ticket_id": e.ticket_id,
                "task_description": e.task_description or "",
                "project_name": e.project_name or "",
                "hours": h,
            })

        daily_view = []
        capacity_per_day = len(all_names) * 8
        day_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
        for idx, d in enumerate(week_days):
            day_planned = sum(planned_by_day[n].get(d, 0) for n in all_names)
            day_actual = sum(actual_by_employee[n].get("by_day", {}).get(d, 0) for n in all_names)
            
            # Build employee breakdown for this day
            employee_breakdown = []
            for name in sorted(all_names):
                emp_planned = planned_by_day[name].get(d, 0)
                emp_actual = actual_by_employee[name].get("by_day", {}).get(d, 0)
                if emp_planned > 0 or emp_actual > 0:
                    emp_planned_tasks = [t for t in planned_by_day_detail.get(d, []) if t["employee_name"] == name]
                    emp_actual_entries = [e for e in actual_by_day_detail.get(d, []) if e["employee_name"] == name]
                    employee_breakdown.append({
                        "employee_name": name,
                        "planned_hours": round(emp_planned, 2),
                        "actual_hours": round(emp_actual, 2),
                        "variance": round(emp_actual - emp_planned, 2),
                        "planned_tasks": emp_planned_tasks,
                        "actual_entries": emp_actual_entries,
                    })
            
            daily_view.append({
                "date": d,
                "day_name": day_names[idx] if idx < len(day_names) else "",
                "total_planned": round(day_planned, 2),
                "total_actual": round(day_actual, 2),
                "total_available": capacity_per_day,
                "variance": round(day_actual - day_planned, 2),
                "planned_tasks": sorted(planned_by_day_detail.get(d, []), key=lambda x: (x["employee_name"], x.get("ticket_id") or 0)),
                "actual_entries": sorted(actual_by_day_detail.get(d, []), key=lambda x: (x["employee_name"], x.get("ticket_id") or "")),
                "employee_breakdown": employee_breakdown,
            })

        return {
            "team": team_lower,
            "week_start": week_start.isoformat(),
            "week_end": week_end.isoformat(),
            "employees": employees,
            "by_day_summary": by_day_summary,
            "daily_view": daily_view,
            "summary": {
                "total_planned_hours": round(total_planned, 2),
                "total_actual_hours": round(total_actual, 2),
                "total_variance": overall_variance,
                "variance_percent": overall_variance_pct,
                "estimation_accuracy": round(100 - abs(overall_variance_pct), 1) if total_planned > 0 else None,
                "employee_count": len(employees),
                "employees_with_planned": employees_with_planned,
                "employees_with_actual": employees_with_actual,
                "employees_with_no_timesheet": employees_with_no_timesheet,
                "capacity_hours": len(all_names) * 40,
            },
        }
    finally:
        db.close()


@app.get("/planning/comparison/monthly")
def get_plan_vs_actual_monthly(
    team: str = Query(..., description="Team: dev or qa"),
    month_str: str = Query(None, description="Month (YYYY-MM). Default: current month"),
    employee_name: str = Query(None, description="Filter by employee name (optional)")
):
    """
    Plan vs Actual comparison for a full month.
    Returns daily breakdown with employee-level details for each day.
    """
    db = SessionLocal()
    try:
        # Parse month
        if month_str:
            try:
                month_date = datetime.strptime(month_str + "-01", "%Y-%m-%d").date()
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid month. Use YYYY-MM")
        else:
            month_date = date.today().replace(day=1)
        
        # Get month start and end
        month_start = month_date.replace(day=1)
        if month_date.month == 12:
            month_end = month_date.replace(year=month_date.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            month_end = month_date.replace(month=month_date.month + 1, day=1) - timedelta(days=1)
        
        team_lower = team.lower()
        if team_lower not in ("dev", "qa"):
            raise HTTPException(status_code=400, detail="team must be 'dev' or 'qa'")
        
        timesheet_team = "DEV" if team_lower == "dev" else "QA"
        
        # Get all planning weeks that overlap with this month
        all_alloc_rows = []
        all_planned_tasks = []
        
        # Find all weeks in this month
        current = month_start
        weeks_in_month = []
        while current <= month_end:
            week_start = current - timedelta(days=current.weekday())  # Monday
            if week_start not in [w[0] for w in weeks_in_month]:
                week_end = week_start + timedelta(days=4)  # Friday
                weeks_in_month.append((week_start, week_end))
            current += timedelta(days=7)
        
        # Get planned data from all weeks
        for week_start, week_end in weeks_in_month:
            if team_lower == "dev":
                pw = db.query(DevPlanningWeek).filter(DevPlanningWeek.week_start == week_start).first()
                if pw:
                    tasks = db.query(DevPlannedTask).filter(
                        DevPlannedTask.planning_week_id == pw.id,
                        DevPlannedTask.status == "active",
                    ).all()
                    all_planned_tasks.extend(tasks)
                    rows = (
                        db.query(DevPlannedTask, DevPlannedAllocation)
                        .join(DevPlannedAllocation, DevPlannedAllocation.task_id == DevPlannedTask.id)
                        .filter(
                            DevPlannedTask.planning_week_id == pw.id,
                            DevPlannedTask.status == "active",
                            DevPlannedAllocation.allocation_date >= month_start,
                            DevPlannedAllocation.allocation_date <= month_end,
                        )
                    ).all()
                    all_alloc_rows.extend(rows)
            else:
                pw = db.query(QAPlanningWeek).filter(QAPlanningWeek.week_start == week_start).first()
                if pw:
                    tasks = db.query(QAPlannedTask).filter(
                        QAPlannedTask.planning_week_id == pw.id,
                        QAPlannedTask.status == "active",
                    ).all()
                    all_planned_tasks.extend(tasks)
                    rows = (
                        db.query(QAPlannedTask, QAPlannedAllocation)
                        .join(QAPlannedAllocation, QAPlannedAllocation.task_id == QAPlannedTask.id)
                        .filter(
                            QAPlannedTask.planning_week_id == pw.id,
                            QAPlannedTask.status == "active",
                            QAPlannedAllocation.allocation_date >= month_start,
                            QAPlannedAllocation.allocation_date <= month_end,
                        )
                    ).all()
                    all_alloc_rows.extend(rows)
        
        # Build planned by employee and by day
        planned_by_employee = defaultdict(lambda: {"hours": 0, "tasks": [], "by_day": defaultdict(float)})
        task_hours_agg = defaultdict(float)
        planned_by_day_detail = defaultdict(list)
        
        for row in all_alloc_rows:
            task, alloc = row[0], row[1]
            name = task.employee_name
            if employee_name and name != employee_name:
                continue
            h = float(alloc.hours or 0)
            d_str = alloc.allocation_date.isoformat()
            task_hours_agg[task.id] += h
            planned_by_employee[name]["hours"] += h
            planned_by_employee[name]["by_day"][d_str] += h
            planned_by_day_detail[d_str].append({
                "employee_name": name,
                "task_id": task.id,
                "ticket_id": task.ticket_id,
                "ticket_title": getattr(task, "ticket_title", None) or "",
                "activity_description": getattr(task, "activity_description", None) or "",
                "generic_category": getattr(task, "generic_category", None),
                "hours": round(h, 2),
            })
        
        # Add task details to employee
        seen_tasks = set()
        for task in all_planned_tasks:
            name = task.employee_name
            if employee_name and name != employee_name:
                continue
            tid = task.id
            if tid in seen_tasks:
                continue
            seen_tasks.add(tid)
            th = round(task_hours_agg.get(tid, 0), 2)
            if th > 0:
                planned_by_employee[name]["tasks"].append({
                    "task_id": tid,
                    "ticket_id": task.ticket_id,
                    "ticket_title": getattr(task, "ticket_title", None) or "",
                    "activity_description": getattr(task, "activity_description", None) or "",
                    "generic_category": getattr(task, "generic_category", None),
                    "planned_hours": th,
                })
        
        for name in planned_by_employee:
            planned_by_employee[name]["hours"] = round(planned_by_employee[name]["hours"], 2)
            planned_by_employee[name]["by_day"] = {k: round(v, 2) for k, v in dict(planned_by_employee[name]["by_day"]).items()}
        
        # Get actual timesheet entries for the month
        actual_query = db.query(EnhancedTimesheet).filter(
            EnhancedTimesheet.date >= month_start,
            EnhancedTimesheet.date <= month_end,
            EnhancedTimesheet.team == timesheet_team,
        )
        if employee_name:
            actual_query = actual_query.filter(EnhancedTimesheet.employee_name == employee_name)
        actual_entries = actual_query.order_by(EnhancedTimesheet.employee_name, EnhancedTimesheet.date).all()
        
        # Build actual by employee and by day
        actual_by_employee = defaultdict(lambda: {"hours": 0, "entries": [], "by_day": defaultdict(float)})
        actual_by_day_detail = defaultdict(list)
        
        for e in actual_entries:
            name = e.employee_name
            h = float(e.hours_logged or 0)
            d_str = e.date.isoformat()
            actual_by_employee[name]["hours"] += h
            actual_by_employee[name]["by_day"][d_str] += h
            actual_by_employee[name]["entries"].append({
                "ticket_id": e.ticket_id,
                "hours": round(h, 2),
                "task_description": e.task_description,
                "date": d_str,
            })
            actual_by_day_detail[d_str].append({
                "employee_name": name,
                "ticket_id": e.ticket_id,
                "task_description": e.task_description or "",
                "project_name": e.project_name or "",
                "hours": round(h, 2),
            })
        
        for name in actual_by_employee:
            actual_by_employee[name]["hours"] = round(actual_by_employee[name]["hours"], 2)
            actual_by_employee[name]["by_day"] = {k: round(v, 2) for k, v in dict(actual_by_employee[name]["by_day"]).items()}
        
        # Get active employees
        if team_lower == "dev":
            active_employees = get_development_employees(db)
        else:
            active_employees = get_qa_employees(db)
        all_names = {e.name for e in active_employees if e.name}
        if employee_name:
            all_names = {n for n in all_names if n == employee_name}
        
        # Build employee comparison
        employees = []
        total_planned = 0
        total_actual = 0
        employees_with_actual = 0
        employees_with_no_timesheet = 0
        emp_lookup = {e.name: e for e in active_employees if e.name}
        
        for name in sorted(all_names):
            planned = planned_by_employee[name]["hours"]
            actual = actual_by_employee[name]["hours"]
            variance = round(actual - planned, 2)
            variance_pct = round((variance / planned * 100), 1) if planned > 0 else (0 if actual == 0 else None)
            total_planned += planned
            total_actual += actual
            if actual > 0:
                employees_with_actual += 1
            else:
                employees_with_no_timesheet += 1
            emp_obj = emp_lookup.get(name)
            # Count tasks and tickets
            planned_task_count = len(planned_by_employee[name]["tasks"])
            actual_entries_list = actual_by_employee[name].get("entries", [])
            actual_entry_count = len(actual_entries_list)
            # Count unique tickets from actual entries
            unique_tickets = set(e.get("ticket_id") for e in actual_entries_list if e.get("ticket_id") is not None)
            actual_ticket_count = len(unique_tickets)
            employees.append({
                "employee_id": emp_obj.employee_id if emp_obj else None,
                "employee_name": name,
                "role": emp_obj.role if emp_obj else None,
                "lead": emp_obj.lead if emp_obj else None,
                "planned_hours": planned,
                "actual_hours": actual,
                "variance": variance,
                "variance_percent": variance_pct,
                "planned_task_count": planned_task_count,
                "actual_entry_count": actual_entry_count,
                "actual_ticket_count": actual_ticket_count,
                "planned_tasks": planned_by_employee[name]["tasks"],
                "actual_entries": actual_entries_list,
                "by_day_planned": planned_by_employee[name].get("by_day", {}),
                "by_day_actual": actual_by_employee[name].get("by_day", {}),
            })
        
        overall_variance = round(total_actual - total_planned, 2)
        overall_variance_pct = round((overall_variance / total_planned * 100), 1) if total_planned > 0 else 0
        
        # Build daily view for all days in month (excluding weekends)
        daily_view = []
        current_date = month_start
        while current_date <= month_end:
            if current_date.weekday() < 5:  # Mon-Fri only
                d_str = current_date.isoformat()
                day_planned = sum(planned_by_employee[n].get("by_day", {}).get(d_str, 0) for n in all_names)
                day_actual = sum(actual_by_employee[n].get("by_day", {}).get(d_str, 0) for n in all_names)
                
                # Build employee breakdown for this day
                employee_breakdown = []
                for name in sorted(all_names):
                    emp_planned = planned_by_employee[name].get("by_day", {}).get(d_str, 0)
                    emp_actual = actual_by_employee[name].get("by_day", {}).get(d_str, 0)
                    if emp_planned > 0 or emp_actual > 0:
                        emp_planned_tasks = [t for t in planned_by_day_detail.get(d_str, []) if t["employee_name"] == name]
                        emp_actual_entries = [e for e in actual_by_day_detail.get(d_str, []) if e["employee_name"] == name]
                        employee_breakdown.append({
                            "employee_name": name,
                            "planned_hours": round(emp_planned, 2),
                            "actual_hours": round(emp_actual, 2),
                            "variance": round(emp_actual - emp_planned, 2),
                            "planned_tasks": emp_planned_tasks,
                            "actual_entries": emp_actual_entries,
                        })
                
                daily_view.append({
                    "date": d_str,
                    "day_name": current_date.strftime("%A"),
                    "total_planned": round(day_planned, 2),
                    "total_actual": round(day_actual, 2),
                    "total_available": len(all_names) * 8,
                    "variance": round(day_actual - day_planned, 2),
                    "planned_tasks": sorted(planned_by_day_detail.get(d_str, []), key=lambda x: (x["employee_name"], x.get("ticket_id") or 0)),
                    "actual_entries": sorted(actual_by_day_detail.get(d_str, []), key=lambda x: (x["employee_name"], x.get("ticket_id") or "")),
                    "employee_breakdown": employee_breakdown,
                })
            current_date += timedelta(days=1)
        
        # Group days by week for weekly summary
        weekly_summary = []
        weeks_grouped = defaultdict(list)
        for day in daily_view:
            d = datetime.strptime(day["date"], "%Y-%m-%d").date()
            week_start = d - timedelta(days=d.weekday())
            weeks_grouped[week_start.isoformat()].append(day)
        
        for week_start_str in sorted(weeks_grouped.keys()):
            days = weeks_grouped[week_start_str]
            week_planned = sum(d["total_planned"] for d in days)
            week_actual = sum(d["total_actual"] for d in days)
            weekly_summary.append({
                "week_start": week_start_str,
                "total_planned": round(week_planned, 2),
                "total_actual": round(week_actual, 2),
                "variance": round(week_actual - week_planned, 2),
                "days": days,
            })
        
        return {
            "team": team_lower,
            "month": month_date.strftime("%Y-%m"),
            "month_name": month_date.strftime("%B %Y"),
            "month_start": month_start.isoformat(),
            "month_end": month_end.isoformat(),
            "employees": employees,
            "daily_view": daily_view,
            "weekly_summary": weekly_summary,
            "summary": {
                "total_planned_hours": round(total_planned, 2),
                "total_actual_hours": round(total_actual, 2),
                "total_variance": overall_variance,
                "variance_percent": overall_variance_pct,
                "estimation_accuracy": round(100 - abs(overall_variance_pct), 1) if total_planned > 0 else None,
                "employee_count": len(employees),
                "employees_with_actual": employees_with_actual,
                "employees_with_no_timesheet": employees_with_no_timesheet,
                "working_days": len(daily_view),
            },
        }
    finally:
        db.close()


@app.get("/planning/comparison/trends")
def get_comparison_trends(
    team: str = Query("ALL", description="Team: QA, DEV, or ALL"),
    weeks: int = Query(4, description="Number of weeks to analyze")
):
    """
    Get historical trends for plan vs actual comparison.
    Shows estimation accuracy over time.
    """
    db = SessionLocal()
    try:
        today = date.today()
        trends = []
        
        for i in range(weeks):
            # Calculate week boundaries
            week_start = today - timedelta(days=today.weekday() + (i * 7))
            week_end = week_start + timedelta(days=6)
            
            # Query planned tasks for this week
            planned_query = db.query(func.sum(PlannedTask.planned_hours)).filter(
                PlannedTask.planned_date >= week_start,
                PlannedTask.planned_date <= week_end
            )
            
            # Query actual entries for this week
            actual_query = db.query(func.sum(EnhancedTimesheet.hours_logged)).filter(
                EnhancedTimesheet.date >= week_start,
                EnhancedTimesheet.date <= week_end
            )
            
            if team.upper() != "ALL":
                planned_query = planned_query.filter(PlannedTask.team == team.upper())
                actual_query = actual_query.filter(EnhancedTimesheet.team == team.upper())
            
            planned_hours = planned_query.scalar() or 0
            actual_hours = actual_query.scalar() or 0
            
            variance = actual_hours - planned_hours
            variance_percent = (variance / planned_hours * 100) if planned_hours > 0 else 0
            accuracy = 100 - abs(variance_percent) if planned_hours > 0 else None
            
            trends.append({
                "week_start": week_start.isoformat(),
                "week_end": week_end.isoformat(),
                "week_number": week_start.isocalendar()[1],
                "planned_hours": round(float(planned_hours), 2),
                "actual_hours": round(float(actual_hours), 2),
                "variance": round(float(variance), 2),
                "variance_percent": round(float(variance_percent), 1),
                "estimation_accuracy": round(accuracy, 1) if accuracy else None
            })
        
        # Reverse to show oldest first
        trends.reverse()
        
        # Calculate average accuracy
        accuracies = [t["estimation_accuracy"] for t in trends if t["estimation_accuracy"] is not None]
        avg_accuracy = sum(accuracies) / len(accuracies) if accuracies else None
        
        return {
            "team": team,
            "weeks_analyzed": weeks,
            "trends": trends,
            "summary": {
                "average_accuracy": round(avg_accuracy, 1) if avg_accuracy else None,
                "best_week": max(trends, key=lambda x: x["estimation_accuracy"] or 0) if trends else None,
                "worst_week": min(trends, key=lambda x: x["estimation_accuracy"] or 100) if trends else None
            }
        }
    finally:
        db.close()


# ===== EMPLOYEE NAME MAPPING ENDPOINTS =====

@app.get("/employee-mappings")
def get_employee_name_mappings():
    """Get all employee name mappings."""
    db = SessionLocal()
    try:
        mappings = db.query(EmployeeNameMapping).filter(
            EmployeeNameMapping.is_active == True
        ).all()
        
        return {
            "mappings": [
                {
                    "id": m.id,
                    "alternate_name": m.alternate_name,
                    "canonical_name": m.canonical_name,
                    "employee_id": m.employee_id,
                    "source": m.source,
                    "notes": m.notes
                }
                for m in mappings
            ]
        }
    finally:
        db.close()


@app.get("/employee-mappings/unmatched")
def get_unmatched_employee_names():
    """Get names in timesheets that don't have a matching Employee record."""
    db = SessionLocal()
    try:
        from sqlalchemy import func
        
        # Get all employee names
        employees = db.query(Employee.name).all()
        emp_names = set(e[0] for e in employees)
        
        # Get all timesheet names with counts
        ts_names = db.query(
            EnhancedTimesheet.employee_name,
            func.count(EnhancedTimesheet.id).label('entry_count'),
            func.min(EnhancedTimesheet.date).label('min_date'),
            func.max(EnhancedTimesheet.date).label('max_date')
        ).group_by(EnhancedTimesheet.employee_name).all()
        
        # Find unmatched
        unmatched = []
        for name, count, min_date, max_date in ts_names:
            if name not in emp_names:
                # Get team info
                sample = db.query(EnhancedTimesheet.team).filter(
                    EnhancedTimesheet.employee_name == name
                ).first()
                
                unmatched.append({
                    "name": name,
                    "entry_count": count,
                    "date_range": f"{min_date} to {max_date}",
                    "team": sample[0] if sample else None
                })
        
        return {
            "unmatched_count": len(unmatched),
            "unmatched": sorted(unmatched, key=lambda x: x["name"])
        }
    finally:
        db.close()


class NameMappingCreate(BaseModel):
    alternate_name: str
    canonical_name: str
    employee_id: Optional[str] = None
    notes: Optional[str] = None


@app.post("/employee-mappings")
def create_employee_name_mapping(mapping: NameMappingCreate):
    """Create a new employee name mapping and update existing records."""
    db = SessionLocal()
    try:
        # Check if mapping already exists
        existing = db.query(EmployeeNameMapping).filter(
            EmployeeNameMapping.alternate_name == mapping.alternate_name
        ).first()
        
        if existing:
            raise HTTPException(status_code=400, detail="Mapping already exists for this name")
        
        # Find employee ID if not provided
        emp_id = mapping.employee_id
        if not emp_id:
            emp = db.query(Employee).filter(Employee.name == mapping.canonical_name).first()
            emp_id = emp.employee_id if emp else None
        
        # Create mapping
        new_mapping = EmployeeNameMapping(
            alternate_name=mapping.alternate_name,
            canonical_name=mapping.canonical_name,
            employee_id=emp_id,
            source='api',
            notes=mapping.notes
        )
        db.add(new_mapping)
        
        # Update existing timesheet entries
        ts_count = db.query(EnhancedTimesheet).filter(
            EnhancedTimesheet.employee_name == mapping.alternate_name
        ).update({
            'employee_name': mapping.canonical_name,
            'employee_id': emp_id
        }, synchronize_session=False)
        
        # Update leave entries
        leave_count = db.query(LeaveEntry).filter(
            LeaveEntry.employee_name == mapping.alternate_name
        ).update({
            'employee_name': mapping.canonical_name,
            'employee_id': emp_id
        }, synchronize_session=False)
        
        db.commit()
        
        return {
            "success": True,
            "message": f"Mapping created. Updated {ts_count} timesheets and {leave_count} leave entries.",
            "mapping": {
                "alternate_name": mapping.alternate_name,
                "canonical_name": mapping.canonical_name,
                "employee_id": emp_id
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


@app.delete("/employee-mappings/{mapping_id}")
def delete_employee_name_mapping(mapping_id: int):
    """Deactivate an employee name mapping."""
    db = SessionLocal()
    try:
        mapping = db.query(EmployeeNameMapping).filter(
            EmployeeNameMapping.id == mapping_id
        ).first()
        
        if not mapping:
            raise HTTPException(status_code=404, detail="Mapping not found")
        
        mapping.is_active = False
        db.commit()
        
        return {"success": True, "message": "Mapping deactivated"}
    finally:
        db.close()