"""Generate QA Weekly and Monthly Excel reports with comparison data."""
import json, os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import Counter, defaultdict
from datetime import date, timedelta

try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(__file__), '.env'))
except ImportError:
    pass

DATA_FILE = os.path.join(os.path.dirname(__file__), 'data', 'pm_tickets_cache.json')
REPORTS_DIR = os.path.join(os.path.dirname(__file__), 'reports')

with open(DATA_FILE) as f:
    tickets = json.load(f)

# Parse custom date range from args: --start=YYYY-MM-DD --end=YYYY-MM-DD
custom_start = None
custom_end = None
for arg in sys.argv[1:]:
    if arg.startswith('--start='):
        custom_start = date.fromisoformat(arg.split('=', 1)[1])
    elif arg.startswith('--end='):
        custom_end = date.fromisoformat(arg.split('=', 1)[1])

today = date.today()
if custom_start and custom_end:
    # Custom date range — use as the "current" period, previous period is same length before it
    period_days = (custom_end - custom_start).days
    d7 = custom_start  # current period start
    d14 = custom_start - timedelta(days=period_days)  # previous period start
    d30 = custom_start  # reuse for monthly too
    d60 = custom_start - timedelta(days=period_days)
    today = custom_end  # report end date
else:
    d7 = today - timedelta(days=7)
    d14 = today - timedelta(days=14)
    d30 = today - timedelta(days=30)
    d60 = today - timedelta(days=60)

def parse_d(v):
    if not v: return None
    try: return date.fromisoformat(str(v)[:10])
    except: return None

def sum_h(tl, f): return round(sum(t.get(f) or 0 for t in tl), 1)

qa_tickets = [t for t in tickets if t.get('qc_tester')]
QC = ('QC Testing', 'QC Testing in Progress', 'QC Testing Hold')
in_qc = [t for t in tickets if t['status'] in QC]
in_bis = [t for t in tickets if t['status'] == 'BIS Testing']
in_appr = [t for t in tickets if t['status'] == 'Approved for Live']
qc_fail = [t for t in tickets if t['status'] == 'QC Review Fail']

# Styles
hf = Font(bold=True, color='FFFFFF', size=11)
sf = Font(bold=True, size=12)
tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
green_fill = PatternFill(start_color='1B5E20', end_color='1B5E20', fill_type='solid')
blue_fill = PatternFill(start_color='0D47A1', end_color='0D47A1', fill_type='solid')
purple_fill = PatternFill(start_color='6A1B9A', end_color='6A1B9A', fill_type='solid')
orange_fill = PatternFill(start_color='E65100', end_color='E65100', fill_type='solid')
red_fill = PatternFill(start_color='B71C1C', end_color='B71C1C', fill_type='solid')
light_green = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
light_red = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
light_blue = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
ca = Alignment(horizontal='center', vertical='center')

def write_header(ws, row, headers, fill):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = hf; cell.fill = fill; cell.alignment = ca; cell.border = tb

def write_row(ws, row, vals):
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.border = tb
        cell.alignment = Alignment(horizontal='center' if isinstance(v, (int, float)) else 'left')

def write_section(ws, row, title):
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = sf; cell.fill = light_blue
    return row + 1

def write_comparison(ws, r, label, curr, prev):
    diff = curr - prev
    sign = '+' if diff >= 0 else ''
    write_row(ws, r, [label, curr, prev, f'{sign}{round(diff, 1)}'])
    if diff > 0: ws.cell(row=r, column=4).fill = light_green
    elif diff < 0: ws.cell(row=r, column=4).fill = light_red

def write_pipeline(ws, r):
    write_header(ws, r, ['Status', 'Count', 'Est Hours', 'Actual Hours', 'Remaining Hours'], purple_fill)
    r += 1
    for label, tl in [
        ('QC Testing (Unassigned)', [t for t in in_qc if t['status'] == 'QC Testing' and not t.get('qc_tester')]),
        ('QC Testing (Assigned)', [t for t in in_qc if t['status'] == 'QC Testing' and t.get('qc_tester')]),
        ('QC Testing in Progress', [t for t in in_qc if t['status'] == 'QC Testing in Progress']),
        ('QC Testing Hold', [t for t in in_qc if t['status'] == 'QC Testing Hold']),
        ('QC Review Fail', qc_fail),
        ('BIS Testing', in_bis),
        ('Approved for Live', in_appr),
    ]:
        est = sum_h(tl, 'qa_estimate_hours'); act = sum_h(tl, 'qa_actual_hours')
        write_row(ws, r, [label, len(tl), est, act, round(max(0, est - act), 1)])
        r += 1
    all_pipeline = in_qc + qc_fail + in_bis + in_appr
    est = sum_h(all_pipeline, 'qa_estimate_hours'); act = sum_h(all_pipeline, 'qa_actual_hours')
    write_row(ws, r, ['TOTAL', len(all_pipeline), est, act, round(max(0, est - act), 1)])
    ws.cell(row=r, column=1).font = Font(bold=True)
    return r + 1

def write_member_sheet(ws, current_list, prev_list, member_active):
    write_header(ws, 1, ['QA Member', 'Closed (Current)', 'Closed (Previous)', 'Change',
                         'Est Hours', 'Actual Hours', 'Efficiency %', 'Overrun Tickets',
                         'Active Tickets', 'Top Module'], orange_fill)
    mc = defaultdict(list); mp = defaultdict(list)
    for t in current_list:
        for n in (x.strip() for x in (t.get('qc_tester') or '').split(',') if x.strip()): mc[n].append(t)
    for t in prev_list:
        for n in (x.strip() for x in (t.get('qc_tester') or '').split(',') if x.strip()): mp[n].append(t)
    # Sort by: most active tickets desc, then most closed desc
    all_names = sorted(set(list(mc.keys()) + list(mp.keys()) + list(member_active.keys())),
                       key=lambda n: (-member_active.get(n, 0), -len(mc.get(n, []))))
    for r, name in enumerate(all_names, 2):
        tl = mc.get(name, []); c_cnt = len(tl); p_cnt = len(mp.get(name, []))
        est = sum_h(tl, 'qa_estimate_hours'); act = sum_h(tl, 'qa_actual_hours')
        eff = round((act / est) * 100, 1) if est else 0
        overrun = sum(1 for t in tl if (t.get('qa_actual_hours') or 0) > (t.get('qa_estimate_hours') or 999))
        mods = Counter(t.get('module') or '-' for t in tl)
        top_mod = mods.most_common(1)[0][0] if mods else '-'
        write_row(ws, r, [name, c_cnt, p_cnt, c_cnt - p_cnt, est, act, f'{eff}%', overrun, member_active.get(name, 0), top_mod])
    ws.column_dimensions['A'].width = 25; ws.column_dimensions['J'].width = 25
    for c in 'BCDEFGHI': ws.column_dimensions[c].width = 16

def write_ticket_sheet(ws, tlist, title_suffix):
    write_header(ws, 1, ['Ticket', 'Title', 'Priority', 'QC Tester', 'Module', 'Platform',
                         'QA Est', 'QA Actual', 'Over/Under', 'Dev Est', 'Dev Actual', 'Type', 'Closed On'], blue_fill)
    for i, t in enumerate(sorted(tlist, key=lambda x: -x['ticket_id']), 2):
        est = t.get('qa_estimate_hours') or 0; act = t.get('qa_actual_hours') or 0
        diff = round(act - est, 1)
        write_row(ws, i, [t['ticket_id'], t['title'], t['priority'], t.get('qc_tester') or '-',
                          t.get('module') or '-', t.get('platform') or '-', est, act,
                          f'+{diff}' if diff > 0 else str(diff),
                          t.get('dev_estimate_hours') or 0, t.get('actual_dev_hours') or 0,
                          t.get('ticket_type') or '-', t.get('closed_on') or '-'])
        if diff > 0: ws.cell(row=i, column=9).fill = light_red
        elif diff < 0: ws.cell(row=i, column=9).fill = light_green
    ws.column_dimensions['A'].width = 10; ws.column_dimensions['B'].width = 55; ws.column_dimensions['E'].width = 22

# Member active tickets
member_active = Counter()
for t in tickets:
    if t['status'] in QC + ('Approved for Live',):
        for n in (x.strip() for x in (t.get('qc_tester') or '').split(',') if x.strip()):
            member_active[n] += 1

# ===== WEEKLY REPORT =====
cw = [t for t in qa_tickets if parse_d(t.get('closed_on')) and parse_d(t['closed_on']) >= d7]
pw = [t for t in qa_tickets if parse_d(t.get('closed_on')) and d14 <= parse_d(t['closed_on']) < d7]

wb = Workbook()
ws = wb.active; ws.title = 'Weekly Summary'
ws.column_dimensions['A'].width = 45; ws.column_dimensions['B'].width = 20; ws.column_dimensions['C'].width = 20; ws.column_dimensions['D'].width = 18

ws.cell(row=1, column=1, value=f'QA Weekly Report - {d7.strftime("%b %d")} to {today.strftime("%b %d, %Y")}').font = Font(bold=True, size=14)

r = write_section(ws, 3, 'Throughput Comparison')
write_header(ws, r, ['Metric', f'This Week ({d7.strftime("%b %d")} - {today.strftime("%b %d")})',
                      f'Prev Week ({d14.strftime("%b %d")} - {d7.strftime("%b %d")})', 'Change'], green_fill)
r += 1
for label, curr, prev in [
    ('Tickets Closed by QA', len(cw), len(pw)),
    ('QA Estimated Hours', sum_h(cw, 'qa_estimate_hours'), sum_h(pw, 'qa_estimate_hours')),
    ('QA Actual Hours', sum_h(cw, 'qa_actual_hours'), sum_h(pw, 'qa_actual_hours')),
    ('Avg QA Hours/Ticket', round(sum_h(cw, 'qa_actual_hours') / max(1, len(cw)), 1), round(sum_h(pw, 'qa_actual_hours') / max(1, len(pw)), 1)),
]:
    write_comparison(ws, r, label, curr, prev); r += 1

r = write_section(ws, r + 1, 'Time Consumption Analysis')
write_header(ws, r, ['Metric', 'This Week', 'Prev Week', 'Change'], blue_fill); r += 1
est_w = sum_h(cw, 'qa_estimate_hours'); act_w = sum_h(cw, 'qa_actual_hours')
est_pw = sum_h(pw, 'qa_estimate_hours'); act_pw = sum_h(pw, 'qa_actual_hours')
eff_w = round((act_w / est_w) * 100, 1) if est_w else 0
eff_pw = round((act_pw / est_pw) * 100, 1) if est_pw else 0
write_row(ws, r, ['QA Time Efficiency (actual/estimate %)', f'{eff_w}%', f'{eff_pw}%', f'{eff_w - eff_pw:+.1f}%']); r += 1
overrun_w = sum(1 for t in cw if (t.get('qa_actual_hours') or 0) > (t.get('qa_estimate_hours') or 999))
overrun_pw = sum(1 for t in pw if (t.get('qa_actual_hours') or 0) > (t.get('qa_estimate_hours') or 999))
write_comparison(ws, r, 'Tickets Over QA Estimate', overrun_w, overrun_pw); r += 1

r = write_section(ws, r + 1, 'Current QA Pipeline')
r = write_pipeline(ws, r)

ws2 = wb.create_sheet('Member Performance')
write_member_sheet(ws2, cw, pw, member_active)

ws3 = wb.create_sheet('Closed Tickets (This Week)')
write_ticket_sheet(ws3, cw, 'This Week')

# QA Member Ticket Mapping (Weekly)
ws_map = wb.create_sheet('QA Ticket Mapping')
ws_map.column_dimensions['A'].width = 12; ws_map.column_dimensions['B'].width = 55; ws_map.column_dimensions['C'].width = 22
ws_map.column_dimensions['D'].width = 18; ws_map.column_dimensions['E'].width = 22; ws_map.column_dimensions['F'].width = 14; ws_map.column_dimensions['G'].width = 14

# Get all QC active tickets grouped by tester
from collections import defaultdict as _dd
qa_by_tester = _dd(list)
for t in qa_tickets:
    if t.get('status') in ('QC Testing', 'QC Testing in Progress', 'QC Testing Hold', 'QC Review Fail', 'BIS Testing', 'Approved for Live'):
        tester = (t.get('qc_tester') or '').strip()
        if tester:
            for n in (x.strip() for x in tester.split(',') if x.strip()):
                qa_by_tester[n].append(t)

mr = 1
purple = PatternFill(start_color='6A1B9A', end_color='6A1B9A', fill_type='solid')
for name in sorted(qa_by_tester.keys()):
    tl = qa_by_tester[name]
    cell = ws_map.cell(row=mr, column=1, value=f'{name} ({len(tl)} tickets)')
    cell.font = Font(bold=True, size=11, color='FFFFFF'); cell.fill = purple
    for c in range(2, 8): ws_map.cell(row=mr, column=c).fill = purple
    mr += 1
    write_header(ws_map, mr, ['Ticket', 'Title', 'Status', 'Priority', 'Module', 'QA Est', 'QA Actual'], PatternFill(start_color='0D47A1', end_color='0D47A1', fill_type='solid'))
    mr += 1
    for t in sorted(tl, key=lambda x: x.get('status', '')):
        write_row(ws_map, mr, [t['ticket_id'], t['title'], t['status'], t['priority'], t.get('module', '-'),
                                t.get('qa_estimate_hours') or 0, t.get('qa_actual_hours') or 0])
        mr += 1
    mr += 1

# ===== RESOURCE PERFORMANCE ANALYSIS SHEET =====
ws_rp = wb.create_sheet('Resource Performance')
ws_rp.column_dimensions['A'].width = 25

# Count per member: how many tickets in each status
QA_STATUS_ORDER = ['QC Testing', 'QC Testing in Progress', 'QC Testing Hold', 'QC Review Fail',
                   'BIS Testing', 'Approved for Live', 'Moved to Live', 'Closed']
member_status = defaultdict(lambda: defaultdict(int))
member_totals = defaultdict(int)
member_hours = defaultdict(lambda: {'est': 0, 'act': 0})

for t in tickets:
    tester = (t.get('qc_tester') or '').strip()
    if not tester:
        continue
    for name in (x.strip() for x in tester.split(',') if x.strip()):
        s = t['status']
        member_status[name][s] += 1
        member_totals[name] += 1
        member_hours[name]['est'] += t.get('qa_estimate_hours') or 0
        member_hours[name]['act'] += t.get('qa_actual_hours') or 0

ws_rp.cell(row=1, column=1, value='QA Resource Performance Analysis').font = Font(bold=True, size=14)
ws_rp.cell(row=2, column=1, value=f'Tickets handled per member by current status — {today.strftime("%b %d, %Y")}').font = Font(size=9, color='666666')

# Headers
rp_headers = ['QA Member'] + QA_STATUS_ORDER + ['Total', 'Est Hours', 'Actual Hours', 'Efficiency %']
rp_r = 4
for c, h in enumerate(rp_headers, 1):
    cell = ws_rp.cell(row=rp_r, column=c, value=h)
    cell.font = hf; cell.fill = blue_fill; cell.alignment = ca; cell.border = tb
    if c > 1: ws_rp.column_dimensions[chr(64 + c) if c <= 26 else 'A'].width = 14
rp_r += 1

# Status color fills
rp_status_fills = {
    'QC Testing in Progress': PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid'),
    'QC Review Fail': PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid'),
    'BIS Testing': PatternFill(start_color='EDE7F6', end_color='EDE7F6', fill_type='solid'),
    'Approved for Live': PatternFill(start_color='E0F7FA', end_color='E0F7FA', fill_type='solid'),
    'Moved to Live': PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid'),
    'Closed': PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid'),
}

# Sort members by total tickets desc
for name in sorted(member_status.keys(), key=lambda x: -member_totals[x]):
    vals = [name]
    for s in QA_STATUS_ORDER:
        vals.append(member_status[name].get(s, 0) or '')
    total = member_totals[name]
    est = round(member_hours[name]['est'], 1)
    act = round(member_hours[name]['act'], 1)
    eff = round(act / est * 100, 1) if est > 0 else 0
    vals += [total, est, act, f'{eff}%']
    write_row(ws_rp, rp_r, vals)
    # Bold total
    ws_rp.cell(row=rp_r, column=len(QA_STATUS_ORDER) + 2).font = Font(bold=True, size=9)
    # Color QC Fail column if > 0
    fail_col = QA_STATUS_ORDER.index('QC Review Fail') + 2
    if member_status[name].get('QC Review Fail', 0) > 0:
        ws_rp.cell(row=rp_r, column=fail_col).fill = light_red
    # Color efficiency
    eff_col = len(rp_headers)
    if eff > 120:
        ws_rp.cell(row=rp_r, column=eff_col).fill = light_red
    rp_r += 1

# Totals row
rp_r += 1
total_vals = ['GRAND TOTAL']
for s in QA_STATUS_ORDER:
    total_vals.append(sum(member_status[n].get(s, 0) for n in member_status))
total_vals += [sum(member_totals.values()),
               round(sum(h['est'] for h in member_hours.values()), 1),
               round(sum(h['act'] for h in member_hours.values()), 1), '']
write_row(ws_rp, rp_r, total_vals)
for c in range(1, len(total_vals) + 1):
    ws_rp.cell(row=rp_r, column=c).font = Font(bold=True, size=9)

ws_rp.freeze_panes = 'B5'

# ===== BUILD QUALITY ANALYSIS SHEET =====
ws_bq = wb.create_sheet('Build Quality Analysis')
ws_bq.column_dimensions['A'].width = 30; ws_bq.column_dimensions['B'].width = 14
ws_bq.column_dimensions['C'].width = 12; ws_bq.column_dimensions['D'].width = 12
ws_bq.column_dimensions['E'].width = 14; ws_bq.column_dimensions['F'].width = 12
ws_bq.column_dimensions['G'].width = 14; ws_bq.column_dimensions['H'].width = 12

# Detect QC failures
qc_fail_tickets = [t for t in tickets if t['status'] == 'QC Review Fail']
DEV_STS = {'Ready For Development', 'In Progress', 'Hold/Pending', 'Start Code Review',
           'Code Review Failed', 'Code Review Passed', 'Express Lane Review', 'Testing In Progress'}
refix_tickets = [t for t in tickets if t['status'] in DEV_STS and t.get('qc_tester')]
all_failed = qc_fail_tickets + refix_tickets
total_tested = len([t for t in tickets if t.get('qc_tester')])
fail_rate = round(len(all_failed) / total_tested * 100, 1) if total_tested else 0

# Summary
ws_bq.cell(row=1, column=1, value='Build Quality Analysis').font = Font(bold=True, size=14)
ws_bq.cell(row=2, column=1, value=f'Generated: {today.strftime("%b %d, %Y")}').font = Font(size=9, color='666666')

bq_r = 4
write_header(ws_bq, bq_r, ['Metric', 'Value'], PatternFill(start_color='B71C1C', end_color='B71C1C', fill_type='solid'))
bq_r += 1
for label, val in [
    ('Total Tickets Tested by QA', total_tested),
    ('QC Review Fail (Current)', len(qc_fail_tickets)),
    ('Refix in Dev Pipeline', len(refix_tickets)),
    ('Total Failed Builds', len(all_failed)),
    ('Fail Rate', f'{fail_rate}%'),
    ('First-time Pass Rate', f'{round(100 - fail_rate, 1)}%'),
]:
    write_row(ws_bq, bq_r, [label, val]); bq_r += 1

# Developer Quality Ranking
bq_r += 1
ws_bq.cell(row=bq_r, column=1, value='Developer Build Quality Ranking').font = Font(bold=True, size=12)
bq_r += 1
write_header(ws_bq, bq_r, ['Developer', 'Tested', 'Failed', 'Fail Rate', 'Bugs', 'Bugs/Ticket', 'Refix', 'Overrun'],
             PatternFill(start_color='E65100', end_color='E65100', fill_type='solid'))
bq_r += 1

dev_quality = defaultdict(lambda: {'total': 0, 'failed': 0, 'bugs': 0, 'refix': 0, 'overrun': 0})
for t in tickets:
    if not t.get('qc_tester'):
        continue
    devs = t.get('developers_str') or t.get('backend_developer') or ''
    is_fail = t['status'] == 'QC Review Fail' or (t['status'] in DEV_STS and t.get('qc_tester'))
    for d in (x.strip() for x in str(devs).split(',') if x.strip() and x.strip() != 'Not Assigned'):
        dev_quality[d]['total'] += 1
        if is_fail: dev_quality[d]['failed'] += 1
        if t['status'] in DEV_STS and t.get('qc_tester'): dev_quality[d]['refix'] += 1
        dev_est = t.get('dev_estimate_hours') or 0
        dev_act = t.get('actual_dev_hours') or 0
        if dev_est > 0 and dev_act > dev_est: dev_quality[d]['overrun'] += 1

for dev_name in sorted(dev_quality.keys(), key=lambda x: -dev_quality[x]['failed']):
    ds = dev_quality[dev_name]
    if ds['total'] < 3:
        continue
    fr = round(ds['failed'] / ds['total'] * 100, 1)
    bd = round(ds['bugs'] / ds['total'], 1) if ds['total'] else 0
    write_row(ws_bq, bq_r, [dev_name, ds['total'], ds['failed'], f'{fr}%', ds['bugs'], bd, ds['refix'], ds['overrun']])
    if fr > 20:
        ws_bq.cell(row=bq_r, column=4).fill = light_red
    elif fr > 10:
        ws_bq.cell(row=bq_r, column=4).fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
    bq_r += 1

# Module Quality
bq_r += 1
ws_bq.cell(row=bq_r, column=1, value='Module-wise Build Quality').font = Font(bold=True, size=12)
bq_r += 1
write_header(ws_bq, bq_r, ['Module', 'Tested', 'Failed', 'Fail Rate', 'Bugs', 'Refix', '', ''],
             PatternFill(start_color='6A1B9A', end_color='6A1B9A', fill_type='solid'))
bq_r += 1

mod_quality = defaultdict(lambda: {'total': 0, 'failed': 0, 'bugs': 0, 'refix': 0})
for t in tickets:
    if not t.get('qc_tester'):
        continue
    mod = t.get('module') or 'Unassigned'
    is_fail = t['status'] == 'QC Review Fail' or (t['status'] in DEV_STS and t.get('qc_tester'))
    mod_quality[mod]['total'] += 1
    if is_fail: mod_quality[mod]['failed'] += 1
    if t['status'] in DEV_STS and t.get('qc_tester'): mod_quality[mod]['refix'] += 1

for mod_name in sorted(mod_quality.keys(), key=lambda x: -mod_quality[x]['failed']):
    ms = mod_quality[mod_name]
    if ms['total'] < 1: continue
    fr = round(ms['failed'] / ms['total'] * 100, 1)
    write_row(ws_bq, bq_r, [mod_name, ms['total'], ms['failed'], f'{fr}%', ms['bugs'], ms['refix'], '', ''])
    if fr > 15: ws_bq.cell(row=bq_r, column=4).fill = light_red
    bq_r += 1

# Current Failures Detail
bq_r += 1
ws_bq.cell(row=bq_r, column=1, value='Current QC Review Failures').font = Font(bold=True, size=12)
bq_r += 1
write_header(ws_bq, bq_r, ['Ticket', 'Title', 'Module', 'Developer', 'QC Tester', 'Priority', 'QA Hrs', 'Verdict'],
             PatternFill(start_color='B71C1C', end_color='B71C1C', fill_type='solid'))
bq_r += 1
for t in sorted(all_failed, key=lambda x: x.get('qa_actual_hours') or 99):
    qa_hrs = t.get('qa_actual_hours') or 0
    if qa_hrs < 1: verdict = 'Critical - Obvious bug'
    elif qa_hrs < 2: verdict = 'Poor - Basic failure'
    else: verdict = 'Moderate'
    write_row(ws_bq, bq_r, [t['ticket_id'], t['title'][:50], t.get('module', ''),
              (t.get('developers_str') or '')[:25], t.get('qc_tester', ''),
              t['priority'], round(qa_hrs, 1), verdict])
    if 'Critical' in verdict: ws_bq.cell(row=bq_r, column=8).fill = light_red
    bq_r += 1

os.makedirs(REPORTS_DIR, exist_ok=True)
wk_path = os.path.join(REPORTS_DIR, f'QA_Report_{today.strftime("%Y%m%d")}.xlsx')
wb.save(wk_path)
print(f'Weekly: {wk_path}')

# ===== MONTHLY REPORT =====
cm = [t for t in qa_tickets if parse_d(t.get('closed_on')) and parse_d(t['closed_on']) >= d30]
pm_list = [t for t in qa_tickets if parse_d(t.get('closed_on')) and d60 <= parse_d(t['closed_on']) < d30]

wb = Workbook()
ws = wb.active; ws.title = 'Monthly Summary'
ws.column_dimensions['A'].width = 45; ws.column_dimensions['B'].width = 20; ws.column_dimensions['C'].width = 20; ws.column_dimensions['D'].width = 18

ws.cell(row=1, column=1, value=f'QA Monthly Report - {d30.strftime("%b %d")} to {today.strftime("%b %d, %Y")}').font = Font(bold=True, size=14)

r = write_section(ws, 3, 'Monthly Throughput Comparison')
write_header(ws, r, ['Metric', f'This Month ({d30.strftime("%b %d")} - {today.strftime("%b %d")})',
                      f'Prev Month ({d60.strftime("%b %d")} - {d30.strftime("%b %d")})', 'Change'], green_fill)
r += 1
for label, curr, prev in [
    ('Tickets Closed by QA', len(cm), len(pm_list)),
    ('QA Estimated Hours', sum_h(cm, 'qa_estimate_hours'), sum_h(pm_list, 'qa_estimate_hours')),
    ('QA Actual Hours', sum_h(cm, 'qa_actual_hours'), sum_h(pm_list, 'qa_actual_hours')),
    ('Avg QA Hours/Ticket', round(sum_h(cm, 'qa_actual_hours') / max(1, len(cm)), 1), round(sum_h(pm_list, 'qa_actual_hours') / max(1, len(pm_list)), 1)),
]:
    write_comparison(ws, r, label, curr, prev); r += 1

r = write_section(ws, r + 1, 'Time Consumption & Efficiency')
write_header(ws, r, ['Metric', 'This Month', 'Prev Month', 'Change'], blue_fill); r += 1
est_m = sum_h(cm, 'qa_estimate_hours'); act_m = sum_h(cm, 'qa_actual_hours')
est_pm = sum_h(pm_list, 'qa_estimate_hours'); act_pm = sum_h(pm_list, 'qa_actual_hours')
eff_m = round((act_m / est_m) * 100, 1) if est_m else 0
eff_pm = round((act_pm / est_pm) * 100, 1) if est_pm else 0
write_row(ws, r, ['QA Time Efficiency (actual/estimate %)', f'{eff_m}%', f'{eff_pm}%', f'{eff_m - eff_pm:+.1f}%']); r += 1
overrun_m = sum(1 for t in cm if (t.get('qa_actual_hours') or 0) > (t.get('qa_estimate_hours') or 999))
overrun_pm = sum(1 for t in pm_list if (t.get('qa_actual_hours') or 0) > (t.get('qa_estimate_hours') or 999))
write_comparison(ws, r, 'Tickets Over QA Estimate', overrun_m, overrun_pm); r += 1
write_row(ws, r, ['Total Remaining QA Hours (pipeline)', round(max(0, sum_h(in_qc, 'qa_estimate_hours') - sum_h(in_qc, 'qa_actual_hours')), 1), '-', '-']); r += 1

r = write_section(ws, r + 1, 'Current QA Pipeline Snapshot')
r = write_pipeline(ws, r)

# Module breakdown
r = write_section(ws, r + 1, 'Tickets Closed by Module (This Month)')
write_header(ws, r, ['Module', 'Closed', 'Est Hours', 'Actual Hours', 'Over/Under'], orange_fill); r += 1
mod_closed = defaultdict(list)
for t in cm: mod_closed[t.get('module') or 'Unassigned'].append(t)
for mod, tl in sorted(mod_closed.items(), key=lambda x: -len(x[1])):
    est = sum_h(tl, 'qa_estimate_hours'); act = sum_h(tl, 'qa_actual_hours')
    write_row(ws, r, [mod, len(tl), est, act, round(act - est, 1)]); r += 1

ws2 = wb.create_sheet('Member Performance')
write_member_sheet(ws2, cm, pm_list, member_active)

ws3 = wb.create_sheet('Closed Tickets (This Month)')
write_ticket_sheet(ws3, cm, 'This Month')

ws4 = wb.create_sheet('BIS & Approved for Live')
write_header(ws4, 1, ['Ticket', 'Title', 'Priority', 'QC Tester', 'Module', 'Status', 'QA Est', 'QA Actual', 'Created'], red_fill)
for i, t in enumerate(sorted(in_bis + in_appr, key=lambda x: -x['ticket_id']), 2):
    write_row(ws4, i, [t['ticket_id'], t['title'], t['priority'], t.get('qc_tester') or '-',
                       t.get('module') or '-', t['status'], t.get('qa_estimate_hours') or 0,
                       t.get('qa_actual_hours') or 0, t.get('created_on') or '-'])
ws4.column_dimensions['A'].width = 10; ws4.column_dimensions['B'].width = 55; ws4.column_dimensions['E'].width = 22

# QA Member Ticket Mapping (Monthly)
ws_mmap = wb.create_sheet('QA Ticket Mapping')
ws_mmap.column_dimensions['A'].width = 12; ws_mmap.column_dimensions['B'].width = 55; ws_mmap.column_dimensions['C'].width = 22
ws_mmap.column_dimensions['D'].width = 18; ws_mmap.column_dimensions['E'].width = 22; ws_mmap.column_dimensions['F'].width = 14; ws_mmap.column_dimensions['G'].width = 14

qa_by_tester_m = _dd(list)
for t in qa_tickets:
    if t.get('status') in ('QC Testing', 'QC Testing in Progress', 'QC Testing Hold', 'QC Review Fail', 'BIS Testing', 'Approved for Live'):
        tester = (t.get('qc_tester') or '').strip()
        if tester:
            for n in (x.strip() for x in tester.split(',') if x.strip()):
                qa_by_tester_m[n].append(t)

mmr = 1
for name in sorted(qa_by_tester_m.keys()):
    tl = qa_by_tester_m[name]
    cell = ws_mmap.cell(row=mmr, column=1, value=f'{name} ({len(tl)} tickets)')
    cell.font = Font(bold=True, size=11, color='FFFFFF'); cell.fill = purple
    for c in range(2, 8): ws_mmap.cell(row=mmr, column=c).fill = purple
    mmr += 1
    write_header(ws_mmap, mmr, ['Ticket', 'Title', 'Status', 'Priority', 'Module', 'QA Est', 'QA Actual'], PatternFill(start_color='0D47A1', end_color='0D47A1', fill_type='solid'))
    mmr += 1
    for t in sorted(tl, key=lambda x: x.get('status', '')):
        write_row(ws_mmap, mmr, [t['ticket_id'], t['title'], t['status'], t['priority'], t.get('module', '-'),
                                  t.get('qa_estimate_hours') or 0, t.get('qa_actual_hours') or 0])
        mmr += 1
    mmr += 1

mn_path = os.path.join(REPORTS_DIR, f'QA_Monthly_Report_{today.strftime("%Y%m%d")}.xlsx')
wb.save(mn_path)
print(f'Monthly: {mn_path}')
